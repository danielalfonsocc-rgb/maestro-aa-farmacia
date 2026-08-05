#!/usr/bin/env python3
"""
servicios_farmaceuticos.py — Recuento mensual de actividades QF (agenda médica SSASUR)
═══════════════════════════════════════════════════════════════════════════════
Toma el Excel/CSV "Listado de Actividades" que descarga AUTO_SSASUR.py del
aplicativo Agenda Médica de SSASUR (index_menu.php → Proceso_nuevo → Hoja
Diaria de Profesional, Tipo Reporte "Principalmente Actividades") y arma un
resumen agregado: cada Químico Farmacéutico del hospital, con el recuento de
cada actividad que registró en el mes.

El reporte crudo trae UNA FILA POR ATENCIÓN DE PACIENTE (incluye RUT) — igual
que informe_completo_recetas*.csv y reporte_de_stock*.xlsx (Ley 19.628, ver
CLAUDE.md). El Excel de salida de este script es SOLO el agregado (QF ×
actividad → conteo), sin RUT ni nombre de paciente, así que es seguro
sincronizarlo a Drive (a diferencia del archivo crudo, que nunca se sube).

Filtra por columna ESPECIALIDAD = 'QUIMICO FARMACEUTICO' (valor confirmado en
el reporte real — no hay columna de estamento separada).

Uso:
  py servicios_farmaceuticos.py [archivo] [--mes MM-AAAA]

  archivo   Ruta al Excel/CSV descargado. Si se omite, toma el más reciente
            'reporte_listado_actividades_*' en la carpeta del proyecto.
  --mes     Fuerza la etiqueta de mes (por defecto: mes más frecuente en
            FECHA EJECUTADA de las filas QF).

Salida: Servicios_Farmaceuticos/<MES AÑO>/Resumen_Servicios_Farmaceuticos_<mes>_<año>.xlsx
"""
import argparse
import glob
import os
import sys
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import aa_colors
from gt_maestro import MESES_ES

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
SALIDA_DIR = os.path.join(WORK_DIR, "Servicios_Farmaceuticos")
ESPECIALIDAD_QF = "QUIMICO FARMACEUTICO"


def _norm(s: str) -> str:
    """Mayúsculas sin tildes/espacios extra — para comparar 'QUÍMICO FARMACÉUTICO'
    con 'QUIMICO FARMACEUTICO' sin importar cómo vino codificado el Excel."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.upper().split())


def _archivo_mas_reciente() -> str | None:
    cand = [
        f for f in glob.glob(os.path.join(WORK_DIR, "reporte_listado_actividades_*"))
        if not os.path.basename(f).startswith("~$") and f.lower().endswith((".xlsx", ".csv"))
    ]
    return max(cand, key=os.path.getmtime) if cand else None


def cargar_reporte(path: str) -> pd.DataFrame:
    """Lee el Excel o CSV crudo de la Agenda Médica. Mismo separador/codificación
    que el resto de reportes SSASUR (';' + latin-1) para el caso CSV."""
    if path.lower().endswith(".csv"):
        for enc in ("latin-1", "utf-8-sig", "cp1252"):
            try:
                return pd.read_csv(path, sep=";", encoding=enc, dtype=str)
            except (UnicodeDecodeError, Exception):
                continue
        raise RuntimeError(f"No se pudo leer el CSV con ninguna codificación conocida: {path}")
    return pd.read_excel(path, dtype=str)


def filtrar_qf(df: pd.DataFrame) -> pd.DataFrame:
    if "ESPECIALIDAD" not in df.columns:
        raise KeyError(
            f"El reporte no tiene columna ESPECIALIDAD (columnas: {list(df.columns)[:10]}...). "
            "¿Es el reporte correcto (Hoja Diaria de Profesional → Principalmente Actividades)?"
        )
    esp_norm = df["ESPECIALIDAD"].map(_norm)
    qf = df[esp_norm == ESPECIALIDAD_QF].copy()
    qf["QF"] = (
        qf["NOMBRE PROFESIONAL"].fillna("").str.strip() + " " +
        qf["PATERNO PROFESIONAL"].fillna("").str.strip() + " " +
        qf["MATERNO PROFESIONAL"].fillna("").str.strip()
    ).str.replace(r"\s+", " ", regex=True).str.strip()
    return qf


def _mes_predominante(qf: pd.DataFrame) -> tuple[int, int]:
    """(mes, año) más frecuente en FECHA EJECUTADA de las filas QF."""
    fechas = pd.to_datetime(qf["FECHA EJECUTADA"], format="%d/%m/%Y", errors="coerce").dropna()
    if fechas.empty:
        hoy = datetime.today()
        return hoy.month, hoy.year
    periodos = fechas.dt.to_period("M")
    top = periodos.value_counts().idxmax()
    return top.month, top.year


def _pivot(qf: pd.DataFrame, columna: str) -> pd.DataFrame:
    """QF (filas) × valores de `columna` (columnas) → conteo, con fila/columna TOTAL."""
    tabla = pd.crosstab(qf["QF"], qf[columna])
    tabla = tabla.reindex(sorted(tabla.columns), axis=1)
    tabla["TOTAL"] = tabla.sum(axis=1)
    tabla = tabla.sort_values("TOTAL", ascending=False)
    fila_total = tabla.sum(axis=0)
    fila_total.name = "TOTAL"
    tabla = pd.concat([tabla, fila_total.to_frame().T])
    return tabla


def _escribir_hoja(ws, tabla: pd.DataFrame, titulo: str):
    header_fill = PatternFill("solid", fgColor=aa_colors.soften(aa_colors.TEAL, keep=0.85))
    header_font = Font(bold=True, color=aa_colors.darken(aa_colors.TEAL, keep=0.65), name="Arial", size=10)
    total_fill = PatternFill("solid", fgColor=aa_colors.GRIS)
    total_font = Font(bold=True, name="Arial", size=10)

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color=aa_colors.darken(aa_colors.TEAL, keep=0.65))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(tabla.columns) + 1)

    fila0 = 3
    ws.cell(row=fila0, column=1, value="QF").font = header_font
    ws.cell(row=fila0, column=1).fill = header_fill
    for j, col in enumerate(tabla.columns, start=2):
        c = ws.cell(row=fila0, column=j, value=str(col))
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (qf_nombre, fila) in enumerate(tabla.iterrows(), start=fila0 + 1):
        es_total = qf_nombre == "TOTAL"
        c0 = ws.cell(row=i, column=1, value=qf_nombre)
        c0.font = total_font if es_total else Font(name="Arial", size=10)
        if es_total:
            c0.fill = total_fill
        for j, col in enumerate(tabla.columns, start=2):
            valor = int(fila[col])
            c = ws.cell(row=i, column=j, value=valor if valor else None)
            c.alignment = Alignment(horizontal="center")
            c.font = total_font if es_total else Font(name="Arial", size=10)
            if es_total:
                c.fill = total_fill

    ws.column_dimensions["A"].width = 32
    for j in range(2, len(tabla.columns) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "B4"


def generar_resumen(qf: pd.DataFrame, mes: int, anio: int) -> str:
    """OJO: en el reporte crudo de SSASUR los nombres de columna son
    contraintuitivos — 'GRUPO ACTIVIDAD' trae el detalle real (Revisión de la
    Medicación, Conciliación Farmacéutica, Educación Farmacéutica, Seguimiento
    Farmacoterapéutico...) mientras que 'ACTIVIDAD' solo distingue Atención
    Abierta/Cerrada (verificado en vivo con el reporte de julio 2026). Por eso
    la hoja principal usa GRUPO ACTIVIDAD."""
    import openpyxl
    wb = openpyxl.Workbook()

    tabla_grupo = _pivot(qf, "GRUPO ACTIVIDAD")
    ws1 = wb.active
    ws1.title = "Resumen por Actividad"
    _escribir_hoja(ws1, tabla_grupo, f"Servicios Farmacéuticos — {MESES_ES[mes - 1]} {anio}")

    tabla_detalle = _pivot(qf, "ACTIVIDAD")
    ws2 = wb.create_sheet("Por Tipo de Atención")
    _escribir_hoja(ws2, tabla_detalle, f"Atención Abierta / Cerrada — {MESES_ES[mes - 1]} {anio}")

    carpeta_mes = f"{MESES_ES[mes - 1]} {anio}"
    destino_dir = os.path.join(SALIDA_DIR, carpeta_mes)
    os.makedirs(destino_dir, exist_ok=True)
    nombre = f"Resumen_Servicios_Farmaceuticos_{MESES_ES[mes - 1].capitalize()}_{anio}.xlsx"
    destino = os.path.join(destino_dir, nombre)
    wb.save(destino)
    return destino


def main():
    ap = argparse.ArgumentParser(description="Recuento mensual de actividades QF desde Agenda Médica SSASUR")
    ap.add_argument("archivo", nargs="?", help="Excel/CSV descargado (default: el más reciente en la carpeta)")
    ap.add_argument("--mes", help="Forzar mes MM-AAAA (default: mes más frecuente en los datos)")
    a = ap.parse_args()

    archivo = a.archivo or _archivo_mas_reciente()
    if not archivo or not os.path.exists(archivo):
        print("[ERROR] No encontré el reporte de actividades. Pasa la ruta explícita:")
        print("  py servicios_farmaceuticos.py \"C:\\ruta\\reporte_listado_actividades_....csv\"")
        sys.exit(1)

    print(f"[Servicios Farmacéuticos] Leyendo {archivo}...")
    df = cargar_reporte(archivo)
    print(f"  {len(df):,} filas totales")

    qf = filtrar_qf(df)
    if qf.empty:
        print(f"  [AVISO] 0 filas con ESPECIALIDAD='{ESPECIALIDAD_QF}' — nada que resumir.")
        sys.exit(0)
    print(f"  {len(qf):,} filas de Químico Farmacéutico ({qf['QF'].nunique()} profesional(es))")

    if a.mes:
        mm, aaaa = a.mes.split("-")
        mes, anio = int(mm), int(aaaa)
    else:
        mes, anio = _mes_predominante(qf)

    destino = generar_resumen(qf, mes, anio)
    print(f"  ✓ {destino}")
    print(f"\n  Resumen por QF ({MESES_ES[mes - 1]} {anio}):")
    for nombre, total in qf.groupby("QF").size().sort_values(ascending=False).items():
        print(f"    {nombre}: {total} actividad(es)")


if __name__ == "__main__":
    main()
