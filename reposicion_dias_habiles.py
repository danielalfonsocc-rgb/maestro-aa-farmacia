# -*- coding: utf-8 -*-
"""
Plan de Reposicion Bodega AA -> Farmacia AT Abierta, RESTRINGIDO A DIAS HABILES.

Fuente unica de verdad de la logica de reposicion en dias habiles: la importan
TANTO la corrida standalone (este __main__, genera el Excel) COMO la app
(app_pedidos.py, pestaña "Reposicion (dias habiles)") -> mismo calculo en ambos.

Por medicamento calcula:
  - Frecuencia de reposicion sugerida (en DIAS HABILES, motor SGLI: IR).
  - Stock de Seguridad y Stock Minimo de Reorden (ROP) que blindan el cierre
    de fin de semana / feriados (la bodega NO repone sab/dom ni feriados).
  - Proxima reposicion = dia habil en que el stock cae al ROP (saltando finde y
    feriados de feriados_chile.csv) + marca "Cubre Cierre".
  - Alertas: REPONER AHORA (stock 0 con consumo), BAJO MINIMO, [ALERTA_ESTRES]
    (la capacidad fisica de la gaveta no alcanza para un ciclo+colchon).

HALLAZGO EMPIRICO QUE FUNDAMENTA EL MODELO (medido sobre Fecha_Entrega real):
  La Farmacia AT Abierta NO dispensa en fin de semana (Sab 0.2% + Dom 0.0%).
  Por eso el "blindaje de fin de semana" NO es stockear el consumo de sab/dom
  (~0), sino dejar la farmacia con stock para REABRIR el lunes (dia completo)
  hasta que el camion vuelva. Ese es el colchon real.
"""
import os, math, datetime as dt
import pandas as pd

WORK_DIR          = os.path.dirname(os.path.abspath(__file__))
CONSOLIDADO       = os.path.join(WORK_DIR, "Consolidado_AA_MAESTRO.xlsx")
FERIADOS_CSV      = os.path.join(WORK_DIR, "feriados_chile.csv")
BUFFER_FINDE_DIAS = 1   # dias habiles de demanda que el SS debe cubrir para REABRIR
                        # tras el cierre. =1 porque la demanda sab/dom es ~0 (medido).
EXTRA_SS_CRITICOS = 1   # dias extra de seguridad para criticidad 1-2 (no quiebran)

# Columnas de salida (orden) -> usadas por la app y por el Excel standalone
COLUMNAS = [
    "Medicamento", "Criticidad", "Stock_Actual", "Stock_Bodega", "CDL",
    "Frecuencia", "Stock_Seguridad", "Stock_Minimo_ROP", "Cap_Fisica",
    "Cobertura_Dias", "Proxima_Reposicion", "Cubre_Cierre", "Alertas",
]

# ───────────────────────── CALENDARIO HABIL ─────────────────────────
def cargar_feriados(path=FERIADOS_CSV):
    """Devuelve dict {date: (nombre, confianza)} desde el CSV (; separado)."""
    fer = {}
    try:
        with open(path, encoding="utf-8") as fh:
            next(fh)
            for ln in fh:
                p = ln.rstrip("\n").split(";")
                if len(p) >= 1 and p[0].strip():
                    fer[dt.date.fromisoformat(p[0].strip())] = (
                        p[1].strip() if len(p) > 1 else "",
                        p[2].strip() if len(p) > 2 else "")
    except FileNotFoundError:
        pass
    return fer

def _es_habil(d, feriados):
    return d.weekday() < 5 and d not in feriados

def _add_habiles(start, n, feriados):
    d = start
    while n > 0:
        d += dt.timedelta(days=1)
        if _es_habil(d, feriados):
            n -= 1
    return d

# ───────────────────────── MOTOR ─────────────────────────
def crit_nivel_local(crit):
    """1..5 desde la etiqueta de criticidad (sin depender de aa_colors)."""
    c = str(crit).strip().upper()
    if len(c) >= 2 and c[0] in "12345" and c[1] == "-":
        return int(c[0])
    if "CRITICO" in c: return 1
    if "URGENTE" in c or "ALTO" in c: return 2
    if "MODERADO" in c: return 3
    if "BAJO" in c: return 4
    return 5

def calcular_reposicion(df_sgli, df_stock, feriados=None, hoy=None,
                        buffer_finde=BUFFER_FINDE_DIAS,
                        extra_ss_criticos=EXTRA_SS_CRITICOS):
    """Calcula el plan de reposicion en dias habiles.

    df_sgli  : DataFrame hoja SGLI_Estres (Criticidad, Dias_Reposicion_IR,
               Cap_Max, Nivel_Objetivo_T, Stock_Farm, Stock_Bod, CDL, ...).
    df_stock : DataFrame hoja Stock_AA (Medicamento, CDL_DiasHab) -> CDL plano.
    Devuelve un DataFrame con las columnas COLUMNAS, ya ordenado.
    """
    if feriados is None:
        feriados = cargar_feriados()
    if hoy is None:
        hoy = dt.date.today()
    if hasattr(hoy, "date"):
        hoy = hoy.date()

    def num(v, d=0.0):
        v = pd.to_numeric(v, errors="coerce")
        return d if pd.isna(v) else float(v)

    # CDL plano por dia habil (steady-state) desde Stock_AA
    cdl_plano = {}
    if df_stock is not None and len(df_stock):
        for _, r in df_stock.iterrows():
            cdl_plano[str(r.get("Medicamento", "")).strip()] = num(r.get("CDL_DiasHab", 0))

    # Freq_Revision per-medicamento desde SGLI (si disponible; si no, usa IR global)
    freq_rev_map = {}
    if "Freq_Revision" in df_sgli.columns:
        for _, r in df_sgli.iterrows():
            med = str(r.get("Medicamento", "")).strip()
            v = pd.to_numeric(r.get("Freq_Revision"), errors="coerce")
            if not pd.isna(v):
                freq_rev_map[med] = int(v)

    filas = []
    for _, r in df_sgli.iterrows():
        med   = str(r.get("Medicamento", "")).strip()
        if not med:
            continue
        crit  = str(r.get("Criticidad", "5-OK")).strip()
        nivel = crit_nivel_local(crit)
        # Usa Freq_Revision per-medicamento si está disponible; si no, IR global
        ir    = freq_rev_map.get(med, int(num(r.get("Dias_Reposicion_IR", 5), 5)) or 5)
        capmx = num(r.get("Cap_Max", 0))
        tobj  = num(r.get("Nivel_Objetivo_T", 0))
        sfarm = num(r.get("Stock_Farm", 0))
        sbod  = num(r.get("Stock_Bod", 0))
        cdl   = cdl_plano.get(med, num(r.get("CDL", 0)))

        sin_consumo = cdl <= 0.0001

        ss_dias = buffer_finde + (extra_ss_criticos if nivel <= 2 else 0)
        ss      = math.ceil(cdl * ss_dias)
        rop     = math.ceil(cdl * ir) + ss if not sin_consumo else 0

        estres  = (not sin_consumo) and capmx > 0 and (rop > capmx or tobj >= capmx)

        alertas = []
        reponer = bajo = False
        if sfarm <= 0 and not sin_consumo:
            alertas.append("REPONER AHORA"); reponer = True
        elif (not sin_consumo) and sfarm <= rop:
            alertas.append("BAJO MINIMO"); bajo = True
        if estres:
            alertas.append("[ALERTA_ESTRES]")
        if sin_consumo:
            alertas.append("SIN CONSUMO")
        if not alertas:
            alertas.append("OK")

        if sin_consumo:
            cob = None; prox = None
        else:
            cob  = max(0, math.floor((sfarm - rop) / cdl)) if cdl > 0 else 0
            prox = _add_habiles(hoy, max(1, cob), feriados)   # siempre dia habil
        cubre = bool(prox) and (not _es_habil(prox + dt.timedelta(days=1), feriados))
        freq  = "A demanda" if sin_consumo else \
                f"Cada {ir} dia habil" + ("es" if ir != 1 else "")

        # ranking de orden: urgencia, luego presencia de [ALERTA_ESTRES] (Sort 2)
        urg = 0 if reponer else 1 if bajo else 3 if sin_consumo else 2
        filas.append({
            "Medicamento": med, "Criticidad": crit, "Stock_Actual": int(sfarm),
            "Stock_Bodega": int(sbod), "CDL": round(cdl, 1), "Frecuencia": freq,
            "Stock_Seguridad": int(ss), "Stock_Minimo_ROP": int(rop),
            "Cap_Fisica": int(capmx),
            "Cobertura_Dias": ("" if cob is None else int(cob)),
            "Proxima_Reposicion": ("" if prox is None else prox.isoformat()),
            "Cubre_Cierre": ("Si" if cubre else ""),
            "Alertas": " + ".join(alertas),
            "_urg": urg, "_estres": 0 if estres else 1, "_nivel": nivel, "_cdl": cdl,
        })

    df = pd.DataFrame(filas)
    if len(df):
        df = df.sort_values(by=["_urg", "_estres", "_nivel", "_cdl"],
                            ascending=[True, True, True, False]).reset_index(drop=True)
        df = df[COLUMNAS]
    else:
        df = pd.DataFrame(columns=COLUMNAS)
    return df

# ───────────────────────── STANDALONE: EXCEL ─────────────────────────
def _generar_excel(df, feriados, hoy, salida):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import aa_colors as C

    HDR = [
        ("Medicamento", 46), ("Criticidad", 12), ("Stock Actual", 12),
        ("Stock Bodega (respaldo)", 14), ("CDL (ud/dia habil)", 14),
        ("Frecuencia Repos. (dias habiles)", 22), ("Stock Seguridad", 13),
        ("Stock Minimo Sugerido (ROP)", 18), ("Cap. Fisica (gaveta)", 14),
        ("Cobertura Actual (dias hab)", 14), ("Proxima Reposicion", 16),
        ("Cubre Cierre", 11), ("Alertas", 30),
    ]
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor=C.soften("0F766E", 0.18))
    hfont = Font(bold=True, color=C.text_on(C.soften("0F766E", 0.18)), name="Arial", size=10)

    owb = openpyxl.Workbook(); ws = owb.active; ws.title = "Plan_Reposicion"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HDR))
    t = ws.cell(1, 1, f"PLAN DE REPOSICION EN DIAS HABILES  -  Bodega AA -> Farmacia AT Abierta"
                      f"   |   Generado {hoy.isoformat()}")
    t.font = Font(bold=True, size=13, color="0F766E", name="Arial")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HDR))
    lg = ws.cell(2, 1, "Alertas: REPONER AHORA = stock 0 con consumo (urgente) | "
        "BAJO MINIMO = stock <= ROP, pedir el proximo dia habil | "
        "[ALERTA_ESTRES] = la gaveta no alcanza para un ciclo+colchon (sobrecarga viernes / ampliar estante) | "
        "Cubre Cierre = esta entrega debe alcanzar para un fin de semana o feriado.")
    lg.font = Font(italic=True, size=9, color="555555", name="Arial")
    lg.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30

    HROW = 3
    for j, (name, w) in enumerate(HDR, 1):
        c = ws.cell(HROW, j, name); c.fill = hfill; c.font = hfont; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, (_, row) in enumerate(df.iterrows()):
        rr = HROW + 1 + i
        hexc = C.crit_hex(row["Criticidad"])
        reponer = "REPONER AHORA" in row["Alertas"]
        estres  = "[ALERTA_ESTRES]" in row["Alertas"]
        for j, (val) in enumerate(row[COLUMNAS], 1):
            c = ws.cell(rr, j, val); c.border = border
            c.fill = PatternFill("solid", fgColor=hexc)
            c.font = Font(name="Arial", size=10, color="1F2937")
            if 3 <= j <= 5 or 7 <= j <= 12:
                c.alignment = Alignment(horizontal="center")
        if reponer:
            ws.cell(rr, 1).font = C.FONT_CRITICO
            ws.cell(rr, len(HDR)).font = C.FONT_CRITICO
        elif estres:
            ws.cell(rr, len(HDR)).font = Font(name="Arial", size=10, bold=True, color="9A3412")
    ws.freeze_panes = f"A{HROW+1}"
    ws.auto_filter.ref = f"A{HROW}:{get_column_letter(len(HDR))}{HROW+len(df)}"

    # Hoja Metodologia
    wm = owb.create_sheet("Metodologia")
    metodo = [
     ("CRITERIO DE BLINDAJE DE FIN DE SEMANA Y FERIADOS", True), ("", False),
     ("1. La bodega repone solo en dias habiles. La demanda real de la Farmacia AT", False),
     ("   Abierta es ~0 en sab/dom (medido sobre Fecha_Entrega: 0.19%). El blindaje", False),
     ("   NO es stockear consumo de sab/dom (~0), sino dejar stock para REABRIR el", False),
     ("   lunes hasta que vuelva el camion.", False), ("", False),
     ("2. CDL = consumo promedio por DIA HABIL (de Fecha_Entrega real, no digitacion", False),
     ("   -> no hay 'pico fantasma' del lunes).", False), ("", False),
     (f"3. Stock de Seguridad = CDL x ({BUFFER_FINDE_DIAS} reapertura + {EXTRA_SS_CRITICOS} extra si criticidad 1-2).", False),
     ("4. Stock Minimo de Reorden (ROP) = CDL x IR + Stock de Seguridad.", False),
     ("   IR = frecuencia de reposicion en dias habiles (motor SGLI: rotacion+capacidad).", False),
     ("5. Fechas en DIAS HABILES, saltando sab/dom y feriados (hoja Feriados): una", False),
     ("   reposicion nunca cae en dia cerrado; el adelanto al viernes es automatico.", False),
     ("6. [ALERTA_ESTRES] = ROP > capacidad fisica de gaveta (Cap_Max): no caben las", False),
     ("   unidades de un ciclo+colchon -> reponer mas seguido (sobrecarga viernes) o", False),
     ("   ampliar estante.", False), ("", False),
     ("ADVERTENCIAS:", True),
     ("- Feriados: este plan solo conoce los de feriados_chile.csv. Mantenerlo al dia.", False),
     ("- Sobrecarga del viernes: revisar 'Cubre Cierre' + [ALERTA_ESTRES] (concentran", False),
     ("  volumen el viernes); validar espacio en estante.", False),
     ("- Consumo fantasma del lunes: mitigado (se usa Fecha_Entrega, no digitacion).", False),
    ]
    for i, (txt, bold) in enumerate(metodo, 1):
        c = wm.cell(i, 1, txt)
        c.font = Font(bold=bold, size=11 if bold else 10, name="Arial",
                      color="0F766E" if bold else "1F2937")
    wm.column_dimensions["A"].width = 95

    # Hoja Feriados
    wf = owb.create_sheet("Feriados")
    wf.append(["Fecha", "Nombre", "Confianza"])
    for c in wf[1]:
        c.font = Font(bold=True); c.fill = hfill
    for d in sorted(feriados):
        nombre, conf = feriados[d]
        wf.append([d.isoformat(), nombre, conf])
    wf.column_dimensions["A"].width = 14
    wf.column_dimensions["B"].width = 48
    wf.column_dimensions["C"].width = 12

    owb.save(salida)

def main():
    hoy = dt.date.today()
    feriados = cargar_feriados()
    print(f"HOY = {hoy} ({['Lun','Mar','Mie','Jue','Vie','Sab','Dom'][hoy.weekday()]})"
          f"  | habil hoy: {_es_habil(hoy, feriados)}")
    print(f"Feriados cargados: {len(feriados)}  | Consolidado: {os.path.basename(CONSOLIDADO)}\n")

    df_sgli  = pd.read_excel(CONSOLIDADO, sheet_name="SGLI_Estres", engine="openpyxl")
    df_stock = pd.read_excel(CONSOLIDADO, sheet_name="Stock_AA",   engine="openpyxl")
    df = calcular_reposicion(df_sgli, df_stock, feriados, hoy)

    n_rep = df["Alertas"].str.contains("REPONER AHORA").sum()
    n_baj = df["Alertas"].str.contains("BAJO MINIMO").sum()
    n_est = df["Alertas"].str.contains(r"\[ALERTA_ESTRES\]").sum()
    n_cub = (df["Cubre_Cierre"] == "Si").sum()
    print(f"Medicamentos analizados : {len(df)}")
    print(f"  REPONER AHORA (stock 0 con consumo): {n_rep}")
    print(f"  BAJO MINIMO (<= ROP)               : {n_baj}")
    print(f"  [ALERTA_ESTRES] (cap. insuficiente): {n_est}")
    print(f"  'Cubre Cierre' (entregas blindaje) : {n_cub}")

    salida = os.path.join(WORK_DIR, f"Reposicion_DiasHabiles_AA_{hoy.strftime('%Y%m%d')}.xlsx")
    _generar_excel(df, feriados, hoy, salida)
    print(f"\nExcel generado: {os.path.basename(salida)}  ({len(df)} medicamentos, 3 hojas)")

if __name__ == "__main__":
    main()
