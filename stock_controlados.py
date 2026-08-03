#!/usr/bin/env python3
"""stock_controlados.py — Stock del sistema para el conteo de controlados.

Cada bodega usa el reporte que le corresponde (controlados_config.json → fuente_stock):
  · AT Cerrada → "Informe Stock" (stock_en_momento_bodega): existencia ACTUAL, con
    el check "Solo controlados". Se registra con la fecha de HOY (día actual).
  · AT Abierta → "Reporte Stock en Fecha" (stockFecha): stock del DÍA ANTERIOR
    (para contar en la mañana contra el cierre del día previo).
Cruza el resultado con el mapeo de controlados_config.json y escribe:

    Conteo_Controlados/stock_sistema.json

que la app conteo_controlados_app.py inyecta al formulario para PRE-LLENAR la
columna "Sistema" con el stock del día anterior. La columna Acta/Vencimiento NO
se toca nunca (la maneja la QF a mano).

Estructura del JSON:
    {
      "fecha": "2026-07-23",                 # día consultado (día anterior)
      "generadoEn": "2026-07-24T08:15:00",
      "porFarmacia": {
        "abierta": { "Alprazolam|0,5 mg cm": 120, ... },
        "cerrada": { "Morfina|10 mg/ml am": 30, ... }
      }
    }

Uso:
  · Normalmente NO se corre suelto: AUTO_SSASUR.py lo llama dentro del módulo
    ABASTECIMIENTO (reutiliza la sesión ya logueada). Ver PASO 4d en AUTO_SSASUR.
  · Suelto para probar/backfill:  py stock_controlados.py            (día anterior)
                                  py stock_controlados.py 23/07/2026 (fecha fija)

NOTA — primera corrida en vivo: los IDs exactos del formulario stockFecha
(select de bodega, campo de fecha, botón Generar XLS) NO están confirmados por
inspección todavía. Este script los busca de forma DEFENSIVA (bodega por texto
de la opción, fecha por el input con formato dd/mm/aaaa, botón por su texto) y,
si algo falla, vuelca [DESCUBRIR …] y guarda debug_stock_controlados.png para
ajustar los selectores — mismo patrón que paso_gt / paso_controlados en
AUTO_SSASUR.py.
"""
import asyncio
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

MAESTRO_DIR = Path(__file__).parent
CONFIG_FILE = MAESTRO_DIR / "controlados_config.json"
SALIDA_DIR  = MAESTRO_DIR / "Conteo_Controlados"
SALIDA_JSON = SALIDA_DIR / "stock_sistema.json"
STOCKFECHA_URL = "https://www.ssasur.cl/abastecimiento/reportes/stockFecha"
# Informe Stock (existencia actual). AT Cerrada usa ESTE reporte (stock del día
# actual, con el check "Solo controlados"); es el mismo que AUTO_SSASUR baja con
# bodega=TODAS para el maestro. AT Abierta usa STOCKFECHA_URL (día anterior).
STOCK_MOMENTO_URL = "https://www.ssasur.cl/abastecimiento/reportes/stock_en_momento_bodega"
TIMEOUT_DESCARGA = 600_000
# stockFecha (AT Abierta) tiene un historial de veces en que el evento de
# descarga simplemente nunca se dispara (27-07 y 31-07-2026) — esperar los
# 10 min completos de TIMEOUT_DESCARGA en cada uno de los 2 intentos deja el
# paso [4d/9] colgado hasta 20 min sin ganar nada. Con el fix de
# _fijar_nivel_reporte_local (dispara 'change' siempre) el caso más común
# queda resuelto; para lo que igual falle, mejor fallar rápido y dejar que
# el 2º intento tenga tiempo real de correr.
TIMEOUT_DESCARGA_STOCKFECHA = 60_000


def _fmt(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def cargar_config() -> dict:
    try:
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  [ERROR] No se pudo leer {CONFIG_FILE.name}: {e}")
        return {}


# ── Descubrimiento defensivo del formulario ────────────────────────────────────
async def _dump_formulario(page, etiqueta="stockFecha"):
    try:
        info = await page.evaluate(r"""() => {
          const sels = [...document.querySelectorAll('select')].map(s => ({
            id: s.id, name: s.name,
            opts: [...s.options].slice(0, 20).map(o => `${o.value}=${(o.textContent||'').trim()}`)
          }));
          const inps = [...document.querySelectorAll('input')]
            .filter(i => !['button','submit','image','hidden'].includes(i.type))
            .map(i => ({id: i.id, name: i.name, type: i.type, value: i.value, ph: i.placeholder || ''}));
          const btns = [...document.querySelectorAll('button,a,input[type=button],input[type=submit]')]
            .map(b => (b.textContent || b.value || '').trim()).filter(Boolean).slice(0, 40);
          return {sels, inps, btns, url: location.href};
        }""")
    except Exception as e:
        print(f"  [dump] no se pudo inspeccionar: {e}")
        return
    print(f"  [DESCUBRIR · {etiqueta}] {info['url']}")
    for s in info["sels"]:
        print(f"    <select id='{s['id']}' name='{s['name']}'>  {s['opts']}")
    for i in info["inps"]:
        print(f"    <input id='{i['id']}' name='{i['name']}' type='{i['type']}'> "
              f"value='{i['value']}' ph='{i['ph']}'")
    print(f"    botones/enlaces: {info['btns']}")


async def _fijar_nivel_reporte_local(page):
    """Fija 'nivel_reporte' en 'Mi establecimiento' (value local) si existe el
    select. En el formulario stockFecha el <select> de bodega se puebla vía
    AJAX disparada por el evento 'change' de este select — NO por su valor
    (bug real 27-07-2026 y 31-07-2026, mismo síntoma en el reintento: <select
    id='bodega'> [] sin opciones). Si la página ya trae 'local' seleccionado
    de fábrica (pasa en un reload, no en la 1ª carga), el código anterior no
    disparaba 'change' por no haber cambio de valor, y el AJAX de bodega
    nunca se ejecutaba — por eso el 1er intento de una sesión suele funcionar
    y el reintento (page.goto de nuevo) llega con bodega vacío. Se dispara el
    evento SIEMPRE, haya cambiado el valor o no."""
    return await page.evaluate(r"""() => {
      const s = document.querySelector('#nivel_reporte, select[name="nivel_reporte"]');
      if (!s) return false;
      const o = [...s.options].find(o => /^local$/i.test(o.value) || /establecimiento/i.test(o.textContent || ''));
      if (!o) return false;
      s.value = o.value;
      s.dispatchEvent(new Event('change', {bubbles: true}));
      return true;
    }""")


async def _esperar_opciones_bodega(page, intentos=18, espera_ms=800):
    """Espera a que el <select> de bodega tenga opciones — se puebla vía AJAX
    tras fijar 'nivel_reporte' y puede tardar más que el wait_for_timeout fijo
    de después del goto. Devuelve True si aparecieron opciones, False si se
    agotaron los intentos (el llamador igual sigue y falla con el error de
    siempre, que ahora queda diagnosticado con [DESCUBRIR]). Ventana ampliada
    03-08-2026: con el fix de disparar 'change' siempre, el bug de bodega
    vacía en el reintento IGUAL reapareció en vivo — la AJAX parece tardar
    más de los ~8s que cubrían los 10 intentos originales en algunas
    corridas."""
    for _ in range(intentos):
        n = await page.evaluate(r"""() => {
          const s = document.querySelector('#bodega, select[name="bodega"]');
          return s ? s.options.length : -1;
        }""")
        if n and n > 0:
            return True
        await page.wait_for_timeout(espera_ms)
    return False


async def _seleccionar_bodega(page, texto_bodega: str):
    """Selecciona en el <select> de bodega la opción cuyo texto contiene
    `texto_bodega` (ej. 'FARMACIA AT ABIERTA'). Busca por texto — no por id/value
    fijo — igual que _seleccionar_bodega_at_abierta() de AUTO_SSASUR."""
    return await page.evaluate(
        r"""(texto) => {
          const norm = s => (s || '').toUpperCase().replace(/\s+/g, ' ').trim();
          const objetivo = norm(texto);
          for (const s of document.querySelectorAll('select')) {
            const o = [...s.options].find(o => norm(o.textContent).includes(objetivo));
            if (o) {
              if (s.value !== o.value) {
                s.value = o.value;
                s.dispatchEvent(new Event('change', {bubbles: true}));
              }
              return {sel: s.id || s.name || 'bodega', val: o.value, label: (o.textContent || '').trim()};
            }
          }
          return null;
        }""",
        texto_bodega,
    )


async def _marcar_solo_controlados(page):
    """Marca el check 'Solo controlados' del Informe Stock (existencia). Busca por
    el texto de la etiqueta / id / name que mencione 'controlado'."""
    return await page.evaluate(r"""() => {
      for (const cb of document.querySelectorAll('input[type=checkbox]')) {
        const lbl  = (cb.labels && cb.labels[0]) ? cb.labels[0].textContent : '';
        const meta = `${lbl} ${cb.id || ''} ${cb.name || ''}`;
        if (/controlado/i.test(meta)) {
          if (!cb.checked) { cb.click(); cb.dispatchEvent(new Event('change', {bubbles: true})); }
          return cb.id || cb.name || 'controlados';
        }
      }
      return null;
    }""")


async def _set_fecha(page, fecha_str: str):
    """Rellena el campo de Fecha (dd/mm/aaaa) tipeando de verdad (click + limpiar
    + escribir + Tab). Busca el input por id 'fecha*', o el primer input de texto
    cuyo valor tenga forma de fecha dd/mm/aaaa. Mismo enfoque de tipeo real que
    _set_fechas() de AUTO_SSASUR (los campos con máscara ignoran .value sintético)."""
    handle = await page.evaluate_handle(r"""() => {
      const cand = [...document.querySelectorAll('input')].filter(i =>
        i.type !== 'hidden' && i.type !== 'button' && i.type !== 'submit');
      // 1) por id/name que mencione 'fecha'
      let el = cand.find(i => /fecha/i.test((i.id || '') + ' ' + (i.name || '')));
      // 2) por valor con forma dd/mm/aaaa
      if (!el) el = cand.find(i => /^\d{2}\/\d{2}\/\d{4}$/.test((i.value || '').trim()));
      // 3) input type=date
      if (!el) el = cand.find(i => i.type === 'date');
      return el || null;
    }""")
    el = handle.as_element()
    if not el:
        return False
    try:
        await el.click()
        await el.press("Control+A")
        await el.press("Delete")
        await el.type(fecha_str, delay=30)
        await page.keyboard.press("Tab")
        await page.wait_for_timeout(300)
        real = await el.input_value()
        return real == fecha_str
    except Exception:
        return False


SELS_XLS = (
    "#generarXLS_stock", "#generarXLS", "#generarXLS_salida",
    'button:has-text("Generar XLS")', 'a:has-text("Generar XLS")',
    'input[value*="XLS" i]', 'button:has-text("XLS")', 'a:has-text("XLS")',
)


async def _descartar_modal(page):
    """La 1ª selección de bodega puede disparar un modal 'Selección Proyecto'
    cuyo backdrop tapa el botón Generar XLS (visto en vivo 2026-07-24: la 1ª
    bodega fallaba y la 2ª ya funcionaba, porque el proyecto quedaba fijado).
    Acepta el modal si aparece; ignora si no hay ninguno."""
    for sel in ('button:has-text("Aceptar e ingresar")',
                'a:has-text("Aceptar e ingresar")',
                'button:has-text("Aceptar")', 'a:has-text("Aceptar")',
                '.modal.show button.btn-primary', '.swal2-confirm'):
        try:
            await page.click(sel, timeout=1_500)
            await page.wait_for_timeout(600)
            return True
        except Exception:
            continue
    # Backdrop residual sin botón claro → forzar cierre por Escape.
    try:
        await page.keyboard.press("Escape")
    except Exception:
        pass
    return False


async def _dump_boton_generar_xls(page, sel: str, etiqueta: str):
    """Vuelca el outerHTML del botón 'Generar XLS' y el action/method/target
    de su <form> contenedor (si tiene uno) — el fix de 31-07/03-08-2026
    (disparar 'change' siempre) NO resolvió el bug de raíz, y la hipótesis
    pendiente es que 'Generar XLS' no dispara un evento 'download' real de
    Playwright, igual que pasaba con 'Ver PDF' en clozapina_hce_hemogramas.py
    (tuvo que resolverse con context.request.post en vez de expect_download).
    Este volcado no requiere sesión interactiva del usuario — se genera solo
    en cualquier corrida real (AUTO_SSASUR --clozapina o el runner suelto) y
    queda listo para diagnosticar la próxima vez que falle."""
    try:
        info = await page.eval_on_selector(sel, r"""el => {
          const f = el.closest('form');
          return {
            outerHTML: el.outerHTML,
            form: f ? {action: f.action, method: f.method, target: f.target} : null,
          };
        }""")
        (MAESTRO_DIR / f"_debug_generarxls_{etiqueta}.txt").write_text(
            f"selector: {sel}\nform: {info.get('form')}\n\nouterHTML:\n{info.get('outerHTML')}\n",
            encoding="utf-8",
        )
    except Exception:
        pass


async def _click_generar_xls(page, etiqueta: str = "stockfecha"):
    """Clic en 'Generar XLS'. Espera primero a que el botón esté visible (la 1ª
    carga del form puede tardar, o el botón puede re-renderizarse tras el
    'change' de bodega/nivel_reporte) y descarta cualquier modal que lo tape."""
    # Espera a que aparezca alguno de los selectores (hasta ~20s).
    for _ in range(20):
        for sel in SELS_XLS:
            try:
                if await page.is_visible(sel):
                    await _dump_boton_generar_xls(page, sel, etiqueta)
                    await _descartar_modal(page)
                    await page.click(sel, timeout=4_000, force=True)
                    return sel
            except Exception:
                continue
        await page.wait_for_timeout(1_000)
    raise RuntimeError(f"No encontré el botón 'Generar XLS' (probé: {SELS_XLS})")


def _reporte_stock_general_reciente() -> "Path | None":
    """El reporte_de_stock_*.xlsx más reciente en la carpeta del proyecto — el
    que baja AUTO_SSASUR en el paso ABASTECIMIENTO (bodega=TODAS) para
    maestro_aa.py, justo antes de llegar a este módulo."""
    candidatos = sorted(MAESTRO_DIR.glob("reporte_de_stock_*.xlsx"),
                         key=lambda p: p.stat().st_mtime, reverse=True)
    return candidatos[0] if candidatos else None


def _parse_stock_general_xlsx(path: Path, bodega_objetivo: str) -> dict:
    """Lee el reporte de stock GENERAL (bodega=TODAS, mismo informe
    'stock_en_momento_bodega' que usa AUTO_SSASUR en el paso ABASTECIMIENTO) y
    devuelve {DESCRIPCIÓN: cantidad} solo para las filas de `bodega_objetivo`.
    Permite no pedirle a SSASUR el MISMO reporte por 2ª vez para AT Cerrada
    (bug real detectado 27-07-2026: 'Informe Stock' se bajaba dos veces en
    cada corrida — una con bodega=TODAS, otra con bodega=CERRADA — mismo
    endpoint, mismos datos, solo cambiaba el filtro). Best-effort; {} si no
    reconoce el formato o no encuentra la bodega."""
    try:
        import openpyxl
    except Exception:
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    ws = wb.worksheets[0]
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    def _norm(v):
        return str(v).strip().lower() if v is not None else ""

    col_desc = col_cant = col_bod = header_i = -1
    for i, fila in enumerate(filas[:25]):
        celdas = [_norm(c) for c in fila]
        cd = next((j for j, c in enumerate(celdas)
                   if any(k in c for k in ("descrip", "producto", "glosa", "artículo", "articulo"))), -1)
        cc = next((j for j, c in enumerate(celdas)
                   if any(k in c for k in ("cantidad", "stock", "saldo", "existencia"))), -1)
        cb = next((j for j, c in enumerate(celdas) if "bodega" in c), -1)
        if cd != -1 and cc != -1 and cb != -1:
            col_desc, col_cant, col_bod, header_i = cd, cc, cb, i
            break
    if header_i == -1:
        return {}

    objetivo_norm = " ".join(bodega_objetivo.upper().split())
    stock = {}
    for fila in filas[header_i + 1:]:
        if col_bod >= len(fila) or col_desc >= len(fila) or col_cant >= len(fila):
            continue
        bod = fila[col_bod]
        if not bod or " ".join(str(bod).upper().split()) != objetivo_norm:
            continue
        desc = fila[col_desc]
        if desc is None or not str(desc).strip():
            continue
        try:
            cant = float(fila[col_cant]) if fila[col_cant] is not None else 0.0
        except (ValueError, TypeError):
            cant = 0.0
        stock[str(desc).strip()] = int(cant) if cant == int(cant) else cant
    return stock


def _stock_desde_reporte_general(bodega: str, mapeo: dict, hoy: date) -> "dict | None":
    """Intenta servir el stock de `bodega` desde el reporte_de_stock_*.xlsx
    GENERAL ya descargado en este mismo run, en vez de descargarlo de nuevo.
    None si no hay reporte de HOY (evita reusar uno viejo en una corrida suelta
    de stock_controlados.py sin AUTO_SSASUR) o si no trae esa bodega — el
    llamador cae a la descarga en vivo de siempre."""
    ruta = _reporte_stock_general_reciente()
    if not ruta:
        return None
    if date.fromtimestamp(ruta.stat().st_mtime) != hoy:
        return None
    crudo = _parse_stock_general_xlsx(ruta, bodega)
    if not crudo:
        return None
    mapeado = _mapear(crudo, mapeo)
    if not mapeado:
        return None
    print(f"    ✓ reutilizado de {ruta.name} (bodega={bodega}, ya descargado en este run — sin pedirlo 2 veces a SSASUR)")
    return mapeado


def _parse_stock_xlsx(path: Path) -> dict:
    """Lee el xlsx del Reporte Stock en Fecha → {DESCRIPCIÓN_SSASUR: cantidad}.
    Detecta la fila de encabezado buscando una columna de nombre (descripción/
    producto/glosa/artículo) y una de cantidad (cantidad/stock/saldo/existencia).
    Best-effort; {} si no se reconoce el formato."""
    try:
        import openpyxl
    except Exception as e:
        print(f"  [ERROR] falta openpyxl: {e}")
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [ERROR] no se pudo abrir {path.name}: {e}")
        return {}
    ws = wb.worksheets[0]
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    def _norm(v):
        return str(v).strip().lower() if v is not None else ""

    col_desc = col_cant = header_i = -1
    for i, fila in enumerate(filas[:25]):
        celdas = [_norm(c) for c in fila]
        cd = next((j for j, c in enumerate(celdas)
                   if any(k in c for k in ("descrip", "producto", "glosa", "artículo", "articulo"))), -1)
        cc = next((j for j, c in enumerate(celdas)
                   if any(k in c for k in ("cantidad", "stock", "saldo", "existencia"))), -1)
        if cd != -1 and cc != -1:
            col_desc, col_cant, header_i = cd, cc, i
            break
    if header_i == -1:
        print(f"  [AVISO] {path.name}: no reconocí columnas Descripción/Cantidad.")
        return {}

    stock = {}
    for fila in filas[header_i + 1:]:
        if col_desc >= len(fila) or col_cant >= len(fila):
            continue
        desc = fila[col_desc]
        if desc is None or not str(desc).strip():
            continue
        try:
            cant = float(fila[col_cant]) if fila[col_cant] is not None else 0.0
        except (ValueError, TypeError):
            cant = 0.0
        stock[str(desc).strip()] = int(cant) if cant == int(cant) else cant
    return stock


def _mapear(stock_ssasur: dict, mapeo: dict) -> dict:
    """Traduce {DESCRIPCIÓN_SSASUR: cant} → {'nombre|presentacion': cant} usando
    el mapeo de controlados_config.json. Compara normalizado (mayúsculas, espacios
    colapsados) para tolerar dobles espacios del reporte."""
    def _n(s):
        return " ".join(str(s).upper().split())
    stock_norm = {_n(k): v for k, v in stock_ssasur.items()}
    out = {}
    for key, nombre_ssasur in mapeo.items():
        v = stock_norm.get(_n(nombre_ssasur))
        if v is not None:
            out[key] = v
    return out


def _finalizar_descarga(tmp: Path, mapeo: dict) -> dict:
    """Parsea el xlsx descargado, lo mapea a los controlados y borra el crudo."""
    print(f"    ✓ {tmp.name} ({tmp.stat().st_size // 1024:,} KB)")
    crudo = _parse_stock_xlsx(tmp)
    mapeado = _mapear(crudo, mapeo)
    print(f"    → {len(mapeado)}/{len(mapeo)} controlados mapeados (de {len(crudo)} filas)")
    try:
        tmp.unlink()   # el xlsx crudo no se versiona
    except Exception:
        pass
    return mapeado


async def _bajar_stockfecha(page, bodega: str, fecha_str: str, fid: str, mapeo: dict) -> dict:
    """AT Abierta: Reporte Stock en Fecha (stockFecha) de `fecha_str` (día anterior).
    Lanza excepción si algo falla (el llamador reintenta)."""
    await page.goto(STOCKFECHA_URL)
    await page.wait_for_load_state("networkidle")
    await page.wait_for_timeout(1_800)

    await _fijar_nivel_reporte_local(page)
    if not await _esperar_opciones_bodega(page):
        await _dump_formulario(page, f"stockFecha-{fid}-bodega-vacio")

    sel = await _seleccionar_bodega(page, bodega)
    if not sel:
        await _dump_formulario(page, f"stockFecha-{fid}")
        await page.screenshot(path=str(MAESTRO_DIR / "debug_stock_controlados.png"))
        raise RuntimeError(f"no encontré la bodega '{bodega}' en el select")
    print(f"    Bodega → {sel['label']}")
    await _descartar_modal(page)   # 1ª selección puede abrir modal Selección Proyecto
    # Elegir bodega dispara su propio 'change' (igual que nivel_reporte), que
    # puede re-renderizar el panel de botones (Mostrar Resultado/Generar XLS)
    # vía AJAX — validado 31-07-2026: con el fix de _fijar_nivel_reporte_local
    # (dispara 'change' siempre) el intento 1 pasó a fallar por "no encontré
    # el botón Generar XLS" en vez de por bodega vacía, señal de que el botón
    # aún no había terminado de re-renderizarse cuando arrancó la búsqueda.
    # Espera a que esa AJAX asiente antes de buscar el botón.
    try:
        await page.wait_for_load_state("networkidle", timeout=8_000)
    except Exception:
        pass

    ok_fecha = await _set_fecha(page, fecha_str)
    if not ok_fecha:
        print(f"    [AVISO] no pude fijar la fecha {fecha_str} — uso la del formulario.")
    await page.wait_for_timeout(600)

    tmp = SALIDA_DIR / f"_stock_{fid}_{fecha_str.replace('/', '-')}.xlsx"
    async with page.expect_download(timeout=TIMEOUT_DESCARGA_STOCKFECHA) as dl_info:
        await _click_generar_xls(page, etiqueta=f"stockfecha-{fid}")
    dl = await dl_info.value
    await dl.save_as(tmp)
    return _finalizar_descarga(tmp, mapeo)


async def _bajar_momento(page, bodega: str, fid: str, mapeo: dict) -> dict:
    """AT Cerrada: Informe Stock (stock_en_momento_bodega) — existencia ACTUAL,
    con el check 'Solo controlados'. Lanza excepción si algo falla."""
    await page.goto(STOCK_MOMENTO_URL)
    await page.wait_for_load_state("networkidle")
    await page.wait_for_timeout(1_800)

    sel = await _seleccionar_bodega(page, bodega)
    if not sel:
        await _dump_formulario(page, f"momento-{fid}")
        await page.screenshot(path=str(MAESTRO_DIR / "debug_stock_controlados.png"))
        raise RuntimeError(f"no encontré la bodega '{bodega}' en el select")
    print(f"    Bodega → {sel['label']}")
    await _descartar_modal(page)

    chk = await _marcar_solo_controlados(page)
    print(f"    Solo controlados → {chk}" if chk
          else "    [AVISO] no encontré el check 'Solo controlados' — sigo (el mapeo igual filtra).")
    await page.wait_for_timeout(400)

    tmp = SALIDA_DIR / f"_stock_{fid}_momento.xlsx"
    async with page.expect_download(timeout=TIMEOUT_DESCARGA) as dl_info:
        await _click_generar_xls(page, etiqueta=f"momento-{fid}")
    dl = await dl_info.value
    await dl.save_as(tmp)
    return _finalizar_descarga(tmp, mapeo)


async def descargar_stock_controlados(page, hoy: "date | None" = None, config: dict | None = None):
    """Descarga el stock de cada bodega de controlados y escribe
    Conteo_Controlados/stock_sistema.json. La FECHA depende de la bodega
    (config → dia_stock): AT Cerrada = día actual, AT Abierta = día anterior.
    Reutiliza `page` YA dentro del módulo ABASTECIMIENTO (el llamador hizo
    entrar_modulo(page,'ABASTECIMIENTO')). Devuelve el Path del JSON, o None."""
    config = config or cargar_config()
    hoy = hoy or date.today()
    farmacias = config.get("farmacias", [])
    bodega_ssasur = config.get("bodega_ssasur", {})
    dia_stock = config.get("dia_stock", {})
    fuente_stock = config.get("fuente_stock", {})
    mapeo = config.get("mapeo_ssasur", {})
    if not mapeo:
        print("  [ERROR] controlados_config.json sin mapeo_ssasur — abortando.")
        return None

    SALIDA_DIR.mkdir(parents=True, exist_ok=True)
    por_farmacia = {}

    for f in farmacias:
        fid = f.get("id")
        bodega = bodega_ssasur.get(fid)
        if not bodega:
            continue  # Urgencia u otra sin bodega → se ingresa a mano en la app
        dia = (dia_stock.get(fid) or "anterior").lower()
        fuente = (fuente_stock.get(fid) or "fecha").lower()
        fecha = hoy if dia == "actual" else hoy - timedelta(days=1)
        fecha_str = _fmt(fecha)
        etq_fuente = "Informe Stock (existencia actual)" if fuente == "momento" else "Reporte Stock en Fecha"
        print(f"  [{bodega}] {etq_fuente} — {fecha_str} "
              f"({'día actual' if dia == 'actual' else 'día anterior'}) ...")
        mapeado = None
        if fuente == "momento":
            # AT Cerrada pide "existencia actual" — el MISMO informe que ya
            # bajó AUTO_SSASUR con bodega=TODAS para maestro_aa.py. Reusarlo
            # evita pedirle a SSASUR el mismo reporte 2 veces por corrida.
            mapeado = _stock_desde_reporte_general(bodega, mapeo, hoy)
        for intento in (1, 2):   # la 1ª bodega puede fallar por el modal → 1 reintento
            if mapeado is not None:
                break
            try:
                if fuente == "momento":
                    mapeado = await _bajar_momento(page, bodega, fid, mapeo)
                else:
                    mapeado = await _bajar_stockfecha(page, bodega, fecha_str, fid, mapeo)
                break
            except Exception as e:
                print(f"    [intento {intento}/2] {bodega}: {e}")
                await page.wait_for_timeout(1_500)
        if mapeado is not None:
            por_farmacia[fid] = {"fecha": fecha.isoformat(), "dia": dia, "fuente": fuente, "stock": mapeado}
        else:
            try:
                await page.screenshot(path=str(MAESTRO_DIR / "debug_stock_controlados.png"))
            except Exception:
                # La página ya puede estar cerrada (mismo error que hizo fallar
                # los 2 intentos) — si el screenshot también revienta sin este
                # try/except, la excepción se propaga fuera de la función y
                # las bodegas que SÍ se descargaron bien en este loop (ej.
                # Cerrada) se pierden entero: nunca se llega a escribir
                # stock_sistema.json. Bug real detectado 27-07-2026: Abierta
                # falló 2/2, el screenshot también falló por página cerrada,
                # y el JSON quedó con la fecha del día anterior.
                pass

    if not por_farmacia:
        print("  [AVISO] No se obtuvo stock de ninguna bodega — no se escribe el JSON.")
        return None

    payload = {
        "generadoEn": datetime.now().isoformat(timespec="seconds"),
        "porFarmacia": por_farmacia,
    }
    SALIDA_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    total = sum(len(v["stock"]) for v in por_farmacia.values())
    resumen = ", ".join(f"{k} ({v['fecha']})" for k, v in por_farmacia.items())
    print(f"  ✓ {SALIDA_JSON.name} — {total} valores · {resumen}")
    return SALIDA_JSON


# ── Runner suelto (para probar/backfill sin AUTO_SSASUR) ────────────────────────
async def _standalone(hoy: date):
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("[ERROR] Playwright no está instalado. Ejecuta AUTO_SSASUR.bat una vez.")
        return
    # Reutiliza el ingreso al dashboard/módulo de AUTO_SSASUR (importar NO corre su main()).
    from AUTO_SSASUR import entrar_modulo, DASHBOARD_URL, SESSION_FILE

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
        )
        if SESSION_FILE.exists():
            context = await browser.new_context(accept_downloads=True, storage_state=str(SESSION_FILE))
        else:
            context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()
        print("Logéate en SSASUR si hace falta (tienes 5 min)...")
        await page.goto(DASHBOARD_URL)
        try:
            await page.wait_for_selector(
                'button:has-text("ABASTECIMIENTO"), div:has-text("ABASTECIMIENTO")',
                timeout=300_000)
        except Exception:
            print("[ERROR] No se detectó el dashboard — ¿login incompleto?")
            await browser.close()
            return
        await context.storage_state(path=str(SESSION_FILE))
        await entrar_modulo(page, "ABASTECIMIENTO")
        await descargar_stock_controlados(page, hoy)
        await browser.close()


if __name__ == "__main__":
    arg = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
    if arg:   # override de "hoy" para backfill: dd/mm/aaaa
        d, m, y = arg.split("/")
        hoy = date(int(y), int(m), int(d))
    else:
        hoy = date.today()
    print(f"Stock de controlados — hoy = {_fmt(hoy)}  "
          f"(Cerrada = día actual, Abierta = día anterior)")
    asyncio.run(_standalone(hoy))
