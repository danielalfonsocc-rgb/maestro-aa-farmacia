#!/usr/bin/env python3
"""
limpiar_gt_maestro.py — Limpia recetas obsoletas de una hoja de mes del
maestro GT, aplicando las 3 reglas identificadas al corregir JULIO 2026
(23-07-2026) contra una copia revisada a mano por el usuario:

  1. BUG: fila de encabezado ("N° de receta"/"Paciente"/...) embebida como
     si fuera una receta más — corrupción de datos, se borra siempre.
  2. YA RESUELTA: la receta figura ENTREGADA / CERRADA-INCOMPLETA / ANULADA
     en el informe_completo_recetas*.csv más reciente — ya no necesita
     seguimiento GT activo.
  3. DUPLICADO POR RE-DIGITACIÓN: mismo RUT, mismo N° de receta ± 1 que otra
     receta que SÍ sigue en la hoja — SSASUR a veces digita la misma cuota
     dos veces con números consecutivos. Se borra la de número MÁS ALTO
     (la redigitada), se deja la más baja (la original).

Lo que NO se borra automáticamente: cualquier receta que siga PENDIENTE en
el CSV y no tenga una hermana consecutiva del mismo RUT en la hoja — no hay
forma de saber por los datos si sigue vigente o quedó obsoleta (paciente
trasladado, tratamiento discontinuado, etc.). Esas quedan en un
Feedback_Limpieza_GT_<MES><AÑO>_<fecha>.xlsx para revisión manual en SSASUR
— NO se tocan en el maestro.

Uso:
  py limpiar_gt_maestro.py --mes "JUNIO 2026" --dry-run
  py limpiar_gt_maestro.py --mes "JUNIO 2026"
"""
import argparse
import csv
import datetime
import glob
import hashlib
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import gt_maestro as GM

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
ESTADOS_RESUELTOS = {"ENTREGADA", "CERRADA / INCOMPLETA", "ANULADA"}


def _estado_actual_csv():
    estado = {}
    for path in sorted(glob.glob(os.path.join(MAESTRO_DIR, "informe_completo_recetas*.csv"))):
        with open(path, encoding="latin-1", newline="") as f:
            for row in csv.DictReader(f, delimiter=";"):
                n = (row.get("Número Receta") or "").strip()
                if n and n not in estado:
                    estado[n] = row.get("Estado")
    return estado


def _rut_hash(rut):
    return hashlib.sha256(str(rut).encode()).hexdigest()[:8] if rut else None


def analizar(mes):
    wb, path = GM.cargar_maestro()
    if mes not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{mes}' en el maestro.")
    ws = wb[mes]

    filas = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[0] is not None and str(row[0]).strip():
            filas.append((i, row))

    por_rut = {}
    for i, row in filas:
        h = _rut_hash(row[2])
        if h:
            por_rut.setdefault(h, []).append(str(row[0]).strip())

    estado_csv = _estado_actual_csv()

    bug, resueltas, duplicadas, ambiguas = [], [], [], []
    for i, row in filas:
        receta = str(row[0]).strip()
        if row[0] == "N° de receta" or row[1] == "Paciente":
            bug.append((i, row))
            continue
        estado = estado_csv.get(receta)
        if estado in ESTADOS_RESUELTOS:
            resueltas.append((i, row, estado))
            continue
        if receta.isdigit():
            h = _rut_hash(row[2])
            hermanos = [r for r in por_rut.get(h, []) if r != receta]
            n = int(receta)
            hermano_menor = next((r for r in hermanos if r.isdigit() and int(r) == n - 1), None)
            if hermano_menor:
                duplicadas.append((i, row, hermano_menor))
                continue
        ambiguas.append((i, row, estado))

    return wb, path, ws, bug, resueltas, duplicadas, ambiguas


def _escribir_feedback(mes, ambiguas, fecha_hoy):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Limpieza GT {mes} — recetas PENDIENTE confirmado sin regla automática (revisar en SSASUR)"
    ws["A1"].font = Font(bold=True, size=13)
    headers = ["N° Receta", "Paciente", "RUT", "Especialidad", "Período", "Estado en el maestro"]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
    # Solo entran las que tienen estado CONFIRMADO en el CSV (típicamente
    # PENDIENTE) — las "sin dato" (receta no encontrada en mi CSV local, ver
    # nota en main()) NO se listan acá: no significa que estén obsoletas,
    # solo que no tengo visibilidad, y listarlas igual sería engañoso.
    confirmadas = [a for a in ambiguas if a[2] is not None]
    for _, row, estado in confirmadas:
        ws.append([row[0], row[1], row[2], row[7], row[6], estado])
    anchos = [12, 30, 14, 22, 10, 26]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    if not confirmadas:
        return None, confirmadas
    mes_compacto = mes.replace(" ", "")
    ruta = os.path.join(os.path.dirname(MAESTRO_DIR), f"Feedback_Limpieza_GT_{mes_compacto}_{fecha_hoy}.xlsx")
    wb.save(ruta)
    return ruta, confirmadas


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mes", required=True, help='Hoja de mes a limpiar, ej. "JUNIO 2026"')
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    wb, path, ws, bug, resueltas, duplicadas, ambiguas = analizar(a.mes)

    print(f"[{a.mes}] Bug (encabezado embebido): {len(bug)}")
    print(f"[{a.mes}] Ya resueltas (ENTREGADA/CERRADA/ANULADA): {len(resueltas)}")
    for _, row, estado in resueltas:
        print(f"  {row[0]} — {estado}")
    print(f"[{a.mes}] Duplicado por re-digitación (N+1, mismo RUT): {len(duplicadas)}")
    for _, row, hermano in duplicadas:
        print(f"  {row[0]} (se deja {hermano})")

    sin_dato = [x for x in ambiguas if x[2] is None]
    confirmadas = [x for x in ambiguas if x[2] is not None]
    print(f"[{a.mes}] Ambiguas con estado CONFIRMADO en CSV (van a feedback): {len(confirmadas)}")
    for _, row, estado in confirmadas:
        print(f"  {row[0]} — estado: {estado}")
    if sin_dato:
        print(f"[{a.mes}] [AVISO] {len(sin_dato)} receta(s) más NO aparecen en mi CSV local en absoluto "
              f"— vacío de cobertura de datos (mes fuera de los bloques descargados, no necesariamente "
              f"obsoletas). NO se tocan ni se reportan como \"para revisar\": no hay base para juzgarlas.")

    # Las "ambiguas" NUNCA se borran automáticamente — julio fue un caso
    # especial donde el usuario ya las revisó una por una y aprobó sacarlas
    # todas. Acá solo se reportan (las que sí tienen estado confirmado) para
    # revisión manual; TODAS las ambiguas quedan intactas en el maestro.
    a_borrar = bug + resueltas + duplicadas
    if a.dry_run:
        print(f"\n[DRY-RUN] Se borrarían {len(a_borrar)} fila(s) de {a.mes} (regla clara). "
              f"Ninguna ambigua se toca. No se modifica nada.")
        return

    for i, *_ in sorted(a_borrar, key=lambda t: t[0], reverse=True):
        ws.delete_rows(i)
    GM.aplicar_formato_maestro(ws)
    GM.guardar(wb, path)
    print(f"\nGuardado: {path}  ({len(a_borrar)} fila(s) borrada(s))")

    if ambiguas:
        fecha_hoy = datetime.date.today().strftime("%Y-%m-%d")
        ruta_fb, listadas = _escribir_feedback(a.mes, ambiguas, fecha_hoy)
        if ruta_fb:
            print(f"Feedback para revisión manual ({len(listadas)} receta(s) con estado confirmado): {ruta_fb}")
        else:
            print("Sin recetas con estado confirmado para reportar — no se generó feedback.")


if __name__ == "__main__":
    main()
