#!/usr/bin/env python3
"""
Generador de Gestión Territorial - Farmacia Hospital de Pitrufquén.

Lee un JSON de registros (ya transcritos y revisados) y produce, POR CADA
establecimiento de destino:
  - <Destino>_Planilla.xlsx  (hoja "Maestra" + hoja "Funcionarios") y su PDF
  - <Destino>_Letrero.xlsx   (etiqueta carta recortable) y su PDF

Uso:
  python generar.py registros.json --salida ./out [--no-pdf]

Estructura del JSON de entrada (ver references/formato.md y ejemplo/):
{
  "fecha_entrega": "26/05/2026",
  "origen": "Farmacia Hospital de Pitrufquén",
  "registros": [
    {"receta","paciente","run","edad","direccion","comuna","telefono",
     "periodo","especialidad","n_presc","refrigerado","pendiente",
     "ventanilla","estab_destino"}, ...
  ]
}
"""
import argparse, json, os, re, subprocess, sys, unicodedata
from datetime import date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

for _s in (sys.stdout, sys.stderr):
    try: _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception: pass

HERE = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(HERE, "..", "assets")
LOGO_SS = os.path.join(ASSETS, "logo_ss_araucania_sur.jpg")
LOGO_HOSP = os.path.join(ASSETS, "logo_hospital_pitrufquen.png")

NAVY="1F3864"; BLUE="2E5496"; LIGHT="D9E1F2"; VENT="FCE4D6"; GREY="F2F2F2"; AMBER="FFE699"
ORANGE="C55A11"; GREEN="1F6F3D"; RED="C00000"
thin=Side(style="thin", color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
V_LABEL="RETIRO EN VENTANILLA"; D_LABEL="GESTIÓN TERRITORIAL"

def slug(s):
    s=unicodedata.normalize("NFKD",s).encode("ascii","ignore").decode()
    return re.sub(r"[^A-Za-z0-9]+","_",s).strip("_")

def fmt_run(run):
    run=re.sub(r"\s+","", str(run or "")).replace(".","")
    if "-" not in run: return run
    body,dv=run.rsplit("-",1)
    try: body=f"{int(body):,}".replace(",",".")
    except ValueError: pass
    return f"{body}-{dv}"

def dv_correcto(run):
    """Valida el dígito verificador chileno (módulo 11). Devuelve True/False, o None si no parsea."""
    r=re.sub(r"\s+","", str(run or "")).replace(".","").replace("-","").upper()
    if len(r)<2 or not r[:-1].isdigit(): return None
    cuerpo,dv=r[:-1],r[-1]; suma=0; mult=2
    for d in reversed(cuerpo):
        suma+=int(d)*mult; mult=2 if mult==7 else mult+1
    resto=11-(suma%11); calc="0" if resto==11 else ("K" if resto==10 else str(resto))
    return calc==dv

def es_ultima(per):
    try:
        n,m=str(per).split("/"); return int(n)==int(m)
    except Exception: return False

# ---- inserción de logos centrados verticalmente, con margen interior ----
def place_logo(ws, path, col_letter, target_h_px, col_off_px, row_off_px, from_row0=0):
    if not path or not os.path.exists(path): return False
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image as PILImage
    img=XLImage(path); w,h=PILImage.open(path).size
    tw=int(w*(target_h_px/h)); th=int(target_h_px)
    img.width=tw; img.height=th
    col0=column_index_from_string(col_letter)-1
    img.anchor=OneCellAnchor(_from=AnchorMarker(col=col0, colOff=pixels_to_EMU(col_off_px),
                             row=from_row0, rowOff=pixels_to_EMU(row_off_px)),
                             ext=XDRPositiveSize2D(pixels_to_EMU(tw), pixels_to_EMU(th)))
    ws.add_image(img); return True

def banda_encabezado(ws, ncols, titulo, subtitulo, h_logo_ss=62, h_logo_hosp=46, off_hosp=70, roff_hosp=22):
    last=get_column_letter(ncols)
    ws.row_dimensions[1].height=34; ws.row_dimensions[2].height=34
    ws.merge_cells("A1:B2"); ws.merge_cells(f"{get_column_letter(ncols-1)}1:{last}2")
    tf=get_column_letter(3); tl=get_column_letter(ncols-2)
    ws.merge_cells(f"{tf}1:{tl}2"); t=ws[f"{tf}1"]; t.value=titulo
    t.font=Font(name="Calibri",size=18,bold=True,color="FFFFFF"); t.alignment=Alignment(horizontal="center",vertical="center")
    for rr in (1,2):
        for cc in range(1,ncols+1): ws.cell(row=rr,column=cc).fill=PatternFill("solid",fgColor=NAVY)
    if not place_logo(ws, LOGO_SS, "A", h_logo_ss, 12, 14):
        c=ws["A1"]; c.value="[ LOGO SS ]"; c.font=Font(size=9,bold=True,color="FFFFFF"); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    if not place_logo(ws, LOGO_HOSP, get_column_letter(ncols-1), h_logo_hosp, off_hosp, roff_hosp):
        c=ws.cell(row=1,column=ncols-1); c.value="[ LOGO HOSP ]"; c.font=Font(size=9,bold=True,color="FFFFFF"); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.merge_cells(f"A3:{last}3"); s=ws["A3"]; s.value=subtitulo
    s.font=Font(name="Calibri",size=11,bold=True,color=NAVY); s.alignment=Alignment(horizontal="center",vertical="center")
    s.fill=PatternFill("solid",fgColor=LIGHT); ws.row_dimensions[3].height=20

def bloque_qf(ws, fila_inicio, ancho_label="A:B", ancho_linea="C:F", sz=10):
    ws.merge_cells(f"A{fila_inicio}:F{fila_inicio}")
    h=ws.cell(row=fila_inicio,column=1,value="VALIDACIÓN QUÍMICO FARMACÉUTICO")
    h.font=Font(name="Calibri",size=sz+1,bold=True,color="FFFFFF"); h.fill=PatternFill("solid",fgColor=NAVY)
    h.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[fila_inicio].height=22
    rr=fila_inicio+2
    la,lb=ancho_label.split(":"); ca,cb=ancho_linea.split(":")
    for lab in ("Nombre QF que valida:","RUT:","Firma:"):
        ws.merge_cells(f"{la}{rr}:{lb}{rr}")
        lc=ws.cell(row=rr,column=1,value=lab); lc.font=Font(name="Calibri",size=sz,bold=True); lc.alignment=Alignment(horizontal="left",vertical="center")
        ws.merge_cells(f"{ca}{rr}:{cb}{rr}")
        for cc in range(column_index_from_string(ca),column_index_from_string(cb)+1):
            ws.cell(row=rr,column=cc).border=Border(bottom=Side(style="thin",color="000000"))
        ws.row_dimensions[rr].height=26 if lab!="Firma:" else 46; rr+=2
    ws.cell(row=rr+1,column=1,value="Fecha de validación: ____/____/________").font=Font(name="Calibri",size=sz)

def carta_landscape(ws, fit_height, print_area=None):
    """Configura la hoja para impresión carta horizontal (landscape).
    fit_height=True  → fitToHeight=1 (contenido comprimido en 1 página vertical)
    fit_height=False → fitToHeight=0 (se expande en varias páginas si hay muchas filas)
    Para las planillas GT siempre usamos fit_height=True para garantizar 1 página carta."""
    ws.page_setup.paperSize=1; ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1  # SIEMPRE 1 página carta horizontal
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.print_options.horizontalCentered=True; ws.print_options.verticalCentered=False
    ws.page_margins.left=0.25; ws.page_margins.right=0.25
    ws.page_margins.top=0.35; ws.page_margins.bottom=0.35
    ws.page_margins.header=0.15; ws.page_margins.footer=0.15
    if print_area: ws.print_area=print_area

# ----------------------------- PLANILLA -----------------------------

def hoja_funcionarios(wb, regs, destino, titulo, subtitulo, modo="normal"):
    ws=wb.active; ws.title="Funcionarios"  # única hoja — sin Maestra
    col_label="Controlados" if modo=="controlados" else "Refrigerados"
    col_color=RED if modo=="controlados" else GREEN
    # ✓ Revisado como col A, a la izquierda de N° Receta
    h2=["✓ Revisado","N° Receta","Paciente","RUN","Especialidad","Período","Estado Receta","N° Presc.",col_label,"Modalidad de Entrega","Pendiente"]
    w2=[6,10,30,13,20,8,13,7,22,18,16]; C2={"✓ Revisado","N° Receta","RUN","Período","Estado Receta","N° Presc."}
    last=get_column_letter(len(h2)); COL={h:i+1 for i,h in enumerate(h2)}
    for c,w in enumerate(w2,start=1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=40; ws.row_dimensions[2].height=40
    # Logos: SS→A-B, Hosp→J-K, Título→C-I  (11 cols)
    ws.merge_cells("A1:B2"); ws.merge_cells(f"{get_column_letter(len(h2)-1)}1:{last}2")
    ws.merge_cells(f"C1:{get_column_letter(len(h2)-2)}2")
    tt=ws["C1"]; tt.value=titulo; tt.font=Font(name="Calibri",size=20,bold=True,color="FFFFFF"); tt.alignment=Alignment(horizontal="center",vertical="center")
    for rr in (1,2):
        for cc in range(1,len(h2)+1): ws.cell(row=rr,column=cc).fill=PatternFill("solid",fgColor=NAVY)
    place_logo(ws, LOGO_SS, "A", 68, 12, 16) or ws["A1"].__setattr__("value","[ LOGO SS ]")
    place_logo(ws, LOGO_HOSP, get_column_letter(len(h2)-1), 50, 60, 28) or ws.cell(row=1,column=len(h2)-1).__setattr__("value","[ LOGO HOSP ]")
    ws.merge_cells(f"A3:{last}3"); s=ws["A3"]; s.value=subtitulo
    s.font=Font(name="Calibri",size=12,bold=True,color=NAVY); s.alignment=Alignment(horizontal="center",vertical="center")
    s.fill=PatternFill("solid",fgColor=LIGHT); ws.row_dimensions[3].height=24
    HR=5
    # NARROW = columnas cuyo ancho no alcanza para el texto del encabezado en 1 línea
    NARROW={"N° Presc.","Período"}  # usan shrink_to_fit para evitar cortes
    for c,h in enumerate(h2,start=1):
        # ✓ Revisado → sólo el símbolo (col demasiado estrecha para el texto completo)
        label="✓" if h=="✓ Revisado" else h
        sz=16 if h=="✓ Revisado" else 11
        al=(Alignment(horizontal="center",vertical="center",shrink_to_fit=True)
            if h in NARROW
            else Alignment(horizontal="center",vertical="center",wrap_text=True))
        cell=ws.cell(row=HR,column=c,value=label)
        cell.font=Font(name="Calibri",size=sz,bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor=BLUE); cell.alignment=al; cell.border=border
    ws.row_dimensions[HR].height=44  # espacio para 2 líneas en cols que hacen wrap
    r=HR+1
    regs=sorted(regs, key=lambda g: unicodedata.normalize("NFKD",g["paciente"]).encode("ascii","ignore").decode().upper())
    chk_border=Border(left=Side(style="medium",color="000000"),right=Side(style="medium",color="000000"),
                      top=Side(style="medium",color="000000"),bottom=Side(style="medium",color="000000"))
    for g in regs:
        est="ÚLTIMA RECETA" if es_ultima(g["periodo"]) else ""; is_v=bool(g.get("ventanilla"))
        refri=g.get("controlado","") or "" if modo=="controlados" else g.get("refrigerado","") or ""
        vals=[None,g["receta"],g["paciente"],fmt_run(g["run"]),g["especialidad"],g["periodo"],est,g["n_presc"],refri,
              V_LABEL if is_v else D_LABEL,g.get("pendiente","") or ""]
        for c,v in enumerate(vals,start=1):
            h=h2[c-1]; cell=ws.cell(row=r,column=c,value=v if v not in ("",None) else None)
            run_bad=(h=="RUN" and dv_correcto(g["run"]) is False)
            bold=h in ("N° Receta","Paciente") or (h=="Modalidad de Entrega" and is_v) or (h==col_label and refri) or (h=="Estado Receta" and est) or run_bad
            color=RED if ((h=="Modalidad de Entrega" and is_v) or run_bad) else (col_color if (h==col_label and refri) else (ORANGE if (h=="Estado Receta" and est) else "000000"))
            cell.font=Font(name="Calibri",size=11,bold=bold,color=color)
            cell.border=chk_border if h=="✓ Revisado" else border
            cell.alignment=Alignment(horizontal=("center" if h in C2 else "left"),vertical="center",wrap_text=True)
        fill=VENT if is_v else (GREY if r%2==0 else "FFFFFF")
        for c in range(1,len(h2)+1): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=fill)
        if est: ws.cell(row=r,column=COL["Estado Receta"]).fill=PatternFill("solid",fgColor=AMBER)
        ws.row_dimensions[r].height=26; r+=1
    DE=r-1; npc=COL["N° Presc."]
    ws.cell(row=r,column=COL["N° Receta"],value="TOTAL").font=Font(size=11,bold=True)
    ws.cell(row=r,column=COL["Paciente"],value=f"{len(regs)} pacientes / recetas").font=Font(size=11,bold=True)
    ws.cell(row=r,column=npc,value=sum(int(g["n_presc"]) for g in regs)).font=Font(size=11,bold=True)
    ws.cell(row=r,column=npc).alignment=Alignment(horizontal="center")
    for c in range(1,len(h2)+1): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=LIGHT); ws.cell(row=r,column=c).border=border
    qf=r+3; bloque_qf(ws, qf, sz=11)
    ws.freeze_panes=f"A{HR+1}"; ws.auto_filter.ref=f"A{HR}:{last}{DE}"; ws.print_title_rows=f"1:{HR}"
    carta_landscape(ws, fit_height=False, print_area=f"A1:{last}{qf+10}")
    ws.evenFooter.right.text=ws.oddFooter.right.text="Página &P de &N"

# ----------------------------- LETRERO -----------------------------
def _bloque_letrero(ws, r0, destino, lleva_refri, NCOL):
    """1 etiqueta = media hoja carta vertical.
    Alturas: 15+76+36+28+16+5×40+32 = 383 pts ≈ 5.3" ≈ mitad de carta.
    Fecha siempre = date.today() (día de uso del skill)."""
    thick=Side(style="thick",color="000000"); box=Border(left=thick,right=thick,top=thick,bottom=thick)
    L=get_column_letter(2); R=get_column_letter(NCOL-1)
    # Banda de logos — 76 pts de alto
    ws.row_dimensions[r0].height=76
    place_logo(ws, LOGO_SS,  "B", 72, 6, 4, from_row0=r0-1)
    place_logo(ws, LOGO_HOSP, get_column_letter(NCOL-2), 60, 4, 8, from_row0=r0-1)
    # Títulos
    ws.merge_cells(f"{L}{r0+1}:{R}{r0+1}"); a=ws[f"{L}{r0+1}"]; a.value="FARMACIA HOSPITAL DE PITRUFQUÉN"
    a.font=Font(name="Calibri",size=19,bold=True); a.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r0+1].height=36
    ws.merge_cells(f"{L}{r0+2}:{R}{r0+2}"); b=ws[f"{L}{r0+2}"]; b.value=f"GESTIÓN TERRITORIAL — {destino.upper()}"
    b.font=Font(name="Calibri",size=13,bold=True,color=NAVY); b.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r0+2].height=28
    ws.row_dimensions[r0+3].height=16  # separador
    # Bloque PARA: 5 filas × 40 pts = 200 pts — centro visual dominante
    ws.merge_cells(f"{L}{r0+4}:{R}{r0+8}"); p=ws[f"{L}{r0+4}"]; p.value=f"PARA:\n{destino.upper()}"
    p.font=Font(name="Calibri",size=36,bold=True); p.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for rr in range(r0+4,r0+9):
        ws.row_dimensions[rr].height=40
        for cc in range(2,NCOL): ws.cell(row=rr,column=cc).border=box
    # Fecha de envío (= fecha del día de uso del skill)
    ws.merge_cells(f"{L}{r0+9}:{R}{r0+9}"); f=ws[f"{L}{r0+9}"]; f.value=f"Fecha de envío: {FECHA}"
    f.font=Font(name="Calibri",size=13,bold=True); f.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r0+9].height=32
    # Total: ~15+76+36+28+16+200+32 = 403 pts ≈ media carta carta (printable 367 pts + margen superior)
    last=r0+9
    if lleva_refri:
        c1,c2=4,NCOL-3
        ws.merge_cells(f"{get_column_letter(c1)}{r0+11}:{get_column_letter(c2)}{r0+12}")
        r=ws.cell(row=r0+11,column=c1,value="❄  REFRIGERADO")
        r.font=Font(name="Calibri",size=22,bold=True,color="FFFFFF"); r.alignment=Alignment(horizontal="center",vertical="center")
        fill=PatternFill("solid",fgColor="2E75B6")
        for rr in range(r0+11,r0+13):
            ws.row_dimensions[rr].height=32
            for cc in range(c1,c2+1): ws.cell(row=rr,column=cc).border=box; ws.cell(row=rr,column=cc).fill=fill
        last=r0+12
    return last

def letrero(destino, lleva_refri, ruta):
    """1 etiqueta que ocupa media hoja carta vertical. Fecha = hoy (date.today)."""
    wb=Workbook(); ws=wb.active; ws.title="LETRERO"
    NCOL=11
    for c in range(1,NCOL+1): ws.column_dimensions[get_column_letter(c)].width=8.5
    last=get_column_letter(NCOL)
    fin=_bloque_letrero(ws, 2, destino, lleva_refri, NCOL)
    # Carta vertical: fitToWidth=1 escala al ancho carta; fitToHeight=0 conserva altura natural
    # Las alturas de fila en _bloque_letrero suman ~375 pts ≈ media hoja carta (5.2")
    ws.page_setup.paperSize=1; ws.page_setup.orientation="portrait"
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.print_options.horizontalCentered=True
    ws.page_margins.left=0.3; ws.page_margins.right=0.3
    ws.page_margins.top=0.35; ws.page_margins.bottom=0.35
    ws.page_margins.header=0.1; ws.page_margins.footer=0.1
    ws.print_area=f"A1:{last}{fin+1}"
    wb.save(ruta)


# ----------------------------- VERIFICACIÓN E IMPORTACIÓN -----------------------------
def norm_run(run):
    return re.sub(r"\s+","", str(run or "")).replace(".","").replace("-","").upper()

def norm_nombre(n):
    import unicodedata as u
    n=u.normalize("NFKD",str(n or "")).encode("ascii","ignore").decode().upper()
    return re.sub(r"\s+"," ",n).strip()

def _key(h):
    import unicodedata as u
    h=u.normalize("NFKD",str(h or "")).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]","",h)

COLS_RUN={"run","rut","runpaciente","rutpaciente","numerodocumento","ndocumento","nrodocumento","documento","identificacion"}
COLS_DV={"dv","digitoverificador","digito"}
COLS_NOMBRE_COMPLETO={"nombrecompleto","paciente","nombresyapellidos","nombreapellido","nombrepaciente"}
COLS_NOMBRES={"nombre","nombres","primernombre","nombre1","nombrepac"}
COLS_AP_PAT={"apellidopaterno","appaterno","apellido1","apaterno","appat"}
COLS_AP_MAT={"apellidomaterno","apmaterno","apellido2","apematerno","apmat"}

def _arma_nombre(fila):
    nombres=appat=apmat=completo=""
    for k,v in fila.items():
        kk=_key(k)
        if kk in COLS_AP_PAT and not appat: appat=v or ""
        elif kk in COLS_AP_MAT and not apmat: apmat=v or ""
        elif kk in COLS_NOMBRES and not nombres: nombres=v or ""
        elif kk in COLS_NOMBRE_COMPLETO and not completo: completo=v or ""
    if appat or apmat:
        return re.sub(r"\s+"," ", f"{nombres} {appat} {apmat}").strip()
    completo=(completo or nombres or "").strip()
    if "," in completo:
        ap,no=completo.split(",",1); completo=f"{no.strip()} {ap.strip()}"
    return completo

def _arma_run(fila):
    run=dv=""
    for k,v in fila.items():
        kk=_key(k)
        if kk in COLS_RUN and not run: run=v or ""
        elif kk in COLS_DV and not dv: dv=v or ""
    run=re.sub(r"\s+","", str(run))
    if dv and "-" not in run: run=f"{run}-{dv}"
    return run

def _leer_csv(path):
    import csv
    for enc in ("utf-8-sig","latin-1"):
        try: raw=open(path, encoding=enc).read()
        except Exception: continue
        try:
            delim=csv.Sniffer().sniff(raw[:4096], delimiters=";,\t|").delimiter
        except Exception:
            delim=";" if raw[:4096].count(";")>=raw[:4096].count(",") else ","
        return list(csv.DictReader(raw.splitlines(), delimiter=delim))
    return []

def cargar_historico(path):
    idx={}
    if not path or not os.path.exists(path): return idx
    if path.lower().endswith(".json"):
        d=json.load(open(path, encoding="utf-8"))
        filas=d.get("pacientes", d if isinstance(d,list) else d.get("registros",[]))
        for r in filas:
            run=norm_run(r.get("run") or r.get("rut") or ""); nom=norm_nombre(r.get("nombre") or r.get("paciente") or "")
            if run and nom: idx.setdefault(run,set()).add(nom)
        return idx
    for fila in _leer_csv(path):
        run=norm_run(_arma_run(fila)); nom=norm_nombre(_arma_nombre(fila))
        if run and nom: idx.setdefault(run,set()).add(nom)
    return idx

def _coincide_nombre(n_planilla, nombres_historico):
    a=set(n_planilla.split())
    if not a: return False
    for h in nombres_historico:
        b=set(h.split())
        if not b: continue
        menor,mayor=(a,b) if len(a)<=len(b) else (b,a)
        if menor and menor<=mayor and len(menor)>=2: return True
        if a==b: return True
    return False

TIPOS_INSULINA=[("GLARGINA","Glargina"),("ASPART","Asparta"),("GLULISINA","Glulisina"),
                ("CRISTALINA","Cristalina"),("LISPRO","Lispro"),("DEGLUDEC","Degludec"),
                ("DETEMIR","Detemir"),("NPH","NPH")]

def _insulina_label(prod):
    """Devuelve 'Insulina <Tipo>' si el producto es una insulina (refrigerado); None si no.
       Excluye dispositivos (jeringas/agujas) que solo mencionan 'insulina'."""
    u=str(prod or "").upper()
    for kw,lab in TIPOS_INSULINA:
        if kw in u: return f"Insulina {lab}"
    if "INSULINA" in u and "JERINGA" not in u and "AGUJA" not in u:
        return "Insulina"
    return None

def cargar_sidra(path):
    """Lee el SIDRA una sola vez y devuelve (idx_identidad, mapa_refrigerados).
       idx_identidad: run_normalizado -> set(nombres).  mapa_refrigerados: n_receta -> 'Insulina X, Insulina Y'."""
    idx={}; refri={}
    if not path or not os.path.exists(path): return idx, refri
    if path.lower().endswith(".json"):
        return cargar_historico(path), {}
    REC={"numeroreceta","nreceta","nroreceta","receta"}
    PRES={"prescripcion","producto","medicamento"}
    for fila in _leer_csv(path):
        run=norm_run(_arma_run(fila)); nom=norm_nombre(_arma_nombre(fila))
        if run and nom: idx.setdefault(run,set()).add(nom)
        rec=None; presc=None
        for k,v in fila.items():
            kk=_key(k)
            if rec is None and kk in REC: rec=str(v or "").strip()
            elif presc is None and kk in PRES: presc=str(v or "")
        if rec and presc:
            lab=_insulina_label(presc)
            if lab: refri.setdefault(rec,set()).add(lab)
    return idx, {k:", ".join(sorted(v)) for k,v in refri.items()}

def verificar_historico(regs, idx):
    out=[]
    for g in regs:
        run=norm_run(g["run"]); nom=norm_nombre(g["paciente"])
        if run in idx:
            estado="COINCIDE" if _coincide_nombre(nom, idx[run]) else "NOMBRE DISTINTO"
            previos="; ".join(sorted(idx[run]))
        else:
            estado="NUEVO"; previos=""
        out.append((g, estado, previos))
    return out

def reporte_verificacion(resultados, destino, ruta):
    wb=Workbook(); ws=wb.active; ws.title="Verificación"
    cols=["N° Receta","Paciente","RUN","Dígito Verificador","Histórico (RUT/nombre)","Nombre(s) en histórico"]
    widths=[12,38,16,18,22,40]
    for c,w in enumerate(widths,start=1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1"); t=ws["A1"]
    t.value=f"VERIFICACIÓN DE IDENTIDAD - {destino}"; t.font=Font(size=14,bold=True,color="FFFFFF")
    t.fill=PatternFill("solid",fgColor=NAVY); t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=26
    for c,h in enumerate(cols,start=1):
        cell=ws.cell(row=2,column=c,value=h); cell.font=Font(size=10,bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor=BLUE); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
    r=3
    for g,estado,previos in resultados:
        dv=dv_correcto(g["run"]); dv_txt="Válido" if dv else ("INVÁLIDO" if dv is False else "No validable")
        vals=[g["receta"],g["paciente"],fmt_run(g["run"]),dv_txt,(estado or "(sin histórico)"),previos]
        for c,v in enumerate(vals,start=1):
            cell=ws.cell(row=r,column=c,value=v); cell.border=border
            cell.alignment=Alignment(horizontal=("left" if c in (2,6) else "center"),vertical="center",wrap_text=True); cell.font=Font(size=9)
        if dv is False or estado=="NOMBRE DISTINTO":
            for c in range(1,len(cols)+1): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor="F8CBAD")
        elif estado=="NUEVO":
            for c in range(1,len(cols)+1): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor="FFF2CC")
        r+=1
    ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:{get_column_letter(len(cols))}{r-1}"
    carta_landscape(ws, fit_height=False)
    wb.save(ruta)

def importar_reporte_xlsx(path):
    from openpyxl import load_workbook
    from collections import OrderedDict
    INS=("INSULINA","GLARGINA","ASPART","CRISTALINA","LISPRO","NPH","DETEMIR","GLULISINA","DEGLUDEC")
    wb=load_workbook(path, read_only=True, data_only=True)
    ws=wb[wb.sheetnames[0]]; rows=[r for r in ws.iter_rows(values_only=True)]
    hi=next(i for i,r in enumerate(rows) if sum(1 for c in r if c not in (None,""))>=5)
    hdr=[str(c).strip() if c is not None else "" for c in rows[hi]]; idx={h:i for i,h in enumerate(hdr)}
    def C(r,n):
        i=idx.get(n); return r[i] if (i is not None and i<len(r)) else None
    dd=OrderedDict(); origenes={}
    for r in rows[hi+1:]:
        if not r or C(r,"N° Receta") in (None,""): continue
        rec=str(C(r,"N° Receta")).strip(); prod=str(C(r,"Producto") or "")
        refri = prod.strip() if any(k in prod.upper() for k in INS) else ""
        org=str(C(r,"Estab. Origen") or "").strip(); origenes[org]=origenes.get(org,0)+1
        if rec not in dd:
            edad=re.sub(r"\D","", str(C(r,"Edad") or "")) or None
            dd[rec]={"receta":rec,"paciente":str(C(r,"Paciente") or "").strip(),
                "run":str(C(r,"Run Paciente") or "").strip(),"edad":int(edad) if edad else None,
                "direccion":str(C(r,"Dirección") or "").strip(),"comuna":str(C(r,"Comuna") or "").strip(),
                "telefono":str(C(r,"Telefono") or "").strip(),"periodo":str(C(r,"Periodo Receta") or "").strip(),
                "especialidad":str(C(r,"Especialidad") or "").strip(),
                "n_presc":int(re.sub(r"\D","",str(C(r,"Número Prescripciones") or "0")) or 0),
                "estab_destino":str(C(r,"Estab. Destino") or "").strip(),
                "estab_origen":org or "Pitrufquén Hosp.","refrigerado":refri,"pendiente":"","ventanilla":str(C(r,"Tipo Retiro") or "").strip().upper()=="PACIENTE",
                "_ref":set([refri]) if refri else set()}
        elif refri:
            dd[rec]["_ref"].add(refri)
    for v in dd.values():
        if v["_ref"]: v["refrigerado"]="; ".join(sorted(v["_ref"]))
        v.pop("_ref")
    fechas=[v for v in (str(C(r,"Fecha Entrega") or "").strip() for r in rows[hi+1:] if r) if v]
    fecha=max(set(fechas), key=fechas.count) if fechas else "(sin fecha)"
    origen=max(origenes, key=origenes.get) if origenes else "Farmacia Hospital de Pitrufquén"
    return {"fecha_entrega":fecha,"origen":origen,"registros":list(dd.values())}

def estadistica_mensual(data, ruta):
    from openpyxl.chart import BarChart, Reference
    from collections import Counter
    regs=data["registros"]; fecha=data.get("fecha_entrega","")
    etiqueta=fecha
    m=re.match(r"\d{1,2}/(\d{1,2})/(\d{4})", fecha or "")
    if m:
        meses=["","enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
        etiqueta=f"{meses[int(m.group(1))]} {m.group(2)}"
    wb=Workbook(); ws=wb.active; ws.title="Estadística"
    ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=16
    ws.merge_cells("A1:C1"); t=ws["A1"]; t.value=f"ESTADÍSTICA GESTIÓN TERRITORIAL - {etiqueta}"
    t.font=Font(size=14,bold=True,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=NAVY)
    t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=26
    tot_rec=len(regs); tot_presc=sum(int(g.get("n_presc",0) or 0) for g in regs)
    ultimas=sum(1 for g in regs if es_ultima(g.get("periodo","")))
    refri=sum(1 for g in regs if (g.get("refrigerado","") or "").strip())
    vent=sum(1 for g in regs if g.get("ventanilla"))
    dest=Counter(g.get("estab_destino","") for g in regs); esp=Counter(g.get("especialidad","") for g in regs)
    def bloque(titulo, filas, r0):
        ws.cell(row=r0,column=1,value=titulo).font=Font(size=12,bold=True,color=NAVY)
        r=r0+1
        for k,v in filas:
            ws.cell(row=r,column=1,value=k).font=Font(size=10)
            c=ws.cell(row=r,column=2,value=v); c.font=Font(size=10,bold=True); c.alignment=Alignment(horizontal="center"); r+=1
        return r+1
    r=3
    r=bloque("Resumen", [("Total recetas (sin duplicados)",tot_rec),("Total prescripciones",tot_presc),
                         ("Establecimientos de destino",len(dest)),("Recetas en último período",ultimas),
                         ("Recetas con refrigerados",refri),("Retiros en ventanilla",vent)], r)
    r=bloque("Recetas por establecimiento de destino", sorted(dest.items(), key=lambda x:-x[1]), r)
    esp_r0=r; r=bloque("Recetas por especialidad", sorted(esp.items(), key=lambda x:-x[1]), r)
    try:
        n=len(esp)
        if n:
            chart=BarChart(); chart.title="Recetas por especialidad"; chart.type="bar"; chart.legend=None
            chart.add_data(Reference(ws, min_col=2, min_row=esp_r0+1, max_row=esp_r0+n))
            chart.set_categories(Reference(ws, min_col=1, min_row=esp_r0+1, max_row=esp_r0+n))
            chart.height=max(6,n*0.5); chart.width=18; ws.add_chart(chart,"E3")
    except Exception: pass
    ws.page_setup.paperSize=1; ws.page_setup.orientation="portrait"; ws.print_options.horizontalCentered=True
    wb.save(ruta)

# ----------------------------- ORQUESTACIÓN -----------------------------
LETTER_W = 612.0   # 8.5 in × 72 pt
LETTER_H = 792.0   # 11 in × 72 pt

def _forzar_carta(pdf_path):
    """Post-procesa el PDF para que cada página quede exactamente en tamaño carta.
    Detecta automáticamente orientación landscape (ancho > alto) o portrait.
    Landscape destino: 792 × 612 pt  |  Portrait destino: 612 × 792 pt."""
    try:
        from pypdf import PdfReader, PdfWriter
        from pypdf.generic import ArrayObject, FloatObject
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        for page in reader.pages:
            orig_w = float(page.mediabox.width)
            orig_h = float(page.mediabox.height)
            if orig_w <= 0 or orig_h <= 0:
                writer.add_page(page); continue
            # Detectar orientación: si la página ya es más ancha que alta → landscape
            if orig_w > orig_h:
                target_w, target_h = LETTER_H, LETTER_W   # 792 × 612 (landscape)
            else:
                target_w, target_h = LETTER_W, LETTER_H   # 612 × 792 (portrait)
            scale = min(target_w / orig_w, target_h / orig_h)
            tx = (target_w - orig_w * scale) / 2
            ty = (target_h - orig_h * scale) / 2
            page.add_transformation([scale, 0, 0, scale, tx, ty])
            page.mediabox = ArrayObject([
                FloatObject(0), FloatObject(0),
                FloatObject(target_w), FloatObject(target_h)
            ])
            page.cropbox = page.mediabox
            writer.add_page(page)
        with open(pdf_path, "wb") as f:
            writer.write(f)
    except Exception as e:
        print(f"  (ajuste carta omitido: {e})")

def to_pdf(path, outdir):
    pdf_path = os.path.join(outdir, os.path.splitext(os.path.basename(path))[0] + ".pdf")
    abs_path = os.path.abspath(path)
    # 1. Intentar LibreOffice (soffice)
    try:
        r = subprocess.run(["soffice","--headless","--convert-to","pdf","--outdir",outdir,path],
                           check=False, capture_output=True, timeout=120)
        if os.path.exists(pdf_path):
            _forzar_carta(pdf_path)
            return
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"  (soffice falló: {e})")
    # 2. Fallback: Excel via win32com (Windows + Excel instalado)
    try:
        import win32com.client
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(abs_path)
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))  # 0 = xlTypePDF
        wb.Close(False)
        xl.Quit()
        if os.path.exists(pdf_path):
            _forzar_carta(pdf_path)
            return
    except Exception as e:
        print(f"  (Excel win32com falló: {e})")
    print(f"  (PDF omitido — sin LibreOffice ni Excel disponible)")

FECHA="(sin fecha)"

def main():
    global FECHA
    ap=argparse.ArgumentParser()
    ap.add_argument("entrada", help="reporte .xlsx de Gestión Territorial, o .json de registros"); ap.add_argument("--salida", default="./out"); ap.add_argument("--no-pdf", action="store_true")
    ap.add_argument("--historico", default=None, help="JSON/CSV de histórico de recetas (run,nombre) para cruce de identidad")
    ap.add_argument("--estadistica", action="store_true", help="Generar estadística mensual del reporte")
    a=ap.parse_args()
    if a.entrada.lower().endswith(".xlsx"):
        data=importar_reporte_xlsx(a.entrada)
        print(f"Reporte importado: {len(data['registros'])} recetas únicas (duplicados eliminados) | origen={data['origen']}")
    else:
        data=json.load(open(a.entrada, encoding="utf-8"))
    FECHA=date.today().strftime("%d/%m/%Y")  # siempre la fecha del día de uso
    origen=data.get("origen","Farmacia Hospital de Pitrufquén")
    os.makedirs(a.salida, exist_ok=True)
    idx_ref, refri_map = cargar_sidra(a.historico)   # identidad + refrigerados, en una sola lectura
    hist_on=bool(a.historico)
    if refri_map:
        n_ref=sum(1 for g in data["registros"] if g["receta"] in refri_map)
        print(f"Refrigerados en SIDRA (activan letrero): {n_ref} recetas — la columna Refrigerados de la planilla solo se llena con datos manuscritos")
    acum={k:set(v) for k,v in idx_ref.items()}   # acumulador para el histórico de salida
    grupos=defaultdict(list)
    # Sectorización salud mental: forzar "refrigerado/controlado" si el campo refrigerado está vacío
    for g in data["registros"]:
        if g.get("sectorizacion"):
            if not (g.get("refrigerado") or "").strip():
                g["refrigerado"] = "refrigerado/controlado"
        grupos[g["estab_destino"]].append(g)
    for destino, regs in grupos.items():
        titulo=f"GESTIÓN TERRITORIAL - {destino.upper()}"
        subtitulo=f"Origen: {origen}   |   Destino: {destino}   |   Fecha de entrega: {FECHA}"
        regs_ctrl=[g for g in regs if (g.get("controlado","") or "").strip()]
        regs_norm=[g for g in regs if not (g.get("controlado","") or "").strip()]
        archivos_dest=[]
        if regs_norm:
            wb=Workbook()
            hoja_funcionarios(wb, regs_norm, destino, titulo, subtitulo, modo="normal")
            pl=os.path.join(a.salida, f"{slug(destino)}_Planilla.xlsx"); wb.save(pl); archivos_dest.append(os.path.basename(pl))
            # Planilla: solo xlsx (editable para tick manual), sin PDF
        if regs_ctrl:
            titulo_ctrl=f"GESTIÓN TERRITORIAL (CONTROLADOS) - {destino.upper()}"
            wb=Workbook()
            hoja_funcionarios(wb, regs_ctrl, destino, titulo_ctrl, subtitulo, modo="controlados")
            pl_ctrl=os.path.join(a.salida, f"{slug(destino)}_Controlados_Planilla.xlsx"); wb.save(pl_ctrl); archivos_dest.append(os.path.basename(pl_ctrl))
            # Controlados: solo xlsx, sin PDF
        lleva=any(((g.get("refrigerado","") or "").strip() or (g["receta"] in refri_map)) for g in regs)
        le=os.path.join(a.salida, f"{slug(destino)}_Letrero.xlsx"); letrero(destino, lleva, le)
        # Letrero: solo PDF (xlsx temporal se elimina)
        if not a.no_pdf:
            to_pdf(le, a.salida)
            try: os.remove(le)
            except OSError: pass
            archivos_dest.append(f"{slug(destino)}_Letrero.pdf")
        else:
            archivos_dest.append(os.path.basename(le))
        resumen=f"{len(regs_norm)} normales" + (f" + {len(regs_ctrl)} controlados" if regs_ctrl else "")
        print(f"OK {destino}: {len(regs)} recetas ({resumen}) | refrigerado={lleva} -> {', '.join(archivos_dest)}")
        # Dígito verificador (módulo 11) - siempre
        malos=[(g["receta"],g["paciente"],g["run"]) for g in regs if dv_correcto(g["run"]) is False]
        nulos=[(g["receta"],g["paciente"],g["run"]) for g in regs if dv_correcto(g["run"]) is None]
        if malos:
            print(f"  ⚠ RUN con dígito verificador INVÁLIDO (revisar): {len(malos)}")
            for rec,pac,run in malos: print(f"     - {rec} {pac}: {run}")
        if nulos:
            print(f"  • RUN no validable (formato/incompleto): {len(nulos)}")
        if not malos and not nulos:
            print("  ✓ Todos los RUN con dígito verificador válido.")
        # Cruce con histórico
        if hist_on:
            resultados = verificar_historico(regs, idx_ref)
            vpath = os.path.join(a.salida, f"{slug(destino)}_Verificacion.xlsx")
            reporte_verificacion(resultados, destino, vpath)
            distintos = [(g["paciente"], estado, previos) for g, estado, previos in resultados if estado == "NOMBRE DISTINTO"]
            nuevos   = [(g["paciente"], g["run"])            for g, estado, _       in resultados if estado == "NUEVO"]
            print(f"  Verificación: {sum(1 for _,e,_ in resultados if e=='COINCIDE')} coinciden | {len(nuevos)} nuevos | {len(distintos)} nombre distinto -> {os.path.basename(vpath)}")
            if distintos:
                print(f"  ⚠ Nombre distinto al histórico (revisar):")
                for pac, est, prev in distintos: print(f"     {pac} → histórico: {prev}")
            # Actualizar acumulador
            for g, estado, _ in resultados:
                run = norm_run(g["run"])
                if run: acum.setdefault(run, set()).add(norm_nombre(g["paciente"]))

    if hist_on:
        out_hist = os.path.join(a.salida, "historico_actualizado.json")
        with open(out_hist, "w", encoding="utf-8") as f:
            json.dump({"pacientes": [{"run": k, "nombre": list(v)} for k, v in acum.items()]}, f, ensure_ascii=False, indent=2)
        print(f"Histórico actualizado guardado: {os.path.basename(out_hist)}")

    if a.estadistica:
        periodo = data.get("fecha_entrega", str(date.today()))
        esp = os.path.join(a.salida, f"Estadistica_{re.sub(r'[/\\\\:]', '-', periodo)}.xlsx")
        estadistica_mensual(data, esp)
        print(f"Estadística generada: {os.path.basename(esp)}")

if __name__ == "__main__":
    main()
