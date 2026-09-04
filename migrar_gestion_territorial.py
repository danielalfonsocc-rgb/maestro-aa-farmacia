#!/usr/bin/env python3
"""
Script: migrar_gestion_territorial.py
Propósito: Migración, depuración y blindaje de la base de datos de Gestión Territorial
           de Farmacia (Hospital de Pitrufquén) al modelo Maestro_Pacientes +
           Despachos_Mensual (Fase 1 — ver plan de migración).
Requisitos: pip install pandas openpyxl pillow
Uso: py migrar_gestion_territorial.py [--origen "ruta\\a\\GT PITRUFQUEN 2026 (N).xlsx"]

Basado en el borrador original del usuario, con 8 correcciones aplicadas tras
revisar los datos reales (ver C:\\Users\\danie\\.claude\\plans\\peaceful-forging-floyd.md):
  1. Enero-Agosto y las 2 hojas SECTORIZACION se archivan (hojas Archivo_*),
     no se descartan.
  2. Data validation real conectando Parametros a las columnas de Despachos.
  3. Detección tolerante de columnas también en el loop de Despachos_Septiembre
     (antes accedía por nombre exacto).
  4. Se quita el chequeo muerto de '.0' en clean_rut (el punto ya se había
     eliminado antes).
  5. Paleta/fuente/logos institucionales del proyecto (skill_gt/scripts/generar.py)
     en vez de una paleta propia.
  6. Fórmulas del Dashboard vía INDIRECT + celda "Mes Activo" (Parametros!G2)
     en vez de hardcodear el nombre de la hoja del mes.
  7. KNOWN_RUT_FIXES y la limpieza de "//" en Periodo Receta quedan auditables:
     el script imprime cada corrección aplicada al correr.
  8. buscar_archivo_origen() falla con mensaje claro si no encuentra un archivo
     que calce con el patrón esperado, en vez de tomar cualquier .xlsx al azar.
"""

import os
import re
import sys
import glob
import argparse
import unicodedata
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "skill_gt", "scripts"))
import generar as G  # noqa: E402  (paleta NAVY/BLUE/GREY/LIGHT + place_logo + logos institucionales)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# --------------------------------------------------------------------------
# Correcciones de datos conocidas — quedan acá porque son puntuales (no una
# homologación recurrente), pero se imprimen cada vez que se aplican (punto 7).
# --------------------------------------------------------------------------
KNOWN_RUT_FIXES = {
    'LUIS SEGUNDO LILLO HERRERA': '7644411-0',
    'NELSON MARCELO ZAPATA SANCHEZ': '11259371-3',
}
_fixes_rut_aplicados = set()
_fixes_periodo_aplicados = 0


def buscar_archivo_origen(ruta_forzada=None):
    """Localiza el archivo crudo. Falla con mensaje claro si no hay un calce
    razonable — nunca toma "cualquier .xlsx" del directorio a ciegas."""
    if ruta_forzada:
        if not os.path.exists(ruta_forzada):
            raise SystemExit(f"[ERROR] No existe el archivo indicado en --origen: {ruta_forzada}")
        return ruta_forzada
    candidatos = [f for f in glob.glob("*GT*PITRUFQUEN*.xlsx")
                  if "DEFINITIVO" not in f.upper() and "BLINDADA" not in f.upper() and "TEST" not in f.upper()]
    if not candidatos:
        raise SystemExit(
            "[ERROR] No se encontró ningún archivo *GT*PITRUFQUEN*.xlsx en el directorio actual.\n"
            "         Pasa la ruta explícita con --origen \"ruta\\al\\archivo.xlsx\"."
        )
    if len(candidatos) > 1:
        candidatos.sort(key=os.path.getmtime, reverse=True)
        print(f"[AVISO] Hay {len(candidatos)} archivos que calzan — se usa el más reciente: {candidatos[0]}")
    return candidatos[0]


def norm_colname(s):
    """Nombre de columna normalizado: sin tildes, minúsculas, sin espacios extra."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


def find_col(df, keywords, exclude=None, required=True):
    """Encuentra la columna de df cuyo nombre normalizado contiene TODAS las
    keywords y ninguna de exclude. Reemplaza el acceso por string exacto
    (row['Paciente '], row['Rut'], ...) que rompía si el encabezado real
    difería en un espacio o una tilde — la planilla real tiene 4 variantes de
    encabezado distintas entre meses."""
    exclude = exclude or []
    for c in df.columns:
        cn = norm_colname(c)
        if all(kw in cn for kw in keywords) and not any(kw in cn for kw in exclude):
            return c
    if required:
        raise KeyError(f"No se encontró columna con keywords {keywords!r} (excluye {exclude!r}) "
                        f"entre {list(df.columns)!r}")
    return None


def clean_rut(val):
    """Sanitiza y normaliza el RUT al estándar chileno (sin puntos, con guion, DV mayúscula)."""
    if pd.isna(val):
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)  # evita que "12345678.0" se transforme mal al sacar el punto
    s = str(val).replace('\xa0', '').replace(' ', '').replace('.', '').strip().upper()
    if '-' not in s and len(s) in (8, 9):
        s = s[:-1] + '-' + s[-1]
    return s


def clean_phone(val):
    """Limpia números de teléfono eliminando decimales (.0) y caracteres extraños."""
    if pd.isna(val):
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    s = str(val).replace('.0', '').replace(' ', '').replace('-', '').strip()
    if s.lower() == 'nan' or s == '0':
        return ""
    return s


def es_si(v):
    """True si v representa 'refrigerado/controlado = sí', tolerante a booleano
    real (como trae la planilla actual) o a texto tipo 'INSULINA x2' (como lo
    escribe gt_maestro.py) — no asume un único formato."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip()
    return s != "" and s.lower() not in ("no", "false", "0", "nan")


def limpiar_periodo(periodo):
    """Colapsa dobles barras ('5//12' -> '5/12'), error de tipeo detectado en
    la planilla real. Generaliza el parche puntual del borrador original a
    cualquier ocurrencia, no solo a '5//12'."""
    global _fixes_periodo_aplicados
    if not periodo:
        return periodo
    nuevo = re.sub(r'/{2,}', '/', periodo)
    if nuevo != periodo:
        _fixes_periodo_aplicados += 1
    return nuevo


def aplicar_fix_rut_conocido(pac, c_rut):
    if pac in KNOWN_RUT_FIXES and (not c_rut or len(c_rut) < 7):
        _fixes_rut_aplicados.add((pac, KNOWN_RUT_FIXES[pac]))
        return KNOWN_RUT_FIXES[pac]
    return c_rut


def nombre_hoja_archivo(original):
    """'SECTORIZACION SEPTIEMBRE' -> 'Archivo_Sectoriz SEPTIEMBRE' (<=31 chars,
    límite de Excel para nombres de hoja)."""
    n = original.replace("SECTORIZACION", "Sectoriz").strip()
    return f"Archivo_{n}"[:31]


def ejecutar_migracion(archivo_origen, incluir_archivo=False):
    print(f"[MIGRACION] Procesando archivo base: {archivo_origen}")
    xls = pd.ExcelFile(archivo_origen)

    monthly_sheets_todas = [
        'ENERO 2026', 'FEBRERO 2026', 'MARZO 2026', 'ABRIL 2026',
        'MAYO 2026', 'JUNIO 2026', 'JULIO 2026', 'AGOSTO 2026', 'SEPTIEMBRE 2026'
    ]
    monthly_sheets_todas = [s for s in monthly_sheets_todas if s in xls.sheet_names]
    MES_ACTIVO = 'SEPTIEMBRE 2026'
    if MES_ACTIVO not in monthly_sheets_todas:
        raise SystemExit(f"[ERROR] No se encontró la hoja del mes activo ({MES_ACTIVO}) en {archivo_origen}")
    meses_archivo = [s for s in monthly_sheets_todas if s != MES_ACTIVO]
    hojas_sectorizacion = [s for s in ['SECTORIZACION AGOSTO', 'SECTORIZACION SEPTIEMBRE'] if s in xls.sheet_names]

    patients = {}

    # 1. Consolidar Maestro de Pacientes único y depurado (sobre TODAS las hojas de mes)
    for sheet in monthly_sheets_todas:
        df = pd.read_excel(xls, sheet_name=sheet)
        rut_c = find_col(df, ['rut'])
        pac_c = find_col(df, ['paciente'])
        dest_c = find_col(df, ['destino'])
        tel_c = find_col(df, ['tel'])
        esp_c = find_col(df, ['especialidad'], required=False)
        ref_c = find_col(df, ['refrigerad'], required=False)
        ctrl_c = find_col(df, ['controlad'], required=False)
        est_c = find_col(df, ['estado'], exclude=['receta'], required=False)

        for _, row in df.iterrows():
            pac = str(row[pac_c]).strip() if pd.notna(row[pac_c]) else ""
            c_rut = clean_rut(row[rut_c])
            c_rut = aplicar_fix_rut_conocido(pac, c_rut)

            if not c_rut or len(c_rut) < 7:
                continue

            dest = str(row[dest_c]).strip() if pd.notna(row[dest_c]) else ""
            tel = clean_phone(row[tel_c])
            esp = str(row[esp_c]).strip() if esp_c and pd.notna(row[esp_c]) else ""
            ref = 1 if ref_c and es_si(row[ref_c]) else 0
            ctrl = 1 if ctrl_c and es_si(row[ctrl_c]) else 0
            estado_raw = str(row[est_c]).strip().lower() if est_c and pd.notna(row[est_c]) else ""

            if c_rut not in patients:
                patients[c_rut] = {
                    'rut': c_rut, 'nombre': pac, 'telefono': tel, 'destino': dest,
                    'especialidades': set([esp]) if esp else set(),
                    'refrigerado': ref, 'controlado': ctrl,
                    'fallecido': 1 if 'fallecid' in estado_raw else 0,
                    'total_despachos': 1,
                }
            else:
                p = patients[c_rut]
                p['total_despachos'] += 1
                if pac and (not p['nombre'] or len(pac) > len(p['nombre'])):
                    p['nombre'] = pac
                if dest:
                    p['destino'] = dest
                if tel:
                    p['telefono'] = tel
                if esp:
                    p['especialidades'].add(esp)
                if ref == 1:
                    p['refrigerado'] = 1
                if ctrl == 1:
                    p['controlado'] = 1
                if 'fallecid' in estado_raw:
                    p['fallecido'] = 1

    # Escanear hojas de sectorización — solo suma pacientes que no aparezcan ya
    # (no traen columna de sector rural real, así que no aportan ese campo).
    for s in hojas_sectorizacion:
        df_sec = pd.read_excel(xls, sheet_name=s)
        run_col = find_col(df_sec, ['run'], required=False) or find_col(df_sec, ['rut'])
        pac_col = find_col(df_sec, ['paciente'])
        dest_col = find_col(df_sec, ['destino'])
        for _, row in df_sec.iterrows():
            pac = str(row[pac_col]).strip() if pd.notna(row[pac_col]) else ""
            c_rut = clean_rut(row[run_col])
            c_rut = aplicar_fix_rut_conocido(pac, c_rut)
            if not c_rut or len(c_rut) < 7:
                continue
            dest = str(row[dest_col]).strip() if pd.notna(row[dest_col]) else ""
            if c_rut not in patients:
                patients[c_rut] = {
                    'rut': c_rut, 'nombre': pac, 'telefono': "", 'destino': dest,
                    'especialidades': set(), 'refrigerado': 0, 'controlado': 0,
                    'fallecido': 0, 'total_despachos': 1,
                }

    print(f"[MIGRACION] Total pacientes únicos consolidados en Maestro: {len(patients)}")
    if _fixes_rut_aplicados:
        print(f"[AUDITORIA] Correcciones KNOWN_RUT_FIXES aplicadas ({len(_fixes_rut_aplicados)}):")
        for pac, rut in sorted(_fixes_rut_aplicados):
            print(f"    - {pac} -> RUT asignado {rut}  (VERIFICAR)")

    # ------------------------------------------------------------------
    # Construcción del libro definitivo con openpyxl
    # ------------------------------------------------------------------
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Paleta institucional del proyecto (skill_gt/scripts/generar.py) — no una
    # paleta propia, para no desentonar con el resto de planillas GT.
    c_navy, c_blue, c_ice, c_zebra = G.NAVY, G.BLUE, G.LIGHT, G.GREY
    c_border = "BFBFBF"
    c_alert_red, c_alert_red_tx = "FFC7CE", "9C0006"
    c_alert_amber, c_alert_amber_tx = "FFEB9C", "9C5700"

    FUENTE = "Calibri"
    f_title = Font(name=FUENTE, size=15, bold=True, color=c_navy)
    f_sub = Font(name=FUENTE, size=9, italic=True, color="595959")
    f_hdr = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
    f_bold = Font(name=FUENTE, size=10, bold=True)
    f_reg = Font(name=FUENTE, size=10)
    f_kpi_val = Font(name=FUENTE, size=20, bold=True, color=c_navy)
    f_kpi_lbl = Font(name=FUENTE, size=9, bold=True, color="595959")

    fill_hdr = PatternFill(start_color=c_navy, end_color=c_navy, fill_type="solid")
    fill_subhdr = PatternFill(start_color=c_blue, end_color=c_blue, fill_type="solid")
    fill_card = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_zebra = PatternFill(start_color=c_zebra, end_color=c_zebra, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=c_border), right=Side(style='thin', color=c_border),
        top=Side(style='thin', color=c_border), bottom=Side(style='thin', color=c_border),
    )

    # -------------------------------------------------------------
    # HOJA: PARAMETROS
    # -------------------------------------------------------------
    ws_param = wb.create_sheet(title="Parametros")
    ws_param.cell(row=1, column=1, value="LISTAS DE VALIDACIÓN DE GESTIÓN TERRITORIAL").font = f_title
    ws_param.cell(row=2, column=1, value="Rangos maestros para dropdowns y control de integridad").font = f_sub

    estab_list = [
        "Cesfam Freire", "Cesfam Quepe", "Cesfam Teodoro", "Hospital Tolten",
        "Psr Queule", "Dsm Loncoche", "Hospital Loncoche", "Dsm Gorbea",
        "Hospital Gorbea", "Entrega a Domicilio",
    ]
    estados_list = [
        "En preparación", "Lista para retiro", "Enviada", "Retiro en ventanilla",
        "pendiente por falta stock", "sin receta vigente", "fallecida", "Digitada",
    ]
    binarios_list = ["SÍ", "NO"]

    ws_param.cell(row=4, column=1, value="Establecimientos de Destino").font = f_hdr
    ws_param.cell(row=4, column=1).fill = fill_hdr
    for idx, val in enumerate(estab_list, start=5):
        c = ws_param.cell(row=idx, column=1, value=val)
        c.font, c.border = f_reg, thin_border

    ws_param.cell(row=4, column=3, value="Estados de Despacho").font = f_hdr
    ws_param.cell(row=4, column=3).fill = fill_hdr
    for idx, val in enumerate(estados_list, start=5):
        c = ws_param.cell(row=idx, column=3, value=val)
        c.font, c.border = f_reg, thin_border

    ws_param.cell(row=4, column=5, value="Opciones Binarias").font = f_hdr
    ws_param.cell(row=4, column=5).fill = fill_hdr
    for idx, val in enumerate(binarios_list, start=5):
        c = ws_param.cell(row=idx, column=5, value=val)
        c.font, c.border = f_reg, thin_border

    # "Mes Activo": única celda que el Dashboard resuelve por INDIRECT (punto 6) —
    # así generarRolloverMes() solo cambia esta celda, no cada fórmula del Dashboard.
    ws_param.cell(row=1, column=7, value="MES ACTIVO").font = f_hdr
    ws_param.cell(row=1, column=7).fill = fill_hdr
    c_mes_activo = ws_param.cell(row=2, column=7, value="Despachos_Septiembre")
    c_mes_activo.font = Font(name=FUENTE, size=11, bold=True, color=c_navy)
    c_mes_activo.fill = fill_card
    c_mes_activo.border = thin_border
    MES_ACTIVO_REF = 'Parametros!$G$2'

    ws_param.column_dimensions['A'].width = 28
    ws_param.column_dimensions['C'].width = 26
    ws_param.column_dimensions['E'].width = 18
    ws_param.column_dimensions['G'].width = 24

    RANGO_DV_ESTAB = f"Parametros!$A$5:$A${4 + len(estab_list)}"
    RANGO_DV_ESTADOS = f"Parametros!$C$5:$C${4 + len(estados_list)}"
    RANGO_DV_BINARIO = f"Parametros!$E$5:$E${4 + len(binarios_list)}"

    # -------------------------------------------------------------
    # HOJA: MAESTRO_PACIENTES
    # -------------------------------------------------------------
    ws_maestro = wb.create_sheet(title="Maestro_Pacientes")
    ws_maestro.freeze_panes = "A4"
    ws_maestro.cell(row=1, column=1, value="Maestro Central de Pacientes - Gestión Territorial").font = f_title
    ws_maestro.cell(row=2, column=1,
                     value=f"Catálogo maestro depurado con {len(patients)} pacientes únicos de la red").font = f_sub

    m_headers = [
        "RUT Normalizado (Clave)", "Nombre Completo Paciente", "Teléfono Principal",
        "Teléfono Cuidador / Respaldo", "Establecimiento Destino Habitual",
        "Sector Rural / Referencia", "Especialidades Frecuentes", "Requiere Frío Habitual",
        "Maneja Controlados", "Estado Paciente", "Observaciones Clínicas", "Total Despachos Históricos 2026",
    ]
    for col_idx, text in enumerate(m_headers, start=1):
        cell = ws_maestro.cell(row=3, column=col_idx, value=text)
        cell.font, cell.fill = f_hdr, fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_maestro.row_dimensions[3].height = 28

    sorted_patients = sorted(patients.values(), key=lambda x: x['nombre'])
    for r_idx, p in enumerate(sorted_patients, start=4):
        use_z = (r_idx % 2 == 0)
        estado_p = "Fallecido" if p['fallecido'] == 1 else "Activo"
        frio_p = "SÍ" if p['refrigerado'] == 1 else "NO"
        ctrl_p = "SÍ" if p['controlado'] == 1 else "NO"
        esp_p = ", ".join(sorted(p['especialidades']))

        row_vals = [p['rut'], p['nombre'], p['telefono'], "", p['destino'], "",
                    esp_p, frio_p, ctrl_p, estado_p, "", p['total_despachos']]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_maestro.cell(row=r_idx, column=c_idx, value=val)
            cell.font, cell.border = f_reg, thin_border
            if use_z:
                cell.fill = fill_zebra
            cell.alignment = Alignment(horizontal="center" if c_idx in (1, 3, 4, 8, 9, 10, 12) else "left",
                                        vertical="center")

    col_widths_m = {'A': 18, 'B': 36, 'C': 16, 'D': 18, 'E': 24, 'F': 25, 'G': 30, 'H': 14, 'I': 14, 'J': 15, 'K': 25, 'L': 15}
    for col, w in col_widths_m.items():
        ws_maestro.column_dimensions[col].width = w

    # Data validation: Establecimiento Destino Habitual (E) contra la misma lista
    dv_estab_maestro = DataValidation(type="list", formula1=f"={RANGO_DV_ESTAB}", allow_blank=True)
    ws_maestro.add_data_validation(dv_estab_maestro)
    dv_estab_maestro.add(f"E4:E{3 + max(len(sorted_patients), 1)}")
    dv_bin_maestro = DataValidation(type="list", formula1=f"={RANGO_DV_BINARIO}", allow_blank=True)
    ws_maestro.add_data_validation(dv_bin_maestro)
    dv_bin_maestro.add(f"H4:I{3 + max(len(sorted_patients), 1)}")

    # -------------------------------------------------------------
    # Encabezados y fórmulas de Despachos (reutilizados en Septiembre y Plantilla)
    # -------------------------------------------------------------
    desp_headers = [
        "N° Envío GT", "N° Receta", "RUT Paciente", "Nombre Paciente (Auto)",
        "Teléfono Contacto (Auto)", "Establecimiento Destino", "Periodo Receta",
        "Alerta Tratamiento (Auto)", "Especialidad Médica", "N° Prescripciones",
        "Estado Despacho", "Refrigerado", "Controlado", "Fármaco Pendiente / Stock",
        "Fecha Retiro Farmacia", "Control Térmico / Termo", "Custodia Receptor (Firma)", "Observaciones Envío",
    ]

    def formula_nombre(r):
        return (f'=IF(C{r}="","",IFNA(XLOOKUP(UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(C{r}),".","")," ","")), '
                f'Maestro_Pacientes!$A$4:$A$1000, Maestro_Pacientes!$B$4:$B$1000), "⚠ RUT No Registrado"))')

    def formula_telefono(r):
        return (f'=IF(C{r}="","",IFNA(XLOOKUP(UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(C{r}),".","")," ","")), '
                f'Maestro_Pacientes!$A$4:$A$1000, Maestro_Pacientes!$C$4:$C$1000), "-"))')

    def formula_destino_default(r):
        return (f'=IF(C{r}="","",IFNA(XLOOKUP(UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(C{r}),".","")," ","")), '
                f'Maestro_Pacientes!$A$4:$A$1000, Maestro_Pacientes!$E$4:$E$1000), "-"))')

    def formula_alerta(r):
        return (f'=IF(K{r}="sin receta vigente","⚠ SIN RECETA VIGENTE",'
                f'IF(OR(G{r}="ULTIMA",G{r}="ultima"),"⚠ ÚLTIMA RECETA / RENOVAR",'
                f'IF(G{r}="","",IFERROR(IF(VALUE(LEFT(G{r},FIND("/",G{r})-1))>=VALUE(MID(G{r},FIND("/",G{r})+1,5)),'
                f'"⚠ ÚLTIMA RECETA / RENOVAR",IF(VALUE(MID(G{r},FIND("/",G{r})+1,5))-VALUE(LEFT(G{r},FIND("/",G{r})-1))<=1,'
                f'"Penúltima Receta (1 mes)","Vigente")),"Vigente"))))')

    def formula_frio_default(r):
        return (f'=IF(C{r}="","",IFNA(XLOOKUP(UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(C{r}),".","")," ","")), '
                f'Maestro_Pacientes!$A$4:$A$1000, Maestro_Pacientes!$H$4:$H$1000), "NO"))')

    def formula_ctrl_default(r):
        return (f'=IF(C{r}="","",IFNA(XLOOKUP(UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(C{r}),".","")," ","")), '
                f'Maestro_Pacientes!$A$4:$A$1000, Maestro_Pacientes!$I$4:$I$1000), "NO"))')

    def escribir_encabezado_despachos(ws):
        for col_idx, text in enumerate(desp_headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font, cell.fill = f_hdr, fill_hdr
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 28

    def aplicar_dv_despachos(ws, hasta_fila):
        dv_estab = DataValidation(type="list", formula1=f"={RANGO_DV_ESTAB}", allow_blank=True)
        ws.add_data_validation(dv_estab)
        dv_estab.add(f"F4:F{hasta_fila}")
        dv_estados = DataValidation(type="list", formula1=f"={RANGO_DV_ESTADOS}", allow_blank=True)
        ws.add_data_validation(dv_estados)
        dv_estados.add(f"K4:K{hasta_fila}")
        dv_bin = DataValidation(type="list", formula1=f"={RANGO_DV_BINARIO}", allow_blank=True)
        ws.add_data_validation(dv_bin)
        dv_bin.add(f"L4:M{hasta_fila}")

    col_widths_sep = {'A': 14, 'B': 14, 'C': 16, 'D': 35, 'E': 16, 'F': 22, 'G': 14, 'H': 26, 'I': 24,
                       'J': 12, 'K': 24, 'L': 14, 'M': 14, 'N': 30, 'O': 16, 'P': 22, 'Q': 24, 'R': 25}

    # -------------------------------------------------------------
    # HOJA: DESPACHOS_SEPTIEMBRE (mes activo real, datos migrados)
    # -------------------------------------------------------------
    ws_sep = wb.create_sheet(title="Despachos_Septiembre")
    ws_sep.freeze_panes = "D4"
    ws_sep.cell(row=1, column=1, value="Registro Operativo de Despachos - Septiembre 2026").font = f_title
    ws_sep.cell(row=2, column=1,
                value="Datos migrados y ordenados con correlativo, control de stock y alerta de fin de receta").font = f_sub
    escribir_encabezado_despachos(ws_sep)

    df_sep_raw = pd.read_excel(xls, sheet_name=MES_ACTIVO)
    col_receta = find_col(df_sep_raw, ['receta'])
    col_pac = find_col(df_sep_raw, ['paciente'])
    col_rut = find_col(df_sep_raw, ['rut'])
    col_dest = find_col(df_sep_raw, ['destino'])
    col_periodo = find_col(df_sep_raw, ['periodo'])
    col_esp = find_col(df_sep_raw, ['especialidad'])
    col_numpx = find_col(df_sep_raw, ['numero', 'prescripcion'])
    col_estado = find_col(df_sep_raw, ['estado'], exclude=['receta'], required=False)
    col_ref = find_col(df_sep_raw, ['refrigerad'], required=False)
    col_ctrl = find_col(df_sep_raw, ['controlad'], required=False)
    col_pend = find_col(df_sep_raw, ['pendiente'], required=False)
    col_fret = find_col(df_sep_raw, ['retiro'], required=False)

    df_sep_raw = df_sep_raw.sort_values(by=[col_dest, col_receta])

    r_final = 3
    for r_idx, (_, row) in enumerate(df_sep_raw.iterrows(), start=4):
        r = r_idx
        use_z = (r % 2 == 0)
        valija_id = ""  # se deja vacío para filas migradas del historial (punto 4 del plan
                         # de diseño): no se inventa un correlativo físico retroactivo.
        num_rec = int(row[col_receta]) if pd.notna(row[col_receta]) else ""
        pac_name = str(row[col_pac]).strip() if pd.notna(row[col_pac]) else ""
        c_rut = clean_rut(row[col_rut])
        c_rut = aplicar_fix_rut_conocido(pac_name, c_rut)

        dest = str(row[col_dest]).strip() if pd.notna(row[col_dest]) else ""
        periodo = limpiar_periodo(str(row[col_periodo]).strip() if pd.notna(row[col_periodo]) else "")
        especialidad = str(row[col_esp]).strip() if pd.notna(row[col_esp]) else ""
        num_px = int(row[col_numpx]) if pd.notna(row[col_numpx]) else ""
        estado = str(row[col_estado]).strip() if col_estado and pd.notna(row[col_estado]) else "Enviada"
        ref_str = "SÍ" if col_ref and es_si(row[col_ref]) else "NO"
        ctrl_str = "SÍ" if col_ctrl and es_si(row[col_ctrl]) else "NO"
        pendiente = str(row[col_pend]).strip() if col_pend and pd.notna(row[col_pend]) else ""

        fecha_ret = ""
        if col_fret and pd.notna(row[col_fret]):
            try:
                fecha_ret = pd.to_datetime(row[col_fret]).strftime('%Y-%m-%d')
            except Exception:
                fecha_ret = str(row[col_fret])[:10]

        ws_sep.cell(row=r, column=1, value=valija_id)
        ws_sep.cell(row=r, column=2, value=num_rec)
        ws_sep.cell(row=r, column=3, value=c_rut)
        ws_sep.cell(row=r, column=4, value=formula_nombre(r))
        ws_sep.cell(row=r, column=5, value=formula_telefono(r))
        ws_sep.cell(row=r, column=6, value=dest)
        ws_sep.cell(row=r, column=7, value=periodo)
        ws_sep.cell(row=r, column=8, value=formula_alerta(r))
        ws_sep.cell(row=r, column=9, value=especialidad)
        ws_sep.cell(row=r, column=10, value=num_px)
        ws_sep.cell(row=r, column=11, value=estado)
        ws_sep.cell(row=r, column=12, value=ref_str)
        ws_sep.cell(row=r, column=13, value=ctrl_str)
        ws_sep.cell(row=r, column=14, value=pendiente)
        ws_sep.cell(row=r, column=15, value=fecha_ret)

        for c in range(1, 19):
            cell = ws_sep.cell(row=r, column=c)
            cell.font, cell.border = f_reg, thin_border
            if use_z:
                cell.fill = fill_zebra
            cell.alignment = Alignment(horizontal="center" if c in (1, 2, 3, 5, 7, 8, 10, 11, 12, 13, 15) else "left",
                                        vertical="center")
        r_final = r

    # Filas "colchón" con fórmula precargada MÁS ALLÁ de los datos migrados —
    # sin esto, cualquier despacho nuevo que se agregue durante septiembre
    # (el mes sigue activo, no es solo un archivo histórico) cae en una fila
    # sin XLOOKUP y no completa Nombre/Teléfono/Alerta. Mismo criterio que
    # Plantilla_Mes_Nuevo, con el mismo tamaño de colchón (150 filas).
    COLCHON_DESPACHOS_ACTIVOS = 150
    inicio_colchon = r_final + 1
    fin_colchon = r_final + COLCHON_DESPACHOS_ACTIVOS
    for r in range(inicio_colchon, fin_colchon + 1):
        use_z = (r % 2 == 0)
        ws_sep.cell(row=r, column=4, value=formula_nombre(r))
        ws_sep.cell(row=r, column=5, value=formula_telefono(r))
        ws_sep.cell(row=r, column=6, value=formula_destino_default(r))
        ws_sep.cell(row=r, column=8, value=formula_alerta(r))
        ws_sep.cell(row=r, column=12, value=formula_frio_default(r))
        ws_sep.cell(row=r, column=13, value=formula_ctrl_default(r))
        for c in range(1, 19):
            cell = ws_sep.cell(row=r, column=c)
            cell.font, cell.border = f_reg, thin_border
            if use_z:
                cell.fill = fill_zebra
            cell.alignment = Alignment(horizontal="center" if c in (1, 2, 3, 5, 7, 8, 10, 11, 12, 13, 15) else "left",
                                        vertical="center")
    r_final = fin_colchon

    aplicar_dv_despachos(ws_sep, max(r_final, 4))
    if _fixes_periodo_aplicados:
        print(f"[AUDITORIA] Se corrigieron {_fixes_periodo_aplicados} valores de Periodo Receta con '//' duplicado.")

    # -------------------------------------------------------------
    # HOJA: PLANTILLA_MES_NUEVO
    # -------------------------------------------------------------
    ws_new = wb.create_sheet(title="Plantilla_Mes_Nuevo")
    ws_new.freeze_panes = "D4"
    ws_new.cell(row=1, column=1, value="Plantilla Operativa de Despachos Mensuales").font = f_title
    ws_new.cell(row=2, column=1,
                value="Duplique esta hoja para iniciar un nuevo mes. Fórmulas precargadas listas para usar.").font = f_sub
    escribir_encabezado_despachos(ws_new)

    ULTIMA_FILA_PLANTILLA = 150
    for r in range(4, ULTIMA_FILA_PLANTILLA + 1):
        use_z = (r % 2 == 0)
        ws_new.cell(row=r, column=4, value=formula_nombre(r))
        ws_new.cell(row=r, column=5, value=formula_telefono(r))
        ws_new.cell(row=r, column=6, value=formula_destino_default(r))
        ws_new.cell(row=r, column=8, value=formula_alerta(r))
        ws_new.cell(row=r, column=12, value=formula_frio_default(r))
        ws_new.cell(row=r, column=13, value=formula_ctrl_default(r))
        for c in range(1, 19):
            cell = ws_new.cell(row=r, column=c)
            cell.font, cell.border = f_reg, thin_border
            if use_z:
                cell.fill = fill_zebra
            cell.alignment = Alignment(horizontal="center" if c in (1, 2, 3, 5, 7, 8, 10, 11, 12, 13, 15) else "left",
                                        vertical="center")

    aplicar_dv_despachos(ws_new, ULTIMA_FILA_PLANTILLA)

    for col, width in col_widths_sep.items():
        ws_sep.column_dimensions[col].width = width
        ws_new.column_dimensions[col].width = width

    # -------------------------------------------------------------
    # HOJAS ARCHIVO_* — Enero-Agosto + Sectorización (opcional, --con-archivo).
    # Por defecto NO se incluyen: a pedido del usuario, el archivo final se
    # queda liviano (Dashboard + Parametros + Maestro_Pacientes + el mes
    # activo) — el historial crudo sigue disponible tal cual en el archivo
    # base (GT PITRUFQUEN 2026 (N).xlsx), no se pierde, solo no se duplica acá.
    # -------------------------------------------------------------
    if incluir_archivo:
        wb_origen_raw = openpyxl.load_workbook(archivo_origen, data_only=True)
        for nombre_original in meses_archivo + hojas_sectorizacion:
            ws_src = wb_origen_raw[nombre_original]
            ws_arch = wb.create_sheet(title=nombre_hoja_archivo(nombre_original))
            for row in ws_src.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    ws_arch.cell(row=cell.row, column=cell.column, value=cell.value)
            if ws_src.max_row >= 1:
                for c in range(1, ws_src.max_column + 1):
                    hc = ws_arch.cell(row=1, column=c)
                    if hc.value is not None:
                        hc.font, hc.fill = f_hdr, fill_hdr
            ws_arch.freeze_panes = "A2"
            print(f"[ARCHIVO] {nombre_original} -> hoja '{ws_arch.title}' ({ws_src.max_row - 1} filas)")

    # -------------------------------------------------------------
    # HOJA: DASHBOARD (primera hoja del libro)
    # -------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Dashboard", index=0)

    G.place_logo(ws_dash, G.LOGO_SS, "A", 40, 4, 4, from_row0=0)
    G.place_logo(ws_dash, G.LOGO_HOSP, "I", 34, 4, 4, from_row0=0)
    ws_dash.row_dimensions[1].height = 42

    ws_dash.cell(row=2, column=2, value="Panel de Control - Gestión Territorial Farmacia").font = f_title
    ws_dash.cell(row=3, column=2,
                 value="Hospital de Pitrufquén | Indicadores consolidados de la red territorial de salud").font = f_sub

    def kpi_indirect(rango):
        return f'"\'" & {MES_ACTIVO_REF} & "\'!{rango}"'

    kpis = [
        ("TOTAL PACIENTES MAESTRO", "=COUNTA(Maestro_Pacientes!A4:A1000)", 2),
        ("DESPACHOS MES ACTIVO", f"=COUNTA(INDIRECT({kpi_indirect('B4:B1000')}))", 4),
        ("ÚLTIMAS RECETAS / RENOVAR", f'=COUNTIF(INDIRECT({kpi_indirect("H4:H1000")}), "⚠ ÚLTIMA RECETA / RENOVAR")', 6),
        ("CADENA DE FRÍO (MES)", f'=COUNTIF(INDIRECT({kpi_indirect("L4:L1000")}), "SÍ")', 8),
        ("QUIEBRES / PENDIENTES", f'=COUNTIF(INDIRECT({kpi_indirect("K4:K1000")}), "pendiente por falta stock")', 10),
    ]
    for label, formula, col in kpis:
        c_lbl = ws_dash.cell(row=5, column=col, value=label)
        c_lbl.font, c_lbl.fill = f_kpi_lbl, fill_card
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_lbl.border = Border(top=Side(style='thin', color="B0C4DE"), left=Side(style='thin', color="B0C4DE"),
                               right=Side(style='thin', color="B0C4DE"))
        c_val = ws_dash.cell(row=6, column=col, value=formula)
        c_val.font, c_val.fill = f_kpi_val, fill_card
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = Border(bottom=Side(style='thin', color="B0C4DE"), left=Side(style='thin', color="B0C4DE"),
                               right=Side(style='thin', color="B0C4DE"))

    ws_dash.cell(row=9, column=2, value="Distribución de Despachos por Centro de Destino (Mes Activo)").font = \
        Font(name=FUENTE, size=11, bold=True, color=c_navy)
    ws_dash.cell(row=11, column=2, value="Establecimiento").font = f_hdr
    ws_dash.cell(row=11, column=2).fill = fill_subhdr
    ws_dash.cell(row=11, column=3, value="Total Despachos").font = f_hdr
    ws_dash.cell(row=11, column=3).fill = fill_subhdr
    ws_dash.cell(row=11, column=3).alignment = Alignment(horizontal="right")

    for idx, est in enumerate(estab_list, start=12):
        ws_dash.cell(row=idx, column=2, value=est).font = f_reg
        ws_dash.cell(row=idx, column=2).border = thin_border
        c_cnt = ws_dash.cell(row=idx, column=3, value=f'=COUNTIF(INDIRECT({kpi_indirect("F4:F1000")}), "{est}")')
        c_cnt.font, c_cnt.alignment, c_cnt.border = f_bold, Alignment(horizontal="right"), thin_border

    ws_dash.cell(row=9, column=5, value="Protocolos de Operación y Blindaje de Calidad").font = \
        Font(name=FUENTE, size=11, bold=True, color=c_navy)
    insts = [
        "1. Doble Identificador Correlativo: la Columna A (N° Envío GT) numera la valija para la posta; "
        "la Col B registra el folio hospitalario. El correlativo se asigna desde Apps Script (Codigo.gs) "
        "y nunca se reasigna a filas ya numeradas.",
        "2. Detección de Duplicados Blindada: la Columna B solo se pinta en ROJO si contiene un número "
        "repetido. Las celdas vacías permanecen limpias.",
        "3. Alerta Robusta de Fin de Receta: la Columna H calcula en tiempo real si el paciente está en su "
        "última cuota (12/12, 6/6, 1/1, ULTIMA) o sin receta vigente.",
        "4. Trazabilidad de Quiebres: al registrar 'pendiente por falta stock', la celda se colorea en ámbar "
        "y se anota el fármaco adeudado en la Columna N.",
        "5. Seguridad Térmica: toda prescripción con insulina o refrigerados (Col L = SÍ) se resalta en azul "
        "hielo para empaque en cooler térmico.",
        "6. Mes Activo: el Dashboard SIEMPRE lee de la hoja indicada en Parametros!G2 — al iniciar un mes "
        "nuevo (Generar Rollover), esa celda se actualiza sola y ninguna fórmula de acá se toca a mano.",
    ]
    for idx, text in enumerate(insts, start=11):
        ws_dash.cell(row=idx, column=5, value=text).font = f_reg

    ws_dash.column_dimensions['A'].width = 4
    ws_dash.column_dimensions['B'].width = 28
    ws_dash.column_dimensions['C'].width = 16
    ws_dash.column_dimensions['D'].width = 24
    ws_dash.column_dimensions['E'].width = 48
    ws_dash.column_dimensions['F'].width = 16
    ws_dash.column_dimensions['G'].width = 14
    ws_dash.column_dimensions['H'].width = 24
    ws_dash.column_dimensions['I'].width = 14
    ws_dash.column_dimensions['J'].width = 24

    # -------------------------------------------------------------
    # FORMATOS CONDICIONALES BLINDADOS
    # -------------------------------------------------------------
    for ws in (ws_sep, ws_new):
        max_r = ws.max_row
        ws.conditional_formatting.add(
            f"B4:B{max_r}",
            FormulaRule(formula=[f'AND(B4<>"", COUNTIF($B$4:$B${max_r}, B4)>1)'],
                        fill=PatternFill(start_color=c_alert_red, end_color=c_alert_red, fill_type="solid"),
                        font=Font(name=FUENTE, size=10, bold=True, color=c_alert_red_tx)))
        ws.conditional_formatting.add(
            f"H4:H{max_r}",
            CellIsRule(operator='equal', formula=['"⚠ ÚLTIMA RECETA / RENOVAR"'],
                       fill=PatternFill(start_color=c_alert_amber, end_color=c_alert_amber, fill_type="solid"),
                       font=Font(name=FUENTE, size=10, bold=True, color=c_alert_amber_tx)))
        ws.conditional_formatting.add(
            f"L4:L{max_r}",
            CellIsRule(operator='equal', formula=['"SÍ"'],
                       fill=PatternFill(start_color=c_ice, end_color=c_ice, fill_type="solid"),
                       font=Font(name=FUENTE, size=10, bold=True, color=c_navy)))
        ws.conditional_formatting.add(
            f"K4:K{max_r}",
            CellIsRule(operator='equal', formula=['"pendiente por falta stock"'],
                       fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                       font=Font(name=FUENTE, size=10, bold=True, color="B25900")))
        ws.conditional_formatting.add(
            f"D4:D{max_r}",
            CellIsRule(operator='equal', formula=['"⚠ RUT No Registrado"'],
                       fill=PatternFill(start_color=c_alert_red, end_color=c_alert_red, fill_type="solid"),
                       font=Font(name=FUENTE, size=10, bold=True, color=c_alert_red_tx)))

    nombre_salida = os.path.join(os.path.dirname(os.path.abspath(archivo_origen)) or ".",
                                  "GT_PITRUFQUEN_SISTEMA_DEFINITIVO_2026.xlsx")
    if os.path.exists(nombre_salida):
        bak = nombre_salida + ".bak"
        try:
            os.replace(nombre_salida, bak)
            print(f"[BACKUP] Versión anterior movida a: {bak}")
        except OSError as e:
            print(f"[AVISO] No se pudo respaldar la versión anterior ({e}) — se sobreescribe igual.")
    wb.save(nombre_salida)
    print(f"[OK] Archivo generado: {nombre_salida} ({os.path.getsize(nombre_salida)/1024:.1f} KB)")
    return nombre_salida


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--origen", help="Ruta al xlsx crudo de Gestión Territorial (default: autodetecta *GT*PITRUFQUEN*.xlsx)")
    ap.add_argument("--con-archivo", dest="con_archivo", action="store_true",
                     help="Incluye las hojas Archivo_* (Enero-Agosto + Sectorización). Por defecto NO se incluyen "
                          "— el archivo base original ya conserva ese historial completo, no se pierde por omitirlo acá.")
    a = ap.parse_args()
    archivo_origen = buscar_archivo_origen(a.origen)
    ejecutar_migracion(archivo_origen, incluir_archivo=a.con_archivo)


if __name__ == "__main__":
    main()
