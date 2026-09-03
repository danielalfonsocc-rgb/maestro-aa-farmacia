"""
Consolidación de hemogramas Clozapina para ingreso en bloque a plataforma MINSAL.
Hospital Pitrufquén (SSAS) — atención abierta.

Flujo:
  1) Filtra despachos de CLOZAPINA (Atención Abierta) del informe_completo_recetas*.csv.
  2) Parsea la posología (columna "Observación Médica Prescripción", formato mañana-tarde-noche
     en comprimidos, ej. "0-0-3") a mg/día.
  3) Extrae el hemograma (BAS/EOS/SEG/LINFO/MONO %, LEUCO, HTO, HB, GLOB ROJOS, PLAQ) desde el
     PDF de resultados de laboratorio HCE — el más reciente disponible para cada paciente
     (si no hay de julio, cae naturalmente al de junio, etc.: siempre se toma el último
     examen sin filtrar por mes). Carpeta local con un PDF por paciente, nombrado <RUN>.pdf.
  4) Agrega DOS hojas NUEVAS al Excel cada vez que corre — NO sobrescribe el archivo ni
     las hojas de días anteriores (si ya hay una hoja de HOY, esa sí se reemplaza, para no
     duplicar al correr dos veces el mismo día):
       - "DD-MM" (ej. "22-07"): columnas EXACTAS al formulario "Ingreso de Hemogramas en
         Bloque" (RUT, FECHA, DOSIS, BAS, EOS, MIELO, JUV, BAC, SEG, LINFO, MONO, LEUCO,
         HTO, HB, GLOB ROJOS, PLAQ, RAE, RAN, Fecha Despacho), ordenadas por Fecha Despacho
         de más antigua a más nueva. La ALARMA (RAN/RAE) la calcula y muestra la propia
         plataforma MINSAL al ingresar — este script no la determina ni la copia.
       - "DD-MM Revisar": RUN, Nombre, Fecha Despacho, Cuota Receta, Fecha Hemograma,
         Antigüedad Hemograma (días), Observaciones — para que la QF revise ANTES de pegar
         (dosis no interpretable, receta fraccionada, sin hemograma, hemograma antiguo, etc.).

ESCALA (confirmada con un caso real ingresado por el usuario 19/05/2026, RAE 1150,64 / RAN 2460,25):
  LEUCO, GLOB ROJOS y PLAQ van como conteo ABSOLUTO, no como aparecen en el PDF de laboratorio
  (que los reporta en miles/millones por uL):
    LEUCO      = leucocitos_pdf   x 1.000       (PDF "7.57" -> 7570)
    GLOB ROJOS = eritrocitos_pdf  x 1.000.000   (PDF "4.49" -> 4.490.000)
    PLAQ       = plaquetas_pdf    x 1.000       (PDF "287"  -> 287.000)
  RAE y RAN NO se copian del valor absoluto del PDF: la plataforma los calcula así, y replicamos
  la misma fórmula para que cuadren exacto:
    RAE = %EOS x LEUCO / 100
    RAN = %SEG x LEUCO / 100
  HB y HTO van tal como aparecen en el PDF (ya son la escala nativa esperada).
"""
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd

# Carpeta LOCAL (no versionada) donde se acumulan los reportes mensuales — trae
# RUT/nombre de paciente (Ley 19.628), así que NUNCA se copia a la carpeta del
# Escritorio (sincronizada a OneDrive); publicar_escritorio.py solo deja un
# acceso directo a esta carpeta, igual que con Recetas Cheque.
CLOZAPINA_REPORTES_DIR = Path(__file__).parent / "Clozapina_Reportes"

_MESES_ES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
             "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def ruta_archivo_mes(fecha: Optional[date] = None) -> Path:
    """Un archivo por mes (ej. 'Consolidado_Hemogramas_Clozapina_Julio_2026.xlsx'
    dentro de Clozapina_Reportes/); en agosto arranca uno nuevo, no se sigue
    acumulando sobre el de julio."""
    fecha = fecha or date.today()
    CLOZAPINA_REPORTES_DIR.mkdir(exist_ok=True)
    return CLOZAPINA_REPORTES_DIR / f"Consolidado_Hemogramas_Clozapina_{_MESES_ES[fecha.month]}_{fecha.year}.xlsx"


MG_POR_COMPRIMIDO_CLOZAPINA = 100.0
FACTOR_LEUCO = 1_000
FACTOR_ERITROCITOS = 1_000_000
FACTOR_PLAQUETAS = 1_000

COLUMNAS_MINSAL = [
    "RUT", "FECHA", "DOSIS", "BAS", "EOS", "MIELO", "JUV", "BAC", "SEG", "LINFO", "MONO",
    "LEUCO", "HTO", "HB", "GLOB ROJOS", "PLAQ", "RAE", "RAN", "Fecha Despacho", "Nombre",
]
# "Nombre" va AL FINAL, después de las 18 columnas exactas del formulario MINSAL +
# "Fecha Despacho" — es solo para identificar al paciente de un vistazo al revisar/
# pegar (la planilla ya es confidencial, gitignored, nunca sube al Escritorio/GitHub);
# no se toca el orden de las columnas que sí se copian/pegan tal cual a la plataforma.
# La alarma (RAN/RAE) la calcula y muestra la propia plataforma MINSAL al ingresar
# los datos — no la determinamos nosotros, por eso no hay columnas de alarma aquí.
COLUMNAS_OBSERVACION = [
    "RUN", "Nombre", "Fecha Despacho", "Cuota Receta", "Fecha Hemograma",
    "Antigüedad Hemograma (días)", "Observaciones",
]


# ============================================================
# 1) PARSEO DE POSOLOGÍA -> mg/día
# ============================================================

_FRACCIONES = {"1/2": 0.5, "1/4": 0.25, "3/4": 0.75}


def _token_a_numero(tok: str) -> Optional[float]:
    tok = tok.strip()
    if not tok:
        return 0.0
    # "2 1/2", "21/2" (sin espacio), "1/2", "2.5", "3"
    m = re.match(r"^(\d+)\s*(1/2|1/4|3/4)$", tok)
    if m:
        return float(m.group(1)) + _FRACCIONES[m.group(2)]
    if tok in _FRACCIONES:
        return _FRACCIONES[tok]
    m = re.match(r"^(\d+)(1/2|1/4|3/4)$", tok)  # "31/2" -> 3 + 1/2
    if m:
        return float(m.group(1)) + _FRACCIONES[m.group(2)]
    try:
        return float(tok.replace(",", "."))
    except ValueError:
        return None


def parsear_posologia(texto: Optional[str], mg_comprimido: float = MG_POR_COMPRIMIDO_CLOZAPINA) -> Optional[float]:
    """'0-0-3' -> 300.0 mg/día (mañana-tarde-noche, en comprimidos).
    Devuelve None si el texto no es un patrón de 3 tomas interpretable (ej. 'SEGÚN INDICACIÓN',
    instrucciones en prosa) — en ese caso, dejar DOSIS en blanco y marcar observación."""
    if not texto or not isinstance(texto, str):
        return None
    t = texto.strip().upper()
    t = re.sub(r"\bVO\b", "", t).strip()
    partes = t.split("-")
    if len(partes) != 3:
        return None
    valores = [_token_a_numero(p) for p in partes]
    if any(v is None for v in valores):
        return None
    return sum(valores) * mg_comprimido


# ============================================================
# 2) DESPACHOS DE CLOZAPINA DESDE EL CSV DE RECETAS
# ============================================================

def _leer_recetas(origen: str) -> pd.DataFrame:
    """Igual convención que maestro_aa.py: si 'origen' es una carpeta, lee TODOS
    los informe_completo_recetas*.csv que haya (bloques de 30 días de AUTO_SSASUR)
    y deduplica por 'ID Receta Detalle' para que los bloques solapados no
    doble-cuenten. Si es un archivo, lo lee directo (modo puntual/pruebas)."""
    ruta = Path(origen)
    if ruta.is_dir():
        archivos = sorted(ruta.glob("informe_completo_recetas*.csv"))
        if not archivos:
            raise FileNotFoundError(f"Sin informe_completo_recetas*.csv en {ruta}")
        partes = [pd.read_csv(f, sep=";", encoding="latin1", dtype=str, on_bad_lines="skip") for f in archivos]
        df = pd.concat(partes, ignore_index=True)
        if "ID Receta Detalle" in df.columns:
            df = df.drop_duplicates(subset=["ID Receta Detalle"], keep="first")
        return df
    return pd.read_csv(ruta, sep=";", encoding="latin1", dtype=str, on_bad_lines="skip")


def cargar_despachos_clozapina(csv_recetas: str, desde: Optional[date] = None, hasta: Optional[date] = None) -> pd.DataFrame:
    df = _leer_recetas(csv_recetas)

    col_presc = next(c for c in df.columns if c.startswith("Prescripci"))
    col_estado_presc = next(c for c in df.columns if c.startswith("Estado Prescripci"))
    col_fentrega = next(c for c in df.columns if c.startswith("Fecha Entrega Receta"))
    col_obs_medica = next(c for c in df.columns if c.startswith("Observaci") and "M" in c and "Prescripci" in c)

    mask = (
        df[col_presc].str.contains("CLOZAPINA", case=False, na=False)
        & (df["Procedencia"].str.strip().str.upper() == "ATENCION ABIERTA")
        & (df[col_estado_presc].str.strip().str.upper() == "ENTREGADO")
    )
    sub = df[mask].copy()
    sub["fecha_despacho"] = pd.to_datetime(sub[col_fentrega], format="%d/%m/%Y", errors="coerce")

    if desde is not None:
        sub = sub[sub["fecha_despacho"] >= pd.Timestamp(desde)]
    if hasta is not None:
        sub = sub[sub["fecha_despacho"] <= pd.Timestamp(hasta)]

    sub["dosis_mg_dia"] = sub[col_obs_medica].apply(parsear_posologia)
    sub["cuota_receta"] = sub["Periodo"]  # ej. "1 de 2"
    sub["run_normalizado"] = sub["RUN"].str.replace(r"\s*-\s*", "-", regex=True).str.strip()

    cols = ["run_normalizado", "Nombre", "Apellido Paterno", "Apellido Materno",
            "fecha_despacho", "dosis_mg_dia", "cuota_receta", col_obs_medica, "Cantidad Pendiente"]
    return sub[cols].rename(columns={col_obs_medica: "posologia_texto"}).sort_values("fecha_despacho")


# ============================================================
# 3) EXTRACCIÓN DE HEMOGRAMA DESDE PDF
# ============================================================

# Alias por campo: distintos laboratorios de la red usan distinta etiqueta para el mismo
# analito ("NEUTROFILOS" en el Sysmex propio del hospital, "SEGMENTADOS" en el formato
# antiguo de CESFAM Freire). Se prueba cada alias en orden hasta encontrar uno que matchee.
_ALIAS_DIFERENCIAL = {
    "basofilos": ["BASOFILOS"],
    "eosinofilos": ["EOSINOFILOS"],
    "neutrofilos": ["NEUTROFILOS", "SEGMENTADOS"],
    "linfocitos": ["LINFOCITOS"],
    "monocitos": ["MONOCITOS"],
}


def _num_pdf(s: str) -> float:
    return float(s.replace(",", "."))


def _quitar_tildes(s: str) -> str:
    """Normaliza tildes ('Basófilos' -> 'Basofilos') para que los regex (escritos sin tilde)
    matcheen también el formato del laboratorio PROPIO del hospital, que a diferencia de los
    laboratorios de la red usa etiquetas acentuadas (ej. 'Neutrófilos %' en vez de
    'NEUTROFILOS %'). Sin esto, el hemograma más común (el del propio hospital) se leía a
    medias: LEUCO/HTO/HB/GLOB ROJOS salían bien (esas etiquetas no llevan tilde) pero
    BAS/EOS/SEG/LINFO/MONO y la fecha de toma de muestra quedaban en blanco."""
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


# Relleno entre la etiqueta y el número: absorbe flechas de alarma ↑/↓, espacios, asteriscos
# u otros glifos que el PDF real pueda usar — pero NO puede cruzar la palabra "PENDIENTE"
# (examen aún sin resultado del laboratorio). Sin ese bloqueo, "Hematocrito Pendiente % 35.0
# - 44.0" saltaba "Pendiente" entero y capturaba 35.0 (el LÍMITE del rango de referencia) como
# si fuera el resultado real — un valor clínico inventado con apariencia normal, peor que
# dejarlo en blanco. Detectado en vivo (29-07, paciente con examen "Pendiente").
_RELLENO = r"(?:(?!PENDIENTE)[^\d\n])*"
_RELLENO_NG = r"(?:(?!PENDIENTE)[^\d\n])*?"


def _buscar(texto: str, patron: str) -> Optional[float]:
    m = re.search(patron, texto, re.IGNORECASE)
    return _num_pdf(m.group(1)) if m else None


def extraer_hemograma_texto(texto: str) -> dict:
    """Parsea el texto plano del PDF de resultados HCE (sección
    'HEMOGRAMA FORMULA DIFERENCIAL AUTOMATIZADA' u homóloga) y devuelve un dict con los
    campos que necesita motor_reglas_clozapina_v3 + la planilla MINSAL. Reconoce el formato
    del laboratorio propio del hospital (Sysmex, etiquetas acentuadas) y el de laboratorios
    de la red con layout más antiguo (ej. CESFAM Freire: sin filas de conteo absoluto,
    'SEGMENTADOS' en vez de 'NEUTROFILOS', recuentos ya en escala absoluta '/mm3')."""
    out: dict = {}
    texto = _quitar_tildes(texto)

    out["hematocrito"] = _buscar(texto, rf"\bHEMATOCRITO{_RELLENO}([\d]+(?:[.,]\d+)?)")
    out["hemoglobina"] = _buscar(texto, rf"\bHEMOGLOBINA{_RELLENO}([\d]+(?:[.,]\d+)?)")
    out["eritrocitos"] = _buscar(texto, rf"RECUENTO DE ERITROCITOS{_RELLENO}([\d]+(?:[.,]\d+)?)")

    # LEUCO y PLAQ: el Sysmex propio reporta ya en la escala x1.000 que usa el resto del
    # script ("RECUENTO DE LEUCOCITOS 6.15 10^3/uL" -> se multiplica x1.000 más abajo). Los
    # laboratorios de red con formato antiguo (CESFAM Freire) reportan el conteo ABSOLUTO
    # directo ("REC. LEUCOCITOS 5230 /mm3") — hay que dividir por 1.000 al leerlo para que
    # quede en la MISMA escala, o el factor x1.000 de construir_fila lo infla 1000 veces.
    out["leucocitos"] = _buscar(texto, rf"RECUENTO DE LEUCOCITOS{_RELLENO}([\d]+(?:[.,]\d+)?)")
    if out["leucocitos"] is None:
        v = _buscar(texto, rf"REC\.?\s*LEUCOCITOS{_RELLENO}([\d]+(?:[.,]\d+)?)\s*/\s*mm3")
        out["leucocitos"] = v / 1000 if v is not None else None
    # El laboratorio propio del hospital omite el "DE" en esta etiqueta puntual ("Recuento
    # plaquetas", a diferencia de "Recuento DE Leucocitos"/"Recuento DE eritrocitos" que sí
    # lo llevan) — "DE" opcional para cubrir ambos.
    out["plaquetas"] = _buscar(texto, rf"RECUENTO\s+(?:DE\s+)?PLAQUETAS{_RELLENO}([\d]+(?:[.,]\d+)?)")
    if out["plaquetas"] is None:
        v = _buscar(texto, rf"REC\.?\s*PLAQUETAS{_RELLENO}([\d]+(?:[.,]\d+)?)\s*/\s*mm3")
        out["plaquetas"] = v / 1000 if v is not None else None

    for campo, alias in _ALIAS_DIFERENCIAL.items():
        pct = abso = None
        for lab in alias:
            if pct is None:
                # Formato de dos filas (Sysmex): "NEUTROFILOS % 27.10 % ..." — el % va
                # PEGADO a la etiqueta, lo que distingue esta fila de la fila de conteo
                # absoluto de más abajo (misma etiqueta, sin % después).
                pct = _buscar(texto, rf"{lab}\s*%{_RELLENO}([\d]+(?:[.,]\d+)?)")
            if pct is None:
                # Formato de una sola fila (CESFAM Freire): "SEGMENTADOS 73.5 % 50 - 68" — el
                # número va ANTES del %, no pegado a la etiqueta.
                pct = _buscar(texto, rf"{lab}(?!\s*%){_RELLENO_NG}([\d]+(?:[.,]\d+)?)\s*%")
            if abso is None:
                abso = _buscar(texto, rf"{lab}(?!\s*%){_RELLENO}([\d]+(?:[.,]\d+)?)\s*10")
            if pct is not None and abso is not None:
                break
        out[f"{campo}_pct"] = pct
        out[f"{campo}_abs"] = abso

    # El PDF real usa GUION en este campo ("17-07-2026"), no barra, a diferencia de casi
    # todas las otras fechas del documento — confirmado en vivo con un caso real. Se acepta
    # cualquiera de los dos separadores por robustez. IGNORECASE porque el laboratorio propio
    # del hospital capitaliza distinto ("Fecha/hora de T. Muestra" en vez de "Fecha/Hora de
    # T. muestra").
    m_fecha = re.search(r"Fecha/Hora de T\.\s*muestra\s*:?\s*(\d{2})[/-](\d{2})[/-](\d{4})", texto, re.IGNORECASE)
    if not m_fecha:
        # Formato antiguo de red (CESFAM Freire): sin el prefijo "Fecha/Hora de T.", solo
        # "Toma muestra : dd/mm/aaaa".
        m_fecha = re.search(r"\bToma\s+muestra\s*:?\s*(\d{2})[/-](\d{2})[/-](\d{4})", texto, re.IGNORECASE)
    out["fecha_muestra"] = date(int(m_fecha.group(3)), int(m_fecha.group(2)), int(m_fecha.group(1))) if m_fecha else None

    m_run = re.search(r"\bRUN\s*:?\s*([\d.]+-?[\dkK])", texto)
    out["run_pdf"] = m_run.group(1) if m_run else None

    return out


def extraer_hemograma_pdf(ruta_pdf: Path) -> dict:
    import pdfplumber

    texto_completo = []
    with pdfplumber.open(ruta_pdf) as pdf:
        for pagina in pdf.pages:
            t = pagina.extract_text() or ""
            texto_completo.append(t)
    return extraer_hemograma_texto("\n".join(texto_completo))


# ============================================================
# 4) CONSOLIDACIÓN + MOTOR DE REGLAS
# ============================================================

@dataclass
class FilaConsolidada:
    run: str
    nombre: str
    fecha_despacho: Optional[date]
    dosis_mg_dia: Optional[float]
    cuota_receta: str
    posologia_texto: str
    hemograma: Optional[dict]  # salida de extraer_hemograma_texto, o None si no se encontró


def _fmt_fecha_minsal(f: Optional[date]) -> str:
    return f.strftime("%d-%m-%Y") if f else ""


def construir_fila(fila: FilaConsolidada, fecha_ref: date) -> tuple[dict, dict]:
    """Devuelve (fila_minsal, fila_observacion) para una fila consolidada."""
    hg = fila.hemograma or {}

    leuco_abs = hg["leucocitos"] * FACTOR_LEUCO if hg.get("leucocitos") is not None else None
    glob_rojos_abs = hg["eritrocitos"] * FACTOR_ERITROCITOS if hg.get("eritrocitos") is not None else None
    plaq_abs = hg["plaquetas"] * FACTOR_PLAQUETAS if hg.get("plaquetas") is not None else None

    eos_pct = hg.get("eosinofilos_pct")
    seg_pct = hg.get("neutrofilos_pct")
    # La plataforma calcula RAE/RAN como %xLEUCO_abs/100 (verificado con caso real: RAE 1150,64 =
    # 15,2 x 7570 / 100; RAN 2460,25 = 32,5 x 7570 / 100) — replicamos la misma fórmula, no se
    # copia el valor absoluto que trae el PDF de laboratorio (puede diferir por redondeo).
    rae = round(eos_pct * leuco_abs / 100, 2) if (eos_pct is not None and leuco_abs is not None) else None
    ran = round(seg_pct * leuco_abs / 100, 2) if (seg_pct is not None and leuco_abs is not None) else None

    fila_minsal = {
        "RUT": fila.run,
        "FECHA": _fmt_fecha_minsal(hg.get("fecha_muestra")),
        "DOSIS": fila.dosis_mg_dia if fila.dosis_mg_dia is not None else "",
        "BAS": hg.get("basofilos_pct", ""),
        "EOS": eos_pct if eos_pct is not None else "",
        "MIELO": 0,
        "JUV": 0,
        "BAC": 0,
        "SEG": seg_pct if seg_pct is not None else "",
        "LINFO": hg.get("linfocitos_pct", ""),
        "MONO": hg.get("monocitos_pct", ""),
        "LEUCO": leuco_abs if leuco_abs is not None else "",
        "HTO": hg.get("hematocrito", ""),
        "HB": hg.get("hemoglobina", ""),
        "GLOB ROJOS": glob_rojos_abs if glob_rojos_abs is not None else "",
        "PLAQ": plaq_abs if plaq_abs is not None else "",
        "RAE": rae if rae is not None else "",
        "RAN": ran if ran is not None else "",
        "Fecha Despacho": _fmt_fecha_minsal(fila.fecha_despacho),
        "Nombre": fila.nombre,
    }

    # Observaciones administrativas para que la QF revise antes de ingresar — la
    # alarma clínica (RAN/RAE) la determina la propia plataforma MINSAL al
    # guardar, no se calcula ni se muestra acá.
    observaciones = []
    if fila.dosis_mg_dia is None:
        observaciones.append(f"DOSIS no interpretable de posología: '{fila.posologia_texto}' — completar manualmente.")
    if fila.cuota_receta and fila.cuota_receta.strip() != "1 de 1":
        observaciones.append(f"Receta fraccionada, cuota {fila.cuota_receta} — verificar continuidad de tratamiento.")

    antiguedad = None
    if fila.hemograma is None:
        observaciones.append("Sin hemograma encontrado en HCE para este paciente — NO ingresar hasta obtenerlo.")
    elif hg.get("fecha_muestra"):
        antiguedad = (fecha_ref - hg["fecha_muestra"]).days
        if antiguedad > 7:
            observaciones.append(f"Hemograma con {antiguedad} días de antigüedad — verificar vigencia según fase de tratamiento.")
    else:
        observaciones.append("No se pudo determinar la fecha de toma de muestra del hemograma — verificar el PDF manualmente.")

    fila_obs = {
        "RUN": fila.run,
        "Nombre": fila.nombre,
        "Fecha Despacho": _fmt_fecha_minsal(fila.fecha_despacho),
        "Cuota Receta": fila.cuota_receta,
        "Fecha Hemograma": _fmt_fecha_minsal(hg.get("fecha_muestra")),
        "Antigüedad Hemograma (días)": antiguedad if antiguedad is not None else "",
        "Observaciones": " | ".join(observaciones),
    }
    return fila_minsal, fila_obs


# ============================================================
# 5) EXCEL DE SALIDA — mismo estilo tinta-económica que el resto del proyecto
#    (aa_colors.py / pedido_fusion.py): título, subtítulo, header en color,
#    bordes finos, filas con color según estado, panes congelados, autofiltro.
# ============================================================

_ANCHOS_MINSAL = {
    "RUT": 13, "FECHA": 11, "DOSIS": 8, "BAS": 7, "EOS": 7, "MIELO": 7, "JUV": 7,
    "BAC": 7, "SEG": 7, "LINFO": 7, "MONO": 7, "LEUCO": 9, "HTO": 7, "HB": 7,
    "GLOB ROJOS": 12, "PLAQ": 10, "RAE": 9, "RAN": 9, "Fecha Despacho": 13, "Nombre": 30,
}
_ANCHOS_REVISAR = {
    "RUN": 13, "Nombre": 30, "Fecha Despacho": 13, "Cuota Receta": 12,
    "Fecha Hemograma": 14, "Antigüedad Hemograma (días)": 12, "Observaciones": 90,
}
_COLS_CENTRADAS_MINSAL = {"FECHA", "DOSIS", "BAS", "EOS", "MIELO", "JUV", "BAC", "SEG",
                           "LINFO", "MONO", "LEUCO", "HTO", "HB", "GLOB ROJOS", "PLAQ",
                           "RAE", "RAN", "Fecha Despacho"}


def _estilos():
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin = Side(style="thin", color="DCDCDC")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "fill_header": PatternFill("solid", fgColor="D1FAF5"),
        "font_header": Font(bold=True, color="065F46", name="Arial", size=10),
        "font_titulo": Font(bold=True, size=12, color="065F46", name="Arial"),
        "font_subtitulo": Font(italic=True, size=9, color="555555", name="Arial"),
        "font_dato": Font(name="Arial", size=10, color="374151"),
        "fill_banda": PatternFill("solid", fgColor="F9FAFB"),
        # "A Revisar": color según severidad — mismo criterio tenue que aa_colors.CRIT_FILL_HEX
        "fill_critico": PatternFill("solid", fgColor="F4B3B3"),
        "font_critico": Font(bold=True, color="7F1D1D", name="Arial", size=10),
        "fill_alerta": PatternFill("solid", fgColor="FFF3E0"),
        "font_alerta": Font(color="7C4A03", name="Arial", size=10),
        "fill_ok": PatternFill("solid", fgColor="E8F5E9"),
        "font_ok": Font(color="1B5E20", name="Arial", size=10),
        "Alignment": Alignment,
    }


def _titulo_subtitulo(ws, titulo: str, subtitulo: str, ncols: int, est: dict):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, titulo)
    c.font = est["font_titulo"]
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, subtitulo)
    c.font = est["font_subtitulo"]
    c.alignment = est["Alignment"](wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30


def _header_fila3(ws, columnas: list[str], est: dict):
    for j, col in enumerate(columnas, 1):
        c = ws.cell(3, j, col)
        c.fill = est["fill_header"]
        c.font = est["font_header"]
        c.border = est["border"]
        c.alignment = est["Alignment"](horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28


def _rematar_hoja(ws, columnas: list[str], anchos: dict, n_filas: int, est: dict):
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(col, max(10, len(col) + 2))
    ws.freeze_panes = "A4"
    if n_filas:
        last = 3 + n_filas
        ws.auto_filter.ref = f"A3:{get_column_letter(len(columnas))}{last}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _escribir_hoja_ingreso(ws, filas_minsal_grupo: list[dict], mes_legible: str, fecha_despacho_txt: str,
                            fecha_generacion: date, est: dict):
    _titulo_subtitulo(
        ws, f"INGRESO HEMOGRAMAS MINSAL — CLOZAPINA · {mes_legible}",
        f"Pacientes despachados el {fecha_despacho_txt} (Atención Abierta) · copiar y pegar tal cual en "
        f"«Ingreso de Hemogramas en Bloque» · generado {fecha_generacion.strftime('%d-%m-%Y')} · la alarma "
        f"RAN/RAE la calcula la propia plataforma MINSAL, no este reporte.",
        len(COLUMNAS_MINSAL), est,
    )
    _header_fila3(ws, COLUMNAS_MINSAL, est)
    for i, fm in enumerate(filas_minsal_grupo, 4):
        for j, col in enumerate(COLUMNAS_MINSAL, 1):
            c = ws.cell(i, j, fm.get(col, ""))
            c.border = est["border"]
            c.font = est["font_dato"]
            if i % 2 == 0:
                c.fill = est["fill_banda"]
            if col in _COLS_CENTRADAS_MINSAL:
                c.alignment = est["Alignment"](horizontal="center")
            if col == "RUT":
                c.number_format = "@"
            elif col in ("LEUCO", "GLOB ROJOS", "PLAQ"):
                c.number_format = "0"
            elif col in ("DOSIS", "BAS", "EOS", "SEG", "LINFO", "MONO", "HTO", "HB", "RAE", "RAN"):
                c.number_format = "0.00"
    _rematar_hoja(ws, COLUMNAS_MINSAL, _ANCHOS_MINSAL, len(filas_minsal_grupo), est)


def _escribir_hoja_revisar(ws, filas_obs_grupo: list[dict], mes_legible: str, fecha_despacho_txt: str,
                            fecha_generacion: date, est: dict):
    _titulo_subtitulo(
        ws, f"A REVISAR — CLOZAPINA · {mes_legible}",
        f"Pacientes despachados el {fecha_despacho_txt} · revisar ANTES de pegar en MINSAL · generado "
        f"{fecha_generacion.strftime('%d-%m-%Y')} · rosa = sin hemograma (no ingresar), ámbar = algo que "
        f"verificar, verde = sin observaciones.",
        len(COLUMNAS_OBSERVACION), est,
    )
    _header_fila3(ws, COLUMNAS_OBSERVACION, est)
    for i, fo in enumerate(filas_obs_grupo, 4):
        obs_txt = str(fo.get("Observaciones", "") or "")
        if "Sin hemograma" in obs_txt:
            fill, font = est["fill_critico"], est["font_critico"]
        elif obs_txt.strip():
            fill, font = est["fill_alerta"], est["font_alerta"]
        else:
            fill, font = est["fill_ok"], est["font_ok"]
        for j, col in enumerate(COLUMNAS_OBSERVACION, 1):
            c = ws.cell(i, j, fo.get(col, ""))
            c.border = est["border"]
            c.fill = fill
            c.font = font
            if col == "RUN":
                c.number_format = "@"
            if col in ("Fecha Despacho", "Cuota Receta", "Fecha Hemograma", "Antigüedad Hemograma (días)"):
                c.alignment = est["Alignment"](horizontal="center")
            if col == "Observaciones":
                c.alignment = est["Alignment"](wrap_text=True, vertical="center")
    _rematar_hoja(ws, COLUMNAS_OBSERVACION, _ANCHOS_REVISAR, len(filas_obs_grupo), est)


def guardar_excel(filas_minsal: list[dict], filas_obs: list[dict], destino: str, fecha_generacion: Optional[date] = None):
    """Crea/reemplaza DOS hojas POR CADA FECHA DE DESPACHO distinta presente en
    los datos (no por el día en que corre el script): 'DD-MM' (listo para
    copiar/pegar, pacientes despachados ese día) y 'DD-MM Revisar' (contexto
    para la QF). Si ya existe la hoja de una fecha de despacho (se re-corrió el
    consolidado más tarde), la reemplaza en vez de duplicarla — así el archivo
    mensual siempre queda con la misma estructura sin importar qué día se
    corrió el script. Estilo visual: mismo criterio tinta-económica del resto
    del proyecto (aa_colors.py / pedido_fusion.py)."""
    from openpyxl import Workbook, load_workbook

    est = _estilos()
    fecha_generacion = fecha_generacion or date.today()

    # Agrupar filas por su propia "Fecha Despacho" (dd-mm-yyyy), no por la
    # fecha de corrida del script.
    grupos: dict[date, list[int]] = {}
    for idx, fm in enumerate(filas_minsal):
        txt = fm.get("Fecha Despacho") or ""
        try:
            fd = datetime.strptime(txt, "%d-%m-%Y").date()
        except ValueError:
            fd = None
        grupos.setdefault(fd, []).append(idx)

    destino_path = Path(destino)
    wb = load_workbook(destino) if destino_path.exists() else Workbook()
    primera_hoja_nueva = not destino_path.exists()
    if primera_hoja_nueva:
        hoja_default = wb.active  # se borra al final si quedaron hojas reales

    for fecha_despacho, indices in sorted(grupos.items(), key=lambda kv: (kv[0] is None, kv[0])):
        if fecha_despacho is None:
            sufijo = "SIN-FECHA"
            fecha_txt = "fecha de despacho no determinada"
            mes_legible = f"{_MESES_ES[fecha_generacion.month]} {fecha_generacion.year}"
        else:
            sufijo = fecha_despacho.strftime("%d-%m")
            fecha_txt = fecha_despacho.strftime("%d-%m-%Y")
            mes_legible = f"{_MESES_ES[fecha_despacho.month]} {fecha_despacho.year}"
        nombre_ingreso, nombre_revisar = sufijo, f"{sufijo} Revisar"

        for nombre in (nombre_ingreso, nombre_revisar):
            if nombre in wb.sheetnames:
                del wb[nombre]  # re-corrida -> reemplaza la hoja de esa fecha de despacho, no duplica

        grupo_minsal = [filas_minsal[i] for i in indices]
        grupo_obs = [filas_obs[i] for i in indices]

        ws1 = wb.create_sheet(nombre_ingreso)
        _escribir_hoja_ingreso(ws1, grupo_minsal, mes_legible, fecha_txt, fecha_generacion, est)
        ws2 = wb.create_sheet(nombre_revisar)
        _escribir_hoja_revisar(ws2, grupo_obs, mes_legible, fecha_txt, fecha_generacion, est)

    if primera_hoja_nueva and hoja_default.title in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb[hoja_default.title]  # borrar la hoja "Sheet" en blanco que crea Workbook() por defecto

    _ordenar_hojas_por_fecha(wb)
    wb.save(destino)


def _ordenar_hojas_por_fecha(wb):
    """Reordena las pestañas 'DD-MM' / 'DD-MM Revisar' cronológicamente — si se
    re-corre un día anterior después de que ya existan días más nuevos (ej. para
    reintentar), la pestaña recién recreada no queda desordenada al final."""
    def _clave(nombre):
        base = nombre.replace(" Revisar", "")
        try:
            dd, mm = base.split("-")
            return (int(mm), int(dd), 1 if nombre.endswith(" Revisar") else 0)
        except ValueError:
            return (99, 99, 0)  # nombres que no calzan el patrón -> al final

    orden = sorted(wb.sheetnames, key=_clave)
    wb._sheets = [wb[nombre] for nombre in orden]


# ============================================================
# CLI
# ============================================================

def main():
    import argparse
    ap = argparse.ArgumentParser(description="Consolida hemogramas de Clozapina para ingreso en bloque al MINSAL.")
    ap.add_argument("csv_recetas", help="informe_completo_recetas*.csv")
    ap.add_argument("--hemogramas-dir", default="_hemogramas_clozapina",
                     help="Carpeta con un PDF por paciente, nombrado <RUN_sin_guion>.pdf")
    ap.add_argument("--desde", default=None)
    ap.add_argument("--hasta", default=None)
    ap.add_argument("--salida", default=None, help="Por defecto: Clozapina_Reportes/Consolidado_..._<Mes>_<Año>.xlsx")
    args = ap.parse_args()

    desde = datetime.strptime(args.desde, "%Y-%m-%d").date() if args.desde else None
    hasta = datetime.strptime(args.hasta, "%Y-%m-%d").date() if args.hasta else None
    hoy = date.today()
    salida = args.salida or str(ruta_archivo_mes(hoy))

    despachos = cargar_despachos_clozapina(args.csv_recetas, desde, hasta)
    por_paciente = despachos.sort_values("fecha_despacho").groupby("run_normalizado").last()
    por_paciente = por_paciente.reset_index().sort_values("fecha_despacho")  # planilla: más antigua -> más nueva

    hemogramas_dir = Path(args.hemogramas_dir)
    filas_minsal, filas_obs = [], []
    for _, row in por_paciente.iterrows():
        run = row["run_normalizado"]
        run_sin_guion = run.replace("-", "").replace(".", "")
        pdf_path = hemogramas_dir / f"{run_sin_guion}.pdf"
        hemograma = extraer_hemograma_pdf(pdf_path) if pdf_path.exists() else None

        fc = FilaConsolidada(
            run=run, nombre=f"{row['Nombre']} {row['Apellido Paterno']} {row['Apellido Materno']}",
            fecha_despacho=row["fecha_despacho"].date() if pd.notna(row["fecha_despacho"]) else None,
            dosis_mg_dia=row["dosis_mg_dia"], cuota_receta=row["cuota_receta"] or "",
            posologia_texto=row["posologia_texto"] or "", hemograma=hemograma,
        )
        fm, fo = construir_fila(fc, hoy)
        filas_minsal.append(fm)
        filas_obs.append(fo)

    guardar_excel(filas_minsal, filas_obs, salida)
    print(f"{len(filas_minsal)} pacientes -> {salida}")
    faltantes = sum(1 for fo in filas_obs if "Sin hemograma" in fo["Observaciones"])
    if faltantes:
        print(f"[AVISO] {faltantes} paciente(s) sin PDF de hemograma en {hemogramas_dir}/ — colócalos ahí (nombre '<RUN_sin_guion>.pdf') y vuelve a correr.")


if __name__ == "__main__":
    main()
