import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import streamlit as st
import pandas as pd
import numpy as np
import os, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from aa_colors import CRIT_FILL_HEX, crit_fill, crit_nivel, crit_hex
from sgli import calcular_sgli, to_markdown, FACTOR_CARGA_DEFAULT, UMBRAL_ESTRES

# PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ─────────────────────────────────────────────
# Ruta dinamica — funciona en cualquier PC Windows
WORK_DIR    = os.path.dirname(os.path.abspath(__file__))
XLS_MAESTRO = os.path.join(WORK_DIR, "Consolidado_AA_MAESTRO.xlsx")

st.set_page_config(
    page_title="Pedidos AA — Farmacia AT Abierta",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ═══════════════════════════════════════════════════════════════
   SISTEMA DE DISEÑO — paleta neutra, color solo para acción/estado
   Fondo gris muy claro · tarjetas blancas · texto carbón ·
   amplio espaciado · esquinas redondeadas · sombras suaves difusas
   ═══════════════════════════════════════════════════════════════ */

/* ── Base tipográfica + lienzo neutro ─────────────────────────── */
html, body, [class*="css"] {
    font-family:-apple-system,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
    color:#1F2937;                       /* gray-800 carbón */
}
.stApp { background:#F9FAFB; }           /* gray-50 lienzo */
.block-container { padding-top:1.4rem; padding-bottom:3rem; max-width:1320px; }
h1,h2,h3,h4 { color:#111827; }           /* gray-900 títulos */
h2 { font-weight:800; letter-spacing:-.01em; }
hr { border:0; border-top:1px solid #EEF0F4; margin:1.4rem 0; }
a { color:#0F766E; }                     /* teal = enlace/acción */

/* ── Encabezado de marca (único bloque con color de marca) ────── */
.app-hero {
    display:flex; align-items:center; gap:18px;
    background:linear-gradient(90deg,#0F766E 0%,#0E7490 100%);
    border-radius:18px; padding:20px 26px; margin-bottom:18px; color:#fff;
    box-shadow:0 10px 30px -14px rgba(15,118,110,.55);
}
.app-hero .ico {
    width:54px; height:54px; border-radius:16px; background:rgba(255,255,255,.18);
    display:flex; align-items:center; justify-content:center; font-size:27px; flex:none;
}
.app-hero .tit { font-size:1.55rem; font-weight:800; line-height:1.1; margin:0; color:#fff; }
.app-hero .sub { font-size:.9rem; opacity:.93; margin:4px 0 0; }
.app-hero .chip {
    margin-left:auto; text-align:right; flex:none;
    background:rgba(255,255,255,.16); border-radius:12px; padding:9px 15px;
}
.app-hero .chip .d { font-size:.78rem; opacity:.9; }
.app-hero .chip .w { font-size:.95rem; font-weight:700; }

/* ── Leyenda de criticidad — badges semánticos suaves ─────────── */
.leyenda { display:flex; flex-wrap:wrap; gap:9px; margin:0 0 22px; }
.leyenda .lg {
    border-radius:999px; padding:5px 13px; font-size:.78rem; font-weight:600;
    background:#F3F4F6; color:#4B5563; border:1px solid #E5E7EB;
}

/* ── Tarjetas de medicamento — blancas, riel de color = estado ── */
.tarjeta {
    background:#fff; border:1px solid #EEF0F4; border-radius:16px;
    padding:20px 24px; margin-bottom:14px;
    box-shadow:0 1px 2px rgba(16,24,40,.04), 0 10px 24px -16px rgba(16,24,40,.18);
    border-left:5px solid #CBD5E1;
}
.t-crit1 { border-left-color:#DC2626; }
.t-crit2 { border-left-color:#EA580C; }
.t-crit3 { border-left-color:#F59E0B; }
.t-ok    { border-left-color:#16A34A; }

/* ── Bloque-med (pestaña Buscar) — blanco, riel de estado ─────── */
.bloque-med {
    background:#fff; border:1px solid #EEF0F4; border-left:5px solid #94A3B8;
    padding:16px 20px; border-radius:14px; margin-bottom:10px;
    box-shadow:0 1px 2px rgba(16,24,40,.04);
}
.crit-1  { border-left-color:#DC2626; }
.crit-2  { border-left-color:#EA580C; }
.crit-3  { border-left-color:#F59E0B; }
.crit-ok { border-left-color:#16A34A; }

/* ── Números y etiquetas ──────────────────────────────────────── */
.num-grande { font-size:2.05rem; font-weight:800; color:#111827; line-height:1; }
.etiqueta   { font-size:.76rem; color:#6B7280; margin-top:4px; }

/* ── Cajas de acción — badges suaves ──────────────────────────── */
.accion-box {
    background:#ECFDF5; border:1px solid #D1FAE5; border-radius:10px;
    padding:8px 14px; margin-top:8px; font-size:.86rem; color:#065F46; font-weight:600;
}
.accion-ext {
    background:#FFF1F2; border:1px solid #FECDD3; border-radius:10px;
    padding:8px 14px; margin-top:6px; font-size:.86rem; color:#9D174D; font-weight:600;
}

/* ── Badges semánticos reutilizables ──────────────────────────── */
.badge { display:inline-block; border-radius:999px; padding:2px 11px;
         font-size:.76rem; font-weight:700; line-height:1.5; }
.badge-red    { background:#FEF2F2; color:#B91C1C; }
.badge-orange { background:#FFF7ED; color:#C2410C; }
.badge-amber  { background:#FFFBEB; color:#B45309; }
.badge-green  { background:#F0FDF4; color:#15803D; }

/* ── Métricas Streamlit — tarjetas blancas espaciosas ─────────── */
div[data-testid="stMetric"], div[data-testid="metric-container"] {
    background:#fff; border:1px solid #EEF0F4; border-radius:16px;
    padding:18px 20px;
    box-shadow:0 1px 2px rgba(16,24,40,.04), 0 8px 20px -16px rgba(16,24,40,.16);
}
div[data-testid="stMetricLabel"] p { color:#6B7280; font-weight:600; }

/* ── Botones — color vibrante reservado para acción ───────────── */
.stButton > button, div[data-testid="stDownloadButton"] > button {
    border-radius:10px; font-weight:600; padding:.55rem 1.1rem;
    border:1px solid #0F766E; background:#0F766E; color:#fff;
    box-shadow:0 1px 2px rgba(16,24,40,.06); transition:filter .12s ease;
}
.stButton > button:hover, div[data-testid="stDownloadButton"] > button:hover {
    filter:brightness(1.08); border-color:#0F766E; color:#fff;
}

/* ── Tablas — limpias, solo separadores horizontales tenues ───── */
div[data-testid="stDataFrame"] {
    border:1px solid #EEF0F4; border-radius:14px; overflow:hidden;
    box-shadow:0 1px 2px rgba(16,24,40,.04);
}

/* ── Pestañas ─────────────────────────────────────────────────── */
button[data-baseweb="tab"] { font-size:.92rem; font-weight:600; color:#6B7280; }
button[data-baseweb="tab"][aria-selected="true"] { color:#0F766E; }
div[data-baseweb="tab-list"] { border-bottom:1px solid #EEF0F4; gap:4px; }
div[data-baseweb="tab-highlight"] { background:#0F766E; }

/* ── Expanders / inputs ───────────────────────────────────────── */
div[data-testid="stExpander"] {
    border:1px solid #EEF0F4; border-radius:14px; background:#fff;
    box-shadow:0 1px 2px rgba(16,24,40,.04);
}
</style>
""", unsafe_allow_html=True)

# ─── Helpers ─────────────────────────────────────────────────────────────────
def semana_mes(d):
    return min((d.day - 1) // 7 + 1, 4)

# Las 4 funciones derivan del nivel unico (aa_colors.crit_nivel), de modo que
# reconocen tanto la escala de pedidos ('1-CRITICO'...) como la de faltantes
# ('[CRITICO]...'). Antes solo entendian el prefijo 'N-', y por eso la pestana
# Faltantes mostraba todo en verde/OK sin importar la criticidad real.
def crit_class(c):
    return {1: 'crit-1', 2: 'crit-2', 3: 'crit-3', 4: 'crit-3'}.get(crit_nivel(c), 'crit-ok')

def crit_emoji(c):
    return {1: '🔴', 2: '🟠', 3: '🟡', 4: '🟡'}.get(crit_nivel(c), '🟢')

def tarjeta_class(c):
    return {1: 't-crit1', 2: 't-crit2', 3: 't-crit3', 4: 't-crit3'}.get(crit_nivel(c), 't-ok')

def orden_crit(c):
    return crit_nivel(c)

# Estilo unificado de tablas: filas blancas + celda 'Criticidad' como badge
# semantico suave (fondo claro + texto del color del estado). Sustituye al
# pintado de fila completa para una lectura limpia tipo "solo lineas".
_BADGE_CRIT = {
    1: 'background-color:#FEF2F2;color:#B91C1C;font-weight:700',
    2: 'background-color:#FFF7ED;color:#C2410C;font-weight:700',
    3: 'background-color:#FFFBEB;color:#B45309;font-weight:700',
    4: 'background-color:#FFFBEB;color:#B45309;font-weight:700',
}
_BADGE_OK = 'background-color:#F0FDF4;color:#15803D;font-weight:700'

def estilo_tabla(df, col_crit='Criticidad'):
    """Devuelve un Styler con filas blancas y la columna de criticidad
    resaltada como badge semantico suave."""
    sty = df.style.set_properties(**{'background-color': '#ffffff', 'color': '#1F2937'})
    if col_crit in df.columns:
        sty = sty.apply(
            lambda col: [_BADGE_CRIT.get(crit_nivel(str(v)), _BADGE_OK) for v in col],
            subset=[col_crit],
        )
    return sty

# ─── Generador PDF reutilizable ──────────────────────────────────────────────
def build_pdf(titulo, subtitulo, df_data, col_config, orientacion='landscape'):
    """
    Genera un PDF carta con tabla formateada y columna 'Solicitado' en blanco.

    col_config: list de dicts con keys:
        name  → nombre columna en df_data
        label → cabecera visible en PDF
        width → ancho en cm
        align → 'left' | 'center' | 'right'
    orientacion: 'landscape' (apaisado) o 'portrait' (vertical)
    """
    buf = io.BytesIO()
    page = landscape(letter) if orientacion == 'landscape' else letter
    pw, ph = page

    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    style_titulo = ParagraphStyle('tit', parent=styles['Heading1'],
                                  fontSize=13, textColor=colors.HexColor('#1A237E'),
                                  spaceAfter=2, alignment=TA_LEFT)
    style_sub    = ParagraphStyle('sub', parent=styles['Normal'],
                                  fontSize=8,  textColor=colors.HexColor('#555555'),
                                  spaceAfter=6, alignment=TA_LEFT)
    style_cell   = ParagraphStyle('cel', parent=styles['Normal'],
                                  fontSize=7.5, leading=9, wordWrap='CJK')
    style_hdr    = ParagraphStyle('hdr', parent=styles['Normal'],
                                  fontSize=8, fontName='Helvetica-Bold',
                                  textColor=colors.white, alignment=TA_CENTER)

    # Columnas de datos + columna SOLICITADO al final
    all_cols    = col_config + [{'name': '_solicitado', 'label': 'Solicitado',
                                  'width': 2.8, 'align': 'center'}]
    col_widths  = [c['width']*cm for c in all_cols]

    # Cabecera
    header_row  = [Paragraph(c['label'], style_hdr) for c in all_cols]

    # Filas 1-CRITICO: texto blanco/negrita para que se lea sobre el rojo fuerte B71C1C
    style_cell_critico = ParagraphStyle('cel_crit', parent=style_cell,
                                         textColor=colors.white, fontName='Helvetica-Bold')

    data_rows = []
    row_crits = []
    for _, row in df_data.iterrows():
        crit = str(row.get('Criticidad', ''))
        base_style = style_cell_critico if crit_nivel(crit) == 1 else style_cell
        cells = []
        for c in col_config:
            v = row.get(c['name'], '')
            v = '' if pd.isna(v) else v
            align = TA_CENTER if c.get('align','left') == 'center' else TA_LEFT
            cells.append(Paragraph(str(v),
                         ParagraphStyle('c', parent=base_style, alignment=align)))
        cells.append('')  # columna Solicitado — vacía
        data_rows.append(cells)
        row_crits.append(crit)

    table_data = [header_row] + data_rows

    # Estilos de tabla
    t_style = [
        # Cabecera
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 8),
        ('ALIGN',       (0,0), (-1,0), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUND',(0,0),(-1,0), colors.HexColor('#1F4E78')),
        # Bordes
        ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
        ('LINEBELOW',   (0,0), (-1,0),  0.8, colors.HexColor('#AAAAAA')),
        # Columna Solicitado (última) — borde izquierdo más grueso
        ('LINEAFTER',   (-2,0), (-2,-1), 1.2, colors.HexColor('#999999')),
        # Padding
        ('TOPPADDING',  (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING',(0,0), (-1,-1), 4),
    ]
    # Colores por fila — misma paleta que los Excel (aa_colors.CRIT_FILL_HEX)
    for i, crit in enumerate(row_crits, 1):
        # Etiquetas de pedidos estan en CRIT_FILL_HEX; las de faltantes
        # ('[CRITICO]'...) caen al color por nivel (crit_hex) en vez de gris.
        t_style.append(('BACKGROUND', (0,i), (-1,i),
                         colors.HexColor('#' + (CRIT_FILL_HEX.get(crit) or crit_hex(crit)))))

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(t_style))

    # Línea separadora pie
    fecha_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    pie = Paragraph(
        f'Farmacia AT Abierta — Pitrufquen  ·  Generado: {fecha_str}  ·  '
        f'Total: {len(df_data)} registros',
        ParagraphStyle('pie', parent=styles['Normal'], fontSize=7,
                       textColor=colors.HexColor('#888888'), alignment=TA_CENTER)
    )

    story = [
        Paragraph(titulo, style_titulo),
        Paragraph(subtitulo, style_sub),
        HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1F4E78'), spaceAfter=6),
        table,
        Spacer(1, 0.4*cm),
        HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#CCCCCC'), spaceBefore=2),
        pie,
    ]

    doc.build(story)
    buf.seek(0)
    return buf


# ─── Carga de datos ───────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def cargar_datos():
    if not os.path.exists(XLS_MAESTRO):
        return None
    try:
        stock    = pd.read_excel(XLS_MAESTRO, sheet_name='Stock_AA',             engine='openpyxl')
        farm     = pd.read_excel(XLS_MAESTRO, sheet_name='Pedido_Farm_Bodega',   engine='openpyxl')
        bod      = pd.read_excel(XLS_MAESTRO, sheet_name='Pedido_Repos_Bodega',  engine='openpyxl')
        falt     = pd.read_excel(XLS_MAESTRO, sheet_name='Faltas_Farmacia_AA',   engine='openpyxl')
        falt_det = pd.read_excel(XLS_MAESTRO, sheet_name='Faltantes_Detalle_AA', engine='openpyxl')
        # Dialisis (solo recetas de nefrologos) — opcional para compatibilidad
        try:
            dial_farm = pd.read_excel(XLS_MAESTRO, sheet_name='Dialisis_Pedido_Farm', engine='openpyxl')
        except Exception:
            dial_farm = pd.DataFrame()
        try:
            dial_bod = pd.read_excel(XLS_MAESTRO, sheet_name='Dialisis_Pedido_Bod', engine='openpyxl')
        except Exception:
            dial_bod = pd.DataFrame()
        # SGLI (reposicion por estres / semana pico) — opcional para compatibilidad
        try:
            sgli = pd.read_excel(XLS_MAESTRO, sheet_name='SGLI_Estres', engine='openpyxl')
        except Exception:
            sgli = pd.DataFrame()
        mtime    = datetime.fromtimestamp(os.path.getmtime(XLS_MAESTRO))
        return {'stock': stock, 'farm': farm, 'bod': bod,
                'falt': falt, 'falt_det': falt_det,
                'dial_farm': dial_farm, 'dial_bod': dial_bod,
                'sgli': sgli, 'mtime': mtime}
    except Exception as e:
        st.error(f"Error cargando datos: {e}")
        return None

def _get(row, col, default=0):
    v = row.get(col, default)
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return default
    return v

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 💊 Pedidos AA")
    st.markdown("**Farmacia AT Abierta**")
    st.markdown("---")

    if st.button("🔄 Recargar datos"):
        st.cache_data.clear()
        st.rerun()

    datos = cargar_datos()
    if datos:
        horas_ant = (datetime.now() - datos['mtime']).total_seconds() / 3600
        if horas_ant < 4:
            st.success(f"✅ Datos al día\n{datos['mtime'].strftime('%d/%m/%Y %H:%M')}")
        elif horas_ant < 24:
            st.warning(f"⚠️ Datos de hace {int(horas_ant)}h\n{datos['mtime'].strftime('%d/%m/%Y %H:%M')}")
        else:
            dias = int(horas_ant // 24)
            st.error(f"🔴 Datos de hace **{dias} día{'s' if dias>1 else ''}**\n"
                     f"{datos['mtime'].strftime('%d/%m/%Y %H:%M')}\n"
                     "Ejecuta AUTO_SSASUR.bat para actualizar.")
    else:
        st.error("No se encontró el archivo maestro.\nEjecuta EJECUTAR_MAESTRO.bat primero.")
        st.stop()

    hoy    = datetime.now()
    semana = semana_mes(hoy)
    st.info(f"📅 {hoy.strftime('%d/%m/%Y')}  ·  Semana **S{semana}**")

    st.markdown("---")
    n_ped = len(st.session_state.get('pedido', {}))
    if n_ped:
        st.markdown(f"🛒 **{n_ped} medicamento{'s' if n_ped>1 else ''} en pedido**")
    else:
        st.caption("Tu pedido aparece en la pestaña 🔍 Buscar")

# ─── Datos ───────────────────────────────────────────────────────────────────
df_stock    = datos['stock']
df_farm     = datos['farm']
df_bod      = datos['bod']
df_falt     = datos['falt']
df_falt_det = datos['falt_det']
df_dial_farm = datos.get('dial_farm', pd.DataFrame())
df_dial_bod  = datos.get('dial_bod',  pd.DataFrame())
df_sgli_base = datos.get('sgli',      pd.DataFrame())
todos_meds = sorted(df_stock['Medicamento'].dropna().unique().tolist())

if 'pedido' not in st.session_state:
    st.session_state.pedido = {}

# ═══════════════════════════════════════════════════════════════════════
# ENCABEZADO DE MARCA + LEYENDA  (visible sobre todas las pestañas)
# ═══════════════════════════════════════════════════════════════════════
_dia_es = {'Mon':'lun','Tue':'mar','Wed':'mié','Thu':'jue','Fri':'vie','Sat':'sáb','Sun':'dom'}
_mes_es = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}
_fecha_chip = f"{_dia_es.get(hoy.strftime('%a'), '')} {hoy.day} {_mes_es.get(hoy.month, '')}"
st.markdown(f"""
<div class='app-hero'>
  <div class='ico'>💊</div>
  <div>
    <p class='tit'>Pedidos AA</p>
    <p class='sub'>Farmacia Atención Abierta · Hospital de Pitrufquén</p>
  </div>
  <div class='chip'>
    <div class='d'>{_fecha_chip}</div>
    <div class='w'>Semana S{semana}</div>
  </div>
</div>
<div class='leyenda'>
  <span class='lg'>🔴 Crítico — sin stock ni respaldo</span>
  <span class='lg'>🟠 Urgente</span>
  <span class='lg'>🟡 Moderado</span>
  <span class='lg'>🟢 Suficiente</span>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# PESTAÑAS PRINCIPALES
# ═══════════════════════════════════════════════════════════════════════
tab_pedido_bod, tab_pedido_farm, tab_dialisis, tab_faltantes, tab_sgli, tab_auditoria, tab_feedback = st.tabs([
    "📝  Pedido a Bodega AA",
    "🏭  Pedido a Bodega Fármacos",
    "💉  Diálisis",
    "🚨  Faltantes",
    "🚦  SGLI · Estrés",
    "🔬  Auditoría de prescripción",
    "💬  Diagnóstico y Sugerencias",
])
with tab_pedido_bod:
    st.markdown("## 📝 Pedido a Bodega AA")
    st.markdown(
        "Lista completa de medicamentos que **Farmacia AA debe solicitar a Bodega AA**, "
        "ordenados por criticidad. Ajusta las cantidades si es necesario y descarga el pedido formal."
    )
    st.markdown("---")

    # Preparar datos del pedido a bodega
    def _col(df, col):
        """Devuelve la columna como Series numérica; cero si no existe."""
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').fillna(0)
        return pd.Series(0, index=df.index, dtype=float)

    df_pb = df_farm.copy()
    df_pb['_nec']  = _col(df_pb, 'Necesidad_5D_Farm')
    df_pb['_tras'] = _col(df_pb, 'Traspaso_Posible')
    df_pb['_ord']  = df_pb['Criticidad'].apply(orden_crit) if 'Criticidad' in df_pb.columns \
                     else pd.Series(5, index=df_pb.index)

    # Solo los que tienen necesidad > 0
    df_pb = df_pb[df_pb['_nec'] > 0].sort_values(
        ['_ord', '_nec'], ascending=[True, False]
    ).reset_index(drop=True)

    if df_pb.empty:
        st.success("✅ No hay medicamentos pendientes de solicitar a Bodega AA en este momento.")
    else:
        # Métricas resumen
        n_crit    = len(df_pb[df_pb['_ord'] <= 1])
        n_urg     = len(df_pb[df_pb['_ord'] <= 2])
        total_ud  = int(df_pb['_nec'].sum())
        con_stock = len(df_pb[df_pb['_tras'] > 0])
        sin_stock = len(df_pb[df_pb['_tras'] == 0])

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Total medicamentos",   f"{len(df_pb)}")
        mc2.metric("Críticos + Urgentes",  f"{n_urg}")
        mc3.metric("Con stock en Bodega",  f"{con_stock}")
        mc4.metric("Sin stock en Bodega",  f"{sin_stock} ⚠️")

        st.markdown("---")

        # ── Tabla editable del pedido ─────────────────────────────────────────
        st.markdown("### Detalle del pedido — edita las cantidades si necesitas ajustar")

        # Construir dataframe para mostrar/editar
        filas_pb = []
        for _, row in df_pb.iterrows():
            nec   = int(float(_get(row, 'Necesidad_5D_Farm', 0)))
            tras  = int(float(_get(row, 'Traspaso_Posible', 0)))
            def_  = int(float(_get(row, 'Deficit_Farm', 0)))
            filas_pb.append({
                'Medicamento'      : str(_get(row, 'Medicamento', '')),
                'Criticidad'       : str(_get(row, 'Criticidad', '5-OK')),
                'Stock Farmacia'   : int(float(_get(row, 'Stock_Farm_Actual', 0))),
                'Cob. actual (dias)': round(float(_get(row, 'Cob_Farm_Actual_Dias', 0)), 1),
                'Disponible Bodega': tras,
                'Solicitar (ud)'   : nec,
                'Deficit externo'  : def_,
                'Accion'           : str(_get(row, 'Accion_1_Traspaso_Bodega', '') or ''),
            })

        df_tabla = pd.DataFrame(filas_pb)

        st.dataframe(
            estilo_tabla(df_tabla),
            use_container_width=True,
            hide_index=True,
            column_config={
                'Medicamento'       : st.column_config.TextColumn("Medicamento",         width="large"),
                'Criticidad'        : st.column_config.TextColumn("Criticidad",          width="small"),
                'Stock Farmacia'    : st.column_config.NumberColumn("Stock Farm.",        format="%d"),
                'Cob. actual (dias)': st.column_config.NumberColumn("Cob. actual",        format="%.1f d"),
                'Disponible Bodega' : st.column_config.NumberColumn("Disponible Bod.",    format="%d"),
                'Solicitar (ud)'    : st.column_config.NumberColumn("A solicitar",        format="%d ud"),
                'Deficit externo'   : st.column_config.NumberColumn("Deficit externo",    format="%d"),
                'Accion'            : st.column_config.TextColumn("Accion sugerida",      width="medium"),
            },
            height=min(50 + len(df_tabla) * 35, 600),
        )

        # ── Alertas importantes ───────────────────────────────────────────────
        sin_bod = df_tabla[df_tabla['Disponible Bodega'] == 0]
        if len(sin_bod):
            with st.expander(f"⚠️ {len(sin_bod)} medicamento(s) SIN STOCK en Bodega — requieren gestión externa", expanded=True):
                for _, r in sin_bod.iterrows():
                    crit = str(r['Criticidad'])
                    emo  = crit_emoji(crit)
                    st.markdown(
                        f"{emo} **{r['Medicamento']}** — "
                        f"Criticidad: `{crit}` — "
                        f"Necesidad: **{r['Solicitar (ud)']} ud.**"
                    )

        # ── Descarga Excel del pedido ─────────────────────────────────────────
        st.markdown("---")
        st.markdown("### Descargar pedido formal")

        def build_pedido_bodega_excel():
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedido a Bodega AA"

            fills_crit = {k: crit_fill(k) for k in CRIT_FILL_HEX}

            # Fila 1 — Título
            ws['A1'] = f'PEDIDO FARMACIA AA → BODEGA AA   |   S{semana}  {hoy.strftime("%d/%m/%Y")}'
            ws['A1'].fill = PatternFill('solid', fgColor='1A237E')
            ws['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
            ws.merge_cells('A1:H1')
            ws.row_dimensions[1].height = 28

            # Fila 2 — Info resumen
            ws['A2'] = f'Total: {len(df_tabla)} medicamentos  |  Con stock bodega: {con_stock}  |  Sin stock: {sin_stock}'
            ws['A2'].font = Font(italic=True, name='Arial', size=10, color='444444')
            ws.merge_cells('A2:H2')
            ws.row_dimensions[2].height = 16

            # Fila 3 — Encabezados
            headers = ['N°','Medicamento','Criticidad','Stock Farmacia',
                       'Cob. actual (dias)','Disponible Bodega','A Solicitar (ud)','Accion sugerida']
            hfill = PatternFill('solid', fgColor='1F4E78')
            hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[3].height = 30

            # Datos
            for ri, (_, row) in enumerate(df_tabla.iterrows(), 4):
                crit = str(row.get('Criticidad', ''))
                fill = fills_crit.get(crit, PatternFill('solid', fgColor='FFFFFF'))
                is_crit1 = crit == '1-CRITICO'
                vals = [ri - 3,
                        row.get('Medicamento', ''),
                        crit,
                        int(row.get('Stock Farmacia', 0) or 0),
                        round(float(row.get('Cob. actual (dias)', 0) or 0), 1),
                        int(row.get('Disponible Bodega', 0) or 0),
                        int(row.get('Solicitar (ud)', 0) or 0),
                        str(row.get('Accion', '') or '')]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.fill = fill
                    c.font = Font(name='Arial', size=10,
                                  bold=is_crit1,
                                  color='FFFFFF' if is_crit1 else '000000')
                    c.alignment = Alignment(vertical='center',
                                            wrap_text=(ci == 2 or ci == 8))

            # Anchos de columna
            for ci, w in enumerate([5, 50, 13, 13, 14, 15, 14, 35], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A4'

            # Sección de firma / autorización al final
            ultima = len(df_tabla) + 5
            ws.cell(row=ultima, column=1,
                    value='Solicitado por:').font = Font(name='Arial', size=10)
            ws.cell(row=ultima, column=4,
                    value='Firma Farmaceutico:').font = Font(name='Arial', size=10)
            ws.cell(row=ultima, column=7,
                    value='Fecha:').font = Font(name='Arial', size=10)

            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            return buf

        col_dl1, col_dl2, col_dl3 = st.columns([2, 2, 2])
        with col_dl1:
            fname_bod = f"Pedido_Bodega_AA_{hoy.strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="📥 Excel",
                data=build_pedido_bodega_excel(),
                file_name=fname_bod,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
        with col_dl2:
            _cols_pdf_bod = [
                {'name':'Medicamento',    'label':'Medicamento',        'width':7.0, 'align':'left'},
                {'name':'Criticidad',     'label':'Criticidad',         'width':2.8, 'align':'center'},
                {'name':'Stock Farmacia', 'label':'Stock Farm.',         'width':2.2, 'align':'center'},
                {'name':'Cob. actual (dias)', 'label':'Cob.(dias)',     'width':1.8, 'align':'center'},
                {'name':'Disponible Bodega','label':'Disp. Bodega',     'width':2.2, 'align':'center'},
                {'name':'Solicitar (ud)', 'label':'A Solicitar',        'width':2.2, 'align':'center'},
                {'name':'Accion',         'label':'Accion',             'width':5.5, 'align':'left'},
            ]
            st.download_button(
                label="📄 PDF Carta",
                data=build_pdf(
                    f"PEDIDO FARMACIA AA → BODEGA AA  |  S{semana} · {hoy.strftime('%d/%m/%Y')}",
                    f"Periodo: {datos['mtime'].strftime('%d/%m/%Y')}  ·  {len(df_tabla)} medicamentos",
                    df_tabla, _cols_pdf_bod, orientacion='landscape'
                ),
                file_name=f"Pedido_Bodega_AA_{hoy.strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        with col_dl3:
            st.caption(
                f"{len(df_tabla)} medicamentos · análisis: {datos['mtime'].strftime('%d/%m/%Y %H:%M')}"
            )

# ══════════════════════════════════════════════════════════════════════
# HELPER NUMÉRICO — usado por las pestañas de abajo
# ══════════════════════════════════════════════════════════════════════
def _num(df, col):
    if col in df.columns:
        return pd.to_numeric(df[col], errors='coerce').fillna(0)
    return pd.Series(0, index=df.index, dtype=float)

# ══════════════════════════════════════════════════════════════════════
# PESTAÑA 4 — PEDIDO A BODEGA FÁRMACOS
# Bodega AA solicita a Bodega Fármacos (única bodega de respaldo)
# ══════════════════════════════════════════════════════════════════════
with tab_pedido_farm:
    st.markdown("## 🏭 Pedido Bodega AA → Bodega Fármacos")
    st.markdown(
        "Medicamentos que **Bodega AA necesita reponer desde Bodega Fármacos** "
        "(única bodega de respaldo), ordenados por criticidad. "
        "Incluye la cantidad disponible en Bodega Fármacos y lo que requiere compra externa."
    )
    st.markdown("---")

    df_pf = df_bod.copy()
    df_pf['_nec']   = _num(df_pf, 'Reponer_Bodega')
    # Stock REAL en Bodega Fármacos (único respaldo válido). Fallback a la
    # columna antigua (suma de todas las bodegas) si el Excel no se regeneró.
    _col_bf = 'Stock_BODEGA_FARMACOS' if 'Stock_BODEGA_FARMACOS' in df_pf.columns else 'Stock_Hospital_Total'
    df_pf['_hosp']  = _num(df_pf, _col_bf)
    df_pf['_ord']   = (df_pf['Criticidad'].apply(orden_crit)
                       if 'Criticidad' in df_pf.columns
                       else pd.Series(5, index=df_pf.index))

    df_pf = df_pf[df_pf['_nec'] > 0].sort_values(
        ['_ord', '_nec'], ascending=[True, False]
    ).reset_index(drop=True)

    if df_pf.empty:
        st.success("✅ Bodega AA no necesita reponer desde Bodega Fármacos en este momento.")
    else:
        # Métricas resumen
        n_con_stk  = len(df_pf[df_pf['_hosp'] > 0])
        n_sin_stk  = len(df_pf[df_pf['_hosp'] == 0])
        total_pedir = int(df_pf['_nec'].sum())
        total_disp  = int(df_pf['_hosp'].sum())

        mp1, mp2, mp3, mp4 = st.columns(4)
        mp1.metric("Total medicamentos",           f"{len(df_pf)}")
        mp2.metric("Con stock en Bod. Fármacos",   f"{n_con_stk} ✅")
        mp3.metric("Sin stock en Bod. Fármacos",   f"{n_sin_stk} ⚠️")
        mp4.metric("Unidades totales a reponer",   f"{total_pedir:,}")
        st.markdown("---")

        # Tabla principal
        filas_pf = []
        for _, row in df_pf.iterrows():
            nec    = int(_get(row, 'Reponer_Bodega', 0))
            hosp   = int(_get(row, _col_bf, 0))
            a_pedir = min(nec, hosp)
            compra  = max(nec - hosp, 0)
            filas_pf.append({
                'Medicamento'            : str(_get(row, 'Medicamento', '')),
                'Criticidad'             : str(_get(row, 'Criticidad', '')),
                'Stock Bodega AA actual' : int(float(_get(row, 'Stock_Bod_Actual', 0))),
                'Cob. actual (dias)'     : round(float(_get(row, 'Cob_Bod_Actual_Dias', 0)), 1),
                'Req. 2 semanas (10d)'   : int(float(_get(row, 'Req_2_Semanas', 0))),
                'Consumo 10D proyectado' : int(float(_get(row, 'Consumo_10D_Trend', 0))),
                'Total a reponer'        : nec,
                'Disponible Bod. Farm.'  : hosp,
                'Solicitar a Bod. Farm.' : a_pedir,
                'Compra externa'         : compra,
                'Accion 1'               : str(_get(row, 'Accion_1_Traspaso_Hospital', '') or ''),
                'Accion 2'               : str(_get(row, 'Accion_2_Compra_Externa', '') or ''),
            })

        df_tabla_pf = pd.DataFrame(filas_pf)

        st.dataframe(
            estilo_tabla(df_tabla_pf),
            use_container_width=True, hide_index=True,
            column_config={
                'Medicamento'           : st.column_config.TextColumn("Medicamento",             width="large"),
                'Criticidad'            : st.column_config.TextColumn("Criticidad",              width="small"),
                'Stock Bodega AA actual': st.column_config.NumberColumn("Stock Bod. AA",         format="%d"),
                'Cob. actual (dias)'    : st.column_config.NumberColumn("Cob. actual",           format="%.1f d"),
                'Req. 2 semanas (10d)'  : st.column_config.NumberColumn("Req. 2 sem. (10d)",     format="%d"),
                'Consumo 10D proyectado': st.column_config.NumberColumn("Consumo 10D tend.",     format="%d"),
                'Total a reponer'       : st.column_config.NumberColumn("Total a reponer",       format="%d ud"),
                'Disponible Bod. Farm.' : st.column_config.NumberColumn("Disp. Bod. Farm.",      format="%d"),
                'Solicitar a Bod. Farm.': st.column_config.NumberColumn("Solicitar",             format="%d ud"),
                'Compra externa'        : st.column_config.NumberColumn("Compra externa",        format="%d"),
                'Accion 1'              : st.column_config.TextColumn("Accion 1",                width="medium"),
                'Accion 2'              : st.column_config.TextColumn("Accion 2",                width="medium"),
            },
            height=min(50 + len(df_tabla_pf) * 35, 620),
        )

        # Alertas: sin stock en Bodega Fármacos
        sin_farm = df_tabla_pf[df_tabla_pf['Disponible Bod. Farm.'] == 0]
        if len(sin_farm):
            with st.expander(f"⚠️ {len(sin_farm)} medicamento(s) sin stock en Bodega Fármacos — requieren compra externa", expanded=False):
                for _, r in sin_farm.iterrows():
                    emo = crit_emoji(str(r['Criticidad']))
                    st.markdown(f"{emo} **{r['Medicamento']}** — Necesidad: **{r['Total a reponer']} ud.** → Compra externa")

        # Descarga Excel
        st.markdown("---")
        def build_pedido_farm_excel():
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedido a Bodega Farmacos"

            ws['A1'] = f'PEDIDO BODEGA AA → BODEGA FARMACOS  (CICLO 2 SEMANAS)  |   S{semana}  {hoy.strftime("%d/%m/%Y")}'
            ws['A1'].fill = PatternFill('solid', fgColor='1B5E20')
            ws['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
            ws.merge_cells('A1:J1')
            ws.row_dimensions[1].height = 28

            ws['A2'] = f'Ciclo pedido: cada 2 semanas (10 días háb.)  |  Total: {len(df_tabla_pf)} meds  |  Con stock Bod.Farm.: {n_con_stk}  |  Requieren compra: {n_sin_stk}'
            ws['A2'].font = Font(italic=True, name='Arial', size=10, color='444444')
            ws.merge_cells('A2:J2')

            hdrs = ['N°','Medicamento','Criticidad','Stock Bod.AA','Cob.(días)',
                    'Req. 2 Semanas','Consumo 10D tend.','Disp.Bod.Farm.','Solicitar','Compra Externa']
            hfill = PatternFill('solid', fgColor='2E7D32')
            hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[3].height = 30

            fills_c = {k: crit_fill(k) for k in CRIT_FILL_HEX}
            for ri, (_, row) in enumerate(df_tabla_pf.iterrows(), 4):
                crit = str(row.get('Criticidad', ''))
                fill = fills_c.get(crit, PatternFill('solid', fgColor='F1F8F1'))
                is_crit1 = crit == '1-CRITICO'
                vals = [ri-3,
                        row.get('Medicamento', ''),
                        crit,
                        int(row.get('Stock Bodega AA actual', 0) or 0),
                        round(float(row.get('Cob. actual (dias)', 0) or 0), 1),
                        int(row.get('Req. 2 semanas (10d)', 0) or 0),
                        int(row.get('Consumo 10D proyectado', 0) or 0),
                        int(row.get('Disponible Bod. Farm.', 0) or 0),
                        int(row.get('Solicitar a Bod. Farm.', 0) or 0),
                        int(row.get('Compra externa', 0) or 0)]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.fill = fill
                    c.font = Font(name='Arial', size=10,
                                  bold=is_crit1,
                                  color='FFFFFF' if is_crit1 else '000000')
                    c.alignment = Alignment(vertical='center', wrap_text=(ci==2))

            for ci, w in enumerate([4,48,13,10,12,13,13,14,12,13], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A4'

            # Firma
            ult = len(df_tabla_pf) + 6
            ws.cell(row=ult, column=1, value='Solicitado por (Bodega AA):').font = Font(name='Arial',size=10)
            ws.cell(row=ult, column=5, value='Recibido por (Bod. Fármacos):').font = Font(name='Arial',size=10)
            ws.cell(row=ult, column=10, value=f'Fecha: {hoy.strftime("%d/%m/%Y")}').font = Font(name='Arial',size=10)

            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return buf

        cpf1, cpf2, cpf3 = st.columns([2, 2, 2])
        with cpf1:
            st.download_button(
                label="📥 Excel",
                data=build_pedido_farm_excel(),
                file_name=f"Pedido_BodFarmacos_{hoy.strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
        with cpf2:
            _cols_pdf_farm = [
                {'name':'Medicamento',            'label':'Medicamento',         'width':6.8, 'align':'left'},
                {'name':'Criticidad',             'label':'Criticidad',          'width':2.8, 'align':'center'},
                {'name':'Stock Bodega AA actual', 'label':'Stock Bod.AA',        'width':2.2, 'align':'center'},
                {'name':'Consumo 10D proyectado', 'label':'Consumo 10D',         'width':2.2, 'align':'center'},
                {'name':'Total a reponer',        'label':'A Reponer',           'width':2.2, 'align':'center'},
                {'name':'Disponible Bod. Farm.',  'label':'Disp.Bod.Farm.',      'width':2.4, 'align':'center'},
                {'name':'Solicitar a Bod. Farm.', 'label':'Solicitar',           'width':2.2, 'align':'center'},
                {'name':'Compra externa',         'label':'Compra Externa',      'width':2.2, 'align':'center'},
            ]
            st.download_button(
                label="📄 PDF Carta",
                data=build_pdf(
                    f"PEDIDO BODEGA AA → BODEGA FÁRMACOS  |  S{semana} · {hoy.strftime('%d/%m/%Y')}",
                    f"Ciclo 2 semanas (10 días hábiles)  ·  {len(df_tabla_pf)} medicamentos",
                    df_tabla_pf, _cols_pdf_farm, orientacion='landscape'
                ),
                file_name=f"Pedido_BodFarmacos_{hoy.strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        with cpf3:
            st.caption(f"Datos al: {datos['mtime'].strftime('%d/%m/%Y %H:%M')}.")

# ══════════════════════════════════════════════════════════════════════
# PESTAÑA 5 — DIÁLISIS  (solo recetas de los nefrólogos)
# Mismo motor de cálculo y misma escala de criticidad que las demás pestañas,
# pero el consumo se restringe a las recetas de:
#   Dr. Yasmani Ortiz Amador · Dra. Martha Peralta · Dra. Mónica Amaya
# ══════════════════════════════════════════════════════════════════════
with tab_dialisis:
    st.markdown("## 💉 Pedido de Diálisis (mensual)")
    st.markdown(
        "Medicamentos para pacientes en **diálisis / nefrología**, identificados por las "
        "recetas de Dr. *Yasmani Ortiz Amador*, Dra. *Martha Peralta* y Dra. *Mónica Amaya*. "
        "Este pedido es **independiente del pedido normal** y se realiza **una vez al mes, "
        "en la 3ª semana**. La cantidad sugerida es el **consumo mensual de diálisis**."
    )

    # Recordatorio de cadencia: el pedido de diálisis va en la 3ª semana del mes
    if semana == 3:
        st.success("📅 **Es la 3ª semana del mes — momento de generar el pedido de diálisis.**")
    elif semana < 3:
        st.info(f"📅 El pedido de diálisis se hace en la **3ª semana** del mes. Vas en la semana **S{semana}**.")
    else:
        st.warning(f"📅 El pedido de diálisis se hace en la **3ª semana** (vas en S{semana}). "
                   "Si aún no lo generaste este mes, hazlo ahora.")

    if df_dial_farm is None or df_dial_farm.empty:
        st.info("No hay datos de diálisis en el maestro actual "
                "(ejecuta AUTO_SSASUR.bat / EJECUTAR_MAESTRO.bat para regenerarlo).")
    else:
        def _cn(df, c):
            return pd.to_numeric(df[c], errors='coerce').fillna(0) if c in df.columns \
                   else pd.Series(0, index=df.index, dtype=float)
        dfa = df_dial_farm.copy()
        dfa['_dial5d'] = _cn(dfa, 'Consumo_5D_Solo_Dialisis')
        dfa['_stock']  = _cn(dfa, 'Stock_Farm_Actual')
        dfa['_factor'] = _cn(dfa, 'Factor_Empaque').replace(0, 1)
        dfa['_ord']    = dfa['Criticidad'].apply(orden_crit) if 'Criticidad' in dfa.columns \
                         else pd.Series(5, index=dfa.index)
        dfa = dfa[dfa['_dial5d'] > 0].copy()
        # Cantidad sugerida = consumo mensual de diálisis (5 días hábiles → 30 días)
        dfa['_mensual'] = (dfa['_dial5d'] / 5 * 30).round().astype(int)
        dfa = dfa.sort_values(['_ord', '_mensual'], ascending=[True, False]).reset_index(drop=True)

        if dfa.empty:
            st.success("✅ No hay consumo de diálisis registrado en el periodo.")
        else:
            filas = []
            for _, row in dfa.iterrows():
                base     = int(row['_mensual'])
                factor   = int(row['_factor']) if int(row['_factor']) > 0 else 1
                # Aproximar al factor de empaque (hacia arriba)
                cantidad = ((base + factor - 1) // factor) * factor if factor > 1 else base
                stock    = int(row['_stock'])
                if stock >= cantidad:
                    obs = f"Se puede sacar directo de Farmacia AA (stock actual: {stock} ud)."
                else:
                    falta = cantidad - stock
                    obs = ("Sacar directo de Bodega AA por la cantidad. "
                           f"Hacer pedido Farmacia AA → Bodega AA (faltan {falta} ud en Farmacia).")
                filas.append({
                    'Medicamento'       : str(_get(row, 'Medicamento', '')),
                    'Cantidad sugerida' : cantidad,
                    'Observaciones'     : obs,
                })
            df_simple = pd.DataFrame(filas)

            m1, m2 = st.columns(2)
            m1.metric("Medicamentos de diálisis", f"{len(df_simple)}")
            m2.metric("Unidades sugeridas (mes)", f"{int(df_simple['Cantidad sugerida'].sum()):,}")
            st.markdown("---")

            st.dataframe(
                df_simple, use_container_width=True, hide_index=True,
                column_config={
                    'Medicamento'       : st.column_config.TextColumn("Medicamento", width="large"),
                    'Cantidad sugerida' : st.column_config.NumberColumn("Cantidad sugerida (mes)", format="%d ud"),
                    'Observaciones'     : st.column_config.TextColumn("Observaciones", width="large"),
                },
                height=min(60 + len(df_simple) * 38, 640),
            )

            def build_dial_simple_excel():
                wb = Workbook(); ws = wb.active; ws.title = "Pedido Dialisis"
                ws['A1'] = f'PEDIDO DE DIÁLISIS (MENSUAL · 3ª SEMANA)   |   {hoy.strftime("%d/%m/%Y")}'
                ws['A1'].fill = PatternFill('solid', fgColor='0F766E')
                ws['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
                ws.merge_cells('A1:C1'); ws.row_dimensions[1].height = 28
                ws['A2'] = 'Consumo segun recetas de: Dr. Yasmani Ortiz Amador - Dra. Martha Peralta - Dra. Monica Amaya'
                ws['A2'].font = Font(italic=True, name='Arial', size=10, color='444444')
                ws.merge_cells('A2:C2')
                hdrs = ['Medicamento', 'Cantidad sugerida', 'Observaciones']
                hfill = PatternFill('solid', fgColor='0F766E')
                hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
                for ci, h in enumerate(hdrs, 1):
                    c = ws.cell(row=3, column=ci, value=h)
                    c.fill = hfill; c.font = hfont
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[3].height = 26
                for ri, (_, r) in enumerate(df_simple.iterrows(), 4):
                    ws.cell(row=ri, column=1, value=r['Medicamento']).alignment = Alignment(vertical='center', wrap_text=True)
                    ws.cell(row=ri, column=2, value=int(r['Cantidad sugerida'])).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=ri, column=3, value=r['Observaciones']).alignment = Alignment(vertical='center', wrap_text=True)
                    for ci in (1, 2, 3):
                        ws.cell(row=ri, column=ci).font = Font(name='Arial', size=10)
                for ci, w in enumerate([48, 18, 62], 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.freeze_panes = 'A4'
                buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

            st.markdown("---")
            cda1, cda2 = st.columns([2, 4])
            with cda1:
                st.download_button(
                    "📥 Excel pedido diálisis", data=build_dial_simple_excel(),
                    file_name=f"Pedido_Dialisis_{hoy.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary", use_container_width=True,
                )
            with cda2:
                st.caption(f"Datos al: {datos['mtime'].strftime('%d/%m/%Y %H:%M')}.")

# ══════════════════════════════════════════════════════════════════════
# PESTAÑA 6 — FALTANTES
# Solo usa Bodega Fármacos como respaldo (única bodega de respaldo real)
# ══════════════════════════════════════════════════════════════════════
with tab_faltantes:
    st.markdown("## 🚨 Faltantes Farmacia AT Abierta")
    st.markdown(
        "Medicamentos prescritos con procedencia **AT Abierta**, sin stock en Farmacia AA ni en Bodega AA. "
        "Se separan según si **Bodega Fármacos** (única bodega de respaldo) tiene o no existencias."
    )
    st.markdown("---")

    # Usar Faltantes_Detalle_AA que tiene Stock_BODEGA_FARMACOS
    df_fd = df_falt_det.copy()
    df_fd['_stk_aa']   = _num(df_fd, 'Stock_AA_Total')
    df_fd['_stk_bf']   = _num(df_fd, 'Stock_BODEGA_FARMACOS')   # único respaldo real
    df_fd['_faltante'] = _num(df_fd, 'Faltante_Neto')
    df_fd['_pac']      = _num(df_fd, 'Pacientes_Afectados')
    df_fd['_rec']      = _num(df_fd, 'N_Recetas')
    df_fd['_ord']      = (df_fd['Criticidad'].apply(orden_crit)
                          if 'Criticidad' in df_fd.columns
                          else pd.Series(5, index=df_fd.index))

    # Solo los que tienen stock 0 en AA
    df_fd = df_fd[df_fd['_stk_aa'] <= 0].copy()

    # Separar: sin respaldo (Bod. Fármacos = 0) vs con respaldo (Bod. Fármacos > 0)
    df_sin = df_fd[df_fd['_stk_bf'] <= 0].sort_values(['_ord','_faltante'], ascending=[True,False])
    df_con = df_fd[df_fd['_stk_bf'] > 0].sort_values(['_ord','_faltante'], ascending=[True,False])

    n_sin  = len(df_sin)
    n_con  = len(df_con)
    pac_tot = int(df_fd['_pac'].sum())
    ud_tot  = int(df_fd['_faltante'].sum())

    mf1, mf2, mf3, mf4 = st.columns(4)
    mf1.metric("Sin respaldo — Compra urgente",      f"{n_sin} 🔴")
    mf2.metric("Con respaldo en Bod. Fármacos",      f"{n_con} 🟡")
    mf3.metric("Pacientes afectados",                f"{pac_tot}")
    mf4.metric("Unidades totales faltantes",         f"{ud_tot:,}")

    # ── BLOQUE 1: SIN RESPALDO (Bodega Fármacos = 0) ──────────────────────────
    st.markdown("---")
    st.markdown(f"### 🔴 Sin respaldo — {n_sin} medicamento(s) — Compra Urgente")
    st.caption("Stock AA = 0 y Bodega Fármacos = 0. Requieren compra externa inmediata.")

    if df_sin.empty:
        st.success("✅ No hay faltantes sin respaldo ahora.")
    else:
        for _, row in df_sin.iterrows():
            med  = str(row.get('Medicamento',''))
            crit = str(row.get('Criticidad',''))
            falt = int(row['_faltante'])
            pac  = int(row['_pac'])
            rec  = int(row['_rec'])
            acc  = str(row.get('Accion_Sugerida','') or '')
            emo  = crit_emoji(crit)
            tc   = tarjeta_class(crit)
            st.markdown(f"""
            <div class='tarjeta {tc}'>
              <div style='display:flex;align-items:center;gap:20px;flex-wrap:wrap'>
                <div style='flex:1;min-width:220px'>
                  <div style='font-size:1.05rem;font-weight:700'>{emo} {med}</div>
                  <div style='font-size:0.82rem;color:#666;margin-top:2px'>Criticidad: <b>{crit}</b></div>
                </div>
                <div style='text-align:center;min-width:85px'>
                  <div class='num-grande' style='color:#E53935'>{falt:,}</div>
                  <div class='etiqueta'>unidades faltantes</div>
                </div>
                <div style='text-align:center;min-width:85px'>
                  <div style='font-size:1.3rem;font-weight:700'>{pac}</div>
                  <div class='etiqueta'>pacientes afectados</div>
                </div>
                <div style='text-align:center;min-width:75px'>
                  <div style='font-size:1.1rem;font-weight:600;color:#555'>{rec}</div>
                  <div class='etiqueta'>recetas pendientes</div>
                </div>
              </div>
              {'<div class=accion-ext>🚨 ' + acc + '</div>' if acc else '<div class=accion-ext>🚨 COMPRA URGENTE</div>'}
            </div>
            """, unsafe_allow_html=True)

    # ── BLOQUE 2: CON RESPALDO EN BODEGA FÁRMACOS ─────────────────────────────
    st.markdown("---")
    st.markdown(f"### 🟡 Con respaldo en Bodega Fármacos — {n_con} medicamento(s)")
    st.caption("Stock AA = 0 pero Bodega Fármacos tiene existencias. Gestionar traspaso urgente.")

    if df_con.empty:
        st.info("No hay faltantes con respaldo en Bodega Fármacos.")
    else:
        filas_con = []
        for _, row in df_con.iterrows():
            filas_con.append({
                'Medicamento'          : str(row.get('Medicamento','')),
                'Criticidad'           : str(row.get('Criticidad','')),
                'Unidades faltantes'   : int(row['_faltante']),
                'Stock Bod. Fármacos'  : int(row['_stk_bf']),
                'Pacientes afectados'  : int(row['_pac']),
                'Recetas pendientes'   : int(row['_rec']),
                'Accion'               : str(row.get('Accion_Sugerida','') or 'TRASPASAR DESDE BODEGA FARMACOS'),
            })
        st.dataframe(
            estilo_tabla(pd.DataFrame(filas_con)),
            use_container_width=True, hide_index=True,
            column_config={
                'Medicamento'         : st.column_config.TextColumn("Medicamento",          width="large"),
                'Criticidad'          : st.column_config.TextColumn("Criticidad",           width="small"),
                'Unidades faltantes'  : st.column_config.NumberColumn("Faltantes",          format="%d ud"),
                'Stock Bod. Fármacos' : st.column_config.NumberColumn("Stock Bod. Farm.",   format="%d"),
                'Pacientes afectados' : st.column_config.NumberColumn("Pacientes",          format="%d"),
                'Recetas pendientes'  : st.column_config.NumberColumn("Recetas",            format="%d"),
                'Accion'              : st.column_config.TextColumn("Accion",               width="medium"),
            },
        )

    # ── Descarga Excel ─────────────────────────────────────────────────────────
    st.markdown("---")

    def build_faltantes_excel():
        wb   = Workbook()
        fh   = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        fc   = {k: crit_fill(k) for k in CRIT_FILL_HEX}

        def _hoja(ws, titulo, color_tit, color_h, df_rows, hdrs, extractor):
            ws['A1'] = titulo
            ws['A1'].fill = PatternFill('solid', fgColor=color_tit)
            ws['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
            ws.merge_cells(f'A1:{get_column_letter(len(hdrs))}1')
            ws.row_dimensions[1].height = 26
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=2, column=ci, value=h)
                c.fill = PatternFill('solid', fgColor=color_h)
                c.font = fh
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 22
            for ri, (_, row) in enumerate(df_rows.iterrows(), 3):
                crit = str(row.get('Criticidad', ''))
                fill = fc.get(crit, PatternFill('solid', fgColor='FFF3F3'))
                vals = extractor(row)
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.fill = fill
                    c.font = Font(name='Arial', size=11)

        # Hoja 1 — Sin respaldo
        ws1 = wb.active
        ws1.title = "SIN RESPALDO - URGENTE"
        _hoja(ws1,
              f'FALTANTES SIN RESPALDO — {hoy.strftime("%d/%m/%Y %H:%M")}',
              'E53935', 'C62828', df_sin,
              ['Medicamento','Criticidad','Unidades Faltantes','Pacientes','Recetas','Accion'],
              lambda r: [r.get('Medicamento',''),
                         str(r.get('Criticidad','')),
                         int(r.get('_faltante',0) or 0),
                         int(r.get('_pac',0) or 0),
                         int(r.get('_rec',0) or 0),
                         str(r.get('Accion_Sugerida','') or 'COMPRA URGENTE')])
        for ci,w in enumerate([50,15,18,12,10,25],1):
            ws1.column_dimensions[get_column_letter(ci)].width = w
        ws1.freeze_panes = 'A3'

        # Hoja 2 — Con respaldo Bod. Fármacos
        ws2 = wb.create_sheet("CON RESPALDO BOD.FARMACOS")
        _hoja(ws2,
              f'FALTANTES CON RESPALDO BODEGA FARMACOS — {hoy.strftime("%d/%m/%Y %H:%M")}',
              'E65100', 'BF360C', df_con,
              ['Medicamento','Criticidad','Unidades Faltantes','Stock Bod. Farmacos','Pacientes','Accion'],
              lambda r: [r.get('Medicamento',''),
                         str(r.get('Criticidad','')),
                         int(r.get('_faltante',0) or 0),
                         int(r.get('_stk_bf',0) or 0),
                         int(r.get('_pac',0) or 0),
                         str(r.get('Accion_Sugerida','') or 'TRASPASAR DESDE BODEGA FARMACOS')])
        for ci,w in enumerate([50,15,18,16,12,30],1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.freeze_panes = 'A3'

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    cff1, cff2, cff3 = st.columns([2, 2, 2])
    with cff1:
        st.download_button(
            label="📥 Excel",
            data=build_faltantes_excel(),
            file_name=f"Faltantes_AA_{hoy.strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True,
        )
    with cff2:
        # PDF faltantes — une sin respaldo y con respaldo
        _df_falt_pdf = pd.concat([df_sin, df_con], ignore_index=True)
        _cols_pdf_falt = [
            {'name':'Medicamento',        'label':'Medicamento',       'width':8.0, 'align':'left'},
            {'name':'Criticidad',         'label':'Criticidad',        'width':3.2, 'align':'center'},
            {'name':'_faltante',          'label':'Unid. Faltantes',   'width':2.4, 'align':'center'},
            {'name':'Stock_BODEGA_FARMACOS','label':'Stock Bod.Farm.', 'width':2.4, 'align':'center'},
            {'name':'_pac',               'label':'Pacientes',         'width':2.0, 'align':'center'},
            {'name':'Accion_Sugerida',    'label':'Accion Sugerida',   'width':5.0, 'align':'left'},
        ]
        st.download_button(
            label="📄 PDF Carta",
            data=build_pdf(
                f"FALTANTES FARMACIA AT ABIERTA  |  {hoy.strftime('%d/%m/%Y')}",
                f"Sin respaldo: {n_sin}  ·  Con respaldo Bod. Fármacos: {n_con}  ·  Total: {len(_df_falt_pdf)}",
                _df_falt_pdf, _cols_pdf_falt, orientacion='landscape'
            ),
            file_name=f"Faltantes_AA_{hoy.strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    with cff3:
        st.caption(
            f"Sin Respaldo: {n_sin} · Con Respaldo Bod. Farm.: {n_con} · "
            f"Datos al {datos['mtime'].strftime('%d/%m/%Y %H:%M')}."
        )
with tab_sgli:
    st.markdown("## 🚦 SGLI · Reposición por estrés / semana pico")
    st.markdown(
        "Motor de **Gestión Logística (SGLI)**: calcula el nivel objetivo de "
        "Farmacia AA bajo un escenario de **estrés de demanda** y decide el "
        "traspaso desde Bodega AA o la compra urgente. Mueve el **factor de "
        "carga** para simular *qué pasa si* la demanda se intensifica."
    )

    if df_sgli_base is None or df_sgli_base.empty:
        st.info(
            "La hoja **SGLI_Estres** aún no está en el Consolidado. Ejecuta "
            "`ACTUALIZAR_DATOS.bat` (o `py maestro_aa.py`) para generarla y vuelve a recargar."
        )
    else:
        with st.expander("ℹ️ Cómo se calcula"):
            st.markdown(
                "- **IR** (días de reposición) = `mín(5, ⌊5 / Factor_Carga⌋)` — nunca supera 5.\n"
                "- **Demanda** = `IR × CDL × Factor_Carga`  ·  **Nivel Objetivo (T)** = `Demanda × 1.25`.\n"
                "- **Déficit** = `máx(0, T − Stock Farmacia)`.\n"
                "- **Decisión:** si hay déficit, primero se **traspasa desde Bodega AA** lo "
                "disponible y el resto va a **compra urgente**.\n"
                "- **Alerta de estrés**: se activa cuando la *semana actual* coincide con la "
                "*semana pico* **y** el Factor_Carga supera 1.15 (configuración global: aplica "
                "a todos los medicamentos)."
            )

        c1, c2, c3 = st.columns([2, 1.4, 1.4])
        with c1:
            factor = st.slider(
                "Factor de carga (estrés)", min_value=1.00, max_value=2.00,
                value=float(FACTOR_CARGA_DEFAULT), step=0.05,
                help="1.15 = sin estrés extra (baseline). Súbelo para simular mayor demanda.",
            )
        with c2:
            _opts_sp = ['S1', 'S2', 'S3', 'S4']
            semana_pico_sel = st.selectbox(
                "Semana pico de la campaña", _opts_sp,
                index=min(semana, 4) - 1,
                help="Semana del mes con mayor demanda esperada.",
            )
        with c3:
            solo_acc = st.toggle("Solo con déficit", value=True,
                                 help="Oculta los medicamentos sin déficit (Déficit = 0).")

        df_calc = calcular_sgli(
            df_sgli_base, factor_carga=factor, semana_pico=semana_pico_sel,
            semana_actual=f"S{semana}",
            col_crit='Criticidad', col_farm='Stock_Farm', col_bod='Stock_Bod',
            col_cdl='CDL', col_sp_hist='Semana_Pico_Hist', col_fc_hist='Factor_Carga_Hist',
        )

        _def = pd.to_numeric(df_calc['Deficit'],   errors='coerce').fillna(0)
        _bod = pd.to_numeric(df_calc['Stock_Bod'], errors='coerce').fillna(0)
        _mask = _def > 0
        ir_val       = int(df_calc['Dias_Reposicion_IR'].iloc[0]) if len(df_calc) else 0
        n_def        = int(_mask.sum())
        traspaso_tot = int(np.minimum(_def, _bod)[_mask].sum())
        compra_tot   = int(np.maximum(_def - _bod, 0)[_mask].sum())
        alerta_on    = bool((df_calc['Alerta_Estres'] != '').any())

        if alerta_on:
            st.error(
                f"🔴 **[ALERTA_ESTRES]** activa — semana actual **S{semana}** = semana pico "
                f"**{semana_pico_sel}** y Factor_Carga **{factor:.2f}** > {UMBRAL_ESTRES:.2f}."
            )
        else:
            st.success(
                f"🟢 Sin alerta de estrés (S{semana} vs pico {semana_pico_sel} · "
                f"Factor {factor:.2f}). Súbelo sobre {UMBRAL_ESTRES:.2f} en la semana pico para activarla."
            )

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Días reposición (IR)", ir_val)
        m2.metric("Con déficit", n_def)
        m3.metric("A traspasar (ud)", f"{traspaso_tot:,}".replace(",", "."))
        m4.metric("Compra urgente (ud)", f"{compra_tot:,}".replace(",", "."))

        df_show = df_calc[_mask].reset_index(drop=True) if solo_acc else df_calc

        if len(df_show) == 0:
            st.info("Ningún medicamento presenta déficit con el factor de carga actual. ✅")
        else:
            st.dataframe(
                estilo_tabla(df_show),
                use_container_width=True, hide_index=True,
                column_config={
                    'Medicamento'        : st.column_config.TextColumn("Medicamento", width="large"),
                    'Criticidad'         : st.column_config.TextColumn("Criticidad", width="small"),
                    'Dias_Reposicion_IR' : st.column_config.NumberColumn("Días Rep. (IR)", format="%d"),
                    'Alerta_Estres'      : st.column_config.TextColumn("Alerta Estrés", width="small"),
                    'Nivel_Objetivo_T'   : st.column_config.NumberColumn("Nivel Obj. (T)", format="%d"),
                    'Deficit'            : st.column_config.NumberColumn("Déficit", format="%d"),
                    'Accion_1_Traspaso'  : st.column_config.TextColumn("Acción 1: Traspaso", width="medium"),
                    'Accion_2_Externa'   : st.column_config.TextColumn("Acción 2: Externa", width="medium"),
                    'Stock_Farm'         : st.column_config.NumberColumn("Stock Farm.", format="%d"),
                    'Stock_Bod'          : st.column_config.NumberColumn("Stock Bod.", format="%d"),
                    'CDL'                : st.column_config.NumberColumn("CDL", format="%.2f"),
                    'Semana_Pico_Hist'   : st.column_config.TextColumn("Sem. pico hist.", width="small"),
                    'Factor_Carga_Hist'  : st.column_config.NumberColumn("Factor hist.", format="%.2f"),
                },
                height=min(50 + len(df_show) * 35, 640),
            )

            def build_sgli_excel():
                wb = Workbook(); ws = wb.active; ws.title = "SGLI_Estres"
                hfill = PatternFill('solid', fgColor='7C2D12')
                hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
                cols = list(df_show.columns)
                for ci, h in enumerate(cols, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.fill = hfill; c.font = hfont
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[1].height = 28
                for ri, (_, row) in enumerate(df_show.iterrows(), 2):
                    crit = str(row.get('Criticidad', '5-OK'))
                    fill = PatternFill('solid', fgColor=crit_hex(crit))
                    is_c1 = crit == '1-CRITICO'
                    for ci, col in enumerate(cols, 1):
                        v = row.get(col, '')
                        v = '' if pd.isna(v) else v
                        c = ws.cell(row=ri, column=ci, value=v)
                        c.fill = fill
                        c.font = Font(name='Arial', size=10, bold=is_c1,
                                      color='FFFFFF' if is_c1 else '000000')
                        c.alignment = Alignment(vertical='center',
                                                wrap_text=(col in ('Medicamento', 'Accion_1_Traspaso',
                                                                   'Accion_2_Externa')))
                for ci, col in enumerate(cols, 1):
                    ltr = get_column_letter(ci)
                    if   col == 'Medicamento':       ws.column_dimensions[ltr].width = 50
                    elif col == 'Accion_1_Traspaso': ws.column_dimensions[ltr].width = 30
                    elif col == 'Accion_2_Externa':  ws.column_dimensions[ltr].width = 26
                    else:                            ws.column_dimensions[ltr].width = 15
                ws.freeze_panes = 'B2'
                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                return buf

            d1, d2, d3 = st.columns([2, 2, 2])
            with d1:
                st.download_button(
                    "📥 Excel", data=build_sgli_excel(),
                    file_name=f"SGLI_Estres_{hoy.strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary", use_container_width=True,
                )
            with d2:
                _cols_pdf_sgli = [
                    {'name': 'Medicamento',        'label': 'Medicamento',     'width': 5.5, 'align': 'left'},
                    {'name': 'Criticidad',         'label': 'Criticidad',      'width': 2.2, 'align': 'center'},
                    {'name': 'Dias_Reposicion_IR', 'label': 'IR (d)',          'width': 1.6, 'align': 'center'},
                    {'name': 'Alerta_Estres',      'label': 'Alerta',          'width': 2.4, 'align': 'center'},
                    {'name': 'Nivel_Objetivo_T',   'label': 'Nivel Obj. (T)',  'width': 1.9, 'align': 'center'},
                    {'name': 'Deficit',            'label': 'Deficit',         'width': 1.6, 'align': 'center'},
                    {'name': 'Accion_1_Traspaso',  'label': 'Accion 1: Traspaso', 'width': 3.3, 'align': 'left'},
                    {'name': 'Accion_2_Externa',   'label': 'Accion 2: Externa',  'width': 3.3, 'align': 'left'},
                ]
                st.download_button(
                    "📄 PDF Carta",
                    data=build_pdf(
                        f"SGLI · REPOSICIÓN POR ESTRÉS  |  S{semana} · {hoy.strftime('%d/%m/%Y')}",
                        f"Factor de carga: {factor:.2f}  ·  Semana pico: {semana_pico_sel}  ·  "
                        f"IR: {ir_val} d  ·  {len(df_show)} medicamentos",
                        df_show, _cols_pdf_sgli, orientacion='landscape',
                    ),
                    file_name=f"SGLI_Estres_{hoy.strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf", use_container_width=True,
                )
            with d3:
                st.caption(
                    f"{len(df_show)} medicamentos · Factor {factor:.2f} · "
                    f"datos al {datos['mtime'].strftime('%d/%m/%Y %H:%M')}"
                )

            with st.expander("📋 Ver como tabla Markdown (Formato de Salida Obligatorio)"):
                st.code(to_markdown(df_show), language="markdown")

with tab_auditoria:
    import json as _json_aud
    st.markdown("## 🔬 Auditoría de prescripción")
    st.markdown(
        "Análisis por medicamento: **consumo mensual** (prescrito vs dispensado), "
        "**mayores prescriptores**, **diagnósticos asociados** y **duplicidad de prescripción**. "
        "Datos pre-calculados de las recetas; se actualizan al regenerar el maestro."
    )

    @st.cache_data(ttl=300)
    def _cargar_auditoria():
        ruta = os.path.join(WORK_DIR, "auditoria_prescripcion.json")
        if not os.path.exists(ruta):
            return None
        with open(ruta, encoding="utf-8") as _f:
            return _json_aud.load(_f)

    _aud = _cargar_auditoria()
    if not _aud or not _aud.get("data"):
        st.info("No hay auditoría disponible todavía. Ejecuta **AUTO_SSASUR.bat** "
                "(o `py auditoria_prescripcion.py`) para generarla.")
    else:
        data = _aud["data"]
        meds = sorted(data.keys())
        _def = next((i for i, m in enumerate(meds) if "EMPAGLIFLOZINA 10" in m), 0)
        med = st.selectbox(f"Medicamento ({len(meds):,} disponibles)", meds, index=_def)
        a = data[med]

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("CMP dispensado", f"{a['cmp_dispensado']:,} ud/mes")
        c2.metric("Prescrito (período)", f"{a['total_prescrito']:,} ud")
        c3.metric("Dispensado", f"{a['total_dispensado']:,} ud", f"{a['pct_dispensado']}% de lo prescrito")
        c4.metric("Pacientes", f"{a['pacientes']:,}")

        if a['pct_dispensado'] and a['pct_dispensado'] < 70:
            st.warning(f"⚠️ Solo se dispensó el **{a['pct_dispensado']}%** de lo prescrito — "
                       "brecha alta prescripción/entrega (demanda no satisfecha o sobreprescripción).")

        if a.get("meses"):
            st.markdown("### 📈 Consumo mensual dispensado (unidades)")
            st.bar_chart(pd.Series(a["meses"]))

        col_p, col_d = st.columns(2)
        with col_p:
            st.markdown("### 👨‍⚕️ Mayores prescriptores")
            if a.get("prescriptores"):
                dfp = pd.DataFrame(a["prescriptores"]).rename(columns={
                    "medico": "Médico", "esp": "Especialidad", "recetas": "Recetas",
                    "pacientes": "Pacientes", "unidades": "Unidades", "pct": "% ud"})
                st.dataframe(dfp, use_container_width=True, hide_index=True,
                             column_config={"% ud": st.column_config.NumberColumn("% ud", format="%.1f %%")})
        with col_d:
            st.markdown("### 🩺 Diagnósticos asociados")
            if a.get("diagnosticos"):
                dfd = pd.DataFrame(a["diagnosticos"]).rename(columns={
                    "dx": "Diagnóstico", "recetas": "Recetas", "pacientes": "Pacientes"})
                st.dataframe(dfd, use_container_width=True, hide_index=True)
            if a.get("sin_diagnostico"):
                st.caption(f"Recetas sin Diagnóstico 1: {a['sin_diagnostico']:,}")

        st.markdown("### 👥 Duplicidad de prescripción")
        d1, d2, d3 = st.columns(3)
        d1.metric("Pacientes distintos", f"{a['pacientes']:,}")
        d2.metric("Con ≥2 médicos distintos", f"{a['dup_2mas_medicos']:,}")
        d3.metric("Con ≥2 recetas el mismo mes", f"{a['dup_2mas_recetas_mes']:,}")
        if a['pacientes']:
            _pdup = round(100 * a['dup_2mas_medicos'] / a['pacientes'])
            if _pdup >= 15:
                st.warning(f"⚠️ El **{_pdup}%** de los pacientes recibieron este medicamento de ≥2 médicos "
                           "distintos — revisar posible duplicidad de prescripción.")
        st.caption(f"Datos al: {datos['mtime'].strftime('%d/%m/%Y %H:%M')}.")

with tab_feedback:
    import json as _json

    st.markdown("## 💬 Diagnóstico y Sugerencias")
    st.markdown(
        "Esta pestaña muestra el **estado actual de los datos** y te permite "
        "registrar sugerencias o reportar problemas para mejorar la aplicación."
    )

    # ── Diagnóstico automático ────────────────────────────────────────────────
    st.markdown("### 🔍 Diagnóstico automático")

    horas_ant_fb = (datetime.now() - datos['mtime']).total_seconds() / 3600
    dias_ant_fb  = horas_ant_fb / 24

    col_d1, col_d2, col_d3, col_d4 = st.columns(4)
    col_d1.metric("Medicamentos universo",  f"{len(todos_meds)}")
    col_d2.metric("Antigüedad del dato",
                  f"{int(horas_ant_fb)}h" if horas_ant_fb < 48 else f"{dias_ant_fb:.1f} días",
                  delta="OK" if horas_ant_fb < 24 else "DESACTUALIZADO",
                  delta_color="normal" if horas_ant_fb < 24 else "inverse")

    n_crit_farm = 0
    n_sin_stock = 0
    if 'Criticidad' in df_farm.columns:
        n_crit_farm = int((df_farm['Criticidad'].isin(['1-CRITICO','2-URGENTE'])).sum())
    if 'Stock_Farm_Actual' in df_farm.columns:
        n_sin_stock = int((pd.to_numeric(df_farm['Stock_Farm_Actual'], errors='coerce').fillna(0) == 0).sum())

    col_d3.metric("Críticos + Urgentes (farm.)", f"{n_crit_farm}",
                  delta="revisar" if n_crit_farm > 5 else "OK",
                  delta_color="inverse" if n_crit_farm > 5 else "normal")
    col_d4.metric("Sin stock en farmacia", f"{n_sin_stock}",
                  delta="alerta" if n_sin_stock > 0 else "OK",
                  delta_color="inverse" if n_sin_stock > 0 else "normal")

    st.markdown("---")

    # Chequeos automáticos
    issues   = []
    ok_items = []

    if horas_ant_fb > 24:
        issues.append(f"🔴 **Datos desactualizados** — el archivo tiene {dias_ant_fb:.1f} días. "
                      "Ejecuta `AUTO_SSASUR.bat` para obtener stock fresco.")
    else:
        ok_items.append(f"✅ Datos actualizados hace {int(horas_ant_fb)}h.")

    if n_sin_stock > 0:
        meds_0 = df_farm[pd.to_numeric(df_farm.get('Stock_Farm_Actual', pd.Series()), errors='coerce').fillna(0) == 0]['Medicamento'].tolist()
        issues.append(f"🔴 **{n_sin_stock} medicamento(s) con stock = 0** en farmacia: "
                      + ", ".join(meds_0[:5]) + ("..." if len(meds_0) > 5 else ""))
    else:
        ok_items.append("✅ Ningún medicamento con stock cero en farmacia.")

    if n_crit_farm > 10:
        issues.append(f"🟠 **{n_crit_farm} medicamentos Crítico/Urgente** — revisar pestaña 📝 Pedido a Bodega AA.")
    elif n_crit_farm > 0:
        ok_items.append(f"🟡 {n_crit_farm} medicamentos en nivel Crítico/Urgente (dentro de rango).")
    else:
        ok_items.append("✅ Ningún medicamento en estado Crítico/Urgente.")

    if not df_dial_farm.empty:
        ok_items.append(f"✅ Datos de diálisis cargados: {len(df_dial_farm)} medicamentos.")
    else:
        issues.append("🟡 **Sin datos de diálisis** — regenera el Consolidado con `EJECUTAR_MAESTRO.bat`.")

    if issues:
        st.markdown("**Problemas detectados:**")
        for iss in issues:
            st.markdown(f"- {iss}")
    if ok_items:
        with st.expander("Ver chequeos OK", expanded=False):
            for ok in ok_items:
                st.markdown(f"- {ok}")

    st.markdown("---")

    # ── Formulario de sugerencias ─────────────────────────────────────────────
    st.markdown("### ✍️ Registrar sugerencia o problema")
    st.caption("Tus comentarios quedan guardados en `feedback.json` dentro de la carpeta de la app.")

    _FEEDBACK_FILE = os.path.join(WORK_DIR, "feedback.json")

    def _load_feedback():
        if os.path.exists(_FEEDBACK_FILE):
            try:
                with open(_FEEDBACK_FILE, encoding='utf-8') as _f:
                    return _json.load(_f)
            except Exception:
                return []
        return []

    def _save_feedback(entries):
        with open(_FEEDBACK_FILE, 'w', encoding='utf-8') as _f:
            _json.dump(entries, _f, ensure_ascii=False, indent=2)

    cat_opts = ["🐛 Error / dato incorrecto", "💡 Mejora sugerida",
                "❓ Consulta / duda", "✅ Funciona bien — comentario positivo"]

    with st.form("form_feedback", clear_on_submit=True):
        fb_cat  = st.selectbox("Categoría", cat_opts)
        fb_pest = st.selectbox("Pestaña afectada (opcional)",
                               ["(ninguna / general)", "Pedido a Bodega AA",
                                "Pedido a Bodega Fármacos", "Diálisis", "Faltantes"])
        fb_txt  = st.text_area("Describe el problema o sugerencia", height=120,
                               placeholder="Ej: El stock de AMLODIPINO aparece diferente al sistema SSASUR...")
        fb_user = st.text_input("Tu nombre (opcional)", placeholder="Ej: María / Turno mañana")
        enviado = st.form_submit_button("💾 Guardar", type="primary", use_container_width=True)

    if enviado and fb_txt.strip():
        entries = _load_feedback()
        entries.append({
            "fecha"    : datetime.now().strftime("%Y-%m-%d %H:%M"),
            "categoria": fb_cat,
            "pestana"  : fb_pest,
            "texto"    : fb_txt.strip(),
            "usuario"  : fb_user.strip() or "Anónimo",
        })
        _save_feedback(entries)
        st.success("✅ Sugerencia guardada. ¡Gracias por el feedback!")
    elif enviado:
        st.warning("Escribe algo en el campo de descripción.")

    # ── Historial de feedback ────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📋 Historial de sugerencias")
    entries_hist = _load_feedback()
    if entries_hist:
        for e in reversed(entries_hist[-20:]):
            with st.expander(f"{e['fecha']}  ·  {e['categoria']}  ·  {e['usuario']}"):
                if e.get('pestana') and e['pestana'] != "(ninguna / general)":
                    st.caption(f"Pestaña: {e['pestana']}")
                st.markdown(e['texto'])
    else:
        st.info("Aún no hay sugerencias registradas. ¡Sé el primero!")

# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.caption(
    f"Maestro AA  ·  Universo: {len(todos_meds)} medicamentos  ·  "
    f"Datos: {datos['mtime'].strftime('%d/%m/%Y %H:%M')}"
)
