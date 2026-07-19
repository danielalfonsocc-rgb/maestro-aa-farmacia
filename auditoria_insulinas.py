#!/usr/bin/env python3
"""
auditoria_insulinas.py
════════════════════════════════════════════════════════════════════════════════
Auditoría de INSULINAS despachadas y/o pendientes del mes — Farmacia AT Abierta
Hospital de Pitrufquén (SSASur)

Para cada prescripción de insulina del mes (despachada o pendiente) extrae la
posología escrita por el médico ("Observación Médica Prescripción") y, cuando
no está disponible en la línea del mes, la infiere de la receta más reciente
del mismo paciente + mismo tipo de insulina en el histórico.

Hojas Excel:
  1. Insulinas Este Mes   — 1 fila por prescripción (Despachado/Pendiente + posología)
  2. Histórico por Paciente — evolución de la posología de cada paciente/insulina
  3. Sin Posología        — casos sin ningún registro de posología (mes ni histórico)
  4. Metodología

Uso:
    py auditoria_insulinas.py
    py auditoria_insulinas.py --mes 2026-07
    py auditoria_insulinas.py --salida mi_reporte.xlsx
"""
import argparse
import math
import os
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aa_colors import TEAL, ROJO, GRIS
from utils_aa import cargar_recetas_csv, setup_stdout

setup_stdout()
WORK = os.path.dirname(os.path.abspath(__file__))

BODEGA_OBJETIVO  = "FARMACIA AT ABIERTA"
ESTADOS_EXCLUIDOS = {"ANULADO", "RECHAZADO", "REEMPLAZADO", "DEVUELTO"}
ESTADOS_DESPACHADO = {"ENTREGADO"}
ESTADOS_PENDIENTE  = {"PENDIENTE", "SOLICITADO"}

TIPOS_INSULINA = [
    "GLARGINA", "GLULISINA", "ASPARTA", "DEGLUDEC", "LISPRO",
    "HUMANA NPH", "NPH", "CRISTALINA", "REGULAR",
]

DIAS_CICLO    = 30   # 1 mes de tratamiento = 30 días, igual que el resto del proyecto.
MEAL_KW       = ["DESAYUNO", "ALMUERZO", "CENA", "ONCE"]
MULTI_TOMA_KW = ["COMIDA", "PRANDIAL"]  # "COMIDA" cubre COMIDA/COMIDAS/CADA COMIDA/etc.

PRIO_COLOR_LAP = {
    "URGENTE": ("DC2626", "FEE2E2"),
    "REVISAR": ("C2410C", "FFF3E0"),
    "OK":      ("15803D", "F0FDF4"),
}
_ORDEN_PRIO_LAP = {"URGENTE": 0, "REVISAR": 1, "OK": 2}


def _capacidad_ui_lapiz(presc: str):
    """UI totales contenidas en 1 lápiz/frasco, según el nombre del producto."""
    p = str(presc).upper()
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*UI\s*/\s*(\d+(?:[.,]\d+)?)\s*ML", p)
    if m:
        return float(m.group(1).replace(",", "."))
    m_conc = re.search(r"(\d+(?:[.,]\d+)?)\s*UI\s*/\s*ML", p)
    m_vol  = re.search(r"FC\s*(\d+(?:[.,]\d+)?)\s*ML", p)
    if m_conc and m_vol:
        return float(m_conc.group(1).replace(",", ".")) * float(m_vol.group(1).replace(",", "."))
    return None


def _dosis_diaria_ui(texto: str):
    """
    Estima la dosis diaria total (UI) desde el texto libre de posología.
    Retorna (dosis_ui | None, confiable: bool, detalle: str).
    Solo calcula cuando el patrón es inequívoco; en cualquier caso ambiguo
    devuelve confiable=False para que el caso se marque URGENTE y lo revise el QF.
    """
    t = str(texto).upper().strip()
    if not t:
        return None, False, "Sin texto de posología"

    # Corta notas de corrección/SOS o umbral de glicemia (ej. "SOS SUBIR 2 U SI
    # GLICEMIAS + 300"): son dosis condicionales, no forman parte de la dosis
    # base diaria y su número ("+ 300") no debe sumarse como si fuera otra toma.
    m_sos = re.search(r"\bSOS\b|GLICEMIA", t)
    t_base = t[:m_sos.start()].strip() if m_sos else t
    if not t_base:
        return None, False, "Posología es solo nota SOS/condicional — verificar dosis base manualmente"

    # 1) Lista de tomas al inicio del texto: "8 - 12 - 4", "10 -10 -10"
    m = re.match(r"^(\d+(?:[.,]\d+)?(?:\s*[-,]\s*\d+(?:[.,]\d+)?){1,4})(?=\s|$)", t_base)
    if m:
        nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", m.group(1))]
        return sum(nums), True, f"Suma de {len(nums)} tomas ({'+'.join(f'{n:g}' for n in nums)} UI)"

    # 2) Tomas separadas por "+": "10 ANTES ALMUERZO + 8 ANTES DE TOMAR ONCE"
    #    (lenient: toma el primer número de cada segmento, aunque no repita la unidad)
    if "+" in t_base:
        nums = []
        for seg in t_base.split("+"):
            m_seg = re.search(r"(\d+(?:[.,]\d+)?)", seg)
            if m_seg:
                nums.append(float(m_seg.group(1).replace(",", ".")))
        if len(nums) >= 2:
            return sum(nums), True, f"Suma de {len(nums)} tomas (+) ({'+'.join(f'{n:g}' for n in nums)} UI)"

    dose_all = re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:U\b|UI\b|UNIDAD(?:ES)?)", t_base)

    # 3) ≥2 menciones explícitas de dosis con unidad, sin conector "+"
    #    (ej. "40U EN LA MAÑANA 8U EN LA NOCHE")
    if len(dose_all) >= 2:
        nums = [float(x.replace(",", ".")) for x in dose_all]
        return sum(nums), True, f"Suma de {len(nums)} tomas ({'+'.join(f'{n:g}' for n in nums)} UI)"

    # 4) Multiplicador explícito: "X VECES AL DIA"
    mult_m = re.search(r"(\d+)\s*VECES", t_base)
    if dose_all and mult_m:
        dosis = float(dose_all[0].replace(",", "."))
        veces = int(mult_m.group(1))
        return dosis * veces, True, f"{dosis:g} UI × {veces} veces/día"

    # 5) Un solo número — ambiguo si menciona ≥2 comidas o "cada comida"/"prandial"
    if len(dose_all) == 1:
        dosis = float(dose_all[0].replace(",", "."))
        n_comidas = sum(1 for kw in MEAL_KW if kw in t_base)
        if n_comidas >= 2 or any(kw in t_base for kw in MULTI_TOMA_KW):
            return None, False, (f"Posología menciona {dosis:g} UI en varias comidas sin "
                                  "detallar cada toma — verificar dosis diaria real")
        return dosis, True, f"{dosis:g} UI, 1 vez/día"

    return None, False, "No se detectó una dosis numérica en el texto"


def _evaluar_lapices(dosis_dia, confiable: bool, detalle_dosis: str,
                     capacidad_ui, recetados: int):
    """Retorna (prioridad, observación) comparando lápices esperados vs. recetados."""
    if capacidad_ui is None:
        return "REVISAR", ("No se pudo determinar el contenido UI del envase para calcular "
                            "lápices esperados — revisar cantidad de lápices manualmente.")
    if not confiable or dosis_dia is None:
        return "URGENTE", (f"{detalle_dosis}. Revisar cantidad de lápices manualmente: "
                            "no se pudo calcular la dosis diaria con certeza.")

    esperados = math.ceil(round(dosis_dia * DIAS_CICLO / capacidad_ui, 6))
    diff = int(recetados) - esperados
    base = (f"{detalle_dosis} → {dosis_dia:g} UI/día × {DIAS_CICLO} días ÷ "
            f"{capacidad_ui:g} UI/lápiz = {esperados} lápiz(ces) esperado(s); "
            f"recetados: {int(recetados)}.")

    if diff == 0:
        return "OK", base
    if abs(diff) == 1:
        return "REVISAR", base + " Diferencia de 1 lápiz (margen habitual) — revisar cantidad de lápices."
    falta_o_exceso = "faltan lápices para cubrir el mes" if diff < 0 else "exceso de lápices recetados"
    return "URGENTE", base + f" No concuerda ({falta_o_exceso}) — revisar cantidad de lápices."

THIN   = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    "ID Receta Detalle", "RUN", "Nombre", "Apellido Paterno", "Apellido Materno",
    "Prescripción", "Número Receta", "Bodega Despacha",
    "Fecha Atención", "Fecha Entrega Receta", "Estado Prescripción",
    "Nombre Profesional", "Apellido Paterno Profesional",
    "Apellido Materno Profesional", "Especialidad",
    "Cod. Diagnóstico 1", "Diagnóstico 1",
    "Cantidad Recetada", "Cantidad Entregada",
    "Observación Médica Prescripción",
]


def _tipo_insulina(presc: str) -> str:
    p = str(presc).upper()
    for kw in TIPOS_INSULINA:
        if kw in p:
            return kw
    return p.strip()


def cargar_y_preparar() -> pd.DataFrame:
    try:
        rec = cargar_recetas_csv(WORK, cols=COLS, solo_ultimo=False)
    except FileNotFoundError:
        print("[AVISO] No hay CSV de recetas. Ejecuta AUTO_SSASUR.bat primero.")
        sys.exit(0)

    rec = rec.copy()
    presc = rec["Prescripción"].fillna("").str.upper()
    es_insulina = presc.str.contains("INSULINA", regex=False) & ~presc.str.contains(
        "JERINGA|AGUJA", regex=True)
    bod = rec["Bodega Despacha"].fillna("").str.upper().str.strip()
    est = rec["Estado Prescripción"].fillna("").str.upper().str.strip()

    rec = rec[es_insulina & (bod == BODEGA_OBJETIVO) & (~est.isin(ESTADOS_EXCLUIDOS))].copy()
    if rec.empty:
        print("[AVISO] No se encontraron prescripciones de insulina en Farmacia AT Abierta.")
        sys.exit(0)

    fa = pd.to_datetime(rec["Fecha Atención"],       dayfirst=True, errors="coerce")
    fe = pd.to_datetime(rec["Fecha Entrega Receta"], dayfirst=True, errors="coerce")
    rec["_fecha"] = fa.fillna(fe)
    rec = rec.dropna(subset=["_fecha"])
    rec["_mes"]   = rec["_fecha"].dt.to_period("M")

    rec["_paciente"] = (
        rec["Nombre"].fillna("") + " " + rec["Apellido Paterno"].fillna("") +
        " " + rec["Apellido Materno"].fillna("")
    ).str.strip().str.title()
    rec["_medico"] = (
        rec["Nombre Profesional"].fillna("") + " " +
        rec["Apellido Paterno Profesional"].fillna("") + " " +
        rec["Apellido Materno Profesional"].fillna("")
    ).str.strip().str.title()
    rec["_tipo"]   = rec["Prescripción"].apply(_tipo_insulina)
    rec["_pos"]    = rec["Observación Médica Prescripción"].fillna("").str.strip()
    rec["_cant_r"] = pd.to_numeric(rec["Cantidad Recetada"],  errors="coerce").fillna(0)
    rec["_cant_e"] = pd.to_numeric(rec["Cantidad Entregada"], errors="coerce").fillna(0)
    rec["_estado_norm"] = est.reindex(rec.index)

    rec = rec.sort_values("_fecha")
    return rec


def posologia_historica(rec: pd.DataFrame, run: str, tipo: str, antes_de: pd.Timestamp):
    """Última posología no vacía del mismo paciente+tipo de insulina, antes de `antes_de`."""
    prev = rec[(rec["RUN"] == run) & (rec["_tipo"] == tipo) &
               (rec["_fecha"] < antes_de) & (rec["_pos"] != "")]
    if prev.empty:
        return None, None
    ult = prev.iloc[-1]
    return ult["_pos"], ult["_fecha"]


def construir_reportes(rec: pd.DataFrame, mes: pd.Period):
    del_mes = rec[rec["_mes"] == mes].copy()
    if del_mes.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    filas = []
    for _, r in del_mes.iterrows():
        estado_norm = r["_estado_norm"]
        if estado_norm in ESTADOS_DESPACHADO or r["_cant_e"] > 0:
            estado_disp = "Despachado"
        elif estado_norm in ESTADOS_PENDIENTE:
            estado_disp = "Pendiente"
        else:
            estado_disp = estado_norm.title() or "—"

        pos_mes = r["_pos"]
        if pos_mes:
            pos_final, fuente = pos_mes, "Receta de este mes"
        else:
            pos_hist, fecha_hist = posologia_historica(rec, r["RUN"], r["_tipo"], r["_fecha"])
            if pos_hist:
                pos_final, fuente = pos_hist, f"Histórico ({fecha_hist:%d-%m-%Y})"
            else:
                pos_final, fuente = "", "Sin registro"

        capacidad = _capacidad_ui_lapiz(r["Prescripción"])
        dosis_dia, confiable, detalle_dosis = _dosis_diaria_ui(pos_final)
        prioridad, observacion = _evaluar_lapices(
            dosis_dia, confiable, detalle_dosis, capacidad, r["_cant_r"])
        esperados = (math.ceil(round(dosis_dia * DIAS_CICLO / capacidad, 6))
                     if (capacidad and dosis_dia is not None) else None)

        filas.append({
            "Prioridad revisión"    : prioridad,
            "RUN"                  : r["RUN"],
            "Paciente"              : r["_paciente"],
            "Medicamento"           : r["Prescripción"],
            "Tipo insulina"         : r["_tipo"],
            "Estado"                : estado_disp,
            "N° Receta"             : r["Número Receta"],
            "Fecha"                 : r["_fecha"],
            "Cant. recetada"        : int(r["_cant_r"]),
            "Cant. entregada"       : int(r["_cant_e"]),
            "Dosis diaria detectada (UI)": dosis_dia if dosis_dia is not None else "—",
            "Lápices esperados (30 días)": esperados if esperados is not None else "—",
            "Diferencia (lápices)"  : (int(r["_cant_r"]) - esperados) if esperados is not None else "—",
            "Observación"           : observacion,
            "Médico"                : r["_medico"],
            "Especialidad"          : r["Especialidad"] or "—",
            "Diagnóstico"           : (str(r["Cod. Diagnóstico 1"]).strip() + " · " +
                                        str(r["Diagnóstico 1"]).strip()).strip(" ·") or "—",
            "Posología (este mes)"  : pos_mes,
            "Posología detectada"   : pos_final,
            "Fuente posología"      : fuente,
        })

    resumen = pd.DataFrame(filas)
    resumen["_o"] = resumen["Prioridad revisión"].map(_ORDEN_PRIO_LAP)
    resumen = resumen.sort_values(
        ["_o", "Estado", "Paciente"], ascending=[True, False, True]).drop(columns=["_o"])

    # ── Histórico por paciente (para pacientes del mes) ─────────────────────
    pares = resumen[["RUN", "Tipo insulina"]].drop_duplicates()
    hist_rows = []
    for _, p in pares.iterrows():
        sub = rec[(rec["RUN"] == p["RUN"]) & (rec["_tipo"] == p["Tipo insulina"]) &
                  (rec["_pos"] != "")].sort_values("_fecha", ascending=False)
        for _, r in sub.iterrows():
            hist_rows.append({
                "RUN"          : r["RUN"],
                "Paciente"     : r["_paciente"],
                "Tipo insulina": r["_tipo"],
                "Fecha"        : r["_fecha"],
                "Medicamento"  : r["Prescripción"],
                "Estado"       : r["_estado_norm"].title(),
                "Posología"    : r["_pos"],
                "Médico"       : r["_medico"],
            })
    historico = pd.DataFrame(hist_rows)

    sin_pos = resumen[resumen["Posología detectada"] == ""].copy()

    return resumen, historico, sin_pos


# ── Escritura Excel ──────────────────────────────────────────────────────────

def _encabezado_hoja(ws, texto: str, color: str, n_cols: int):
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = texto
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _escribir_tabla(ws, df: pd.DataFrame, fila_inicio: int,
                    anchos: dict | None = None, fechas: list | None = None,
                    wrap_cols: tuple = (), estado_col: str | None = None,
                    prio_col: str | None = None):
    fechas = fechas or []
    cols   = list(df.columns)
    col_ix = {c: i + 1 for i, c in enumerate(cols)}

    for c_i, col in enumerate(cols, 1):
        cell = ws.cell(row=fila_inicio, column=c_i, value=col)
        cell.fill      = PatternFill("solid", fgColor=TEAL)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[fila_inicio].height = 28
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)

    for r_off, (_, row) in enumerate(df.iterrows()):
        r = fila_inicio + 1 + r_off
        ws.append(list(row.values))
        for c_i, col in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c_i)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col in wrap_cols))
            if prio_col and col == prio_col:
                txt_c, bg_c = PRIO_COLOR_LAP.get(str(cell.value), ("1F2937", "FFFFFF"))
                cell.font = Font(bold=True, color=txt_c)
                cell.fill = PatternFill("solid", fgColor=bg_c)
            elif estado_col and col == estado_col:
                if str(cell.value) == "Pendiente":
                    cell.font = Font(bold=True, color=ROJO)
                elif str(cell.value) == "Despachado":
                    cell.font = Font(bold=True, color="15803D")
            elif r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRIS)
        ws.row_dimensions[r].height = 32 if any(col in wrap_cols for col in cols) else 20

    for fcol in fechas:
        if fcol in col_ix:
            for r in range(fila_inicio + 1, ws.max_row + 1):
                ws.cell(row=r, column=col_ix[fcol]).number_format = "DD-MM-YYYY"

    for i, col in enumerate(cols, 1):
        w = (anchos or {}).get(col)
        if w is None:
            vals = df[col].astype(str)
            w = min(max(len(str(col)) + 2, vals.str.len().max() + 2 if len(df) else 12), 55)
        ws.column_dimensions[get_column_letter(i)].width = w

    ultima = ws.max_row
    ws.auto_filter.ref = f"A{fila_inicio}:{get_column_letter(len(cols))}{ultima}"


ANCHOS_RES = {
    "Prioridad revisión": 14, "Paciente": 26, "Medicamento": 42, "Tipo insulina": 14,
    "Estado": 12, "N° Receta": 12, "Fecha": 13,
    "Dosis diaria detectada (UI)": 14, "Lápices esperados (30 días)": 16,
    "Diferencia (lápices)": 14, "Observación": 55,
    "Médico": 30, "Especialidad": 20,
    "Diagnóstico": 34, "Posología (este mes)": 40, "Posología detectada": 40,
    "Fuente posología": 20,
}
ANCHOS_HIST = {
    "Paciente": 26, "Tipo insulina": 14, "Fecha": 13, "Medicamento": 42,
    "Estado": 12, "Posología": 40, "Médico": 30,
}


def exportar_excel(resumen, historico, sin_pos, mes: pd.Period, dest: str):
    wb = Workbook()

    n_total = len(resumen)
    n_desp  = int((resumen["Estado"] == "Despachado").sum()) if n_total else 0
    n_pend  = int((resumen["Estado"] == "Pendiente").sum()) if n_total else 0
    n_hist  = int((resumen["Fuente posología"].str.startswith("Histórico")).sum()) if n_total else 0
    n_sin   = len(sin_pos)
    n_urg   = int((resumen["Prioridad revisión"] == "URGENTE").sum()) if n_total else 0
    n_rev   = int((resumen["Prioridad revisión"] == "REVISAR").sum()) if n_total else 0
    n_ok    = int((resumen["Prioridad revisión"] == "OK").sum()) if n_total else 0

    ws1 = wb.active
    ws1.title = "Insulinas Este Mes"
    res_disp = resumen.drop(columns=["RUN"], errors="ignore")
    _encabezado_hoja(ws1,
                     f"INSULINAS — DESPACHADAS Y PENDIENTES · {mes}  ·  "
                     f"{n_total} prescripciones ({n_desp} despachadas, {n_pend} pendientes)  ·  "
                     f"URGENTE: {n_urg}  ·  REVISAR: {n_rev}  ·  OK: {n_ok}",
                     TEAL, max(len(res_disp.columns), 10))
    _escribir_tabla(ws1, res_disp, fila_inicio=2, anchos=ANCHOS_RES, fechas=["Fecha"],
                    wrap_cols=("Posología (este mes)", "Posología detectada", "Médico", "Observación"),
                    estado_col="Estado", prio_col="Prioridad revisión")

    ws2 = wb.create_sheet("Histórico por Paciente")
    hist_disp = historico.drop(columns=["RUN"], errors="ignore")
    _encabezado_hoja(ws2, "EVOLUCIÓN DE POSOLOGÍA POR PACIENTE (histórico completo)",
                     "0F766E", max(len(hist_disp.columns), 8) if len(hist_disp) else 8)
    if len(hist_disp):
        _escribir_tabla(ws2, hist_disp, fila_inicio=2, anchos=ANCHOS_HIST, fechas=["Fecha"],
                        wrap_cols=("Posología", "Médico"))
    else:
        ws2.cell(row=2, column=1, value="Sin histórico disponible.").font = Font(italic=True, color="6B7280")

    ws3 = wb.create_sheet("Sin Posología")
    sp_disp = sin_pos.drop(columns=["RUN", "Posología (este mes)", "Posología detectada",
                                     "Fuente posología"], errors="ignore")
    _encabezado_hoja(ws3, f"CASOS SIN REGISTRO DE POSOLOGÍA (mes ni histórico) · {n_sin} caso(s)",
                     ROJO, max(len(sp_disp.columns), 8) if len(sp_disp) else 8)
    if len(sp_disp):
        _escribir_tabla(ws3, sp_disp, fila_inicio=2, anchos=ANCHOS_RES, fechas=["Fecha"],
                        wrap_cols=("Observación", "Médico"),
                        estado_col="Estado", prio_col="Prioridad revisión")
    else:
        ws3.cell(row=2, column=1, value="Todos los casos tienen posología registrada.").font = Font(
            italic=True, color="15803D")

    ws4 = wb.create_sheet("Metodología")
    txt = [
        ("AUDITORÍA DE INSULINAS — Despachadas y Pendientes del Mes", True),
        (f"Generado: {datetime.now():%d-%m-%Y %H:%M}  ·  Hospital de Pitrufquén (SSASur, Chile)", False),
        ("", False),
        ("Alcance", True),
        (f"• Mes analizado: {mes}  ·  Bodega: {BODEGA_OBJETIVO}", False),
        ("• Incluye medicamentos cuyo nombre contiene 'INSULINA' (excluye jeringas y agujas).", False),
        (f"• Excluidos estados: {', '.join(sorted(ESTADOS_EXCLUIDOS))}", False),
        ("• Despachado = Estado 'Entregado' o Cantidad Entregada > 0.", False),
        ("• Pendiente  = Estado 'Pendiente' o 'Solicitado' (aún no retirado).", False),
        ("", False),
        ("Detección de posología", True),
        ("• Fuente primaria: campo 'Observación Médica Prescripción' de la receta del mes.", False),
        ("• Si está vacío, se usa la posología no vacía más reciente del mismo paciente + mismo", False),
        ("  tipo de insulina (Glargina, Glulisina, Asparta, NPH, Cristalina, etc.) en el histórico", False),
        ("  completo de recetas, indicando la fecha de esa receta como fuente.", False),
        ("• El texto se muestra tal como lo escribió el médico — no se reinterpreta la dosis.", False),
        ("• Hoja 'Histórico por Paciente': todas las posologías registradas históricamente para", False),
        ("  cada paciente/tipo de insulina del mes, para ver la evolución de la dosis.", False),
        ("", False),
        ("Chequeo de cantidad de lápices/frascos (columna 'Prioridad revisión')", True),
        ("• Dosis diaria (UI): se estima desde el texto de posología solo cuando el patrón es", False),
        ("  inequívoco (una toma diaria, suma explícita de varias tomas, o multiplicador 'X veces/día').", False),
        ("  Si la posología menciona varias comidas sin detallar cada toma (ej. 'antes de cada comida'),", False),
        ("  no se calcula — se marca URGENTE para que el QF confirme la dosis real.", False),
        ("• Contenido UI del envase: se extrae del nombre del producto (ej. lápiz de 300 UI, frasco de", False),
        ("  1000 UI).", False),
        (f"• Lápices esperados = ceil(dosis diaria × {DIAS_CICLO} días ÷ UI por envase). Se compara contra", False),
        ("  la Cantidad Recetada del mes:", False),
        ("    OK       — coincide exactamente.", False),
        ("    REVISAR  — difiere en 1 lápiz (margen habitual, igual conviene mirarlo).", False),
        ("    URGENTE  — no se pudo calcular la dosis con certeza, o difiere en ≥2 lápices", False),
        ("               (faltan lápices para cubrir el mes, o exceso importante).", False),
        ("• Esto es un TAMIZAJE automático, no un cálculo clínico definitivo — siempre confirmar con", False),
        ("  la ficha clínica y el criterio del QF antes de ajustar una receta.", False),
        ("", False),
        ("Resultados", True),
        (f"• Prescripciones del mes    : {n_total:,}", False),
        (f"• Despachadas               : {n_desp:,}", False),
        (f"• Pendientes                : {n_pend:,}", False),
        (f"• Posología inferida del histórico: {n_hist:,}", False),
        (f"• Sin ningún registro de posología: {n_sin:,}", False),
        (f"• URGENTE (revisar lápices) : {n_urg:,}", False),
        (f"• REVISAR (diferencia leve) : {n_rev:,}", False),
        (f"• OK (cantidad coincide)    : {n_ok:,}", False),
        ("", False),
        ("Privacidad", True),
        ("• Este archivo contiene datos de salud de pacientes (Ley 19.628).", False),
        ("• No distribuir fuera de la institución ni publicar en redes compartidas.", False),
    ]
    for i, (line, bold) in enumerate(txt, 1):
        c = ws4.cell(row=i, column=1, value=line)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 10, color=TEAL if bold else "1F2937")
    ws4.column_dimensions["A"].width = 100

    wb.save(dest)


def main():
    ap = argparse.ArgumentParser(
        description="Auditoría de insulinas despachadas/pendientes del mes con detección de posología.")
    ap.add_argument("--mes", default=None, help="Mes a analizar, formato YYYY-MM (default: mes actual).")
    ap.add_argument("--salida", default=None, help="Ruta del Excel de salida.")
    args = ap.parse_args()

    mes = pd.Period(args.mes) if args.mes else pd.Period(datetime.now(), freq="M")
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    dest = args.salida or os.path.join(WORK, f"Auditoria_Insulinas_{mes}_{ts}.xlsx")

    print("=" * 70)
    print("  AUDITORÍA DE INSULINAS — DESPACHADAS Y PENDIENTES DEL MES")
    print(f"  Farmacia AT Abierta · Hospital de Pitrufquén · Mes: {mes}")
    print("=" * 70)

    rec = cargar_y_preparar()
    print(f"Líneas de insulina (histórico, Farmacia AT Abierta): {len(rec):,}")

    resumen, historico, sin_pos = construir_reportes(rec, mes)
    if resumen.empty:
        print(f"\n[AVISO] No hay prescripciones de insulina en {mes}.")
        sys.exit(0)

    n_desp = int((resumen["Estado"] == "Despachado").sum())
    n_pend = int((resumen["Estado"] == "Pendiente").sum())
    n_hist = int(resumen["Fuente posología"].str.startswith("Histórico").sum())

    print(f"\nPrescripciones del mes : {len(resumen):,}")
    print(f"  Despachadas          : {n_desp:,}")
    print(f"  Pendientes           : {n_pend:,}")
    print(f"  Posología del histórico usada: {n_hist:,}")
    print(f"  Sin posología alguna : {len(sin_pos):,}")

    print("\n  Exportando Excel...")
    exportar_excel(resumen, historico, sin_pos, mes, dest)
    print(f"\n[OK] Excel generado: {dest}  ({os.path.getsize(dest) // 1024} KB)")


if __name__ == "__main__":
    main()
