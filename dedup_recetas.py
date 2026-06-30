#!/usr/bin/env python3
"""
dedup_recetas.py — Detecta y limpia recetas duplicadas por sobre-extracción.

Problema: al descargar el informe GT varias veces con rangos solapados, la misma
Nº Receta aparece en múltiples archivos reporteGestionTerritorial_*.xlsx. Al correr
cruce_gt.py sobre cada uno por separado, se generan planillas con pacientes repetidos.

Solución:
  1. GT: detecta Nº Receta en >1 archivo → las elimina del archivo MÁS ANTIGUO,
     dejando cada receta solo en el Excel más reciente.
  2. CSV sábana: detecta archivos con rangos de fecha solapados → reporta cuáles
     son redundantes (la carga ya deduplica por ID Receta Detalle, pero acumulan
     espacio innecesario).

Uso:
  py dedup_recetas.py                 # solo analizar (sin cambios)
  py dedup_recetas.py --limpiar       # limpiar (crea .bak antes de modificar)
  py dedup_recetas.py --gt-dir <ruta> # usar otra carpeta GT
"""
import argparse, glob, os, re, shutil, sys, unicodedata
from collections import defaultdict
from datetime import datetime

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
GT_DIR_DEFAULT = os.path.join(os.path.dirname(MAESTRO_DIR), "04_Farmacia_Gestion_Territorial")


def _key(h):
    h = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", h)


def _cargar_gt_xlsx(path):
    """Retorna (set_numeros_receta, headers, rows_data) del GT Excel."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return set(), [], []
    hi = next((i for i, r in enumerate(rows)
               if sum(1 for c in r if c not in (None, "")) >= 5), None)
    if hi is None:
        return set(), [], []
    hdr = [str(c).strip() if c is not None else "" for c in rows[hi]]
    K = {_key(h): i for i, h in enumerate(hdr)}
    c_rec = next((K[k] for k in ("nreceta", "noreceta", "numeroreceta") if k in K), 0)
    nums = set()
    data_rows = []
    for r in rows[hi + 1:]:
        if not r:
            continue
        v = str(r[c_rec]).strip() if c_rec < len(r) and r[c_rec] is not None else ""
        if v and v.upper() not in ("TOTAL", "NONE", ""):
            nums.add(v)
            data_rows.append((v, r))
    return nums, hdr, data_rows


def _guardar_gt_xlsx_sin(path_orig, path_dest, excluir_recetas):
    """Guarda una copia del GT Excel omitiendo las recetas en excluir_recetas."""
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb_in = load_workbook(path_orig, read_only=True, data_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    rows = [r for r in ws_in.iter_rows(values_only=True)]
    wb_in.close()

    hi = next((i for i, r in enumerate(rows)
               if sum(1 for c in r if c not in (None, "")) >= 5), None)
    if hi is None:
        return 0
    hdr_row = rows[hi]
    K = {_key(str(c).strip() if c else ""): i for i, c in enumerate(hdr_row)}
    c_rec = next((K[k] for k in ("nreceta", "noreceta", "numeroreceta") if k in K), 0)

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title if hasattr(ws_in, "title") else "Hoja1"

    # Copiar filas anteriores al encabezado (título del reporte, etc.)
    for r in rows[:hi]:
        ws_out.append(list(r))
    ws_out.append(list(hdr_row))

    eliminadas = 0
    for r in rows[hi + 1:]:
        if not r:
            continue
        v = str(r[c_rec]).strip() if c_rec < len(r) and r[c_rec] is not None else ""
        if v and v.upper() not in ("TOTAL", "NONE", "") and v in excluir_recetas:
            eliminadas += 1
            continue
        ws_out.append(list(r))

    wb_out.save(path_dest)
    return eliminadas


# ── GT dedup ──────────────────────────────────────────────────────────────────
def analizar_gt(gt_dir, limpiar=False):
    patron = os.path.join(gt_dir, "reporteGestionTerritorial_*.xlsx")
    archivos = sorted(glob.glob(patron), key=os.path.getmtime)  # orden cronológico
    if not archivos:
        print(f"[GT] No hay archivos GT en {gt_dir}")
        return

    print(f"\n[GT] {len(archivos)} archivo(s) — analizando duplicados entre rangos:")
    receta_a_archivos = defaultdict(list)  # Nº Receta → [archivos que la contienen]
    info = {}
    for arch in archivos:
        nombre = os.path.basename(arch)
        nums, _, _ = _cargar_gt_xlsx(arch)
        info[arch] = nums
        mtime = datetime.fromtimestamp(os.path.getmtime(arch)).strftime("%d/%m %H:%M")
        print(f"  {nombre}: {len(nums):>4} recetas  ({mtime})")
        for n in nums:
            receta_a_archivos[n].append(arch)

    duplicadas = {r: files for r, files in receta_a_archivos.items() if len(files) > 1}
    if not duplicadas:
        print("  ✓ Sin duplicados entre archivos GT")
        return 0

    # Agrupar por pares de archivos que comparten recetas
    pares = defaultdict(list)
    for rec, files in duplicadas.items():
        clave = tuple(sorted(files))
        pares[clave].append(rec)

    n_dup = len(duplicadas)
    print(f"\n  ⚠  {n_dup} receta(s) en más de un archivo:")
    for clave, recetas in sorted(pares.items()):
        nombres = " ∩ ".join(os.path.basename(a) for a in clave)
        print(f"    {nombres}: {len(recetas)} receta(s)")

    if limpiar:
        print("\n  [LIMPIAR] Eliminando duplicados — cada receta queda solo en el más reciente:")
        # Para cada receta duplicada: conservar en el archivo MÁS NUEVO (último en la lista
        # ordenada por mtime), eliminar de los anteriores.
        a_excluir = defaultdict(set)  # archivo → set de recetas a quitar
        for rec, files in duplicadas.items():
            # files está en orden cronológico (sorted por mtime arriba)
            for arch in files[:-1]:   # todos menos el más reciente
                a_excluir[arch].add(rec)

        for arch, excluir in sorted(a_excluir.items()):
            nombre = os.path.basename(arch)
            bak = arch + ".bak"
            shutil.copy2(arch, bak)
            eliminadas = _guardar_gt_xlsx_sin(arch, arch, excluir)
            print(f"    {nombre}: {eliminadas} fila(s) eliminadas (.bak guardado)")

    return n_dup


# ── CSV sábana dedup ──────────────────────────────────────────────────────────
def analizar_csv(maestro_dir, limpiar=False):
    archivos = sorted(glob.glob(os.path.join(maestro_dir, "informe_completo_recetas*.csv")))
    if not archivos:
        print("\n[CSV] No hay sábanas CSV")
        return

    print(f"\n[CSV] {len(archivos)} sábana(s):")
    formato_nuevo = [f for f in archivos if "_b" in os.path.basename(f)]
    formato_viejo = [f for f in archivos if "_b" not in os.path.basename(f)]

    for f in archivos:
        kb = os.path.getsize(f) // 1024
        tag = "(nuevo _b)" if "_b" in os.path.basename(f) else "(formato antiguo)"
        print(f"  {os.path.basename(f)}: {kb:,} KB  {tag}")

    # Detectar archivos del formato antiguo que ya están cubiertos por los bloques nuevos
    if formato_viejo and formato_nuevo:
        print(f"\n  Hay {len(formato_viejo)} archivo(s) en formato antiguo (pre-bloques).")
        print("  cargar_recetas_csv() los carga todos y deduplica por ID Receta Detalle.")
        print("  Ocupan espacio extra pero NO generan conteos incorrectos.")
        if limpiar:
            total_kb = sum(os.path.getsize(f) for f in formato_viejo) // 1024
            print(f"  [LIMPIAR] Moviendo {len(formato_viejo)} CSV antiguo(s) a _csv_bak/ ({total_kb:,} KB)...")
            bak_dir = os.path.join(maestro_dir, "_csv_bak")
            os.makedirs(bak_dir, exist_ok=True)
            for f in formato_viejo:
                dest = os.path.join(bak_dir, os.path.basename(f))
                shutil.move(f, dest)
                print(f"    Movido: {os.path.basename(f)}")
            print(f"  Backup en: {bak_dir}")
    else:
        print("  ✓ Sin mezcla de formatos CSV")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Detecta duplicados en recetas GT y CSV")
    ap.add_argument("--limpiar", action="store_true",
                    help="Aplicar limpieza activa (crea .bak antes de modificar)")
    ap.add_argument("--gt-dir", default=GT_DIR_DEFAULT,
                    help="Carpeta donde están los GT Excel")
    a = ap.parse_args()

    print("=" * 62)
    print("  DEDUP RECETAS — Análisis de duplicados por sobre-extracción")
    print("=" * 62)

    n_dup_gt = analizar_gt(a.gt_dir, a.limpiar) or 0
    analizar_csv(MAESTRO_DIR, a.limpiar)

    print("\n" + "=" * 62)
    if n_dup_gt:
        accion = "Limpieza aplicada" if a.limpiar else "Ejecuta con --limpiar para corregir"
        print(f"  ⚠  {n_dup_gt} receta(s) GT duplicadas entre archivos. {accion}.")
    else:
        print("  ✓ Sin duplicados GT detectados")
    print("=" * 62)


if __name__ == "__main__":
    main()
