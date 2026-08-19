#!/usr/bin/env python3
"""
agregar_gt_manual.py — Registra a mano recetas que SSASUR NO etiqueta con un
establecimiento de destino (el campo "Gestión Territorial Despacho" viene
vacío), pero que igual se están despachando. Como no traen destino, no
aparecen en el reporteGestionTerritorial_*.xlsx ni se sincronizan solas con
gt_maestro.sincronizar_gt_report() — hay que registrarlas a mano una vez.

Hace dos cosas con cada fila del CSV:
  1. Las agrega/actualiza en el maestro (gt_maestro.upsert_receta_maestro) —
     quedan trackeadas igual que todo lo demás: color por establecimiento,
     MAYÚSCULAS, Historial si cambia el Estado, Sheets/Drive al republicar.
  2. Genera la Nómina de Envío (mismo formato que usa el skill_gt para las
     planillas automáticas) agrupada por establecimiento, guardada en
     04_Farmacia_Gestion_Territorial/<ESTABLECIMIENTO>/Nóminas de Envío/

Uso — dos formas de entrada (ver también GT_NOMINA_PARTICULAR.bat):

  A) CSV a mano — llena una copia de _gt_manual_plantilla.csv (una fila por
     receta): receta,paciente,rut,destino,periodo,especialidad,n_presc,telefono,pendiente,refrigerado
     - destino: nombre del establecimiento tal como quieres que aparezca
       (ej. "CESFAM TEODORO SCHMIDT", "HOSPITAL TOLTEN").
     - pendiente: opcional, texto libre (ej. medicamento que falta).
     - refrigerado: opcional, nombre(s) del/los medicamento(s) termolábil(es)
       de la receta, con cantidad si se conoce (ej. "Insulina Glargina x1";
       varios separados por ";"). Si se completa, el letrero del
       establecimiento sale con el aviso ❄ REFRIGERADO y el detalle de
       paciente/RUN/receta/medicamento(+cantidad).
       py agregar_gt_manual.py --csv mi_lote.csv --dry-run   (preview)
       py agregar_gt_manual.py --csv mi_lote.csv             (aplica + genera nóminas)

  B) Excel de Modalidad de Despacho ya descargado (reporteGestionTerritorial_*.xlsx,
     el mismo que baja AUTO_SSASUR.py) — detecta SOLO las recetas que todavía
     no están en el maestro Y que además NO traen Estab. Destino en el informe
     (las que ya traen destino las toma sola cruce_gt.py --generar, se omiten
     acá para no duplicar la Nómina de Envío del mismo destino el mismo día):
       py agregar_gt_manual.py --gt-excel reporte.xlsx --dry-run
       py agregar_gt_manual.py --gt-excel reporte.xlsx
"""
import argparse
import csv
import datetime
import os
import sys

import openpyxl

import gt_maestro as GM
import generar as G

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
GT_SOLICITUDES_DIR = os.path.join(os.path.dirname(MAESTRO_DIR), "04_Farmacia_Gestion_Territorial")

# Mismo mapeo que publicar_drive.py — nombre de carpeta local (gt_maestro)
# vs. nombre de carpeta ya usado en Drive (convención SSASUR/skill_gt).
# Incluye AMBOS órdenes de palabra ("GORBEA DSM" y "DSM GORBEA", etc.) porque
# la columna "Establecimiento de destino" del maestro gt_maestro usa
# "TIPO LUGAR" (DSM GORBEA, HOSPITAL LONCOCHE) mientras que esta convención
# histórica (Drive/skill_gt) usa "LUGAR TIPO" (GORBEA DSM, LONCOCHE HOSP) —
# confirmado 23-07-2026 al revisar los valores reales del maestro.
_CARPETA_LOCAL = {
    "CESFAM FREIRE": "CESFAM_FREIRE",
    "CESFAM HUALPIN": "CESFAM_HUALPIN",
    "CESFAM QUEPE": "CESFAM_QUEPE",
    "CESFAM TEODORO SCHMIDT": "CESFAM_TEODORO_SCHMIDT",
    "GORBEA DSM": "DSM_GORBEA",
    "DSM GORBEA": "DSM_GORBEA",
    "LONCOCHE DSM": "DSM_LONCOCHE",
    "DSM LONCOCHE": "DSM_LONCOCHE",
    "TOLTEN DSM": "DSM_TOLTEN",
    "DSM TOLTEN": "DSM_TOLTEN",
    "GORBEA HOSP": "HOSPITAL_GORBEA",
    "HOSPITAL GORBEA": "HOSPITAL_GORBEA",
    "LONCOCHE HOSP": "HOSPITAL_LONCOCHE",
    "HOSPITAL LONCOCHE": "HOSPITAL_LONCOCHE",
    "TOLTEN HOSP": "HOSPITAL_TOLTEN",
    "HOSPITAL TOLTEN": "HOSPITAL_TOLTEN",
    "PSR COMUY": "PSR_COMUY",
    "PSR QUEULE": "PSR_QUEULE",
}


def _leer_csv(path):
    filas = []
    with open(path, encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            receta = (row.get("receta") or "").strip()
            if not receta:
                continue
            filas.append({
                "receta": receta,
                "paciente": (row.get("paciente") or "").strip(),
                "rut": (row.get("rut") or "").strip(),
                "destino": (row.get("destino") or "").strip(),
                "periodo": (row.get("periodo") or "").strip(),
                "especialidad": (row.get("especialidad") or "").strip(),
                "n_presc": (row.get("n_presc") or "1").strip(),
                "telefono": (row.get("telefono") or "").strip(),
                "pendiente": (row.get("pendiente") or "").strip(),
                "refrigerado": (row.get("refrigerado") or "").strip(),
            })
    return filas


def _leer_gt_excel(path, wb_maestro):
    """Lee el reporteGestionTerritorial_*.xlsx crudo (Informe Modalidad de
    Despacho, el mismo que baja AUTO_SSASUR.py) y devuelve SOLO las recetas
    que todavía NO están en el maestro GT Y que NO traen Estab. Destino en el
    informe — las genuinamente "sin destino en SSASUR" (ver docstring del
    archivo). Las que ya traen destino se omiten (no solo se avisan, fix
    19-08-2026): esas las toma sola cruce_gt.py --generar en la misma corrida
    automática, y dejarlas pasar acá generaba una segunda Nómina de Envío
    aparte para el mismo destino el mismo día (caso real: PSR Queule 19-08,
    3 recetas con destino ya asignado procesadas a mano ANTES de que corriera
    el automático, que las excluyó por dedup al encontrarlas ya en el maestro).
    Formato real confirmado 24-07-2026 (fila 1 = título fusionado, fila 2 =
    encabezado real): N° Receta, Paciente, Run Paciente, Edad, Dirección,
    Comuna, Telefono, Estab. Origen, Farmacia, Estab. Destino, Fecha Entrega,
    Periodo Receta, Especialidad, Número Prescripciones, Producto, Cantidad,
    Estab. Prepara, Estab. Transito, Fecha Recepción, Estab. Despacha,
    Estado, Tipo Retiro, Información Retiro, Cantidad Entregada — una fila
    por PRODUCTO, hay que agrupar por N° Receta."""
    wb_in = openpyxl.load_workbook(path, data_only=True)
    ws_in = wb_in.active
    filas_raw = list(ws_in.iter_rows(min_row=1, values_only=True))
    header_idx = next((i for i, row in enumerate(filas_raw) if row and row[0] == "N° Receta"), None)
    if header_idx is None:
        raise ValueError(f"No encontré la fila de encabezado ('N° Receta') en {path} — "
                          f"¿es un Informe Modalidad de Despacho real?")
    header = [str(c).strip() if c else "" for c in filas_raw[header_idx]]

    def _col(nombre):
        return header.index(nombre) if nombre in header else None

    c_receta, c_paciente, c_rut = _col("N° Receta"), _col("Paciente"), _col("Run Paciente")
    c_destino, c_periodo, c_especialidad = _col("Estab. Destino"), _col("Periodo Receta"), _col("Especialidad")
    c_n_presc, c_telefono = _col("Número Prescripciones"), _col("Telefono")
    c_producto, c_cantidad = _col("Producto"), _col("Cantidad")

    por_receta = {}
    refri_por_receta = {}  # receta -> {nombre_insulina: cantidad|None}
    for row in filas_raw[header_idx + 1:]:
        if not row or not row[c_receta]:
            continue
        receta = str(row[c_receta]).strip()
        por_receta.setdefault(receta, row)  # primera fila del grupo alcanza (mismos datos de cabecera por receta)
        # el informe trae una fila por producto — hay que recorrerlas todas
        # (no solo la primera del grupo) para no perder refrigerados.
        if c_producto is not None:
            lab = G._insulina_label(row[c_producto])
            if lab:
                cant = None
                if c_cantidad is not None and row[c_cantidad] not in (None, ""):
                    try: cant = int(row[c_cantidad])
                    except (TypeError, ValueError): cant = None
                refri_por_receta.setdefault(receta, {})[lab] = cant

    nuevas, ya_registradas, con_destino = [], [], []
    for receta, row in por_receta.items():
        if GM.buscar_receta_en_maestro(wb_maestro, receta) is not None:
            ya_registradas.append(receta)
            continue
        destino = str(row[c_destino] or "").strip().upper().rstrip(".")
        if destino:
            con_destino.append(receta)
            continue
        meds = refri_por_receta.get(receta, {})
        refrigerado = "; ".join(sorted(f"{n} x{c}" if c else n for n, c in meds.items()))
        nuevas.append({
            "receta": receta,
            "paciente": str(row[c_paciente] or "").strip(),
            "rut": str(row[c_rut] or "").strip(),
            "destino": destino,  # siempre "" acá (con destino ya se filtró arriba)
            "periodo": str(row[c_periodo] or "").strip(),
            "especialidad": str(row[c_especialidad] or "").strip(),
            "n_presc": str(row[c_n_presc] or "1").strip(),
            "telefono": str(row[c_telefono] or "").strip(),
            "pendiente": "",
            "refrigerado": refrigerado,
        })
    print(f"  {len(por_receta)} receta(s) en el informe · {len(ya_registradas)} ya estaban en el maestro "
          f"(cruce_gt.py sí las tomó) · {len(con_destino)} OMITIDA(S) por ya traer Estab. Destino "
          f"(las toma sola cruce_gt.py --generar, no se duplica nómina) · {len(nuevas)} NUEVA(S) "
          f"sin destino — quedan para completar a mano.")
    if con_destino:
        print(f"  [OMITIDAS] {len(con_destino)} receta(s) ya traen Estab. Destino en el informe "
              f"({', '.join(con_destino)}) — no se registran acá. Las toma sola cruce_gt.py --generar "
              f"(parte de AUTO_SSASUR.py) en la Nómina de Envío automática de ese destino. Si es "
              f"urgente y no puede esperar esa corrida, corre en vez de esto: "
              f"py cruce_gt.py <reporte> --salida out_gt/<rango> --generar")
    return nuevas


def _leer_nomina_existente(ruta):
    """Lee una Nómina de Envío ya generada por esta misma función (hoja
    'Funcionarios', modo normal) y devuelve sus filas en el mismo shape que
    arma _generar_nomina — para poder FUSIONAR en vez de pisar cuando ya
    existe una nómina del mismo destino/fecha.
    Bug real detectado 10-08-2026: el nombre de archivo es determinístico
    por destino+fecha (Nomina_Manual_<destino>_<fecha>.xlsx); una segunda
    corrida el mismo día (ej. procesando un informe distinto) SOBRESCRIBÍA
    la nómina anterior en vez de agregarle filas — se perdieron del disco
    23 recetas de CESFAM QUEPE y 15 de TOLTEN HOSP. de la corrida de esa
    misma mañana (los datos seguían íntegros en gt_maestro.xlsx, que nunca
    se pisa, así que se pudieron reconstruir — pero no debería depender de
    eso)."""
    if not os.path.exists(ruta):
        return []
    try:
        wb = openpyxl.load_workbook(ruta)
    except Exception:
        return []
    if "Funcionarios" not in wb.sheetnames:
        return []
    ws = wb["Funcionarios"]
    HR = 5   # fila de encabezado real de hoja_funcionarios (ver generar.py)
    regs = []
    for row in ws.iter_rows(min_row=HR + 1, values_only=True):
        receta = row[1] if len(row) > 1 else None
        if receta in (None, "", "TOTAL"):
            continue
        regs.append({
            "receta": str(receta), "paciente": row[2] or "", "run": row[3] or "",
            "especialidad": row[4] or "", "periodo": row[5] or "",
            "n_presc": row[7] or 1, "refrigerado": row[8] or "",
            "pendiente": row[10] if len(row) > 10 else "",
        })
    return regs


def _generar_nomina(destino, filas_destino, fecha_hoy):
    """Genera la Nómina de Envío (.xlsx, mismo formato que skill_gt) y,
    cuando alguna receta trae refrigerado, también el Letrero (.pdf) para
    un establecimiento, con las filas manuales de esta corrida. Antes esta
    función solo generaba la planilla — un envío armado a mano con
    refrigerados salía sin el aviso físico ❄ REFRIGERADO en la caja.

    Si ya existe una nómina del mismo destino/fecha (dos corridas el mismo
    día), FUSIONA con la existente en vez de pisarla — ver
    _leer_nomina_existente."""
    regs_nuevos = {f["receta"]: {
        "receta": f["receta"], "paciente": f["paciente"], "run": f["rut"],
        "especialidad": f["especialidad"], "periodo": f["periodo"],
        "n_presc": int(f["n_presc"] or 1), "pendiente": f["pendiente"],
        "refrigerado": f.get("refrigerado", ""),
    } for f in filas_destino}

    # "Nóminas de Envío" se organiza en dos niveles: carpeta por mes
    # ("JULIO 2026") y adentro carpeta por fecha de extracción (DD-MM-YYYY) —
    # mismo esquema que "Revisión de Solicitudes" (ver
    # revision_solicitudes._carpeta_salida). La nómina manual queda ahí, con
    # prefijo "Nomina_Manual_" para distinguirla de las que baja skill_gt.
    carpeta_mes = f"{GM.MESES_ES[fecha_hoy.month - 1]} {fecha_hoy.year}"
    carpeta_fecha = fecha_hoy.strftime("%d-%m-%Y")
    carpeta_local = _CARPETA_LOCAL.get(destino)
    if carpeta_local:
        destino_dir = os.path.join(GT_SOLICITUDES_DIR, carpeta_local, "Nóminas de Envío", carpeta_mes, carpeta_fecha)
    else:
        destino_dir = os.path.join(GT_SOLICITUDES_DIR, "_sin_carpeta_conocida", carpeta_mes, carpeta_fecha)
        print(f"  [AVISO] '{destino}' no está en el mapeo de carpetas — guardando en {destino_dir}")
    os.makedirs(destino_dir, exist_ok=True)

    nombre = f"Nomina_Manual_{GM._norm(destino)}_{fecha_hoy.strftime('%Y-%m-%d')}.xlsx"
    ruta = os.path.join(destino_dir, nombre)

    regs_existentes = {r["receta"]: r for r in _leer_nomina_existente(ruta)}
    n_fusionadas = len(regs_existentes) - len(set(regs_existentes) & set(regs_nuevos))
    if n_fusionadas:
        print(f"  [FUSIÓN] {destino}: ya había una nómina de hoy con {len(regs_existentes)} receta(s) — "
              f"se conservan {n_fusionadas} que no venían en esta corrida.")
    regs = list({**regs_existentes, **regs_nuevos}.values())   # los nuevos pisan si hay choque de receta

    wb = openpyxl.Workbook()
    titulo = f"GESTIÓN TERRITORIAL — {destino}"
    subtitulo = (f"Origen: Farmacia Hospital de Pitrufquén   |   Destino: {destino}   |   "
                 f"Nómina manual (sin destino en SSASUR) — {fecha_hoy.strftime('%d/%m/%Y')}")
    G.hoja_funcionarios(wb, regs, destino, titulo, subtitulo)
    wb.save(ruta)

    # Letrero — mismo criterio que skill_gt: solo si algún registro trae
    # refrigerado. Aquí no hay cruce con histórico SIDRA (refri_map vacío);
    # el detalle sale de la columna "refrigerado" del CSV/informe de entrada.
    ruta_letrero = None
    if any((r.get("refrigerado") or "").strip() for r in regs):
        G.FECHA = fecha_hoy.strftime("%d/%m/%Y")
        detalle_refri = G._detalle_refrigerados(regs, {})
        nombre_letrero = f"Nomina_Manual_{GM._norm(destino)}_{fecha_hoy.strftime('%Y-%m-%d')}_Letrero.xlsx"
        ruta_letrero = os.path.join(destino_dir, nombre_letrero)
        G.letrero(destino, True, ruta_letrero, detalle_refri)
        if G.to_pdf(ruta_letrero, destino_dir):
            try:
                os.remove(ruta_letrero)
            except OSError:
                pass
            ruta_letrero = os.path.splitext(ruta_letrero)[0] + ".pdf"

    return ruta, ruta_letrero


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--csv", help="CSV con las recetas sin destino en el sistema")
    ap.add_argument("--gt-excel", help="reporteGestionTerritorial_*.xlsx crudo (Informe Modalidad de Despacho) — "
                                        "registra solo las recetas que AÚN NO están en el maestro (alternativa a --csv)")
    ap.add_argument("--estado", default="EN PREPARACIÓN",
                     help="Estado a asignar (default: EN PREPARACIÓN)")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    if not a.csv and not a.gt_excel:
        print("[ERROR] Pasa --csv o --gt-excel.")
        return

    wb, path = GM.cargar_maestro()

    if a.csv:
        filas = _leer_csv(a.csv)
        if not filas:
            print(f"[ERROR] {a.csv} no tiene filas válidas (falta la columna 'receta' o está vacía).")
            return
        sin_destino = [f["receta"] for f in filas if not f["destino"]]
        if sin_destino:
            print(f"[ERROR] Estas recetas no tienen 'destino' en el CSV — complétalo antes de correr: {sin_destino}")
            return
        print(f"{'[DRY-RUN] ' if a.dry_run else ''}{len(filas)} receta(s) en {a.csv}")
    else:
        filas = _leer_gt_excel(a.gt_excel, wb)
        if not filas:
            print("Nada que registrar — todas las recetas del informe ya estaban en el maestro.")
            return
        sin_destino = [f["receta"] for f in filas if not f["destino"]]
        if sin_destino:
            print(f"[AVISO] Estas recetas no traen Estab. Destino en el informe — se guardan igual, "
                  f"revisa el destino a mano después: {sin_destino}")

    hoy = datetime.date.today()
    hojas_tocadas = {}
    for f in filas:
        receta_dict = {k: v for k, v in f.items() if k != "receta" and v} | {"receta": f["receta"]}
        if a.dry_run:
            ws_existente = GM.buscar_receta_en_maestro(wb, f["receta"])
            print(f"  {'actualizaría' if ws_existente else 'nueva'} -> {f['receta']} | {f['paciente']} | destino={f['destino']}")
            continue
        ws, resultado, _ = GM.upsert_receta_maestro(wb, receta_dict, estado=a.estado, fecha_fallback=hoy)
        hojas_tocadas[ws.title] = ws
        print(f"  {resultado} -> {ws.title} | {f['receta']} | {f['paciente']}")

    if a.dry_run:
        print("\nNada escrito (dry-run). Nóminas no generadas en dry-run.")
        return

    for ws in hojas_tocadas.values():
        GM.aplicar_formato_maestro(ws)
    GM.formatear_historial(wb)
    GM.guardar(wb, path)
    print(f"\nGuardado: {path}")

    por_destino = {}
    for f in filas:
        por_destino.setdefault(f["destino"], []).append(f)
    print("\nGenerando Nóminas de Envío:")
    for destino, filas_destino in por_destino.items():
        ruta, ruta_letrero = _generar_nomina(destino, filas_destino, hoy)
        if ruta_letrero:
            print(f"  {destino}: {len(filas_destino)} receta(s) -> {ruta}, {ruta_letrero}")
        else:
            print(f"  {destino}: {len(filas_destino)} receta(s) -> {ruta}")


if __name__ == "__main__":
    main()
