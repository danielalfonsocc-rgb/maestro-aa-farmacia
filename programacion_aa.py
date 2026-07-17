# -*- coding: utf-8 -*-
"""
programacion_aa.py — Planilla de Programación del ciclo Bodega AA
===========================================================================
Al reiniciarse el ciclo de pedidos Bodega AA → Bodega Fármacos (mismo ciclo
de 2 semanas / 10 días hábiles que usa pedido_fusion.py), genera una planilla
con:

    Medicamento (orden alfabético) | Cantidad Programada | Cantidad Solicitada
    | Stock Bodega AA | Stock Real (en blanco, a contar) | Consumo Promedio
    Mensual | Sugerencia

  - Cantidad Programada / Cantidad Solicitada: salen del reporte mensual de
    SSASUR "Consumos por centro de costo" (Reportes → Reporte de consumo por
    centro de costo → Centro de Costo = FARMACIA → Generar XLS). Se descarga
    a mano cada mes y se detecta automáticamente el más reciente.
  - Stock Bodega AA: Consolidado_AA_MAESTRO → Pedido_Repos_Bodega → Stock_Bod_Actual.
  - Stock Real: se llena a mano tras el conteo físico y luego se aplica con
    --aplicar-conteo (ver más abajo).
  - Consumo Promedio Mensual: CMP_Mensual_22d, ya calculado/afinado en el
    Consolidado — es el requerimiento real que sustenta la Sugerencia.
  - Sugerencia: compara el Consumo Promedio Mensual contra la Cantidad
    Programada del reporte, e incluye la cantidad sugerida de programación
    (= el propio Consumo Promedio Mensual, redondeado):
      · "Subir programación a <N> ud"  si el real supera lo programado (+15%)
      · "Bajar programación a <N> ud"  si el real es menor a lo programado (-15%)
      · en blanco                       si están dentro del rango
      · "Incorporar a programación: <N> ud" si el medicamento NO aparece en
        el reporte de programación y lleva ≥3 meses consecutivos con demanda
        real (CMP_Mensual_22d > 0) sin estar programado — la racha se
        guarda en _historial_programacion.json.

Sin llamadas a IA — solo pandas + openpyxl.

Uso:
    py programacion_aa.py                        # solo genera si hoy es inicio de ciclo
    py programacion_aa.py --forzar                # genera igual, fuera del inicio de ciclo
    py programacion_aa.py --reporte ruta.xlsx      # fuerza el reporte SSASUR a usar

Segundo paso (tras contar físicamente y escanear la planilla impresa):
    py programacion_aa.py --aplicar-conteo conteo.json
        conteo.json = {"MEDICAMENTO TAL COMO SALE EN LA PLANILLA": 123, ...}
    Genera Resumen_Programacion_AA_<fecha>.xlsx en la carpeta Programacion_AA,
    con Diferencia (Stock Bodega AA − Stock Real) y las mismas Sugerencias.
    Esta es la salida que se sube a Drive y Escritorio.
"""
import os
import re
import sys
import json
import glob
import argparse
import datetime as dt

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, WORK_DIR)
from utils_aa import norm_erp, HOMOLOGACION, setup_stdout  # noqa: E402
from pedido_fusion import _feriados, _dias_ciclo  # noqa: E402

setup_stdout()

OUT_DIR = os.path.join(WORK_DIR, 'Programacion_AA')
HIST_JSON = os.path.join(WORK_DIR, '_historial_programacion.json')
TOL_PCT = 0.15   # ±15% de tolerancia entre requerimiento real y lo programado
RACHA_MIN = 3    # meses consecutivos de demanda sin programación → incorporar

MESES_ES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11,
    'DICIEMBRE': 12,
}

THIN  = Side(style='thin', color='DCDCDC')
BRD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFILL = PatternFill('solid', fgColor='D1FAF5')
HFONT = Font(bold=True, color='065F46', name='Arial', size=10)

SUG_COL = {
    'Subir programación'          : ('FFE0B2', 'B45309'),
    'Bajar programación'          : ('E1F5FE', '01579B'),
    'Incorporar a programación'   : ('F4B3B3', '7F1D1D'),
    ''                             : ('F9FAFB', '374151'),
}


def _color_sugerencia(texto):
    """La Sugerencia incluye la cantidad sugerida (ej. 'Subir programación a 1200 ud'),
    así que el color se resuelve por el prefijo, no por igualdad exacta."""
    texto = texto or ''
    for prefijo, colores in SUG_COL.items():
        if prefijo and texto.startswith(prefijo):
            return colores
    return SUG_COL['']


def _pfill(hx):
    h = str(hx).lstrip('#')
    return PatternFill('solid', fgColor=('FF' + h) if len(h) == 6 else h)


def _key(nombre):
    n = norm_erp(nombre)
    return HOMOLOGACION.get(n, n)


# ─────────────── helpers de archivos ────────────────────────────────────────

def _mas_reciente(patron, extra_dirs=()):
    dirs = [WORK_DIR, *extra_dirs]
    cand = []
    for d in dirs:
        cand += [f for f in glob.glob(os.path.join(d, patron))
                 if not os.path.basename(f).startswith('~$')]
    return max(cand, key=os.path.getmtime) if cand else None


def _downloads_dir():
    perfil = os.environ.get('USERPROFILE', os.path.expanduser('~'))
    d = os.path.join(perfil, 'Downloads')
    return d if os.path.isdir(d) else WORK_DIR


def _universo_bodega():
    mae = _mas_reciente('Consolidado_AA_MAESTRO*.xlsx')
    if not mae:
        print('[ERROR] No se encontró Consolidado_AA_MAESTRO*.xlsx — corre maestro_aa.py primero.')
        sys.exit(1)
    df = pd.read_excel(mae, sheet_name='Pedido_Repos_Bodega', engine='openpyxl')
    df = df[['Medicamento', 'Stock_Bod_Actual', 'CMP_Mensual_22d']].copy()
    df['Medicamento'] = df['Medicamento'].astype(str).str.strip()
    df = df[df['Medicamento'] != ''].drop_duplicates(subset=['Medicamento'])
    return df.sort_values('Medicamento').reset_index(drop=True), mae


def _leer_reporte_ssasur(ruta):
    """Lee el reporte 'Consumos por centro de costo' de SSASUR (Generar XLS).
    Fila 0 = título, fila 1 = metadata con el mes/año, fila 2 = encabezados."""
    meta = pd.read_excel(ruta, header=None, nrows=2, engine='openpyxl')
    texto_meta = str(meta.iloc[1, 0]) if meta.shape[0] > 1 else ''
    m = re.search(r'mes de (\w+) de (\d{4})', texto_meta, re.IGNORECASE)
    periodo = None
    if m:
        mes_num = MESES_ES.get(m.group(1).strip().upper())
        if mes_num:
            periodo = f'{m.group(2)}-{mes_num:02d}'

    df = pd.read_excel(ruta, header=2, engine='openpyxl')
    df = df.rename(columns=lambda c: str(c).strip())
    if 'Centro Costo' in df.columns:
        df = df[df['Centro Costo'].astype(str).str.strip().str.upper() == 'FARMACIA']
    df['_key'] = df['Producto'].astype(str).map(_key)
    prog = dict(zip(df['_key'], pd.to_numeric(df['Total de Productos Programados'], errors='coerce')))
    sol  = dict(zip(df['_key'], pd.to_numeric(df['Productos Solicitado'], errors='coerce')))
    return prog, sol, periodo, texto_meta.strip()


def _cargar_historial():
    if os.path.isfile(HIST_JSON):
        try:
            with open(HIST_JSON, encoding='utf-8') as fh:
                return json.load(fh)
        except (OSError, json.JSONDecodeError):
            pass
    return {}


def _guardar_historial(hist):
    try:
        with open(HIST_JSON, 'w', encoding='utf-8') as fh:
            json.dump(hist, fh, ensure_ascii=False, indent=2)
    except OSError as e:
        print(f'  [aviso] no se pudo guardar {os.path.basename(HIST_JSON)}: {e}')


def _sugerencia(key, requerimiento_real, programado, periodo, hist):
    """Devuelve el texto de Sugerencia, con la cantidad sugerida de programación
    (= el Consumo Promedio Mensual ya calculado/afinado, redondeado a unidades)."""
    cant = round(requerimiento_real) if requerimiento_real else 0
    if programado is None or (isinstance(programado, float) and pd.isna(programado)):
        # No está en el reporte de programación: rastrea la racha de demanda.
        entry = hist.get(key, {'racha': 0, 'ultimo_mes': None})
        if requerimiento_real and requerimiento_real > 0:
            if periodo and entry.get('ultimo_mes') != periodo:
                entry['racha'] = entry.get('racha', 0) + 1
                entry['ultimo_mes'] = periodo
            hist[key] = entry
            if entry['racha'] >= RACHA_MIN:
                return f'Incorporar a programación: {cant} ud'
            return ''
        else:
            hist.pop(key, None)
            return ''
    else:
        # Está programado: se resetea cualquier racha acumulada de antes.
        hist.pop(key, None)
        if requerimiento_real is None or pd.isna(requerimiento_real):
            return ''
        if programado <= 0:
            return f'Subir programación a {cant} ud' if requerimiento_real > 0 else ''
        ratio = requerimiento_real / programado
        if ratio > 1 + TOL_PCT:
            return f'Subir programación a {cant} ud'
        if ratio < 1 - TOL_PCT:
            return f'Bajar programación a {cant} ud'
        return ''


# ─────────────── modo 1: generar planilla del ciclo ────────────────────────

HDRS = [
    ('Medicamento',                46),
    ('Cantidad Programada',        17),
    ('Cantidad Solicitada',        17),
    ('Stock Bodega AA',            15),
    ('Stock Real',                 13),
    ('Consumo Promedio Mensual',   20),
    ('Sugerencia',                 30),
]


def _es_inicio_ciclo(hoy, fer):
    """True si hoy es el primer día hábil del ciclo Bod_Farmacos (10d hábiles)."""
    return _dias_ciclo(hoy, fer) == 10


def generar(ruta_reporte=None, forzar=False):
    hoy = dt.date.today()
    fer = _feriados()
    if not forzar and not _es_inicio_ciclo(hoy, fer):
        print(f'Hoy ({hoy.strftime("%d/%m/%Y")}) no es el inicio del ciclo Bodega AA → Bodega Fármacos.')
        print('No se genera planilla. Usa --forzar para generarla igual.')
        return

    universo, mae = _universo_bodega()
    ruta_reporte = ruta_reporte or _mas_reciente(
        'cantidad_de_productos_consumidos_en_centro_de_costo_farmacia*.xlsx',
        extra_dirs=[_downloads_dir()])
    if not ruta_reporte:
        print('[ERROR] No se encontró el reporte de SSASUR '
              '"cantidad_de_productos_consumidos_en_centro_de_costo_farmacia*.xlsx" '
              'en la carpeta del proyecto ni en Descargas.')
        print('  Descárgalo desde SSASUR → Reportes → Consumo por centro de costo → '
              'Centro de Costo = FARMACIA → Generar XLS.')
        sys.exit(1)

    prog, sol, periodo, meta_txt = _leer_reporte_ssasur(ruta_reporte)
    hist = _cargar_historial()

    filas = []
    n_sin_reporte = 0
    for r in universo.itertuples(index=False):
        key = _key(r.Medicamento)
        programado = prog.get(key)
        solicitado = sol.get(key)
        req_real = r.CMP_Mensual_22d if not pd.isna(r.CMP_Mensual_22d) else 0
        if key not in prog:
            n_sin_reporte += 1
        sugerencia = _sugerencia(key, req_real, programado, periodo, hist)
        filas.append({
            'Medicamento': r.Medicamento,
            'Cantidad Programada': None if programado is None or pd.isna(programado) else int(programado),
            'Cantidad Solicitada': None if solicitado is None or pd.isna(solicitado) else int(solicitado),
            'Stock Bodega AA': int(r.Stock_Bod_Actual) if not pd.isna(r.Stock_Bod_Actual) else 0,
            'Stock Real': None,
            'Consumo Promedio Mensual': round(req_real),
            'Sugerencia': sugerencia,
        })

    _guardar_historial(hist)
    os.makedirs(OUT_DIR, exist_ok=True)
    sal = os.path.join(OUT_DIR, f'Programacion_AA_{hoy.strftime("%Y%m%d")}.xlsx')
    _escribir_planilla(sal, filas, mae, ruta_reporte, meta_txt, hoy, forzar)

    n_subir = sum(1 for f in filas if f['Sugerencia'].startswith('Subir programación'))
    n_bajar = sum(1 for f in filas if f['Sugerencia'].startswith('Bajar programación'))
    n_incorp = sum(1 for f in filas if f['Sugerencia'].startswith('Incorporar a programación'))
    print(f'HOY = {hoy}')
    print(f'Sistema  : {os.path.basename(mae)}')
    print(f'Reporte  : {os.path.basename(ruta_reporte)}  ({meta_txt or "sin metadata de mes"})')
    print(f'{len(filas)} medicamentos | {n_sin_reporte} sin programación en el reporte')
    print(f'Sugerencias: Subir {n_subir} | Bajar {n_bajar} | Incorporar a programación {n_incorp}')
    print(f'\nExcel: {os.path.basename(sal)}  (carpeta {os.path.basename(OUT_DIR)}\\)')
    print('\nImprime esta planilla, cuenta físicamente Bodega AA y llena "Stock Real" a mano.')
    print('Cuando esté escaneada, avisa para transcribirla y correr --aplicar-conteo.')


def _escribir_planilla(sal, filas, mae, ruta_reporte, meta_txt, hoy, forzar,
                         es_resumen=False, n_diferencias=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumen' if es_resumen else 'Programacion'

    hdrs = list(HDRS)
    if es_resumen:
        hdrs = hdrs + [('Diferencia (Bod.AA - Real)', 20)]
    ncols = len(hdrs)

    titulo = ('RESUMEN CONTEO vs PROGRAMACIÓN — Bodega AA' if es_resumen
              else 'PLANILLA DE PROGRAMACIÓN — Bodega AA')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, f'{titulo}  ·  {hoy.strftime("%d/%m/%Y")}')
    ws.cell(1, 1).font = Font(bold=True, size=12, color='065F46', name='Arial')
    ws.row_dimensions[1].height = 22

    sub = (f'Sistema: {os.path.basename(mae)}  ·  Reporte SSASUR: {os.path.basename(ruta_reporte)} '
           f'({meta_txt or "sin metadata de mes"})  ·  Tolerancia sugerencia: ±{int(TOL_PCT*100)}%  ·  '
           f'"Incorporar a programación" = ≥{RACHA_MIN} meses seguidos con demanda sin estar programado')
    if forzar and not es_resumen:
        sub += '  ·  generada fuera del inicio de ciclo (--forzar)'
    if es_resumen and n_diferencias is not None:
        sub += f'  ·  {n_diferencias} medicamento(s) con diferencia entre Stock Bodega AA y Stock Real'
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, sub)
    ws.cell(2, 1).font = Font(italic=True, size=9, color='555555', name='Arial')
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 30

    for j, (label, w) in enumerate(hdrs, 1):
        c = ws.cell(3, j, label)
        c.fill = HFILL; c.font = HFONT; c.border = BRD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[3].height = 26

    for i, f in enumerate(filas, 4):
        vals = [f['Medicamento'], f['Cantidad Programada'], f['Cantidad Solicitada'],
                f['Stock Bodega AA'], f['Stock Real'], f['Consumo Promedio Mensual'],
                f['Sugerencia']]
        if es_resumen:
            vals.append(f.get('Diferencia'))
        bg, fg = _color_sugerencia(f['Sugerencia'])
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            c.border = BRD
            if j >= 2:
                c.fill = _pfill(bg)
                c.font = Font(name='Arial', size=10, color=fg)
                c.alignment = Alignment(horizontal='center')
            else:
                c.font = Font(name='Arial', size=10)
        if es_resumen and f.get('Diferencia') not in (None, 0):
            dc = ws.cell(i, len(hdrs))
            dc.font = Font(name='Arial', size=10, color='7F1D1D', bold=True)
            dc.fill = _pfill('F4B3B3')

    last = 3 + len(filas)
    ws.freeze_panes = 'A4'
    if filas:
        ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}{last}'
        ws.print_area = f'A1:{get_column_letter(ncols)}{last}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    wb.save(sal)


# ─────────────── modo 2: aplicar conteo (tras escanear) ────────────────────

def aplicar_conteo(ruta_json):
    plantilla = _mas_reciente('Programacion_AA_*.xlsx', extra_dirs=[OUT_DIR])
    if not plantilla:
        print('[ERROR] No hay ninguna Programacion_AA_*.xlsx generada todavía. '
              'Corre primero: py programacion_aa.py --forzar')
        sys.exit(1)
    try:
        with open(ruta_json, encoding='utf-8') as fh:
            valores = json.load(fh)
    except (OSError, json.JSONDecodeError) as e:
        print(f'[ERROR] No se pudo leer {ruta_json}: {e}')
        sys.exit(1)

    valores_key = {_key(k): v for k, v in valores.items()}

    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active
    hoy = dt.date.today()
    filas = []
    n_diff = 0
    for row in ws.iter_rows(min_row=4, max_col=7):
        med_c, prog_c, sol_c, sbod_c, sreal_c, cpm_c, sug_c = row
        if med_c.value is None:
            continue
        key = _key(med_c.value)
        stock_real = valores_key.get(key)
        sbod = sbod_c.value or 0
        diff = None if stock_real is None else sbod - stock_real
        if diff:
            n_diff += 1
        filas.append({
            'Medicamento': med_c.value,
            'Cantidad Programada': prog_c.value,
            'Cantidad Solicitada': sol_c.value,
            'Stock Bodega AA': sbod,
            'Stock Real': stock_real,
            'Consumo Promedio Mensual': cpm_c.value,
            'Sugerencia': sug_c.value or '',
            'Diferencia': diff,
        })

    mae = _mas_reciente('Consolidado_AA_MAESTRO*.xlsx')
    os.makedirs(OUT_DIR, exist_ok=True)
    sal = os.path.join(OUT_DIR, f'Resumen_Programacion_AA_{hoy.strftime("%Y%m%d_%H%M")}.xlsx')
    _escribir_planilla(sal, filas, mae or plantilla, plantilla, '', hoy, False,
                        es_resumen=True, n_diferencias=n_diff)

    print(f'Planilla base : {os.path.basename(plantilla)}')
    print(f'Conteo aplicado: {len(valores_key)} medicamento(s) en {os.path.basename(ruta_json)}')
    print(f'Diferencias detectadas: {n_diff} de {len(filas)} medicamentos')
    print(f'\nExcel: {os.path.basename(sal)}  (carpeta {os.path.basename(OUT_DIR)}\\)')
    print('Sube esta carpeta a Drive y Escritorio con publicar_drive.py / publicar_escritorio.py.')


# ─────────────── main ────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description='Planilla de Programación del ciclo Bodega AA. Sin IA.')
    ap.add_argument('--forzar', action='store_true',
                     help='Genera la planilla aunque hoy no sea el inicio del ciclo')
    ap.add_argument('--reporte', default=None,
                     help='Ruta a un reporte SSASUR específico (si no, usa el más reciente)')
    ap.add_argument('--aplicar-conteo', default=None, metavar='JSON',
                     help='Aplica el conteo físico (JSON medicamento→cantidad) y genera el Resumen final')
    args = ap.parse_args()

    if args.aplicar_conteo:
        aplicar_conteo(args.aplicar_conteo)
    else:
        generar(args.reporte, args.forzar)


if __name__ == '__main__':
    main()
