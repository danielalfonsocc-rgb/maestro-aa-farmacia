# -*- coding: utf-8 -*-
"""
actualizar_programacion_sept.py
Actualiza Cantidad Programada / Cantidad Solicitada / Sugerencia en una
planilla Programacion_AA existente usando el reporte SSASUR de septiembre.

Uso:
    py actualizar_programacion_sept.py [planilla.xlsx] [reporte_ssasur.xlsx]

Si no se pasan argumentos, usa la planilla más reciente en Programacion_AA\
y el reporte más reciente disponible.
"""
import os, sys, glob, re, json, datetime as dt
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, WORK_DIR)
from utils_aa import norm_erp, HOMOLOGACION  # noqa

MESES_ES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11,
    'DICIEMBRE': 12,
}
TOL_PCT   = 0.15
RACHA_MIN = 3
HIST_JSON = os.path.join(WORK_DIR, '_historial_programacion.json')

SUG_COL = {
    'Subir programación'         : ('FFE0B2', 'B45309'),
    'Bajar programación'         : ('E1F5FE', '01579B'),
    'Incorporar a programación'  : ('F4B3B3', '7F1D1D'),
    ''                           : ('F9FAFB', '374151'),
}


def _key(nombre):
    n = norm_erp(str(nombre))
    return HOMOLOGACION.get(n, n)


def _pfill(hx):
    h = str(hx).lstrip('#')
    return PatternFill('solid', fgColor=('FF' + h) if len(h) == 6 else h)


def _color_sugerencia(texto):
    texto = texto or ''
    for prefijo, colores in SUG_COL.items():
        if prefijo and texto.startswith(prefijo):
            return colores
    return SUG_COL['']


def _sugerencia(key, req_real, programado, periodo, hist):
    cant = round(req_real) if req_real else 0
    if programado is None or (isinstance(programado, float) and pd.isna(programado)):
        entry = hist.get(key, {'racha': 0, 'ultimo_mes': None})
        if req_real and req_real > 0:
            if periodo and entry.get('ultimo_mes') != periodo:
                entry['racha'] = entry.get('racha', 0) + 1
                entry['ultimo_mes'] = periodo
            hist[key] = entry
            if entry['racha'] >= RACHA_MIN:
                return f'Incorporar a programación: {cant} ud'
            return ''
        else:
            hist.pop(key, None)
            return ''
    else:
        hist.pop(key, None)
        if req_real is None or pd.isna(req_real):
            return ''
        if programado <= 0:
            return f'Subir programación a {cant} ud' if req_real > 0 else ''
        ratio = req_real / programado
        if ratio > 1 + TOL_PCT:
            return f'Subir programación a {cant} ud'
        if ratio < 1 - TOL_PCT:
            return f'Bajar programación a {cant} ud'
        return ''


def _leer_reporte(ruta):
    meta = pd.read_excel(ruta, header=None, nrows=2, engine='openpyxl')
    texto_meta = str(meta.iloc[1, 0]) if meta.shape[0] > 1 else ''
    m = re.search(r'mes de (\w+) de (\d{4})', texto_meta, re.IGNORECASE)
    periodo = None
    if m:
        mes_num = MESES_ES.get(m.group(1).strip().upper())
        if mes_num:
            periodo = f'{m.group(2)}-{mes_num:02d}'

    df = pd.read_excel(ruta, header=2, engine='openpyxl')
    df = df.rename(columns=lambda c: str(c).strip())
    if 'Centro Costo' in df.columns:
        df = df[df['Centro Costo'].astype(str).str.strip().str.upper() == 'FARMACIA']
    df['_key'] = df['Producto'].astype(str).map(_key)
    prog = dict(zip(df['_key'], pd.to_numeric(df['Total de Productos Programados'], errors='coerce')))
    sol  = dict(zip(df['_key'], pd.to_numeric(df['Productos Solicitado'], errors='coerce')))
    return prog, sol, periodo, texto_meta.strip()


def _mas_reciente(patron, dirs):
    cands = []
    for d in dirs:
        cands += [f for f in glob.glob(os.path.join(d, patron))
                  if not os.path.basename(f).startswith('~$')]
    return max(cands, key=os.path.getmtime) if cands else None


def main():
    # ── 1. Resolver rutas ────────────────────────────────────────────────────
    buscar_dirs = [
        WORK_DIR,
        os.path.join(os.environ.get('USERPROFILE', ''), 'Downloads'),
    ]

    if len(sys.argv) >= 3:
        ruta_planilla = sys.argv[1]
        ruta_reporte  = sys.argv[2]
    elif len(sys.argv) == 2:
        ruta_planilla = sys.argv[1]
        ruta_reporte  = _mas_reciente(
            'cantidad_de_productos_consumidos_en_centro_de_costo_farmacia*.xlsx',
            buscar_dirs)
    else:
        # Planilla más reciente en Programacion_AA\
        ruta_planilla = _mas_reciente(
            os.path.join(WORK_DIR, 'Programacion_AA', 'Programacion_AA_*.xlsx'),
            [''])
        if not ruta_planilla:
            ruta_planilla = _mas_reciente(
                'Programacion_AA_*.xlsx',
                [os.path.join(WORK_DIR, 'Programacion_AA')])
        ruta_reporte = _mas_reciente(
            'cantidad_de_productos_consumidos_en_centro_de_costo_farmacia*.xlsx',
            buscar_dirs)

    if not ruta_planilla or not os.path.isfile(ruta_planilla):
        print('[ERROR] No se encontró planilla Programacion_AA_*.xlsx')
        sys.exit(1)
    if not ruta_reporte or not os.path.isfile(ruta_reporte):
        print('[ERROR] No se encontró reporte SSASUR '
              'cantidad_de_productos_consumidos_en_centro_de_costo_farmacia*.xlsx')
        sys.exit(1)

    print(f'Planilla : {os.path.basename(ruta_planilla)}')
    print(f'Reporte  : {os.path.basename(ruta_reporte)}')

    # ── 2. Leer reporte SSASUR ───────────────────────────────────────────────
    prog, sol, periodo, meta_txt = _leer_reporte(ruta_reporte)
    print(f'Período  : {periodo or "no detectado"}  ({meta_txt or "sin metadata"})')

    # ── 3. Cargar historial de rachas ────────────────────────────────────────
    hist = {}
    if os.path.isfile(HIST_JSON):
        try:
            with open(HIST_JSON, encoding='utf-8') as fh:
                hist = json.load(fh)
        except (OSError, json.JSONDecodeError):
            pass

    # ── 4. Abrir planilla con openpyxl para actualizar en lugar ─────────────
    wb = openpyxl.load_workbook(ruta_planilla)
    ws = wb.active

    # Detectar columnas por cabecera (fila 3)
    col_map = {}
    for cell in ws[3]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    COL_PROG = col_map.get('Cantidad Programada')
    COL_SOL  = col_map.get('Cantidad Solicitada')
    COL_MED  = col_map.get('Medicamento')
    COL_CMP  = col_map.get('Consumo Promedio Mensual')
    COL_SUG  = col_map.get('Sugerencia')

    if not all([COL_PROG, COL_SOL, COL_MED]):
        print(f'[ERROR] No se encontraron columnas esperadas. '
              f'Cabeceras detectadas: {list(col_map.keys())}')
        sys.exit(1)

    # ── 5. Actualizar fila a fila ────────────────────────────────────────────
    n_act = 0
    n_sin = 0
    for row in ws.iter_rows(min_row=4):
        med_cell = row[COL_MED - 1]
        if med_cell.value is None:
            continue
        med_nombre = str(med_cell.value).strip()
        if not med_nombre:
            continue

        key = _key(med_nombre)
        programado = prog.get(key)
        solicitado = sol.get(key)

        prog_val = None if programado is None or pd.isna(programado) else int(programado)
        sol_val  = None if solicitado is None  or pd.isna(solicitado)  else int(solicitado)

        ws.cell(row[0].row, COL_PROG).value = prog_val
        ws.cell(row[0].row, COL_SOL).value  = sol_val

        if key not in prog:
            n_sin += 1

        # Recalcular sugerencia si la columna existe
        if COL_SUG and COL_CMP:
            cmp_val = ws.cell(row[0].row, COL_CMP).value
            req_real = float(cmp_val) if cmp_val is not None else 0
            sug = _sugerencia(key, req_real, programado, periodo, hist)
            sug_cell = ws.cell(row[0].row, COL_SUG)
            sug_cell.value = sug
            bg, fg = _color_sugerencia(sug)
            # Actualizar color solo en col programada, solicitada y sugerencia
            for ci in [COL_PROG, COL_SOL, COL_SUG]:
                c = ws.cell(row[0].row, ci)
                c.fill = _pfill(bg)
                c.font = Font(
                    name='Arial', size=10, color=fg,
                    bold=(c.font.bold if c.font else False)
                )

        n_act += 1

    # ── 6. Actualizar subtítulo (fila 2) con info del reporte nuevo ──────────
    sub_cell = ws.cell(2, 1)
    if sub_cell.value:
        # Reemplaza la mención del reporte anterior
        nuevo_sub = re.sub(
            r'Reporte SSASUR: [\w\.\-]+\.xlsx[^·]*',
            f'Reporte SSASUR: {os.path.basename(ruta_reporte)} ({meta_txt or "sin metadata de mes"})  ·  ',
            str(sub_cell.value)
        )
        sub_cell.value = nuevo_sub

    # ── 7. Guardar historial y planilla ──────────────────────────────────────
    try:
        with open(HIST_JSON, 'w', encoding='utf-8') as fh:
            json.dump(hist, fh, ensure_ascii=False, indent=2)
    except OSError as e:
        print(f'  [aviso] no se pudo guardar historial: {e}')

    wb.save(ruta_planilla)

    print(f'\n{n_act} filas actualizadas | {n_sin} sin datos en el reporte')
    print(f'Guardado: {ruta_planilla}')


if __name__ == '__main__':
    main()
