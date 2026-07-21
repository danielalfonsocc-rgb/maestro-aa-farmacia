#!/usr/bin/env python3
"""
gt_maestro.py — Mantiene la planilla maestra de Gestión Territorial (histórico
por mes) sin perder datos de meses anteriores y con el mismo formato visual
que las planillas por destino de skill_gt/scripts/generar.py.

Reglas:
  - 1 hoja por mes calendario ("ENERO 2026", "FEBRERO 2026", ...). Si el mes en
    curso no tiene hoja, se crea (encabezados canónicos), pero NUNCA se borra ni
    sobreescribe una hoja de un mes anterior.
  - upsert_receta() busca por Nº de receta dentro de la hoja del mes indicado;
    si existe, actualiza; si no, agrega una fila nueva.
  - El Estado no retrocede solo: si una receta ya está "Enviada" y llega un
    upsert con estado "En revisión", se ignora el cambio de estado (salvo
    forzar_estado=True).
  - aplicar_formato_maestro() es idempotente: se puede llamar tantas veces como
    se quiera sobre la misma hoja sin duplicar la banda de título.

Uso como script:
  py gt_maestro.py --registrar --receta 12345678 --paciente "NOMBRE APELLIDO PATERNO MATERNO" \
      --rut 11111111-1 --destino "Hospital Tolten" --periodo 3/6 --especialidad "MEDICINA INTERNA" \
      --estado "En revisión"

  py gt_maestro.py --estado --receta 12345678 --nuevo-estado "En preparación"

Uso como módulo:
  from gt_maestro import cargar_maestro, obtener_hoja_mes, upsert_receta, aplicar_formato_maestro, guardar
"""
import argparse
import colorsys
import datetime
import glob
import os
import re
import shutil
import sys
import unicodedata
import zlib
from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "skill_gt", "scripts"))
import generar as G  # noqa: E402  (reusa paleta NAVY/BLUE/GREY/... y place_logo + logos)
import cruce_gt as CG  # noqa: E402  (reusa clasificación refrigerado/controlado vs. histórico)

from utils_aa import GT_MAESTRO_XLSX, setup_stdout  # noqa: E402

setup_stdout()

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))

MESES_ES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
            "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

HEADERS = ["Nº de receta", "Paciente", "Rut", "Establecimiento de destino", "Teléfono",
           "Fecha de solicitud", "Periodo Receta",
           "Especialidad", "Número Prescripciones", "Estado",
           "Fecha de retiro en Farmacia", "Refrigerado", "Controlado", "PENDIENTE"]

COLUMN_MAP = {
    "receta": {"ndereceta", "nreceta", "nroreceta"},
    "paciente": {"paciente"},
    "rut": {"rut", "runpaciente", "run"},
    "destino": {"establecimientodedestino", "estabdestino"},
    "telefono": {"telefono", "telefono"},
    "fecha_solicitud": {"fechadesolicitud"},
    "periodo": {"periodoreceta"},
    "especialidad": {"especialidad"},
    "n_presc": {"numeroprescripciones"},
    "estado": {"estado"},
    "fecha_retiro": {"fechaderetiroenfarmacia"},
    "refrigerado": {"refrigerado"},
    "controlado": {"controlado"},
    "pendiente": {"pendiente"},
}

# Orden de avance del estado: un upsert con estado "anterior" en esta lista NO
# sobreescribe un estado ya más avanzado (salvo forzar_estado=True).
ORDEN_ESTADOS = ["enrevision", "enpreparacion", "listapararetiro",
                  "retiroenventanilla", "enviada", "entregada"]

# "Fecha de retiro en Farmacia" es manual a propósito: depende de que alguien
# en la farmacia confirme el hecho físico (que ya la retiraron), algo que el
# script no puede saber ni inventar con la fecha de hoy. Las columnas de fecha
# por estado (En Revisión/En Preparación/Lista para Retiro) y "Fecha de
# entrega último retiro" se eliminaron el 19-07-2026 (usuario): la primera
# quedaba casi siempre vacía y la segunda tomaba la fecha del sistema, que no
# aplica cuando el envío es por Gestión Territorial.

# Paleta de colores por establecimiento, mismo lenguaje visual "tenue" que
# pedido_fusion.py (CPAL): fondo casi blanco + texto oscuro saturado del
# mismo matiz, en vez de pasteles Material Design más vivos. 12 tonos
# repartidos uniformemente en el círculo cromático (cada 30°); cada uno se
# asigna por hash estable (crc32) del nombre normalizado del establecimiento
# — no por orden de aparición — así el mismo establecimiento saca siempre el
# mismo color en cualquier hoja/archivo (Excel real o la prueba de Sheets),
# sin necesidad de compartir estado entre corridas.
_N_TONOS_ESTABLECIMIENTO = 12


def _hsl_a_hex(h, s, l):
    r, g, b = colorsys.hls_to_rgb(h / 360, l / 100, s / 100)
    return f"{round(r * 255):02X}{round(g * 255):02X}{round(b * 255):02X}"


PALETA_ESTABLECIMIENTOS = [_hsl_a_hex(i * 360 / _N_TONOS_ESTABLECIMIENTO, 70, 97)
                           for i in range(_N_TONOS_ESTABLECIMIENTO)]
PALETA_ESTABLECIMIENTOS_TEXTO = [_hsl_a_hex(i * 360 / _N_TONOS_ESTABLECIMIENTO, 75, 33)
                                 for i in range(_N_TONOS_ESTABLECIMIENTO)]


def _indice_establecimiento(nombre):
    clave = _norm(nombre)
    if not clave:
        return None
    return zlib.crc32(clave.encode()) % _N_TONOS_ESTABLECIMIENTO


def color_establecimiento(nombre):
    """Hex de fondo (sin '#', casi blanco) asignado a ese establecimiento, estable entre corridas."""
    i = _indice_establecimiento(nombre)
    return PALETA_ESTABLECIMIENTOS[i] if i is not None else None


def texto_establecimiento(nombre):
    """Hex de texto (oscuro, mismo matiz que color_establecimiento) para ese establecimiento."""
    i = _indice_establecimiento(nombre)
    return PALETA_ESTABLECIMIENTOS_TEXTO[i] if i is not None else None


THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _norm(s):
    """Minúsculas, sin tildes/ñ, sin espacios ni puntuación — así 'Estab. Destino'
    y 'Establecimiento de destino' se comparan igual de forma robusta. 'º' (U+00BA)
    y '°' (U+00B0) se ven casi idénticos en Excel pero NFKD los trata distinto
    (uno se descompone a 'o', el otro se descarta) — se sacan ambos a mano antes
    de normalizar para que "Nº de receta" y "N° de receta" calcen igual."""
    s = str(s or "").replace("º", "").replace("°", "")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", s)


def _mes_sheet_name(fecha):
    return f"{MESES_ES[fecha.month - 1]} {fecha.year}"


def cargar_maestro(path=None):
    path = path or GT_MAESTRO_XLSX
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe la planilla maestra GT: {path}")
    return load_workbook(path), path


def _backup(path):
    bak = path + ".bak"
    shutil.copy2(path, bak)
    return bak


def guardar(wb, path, backup=True):
    if backup and os.path.exists(path):
        _backup(path)
    wb.save(path)


def _tiene_titulo(ws):
    v = ws["A1"].value
    return isinstance(v, str) and v.startswith("GESTIÓN TERRITORIAL")


def _headers_de_hoja(ws):
    fila = 2 if _tiene_titulo(ws) else 1
    headers = []
    for cell in ws[fila]:
        if cell.value is None:
            break
        headers.append(cell.value)
    return fila, headers


def asegurar_columnas(ws):
    """Agrega al final las columnas de HEADERS que falten en la hoja (migración
    segura: solo agrega, nunca reordena ni borra columnas existentes). Necesario
    para hojas creadas antes de sumar un campo nuevo a HEADERS. Devuelve la
    lista de encabezados agregados (vacía si la hoja ya tenía todo)."""
    fila_header, headers_actuales = _headers_de_hoja(ws)
    existentes_norm = {_norm(h) for h in headers_actuales if h}
    agregadas = []
    col = len(headers_actuales)
    for h in HEADERS:
        if _norm(h) not in existentes_norm:
            col += 1
            ws.cell(row=fila_header, column=col, value=h)
            agregadas.append(h)
    return agregadas


def obtener_hoja_mes(wb, fecha=None):
    """Devuelve (ws, es_nueva). Crea la hoja del mes si no existe — NUNCA toca meses previos."""
    fecha = fecha or datetime.date.today()
    nombre = _mes_sheet_name(fecha)
    if nombre in wb.sheetnames:
        ws = wb[nombre]
        asegurar_columnas(ws)
        return ws, False
    ws = wb.create_sheet(nombre)
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    return ws, True


def _col_index(headers, key):
    candidatos = COLUMN_MAP[key]
    for i, h in enumerate(headers):
        if _norm(h) in candidatos:
            return i
    return None


def buscar_receta_en_maestro(wb, receta_num):
    """Busca un Nº de receta en TODAS las hojas del libro. Devuelve la worksheet
    donde ya existe, o None. Necesario porque el mes de una receta se decide por
    Fecha Entrega, y SSASUR puede corregir esa fecha entre sincronizaciones —sin
    esto, la misma receta termina duplicada en dos hojas de mes distintas."""
    receta_num = str(receta_num).strip()
    for ws in wb.worksheets:
        try:
            fila_header, headers = _headers_de_hoja(ws)
        except Exception:
            continue
        idx_receta = _col_index(headers, "receta")
        if idx_receta is None:
            continue
        for r in range(fila_header + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=idx_receta + 1).value
            if v is not None and str(v).strip() == receta_num:
                return ws
    return None


def upsert_receta_maestro(wb, receta, estado=None, forzar_estado=False, fecha_fallback=None):
    """Punto de entrada recomendado (en vez de upsert_receta + obtener_hoja_mes
    a mano): si la receta ya existe en CUALQUIER hoja se actualiza ahí — el mes
    donde se vio primero manda, nunca se mueve ni se duplica entre hojas. Solo
    si es genuinamente nueva se crea en la hoja del mes de fecha_fallback (o
    la fecha de hoy). Devuelve (ws, resultado, hoja_nueva_creada)."""
    receta_num = str(receta.get("receta") or "").strip()
    if not receta_num:
        raise ValueError("receta['receta'] (Nº de receta) es obligatorio.")

    ws_existente = buscar_receta_en_maestro(wb, receta_num)
    if ws_existente is not None:
        resultado = upsert_receta(ws_existente, receta, estado=estado, forzar_estado=forzar_estado)
        return ws_existente, resultado, False

    fecha = fecha_fallback or datetime.date.today()
    ws, es_nueva_hoja = obtener_hoja_mes(wb, fecha)
    resultado = upsert_receta(ws, receta, estado=estado, forzar_estado=forzar_estado)
    return ws, resultado, es_nueva_hoja


def _set_medicamentos(receta):
    meds = receta.get("medicamentos") or []
    return frozenset(_norm(m) for m in meds if m)


def detectar_alertas_mismo_rut(recetas):
    """recetas: lista de dicts con al menos 'rut', 'receta', 'periodo', 'especialidad',
    'paciente' y opcionalmente 'medicamentos' (lista/set de nombres de producto —
    lo agrega leer_reporte_gt_crudo() a partir de la columna Producto del reporte
    crudo) y 'fecha_digitacion' (string tal como aparece en SSASUR — columna
    "Fecha Ingreso" en Consultar Receta o "Fecha Digitación" en el PDF impreso;
    solo se consigue revisando SSASUR a mano, no viene en el reporte GT crudo).
    Dos recetas ACTIVAS de un mismo RUT son normales (tratamientos distintos)
    salvo tres casos sospechosos:

      1. 'ambiguo_mismo_periodo_especialidad': mismo período Y misma especialidad
         — puede ser un error de tipeo o una duplicación real de la solicitud.
      2. 'especialidad_distinta_mismos_medicamentos': la especialidad cambió (ej.
         Medicina Interna -> Medicina General) pero los medicamentos son idénticos
         — probable error de clasificación, no un tratamiento nuevo.
      3. 'duplicado_informatico': mismo período Y misma Fecha Digitación pero Nº
         de receta DISTINTO — SSASUR a veces digita dos veces la misma cuota por
         un problema del sistema (caso real verificado 18-07-2026: cuotas 1/6
         y 2/6 de una serie completa duplicadas con Nº de receta distintos,
         ambas con Fecha Ingreso 18/06/2026). Solo se detecta si 'fecha_digitacion' viene
         informada en ambos registros — si no se tiene ese dato, esta regla no
         dispara (no es un falso negativo, es información que falta).

    Los tres casos vienen con bloquear_impresion=True: no se imprime/prepara esa
    receta hasta que un QF lo confirme. El caller es responsable de excluirla del
    lote de impresión y de escribir la nota en la planilla maestra (ver
    aplicar_alertas_a_maestro) y en el feedback."""
    por_rut = {}
    for r in recetas:
        rut = str(r.get("rut") or "").strip()
        if rut:
            por_rut.setdefault(rut, []).append(r)

    alertas = []
    for rut, grupo in por_rut.items():
        if len(grupo) < 2:
            continue

        # OJO — falso positivo real detectado y corregido (18-07-2026, caso
        # verificado en SSASUR): un mismo médico puede escribir DOS recetas
        # crónicas de medicamentos distintos en la misma consulta, con el mismo
        # período y la misma especialidad (ej. Fenitoína y Dabigatrán+Calcio,
        # ambas "MEDICINA GENERAL" 1/6). Eso NO es ambiguo, son dos tratamientos
        # paralelos legítimos. Por eso, cuando el dato de medicamentos está
        # disponible en TODOS los candidatos del grupo, se subagrupa por
        # medicamentos y solo se alerta si además coinciden — si difieren, no.
        vistos = {}
        for r in grupo:
            periodo_r, esp_r = _norm(r.get("periodo")), _norm(r.get("especialidad"))
            if not periodo_r or not esp_r:
                continue  # dato incompleto — no comparar vacíos contra vacíos
            vistos.setdefault((periodo_r, esp_r), []).append(r)
        for (periodo, especialidad), dup in vistos.items():
            if len(dup) < 2:
                continue
            if all(_set_medicamentos(d) for d in dup):
                por_meds = {}
                for d in dup:
                    por_meds.setdefault(_set_medicamentos(d), []).append(d)
                subgrupos = [g for g in por_meds.values() if len(g) >= 2]
            else:
                subgrupos = [dup]  # medicamentos no verificables en todos — no se puede descartar
            for conf in subgrupos:
                paciente = conf[0].get("paciente") or rut
                alertas.append({
                    "tipo": "ambiguo_mismo_periodo_especialidad",
                    "rut": rut, "paciente": paciente,
                    "recetas": [str(d.get("receta")) for d in conf],
                    "bloquear_impresion": True,
                    "nota": (
                        f"REVISAR: {len(conf)} recetas con el mismo período/especialidad "
                        f"({conf[0].get('periodo')} / {conf[0].get('especialidad')}) — "
                        "confirmar con SSASUR si es duplicado antes de preparar."
                    ),
                })

        for i in range(len(grupo)):
            for j in range(i + 1, len(grupo)):
                a, b = grupo[i], grupo[j]

                # Regla 2: especialidad distinta pero medicamentos idénticos.
                esp_a, esp_b = _norm(a.get("especialidad")), _norm(b.get("especialidad"))
                if esp_a and esp_b and esp_a != esp_b:
                    meds_a, meds_b = _set_medicamentos(a), _set_medicamentos(b)
                    if meds_a and meds_a == meds_b:
                        paciente = a.get("paciente") or rut
                        alertas.append({
                            "tipo": "especialidad_distinta_mismos_medicamentos",
                            "rut": rut, "paciente": paciente,
                            "recetas": [str(a.get("receta")), str(b.get("receta"))],
                            "bloquear_impresion": True,
                            "nota": (
                                f"REVISAR: receta {a.get('receta')} ({a.get('especialidad')}) y "
                                f"{b.get('receta')} ({b.get('especialidad')}) tienen los MISMOS medicamentos "
                                "pero especialidad distinta — no preparar/imprimir hasta confirmar con SSASUR/QF."
                            ),
                        })

                # Regla 3: duplicado informático — mismo período + misma Fecha
                # Digitación, Nº de receta distinto. Independiente de la
                # especialidad (el duplicado real verificado tenía la
                # MISMA especialidad en ambas copias). Solo dispara si el dato de
                # fecha_digitacion viene informado en ambos lados.
                #
                # OJO — falso positivo real detectado y corregido (18-07-2026,
                # caso verificado en SSASUR): un mismo médico puede escribir
                # DOS recetas crónicas de medicamentos distintos en la misma
                # consulta (misma Cuenta Corriente, mismo período, misma Fecha
                # Digitación) — ej. una serie de un medicamento y otra serie
                # de otro medicamento distinto, ambas 1/6→2/6 el 18/06/2026.
                # Eso NO es un duplicado, son dos tratamientos paralelos legítimos.
                # Por eso esta regla exige que los medicamentos TAMBIÉN coincidan
                # cuando el dato está disponible — si difieren, no se alerta.
                receta_a, receta_b = str(a.get("receta") or ""), str(b.get("receta") or "")
                if receta_a == receta_b:
                    continue
                periodo_a, periodo_b = _norm(a.get("periodo")), _norm(b.get("periodo"))
                fecha_a, fecha_b = a.get("fecha_digitacion"), b.get("fecha_digitacion")
                if not (periodo_a and periodo_a == periodo_b and fecha_a and fecha_b and str(fecha_a).strip() == str(fecha_b).strip()):
                    continue
                meds_a, meds_b = _set_medicamentos(a), _set_medicamentos(b)
                if meds_a and meds_b and meds_a != meds_b:
                    continue  # medicamentos distintos -> tratamientos paralelos legítimos, no duplicado
                if meds_a and meds_b:
                    confianza = "medicamentos idénticos confirmados"
                else:
                    confianza = "medicamentos no verificados (dato no disponible) — revisar manualmente en SSASUR antes de descartar"
                paciente = a.get("paciente") or rut
                alertas.append({
                    "tipo": "duplicado_informatico",
                    "rut": rut, "paciente": paciente,
                    "recetas": [receta_a, receta_b],
                    "bloquear_impresion": True,
                    "nota": (
                        f"REVISAR: recetas {receta_a} y {receta_b} tienen el mismo período ({a.get('periodo')}) "
                        f"y la misma Fecha Digitación ({fecha_a}) — probable duplicado informático de SSASUR ({confianza}). "
                        "Confirmar en SSASUR cuál es la vigente antes de preparar/imprimir cualquiera de las dos."
                    ),
                    })
    return alertas


def aplicar_alertas_a_maestro(wb, alertas):
    """Escribe la nota de cada alerta en la columna Pendiente de las filas
    afectadas (busca por Nº de receta en TODO el libro). No toca el Estado —
    la alerta bloquea la impresión, no el seguimiento. Devuelve cuántas filas
    se anotaron (0 si la receta aún no existe en ninguna hoja del maestro)."""
    tocadas = 0
    for alerta in alertas:
        for receta_num in alerta["recetas"]:
            ws = buscar_receta_en_maestro(wb, receta_num)
            if ws is None:
                continue
            upsert_receta(ws, {"receta": receta_num, "pendiente": alerta["nota"]})
            aplicar_formato_maestro(ws)
            tocadas += 1
    return tocadas


def _rango_avance(estado_norm):
    try:
        return ORDEN_ESTADOS.index(estado_norm)
    except ValueError:
        return -1


HOJA_HISTORIAL = "HISTORIAL"
HISTORIAL_HEADERS = ["Fecha y Hora", "Nº de receta", "Paciente", "Hoja",
                      "Estado Anterior", "Estado Nuevo"]


def _obtener_hoja_historial(wb):
    """Hoja de registro append-only: 1 fila por CAMBIO DE ESTADO real (no por
    cualquier edición — teléfonos, refrigerado/controlado, etc. no se
    registran acá, solo movimientos del Estado de una receta)."""
    if HOJA_HISTORIAL in wb.sheetnames:
        return wb[HOJA_HISTORIAL]
    ws = wb.create_sheet(HOJA_HISTORIAL)
    for c, h in enumerate(HISTORIAL_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    return ws


def _registrar_historial(ws_origen, receta_num, paciente, estado_anterior, estado_nuevo):
    ws_h = _obtener_hoja_historial(ws_origen.parent)
    fila = ws_h.max_row + 1
    ws_h.cell(row=fila, column=1, value=datetime.datetime.now())
    ws_h.cell(row=fila, column=2, value=receta_num)
    ws_h.cell(row=fila, column=3, value=paciente)
    ws_h.cell(row=fila, column=4, value=ws_origen.title)
    ws_h.cell(row=fila, column=5, value=estado_anterior or "(VACÍO)")
    ws_h.cell(row=fila, column=6, value=estado_nuevo)


def formatear_historial(wb):
    """Formato liviano de la hoja Historial: encabezado azul + bordes + anchos.
    No usa aplicar_formato_maestro() (esquema de columnas distinto: sin
    Establecimiento/Estado, no aplica el color por establecimiento)."""
    if HOJA_HISTORIAL not in wb.sheetnames:
        return
    ws = wb[HOJA_HISTORIAL]
    anchos = [18, 12, 28, 14, 16, 16]
    for c, (h, ancho) in enumerate(zip(HISTORIAL_HEADERS, anchos), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=G.BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = ancho
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(HISTORIAL_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor=(G.GREY if r % 2 == 0 else "FFFFFF"))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HISTORIAL_HEADERS))}{ws.max_row}"


def upsert_receta(ws, receta, estado=None, forzar_estado=False):
    """receta: dict con claves de COLUMN_MAP (recibe 'receta' obligatorio + las que se tengan).
    Devuelve 'nueva', 'actualizada' o 'sin_cambio_estado' (fila tocada igual si hay otros datos)."""
    fila_header, headers = _headers_de_hoja(ws)
    idx = {k: _col_index(headers, k) for k in COLUMN_MAP}
    faltantes = [k for k in ("receta", "estado") if idx[k] is None]
    if faltantes:
        raise ValueError(
            f"Hoja '{ws.title}': no reconozco la columna de {faltantes} entre los encabezados {headers!r}. "
            "Revisa _norm()/COLUMN_MAP antes de escribir — mejor fallar acá que guardar datos en la columna equivocada."
        )

    receta_num = str(receta.get("receta") or "").strip()
    if not receta_num:
        raise ValueError("receta['receta'] (Nº de receta) es obligatorio.")

    fila_encontrada = None
    for r in range(fila_header + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=idx["receta"] + 1).value
        if v is not None and str(v).strip() == receta_num:
            fila_encontrada = r
            break

    resultado = "actualizada"
    if fila_encontrada is None:
        fila_encontrada = ws.max_row + 1
        resultado = "nueva"

    ws.cell(row=fila_encontrada, column=idx["receta"] + 1, value=receta_num)
    for key, valor in receta.items():
        if key == "receta" or valor in (None, ""):
            continue
        ci = idx.get(key)
        if ci is not None:
            ws.cell(row=fila_encontrada, column=ci + 1, value=valor.upper() if isinstance(valor, str) else valor)

    if estado is not None and idx["estado"] is not None:
        celda_estado = ws.cell(row=fila_encontrada, column=idx["estado"] + 1)
        estado_previo = celda_estado.value
        actual_norm = _norm(estado_previo)
        nuevo_norm = _norm(estado)
        if forzar_estado or _rango_avance(nuevo_norm) >= _rango_avance(actual_norm):
            nuevo_valor = str(estado).strip().upper()
            if actual_norm != nuevo_norm:
                paciente_fila = ws.cell(row=fila_encontrada, column=idx["paciente"] + 1).value if idx["paciente"] is not None else None
                _registrar_historial(ws, receta_num, paciente_fila, estado_previo, nuevo_valor)
            celda_estado.value = nuevo_valor
        elif resultado == "actualizada":
            resultado = "sin_cambio_estado"

    return resultado


def aplicar_formato_maestro(ws):
    """Idempotente: banda de título navy + logos, encabezado azul, filas
    coloreadas por establecimiento de destino (incluida la columna Estado),
    freeze panes + autofiltro. Mismo lenguaje visual que skill_gt/scripts/generar.py.
    La hoja HISTORIAL tiene otro esquema (sin Establecimiento/Estado propios) —
    se formatea aparte con formatear_historial(), nunca acá."""
    if ws.title == HOJA_HISTORIAL:
        return
    ya_formateada = _tiene_titulo(ws)

    if not ya_formateada:
        headers = []
        for cell in ws[1]:
            if cell.value is None:
                break
            headers.append(cell.value)
        while headers and headers[-1] is None:
            headers.pop()
        ncols = len(headers)
        last_col = get_column_letter(ncols)
        ws.insert_rows(1, amount=1)
        ws.merge_cells(f"A1:{last_col}1")
        t = ws["A1"]
        t.value = f"GESTIÓN TERRITORIAL — FARMACIA HOSPITAL DE PITRUFQUÉN — {ws.title}"
        t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        t.fill = PatternFill("solid", fgColor=G.NAVY)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        G.place_logo(ws, G.LOGO_SS, "A", 26, 6, 2)
        G.place_logo(ws, G.LOGO_HOSP, last_col, 22, 6, 4)
    else:
        headers = []
        for cell in ws[2]:
            if cell.value is None:
                break
            headers.append(cell.value)
        ncols = len(headers)
        last_col = get_column_letter(ncols)

    HR = 2
    WIDTHS = {
        "ndereceta": 12, "nreceta": 12, "paciente": 26, "rut": 13, "establecimientodedestino": 20,
        "telefono": 13, "fechadesolicitud": 13, "periodoreceta": 10, "especialidad": 20,
        "numeroprescripciones": 10, "estado": 16,
        "fechaderetiroenfarmacia": 16,
        "refrigerado": 12, "controlado": 12, "pendiente": 24,
    }
    # WIDTHS de arriba es un PISO mínimo, no un valor fijo: si el contenido real
    # de la columna es más largo (ej. un nombre largo o una nota de Pendiente),
    # se ensancha para que no quede cortado visualmente. Tope distinto para
    # Pendiente (texto libre, puede ser una frase larga) que para el resto.
    TOPE_ANCHO, TOPE_PENDIENTE = 50, 65
    max_row_datos = ws.max_row
    anchos_contenido = {}
    for c in range(1, len(headers) + 1):
        largo = 0
        for r in range(HR + 1, max_row_datos + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            texto = v.strftime("%d/%m/%Y") if hasattr(v, "strftime") else str(v)
            largo = max(largo, len(texto))
        anchos_contenido[c] = largo

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=HR, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=G.BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        clave = _norm(h)
        piso = WIDTHS.get(clave, 14)
        tope = TOPE_PENDIENTE if clave == "pendiente" else TOPE_ANCHO
        ancho = min(max(piso, anchos_contenido.get(c, 0) + 2), tope)
        ws.column_dimensions[get_column_letter(c)].width = ancho
    ws.row_dimensions[HR].height = 30

    idx_estado = None
    idx_destino = None
    for i, h in enumerate(headers):
        hn = _norm(h)
        if hn == "estado":
            idx_estado = i
        elif hn == "establecimientodedestino":
            idx_destino = i

    r = HR + 1
    max_row = ws.max_row
    while r <= max_row:
        any_val = any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, len(headers) + 1))
        if not any_val:
            r += 1
            continue

        # Todo el texto de la planilla va en MAYÚSCULAS (ver upsert_receta para
        # las escrituras nuevas); esto normaliza a mano filas antiguas cargadas
        # antes de ese cambio. Fechas/números no son str, así que no los toca.
        for c in range(1, len(headers) + 1):
            cell_norm = ws.cell(row=r, column=c)
            if isinstance(cell_norm.value, str) and cell_norm.value:
                cell_norm.value = cell_norm.value.upper()

        # Toda la fila (incluida la columna Estado) se pinta del color del
        # establecimiento de destino — fondo casi blanco + texto oscuro del
        # mismo matiz, estilo pedido_fusion.py. Si la fila no tiene
        # establecimiento informado, se usa la cebra gris/blanco de siempre
        # con texto negro por defecto.
        destino_val = ws.cell(row=r, column=idx_destino + 1).value if idx_destino is not None else None
        color_fila = color_establecimiento(destino_val) if destino_val else None
        texto_fila = texto_establecimiento(destino_val) if destino_val else None
        fill_fila = color_fila or (G.GREY if (r - HR) % 2 == 0 else "FFFFFF")
        font_color = texto_fila or "000000"

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10, color=font_color,
                              bold=(idx_estado is not None and c == idx_estado + 1))
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill("solid", fgColor=fill_fila)
        ws.row_dimensions[r].height = 16
        r += 1

    if idx_estado is not None:
        col_estado = get_column_letter(idx_estado + 1)
        # La hoja puede traer una validación de Estado previa (de antes de
        # gt_maestro.py) con valores propios del flujo real (ej. "Validado QF",
        # "sin receta vigente", "pendiente por falta stock") que ORDEN_ESTADOS no
        # conoce. Se recogen y se fusionan con las nuestras en una sola lista —
        # nunca se descartan valores reales que ya estaban en uso.
        opciones_estado = ["EN REVISIÓN", "EN PREPARACIÓN", "LISTA PARA RETIRO",
                            "RETIRO EN VENTANILLA", "ENVIADA", "ENTREGADA"]
        vistas = {_norm(o) for o in opciones_estado}
        previas = [dv for dv in list(ws.data_validations.dataValidation)
                   if dv.type == "list" and any(str(r).startswith(col_estado) for r in dv.sqref.ranges)]
        for dv_previa in previas:
            formula = (dv_previa.formula1 or "").strip('"')
            for valor in formula.split(","):
                valor = valor.strip().upper()
                if valor and _norm(valor) not in vistas:
                    opciones_estado.append(valor)
                    vistas.add(_norm(valor))
            ws.data_validations.dataValidation.remove(dv_previa)

        dv = DataValidation(type="list", formula1='"' + ",".join(opciones_estado) + '"', allow_blank=True)
        dv.error = f"Elige un valor de la lista ({' / '.join(opciones_estado)})."
        dv.errorTitle = "Estado no válido"
        ws.add_data_validation(dv)
        dv.add(f"{col_estado}{HR + 1}:{col_estado}{max_row}")

    ws.freeze_panes = f"A{HR + 1}"
    ws.auto_filter.ref = f"A{HR}:{last_col}{max_row}"


# ----------------------- sincronización con el reporte GT crudo -----------------------
# El reporte "reporteGestionTerritorial_<desde>_<hasta>.xlsx" (el que baja
# AUTO_SSASUR.py --gt a 04_Farmacia_Gestion_Territorial/) trae 1 fila por
# PRODUCTO de la receta, no por receta. Se agrupa por Nº Receta antes de hacer
# upsert. Usar este archivo (no gt_enriquecido.json) porque es el único que
# trae el RUT — cruce_gt.py lo omite a propósito del JSON público por Ley 19.628.

_FMT_FECHA = ("%d/%m/%Y", "%d-%m-%Y")


def _parsear_fecha(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    for fmt in _FMT_FECHA:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def leer_reporte_gt_crudo(path):
    """Agrupa el reporte crudo por Nº Receta -> 1 dict por receta (primer valor
    no vacío / no 'Sin Información' encontrado para cada campo)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = [c.value for c in ws[2]]  # fila 1 = título del período, fila 2 = encabezados

    def col(*cands):
        for i, h in enumerate(headers):
            if _norm(h) in cands:
                return i
        return None

    c_receta = col("nreceta", "ndereceta")
    c_paciente = col("paciente")
    c_run = col("runpaciente", "run")
    c_destino = col("estabdestino")
    c_fentrega = col("fechaentrega")
    c_periodo = col("periodoreceta")
    c_esp = col("especialidad")
    c_npresc = col("numeroprescripciones")
    c_tel = col("telefono")
    c_producto = col("producto")
    if c_receta is None:
        raise ValueError(f"No se encontró la columna 'Nº Receta' en {path}")

    grupos = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if c_receta >= len(row) or row[c_receta] in (None, ""):
            continue
        crudo = row[c_receta]
        receta_num = str(int(crudo)) if isinstance(crudo, float) else str(crudo).strip()
        g = grupos.setdefault(receta_num, {"receta": receta_num, "medicamentos": set()})

        def _set(key, idxcol):
            if idxcol is None or idxcol >= len(row):
                return
            v = row[idxcol]
            if v not in (None, "", "Sin Información") and not g.get(key):
                g[key] = v

        _set("paciente", c_paciente)
        _set("rut", c_run)
        _set("destino", c_destino)
        _set("fecha_entrega", c_fentrega)
        _set("periodo", c_periodo)
        _set("especialidad", c_esp)
        _set("n_presc", c_npresc)
        _set("telefono", c_tel)

        if c_producto is not None and c_producto < len(row):
            prod = row[c_producto]
            if prod not in (None, "", 0, "0"):
                g["medicamentos"].add(str(prod).strip())

    return list(grupos.values())


def _clasificar_refrigerado_controlado(grupos):
    """Cruza cada receta contra el histórico (informe_completo_recetas*.csv) con
    la misma lógica que usan las planillas de envío (cruce_gt.py), y agrega
    'refrigerado'/'controlado' (texto "Medicamento xCantidad") a cada grupo
    cuando corresponda. NUNCA agrega 'pendiente' — esa clave la usa gt_maestro
    para las notas de alerta de duplicados (columna PENDIENTE), no para
    medicamentos pendientes; mezclarlas pisaría esas notas."""
    recetas_set = {g["receta"] for g in grupos}
    archivos = sorted(glob.glob(os.path.join(MAESTRO_DIR, "informe_completo_recetas*.csv")))
    if not archivos:
        return
    det = CG.cruzar_historico(recetas_set, archivos)
    for g in grupos:
        reg_tmp = {}
        CG.clasificar(reg_tmp, det.get(g["receta"], {"tipo_receta": "", "lineas": OrderedDict()}))
        if reg_tmp["refrigerado"]:
            g["refrigerado"] = reg_tmp["refrigerado"]
        if reg_tmp["controlado"]:
            g["controlado"] = reg_tmp["controlado"]


def sincronizar_gt_report(wb, path, estado_destino="EN PREPARACIÓN", forzar_estado=False):
    """Hace upsert de cada receta del reporte crudo. Si la receta ya existe en
    CUALQUIER hoja del maestro, se actualiza ahí (nunca se duplica entre meses
    aunque SSASUR corrija la Fecha Entrega de un mes a otro); si es nueva, se
    crea en la hoja del mes de su Fecha Entrega."""
    grupos = leer_reporte_gt_crudo(path)
    _clasificar_refrigerado_controlado(grupos)
    hojas_tocadas = {}
    resumen = {"nueva": 0, "actualizada": 0, "sin_cambio_estado": 0}
    for g in grupos:
        fecha = _parsear_fecha(g.get("fecha_entrega")) or datetime.date.today()
        ws, resultado, _ = upsert_receta_maestro(wb, g, estado=estado_destino, forzar_estado=forzar_estado, fecha_fallback=fecha)
        hojas_tocadas[ws.title] = ws
        resumen[resultado] = resumen.get(resultado, 0) + 1
    for ws in hojas_tocadas.values():
        aplicar_formato_maestro(ws)
    formatear_historial(wb)
    return resumen, list(hojas_tocadas.keys())


def _cli():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--maestro", help="Ruta al xlsx maestro (default: GT_MAESTRO_XLSX de utils_aa.py)")
    modo = ap.add_mutually_exclusive_group(required=True)
    modo.add_argument("--registrar", action="store_true", help="Agrega/actualiza una receta")
    modo.add_argument("--estado", action="store_true", help="Cambia el estado de una receta ya existente")
    modo.add_argument("--sincronizar-gt-report", dest="reporte",
                       help="Ruta al reporteGestionTerritorial_*.xlsx crudo (el que baja AUTO_SSASUR.py --gt)")
    ap.add_argument("--receta", help="Nº de receta (requerido en --registrar / --estado)")
    ap.add_argument("--paciente")
    ap.add_argument("--rut")
    ap.add_argument("--destino")
    ap.add_argument("--periodo")
    ap.add_argument("--especialidad")
    ap.add_argument("--telefono")
    ap.add_argument("--estado-valor", dest="estado_valor", default="EN REVISIÓN",
                     help="Estado a asignar en --registrar (default: EN REVISIÓN)")
    ap.add_argument("--nuevo-estado", dest="nuevo_estado", help="Estado a asignar en --estado")
    ap.add_argument("--estado-sync", dest="estado_sync", default="EN PREPARACIÓN",
                     help="Estado a asignar en --sincronizar-gt-report (default: EN PREPARACIÓN)")
    ap.add_argument("--forzar-estado", action="store_true")
    ap.add_argument("--mes", help="AAAA-MM de la hoja destino (default: mes en curso; ignorado en --sincronizar-gt-report)")
    ap.add_argument("--dry-run", action="store_true",
                     help="Procesa todo y muestra qué habría pasado, pero NO guarda el archivo. Úsalo para probar antes de tocar el maestro real.")
    a = ap.parse_args()

    if (a.registrar or a.estado) and not a.receta:
        ap.error("--receta es requerido con --registrar o --estado")

    wb, path = cargar_maestro(a.maestro)
    print(f"[MAESTRO] {path}" + ("  (DRY-RUN — no se va a guardar nada)" if a.dry_run else ""))

    if a.reporte:
        resumen, hojas = sincronizar_gt_report(wb, a.reporte, estado_destino=a.estado_sync, forzar_estado=a.forzar_estado)
        if a.dry_run:
            print(f"[DRY-RUN] sincronizaría -> {resumen} | hojas tocadas: {', '.join(hojas)}")
        else:
            guardar(wb, path)
            print(f"OK sincronizado -> {resumen} | hojas tocadas: {', '.join(hojas)}")
        return

    if a.mes:
        anio, mes = a.mes.split("-")
        fecha_fallback = datetime.date(int(anio), int(mes), 1)
    else:
        fecha_fallback = datetime.date.today()

    if a.estado:
        # --estado solo tiene sentido sobre una receta que ya existe en alguna
        # hoja; si no la encuentra en ninguna, es un Nº de receta mal tipeado —
        # mejor fallar que crear una fila fantasma casi vacía.
        ws_existente = buscar_receta_en_maestro(wb, a.receta)
        if ws_existente is None:
            print(f"[ERROR] Nº de receta {a.receta!r} no existe en ninguna hoja del maestro. Nada que actualizar.")
            return
        resultado = upsert_receta(ws_existente, {"receta": a.receta}, estado=a.nuevo_estado, forzar_estado=a.forzar_estado)
        ws = ws_existente
        es_nueva = False
    else:
        receta = {"receta": a.receta, "paciente": a.paciente, "rut": a.rut, "destino": a.destino,
                  "periodo": a.periodo, "especialidad": a.especialidad, "telefono": a.telefono}
        ws, resultado, es_nueva = upsert_receta_maestro(wb, receta, estado=a.estado_valor,
                                                          forzar_estado=a.forzar_estado, fecha_fallback=fecha_fallback)

    if es_nueva:
        print(f"Hoja {'que se crearía' if a.dry_run else 'nueva creada'}: {ws.title}")

    aplicar_formato_maestro(ws)
    formatear_historial(wb)
    if a.dry_run:
        print(f"[DRY-RUN] ({resultado}) -> {ws.title} — no se guardó")
    else:
        guardar(wb, path)
        print(f"OK ({resultado}) -> {ws.title} [{path}]")


if __name__ == "__main__":
    _cli()
