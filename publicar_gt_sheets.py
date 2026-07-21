#!/usr/bin/env python3
"""Publica el maestro GT COMPLETO (las 7 hojas de mes + HISTORIAL) a UN
Google Sheet permanente y compartido — versión DEFINITIVA aprobada por el
usuario el 19-07-2026 para poder editar desde cualquier equipo de la
farmacia sin depender de un .xlsx local sin sincronizar.

A diferencia de subir_prueba_sheets.py (crea una copia descartable nueva
cada vez, para previsualizar cambios de formato antes de aprobarlos), este
script SIEMPRE actualiza el mismo documento: su ID queda guardado en
_gt_sheets_id.json (versionado en el repo — así cualquier equipo que hace
`git pull` apunta automáticamente al mismo Sheet compartido).

Uso:
  py publicar_gt_sheets.py
"""
import json
import os
import sys

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, MAESTRO_DIR)
os.chdir(MAESTRO_DIR)

import openpyxl  # noqa: E402
import gt_maestro as GM  # noqa: E402
import generar as G  # noqa: E402

from google.oauth2.credentials import Credentials  # noqa: E402
from google.auth.transport.requests import Request  # noqa: E402
import gspread  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

TOKEN_FILE = os.path.join(MAESTRO_DIR, "token_drive.json")
SCOPES = ["https://www.googleapis.com/auth/drive"]
ID_FILE = os.path.join(MAESTRO_DIR, "_gt_sheets_id.json")
# "2 - Gestion Territorial" en Drive (ver _drive_folders.json)
CARPETA_DRIVE_GT = "1DIGX7gF2e_hi-8HffM3glQGHBnCkHxul"
TITULO_DOC = "GT PITRUFQUEN 2026 (Maestro Definitivo)"

creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
if not creds.valid:
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(TOKEN_FILE, "w", encoding="utf-8") as fh:
            fh.write(creds.to_json())
    else:
        raise SystemExit("Token inválido y sin refresh_token — hay que re-autorizar con publicar_drive.py --setup")

gc = gspread.authorize(creds)


def _hex_a_float(hexval):
    if not hexval or len(str(hexval)) < 6:
        return None
    h = str(hexval)[-6:]
    return {"red": int(h[0:2], 16) / 255, "green": int(h[2:4], 16) / 255, "blue": int(h[4:6], 16) / 255}


BLANCO = {"red": 1, "green": 1, "blue": 1}
NAVY = _hex_a_float(G.NAVY)
BLUE = _hex_a_float(G.BLUE)
GREY = _hex_a_float(G.GREY)
BORDE_GRIS = {"style": "SOLID", "color": {"red": 0xBF / 255, "green": 0xBF / 255, "blue": 0xBF / 255}}

CENTRADAS = {"nreceta", "ndereceta", "rut", "periodoreceta", "numeroprescripciones", "estado",
             "refrigerado", "controlado", "fechadesolicitud", "fechaderetiroenfarmacia", "telefono"}

OPCIONES_ESTADO = ["EN REVISIÓN", "EN PREPARACIÓN", "LISTA PARA RETIRO", "RETIRO EN VENTANILLA",
                   "ENVIADA", "ENTREGADA", "VALIDADO QF", "SIN RECETA VIGENTE", "PENDIENTE POR FALTA STOCK"]


def _obtener_o_crear_spreadsheet():
    if os.path.exists(ID_FILE):
        with open(ID_FILE, encoding="utf-8") as fh:
            data = json.load(fh)
        try:
            return gc.open_by_key(data["id"])
        except Exception as e:
            print(f"[AVISO] No se pudo abrir el Sheet guardado ({data.get('id')}): {e} — se crea uno nuevo.")
    sh = gc.create(TITULO_DOC, folder_id=CARPETA_DRIVE_GT)
    with open(ID_FILE, "w", encoding="utf-8") as fh:
        json.dump({"id": sh.id, "url": sh.url, "titulo": TITULO_DOC}, fh, ensure_ascii=False, indent=2)
    return sh


def _limpiar_bandings(sh, sheet_id):
    """Quita cualquier "banding" (cebra de colores alternos) ya existente en
    la hoja — necesario antes de volver a llamar addBanding() en una
    republicación, porque la API rechaza agregar bandas sobre un rango que
    ya las tiene ("No se pueden añadir colores de fondo alternos a un
    intervalo que ya los tiene")."""
    meta = sh.fetch_sheet_metadata()
    for hoja in meta.get("sheets", []):
        if hoja["properties"]["sheetId"] != sheet_id:
            continue
        bandings = hoja.get("bandedRanges", [])
        if bandings:
            sh.batch_update({"requests": [
                {"deleteBanding": {"bandedRangeId": b["bandedRangeId"]}} for b in bandings
            ]})


def _obtener_o_crear_tab(sh, titulo):
    for ws in sh.worksheets():
        if ws.title == titulo:
            ws.clear()
            _limpiar_bandings(sh, ws.id)
            return ws, False
    return sh.add_worksheet(title=titulo, rows=20, cols=20), True


def _publicar_hoja_mes(sh, ws_source, titulo):
    """Sube una hoja de mes (ENERO..JULIO) del Excel al Sheet, con el mismo
    formato que aplicar_formato_maestro (color por establecimiento, Estado
    incluido, dropdown, filtro, congelado)."""
    headers = [c.value for c in ws_source[2]]
    ncols = len(headers)
    idx_estado = next((i for i, h in enumerate(headers) if GM._norm(h) == "estado"), None)
    idx_pendiente = next((i for i, h in enumerate(headers) if GM._norm(h) == "pendiente"), None)

    filas_datos, fills_fila = [], []
    for r in range(3, ws_source.max_row + 1):
        vals = [ws_source.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        if all(v in (None, "") for v in vals):
            continue
        fila = [v.strftime("%d/%m/%Y") if hasattr(v, "strftime") else ("" if v is None else v) for v in vals]
        filas_datos.append(fila)
        fills_fila.append(_hex_a_float(ws_source.cell(row=r, column=1).fill.fgColor.rgb))

    wsheet, _ = _obtener_o_crear_tab(sh, titulo)
    FILA_TITULO, FILA_LEYENDA, FILA_HEADER, FILA_DATOS_INICIO = 1, 2, 3, 4
    total_filas = len(filas_datos) + 3
    wsheet.resize(rows=max(total_filas + 5, 20), cols=max(ncols, 1))

    titulo_banda = f"🏥 GESTIÓN TERRITORIAL — FARMACIA HOSPITAL DE PITRUFQUÉN — {titulo}"
    wsheet.update("A1", [[titulo_banda]], value_input_option="USER_ENTERED")
    wsheet.update(f"A{FILA_HEADER}", [headers] + filas_datos, value_input_option="USER_ENTERED")

    sheet_id = wsheet.id
    requests = [
        {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                                   "startColumnIndex": 0, "endColumnIndex": ncols}, "mergeType": "MERGE_ALL"}},
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {
                "backgroundColor": NAVY, "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE",
                "textFormat": {"bold": True, "fontSize": 14, "fontFamily": "Calibri", "foregroundColor": BLANCO},
            }},
            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)",
        }},
        {"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 0, "endIndex": 1},
            "properties": {"pixelSize": 40}, "fields": "pixelSize",
        }},
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.93, "green": 0.94, "blue": 0.97},
                                            "verticalAlignment": "MIDDLE"}},
            "fields": "userEnteredFormat(backgroundColor,verticalAlignment)",
        }},
        {"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 1, "endIndex": 2},
            "properties": {"pixelSize": 26}, "fields": "pixelSize",
        }},
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": 3, "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {
                "backgroundColor": BLUE, "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE", "wrapStrategy": "WRAP",
                "textFormat": {"bold": True, "fontSize": 10, "fontFamily": "Calibri", "foregroundColor": BLANCO},
            }},
            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,wrapStrategy,textFormat)",
        }},
        {"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 2, "endIndex": 3},
            "properties": {"pixelSize": 34}, "fields": "pixelSize",
        }},
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {"textFormat": {"fontFamily": "Calibri", "fontSize": 10}, "verticalAlignment": "MIDDLE"}},
            "fields": "userEnteredFormat(textFormat,verticalAlignment)",
        }},
        {"addBanding": {
            "bandedRange": {
                "range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": ncols},
                "rowProperties": {"headerColor": BLUE, "firstBandColor": BLANCO, "secondBandColor": GREY},
            }
        }},
        {"updateBorders": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": ncols},
            "top": BORDE_GRIS, "bottom": BORDE_GRIS, "left": BORDE_GRIS, "right": BORDE_GRIS,
            "innerHorizontal": BORDE_GRIS, "innerVertical": BORDE_GRIS,
        }},
        {"updateSheetProperties": {
            "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 3}},
            "fields": "gridProperties.frozenRowCount",
        }},
        {"setBasicFilter": {
            "filter": {"range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": total_filas,
                                  "startColumnIndex": 0, "endColumnIndex": ncols}}
        }},
        {"setDataValidation": {
            "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas,
                       "startColumnIndex": idx_estado, "endColumnIndex": idx_estado + 1},
            "rule": {
                "condition": {"type": "ONE_OF_LIST", "values": [{"userEnteredValue": v} for v in OPCIONES_ESTADO]},
                "showCustomUi": True, "strict": False,
            },
        }} if idx_estado is not None else None,
        {"addConditionalFormatRule": {
            "rule": {
                "ranges": [{"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas,
                            "startColumnIndex": idx_pendiente, "endColumnIndex": idx_pendiente + 1}],
                "booleanRule": {
                    "condition": {"type": "NOT_BLANK"},
                    "format": {"textFormat": {"bold": True, "foregroundColor": {"red": 0xC5 / 255, "green": 0x5A / 255, "blue": 0x11 / 255}}},
                },
            },
            "index": 0,
        }} if idx_pendiente is not None else None,
    ]
    requests = [r for r in requests if r is not None]

    for i, h in enumerate(headers):
        clave = GM._norm(h)
        ancho_excel = ws_source.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width or 14
        ancho = max(round(ancho_excel * 7 + 12), 50)
        requests.append({"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": i, "endIndex": i + 1},
            "properties": {"pixelSize": ancho}, "fields": "pixelSize",
        }})
        if clave in CENTRADAS:
            requests.append({"repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas, "startColumnIndex": i, "endColumnIndex": i + 1},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat.horizontalAlignment",
            }})

    sh.batch_update({"requests": requests})

    color_requests = []
    for i, fill in enumerate(fills_fila):
        if fill is None:
            continue
        fila_sheet = FILA_DATOS_INICIO + i
        color_requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": fila_sheet - 1, "endRowIndex": fila_sheet,
                           "startColumnIndex": 0, "endColumnIndex": ncols},
                "cell": {"userEnteredFormat": {"backgroundColor": fill}},
                "fields": "userEnteredFormat.backgroundColor",
            }
        })
    TANDA = 400
    for ini in range(0, len(color_requests), TANDA):
        sh.batch_update({"requests": color_requests[ini:ini + TANDA]})

    print(f"  {titulo}: {len(filas_datos)} fila(s)")


def _publicar_historial(sh, wb_source):
    if GM.HOJA_HISTORIAL not in wb_source.sheetnames:
        return
    ws_hist = wb_source[GM.HOJA_HISTORIAL]
    hist_headers = [c.value for c in ws_hist[1]]
    hist_ncols = len(hist_headers)
    filas_hist = []
    for r in range(2, ws_hist.max_row + 1):
        vals = [ws_hist.cell(row=r, column=c).value for c in range(1, hist_ncols + 1)]
        if all(v in (None, "") for v in vals):
            continue
        filas_hist.append([v.strftime("%d/%m/%Y %H:%M") if hasattr(v, "strftime") else ("" if v is None else v) for v in vals])

    wsheet, _ = _obtener_o_crear_tab(sh, GM.HOJA_HISTORIAL)
    wsheet.resize(rows=max(len(filas_hist) + 5, 20), cols=hist_ncols)
    wsheet.update("A1", [hist_headers] + filas_hist, value_input_option="USER_ENTERED")

    sheet_id = wsheet.id
    total_filas = len(filas_hist) + 1
    requests = [
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": hist_ncols},
            "cell": {"userEnteredFormat": {
                "backgroundColor": BLUE, "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE",
                "textFormat": {"bold": True, "fontSize": 10, "fontFamily": "Calibri", "foregroundColor": BLANCO},
            }},
            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)",
        }},
        {"addBanding": {
            "bandedRange": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": hist_ncols},
                "rowProperties": {"headerColor": BLUE, "firstBandColor": BLANCO, "secondBandColor": GREY},
            }
        }},
        {"updateBorders": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": hist_ncols},
            "top": BORDE_GRIS, "bottom": BORDE_GRIS, "left": BORDE_GRIS, "right": BORDE_GRIS,
            "innerHorizontal": BORDE_GRIS, "innerVertical": BORDE_GRIS,
        }},
        {"updateSheetProperties": {
            "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount",
        }},
        {"setBasicFilter": {
            "filter": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": total_filas,
                                  "startColumnIndex": 0, "endColumnIndex": hist_ncols}}
        }},
    ]
    sh.batch_update({"requests": requests})
    print(f"  {GM.HOJA_HISTORIAL}: {len(filas_hist)} fila(s)")


def main():
    wb, path = GM.cargar_maestro()
    for ws in wb.worksheets:
        GM.aplicar_formato_maestro(ws)
    GM.formatear_historial(wb)
    GM.guardar(wb, path)

    wb2 = openpyxl.load_workbook(path, data_only=True)

    sh = _obtener_o_crear_spreadsheet()
    print(f"Publicando en: {sh.url}")

    tabs_objetivo = []
    for ws in wb2.worksheets:
        if ws.title == GM.HOJA_HISTORIAL:
            continue
        tabs_objetivo.append(ws.title)
        _publicar_hoja_mes(sh, ws, ws.title)

    _publicar_historial(sh, wb2)

    # Limpieza: borra la hoja "Hoja 1" / "Sheet1" que deja gspread por
    # defecto al crear el documento la primera vez.
    for ws_sheets in sh.worksheets():
        if ws_sheets.title not in tabs_objetivo and ws_sheets.title != GM.HOJA_HISTORIAL:
            sh.del_worksheet(ws_sheets)
            print(f"  (borrada hoja por defecto: {ws_sheets.title})")

    print("\nURL definitiva:", sh.url)


if __name__ == "__main__":
    main()
