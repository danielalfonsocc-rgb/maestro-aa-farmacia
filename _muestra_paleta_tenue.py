"""Genera un Excel de muestra con la paleta tenue (tinta-economica) ya aplicada
al proyecto, para revisar/imprimir antes de regenerar los reportes reales.
Usa los MISMOS helpers que maestro_aa.py / app_pedidos.py (aa_colors)."""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from aa_colors import CRIT_FILL_HEX, crit_fill, FONT_CRITICO, soften, darken

wb = Workbook()
ws = wb.active
ws.title = "Muestra tenue"
ws.sheet_view.showGridLines = False

thin = Side(style='thin', color='D9D9D9')
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

# Banda de titulo (matiz azul, atenuado) — como _titulo()/banner
TIT = '1F4E78'
ws.merge_cells('A1:D1')
ws['A1'] = 'PEDIDO FARMACIA AA → BODEGA AA  ·  MUESTRA IMPRESIÓN CARTA'
ws['A1'].fill = PatternFill('solid', fgColor=soften(TIT))
ws['A1'].font = Font(bold=True, color=darken(TIT), name='Arial', size=12)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 24

# Encabezado de columnas (atenuado) — como style_sheet()
hdrs = ['Medicamento', 'Criticidad', 'A solicitar', 'Cobertura (días)']
for ci, h in enumerate(hdrs, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.fill = PatternFill('solid', fgColor=soften(TIT))
    c.font = Font(bold=True, color=darken(TIT), name='Arial', size=11)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = borde
ws.row_dimensions[2].height = 22

# Filas: una por nivel de criticidad, con su color real del proyecto
filas = [
    ('Insulina NPH 100UI', '1-CRITICO', 120, 0.0),
    ('Losartán 50mg',      '2-URGENTE', 300, 2.5),
    ('Enalapril 10mg',      '3-ALTO',    180, 4.0),
    ('Metformina 850mg',    '3-MODERADO',150, 7.0),
    ('Atorvastatina 20mg',  '4-BAJO',    100, 12.0),
    ('Paracetamol 500mg',   '5-OK',      0,   25.0),
]
for ri, (med, crit, sol, cob) in enumerate(filas, 3):
    fill = crit_fill(crit)
    es_crit = crit == '1-CRITICO'
    vals = [med, crit, sol, cob]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.fill = fill
        # Critico: texto vino + negrita (FONT_CRITICO). Resto: negro normal.
        c.font = Font(name='Arial', size=11, bold=es_crit,
                      color='7F1D1D' if es_crit else '000000')
        c.alignment = Alignment(horizontal='left' if ci == 1 else 'center',
                                vertical='center')
        c.border = borde

# Leyenda del semaforo
ws.cell(row=10, column=1, value='LEYENDA DE COLORES').font = Font(bold=True, size=10, color='334155')
leyenda = [('1-CRITICO', 'CRÍTICO — sin stock ni respaldo'),
           ('2-URGENTE', 'URGENTE — cobertura crítica'),
           ('3-ALTO',    'ALTO / MODERADO — reponer pronto'),
           ('4-BAJO',    'BAJO — vigilar'),
           ('5-OK',      'SUFICIENTE — sin necesidad')]
for i, (k, txt) in enumerate(leyenda, 11):
    c = ws.cell(row=i, column=1, value=txt)
    c.fill = crit_fill(k)
    c.font = Font(size=10, bold=(k == '1-CRITICO'),
                  color='7F1D1D' if k == '1-CRITICO' else '000000')
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

for ci, w in enumerate([42, 16, 14, 16], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

wb.save('Muestra_Paleta_Tenue.xlsx')
print('OK -> Muestra_Paleta_Tenue.xlsx')
