#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inventario_rapido.py — Instrumento de conteo + diferencia + pedido, SIN Consolidado
=====================================================================================
Versión liviana de `programacion_aa.py` para cuando NO se tiene acceso al
Consolidado_AA_MAESTRO.xlsx ni al historial de recetas (p.ej. en un
computador sin el repo Maestro AA clonado). Solo necesita los dos reportes
que se descargan directo de SSASUR y que cualquiera puede tener a mano:

  1. Reporte de Programación ("cantidad_de_productos_consumidos_en_centro_de
     _costo_farmacia*.xlsx", Reportes → Consumo por centro de costo →
     Centro de Costo = FARMACIA → Generar XLS)
       → Cantidad Programada / Cantidad Solicitada del ciclo.
  2. Reporte de stock ("reporte_de_stock_*.xlsx")
       → Stock Sistema de la bodega (por defecto BODEGA AT ABIERTA).

A diferencia de programacion_aa.py, este script NO calcula Consumo Promedio
Mensual ni Sugerencia de programación (eso requiere el historial completo de
recetas vía maestro_aa.py) — solo entrega el instrumento de conteo y, tras
contar físicamente, la Diferencia y una Cantidad a Pedir simple.

Si hay un Consolidado_AA_MAESTRO*.xlsx en la carpeta del proyecto (aunque sea
de días atrás), se autodetecta y se usa su lista de medicamentos AA (hoja
Pedido_Repos_Bodega) para acotar el instrumento a lo que Farmacia AA
realmente dispensa — los reportes crudos de SSASUR traen de todo (inyectables
de pabellón, hospitalización cerrada, etc.). Lo que quede fuera de ese
universo se guarda aparte, en la hoja "Fuera_Universo_AA" del mismo Excel,
no se pierde. Sin Consolidado no hay forma de saber qué es AA y qué no —
el instrumento queda como la unión cruda de ambos reportes, con aviso.

Uso — paso 1, generar el instrumento de conteo:
    py inventario_rapido.py --programacion reporte_prog.xlsx --stock reporte_stock.xlsx

    Genera Programacion_AA/Instrumento_Conteo_AA_<fecha>.xlsx. Imprímelo (o
    ábrelo en Excel/Sheets) y cuenta físicamente, anotando en la columna
    "Stock Real".

Uso — paso 2, aplicar el conteo:
    py inventario_rapido.py --aplicar-conteo Instrumento_Conteo_AA_<fecha>.xlsx
        (si ya llenaste la columna "Stock Real" directo en ese Excel)
  o
    py inventario_rapido.py --aplicar-conteo conteo.json
        conteo.json = {"MEDICAMENTO TAL COMO SALE EN LA PLANILLA": 123, ...}
        (si prefieres pasar los conteos como texto/JSON; se aplican sobre el
        Instrumento_Conteo_AA_*.xlsx más reciente, o el que indiques con
        --plantilla)

    Genera Programacion_AA/Resumen_Conteo_AA_<fecha_hora>.xlsx con:
      - Diferencia = Stock Real − Stock Sistema (discrepancia de inventario)
      - Cantidad a Pedir = max(0, objetivo − Stock Real), donde objetivo es
        Cantidad Programada (o Solicitada si no hay Programada). Es una
        fórmula simple de "reponer hasta lo programado" — no reemplaza el
        modelo SGLI completo (sgli.py), que sí usa consumo histórico.

Otras opciones:
    --bodega "NOMBRE"   Filtra el reporte de stock por otra bodega (default:
                         BODEGA AT ABIERTA — debe calzar con BODEGA_FISICA_AA
                         de programacion_aa.py). Si el nombre no aparece en
                         el archivo, el script lista las bodegas disponibles.

Sin llamadas a IA — solo pandas + openpyxl.
"""
import os
import sys
import json
import glob
import argparse
import datetime as dt

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, WORK_DIR)
from utils_aa import norm_erp, HOMOLOGACION, setup_stdout  # noqa: E402

setup_stdout()

OUT_DIR = os.path.join(WORK_DIR, 'Programacion_AA')
BODEGA_DEFAULT = 'BODEGA AT ABIERTA'  # debe calzar con BODEGA_FISICA_AA de programacion_aa.py

THIN  = Side(style='thin', color='DCDCDC')
BRD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFILL = PatternFill('solid', fgColor='D1FAF5')
HFONT = Font(bold=True, color='065F46', name='Arial', size=10)
PEDIR_FILL = PatternFill('solid', fgColor='FFE0B2')
PEDIR_FONT = Font(name='Arial', size=10, color='B45309', bold=True)
DIFF_FILL  = PatternFill('solid', fgColor='F4B3B3')
DIFF_FONT  = Font(name='Arial', size=10, color='7F1D1D', bold=True)
FALTA_FILL = PatternFill('solid', fgColor='E5E7EB')


def _key(nombre):
    n = norm_erp(nombre)
    return HOMOLOGACION.get(n, n)


# ─────────────── lectura de los dos reportes crudos ─────────────────────────

def _leer_programacion(ruta):
    """Lee el reporte 'Consumos por centro de costo' de SSASUR (Generar XLS).
    Fila 0 = título, fila 1 = metadata con el mes/año, fila 2 = encabezados.
    Devuelve {key: {'Medicamento': nombre_original, 'Cantidad Programada': x,
    'Cantidad Solicitada': y}} y el texto de metadata (mes/año) para mostrar."""
    meta = pd.read_excel(ruta, header=None, nrows=2, engine='openpyxl')
    meta_txt = str(meta.iloc[1, 0]).strip() if meta.shape[0] > 1 else ''

    df = pd.read_excel(ruta, header=2, engine='openpyxl')
    df = df.rename(columns=lambda c: str(c).strip())
    faltantes = [c for c in ('Producto', 'Total de Productos Programados', 'Productos Solicitado')
                 if c not in df.columns]
    if faltantes:
        print(f'[ERROR] El reporte de Programación no tiene las columnas esperadas {faltantes}.')
        print(f'  Columnas encontradas: {list(df.columns)}')
        print('  ¿Es el archivo correcto? (Reportes → Consumo por centro de costo → Generar XLS)')
        sys.exit(1)
    if 'Centro Costo' in df.columns:
        df = df[df['Centro Costo'].astype(str).str.strip().str.upper() == 'FARMACIA']
    df['Producto'] = df['Producto'].astype(str).str.strip()
    df = df[df['Producto'] != '']
    df['_key'] = df['Producto'].map(_key)
    df['Total de Productos Programados'] = pd.to_numeric(df['Total de Productos Programados'], errors='coerce')
    df['Productos Solicitado'] = pd.to_numeric(df['Productos Solicitado'], errors='coerce')

    prog = {}
    for row in df.to_dict('records'):
        k = row['_key']
        cur = prog.setdefault(k, {'Medicamento': row['Producto'], 'Cantidad Programada': 0.0, 'Cantidad Solicitada': 0.0})
        p = row['Total de Productos Programados']
        s = row['Productos Solicitado']
        if not pd.isna(p):
            cur['Cantidad Programada'] += p
        if not pd.isna(s):
            cur['Cantidad Solicitada'] += s
    return prog, meta_txt


def _leer_stock(ruta, bodega):
    """Lee un reporte de stock de SSASUR y devuelve
    {key: {'Medicamento': nombre_original, 'Stock Sistema': cantidad}},
    sumando por lote si corresponde.

    Soporta dos formatos:
      - Volcado crudo multi-bodega ("reporte_de_stock_*.xlsx"): columnas
        Descripción/Bodega/Cantidad — se filtra por `bodega`. Si el nombre
        no aparece, lista las disponibles y aborta.
      - Reporte "Existencias" (ya viene acotado a una sola bodega/farmacia
        al generarlo en SSASUR): columnas Producto/Stock Disponible, sin
        columna de Bodega — se usa tal cual, sin filtrar.
    """
    df = pd.read_excel(ruta, header=2, engine='openpyxl')
    df.columns = [str(c).strip() for c in df.columns]

    bod_col = next((c for c in df.columns if 'Bodega' in c), None)
    descr_col = next((c for c in df.columns if 'Descrip' in c or c == 'Producto'), None)
    cant_col = next((c for c in df.columns
                      if c == 'Cantidad' or 'Stock Disponible' in c), None)

    if descr_col is None or cant_col is None:
        print('[ERROR] El reporte de stock no tiene columnas reconocibles '
              '(se esperaba Descripción/Bodega/Cantidad, o Producto/Stock Disponible).')
        print(f'  Columnas encontradas: {list(df.columns)}')
        sys.exit(1)

    if bod_col is not None:
        df[bod_col] = df[bod_col].astype(str).str.strip().str.upper()
        bodega_norm = bodega.strip().upper()
        disponibles = sorted(df[bod_col].dropna().unique())
        if bodega_norm not in disponibles:
            print(f'[ERROR] No se encontró la bodega "{bodega}" en el reporte de stock.')
            print('  Bodegas disponibles en el archivo:')
            for b in disponibles:
                print(f'    - {b}')
            print('  Reintenta con --bodega "NOMBRE EXACTO"')
            sys.exit(1)
        df = df[df[bod_col] == bodega_norm]
    else:
        # Sin columna de Bodega: el reporte "Existencias" ya viene acotado a
        # una sola bodega/farmacia (la que se eligió al generarlo en SSASUR)
        # — se usa completo, sin filtrar.
        print(f'  [aviso] el reporte de stock no trae columna de Bodega — se asume que ya '
              f'está acotado a "{bodega}" (elegida al generarlo en SSASUR). '
              f'Verifica que sea así antes de confiar en el instrumento.')

    sub = df.copy()
    sub['_cantidad'] = pd.to_numeric(sub[cant_col], errors='coerce').fillna(0)
    sub['_key'] = sub[descr_col].astype(str).str.strip().map(_key)

    stock = {}
    for row in sub.to_dict('records'):
        k = row['_key']
        cur = stock.setdefault(k, {'Medicamento': str(row[descr_col]).strip(), 'Stock Sistema': 0.0})
        cur['Stock Sistema'] += row['_cantidad']
    return stock


# ─────────────── modo 1: generar instrumento de conteo ──────────────────────

HDRS = [
    ('Medicamento',            48),
    ('Cantidad Programada',    18),
    ('Cantidad Solicitada',    18),
    ('Stock Sistema',          16),
    ('Stock Real',             18),
]
FILA_ALTO = 26      # alto de fila (pt) — deja espacio cómodo para escribir a mano
FUENTE_DATOS = 12   # tamaño de fuente de los datos — legible al imprimir


def _mas_reciente_consolidado():
    cand = [f for f in glob.glob(os.path.join(WORK_DIR, 'Consolidado_AA_MAESTRO*.xlsx'))
            if not os.path.basename(f).startswith('~$')]
    return max(cand, key=os.path.getmtime) if cand else None


def _leer_universo_aa(ruta):
    """Lee la lista oficial de medicamentos que sí dispensa Farmacia AT
    Abierta desde Consolidado_AA_MAESTRO (hoja Pedido_Repos_Bodega — la
    misma que usa programacion_aa.py). Devuelve {key: nombre_original}."""
    df = pd.read_excel(ruta, sheet_name='Pedido_Repos_Bodega', engine='openpyxl')
    df['Medicamento'] = df['Medicamento'].astype(str).str.strip()
    df = df[df['Medicamento'] != '']
    return {_key(m): m for m in df['Medicamento']}


def generar(ruta_prog, ruta_stock, bodega, ruta_consolidado=None):
    prog, meta_txt = _leer_programacion(ruta_prog)
    stock = _leer_stock(ruta_stock, bodega)

    # Si hay un Consolidado_AA_MAESTRO a mano (aunque sea de días atrás), se usa
    # su lista de 441 medicamentos AA (hoja Pedido_Repos_Bodega) para acotar el
    # instrumento a lo que Farmacia AA realmente dispensa — los reportes crudos
    # de SSASUR (Programación/Existencias) traen de todo (inyectables de
    # pabellón, hospitalización cerrada, etc.) y sin este filtro el instrumento
    # queda "sucio". Sin Consolidado no hay forma de saber cuáles son AA y
    # cuáles no — se usa la unión cruda de ambos reportes, con aviso.
    ruta_consolidado = ruta_consolidado or _mas_reciente_consolidado()
    universo_aa = None
    if ruta_consolidado:
        try:
            universo_aa = _leer_universo_aa(ruta_consolidado)
        except (OSError, ValueError, KeyError) as e:
            print(f'  [aviso] no se pudo leer el universo AA de {os.path.basename(ruta_consolidado)}: {e}')

    universo = {}
    fuera_universo = []
    if universo_aa is not None:
        for k, nombre in universo_aa.items():
            universo[k] = {'Medicamento': nombre, 'Cantidad Programada': None,
                            'Cantidad Solicitada': None, 'Stock Sistema': None, 'Stock Real': None}

    for k, v in prog.items():
        if universo_aa is not None and k not in universo:
            fuera_universo.append({'Medicamento': v['Medicamento'], 'Fuente': 'Programación',
                                    'Cantidad': v['Cantidad Programada']})
            continue
        cur = universo.setdefault(k, {'Medicamento': v['Medicamento'], 'Cantidad Programada': None,
                                       'Cantidad Solicitada': None, 'Stock Sistema': None, 'Stock Real': None})
        cur['Cantidad Programada'] = v['Cantidad Programada']
        cur['Cantidad Solicitada'] = v['Cantidad Solicitada']

    for k, v in stock.items():
        if universo_aa is not None and k not in universo:
            fuera_universo.append({'Medicamento': v['Medicamento'], 'Fuente': 'Stock',
                                    'Cantidad': v['Stock Sistema']})
            continue
        cur = universo.setdefault(k, {'Medicamento': v['Medicamento'], 'Cantidad Programada': None,
                                       'Cantidad Solicitada': None, 'Stock Sistema': None, 'Stock Real': None})
        cur['Stock Sistema'] = v['Stock Sistema']

    filas = sorted(universo.values(), key=lambda f: f['Medicamento'])
    fuera_universo.sort(key=lambda f: f['Medicamento'])
    n_sin_datos = sum(1 for f in filas if f['Cantidad Programada'] is None and f['Stock Sistema'] is None)

    hoy = dt.date.today()
    os.makedirs(OUT_DIR, exist_ok=True)
    sal = os.path.join(OUT_DIR, f'Instrumento_Conteo_AA_{hoy.strftime("%Y%m%d")}.xlsx')
    sub = (f'Programación: {os.path.basename(ruta_prog)} ({meta_txt or "sin metadata de mes"})  ·  '
           f'Stock: {os.path.basename(ruta_stock)}  ·  Bodega: {bodega}')
    if ruta_consolidado:
        sub += f'  ·  Universo AA: {os.path.basename(ruta_consolidado)} ({len(universo_aa)} medicamentos)'
    extra_sheet = None
    if fuera_universo:
        extra_sheet = ('Fuera_Universo_AA', fuera_universo)
    _escribir(sal, filas, 'INSTRUMENTO DE CONTEO — Bodega AA', sub, hoy, resumen=False, extra_sheet=extra_sheet)

    print(f'HOY = {hoy}')
    if ruta_consolidado:
        print(f'Universo AA         : {os.path.basename(ruta_consolidado)} ({len(universo_aa)} medicamentos)')
        print(f'{len(filas)} medicamentos en el instrumento ({n_sin_datos} sin Cantidad Programada ni Stock Sistema este ciclo)')
        if fuera_universo:
            print(f'{len(fuera_universo)} fila(s) de los reportes NO pertenecen al universo AA — '
                  f'quedaron en la hoja "Fuera_Universo_AA" del Excel, no en el instrumento de conteo.')
    else:
        n_solo_stock = sum(1 for k in stock if k not in prog)
        print(f'  [aviso] no se encontró Consolidado_AA_MAESTRO*.xlsx — no se pudo acotar al universo AA real. '
              f'El instrumento es la unión cruda de ambos reportes y puede incluir medicamentos que Farmacia AA '
              f'no dispensa (p.ej. inyectables de pabellón, hospitalización cerrada).')
        print(f'{len(filas)} medicamentos ({len(prog)} en Programación, {n_solo_stock} solo en Stock — sin programar)')
    print(f'\nExcel: {os.path.basename(sal)}  (carpeta {os.path.basename(OUT_DIR)}\\)')
    print('Imprime o abre esta planilla, cuenta físicamente y llena "Stock Real" a mano.')
    print(f'Luego corre: py inventario_rapido.py --aplicar-conteo "{os.path.basename(sal)}"')
    return sal


# ─────────────── modo 2: aplicar conteo → resumen con pedido ────────────────

def _mas_reciente_instrumento():
    cand = [f for f in glob.glob(os.path.join(OUT_DIR, 'Instrumento_Conteo_AA_*.xlsx'))
            if not os.path.basename(f).startswith('~$')]
    return max(cand, key=os.path.getmtime) if cand else None


def _cantidad_a_pedir(programada, solicitada, stock_real):
    if stock_real is None:
        return None
    objetivo = programada if programada not in (None,) and not (isinstance(programada, float) and pd.isna(programada)) else solicitada
    if objetivo is None or (isinstance(objetivo, float) and pd.isna(objetivo)):
        return None
    return max(0, round(objetivo - stock_real))


def aplicar_conteo(ruta, ruta_plantilla=None):
    hoy = dt.date.today()

    if ruta.lower().endswith('.xlsx'):
        # El propio instrumento, ya con "Stock Real" llenado a mano.
        plantilla = ruta
        wb = openpyxl.load_workbook(plantilla, data_only=True)
        ws = wb.active
        filas_base = []
        for row in ws.iter_rows(min_row=4, max_col=5):
            med_c, prog_c, sol_c, sbod_c, sreal_c = row
            if med_c.value is None:
                continue
            filas_base.append({
                'Medicamento': med_c.value,
                'Cantidad Programada': prog_c.value,
                'Cantidad Solicitada': sol_c.value,
                'Stock Sistema': sbod_c.value,
                'Stock Real': sreal_c.value,
            })
    else:
        # JSON medicamento→cantidad, aplicado sobre el instrumento más reciente.
        plantilla = ruta_plantilla or _mas_reciente_instrumento()
        if not plantilla:
            print('[ERROR] No hay ningún Instrumento_Conteo_AA_*.xlsx generado todavía.')
            print('  Corre primero: py inventario_rapido.py --programacion ... --stock ...')
            sys.exit(1)
        try:
            with open(ruta, encoding='utf-8') as fh:
                valores = json.load(fh)
        except (OSError, json.JSONDecodeError) as e:
            print(f'[ERROR] No se pudo leer {ruta} como JSON: {e}')
            sys.exit(1)
        valores_key = {_key(k): v for k, v in valores.items()}

        wb = openpyxl.load_workbook(plantilla, data_only=True)
        ws = wb.active
        filas_base = []
        for row in ws.iter_rows(min_row=4, max_col=5):
            med_c, prog_c, sol_c, sbod_c, sreal_c = row
            if med_c.value is None:
                continue
            k = _key(med_c.value)
            filas_base.append({
                'Medicamento': med_c.value,
                'Cantidad Programada': prog_c.value,
                'Cantidad Solicitada': sol_c.value,
                'Stock Sistema': sbod_c.value,
                'Stock Real': valores_key.get(k, sreal_c.value),
            })

    filas = []
    n_diff = 0
    n_pedir = 0
    n_falta = 0
    total_pedir = 0
    for f in filas_base:
        sreal = f['Stock Real']
        sbod = f['Stock Sistema']
        diff = None
        if sreal is not None and sbod is not None:
            diff = sreal - sbod
            if diff:
                n_diff += 1
        pedir = _cantidad_a_pedir(f['Cantidad Programada'], f['Cantidad Solicitada'], sreal)
        if sreal is None:
            n_falta += 1
        elif pedir:
            n_pedir += 1
            total_pedir += pedir
        filas.append({**f, 'Diferencia': diff, 'Cantidad a Pedir': pedir})

    os.makedirs(OUT_DIR, exist_ok=True)
    sal = os.path.join(OUT_DIR, f'Resumen_Conteo_AA_{hoy.strftime("%Y%m%d_%H%M")}.xlsx')
    sub = f'Basado en {os.path.basename(plantilla)}  ·  {n_falta} sin contar aún'
    _escribir(sal, filas, 'RESUMEN CONTEO vs PROGRAMACIÓN — Bodega AA', sub, hoy, resumen=True)

    print(f'Planilla base       : {os.path.basename(plantilla)}')
    print(f'Medicamentos        : {len(filas)}  ({n_falta} sin "Stock Real" todavía)')
    print(f'Con diferencia stock: {n_diff}  (Stock Real ≠ Stock Sistema)')
    print(f'A pedir             : {n_pedir} medicamento(s), {total_pedir} unidades en total')
    print(f'\nExcel: {os.path.basename(sal)}  (carpeta {os.path.basename(OUT_DIR)}\\)')
    if n_pedir:
        print('\nTop qué pedir (mayor cantidad primero):')
        top = sorted([f for f in filas if f['Cantidad a Pedir']],
                     key=lambda f: f['Cantidad a Pedir'], reverse=True)[:15]
        for f in top:
            print(f'  - {f["Medicamento"]}: {f["Cantidad a Pedir"]} ud')
    return sal


# ─────────────── escritura de Excel (instrumento y resumen comparten look) ──

def _escribir(sal, filas, titulo, subtitulo, hoy, resumen, extra_sheet=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumen' if resumen else 'Conteo'

    hdrs = list(HDRS)
    if resumen:
        hdrs = hdrs + [('Diferencia (Real − Sistema)', 20), ('Cantidad a Pedir', 16)]
    ncols = len(hdrs)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, f'{titulo}  ·  {hoy.strftime("%d/%m/%Y")}')
    ws.cell(1, 1).font = Font(bold=True, size=14, color='065F46', name='Arial')
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, subtitulo)
    ws.cell(2, 1).font = Font(italic=True, size=10, color='555555', name='Arial')
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 30

    for j, (label, w) in enumerate(hdrs, 1):
        c = ws.cell(3, j, label)
        c.fill = HFILL; c.font = Font(bold=True, color='065F46', name='Arial', size=12); c.border = BRD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[3].height = 30

    for i, f in enumerate(filas, 4):
        vals = [f['Medicamento'], f['Cantidad Programada'], f['Cantidad Solicitada'],
                f['Stock Sistema'], f['Stock Real']]
        if resumen:
            vals += [f.get('Diferencia'), f.get('Cantidad a Pedir')]
        ws.row_dimensions[i].height = FILA_ALTO
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            c.border = BRD
            c.font = Font(name='Arial', size=FUENTE_DATOS)
            if j >= 2:
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(vertical='center')
        if resumen:
            if f['Stock Real'] is None:
                for j in range(1, ncols + 1):
                    ws.cell(i, j).fill = FALTA_FILL
            else:
                if f.get('Diferencia'):
                    dc = ws.cell(i, len(HDRS) + 1)
                    dc.fill = DIFF_FILL
                    dc.font = Font(name='Arial', size=FUENTE_DATOS, color='7F1D1D', bold=True)
                if f.get('Cantidad a Pedir'):
                    pc = ws.cell(i, len(HDRS) + 2)
                    pc.fill = PEDIR_FILL
                    pc.font = Font(name='Arial', size=FUENTE_DATOS, color='B45309', bold=True)

    last = 3 + len(filas)
    ws.freeze_panes = 'A4'
    if filas:
        ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}{last}'
        ws.print_area = f'A1:{get_column_letter(ncols)}{last}'

    # Impresión en Carta (Letter), horizontal, con encabezado repetido en cada
    # página y filas altas para que sea cómodo escribir a mano el conteo.
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.page_margins.header = ws.page_margins.footer = 0.2
    ws.print_title_rows = '3:3'
    ws.oddFooter.center.text = 'Página &P de &N'
    ws.oddFooter.center.size = 9

    if extra_sheet:
        titulo_hoja, filas_extra = extra_sheet
        ws2 = wb.create_sheet(titulo_hoja[:31])
        ehdrs = [('Medicamento', 46), ('Fuente', 14), ('Cantidad', 12)]
        ws2.cell(1, 1, 'NO pertenecen al universo AA (Consolidado) — no incluidos en el instrumento de conteo')
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ehdrs))
        ws2.cell(1, 1).font = Font(bold=True, size=11, color='7F1D1D', name='Arial')
        for j, (label, w) in enumerate(ehdrs, 1):
            c = ws2.cell(2, j, label)
            c.fill = HFILL; c.font = HFONT; c.border = BRD
            ws2.column_dimensions[get_column_letter(j)].width = w
        for i, f in enumerate(filas_extra, 3):
            for j, key in enumerate(('Medicamento', 'Fuente', 'Cantidad'), 1):
                c = ws2.cell(i, j, f.get(key))
                c.border = BRD
                c.font = Font(name='Arial', size=10)
        ws2.freeze_panes = 'A3'

    wb.save(sal)


# ─────────────── main ────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description='Instrumento de conteo + diferencia + pedido, sin depender del Consolidado_AA_MAESTRO.')
    ap.add_argument('--programacion', default=None, metavar='XLSX',
                     help='Reporte SSASUR de Programación (Consumo por centro de costo → FARMACIA)')
    ap.add_argument('--stock', default=None, metavar='XLSX',
                     help='Reporte de stock crudo de SSASUR (reporte_de_stock_*.xlsx)')
    ap.add_argument('--bodega', default=BODEGA_DEFAULT,
                     help=f'Nombre de la bodega a filtrar en el reporte de stock (default: {BODEGA_DEFAULT})')
    ap.add_argument('--consolidado', default=None, metavar='XLSX',
                     help='Consolidado_AA_MAESTRO.xlsx a usar para acotar el instrumento al universo AA real '
                          '(si no, se autodetecta el más reciente en la carpeta del proyecto; si no hay ninguno, '
                          'el instrumento queda como la unión cruda de los dos reportes)')
    ap.add_argument('--aplicar-conteo', default=None, metavar='XLSX_O_JSON',
                     help='Aplica el conteo: pásale el Instrumento_Conteo_AA_*.xlsx ya lleno, '
                          'o un JSON medicamento→cantidad')
    ap.add_argument('--plantilla', default=None, metavar='XLSX',
                     help='Fuerza el Instrumento_Conteo_AA_*.xlsx a usar con --aplicar-conteo JSON '
                          '(si no, usa el más reciente)')
    args = ap.parse_args()

    if args.aplicar_conteo:
        aplicar_conteo(args.aplicar_conteo, args.plantilla)
    else:
        if not args.programacion or not args.stock:
            print('[ERROR] Falta --programacion y/o --stock. Usa --help para ver el uso.')
            sys.exit(1)
        generar(args.programacion, args.stock, args.bodega, args.consolidado)


if __name__ == '__main__':
    main()
