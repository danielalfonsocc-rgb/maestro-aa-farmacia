#!/usr/bin/env python3
"""
revision_solicitudes.py — Automatiza el paso 4 del workflow de Gestión
Territorial: combinar los PDF individuales de una solicitud en un solo
documento y generar la planilla de feedback para el QF.

Antes (manual): juntar a mano los PDF (Receta_*.pdf / Historico_*.pdf) en un
Recetas_Combinadas_<ESTAB>_<fecha>.pdf y armar el Feedback_Solicitud_*.xlsx.

Ahora: dejas los PDF sueltos en
  04_Farmacia_Gestion_Territorial/<ESTABLECIMIENTO>/Revisión de Solicitudes/
y este script:
  1. Extrae el Nº de receta de cada nombre de archivo (Receta_<n>_... /
     Historico_<n>_...) y busca esa receta en informe_completo_recetas*.csv
     para obtener RUT, paciente, especialidad, período, medicamentos y
     Fecha Digitación reales (no adivinados del nombre del archivo).
  2. Por cada RUT encontrado, busca TODAS sus recetas vigentes en el CSV
     (no solo la que llegó en el PDF) y corre gt_maestro.detectar_alertas_mismo_rut()
     sobre el set completo — si el paciente tiene 2+ recetas vigentes, se
     revisan las dos juntas aplicando las reglas ya definidas (duplicado
     informático / especialidad distinta mismos medicamentos / ambiguo
     mismo período-especialidad) para no confundir tratamientos paralelos
     legítimos con duplicados reales.
  3. Combina los PDF sueltos en un solo Recetas_Combinadas_<ESTAB>_<fecha>.pdf.
  4. Genera Feedback_Solicitud_<ESTAB>_<fecha>.xlsx con una fila por receta
     vigente (la que llegó en el PDF + las adicionales que aparezcan), con
     la alerta (si la hay) volcada en la columna Observación, y columnas en
     blanco para que el QF registre Acción tomada / Medicamento pendiente /
     Fecha de revisión.
  5. Mueve los PDF procesados a la subcarpeta "PDFs individuales".

También soporta el formato real en que suelen llegar las solicitudes: una
planilla xlsx con columnas FECHA/RUT/NOMBRE (sin Nº de receta) — para ese
caso usa --xlsx: busca cada RUT en informe_completo_recetas*.csv, trae TODAS
sus recetas vigentes, corre las mismas reglas de alerta, y archiva la
planilla original recibida en la carpeta del establecimiento para que quede
como registro (no se pierde en revisiones futuras).

Uso:
  py revision_solicitudes.py --estab "HOSPITAL TOLTEN" --dry-run   (PDFs sueltos, preview)
  py revision_solicitudes.py --estab "HOSPITAL TOLTEN"             (PDFs sueltos, aplica)
  py revision_solicitudes.py --estab "CESFAM FREIRE" --xlsx "C:/ruta/solicitud.xlsx" --dry-run
  py revision_solicitudes.py --estab "CESFAM FREIRE" --xlsx "C:/ruta/solicitud.xlsx"
"""
import argparse
import csv
import datetime
import glob
import os
import re
import sys
import unicodedata

import openpyxl
from openpyxl.styles import Font, PatternFill
from pypdf import PdfWriter

import gt_maestro as GM
from agregar_gt_manual import _CARPETA_LOCAL, GT_SOLICITUDES_DIR

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
NOMBRE_PDF_RE = re.compile(r"^(Historico|Receta)_(\d+)_.+\.pdf$", re.IGNORECASE)


def _cargar_recetas_csv():
    """Lee todos los informe_completo_recetas*.csv del repo y arma un índice
    por Nº de receta -> lista de filas (una receta puede tener varias filas,
    una por medicamento prescrito)."""
    por_receta = {}
    for path in sorted(glob.glob(os.path.join(MAESTRO_DIR, "informe_completo_recetas*.csv"))):
        with open(path, encoding="latin-1", newline="") as fh:
            for row in csv.DictReader(fh, delimiter=";"):
                receta = (row.get("Número Receta") or "").strip()
                if receta:
                    por_receta.setdefault(receta, []).append(row)
    return por_receta


def _fila_a_dict(filas_receta):
    """Colapsa las filas (una por medicamento) de una misma receta en un
    solo dict con la lista de medicamentos, formato esperado por
    gt_maestro.detectar_alertas_mismo_rut()."""
    r0 = filas_receta[0]
    paciente = f"{r0.get('Nombre','').strip()} {r0.get('Apellido Paterno','').strip()} {r0.get('Apellido Materno','').strip()}".strip()
    return {
        "receta": (r0.get("Número Receta") or "").strip(),
        "rut": (r0.get("RUN") or "").strip(),
        "paciente": paciente,
        "periodo": (r0.get("Periodo") or "").strip(),
        "especialidad": (r0.get("Especialidad") or "").strip(),
        "tipo_receta": (r0.get("Tipo Receta") or "").strip(),
        "estado": (r0.get("Estado") or "").strip(),
        # OJO: la columna "Fecha Digitación" del CSV es la fecha de la serie
        # completa, NO por cuota — dos cuotas mensuales normales de una misma
        # receta crónica comparten el mismo valor. Si se pasa a
        # detectar_alertas_mismo_rut() dispara Regla 3 (duplicado informático)
        # en CUALQUIER serie crónica normal, no solo en duplicados reales.
        # Se omite a propósito (sin este dato, Regla 3 no dispara — comportamiento
        # documentado en gt_maestro.detectar_alertas_mismo_rut).
        "medicamentos": [(f.get("Prescripción") or "").strip() for f in filas_receta if f.get("Prescripción")],
    }


def _norm_rut(s):
    """CSV trae 'NNNNNNNN - X' y las planillas de solicitud traen
    'N.NNN.NNN-X' — normaliza ambos a solo dígitos+dígito verificador para
    poder cruzarlos."""
    return re.sub(r"[.\s-]", "", str(s or "")).upper()


def _norm_texto(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().upper()


def _elegir_receta(todas, especialidad_pedida=None):
    """Elige qué receta imprimir entre TODAS las vigentes de un RUT. Bug real
    detectado 25-07-2026 (caso Toltén, 8 de 37 pacientes con receta de
    especialidad equivocada): elegir a ciegas la de N° más alto ignora tanto
    el Tipo Receta (a veces la más nueva es CONTROLADA — formulario ISP
    aparte, ver recetas_cheque.py — cuando lo que se necesita es la NORMAL)
    como la Especialidad (un paciente crónico puede tener recetas vigentes de
    varias especialidades a la vez; Tolten pide una en concreto).

    Prioridad: 1) Tipo Receta = NORMAL sobre CONTROLADA (salvo que sea la
    única opción), 2) especialidad pedida si se informó, 3) N° más alto como
    desempate. Devuelve (receta_elegida, nota_o_None) — la nota se llena
    cuando ninguna candidata coincide con la especialidad pedida, para que
    quede visible en el feedback en vez de fallar en silencio."""
    normales = [r for r in todas if _norm_texto(r.get("tipo_receta")) != "CONTROLADA"]
    pool = normales or todas

    if especialidad_pedida:
        esp_pedida = _norm_texto(especialidad_pedida)
        match = [r for r in pool
                 if esp_pedida in _norm_texto(r.get("especialidad")) or _norm_texto(r.get("especialidad")) in esp_pedida]
        if match:
            elegida = max(match, key=lambda r: int(r["receta"]) if r["receta"].isdigit() else -1)
            return elegida, None
        elegida = max(pool, key=lambda r: int(r["receta"]) if r["receta"].isdigit() else -1)
        nota = (f"Sin receta vigente de {especialidad_pedida} en el sistema — se adjunta la más reciente "
                f"disponible ({elegida.get('especialidad') or 'especialidad desconocida'}, "
                f"{elegida.get('estado') or 'estado desconocido'}). Revisar manualmente.")
        return elegida, nota

    elegida = max(pool, key=lambda r: int(r["receta"]) if r["receta"].isdigit() else -1)
    return elegida, None


def _indice_por_rut(por_receta):
    """RUT normalizado -> lista de dicts de receta (una vez por Nº de receta,
    sin repetir por cada línea de medicamento)."""
    indice = {}
    for filas in por_receta.values():
        rut_norm = _norm_rut(filas[0].get("RUN"))
        if rut_norm:
            indice.setdefault(rut_norm, []).append(_fila_a_dict(filas))
    return indice


def _todas_las_recetas_del_rut(rut, indice_rut):
    return indice_rut.get(_norm_rut(rut), [])


def _carpeta_salida(base_dir, fecha):
    """<Revisión de Solicitudes>/<MES AÑO>/<DD-MM-YYYY>/ — mismo esquema de
    dos niveles (mes -> fecha de extracción) que Nóminas de Envío."""
    carpeta_mes = f"{GM.MESES_ES[fecha.month - 1]} {fecha.year}"
    carpeta = os.path.join(base_dir, carpeta_mes, fecha.strftime("%d-%m-%Y"))
    os.makedirs(carpeta, exist_ok=True)
    return carpeta


def _procesar_estab(estab, dry_run):
    carpeta_local = _CARPETA_LOCAL.get(estab.upper())
    if not carpeta_local:
        print(f"[ERROR] '{estab}' no está en el mapeo de establecimientos (ver agregar_gt_manual._CARPETA_LOCAL).")
        return
    base_dir = os.path.join(GT_SOLICITUDES_DIR, carpeta_local, "Revisión de Solicitudes")
    if not os.path.isdir(base_dir):
        print(f"[ERROR] No existe {base_dir}")
        return

    sueltos = [f for f in sorted(os.listdir(base_dir))
               if os.path.isfile(os.path.join(base_dir, f)) and NOMBRE_PDF_RE.match(f)]
    if not sueltos:
        print(f"[{estab}] No hay PDF sueltos por procesar en {base_dir}")
        return

    print(f"[{estab}] {len(sueltos)} PDF suelto(s) encontrado(s)")
    por_receta = _cargar_recetas_csv()
    indice_rut = _indice_por_rut(por_receta)

    recetas_en_pdfs = []
    sin_match = []
    for nombre in sueltos:
        m = NOMBRE_PDF_RE.match(nombre)
        n_receta = m.group(2)
        if n_receta in por_receta:
            recetas_en_pdfs.append((nombre, _fila_a_dict(por_receta[n_receta])))
        else:
            sin_match.append(nombre)
            print(f"  [AVISO] {nombre}: receta {n_receta} no encontrada en informe_completo_recetas*.csv")

    ruts_vistos = {}
    set_completo = []
    for nombre, receta in recetas_en_pdfs:
        rut = receta["rut"]
        if rut and rut not in ruts_vistos:
            todas = _todas_las_recetas_del_rut(rut, indice_rut)
            ruts_vistos[rut] = todas
            set_completo.extend(todas)
            if len(todas) > 1:
                print(f"  [{receta['paciente']}] {len(todas)} recetas vigentes en el hospital — se revisan todas juntas.")

    alertas = GM.detectar_alertas_mismo_rut(set_completo)
    alerta_por_receta = {}
    for al in alertas:
        for r in al["recetas"]:
            alerta_por_receta.setdefault(r, []).append(al["nota"])

    fecha_hoy = datetime.date.today()
    fecha_str = fecha_hoy.strftime("%Y-%m-%d")

    filas_feedback = []
    recetas_con_pdf = {r["receta"] for _, r in recetas_en_pdfs}
    for nombre, receta in recetas_en_pdfs:
        filas_feedback.append((receta, "Solicitud recibida (PDF adjunto)"))
    for rut, todas in ruts_vistos.items():
        for receta in todas:
            if receta["receta"] not in recetas_con_pdf:
                filas_feedback.append((receta, "Adicional — misma paciente, sin PDF de solicitud"))
    for nombre in sin_match:
        m = NOMBRE_PDF_RE.match(nombre)
        filas_feedback.append(({
            "paciente": "(no encontrado en CSV — completar a mano)", "rut": "",
            "receta": m.group(2), "especialidad": "", "periodo": "",
        }, f"PDF adjunto ({nombre}) sin match en informe_completo_recetas — revisar manualmente"))

    print(f"  {len(alertas)} alerta(s) de posible duplicado/ambigüedad detectada(s).")

    if dry_run:
        print(f"  [DRY-RUN] Generaría Recetas_Combinadas_{carpeta_local}_{fecha_str}.pdf "
              f"({len(recetas_en_pdfs) + len(sin_match)} PDF) y Feedback_Solicitud_{carpeta_local}_{fecha_str}.xlsx "
              f"({len(filas_feedback) + len(sin_match)} fila(s)). No se mueve nada.")
        return

    salida_dir = _carpeta_salida(base_dir, fecha_hoy)

    combinado = os.path.join(salida_dir, f"Recetas_Combinadas_{carpeta_local}_{fecha_str}.pdf")
    writer = PdfWriter()
    for nombre, _ in recetas_en_pdfs:
        writer.append(os.path.join(base_dir, nombre))
    for nombre in sin_match:
        writer.append(os.path.join(base_dir, nombre))
    with open(combinado, "wb") as fh:
        writer.write(fh)
    print(f"  Guardado: {combinado}")

    feedback_path = os.path.join(salida_dir, f"Feedback_Solicitud_{carpeta_local}_{fecha_str}.xlsx")
    _escribir_feedback(feedback_path, estab, fecha_hoy, filas_feedback, alerta_por_receta)
    print(f"  Guardado: {feedback_path}")

    pdfs_dir = os.path.join(salida_dir, "PDFs individuales")
    os.makedirs(pdfs_dir, exist_ok=True)
    for nombre in sueltos:
        os.replace(os.path.join(base_dir, nombre), os.path.join(pdfs_dir, nombre))
    print(f"  {len(sueltos)} PDF individual(es) movido(s) a 'PDFs individuales'.")


def _escribir_feedback(path, estab, fecha_hoy, filas_feedback, alerta_por_receta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Feedback solicitud Gestión Territorial — {estab} ({fecha_hoy.strftime('%d-%m-%Y')})"
    ws["A1"].font = Font(bold=True, size=13)
    headers = ["Paciente", "RUT", "N° Receta", "Especialidad", "Período",
               "Observación", "Acción tomada", "Fecha de revisión"]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for receta, origen in filas_feedback:
        obs = "; ".join(alerta_por_receta.get(receta["receta"], [])) or origen
        ws.append([receta["paciente"], receta["rut"], receta["receta"],
                   receta["especialidad"], receta["periodo"], obs, "", ""])
    anchos = [30, 14, 12, 22, 10, 45, 20, 16]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)


def _procesar_estab_xlsx(estab, xlsx_path, dry_run):
    """Solicitud en el formato real que llega de los establecimientos: una
    planilla con columnas FECHA/RUT/NOMBRE (sin Nº de receta). Busca cada RUT
    en el histórico, trae TODAS sus recetas vigentes y corre las mismas
    reglas de alerta. No hay PDFs que combinar — la planilla original recibida
    se archiva en la carpeta del establecimiento como registro."""
    carpeta_local = _CARPETA_LOCAL.get(estab.upper())
    if not carpeta_local:
        print(f"[ERROR] '{estab}' no está en el mapeo de establecimientos (ver agregar_gt_manual._CARPETA_LOCAL).")
        return
    base_dir = os.path.join(GT_SOLICITUDES_DIR, carpeta_local, "Revisión de Solicitudes")
    if not os.path.isdir(base_dir):
        print(f"[ERROR] No existe {base_dir}")
        return
    if not os.path.isfile(xlsx_path):
        print(f"[ERROR] No existe {xlsx_path}")
        return

    wb_in = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_in = wb_in.active
    header = [str(c).strip().upper() if c else "" for c in next(ws_in.iter_rows(max_row=1, values_only=True))]
    try:
        col_rut = header.index("RUT")
    except ValueError:
        print(f"[ERROR] No encontré columna 'RUT' en {xlsx_path} (headers: {header})")
        return
    col_nombre = header.index("NOMBRE") if "NOMBRE" in header else None
    col_especialidad = header.index("ESPECIALIDAD") if "ESPECIALIDAD" in header else None

    pacientes = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        rut = row[col_rut] if col_rut < len(row) else None
        if not rut:
            continue
        nombre = row[col_nombre] if col_nombre is not None and col_nombre < len(row) else None
        especialidad = row[col_especialidad] if col_especialidad is not None and col_especialidad < len(row) else None
        pacientes.append((str(rut).strip(), str(nombre).strip() if nombre else "",
                           str(especialidad).strip() if especialidad else ""))

    if not pacientes:
        print(f"[{estab}] La planilla no tiene filas con RUT.")
        return
    print(f"[{estab}] {len(pacientes)} paciente(s) en la solicitud recibida ({os.path.basename(xlsx_path)})")

    por_receta = _cargar_recetas_csv()
    indice_rut = _indice_por_rut(por_receta)

    set_completo = []
    filas_feedback = []
    for rut, nombre, especialidad_pedida in pacientes:
        todas = _todas_las_recetas_del_rut(rut, indice_rut)
        if not todas:
            filas_feedback.append(({
                "paciente": nombre or "(sin nombre en la solicitud)", "rut": rut,
                "receta": f"SIN_RECETA_{rut}", "especialidad": "", "periodo": "",
            }, "No se encontraron recetas vigentes en el sistema para este RUT — revisar manualmente"))
            continue
        set_completo.extend(todas)
        # El combinado a imprimir/despachar lleva SOLO UNA receta por
        # paciente — no todas las vigentes. Las demás se revisan igual
        # (arriba, para alertas de duplicado/ambigüedad) pero no se imprimen
        # todas, para no inflar el combinado con recetas que no corresponden
        # a esta solicitud puntual (corregido 23-07-2026 tras detectarlo en
        # Freire: 26 pacientes -> 47 páginas por incluir todas las vigentes).
        #
        # OJO — "la de N° más alto" a secas NO basta (bug real 25-07-2026,
        # caso Toltén: 8 de 37 pacientes con receta de especialidad
        # equivocada porque un paciente crónico puede tener recetas vigentes
        # de VARIAS especialidades a la vez, y a veces la más nueva es
        # CONTROLADA cuando se necesita la NORMAL). _elegir_receta() filtra
        # por Tipo Receta=NORMAL y por especialidad pedida (columna
        # ESPECIALIDAD opcional en la planilla de entrada) antes de usar el
        # N° más alto como desempate.
        a_imprimir, nota_mismatch = _elegir_receta(todas, especialidad_pedida)
        notas = []
        if len(todas) > 1:
            otras = [r["receta"] for r in todas if r is not a_imprimir]
            print(f"  [{todas[0]['paciente']}] {len(todas)} recetas vigentes — se revisan todas, "
                  f"se imprime solo {a_imprimir['receta']} ({a_imprimir.get('especialidad')}); "
                  f"otras: {', '.join(otras)}.")
            notas.append(f"Paciente tiene otra(s) receta(s) vigente(s) no incluida(s) en el combinado: {', '.join(otras)}")
        if nota_mismatch:
            print(f"  [AVISO] {a_imprimir['paciente']}: {nota_mismatch}")
            notas.append(nota_mismatch)
        filas_feedback.append((a_imprimir, "Solicitud recibida (planilla establecimiento)"
                                + (" — " + " | ".join(notas) if notas else "")))

    alertas = GM.detectar_alertas_mismo_rut(set_completo)
    alerta_por_receta = {}
    for al in alertas:
        for r in al["recetas"]:
            alerta_por_receta.setdefault(r, []).append(al["nota"])
    print(f"  {len(alertas)} alerta(s) de posible duplicado/ambigüedad detectada(s).")

    fecha_hoy = datetime.date.today()
    fecha_str = fecha_hoy.strftime("%Y-%m-%d")
    archivo_original = os.path.basename(xlsx_path)

    if dry_run:
        salida_prevista = os.path.join(base_dir, f"{GM.MESES_ES[fecha_hoy.month - 1]} {fecha_hoy.year}", fecha_hoy.strftime("%d-%m-%Y"))
        print(f"  [DRY-RUN] Generaría Feedback_Solicitud_{carpeta_local}_{fecha_str}.xlsx "
              f"({len(filas_feedback)} fila(s)) y archivaría '{archivo_original}' en {salida_prevista}. No se mueve nada.")
        return

    salida_dir = _carpeta_salida(base_dir, fecha_hoy)

    feedback_path = os.path.join(salida_dir, f"Feedback_Solicitud_{carpeta_local}_{fecha_str}.xlsx")
    _escribir_feedback(feedback_path, estab, fecha_hoy, filas_feedback, alerta_por_receta)
    print(f"  Guardado: {feedback_path}")

    import shutil
    destino_original = os.path.join(salida_dir, archivo_original)
    if os.path.abspath(destino_original) != os.path.abspath(xlsx_path):
        shutil.copy2(xlsx_path, destino_original)
        print(f"  Solicitud original archivada: {destino_original}")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--estab", required=True, help='Establecimiento, ej. "HOSPITAL TOLTEN"')
    ap.add_argument("--xlsx", help="Planilla de solicitud recibida (columnas FECHA/RUT/NOMBRE), en vez de PDFs sueltos")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    if a.xlsx:
        _procesar_estab_xlsx(a.estab, a.xlsx, a.dry_run)
    else:
        _procesar_estab(a.estab, a.dry_run)


if __name__ == "__main__":
    main()
