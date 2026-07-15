#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sgli_historico.py — Planilla SGLI basada en historial real de prescripciones.

Metodología ABC-XYZ (referenciada en literatura de gestión de inventario):
  · ABC — Pareto acumulado sobre CDL_pico semanal (NOT umbral fijo).
    Ordenar meds por CDL pico descendente → acumular % del total CDL_pico:
        A: primeros ítems hasta el 70% acumulado  (~10-15% de ítems)
        B: siguiente tramo hasta el 90%            (~20-30% de ítems)
        C: restantes                               (~55-65% de ítems)
    Evita que un umbral fijo (ej. 30 ud/día) meta el 40% de los meds en A
    cuando la distribución es muy asimétrica (Paracetamol: 5317 ud/día).
    Fuente: Silver, Pyke & Thomas "Inventory Mgmt" §3.2
  · XYZ por Coeficiente de Variación semanal (CV = σ/μ):
        X: CV < 0.5 (demanda estable)
        Y: 0.5 ≤ CV < 1.0 (variabilidad moderada)
        Z: CV ≥ 1.0 (demanda errática / estacional)

Índice de reposición (días extra de CDL como stock de seguridad por grupo):
    AX=1.0  AY=1.5  AZ=2.0  BX=1.25  BY=1.75  BZ=2.5  CX=1.5  CY=2.0  CZ=3.0

Fórmula principal:
    STOCK_A_PEDIR = CDL_hist × DIAS_cobertura + CDL_hist × Indice_Repo
    NECESIDAD     = max(0, STOCK_A_PEDIR − Stock_Farmacia)

DIAS_cobertura según día hábil del reporte: Lunes=5, Martes=4, Miércoles=3, Jueves=2, Viernes=1

Fuentes metodológicas:
  · ABC XYZ analysis — EazyStock (eazystock.com/blog/abc-xyz-analysis)
  · Safety Stock Calculation Guide — supplychainmath.com/en/safety-stock-guide.html
  · ISM: Mastering Safety Stock Calculations (ism.ws/logistics/)
  · PBA Health: 6 Inventory Management Formulas for Pharmacies (pbahealth.com)
"""

import glob
import math
import os
import re
import sys
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, WORK_DIR)
from utils_aa import norm_erp, HOMOLOGACION, setup_stdout

setup_stdout()

# ─── Parámetros ────────────────────────────────────────────────────────────────

# Clasificación ABC — Pareto acumulado sobre CDL_pico semanal
# (ordenar por CDL_pico desc → % acumulado del total CDL_pico)
# Fuente: Silver, Pyke & Thomas "Inventory Mgmt & Production Planning" cap. 3
#         EazyStock ABC analysis — eazystock.com/blog/abc-xyz-analysis/
ABC_PARETO_A = 70.0   # % acumulado hasta grupo A (~10-15% de ítems)
ABC_PARETO_B = 90.0   # % acumulado hasta grupo B (~25-35% de ítems)
#                       resto → C (~50-65% de ítems)

# CV para clasificación XYZ
CV_X_MAX = 0.5   # X: estable
CV_Y_MAX = 1.0   # Y: moderado | ≥1.0 = Z

# Índice de reposición (días de CDL extra = stock de seguridad)
# Fuente: ABC-XYZ inventory matrix, EazyStock / ISM best-practice
INDICE_REPO: dict[str, float] = {
    "AX": 1.00, "AY": 1.50, "AZ": 2.00,
    "BX": 1.25, "BY": 1.75, "BZ": 2.50,
    "CX": 1.50, "CY": 2.00, "CZ": 3.00,
}

# Días de cobertura Farmacia AA → Bodega AA (ciclo semanal)
# Lun=5 (cubre toda la semana), Mar=4, Mié=3, Jue=2, Vie=1
# El lunes se pide para los 5 días hábiles de la semana; cada día siguiente se descuenta 1.
# Sáb/Dom: equivale a lunes (prepara el ciclo de la próxima semana).
DIAS_REPORTE = {0: 5, 1: 4, 2: 3, 3: 2, 4: 1, 5: 5, 6: 5}

# Días de cobertura Bodega AA → Bodega de Fármacos (ciclo quincenal = 10 días hábiles / 2 semanas)
# El ciclo parte siempre el lunes (Lun=10 días completos, Mar=9, ..., Vie=6).
# Sáb/Dom: prepara el lunes siguiente con 10 días.
DIAS_CICLO_BODEGA = {0: 10, 1: 9, 2: 8, 3: 7, 4: 6, 5: 10, 6: 10}

# Umbral mensual (22 d.h.) para sugerir fuente de diálisis
DIALISIS_UMBRAL_BODEGA = 50   # Req mensual ≥ 50 ud → Bodega AA; < 50 → Farmacia AA

# ─── Factor de empaque CENABAST ───────────────────────────────────────────────
# Igual lógica que maestro_aa.py; fuente: cenabast_intermediacion.csv
_FORMAS_EMP = {
    'CM','CP','COMPRIMIDO','COMPRIMIDOS','COM','C','CAPSULA','CAPSULAS','CAPS','CAP',
    'REC','ENT','UD','GR','G','TAB','TABLETA','SOL','INY','AM','FAM','FA','FAMP','FRA',
    'FCO','FC','POMO','SUSP','JBE','JRP','CREMA','GEL','UN','SOBRE','SOB','LIOF','P',
    'INYECTABLE','ORAL','TOPICO','OFTALMICO','NEB','CAJ','CJ','AMP','PERF','IV',
}

def _clave_empaque(nombre: str) -> str:
    s = norm_erp(nombre).replace('/', ' ').replace('-', ' ')
    m = re.search(r'(\d+[.,]?\d*)\s*(MG|MCG|UI|G|ML|%)', s)
    conc = (m.group(1).replace(',', '.') + m.group(2)) if m else ''
    pre  = s[:m.start()] if m else s
    toks = [t for t in re.split(r'[ .,]+', pre)
            if t and t not in _FORMAS_EMP and not any(ch.isdigit() for ch in t)]
    return (' '.join(toks[:3]) + ' ' + conc).strip()

def _cargar_factores_empaque() -> dict:
    files = glob.glob(os.path.join(WORK_DIR, 'cenabast_intermediacion.csv'))
    if not files:
        return {}
    try:
        df = pd.read_csv(files[0], encoding='latin1', sep=';', skiprows=3, dtype=str)
    except Exception:
        return {}
    if 'NOMBRE COMERCIAL DEL PRODUCTO' not in df.columns or 'NOMBRE GENERICO' not in df.columns:
        return {}
    def _pack(nombre):
        if not isinstance(nombre, str): return None
        s = nombre.upper()
        m = re.search(r'CAJ\s*(\d+)\s*[A-Z]', s) or re.search(r'\bX\s*(\d+)\b', s)
        return int(m.group(1)) if m else None
    out: dict = {}
    for _, r in df.iterrows():
        fac = _pack(r['NOMBRE COMERCIAL DEL PRODUCTO'])
        if not fac or fac <= 1:
            continue
        k = _clave_empaque(r['NOMBRE GENERICO'])
        if k and k not in out:
            out[k] = fac
    return out

def _redondear_empaque(cantidad: float, medicamento: str) -> int:
    """Redondea HACIA ARRIBA al múltiplo del factor de empaque ICP CENABAST."""
    c = float(cantidad)
    if c <= 0:
        return 0
    fac = FACTORES_EMPAQUE.get(_clave_empaque(medicamento))
    if not fac or fac <= 1:
        return int(math.ceil(c))
    return int(math.ceil(c / fac) * fac)

FACTORES_EMPAQUE: dict = _cargar_factores_empaque()

# ─── Estilos Excel ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _side() -> Side:
    return Side(style="thin", color="BFBFBF")

def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

FILL_HEADER = _fill("2F5496")
FILL_A_ROW  = _fill("E2EFDA")   # verde muy claro
FILL_B_ROW  = _fill("FFF2CC")   # amarillo muy claro
FILL_C_ROW  = _fill("FCE4D6")   # salmón muy claro
FILL_ALT    = _fill("F7F7F7")   # gris alternado
FILL_ALERTA = _fill("FFD7D7")   # rojo claro para necesidad crítica

FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORM   = Font(size=10)
FONT_RED    = Font(bold=True, color="C00000", size=10)
FONT_OK     = Font(color="375623", size=10)


# ─── Carga de datos ────────────────────────────────────────────────────────────

def _cargar_todos_csv() -> pd.DataFrame:
    """Carga todos los CSV de recetas (root + _csv_bak), filtra ENTREGADO."""
    cols_usar = [
        "Prescripci\xf3n",           # Prescripción
        "Cantidad Entregada",
        "Fecha Entrega Receta",
        "Estado Prescripci\xf3n",    # Estado Prescripción
        "ID Receta Detalle",
    ]
    files = sorted(
        glob.glob(os.path.join(WORK_DIR, "informe_completo_recetas*.csv"))
        + glob.glob(os.path.join(WORK_DIR, "_csv_bak", "informe_completo_recetas*.csv"))
    )
    if not files:
        raise FileNotFoundError(
            "No hay archivos informe_completo_recetas*.csv.\n"
            "Ejecuta AUTO_SSASUR.bat primero."
        )
    chunks = []
    for f in files:
        try:
            df = pd.read_csv(
                f, sep=";", encoding="latin1", dtype=str,
                on_bad_lines="skip",
                usecols=lambda c, _c=cols_usar: c in _c,
            )
            chunks.append(df)
        except Exception as e:
            print(f"  [WARN] {os.path.basename(f)}: {e}")
    if not chunks:
        raise RuntimeError("No se pudo leer ningún CSV.")

    big = (
        pd.concat(chunks, ignore_index=True)
        .drop_duplicates(subset=["ID Receta Detalle"], keep="first")
    )

    # Filtrar ENTREGADO
    col_estado = "Estado Prescripci\xf3n"
    mask = big[col_estado].str.strip().str.upper() == "ENTREGADO"
    big = big[mask].copy()

    big["Medicamento"] = big["Prescripci\xf3n"].apply(lambda x: HOMOLOGACION.get(norm_erp(str(x)), norm_erp(str(x))))
    big["Unidades"] = (
        pd.to_numeric(
            big["Cantidad Entregada"].astype(str).str.replace(",", "."),
            errors="coerce",
        ).fillna(0)
    )
    big["Fecha"] = pd.to_datetime(
        big["Fecha Entrega Receta"], dayfirst=True, errors="coerce"
    )
    big = big.dropna(subset=["Fecha", "Medicamento"])
    big = big[big["Unidades"] > 0]
    return big[["Medicamento", "Fecha", "Unidades"]]


def _cargar_stock() -> pd.DataFrame:
    """Lee stock actual desde SGLI_Estres o Stock_AA del Consolidado."""
    path = os.path.join(WORK_DIR, "Consolidado_AA_MAESTRO.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame()
    xl = pd.ExcelFile(path, engine="openpyxl")
    if "SGLI_Estres" in xl.sheet_names:
        df = xl.parse("SGLI_Estres")
        keep = [c for c in ["Medicamento", "CDL", "Stock_Farm", "Stock_Bod", "Criticidad"] if c in df.columns]
        return df[keep].copy()
    if "Stock_AA" in xl.sheet_names:
        df = xl.parse("Stock_AA")
        df = df.rename(columns={"Stock_Farmacia_AA": "Stock_Farm", "Stock_Bodega_AA": "Stock_Bod"})
        keep = [c for c in ["Medicamento", "CDL", "Stock_Farm", "Stock_Bod", "Criticidad"] if c in df.columns]
        return df[keep].copy()
    return pd.DataFrame()


# ─── Análisis ABC-XYZ ──────────────────────────────────────────────────────────

def _dias_habiles_semana(semana_period, feriados: set) -> int:
    """Días hábiles de lunes a viernes en la semana (descontando feriados)."""
    lunes = semana_period.start_time.date()
    dias = 0
    for d in range(5):
        dia = lunes + pd.Timedelta(days=d)
        if dia not in feriados:
            dias += 1
    return max(1, dias)


def _cargar_feriados() -> set:
    path = os.path.join(WORK_DIR, "feriados_chile.csv")
    if not os.path.exists(path):
        return set()
    df = pd.read_csv(path, sep=";", encoding="latin1", dtype=str)
    col = df.columns[0]
    fechas = pd.to_datetime(df[col], dayfirst=True, errors="coerce").dropna()
    return set(fechas.dt.date)


def calcular_abc_xyz(df_rec: pd.DataFrame, feriados: set) -> pd.DataFrame:
    """
    Agrupa por semana ISO, calcula CDL semanal y estadísticas,
    clasifica ABC (Pareto acumulado sobre CDL_pico) y XYZ (CV = σ/μ).

    ABC — Método Pareto sobre CDL pico (semana de mayor demanda):
      Ordenar medicamentos por CDL_pico descendente.
      Acumular su contribución al total de CDL_pico de todos los meds.
      A: primeros ítems hasta alcanzar ABC_PARETO_A% del acumulado (~10-15% de meds)
      B: siguientes hasta ABC_PARETO_B%  (~20-30% de meds)
      C: restantes                         (~55-65% de meds)
    → Evita que un umbral fijo clasifique el 40% como A cuando la distribución
      es muy asimétrica (como en esta farmacia: PARACETAMOL=5317 vs mediana~14).

    Fuente: Silver, Pyke & Thomas "Inventory Mgmt and Production Planning" §3.2;
            EazyStock ABC analysis (eazystock.com/blog/abc-xyz-analysis/)
    """
    df = df_rec.copy()
    df["Semana"] = df["Fecha"].dt.to_period("W-SUN")

    # CDL por semana = unidades / días hábiles de esa semana
    semanal = df.groupby(["Medicamento", "Semana"])["Unidades"].sum().reset_index()
    semanal["DH"] = semanal["Semana"].apply(
        lambda s: _dias_habiles_semana(s, feriados)
    )
    semanal["CDL_sem"] = semanal["Unidades"] / semanal["DH"]

    # Estadísticas por medicamento
    stats = (
        semanal.groupby("Medicamento")["CDL_sem"]
        .agg(CDL_Prom="mean", CDL_Std="std", N_Semanas="count", CDL_Pico="max")
        .reset_index()
    )
    stats["CDL_Std"] = stats["CDL_Std"].fillna(0)
    stats["CV"] = np.where(
        stats["CDL_Prom"] > 0, stats["CDL_Std"] / stats["CDL_Prom"], 0.0
    )

    # Total dispensado (solo informativo)
    total = df.groupby("Medicamento")["Unidades"].sum().reset_index(name="Total_Unidades")
    stats = stats.merge(total, on="Medicamento")

    # ── Clasificación ABC — Pareto acumulado sobre CDL_pico ─────────────────
    # Ordenar por CDL_pico desc → acumular % del total → cortar en 70/90/100
    stats_s = stats.sort_values("CDL_Pico", ascending=False).copy()
    total_pico = stats_s["CDL_Pico"].sum()
    if total_pico > 0:
        stats_s["_acum_pct"] = stats_s["CDL_Pico"].cumsum() / total_pico * 100
    else:
        stats_s["_acum_pct"] = 100.0

    # La primera fila puede ya superar ABC_PARETO_A si un ítem domina mucho
    # → pd.cut con include_lowest para asignar A al primer ítem siempre
    stats_s["ABC"] = pd.cut(
        stats_s["_acum_pct"],
        bins=[-0.001, ABC_PARETO_A, ABC_PARETO_B, 100.0],
        labels=["A", "B", "C"],
    ).astype(str)

    # Calcular umbrales reales de CDL_pico para informar en el reporte
    limite_a = stats_s.loc[stats_s["ABC"] == "A", "CDL_Pico"].min()
    limite_b = stats_s.loc[stats_s["ABC"] == "B", "CDL_Pico"].min()
    n_a = (stats_s["ABC"] == "A").sum()
    n_b = (stats_s["ABC"] == "B").sum()
    n_c = (stats_s["ABC"] == "C").sum()
    print(f"      ABC Pareto 70/90/100: A={n_a} meds (CDL pico ≥{limite_a:.0f}), "
          f"B={n_b} (≥{limite_b:.0f}), C={n_c} (<{limite_b:.0f})")

    # Volver al orden original para merge
    stats = stats.merge(
        stats_s[["Medicamento", "ABC", "_acum_pct"]].rename(columns={"_acum_pct": "Acum_Pct_CDLPico"}),
        on="Medicamento", how="left"
    )

    # ── Clasificación XYZ por CV ─────────────────────────────────────────────
    def _xyz(cv: float) -> str:
        if cv < CV_X_MAX:
            return "X"
        if cv < CV_Y_MAX:
            return "Y"
        return "Z"

    stats["XYZ"] = stats["CV"].apply(_xyz)
    stats["Grupo_ABCXYZ"] = stats["ABC"] + stats["XYZ"]

    # ── Índice de reposición (días de CDL) ───────────────────────────────────
    stats["Indice_Repo"] = stats["Grupo_ABCXYZ"].map(INDICE_REPO).fillna(1.5)

    return stats


# ─── Generación de planilla ────────────────────────────────────────────────────

def _celda_accion(necesidad: int, stock_bod: int, critico: bool) -> str:
    if necesidad <= 0:
        return "✓ STOCK SUFICIENTE"
    if critico and necesidad > 0:
        prefijo = "⚠ CRÍTICO — "
    else:
        prefijo = ""
    if stock_bod >= necesidad:
        return f"{prefijo}TRASPASAR {necesidad} DESDE BODEGA"
    if stock_bod > 0:
        diff = necesidad - stock_bod
        return f"{prefijo}TRASPASAR {stock_bod} + COMPRA {diff}"
    return f"{prefijo}COMPRA URGENTE {necesidad}"


def _hoja_titulo(ws, texto: str, color_hex: str, n_cols: int, fila: int = 1) -> None:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = _fill(color_hex)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 20


def _hoja_cabecera(ws, headers: list[str], fila: int = 2) -> None:
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=fila, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[fila].height = 30


def _agregar_hoja_pedido_farm_bodega(wb: Workbook, df_final: pd.DataFrame, hoy: date) -> None:
    """Hoja resumen: lo que Farmacia AA solicita a Bodega AA (traspasos internos)."""
    ws = wb.create_sheet("Pedido_Farm→Bodega")

    df = df_final[df_final["Necesidad"] > 0].copy()
    # Redondear necesidad al empaque ICP CENABAST antes de calcular traspaso
    df["Necesidad"] = df.apply(
        lambda r: _redondear_empaque(float(r["Necesidad"]), str(r["Medicamento"])), axis=1
    )
    df["Factor_Emp"] = df["Medicamento"].apply(
        lambda m: FACTORES_EMPAQUE.get(_clave_empaque(str(m)), 1)
    )
    # Recalcular Accion con la Necesidad ya redondeada al empaque — si no, el texto
    # de Accion (cantidad cruda) queda desincronizado de la columna Necesidad mostrada.
    df["Accion"] = df.apply(
        lambda r: _celda_accion(
            int(r["Necesidad"]), int(r["Stock_Bod"]),
            str(r["Criticidad"]).startswith("1-")
        ),
        axis=1,
    )
    df["A_Traspasar"] = df.apply(
        lambda r: min(int(r["Necesidad"]), int(r["Stock_Bod"])), axis=1
    )
    df["Faltante"] = df.apply(
        lambda r: max(0, int(r["Necesidad"]) - int(r["Stock_Bod"])), axis=1
    )

    orden_abc = {"A": 0, "B": 1, "C": 2}
    df["_ord"] = df["ABC"].map(orden_abc).fillna(3)
    df.sort_values(["_ord", "Necesidad"], ascending=[True, False], inplace=True)
    df.drop(columns=["_ord"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    HEADERS = ["Medicamento", "Criticidad", "ABC-XYZ", "CDL\n(ud/día)", "Stock\nFarmacia",
               "Stock\nBodega AA", "Necesidad\n(redondeada\nempaque)", "A Traspasar\nDesde Bodega",
               "Faltante Post-\nTraspaso", "Factor\nEmpaque\nICP", "Acción"]
    COLS   = ["Medicamento", "Criticidad", "Grupo_ABCXYZ", "CDL_Prom", "Stock_Farm",
              "Stock_Bod", "Necesidad", "A_Traspasar", "Faltante", "Factor_Emp", "Accion"]
    N = len(HEADERS)

    _hoja_titulo(ws, f"PEDIDO FARMACIA AA → BODEGA AA  |  {hoy.strftime('%d/%m/%Y')}  |  "
                 f"{len(df)} medicamentos con necesidad de reposición", "1F3864", N)
    _hoja_cabecera(ws, HEADERS)

    FILL_ABC = {"A": FILL_A_ROW, "B": FILL_B_ROW, "C": FILL_C_ROW}
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        fill = FILL_ABC.get(str(row.get("ABC", "")), FILL_ALT)
        for ci, col in enumerate(COLS, 1):
            v = row.get(col, "")
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                v = 0 if col not in ("Accion", "Grupo_ABCXYZ") else ""
            if col in ("CDL_Prom",):
                v = round(float(v), 2)
            elif col in ("Stock_Farm", "Stock_Bod", "Necesidad", "A_Traspasar", "Faltante", "Factor_Emp"):
                v = int(v) if v != "" else 0
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = _border()
            c.font = FONT_NORM
            c.fill = fill
            c.alignment = Alignment(
                horizontal="left" if col in ("Medicamento", "Accion") else "center",
                wrap_text=(col == "Accion"),
            )
            if col == "Faltante" and isinstance(v, int) and v > 0:
                c.font = FONT_RED
            if col == "A_Traspasar" and isinstance(v, int) and v > 0:
                c.font = Font(bold=True, color="375623", size=10)

    anchos = [42, 12, 9, 9, 9, 9, 10, 12, 12, 8, 38]
    for ci, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(N)}2"

    # Totales al pie
    pie_row = len(df) + 3
    ws.cell(row=pie_row, column=1, value="TOTALES").font = FONT_BOLD
    for ci, col in enumerate(COLS, 1):
        if col in ("Necesidad", "A_Traspasar", "Faltante"):
            val = int(df[col].sum())
            c = ws.cell(row=pie_row, column=ci, value=val)
            c.font = FONT_BOLD
            c.border = _border()


def _agregar_hoja_pedido_bod_farmacos(wb: Workbook, df_final: pd.DataFrame,
                                       hoy: date, dias_ciclo: int, nombre_dia: str) -> None:
    """Hoja resumen: lo que Bodega AA debe comprar a Bodega de Fármacos.

    Fórmula: Compra = max(0, CDL × dias_ciclo − (Stock_Farm + Stock_Bod))
    días_ciclo: Lun=10, Mar=9, Mié=8, Jue=7, Vie=6 (ciclo 2 semanas desde lunes)
    """
    ws = wb.create_sheet("Pedido_Bod→BodFarmacos")

    df = df_final.copy()
    df["Stock_Total"] = (df["Stock_Farm"] + df["Stock_Bod"]).apply(int)
    df["Req_Ciclo"]   = (df["CDL_Prom"] * dias_ciclo).apply(math.ceil)
    df["Compra_Nec"]  = df.apply(
        lambda r: _redondear_empaque(
            max(0.0, float(r["Req_Ciclo"]) - float(r["Stock_Total"])),
            str(r["Medicamento"])
        ), axis=1
    )
    df["Factor_Emp"]  = df["Medicamento"].apply(
        lambda m: FACTORES_EMPAQUE.get(_clave_empaque(str(m)), 1)
    )
    df["Cob_Actual"]  = df.apply(
        lambda r: round(r["Stock_Total"] / r["CDL_Prom"], 1) if r["CDL_Prom"] > 0 else 999, axis=1
    )

    df = df[df["Compra_Nec"] > 0].copy()
    orden_abc = {"A": 0, "B": 1, "C": 2}
    df["_ord"] = df["ABC"].map(orden_abc).fillna(3)
    df.sort_values(["_ord", "Compra_Nec"], ascending=[True, False], inplace=True)
    df.drop(columns=["_ord"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    HEADERS = ["Medicamento", "Criticidad", "ABC-XYZ", "CDL\n(ud/día)",
               "Stock\nFarmacia", "Stock\nBodega AA", f"Stock Total\n(Farm+Bod)",
               f"Req {dias_ciclo} días\nhábiles", "Compra\nNecesaria\n(redondeada)",
               "Factor\nEmpaque\nICP", "Cobertura\nActual (días)"]
    COLS   = ["Medicamento", "Criticidad", "Grupo_ABCXYZ", "CDL_Prom",
              "Stock_Farm", "Stock_Bod", "Stock_Total", "Req_Ciclo",
              "Compra_Nec", "Factor_Emp", "Cob_Actual"]
    N = len(HEADERS)

    _hoja_titulo(ws,
        f"PEDIDO BODEGA AA → BODEGA DE FÁRMACOS  |  {hoy.strftime('%d/%m/%Y')} ({nombre_dia})  |  "
        f"Ciclo quincenal: {dias_ciclo} d.h.  |  {len(df)} meds  |  Compra redondeada al empaque ICP CENABAST", "C00000", N)

    nota_row = 2
    ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=N)
    nota = (f"Fórmula: Compra = CDL × {dias_ciclo} días − (Stock Farmacia + Stock Bodega AA)  |  "
            f"Ciclo 10 d.h. desde lunes: Lun=10, Mar=9, Mié=8, Jue=7, Vie=6")
    cn = ws.cell(row=nota_row, column=1, value=nota)
    cn.font = Font(italic=True, size=9, color="595959")
    cn.alignment = Alignment(horizontal="left")
    ws.row_dimensions[nota_row].height = 13

    _hoja_cabecera(ws, HEADERS, fila=3)

    FILL_ABC = {"A": FILL_A_ROW, "B": FILL_B_ROW, "C": FILL_C_ROW}
    for ri, (_, row) in enumerate(df.iterrows(), 4):
        fill = FILL_ABC.get(str(row.get("ABC", "")), FILL_ALT)
        for ci, col in enumerate(COLS, 1):
            v = row.get(col, "")
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                v = 0
            if col == "CDL_Prom":
                v = round(float(v), 2)
            elif col in ("Stock_Farm","Stock_Bod","Stock_Total","Req_Ciclo","Compra_Nec","Factor_Emp"):
                v = int(v) if v != "" else 0
            elif col == "Cob_Actual":
                v = round(float(v), 1) if v != "" else 0
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = _border()
            c.font = FONT_NORM
            c.fill = fill
            c.alignment = Alignment(
                horizontal="left" if col == "Medicamento" else "center"
            )
            if col == "Compra_Nec" and isinstance(v, int) and v > 0:
                c.font = FONT_RED
            if col == "Cob_Actual" and isinstance(v, (int, float)) and v < dias_ciclo:
                c.fill = FILL_ALERTA

    anchos = [42, 12, 9, 9, 9, 10, 10, 10, 11, 8, 12]
    for ci, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(N)}3"

    pie_row = len(df) + 4
    ws.cell(row=pie_row, column=1, value="TOTALES").font = FONT_BOLD
    for ci, col in enumerate(COLS, 1):
        if col in ("Req_Ciclo", "Compra_Nec"):
            val = int(df[col].sum())
            c = ws.cell(row=pie_row, column=ci, value=val)
            c.font = FONT_BOLD
            c.border = _border()


def _agregar_hoja_dialisis(wb: Workbook, hoy: date) -> None:
    """Hoja: requerimiento mensual de diálisis con sugerencia de fuente.

    Carga Dialisis_Medicamentos del Consolidado — universo COMPLETO de
    medicamentos prescritos por nefrólogos (no solo los que necesitan pedido;
    Dialisis_Pedido_Farm/Bod filtran a Necesidad>0 a propósito y por eso
    omitían medicamentos con stock suficiente, ej. Furosemida).
    Requerimiento mensual en DÍAS CORRIDOS (30), no días hábiles — el pedido
    de diálisis es mensual calendario, igual que en pedido_fusion.py calc_h3.
    Fuente: Bodega AA si req mensual ≥ DIALISIS_UMBRAL_BODEGA ud; Farmacia AA si menor.
    """
    ws = wb.create_sheet("Dialisis_Mensual")
    cons_path = os.path.join(WORK_DIR, "Consolidado_AA_MAESTRO.xlsx")

    if not os.path.exists(cons_path):
        ws.cell(row=1, column=1, value="Consolidado_AA_MAESTRO.xlsx no encontrado.")
        return

    xl = pd.ExcelFile(cons_path, engine="openpyxl")
    if "Dialisis_Medicamentos" not in xl.sheet_names:
        ws.cell(row=1, column=1, value="No hay hoja Dialisis_Medicamentos en el Consolidado.")
        return

    df = xl.parse("Dialisis_Medicamentos")
    df = df.drop_duplicates(subset=["Medicamento"], keep="first")

    # CDL diálisis exclusivo
    df["CDL_Dialisis"] = pd.to_numeric(
        df.get("Consumo_5D_Solo_Dialisis", 0), errors="coerce"
    ).fillna(0) / 5
    df = df[df["CDL_Dialisis"] > 0].copy()

    df["Fe"] = pd.to_numeric(df.get("Factor_Empaque", 1), errors="coerce").fillna(1).clip(lower=1).apply(int)
    df["Req_Mensual_30d"] = df.apply(
        lambda r: int(math.ceil(r["CDL_Dialisis"] * 30 / r["Fe"]) * r["Fe"]), axis=1
    )

    # Stock actual (ya viene de Dialisis_Medicamentos con nombres propios)
    for col_map in [("Stock_Farmacia_AA", "Stock_Farm"), ("Stock_Bodega_AA", "Stock_Bod")]:
        src, dst = col_map
        if src in df.columns:
            df[dst] = pd.to_numeric(df[src], errors="coerce").fillna(0).apply(int)
        elif dst not in df.columns:
            df[dst] = 0

    df["Stock_Total"] = df["Stock_Farm"] + df["Stock_Bod"]
    df["Cob_Actual_Dias"] = df.apply(
        lambda r: round(r["Stock_Total"] / r["CDL_Dialisis"], 1) if r["CDL_Dialisis"] > 0 else 999,
        axis=1,
    )

    # CDL Total (medicamento completo, combinado farm no-dial + diálisis)
    if "CDL_Combinado" in df.columns:
        df["CDL_Total"] = pd.to_numeric(df["CDL_Combinado"], errors="coerce").fillna(0).round(2)
    else:
        df["CDL_Total"] = df["CDL_Dialisis"]

    # Req mensual total (30 días corridos × CDL total combinado)
    df["Req_Mensual_Total"] = (df["CDL_Total"] * 30).apply(math.ceil)

    # Fuente sugerida
    def _fuente(req_dial: int, req_total: int) -> tuple[str, str]:
        if req_total >= 500:
            return ("BODEGA AA",
                    f"Alto volumen total ({req_total} ud/mes): gestionar vía Bodega AA en pedido mensual")
        if req_dial >= DIALISIS_UMBRAL_BODEGA:
            return ("BODEGA AA",
                    f"Consumo diálisis ({req_dial} ud/mes) ≥ {DIALISIS_UMBRAL_BODEGA} ud: pedir desde Bodega AA")
        return ("FARMACIA AA",
                f"Consumo diálisis bajo ({req_dial} ud/mes): puede salir directamente de Farmacia AA")

    fuentes = df.apply(
        lambda r: _fuente(int(r["Req_Mensual_30d"]), int(r["Req_Mensual_Total"])),
        axis=1,
    )
    df["Fuente_Sugerida"] = [f[0] for f in fuentes]
    df["Observacion"]     = [f[1] for f in fuentes]

    df.sort_values(["Fuente_Sugerida", "Req_Mensual_30d"], ascending=[True, False], inplace=True)
    df.reset_index(drop=True, inplace=True)

    HEADERS = ["Medicamento", "Fe\n(ud/envase)",
               "Req Mensual\nDiálisis (30d)", "Req Mensual\nTotal (30d)",
               "Stock\nFarmacia", "Stock\nTotal", "Cobertura\nActual (días)",
               "Fuente\nSugerida", "Observación"]
    COLS   = ["Medicamento", "Fe",
              "Req_Mensual_30d", "Req_Mensual_Total",
              "Stock_Farm", "Stock_Total", "Cob_Actual_Dias",
              "Fuente_Sugerida", "Observacion"]

    N = len(HEADERS)
    _hoja_titulo(ws,
        f"REQUERIMIENTO MENSUAL DIÁLISIS (30 días corridos)  |  {hoy.strftime('%d/%m/%Y')}  |  "
        f"{(df['Fuente_Sugerida']=='BODEGA AA').sum()} desde Bodega AA  |  "
        f"{(df['Fuente_Sugerida']=='FARMACIA AA').sum()} desde Farmacia AA", "5B2878", N)

    nota_row = 2
    ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=N)
    nota = (f"Req Mensual = CDL × 30 días corridos (el pedido de diálisis es mensual "
            f"calendario, no días hábiles), redondeado al empaque (Fe)  |  "
            f"Fuente: BODEGA AA si consumo diálisis ≥ {DIALISIS_UMBRAL_BODEGA} ud/mes "
            f"o req total ≥ 500 ud/mes; FARMACIA AA si menor  |  "
            f"Universo completo (incluye meds con stock suficiente, ej. Furosemida)")
    ws.cell(row=nota_row, column=1, value=nota).font = Font(italic=True, size=9, color="595959")
    ws.row_dimensions[nota_row].height = 13

    _hoja_cabecera(ws, HEADERS, fila=3)

    FILL_FARM = _fill("E2EFDA")   # verde = Farmacia AA
    FILL_BOD  = _fill("DEEAF1")   # azul claro = Bodega AA
    for ri, (_, row) in enumerate(df.iterrows(), 4):
        fill = FILL_FARM if row.get("Fuente_Sugerida") == "FARMACIA AA" else FILL_BOD
        for ci, col in enumerate(COLS, 1):
            v = row.get(col, "")
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                v = 0 if col != "Observacion" else ""
            if col in ("Fe", "Req_Mensual_30d", "Req_Mensual_Total", "Stock_Farm", "Stock_Bod", "Stock_Total"):
                v = int(v) if v != "" else 0
            elif col == "Cob_Actual_Dias":
                v = round(float(v), 1) if v != "" else 0
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = _border()
            c.font = FONT_NORM
            c.fill = fill
            c.alignment = Alignment(
                horizontal="left" if col in ("Medicamento", "Observacion", "Fuente_Sugerida") else "center",
                wrap_text=(col == "Observacion"),
            )
            if col == "Fuente_Sugerida":
                c.font = Font(bold=True, size=10,
                              color="375623" if v == "FARMACIA AA" else "17375E")

    anchos = [42, 11, 13, 12, 9, 9, 12, 14, 46]
    for ci, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(N)}3"


def generar_planilla(df_final: pd.DataFrame, hoy: date, dias_cobertura: int,
                     dias_ciclo: int = 10, nombre_dia: str = "") -> str:
    from datetime import datetime
    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    out = os.path.join(WORK_DIR, f"SGLI_Historico_{ts}.xlsx")
    wb  = Workbook()

    # ── Hoja principal ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SGLI_Historico"

    _nombres_dia = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}
    _dia_str = nombre_dia or _nombres_dia.get(hoy.weekday(), "")
    encabezado = (
        f"PLANILLA SGLI — HISTORIAL DE PRESCRIPCIONES | "
        f"{hoy.strftime('%d/%m/%Y')} ({_dia_str}) | "
        f"Cobertura: {dias_cobertura} días"
    )
    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = encabezado
    c.font  = Font(bold=True, size=12, color="FFFFFF")
    c.fill  = _fill("1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Nota metodológica
    ws.merge_cells("A2:S2")
    nota = (
        "ABC: Pareto acumulado CDL pico (A=top 70%, B=70-90%, C=90-100%)  |  "
        "XYZ: X=CV<0.5 (estable), Y=0.5-1.0 (moderado), Z≥1.0 (errático)  |  "
        "STOCK A PEDIR = CDL × Días + CDL × Índice_Repo"
    )
    cn = ws["A2"]
    cn.value = nota
    cn.font  = Font(italic=True, size=9, color="595959")
    cn.alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 14

    # Cabeceras fila 3
    HEADERS = [
        "Medicamento",
        "Criticidad",
        "ABC",
        "XYZ",
        "Grupo\nABC-XYZ",
        "CDL Hist.\n(ud/día)",
        "CDL Pico\n(ud/día)",
        "CV\n(variab.)",
        "N°\nSemanas",
        "Índice\nRepo (días)",
        "Stock\nFarmacia",
        "Stock\nBodega",
        "Stock\nSeguridad",
        f"Cant. a Pedir\n(CDL×{dias_cobertura}d)",
        "Stock\nObjetivo",
        "Necesidad\nReposición",
        "Total\nDispensado",
        "Período\nHistórico",
        "Acción a Realizar",
    ]
    COLS_DATA = [
        "Medicamento", "Criticidad", "ABC", "XYZ", "Grupo_ABCXYZ",
        "CDL_Prom", "CDL_Pico", "CV",
        "N_Semanas", "Indice_Repo",
        "Stock_Farm", "Stock_Bod",
        "Stock_Seguridad", "Cant_A_Pedir", "Stock_Objetivo",
        "Necesidad", "Total_Unidades",
        "Periodo_Hist", "Accion",
    ]
    assert len(HEADERS) == len(COLS_DATA)

    ROW_HDR = 3
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=ROW_HDR, column=ci, value=h)
        c.font      = FONT_HEADER
        c.fill      = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[ROW_HDR].height = 32

    # Filas de datos (desde fila 4)
    FILL_ABC = {"A": FILL_A_ROW, "B": FILL_B_ROW, "C": FILL_C_ROW}
    for ri, (_, row) in enumerate(df_final.iterrows(), ROW_HDR + 1):
        abc_val = str(row.get("ABC", ""))
        fila_par = ri % 2 == 0
        fill_base = FILL_ABC.get(abc_val, FILL_ALT)
        for ci, col in enumerate(COLS_DATA, 1):
            v = row.get(col, "")
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                v = 0 if col not in ("Accion", "Periodo_Hist", "Grupo_ABCXYZ") else ""
            # Formateo por tipo
            if col in ("CDL_Prom", "CDL_Pico"):
                v = round(float(v), 2) if v else 0.0
            elif col == "CV":
                v = round(float(v), 3) if v else 0.0
            elif col == "Indice_Repo":
                v = round(float(v), 2) if v else 0.0
            elif col in ("Stock_Farm", "Stock_Bod", "Stock_Seguridad",
                         "Cant_A_Pedir", "Stock_Objetivo", "Necesidad",
                         "Total_Unidades", "N_Semanas"):
                try:
                    v = int(round(float(v))) if v != "" else 0
                except Exception:
                    v = 0

            c = ws.cell(row=ri, column=ci, value=v)
            c.border    = _border()
            c.font      = FONT_NORM
            c.fill      = fill_base if fila_par else FILL_ALT
            c.alignment = Alignment(horizontal="center" if col not in ("Medicamento","Accion") else "left")

            # Color especial por columna
            if col == "Necesidad":
                nec = row.get("Necesidad", 0)
                try:
                    nec = int(float(nec))
                except Exception:
                    nec = 0
                crit = str(row.get("Criticidad","")).startswith("1-")
                if nec > 0 and crit:
                    c.fill = FILL_ALERTA
                    c.font = FONT_RED
                elif nec > 0:
                    c.font = Font(bold=True, size=10, color="7B2000")
                elif nec == 0:
                    c.font = FONT_OK
            elif col == "Accion":
                c.alignment = Alignment(horizontal="left", wrap_text=True)

    # Anchos de columna
    anchos = [42,14,5,5,9,9,9,8,8,10,9,9,10,12,10,12,12,20,38]
    for ci, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    # ── Hoja Metodología ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Metodología_ABC_XYZ")
    lineas = [
        ("METODOLOGÍA ABC-XYZ PARA PLANILLA SGLI HISTÓRICO", True),
        ("", False),
        ("CLASIFICACIÓN ABC — Pareto acumulado sobre CDL pico semanal", True),
        ("Principio", "Ordenar meds por CDL_pico desc → acumular % del total CDL_pico de todos los meds"),
        ("A", f"Primeros ítems hasta {int(ABC_PARETO_A)}% acumulado  →  ~10-15% de meds  (esenciales críticos de stock)"),
        ("B", f"Siguiente tramo hasta {int(ABC_PARETO_B)}% acumulado  →  ~20-30% de meds  (uso frecuente)"),
        ("C", f"Ítems restantes            →  ~55-65% de meds  (baja rotación / especialidad)"),
        ("Por qué Pareto y no umbral fijo", "Un umbral ≥30 ud/día meta el 40% de los meds en A (distribución muy asimétrica). "
         "El Pareto acumulado se adapta automáticamente a CUALQUIER distribución de demanda."),
        ("", False),
        ("CLASIFICACIÓN XYZ — por Coeficiente de Variación semanal (CV = σ/μ)", True),
        ("X", "CV < 0.5  →  demanda estable y predecible (stock de seguridad mínimo)"),
        ("Y", "0.5 ≤ CV < 1.0  →  variabilidad moderada"),
        ("Z", "CV ≥ 1.0  →  demanda errática o estacional (mayor cobertura requerida)"),
        ("", False),
        ("ÍNDICE DE REPOSICIÓN (días extra de CDL como stock de seguridad)", True),
        ("AX = 1.0 d", "AY = 1.5 d", "AZ = 2.0 d"),
        ("BX = 1.25 d", "BY = 1.75 d", "BZ = 2.5 d"),
        ("CX = 1.5 d", "CY = 2.0 d", "CZ = 3.0 d"),
        ("", False),
        ("FÓRMULA PRINCIPAL", True),
        ("Stock a Pedir = CDL_hist × Días_cobertura + CDL_hist × Índice_Repo",
         "→  (cobertura básica + stock de seguridad proporcional al grupo)"),
        ("Días_cobertura: Lun=1, Mar=2, Mié=3, Jue=4, Vie=5",
         "→  varía según el día en que se ejecuta el reporte"),
        ("Necesidad = max(0, Stock_a_Pedir − Stock_Farmacia)", "→  cantidad real a reponer"),
        ("", False),
        ("FUENTES METODOLÓGICAS", True),
        ("EazyStock", "ABC-XYZ Analysis for Inventory — eazystock.com/blog/abc-xyz-analysis"),
        ("supplychainmath.com", "Safety Stock Calculation Guide — supplychainmath.com/en/safety-stock-guide.html"),
        ("ISM", "Mastering Safety Stock Calculations — ism.ws/logistics/how-to-calculate-safety-stock/"),
        ("PBA Health", "6 Inventory Management Formulas for Pharmacies — pbahealth.com/elements/"),
        ("", False),
        ("INTERPRETACIÓN DE ACCIONES", True),
        ("✓ STOCK SUFICIENTE", "Stock_Farm ≥ Stock_Objetivo → no se requiere acción"),
        ("TRASPASAR N DESDE BODEGA", "Stock_Bod cubre la necesidad → traspaso interno"),
        ("TRASPASAR N + COMPRA M", "Stock_Bod cubre parcialmente → traspaso + compra complementaria"),
        ("⚠ CRÍTICO — COMPRA URGENTE N", "Criticidad 1 sin stock bodega → gestión urgente requerida"),
    ]
    fill_titulo = _fill("2F5496")
    for ri, line in enumerate(lineas, 1):
        if isinstance(line, tuple) and len(line) == 2 and line[1] is True:
            c = ws2.cell(row=ri, column=1, value=line[0])
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = fill_titulo
        elif isinstance(line, tuple) and len(line) == 2:
            ws2.cell(row=ri, column=1, value=line[0]).font = Font(bold=True, size=10)
            ws2.cell(row=ri, column=2, value=line[1]).font = Font(size=10)
        elif isinstance(line, tuple) and len(line) == 3:
            for ci, v in enumerate(line, 1):
                ws2.cell(row=ri, column=ci, value=v).font = Font(bold=True, size=10, color="1F3864")
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 70
    ws2.column_dimensions["C"].width = 20

    # ── Hoja Top Prescripciones ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Top_Prescripciones")
    top50 = df_final.nlargest(50, "Total_Unidades")
    ws3.merge_cells("A1:G1")
    c3 = ws3["A1"]
    c3.value = f"TOP 50 MEDICAMENTOS POR VOLUMEN TOTAL DISPENSADO | Período: {df_final['Periodo_Hist'].iloc[0] if len(df_final) else ''}"
    c3.font  = Font(bold=True, size=11, color="FFFFFF")
    c3.fill  = _fill("1F3864")
    c3.alignment = Alignment(horizontal="center")

    h3 = ["#","Medicamento","ABC","XYZ","Grupo","CDL Prom\n(ud/día)","Total\nUnidades"]
    for ci, h in enumerate(h3, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _border()
    ws3.row_dimensions[2].height = 28
    for ri, (_, row) in enumerate(top50.iterrows(), 3):
        ws3.cell(row=ri, column=1, value=ri - 2)
        ws3.cell(row=ri, column=2, value=row["Medicamento"])
        ws3.cell(row=ri, column=3, value=row["ABC"])
        ws3.cell(row=ri, column=4, value=row["XYZ"])
        ws3.cell(row=ri, column=5, value=row["Grupo_ABCXYZ"])
        ws3.cell(row=ri, column=6, value=round(float(row["CDL_Prom"]), 2))
        ws3.cell(row=ri, column=7, value=int(row["Total_Unidades"]))
        fill = FILL_ABC.get(str(row["ABC"]), FILL_ALT)
        for ci in range(1, 8):
            ws3.cell(row=ri, column=ci).border = _border()
            ws3.cell(row=ri, column=ci).fill   = fill
            ws3.cell(row=ri, column=ci).font   = FONT_NORM
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 45
    for ci in range(3, 8):
        ws3.column_dimensions[get_column_letter(ci)].width = 12
    ws3.freeze_panes = "A3"

    # ── Hojas operacionales adicionales ──────────────────────────────────────
    _agregar_hoja_pedido_farm_bodega(wb, df_final, hoy)
    _agregar_hoja_pedido_bod_farmacos(wb, df_final, hoy, dias_ciclo, nombre_dia)
    _agregar_hoja_dialisis(wb, hoy)

    wb.save(out)
    return out


# ─── Pipeline principal ────────────────────────────────────────────────────────

def main():
    hoy          = date.today()
    dia_semana   = hoy.weekday()
    dias_cober   = DIAS_REPORTE.get(dia_semana, 3)
    dias_ciclo   = DIAS_CICLO_BODEGA.get(dia_semana, 10)
    _ndia        = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",
                    5:"Sábado",6:"Domingo"}
    nombre_dia   = _ndia[dia_semana]
    print(f"\n{'='*65}")
    print(f" SGLI HISTÓRICO — {hoy.strftime('%d/%m/%Y')} ({nombre_dia})")
    print(f" Días cobertura Farmacia: {dias_cober} | Días ciclo Bodega AA: {dias_ciclo}")
    print(f"{'='*65}")

    # 1. Cargar recetas
    print("\n[1/4] Cargando historial de prescripciones...")
    df_rec = _cargar_todos_csv()
    fecha_min = df_rec["Fecha"].min().date()
    fecha_max = df_rec["Fecha"].max().date()
    periodo_str = f"{fecha_min.strftime('%d/%m/%Y')} – {fecha_max.strftime('%d/%m/%Y')}"
    print(f"      {len(df_rec):,} registros ENTREGADO | {df_rec['Medicamento'].nunique()} meds únicos")
    print(f"      Período: {periodo_str}")

    # 2. Clasificación ABC-XYZ
    print("\n[2/4] Calculando CDL semanal y clasificación ABC-XYZ...")
    feriados = _cargar_feriados()
    df_abc   = calcular_abc_xyz(df_rec, feriados)
    print(f"      {len(df_abc)} medicamentos procesados")
    for g in ["A","B","C"]:
        sub = df_abc[df_abc["ABC"] == g]
        print(f"      Grupo {g}: {len(sub)} meds")

    # 3. Unir con stock actual
    print("\n[3/4] Cargando stock actual del Consolidado...")
    df_stock = _cargar_stock()
    if not df_stock.empty:
        df_stock["_med"] = df_stock["Medicamento"].apply(lambda x: HOMOLOGACION.get(norm_erp(str(x)), norm_erp(str(x))))
        df_abc["_med"]   = df_abc["Medicamento"]
        merged = df_abc.merge(
            df_stock[["_med","Stock_Farm","Stock_Bod","Criticidad"]].rename(columns={}),
            on="_med", how="left"
        )
        merged.drop(columns=["_med"], inplace=True)
        print(f"      Stock cruzado: {(~merged['Stock_Farm'].isna()).sum()} coincidencias")
    else:
        merged = df_abc.copy()
        print("      [WARN] Consolidado no encontrado; stock = 0")
    if "Stock_Farm" not in merged.columns:
        merged["Stock_Farm"] = 0.0
    if "Stock_Bod" not in merged.columns:
        merged["Stock_Bod"] = 0.0
    if "Criticidad" not in merged.columns:
        merged["Criticidad"] = "5-OK"
    merged["Stock_Farm"] = pd.to_numeric(merged["Stock_Farm"], errors="coerce").fillna(0)
    merged["Stock_Bod"]  = pd.to_numeric(merged["Stock_Bod"],  errors="coerce").fillna(0)
    merged["Criticidad"] = merged["Criticidad"].fillna("5-OK")

    # 4. Calcular stock a pedir
    print("\n[4/4] Calculando stock a pedir y generando planilla...")
    merged["CDL_Prom"]       = merged["CDL_Prom"].round(2)
    merged["CDL_Pico"]       = merged["CDL_Pico"].round(2)
    merged["Stock_Seguridad"] = (merged["CDL_Prom"] * merged["Indice_Repo"]).apply(math.ceil)
    merged["Cant_A_Pedir"]    = (merged["CDL_Prom"] * dias_cober).apply(math.ceil)
    merged["Stock_Objetivo"]  = merged["Cant_A_Pedir"] + merged["Stock_Seguridad"]
    merged["Necesidad"]       = (merged["Stock_Objetivo"] - merged["Stock_Farm"]).clip(lower=0).apply(
        lambda x: math.ceil(float(x))
    )
    merged["Periodo_Hist"] = periodo_str
    merged["Accion"] = merged.apply(
        lambda r: _celda_accion(
            int(r["Necesidad"]), int(r["Stock_Bod"]),
            str(r["Criticidad"]).startswith("1-")
        ),
        axis=1,
    )

    # Ordenar: grupo ABC (A→C) luego CDL_Pico desc (mayor rotación primero)
    orden_abc = {"A": 0, "B": 1, "C": 2}
    merged["_ord"] = merged["ABC"].map(orden_abc).fillna(3)
    merged.sort_values(["_ord", "CDL_Pico"], ascending=[True, False], inplace=True)
    merged.drop(columns=["_ord"], inplace=True)
    merged.reset_index(drop=True, inplace=True)

    # Generar Excel
    out = generar_planilla(merged, hoy, dias_cober, dias_ciclo, nombre_dia)

    # Resumen final
    print(f"\n{'─'*65}")
    print(f" Planilla generada: {os.path.basename(out)}")
    print(f" Medicamentos: {len(merged)} | Con necesidad: {(merged['Necesidad']>0).sum()}")
    print()
    for g in ["A", "B", "C"]:
        sub = merged[merged["ABC"] == g]
        nec = (sub["Necesidad"] > 0).sum()
        print(f"  Grupo {g}: {len(sub):>3} meds | Necesitan reposición: {nec}")
    xyz_dist = merged.groupby("XYZ")["Medicamento"].count()
    print()
    for xyz in ["X","Y","Z"]:
        print(f"  Variabilidad {xyz}: {xyz_dist.get(xyz,0)} meds")
    print(f"{'─'*65}\n")
    print(f"Abrir: {out}")
    return out


if __name__ == "__main__":
    main()
