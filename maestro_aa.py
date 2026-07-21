import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
"""
MAESTRO AA — Pipeline de Consolidación Farmacia Atención Abierta
Hospital de Pitrufquén · Gestión de Pedidos de Medicamentos
═══════════════════════════════════════════════════════════════════

FLUJO (secciones numeradas en el código):
  1-2.  Configuración: paths, umbrales, bodegas, médicos de diálisis
  3-7.  Carga de datos: recetas CSV + stock Excel → universo AA → consumo
  8-12. Indicadores: faltantes, quiebres, pendientes, consumo mensual
  13.   Tendencia semanal (factor de carga por semana del mes)
  14.   Pedidos Farmacia ↔ Bodega AA (ciclo 5 días hábiles)
  15.   Pedidos Diálisis — solo recetas de nefrólogos (mismo motor)
  16.   Auditoría de homologación de nombres
  17.   Comparación de stocks por bodega del hospital
  18.   Resumen ejecutivo KPIs
  19.   Escritura Excel → Consolidado_AA_MAESTRO.xlsx (19 hojas)
  20.   Resumen Semanal operativo → Resumen_Pedidos_AA.xlsx (5 hojas)

ARCHIVOS DE ENTRADA (en la misma carpeta):
  - informe_completo_recetas*.csv   → recetas del ERP SSASUR (módulo RECETA)
  - reporte_de_stock_*.xlsx         → stock actual (módulo ABASTECIMIENTO)
  Ambos se descargan con AUTO_SSASUR.bat → AUTO_SSASUR.py (Playwright)

ARCHIVOS DE SALIDA:
  - Consolidado_AA_MAESTRO.xlsx     → leído por la app Streamlit (19 hojas)
  - Resumen_Pedidos_AA.xlsx         → Excel operativo simplificado (5 hojas)

GLOSARIO:
  CDL     Consumo Diario Laboral (ud/día hábil) — base de todos los cálculos
  CMP     Consumo Mensual Promedio (22 días hábiles)
  COB     Cobertura en días hábiles con el stock actual
  Farmacia AA      FARMACIA AT ABIERTA (mostrador, despacha a pacientes)
  Bodega AA        BODEGA AT ABIERTA (respaldo inmediato de farmacia)
  Bodega Fármacos  Bodega central del hospital (abastece a Bodega AA)
═══════════════════════════════════════════════════════════════════
"""
import pandas as pd
import numpy as np
import re, os, glob, math, warnings, unicodedata
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from aa_colors import CRIT_FILL_HEX, crit_fill, crit_hex, crit_nivel, soften, darken, fill_hex
from sgli import calcular_sgli, cargar_tallas, FACTOR_CARGA_DEFAULT
from utils_aa import norm_erp, HOMOLOGACION

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# PATHS — ruta dinamica: funciona en cualquier PC Windows
# ─────────────────────────────────────────────
WORK_DIR   = os.path.dirname(os.path.abspath(__file__))
# Nombre de salida fijo — se sobreescribe en cada ejecución
# Si está abierto en Excel, genera uno con fecha como respaldo
_OUT_BASE  = os.path.join(WORK_DIR, "Consolidado_AA_MAESTRO.xlsx")
_OUT_DATED = os.path.join(WORK_DIR,
    f"Consolidado_AA_MAESTRO_{datetime.now().strftime('%Y%m%d')}.xlsx")

def _output_path():
    """Devuelve el path de salida; si el archivo está bloqueado usa el dated."""
    if not os.path.exists(_OUT_BASE):
        return _OUT_BASE
    try:
        with open(_OUT_BASE, 'a'):
            pass
        return _OUT_BASE
    except PermissionError:
        print(f"  [aviso] {os.path.basename(_OUT_BASE)} esta abierto en Excel.")
        print(f"          Generando respaldo: {os.path.basename(_OUT_DATED)}")
        return _OUT_DATED

OUTPUT_XLS = _output_path()

# ─────────────────────────────────────────────
# COLUMN NAME NORMALISER  (strips accents -> ASCII)
# ─────────────────────────────────────────────
def norm_col(name: str) -> str:
    nfkd = unicodedata.normalize('NFD', str(name))
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).strip()

# ─────────────────────────────────────────────
# RULE 2 – ERP NAME NORMALISER  (importado desde utils_aa)
# RULE 3 – HOMOLOGATION TABLE   (importado desde utils_aa)
# ─────────────────────────────────────────────
# norm_erp, HOMOLOGACION vienen del import de arriba.

# ─────────────────────────────────────────────
# FACTOR DE EMPAQUE (CENABAST) — para aproximar pedidos a empaques completos
# ─────────────────────────────────────────────
# Formas farmacéuticas a descartar al construir la clave de emparejamiento
_FORMAS_EMP = {
    'CM','CP','COMPRIMIDO','COMPRIMIDOS','COM','C','CAPSULA','CAPSULAS','CAPS','CAP',
    'REC','ENT','UD','GR','G','TAB','TABLETA','SOL','INY','AM','FAM','FA','FAMP','FRA',
    'FCO','FC','POMO','SUSP','JBE','JRP','CREMA','GEL','UN','SOBRE','SOB','LIOF','P',
    'INYECTABLE','ORAL','TOPICO','OFTALMICO','NEB','CAJ','CJ','AMP','PERF','IV',
}

def _clave_empaque(nombre: str) -> str:
    """Clave de emparejamiento: ingrediente (≤3 palabras) + primera concentración."""
    s = norm_erp(nombre).replace('/', ' ').replace('-', ' ')
    m = re.search(r'(\d+[.,]?\d*)\s*(MG|MCG|UI|G|ML|%)', s)
    conc = (m.group(1).replace(',', '.') + m.group(2)) if m else ''
    pre  = s[:m.start()] if m else s
    toks = [t for t in re.split(r'[ .,]+', pre)
            if t and t not in _FORMAS_EMP and not any(ch.isdigit() for ch in t)]
    return (' '.join(toks[:3]) + ' ' + conc).strip()

def cargar_factores_empaque(work_dir):
    """Lee el CSV de intermediación CENABAST y devuelve {clave_empaque: factor}."""
    files = (glob.glob(os.path.join(work_dir, 'cenabast_intermediacion.csv')) +
             glob.glob(os.path.join(work_dir, 'ICP-Intermediacion*.csv')))
    if not files:
        return {}
    f = max(files, key=os.path.getmtime)
    try:
        df = pd.read_csv(f, encoding='latin1', sep=';', skiprows=3, dtype=str)
    except Exception:
        return {}
    if 'NOMBRE COMERCIAL DEL PRODUCTO' not in df.columns or 'NOMBRE GENERICO' not in df.columns:
        return {}
    def _pack(nombre):
        if not isinstance(nombre, str): return None
        s = nombre.upper()
        m = re.search(r'CAJ\s*(\d+)\s*[A-Z]', s) or re.search(r'\bX\s*(\d+)\b', s)
        return int(m.group(1)) if m else None
    out = {}
    for _, r in df.iterrows():
        fac = _pack(r['NOMBRE COMERCIAL DEL PRODUCTO'])
        if not fac or fac <= 1:
            continue
        k = _clave_empaque(r['NOMBRE GENERICO'])
        if k and k not in out:
            out[k] = fac
    return out

def redondear_empaque(cantidad, medicamento, factores):
    """Aproxima HACIA ARRIBA al múltiplo del factor de empaque (mínimo 1 empaque).
    Si no hay factor conocido, redondea hacia arriba a la unidad."""
    c = float(cantidad)
    if c <= 0:
        return 0
    fac = factores.get(_clave_empaque(medicamento))
    if not fac or fac <= 1:
        return int(math.ceil(c))
    return int(math.ceil(c / fac) * fac)

# ─────────────────────────────────────────────
# UNIVERSE FILTERS
# ─────────────────────────────────────────────
EXCLUIR_KW = [
    'ACCESORIO','COMPUTACION','IMPRESION','ABREBOCA','ACEITE DE INMERSION',
    'ACELGA','VERDURA','CLIP','PORTA','LLAVE','SILICONA','MATERIAL DENTAL',
    'IMPLEMENTO','MESA','SILLA','TECLADO','MONITOR','CABLE','OTOSCOPIO',
    'OFTALMOSCOPIO','EJERCITADOR','ELECTRODOS','GUANTE','MASCARILLA',
    'APOSITO','VENDA','SONDA','CATETER','SUTURA','ALIMENTO DIETETICO',
    'ALIMENTO FRESUBIN','PROTEINEX','FORMULA ENTERAL','AGUJA PARA LAPIZ',
    'EQUIPO LECTOR','LECTOR DE GLICEMIA',
    # CINTA PARA DETERMINACION / JERINGA P/INSULINA se dispensan a pacientes
    # en Farmacia AA — se sacaron de aqui para que entren al universo (14-07-2026)
]
INCLUIR_KW = [
    'MG','G ','MCG','UI ','UI\t',' ML','AMP','FA ','COMPRIMIDO','CAPSULA',
    'TAB','JARABE','SOLUCION','CREMA','UNGUENTO','INHALADOR','INSULINA',
    'SUSPENSION','GOTAS','INY','VIAL','LAPICERA','COLIRIO','ELIXIR',
    ' CM',' CP',' UD','SACHET','PARCHE','SOBRE','GEL','OVULO',
]
EXCEP_KW = [
    'TRIAMCINOLONA','CEFTRIAXONA','HALOPERIDOL DECANOATO','ACIDO PAMIDRONICO',
    'LIDOCAINA','ONDANSETRON','DICLOFENACO','METAMIZOL','KETOROLACO',
    'MEDROXIPROGESTERONA','FLUFENAZINA','PALIPERIDONA','RISPERIDONA',
    'ADALIMUMAB','LIRAGLUTIDA','INSULINA','ERITROPOYETINA','EPOETINA',
    'DARBEPOETINA','HEPARINA','ENOXAPARINA','DALTEPARINA','TINZAPARINA',
    'NADROPARINA',
]

# ─────────────────────────────────────────────
# EXCLUSIÓN EXPLÍCITA — NO SON AT ABIERTA
# ─────────────────────────────────────────────
EXCLUIR_ESPECIFICOS_RAW = [
    'ACIDO FOLICO CM 0,4 MG',
    'SODIO CLORURO 0.9% AMP 10 ML',
    'DAPAGLIFLOZINA 10 MG CR',
    'METRONIDAZOL 500 MG/100 ML (5 MG/ML) INYEC',
    'PARACETAMOL ENDOVENOSO 10 MG/ML FC 100 ML.',
    'LACOSAMIDA 100 MG COMPRIMIDO',
    'PENICILINA-G BENZATINA FA 1.200.000 UI',
    # Sin consumo AA / no pertenece a AT Abierta — excluidos manualmente
    'ACETILCISTEINA N 100 MG/ML PARA NEB. FC X 60 ML',
    'ACIDO VALPROICO 500 MG COMPRIMIDO',
    'ACIDO VALPROICO 500 MG COMPRIMIDO LIBERACION MODIFICADA',
    'ALBUMINA HUMANA 20% (200 MG/ML) INYEC',
    'ATROPINA SULFATO 1 MG/ML AMPOLLA',
    'BUPIVACAINA FA 50 MG/10 ML',
    'CLOSTRIDIUM BOTULINUM TOXINA TIPO A 100 U INYECTABLE',
    'DEXMEDETOMIDINA 200 MCG/ 2 ML',
    'EMULSION CON UREA 10 % UD 100 ML',
    'ESCOPOLAMINA N-BUTIL BROMURO 20 MG/ML AMPOLLA.',
    'ETOMIDATO FA 20 MG/10 ML',
    'FLUFENAZINA 250 MG/10 ML SOL. INY. FA.',
    'GAMAGLOBULINA 5 % FA 5 GR EV',
    'GENTAMICINA 0,3% (3 MG/ML) COLIRIO.',
    'GLUCONATO DE CALCIO AMP 10%',
    'HALOPERIDOL AMP 5 MG/ML',
    'KETAMINA 500 MG/10 ML (50 MG/ML) INYECTABL',
    'LABETALOL 5 MG/ ML AM 20 ML',
    'LANATOSIDO C 400 MCG/2 ML (200 MCG/ML) INYECTABLE',
    'LEVETIRACETAM 100 MG/ML AMP 5 ML',
    'LEVOFLOXACINO 500 MG/100 ML (5 MG/ML) INYECTABLE',
    'LEVOSIMENDAN 12,5 MG/5 ML (2,5 MG/ML) INYECTABLE',
    'LIDOCAINA 10% (100 MG/ML) SOLUCION USO EXTERNO',
    'LIDOCAINA 2 % FC 1 UD GEL ORAL',
    'METRONIDAZOL 0,75% CREMA',
    'NALOXONA 0,4 MG/ML INYECTABLE',
    'NITROGLICERINA 50 MG/10 ML (5 MG/ML) INYECTABLE',
    'NOREPINEFRINA 4 MG/4 ML (1 MG/ML) INYECTABLE',
    'OLANZAPINA 10 MG INYECTABLE',
    'OMEPRAZOL FA 40 MG',
    'PENICILINA-G SODICA 2.000.000 U.I IV/IM FA',
    'PILOCARPINA 2% (20 MG/ML) COLIRIO',
    'PIPERACILINA/TAZOBACTAM 4G/500 MG FAM',
    'PROPANOLOL AM 1MG/1ML',
    'PROPOFOL 1 % UD 20 ML',
    'PROPOFOL 1% F.A 50 ML CON SET DE ADMINISTRACION',
    'PROTAMINA SULFATO 50 MG/5 ML (10 MG/ML) INYECTABLE',
    'RANITIDINA 50MG/2ML SOL INY IV-IM AM/FAM',
    'SACCHAROMYCES BOULARDII 250 MG CAP',
    'SIMETICONA 40 MG/ML GOTAS ORALES',
    'SODIO CLORURO 0.9 % AMP 20 ML',
    'SODIO CLORURO 10% AMP 10 ML',
    'SODIO CLORURO 10% AMP 20 ML',
    'SUCCINILCOLINA (SUXAMETONIO) AMP 100 MG/5 ML',
    'SULBACTAM + AMPICILINA 500 MG/1000 MG',
    'SULFAMETOXAZOL 400 MG/5 ML (80 MG/ML) + TRIMETOPRIMA 80 MG/5 ML (16 MG/ML) INYECTABLE',
    'SULFATO DE MAGNESIO AMP 25%',
    'SUSTITUTO DEL PLASMA 6% 500 ML',
    'TRAMADOL AMPOLLAS 100 MG/2 ML',
    'URAPIDIL 5 MG/ML INYECTABLE AMP 10 ML',
    'URAPIDIL 5 MG/ML INYECTABLE AMP 20 ML',
    'VASELINA LIQUIDA ESTERIL AMPOLLA 10 ML.',
    'VITAMINAS HIDROSOLUBLES EV 10 ML FA ( TIAMINA/ RIBOFLAVINA/ NICOTINAMIDA/ PIRIDOXINA/ PANTOTENATO DE SODIO/ ASCORBATO DE SODIO/ BIOTINA/ ACIDO FOLICO/ CIANOCOBALAMINA)',
    # Batch 28/05/2026
    'ACENOCUMAROL 4 MG CM',
    'ACIDO TRANEXAMICO AMP 1000 MG/10 ML',
    'AMIKACINA FA 500 MG/2 ML',
    'AZUL DE METILENO 1% AMP 10 ML',
    'CEFAZOLINA 1000 MG INYECTABLE',
    'CLINDAMICINA FA 600 MG/4 ML',
    'CLONIXINATO DE LISINA AMP 100 MG/2 ML',
    'CLORFENAMINA 10 MG/ML INYECTABLE',
    'CLORURO DE POTASIO 10 % AM 10 ML SOLUCION INYECTABLE.',
    'CLOSTRIDIUM BOTULINUM TOXINA TIPO A 200 U INYECTABLE',
    'DOMPERIDONA 10 MG/2 ML (5 MG/ML) INYECTABLE',
    'DROPERIDOL AMP 5 MG/2 ML',
    'EFEDRINA 6% (60 MG/ML) INYECTABLE',
    'EPINEFRINA 0,1% (1 MG/ML) INYECTABLE',
    'FITOMENADIONA 10MG/ML AM ENDOVENOSA U ORAL',
    'FLUMAZENIL 500 MCG/5 ML (100 MCG/ML) INYECTABLE',
    'FUROSEMIDA 20 MG/2 ML (10 MG/ML) INYECTABLE',
    'GLUCOSA 30% AMP 20 ML',
    'HIDROCORTISONA SUCCINATO SODICO 100 MG INYECTABLE',
    'MEROPENEM 1000 MG INYECTABLE',
    'PROPINOXATO 5 MG / 1ML AM',
]
EXCLUIR_ESPECIFICOS = {norm_erp(x) for x in EXCLUIR_ESPECIFICOS_RAW}

# ─────────────────────────────────────────────
# FORZAR UNIVERSO — incluir aunque no aparezcan en recetas historicas
# (tienen stock AA pero sin prescripciones en el periodo cargado)
# Se usan los nombres EXACTOS del reporte de stock para que el pivot de
# stock los encuentre correctamente.
# ─────────────────────────────────────────────
FORZAR_UNIVERSO_RAW = [
    'MICONAZOL 100 MG + TINIDAZOL 150 MG CAPSULA VAGINAL',
    'LEVONORGESTREL 750 MCG COMPRIMIDO',
    'LEVONORGESTREL 1,5 MG CM',
    'NORETISTERONA/ESTRADIOL 50/5MG JRP',
    'MEDROXIPROGESTERONA 150 MG/1 ML',
    'TOLTERODINA CM 2 MG',
    'TOLTERODINA L-TARTRATO 2 MG CM',
    'BETAHISTINA 16 MG COMPRIMIDO',
    'CABERGOLINA 0,5 MG COMPRIMIDO',
    'ACIDO PAMIDRONICO FA 90 MG',
    'LIDOCAINA 2% AMP 5 ML',
    'ONDANSETRON 4 MG/2 ML (2 MG/ML) INYECTABLE',
    'PROGESTERONA 200 MG CM MICRONIZADA',
    'ESTRIOL 0,5 MG OVULO',
    'PRESERVATIVOS LUBRICADOS 1 UD',
    'ADIFENINA 50 MG + PROPIFENAZONA 440 MG SUPOSITORIO',
    'ISOTRETINOINA 10 MG CAPSULA',
    'MICOFENOLATO 500 MG CR',
    'DICLOFENACO 75 MG/3ML SOL. IM AM',
    'METAMIZOL SODICO AMP 1000 MG/2 ML',
    'KETOROLACO AMP 30 MG/ML',
    # Paracetamol gotas — sin stock AA, solicitado activamente
    'PARACETAMOL GOTAS FC 15ML',
    'PARACETAMOL 100 MG/ML GOTAS ORALES',
    # Batch 14/07/2026 — tienen stock real en Farmacia/Bodega AT Abierta
    # pero sin recetas en el periodo cargado (detectados via auditoria acetazolamida)
    'ACICLOVIR 3% UNGUENTO OFTALMICO',
    'ADHESIVO HEMOSTATICO FISIOLOGICO 1 ML AM',
    'ARIPIPRAZOL 1 MG/ML FC 150 ML',
    'AZITROMICINA 200 MG/5 ML POLVO PARA PREPARADO ORAL',
    'AZITROMICINA 400 MG/5 ML POLVO PARA PREPARADO ORAL',
    'CEFUROXIMA 250 MG/5 ML X 100 ML',
    'CICLOSPORINA CP 50 MG',
    'CIPROFLOXACINO 0,3% (3 MG/ML) + DEXAMETASONA 0,1% (1 MG/ML) GOTAS OTICAS',
    'CIPROFLOXACINO 0,3% GOTAS OTICAS 5 ML',
    'DEXAMETASONA 0,1% (1 MG/ML) + TOBRAMICINA 0,3% (3 MG/ML) UNGUENTO OFTALMICO',
    'DICLOFENACO SODICO 0,1% SOL. OFT FC',
    'DICLOFENACO SODICO 12,5 MG SUPOSITORIO',
    'ERITROMICINA 500 MG COMPRIMIDO',
    'FENILEFRINA 2,5 % COLIRIO X 2 ML',
    'FENOBARBITAL 15 MG COMPRIMIDO',
    'FENOBARBITAL 200 MG/ML INYECTABLE',
    'FENTANILO 8,4 MG (50 MCG/HORA) PARCHE',
    'INMUNOGLOBULINA HUMANA ANTI RHO (D) 300 MCG/2 ML (150 MCG/ML) INYECTABLE',
    'ISOPROTERENOL AMP 1 MG/ ML',
    'LAMIVUDINA 10 MG/ML SOLUCION ORAL (JBE) UD.',
    'LAMIVUDINA 150 MG + ZIDOVUDINA 300 MG COMPRIMIDO',
    'LEVONORGESTREL 0,03 MG CICLO 28 O 35 CM',
    'LEVONORGESTREL 20 MCG/24 HRS SISTEMA INTRAUTERINO',
    'MEMANTINA CLORHIDRATO 20 MG COMPRIMIDO RECUBIERTO',
    'METADONA AMP 10 MG/2 ML',
    'MICOFENOLATO MOFETILO 250 MG CAPSULA',
    'MOXIFLOXACINO 0,5% (5 MG/ML) COLIRIO',
    'NIMODIPINO 30 MG CM',
    'NISTATINA 100.000 UI/ML SUSPENSION ORAL',
    'OXCARBAZEPINA CM 300 MG',
    'PALIPERIDONA 75 MG JER.PC',
    'PARACETAMOL 125 MG SUPOSITORIO',
    'PARACETAMOL 80 MG COMPRIMIDO',
    'PETIDINA 100 MG/2 ML (50 MG/ML) INYECTABLE',
    'PILOCARPINA COLIRIO 4% 10ML',
    'POLIVITAMINICO ACD GOTAS FC 30 ML',
    'PROPARACAINA 0,5% (5 MG/ML) COLIRIO',
    'RIFAMPICINA/ISONIAZIDA 150/75 CM',
    'TALIDOMIDA 100 MG COMPRIMIDO',
    'TENECTEPLASA 50 MG INYECTABLE',
    'TENOFOVIR / LAMIVUDINA/ DOLUTEGRAVIR 300 MG/300 MG/50MG CM UD',
    # Batch 14/07/2026 (2) — mismo motivo, formato de dosis no reconocido
    # por INCLUIR_KW (POMADA, SOL abreviado, INH, GR con punto, MEQ, combos con "/")
    'ATROPINA 1% SOL OFTALMICA FC',
    'HIDROCORTISONA 1 % POMADA 10 A 30 G',
    'TOBRAMICINA 0.3% SOL. OFTALMICA FC',
    'UMECLIDINIO/VILANTEROL 55/22 INH',
    'VASELINA AZUFRADA 6% PT 40 GR.',
    'POLIESTIRENO SULFONATO CALCICO 15 GR SO',
    'RIFA/ISO/PIRA/ETA 150/75/400/2',
    'SALES PARA REHIDRATACION ORAL 60 A 75 MEQ/LT',
    'SALES PARA REHIDRATACION ORAL 90 MEQ',
]
FORZAR_UNIVERSO = {norm_erp(x) for x in FORZAR_UNIVERSO_RAW}

def es_farmaco_aa(nombre: str) -> bool:
    n = norm_erp(nombre)
    if n in EXCLUIR_ESPECIFICOS:
        return False
    # MESALAZINA colisiona con la palabra clave 'MESA' (mobiliario) de EXCLUIR_KW
    if n.startswith('MESALAZINA'):
        return True
    if any(kw in n for kw in EXCLUIR_KW):
        return False
    if any(kw in n for kw in EXCEP_KW):
        return True
    return any(kw in n for kw in INCLUIR_KW)

# ─────────────────────────────────────────────
# BODEGA SCOPES
# ─────────────────────────────────────────────
BODEGAS_AA_FARMACIA = {'FARMACIA AT ABIERTA'}
BODEGAS_AA_BODEGA   = {'BODEGA AT ABIERTA'}
BODEGAS_AA          = BODEGAS_AA_FARMACIA | BODEGAS_AA_BODEGA

BODEGAS_EXCLUIDAS = {
    'BODEGA CUARENTENA','BODEGA REENVASADO FARMACOS',
    'BODEGA EQUIPAMIENTO','BODEGA ORTESIS','BODEGA HABILITACION SEGUNDA FASE',
}
BODEGAS_HOSPITAL = {
    'FARMACIA AT CERRADA','BODEGA AT CERRADA',
    'BODEGA ACTIVA FARMACIA URGENCIA','BODEGA FARMACOS',
}
ORDEN_TRASPASO = [
    'BODEGA FARMACOS','BODEGA AT CERRADA',
    'FARMACIA AT CERRADA','BODEGA ACTIVA FARMACIA URGENCIA',
]

# ─────────────────────────────────────────────
# MEDICOS PRESCRIPTORES DE DIALISIS (nefrologia / ERC-5D / hemodialisis)
# Solo las recetas firmadas por estos profesionales definen el universo y el
# consumo de la pestaña/hojas de Dialisis. Nombre completo normalizado (norm_erp).
# ─────────────────────────────────────────────
MEDICOS_DIALISIS_RAW = [
    'YASMANI RAMIRO ORTIZ AMADOR',
    'MARTHA PERALTA BELTRAN',
    'MONICA ANYUL AMAYA FORGIONNI',
]
MEDICOS_DIALISIS = {norm_erp(x) for x in MEDICOS_DIALISIS_RAW}

# ─────────────────────────────────────────────
# COVERAGE TARGET (Rule 6)
# ─────────────────────────────────────────────
COB_LABORAL = {0:5, 1:4, 2:4, 3:4, 4:6, 5:5, 6:5}  # Mon=0 … Sun=6

# ═══════════════════════════════════════════════
# 1. LOAD STOCK
# ═══════════════════════════════════════════════
_T_INICIO = datetime.now()
print("=" * 60)
print(f"  MAESTRO AA  —  {_T_INICIO.strftime('%d/%m/%Y  %H:%M:%S')}")
print("=" * 60)

# ── Stock: usa siempre el archivo MÁS RECIENTE ────────────────────────────────
print("Cargando stock...")
_stk_files = glob.glob(os.path.join(WORK_DIR, 'reporte_de_stock_*.xlsx'))
if not _stk_files:
    raise FileNotFoundError("No se encontro ningun archivo reporte_de_stock_*.xlsx")
stk_file = max(_stk_files, key=os.path.getmtime)
print(f"  Archivo stock : {os.path.basename(stk_file)}")
df_stk_raw = pd.read_excel(stk_file, header=2, engine='openpyxl')
df_stk_raw.columns = [norm_col(c) for c in df_stk_raw.columns]

# Rename to canonical
stk_rename = {}
for c in df_stk_raw.columns:
    cu = c.upper()
    if 'DESCRIPCI' in cu:  stk_rename[c] = 'Descripcion'
    elif 'BODEGA'   in cu:  stk_rename[c] = 'Bodega'
    elif 'CANTIDAD' in cu:  stk_rename[c] = 'Cantidad'
    elif 'CODIGO'   in cu:  stk_rename[c] = 'Codigo'
df_stk_raw.rename(columns=stk_rename, inplace=True)

df_stk = df_stk_raw[['Codigo','Bodega','Descripcion','Cantidad']].copy()
df_stk.dropna(subset=['Descripcion','Bodega'], inplace=True)
df_stk['Cantidad'] = pd.to_numeric(df_stk['Cantidad'], errors='coerce').fillna(0)
df_stk['Bodega']   = df_stk['Bodega'].str.strip().str.upper()

# Rule 2 – normalise ERP description
df_stk['Descripcion_norm'] = df_stk['Descripcion'].apply(norm_erp)

print(f"  Stock: {len(df_stk):,} filas, {df_stk['Bodega'].nunique()} bodegas")

# ═══════════════════════════════════════════════
# 2. LOAD & DEDUPLICATE RECETAS  (Rule 1)
# ═══════════════════════════════════════════════
print("Cargando recetas...")
csv_files = sorted(glob.glob(os.path.join(WORK_DIR, 'informe_completo_recetas*.csv')))
if not csv_files:
    raise FileNotFoundError(
        "No se encontro ningun archivo informe_completo_recetas*.csv\n"
        "Ejecuta AUTO_SSASUR.bat primero."
    )
chunks = []
for f in csv_files:
    tmp = pd.read_csv(f, encoding='latin1', sep=';', on_bad_lines='skip', dtype=str)
    tmp.columns = [norm_col(c) for c in tmp.columns]
    chunks.append(tmp)
    print(f"  {os.path.basename(f)}: {len(tmp):,} filas")

df_rec = pd.concat(chunks, ignore_index=True)
antes = len(df_rec)
df_rec = df_rec.drop_duplicates(subset=['ID Receta Detalle'], keep='first')
print(f"  Deduplicacion: {antes:,} -> {len(df_rec):,} (eliminados {antes-len(df_rec):,})")

# ─── Rename columns ───
rec_map = {}
for c in df_rec.columns:
    cu = c.upper()
    if cu in ('PRESCRIPCION', 'PRESCRIPCI\xd3N'):
        rec_map[c] = 'Prescripcion'
    elif 'ESTADO PRESCRIPCI' in cu:    rec_map[c] = 'Estado_Prescripcion'
    elif c == 'Estado':                 rec_map[c] = 'Estado_Receta'
    elif 'FECHA ATENCI' in cu:                        rec_map[c] = 'Fecha_Atencion'
    elif cu == 'FECHA ENTREGA RECETA':               rec_map[c] = 'Fecha_Entrega'
    elif 'CANTIDAD PENDIENTE' in cu:    rec_map[c] = 'Cantidad_Pendiente'
    elif 'CANTIDAD ENTREGADA' in cu:    rec_map[c] = 'Cantidad_Entregada'
    elif 'CANTIDAD RECETADA' in cu:     rec_map[c] = 'Cantidad_Recetada'
    elif 'BODEGA DESPACHA' in cu:       rec_map[c] = 'Bodega_Despacha'
    elif c == 'Procedencia':            rec_map[c] = 'Procedencia'
    elif c == 'RUN':                    rec_map[c] = 'RUN'
    elif c == 'Nombre':                 rec_map[c] = 'Nombre'
    elif 'APELLIDO PATERNO' in cu and 'PROFESIONAL' not in cu: rec_map[c] = 'Apellido_Paterno'
    elif 'APELLIDO MATERNO' in cu and 'PROFESIONAL' not in cu: rec_map[c] = 'Apellido_Materno'
    elif 'NUMERO RECETA' in cu:         rec_map[c] = 'Numero_Receta'
    elif 'NUMERO FOLIO' in cu:          rec_map[c] = 'Numero_Folio'
    elif 'ID RECETA DETALLE' in cu:     rec_map[c] = 'ID_Receta_Detalle'
    elif cu == 'PERIODO' and '.1' not in c: rec_map[c] = 'Periodo'
df_rec.rename(columns=rec_map, inplace=True)

# Types
df_rec['Fecha_Atencion'] = pd.to_datetime(df_rec['Fecha_Atencion'], dayfirst=True, errors='coerce')
# Fecha_Entrega: fecha real de despacho en farmacia (usada para consumo)
# Para PENDIENTES sin entrega usa Fecha_Atencion como fallback
_col_fe = df_rec['Fecha_Entrega'] if 'Fecha_Entrega' in df_rec.columns \
          else pd.Series(pd.NaT, index=df_rec.index)
df_rec['Fecha_Entrega_raw'] = pd.to_datetime(_col_fe, dayfirst=True, errors='coerce')
df_rec['Fecha_Entrega'] = df_rec['Fecha_Entrega_raw'].fillna(df_rec['Fecha_Atencion'])
df_rec['Cantidad_Pendiente'] = pd.to_numeric(df_rec['Cantidad_Pendiente'], errors='coerce').fillna(0)
df_rec['Cantidad_Entregada'] = pd.to_numeric(df_rec['Cantidad_Entregada'], errors='coerce').fillna(0)
df_rec['Cantidad_Recetada']  = pd.to_numeric(df_rec['Cantidad_Recetada'],  errors='coerce').fillna(0)
df_rec['Prescripcion']       = df_rec['Prescripcion'].fillna('').astype(str).str.strip()
df_rec['Procedencia']        = df_rec['Procedencia'].fillna('').str.strip().str.upper()
df_rec['Estado_Prescripcion']= df_rec['Estado_Prescripcion'].fillna('').str.strip().str.upper()
df_rec['Bodega_Despacha']    = df_rec['Bodega_Despacha'].fillna('').str.strip().str.upper()

# ─── Prescripcion_norm (Rule 2 + homologation Rule 3) ───
df_rec['Prescripcion_norm'] = df_rec['Prescripcion'].apply(norm_erp)
df_rec['Prescripcion_norm'] = df_rec['Prescripcion_norm'].map(lambda x: HOMOLOGACION.get(x, x))


# ═══════════════════════════════════════════════
# 3. FECHA_MAX, PERÍODOS (Rule 4+5)
# ═══════════════════════════════════════════════
HOY             = pd.Timestamp.today().normalize()
# FECHA_INICIO: primera Fecha_Entrega_raw real del histórico (dinámica — no hardcoded)
# Usa Fecha_Entrega_raw (campo original del CSV) para evitar que el fallback a
# Fecha_Atencion de registros PENDIENTE arrastre fechas antiguas.
_ini_raw        = df_rec['Fecha_Entrega_raw'].dropna().min()
FECHA_INICIO_OP = _ini_raw if pd.notna(_ini_raw) else df_rec['Fecha_Entrega'].dropna().min()
# FECHA_MAX: última Fecha_Entrega del histórico (dinámica)
FECHA_MAX_DATA  = df_rec['Fecha_Entrega'].dropna().max()
FECHA_MAX       = min(FECHA_MAX_DATA, HOY - pd.Timedelta(days=1))

# Período real del histórico: 1 ene 2026 → FECHA_MAX
DIAS_CAL_OP  = (FECHA_MAX - FECHA_INICIO_OP).days + 1
MESES_TOT_OP = DIAS_CAL_OP / 30.4375          # meses reales del período

# Días laborales del período y ventanas ponderadas (sobre Fecha_Entrega)
DIAS_LAB_OP  = len(pd.bdate_range(FECHA_INICIO_OP, FECHA_MAX))
DIAS_LAB_15  = len(pd.bdate_range(FECHA_MAX - timedelta(days=15), FECHA_MAX))
DIAS_LAB_30  = len(pd.bdate_range(FECHA_MAX - timedelta(days=30), FECHA_MAX))
DIAS_LAB_60  = len(pd.bdate_range(FECHA_MAX - timedelta(days=60), FECHA_MAX))

# Estándar de días hábiles por mes (base para CMP_Mensual y Consumo_5D)
# Todas las métricas derivan de CDL → son siempre proporcionales:
#   CMP_Mensual = CDL × 22   |   Consumo_5D = CDL × 5   |   CMP/22 = CDL ✓
DIAS_LAB_MES = 22

print(f"  FECHA_INICIO operacional : {FECHA_INICIO_OP.date()}")
print(f"  FECHA_MAX operacional    : {FECHA_MAX.date()}")
print(f"  Dias calendario periodo: {DIAS_CAL_OP}")
print(f"  Meses reales periodo   : {MESES_TOT_OP:.4f}")
print(f"  Dias laborales periodo : {DIAS_LAB_OP}")
print(f"  Dias habiles/mes std   : {DIAS_LAB_MES}")

# Coverage target based on FECHA_MAX weekday
dow_hoy          = HOY.weekday()
COB_OBJETIVO     = COB_LABORAL[dow_hoy]
print(f"  Cobertura objetivo: {COB_OBJETIVO} días laborales")

# ═══════════════════════════════════════════════
# 4. FILTER AA RECETAS
# ═══════════════════════════════════════════════
mask_proc = df_rec['Procedencia'].isin(['ATENCION ABIERTA', 'SIN RESPALDO'])
df_aa = df_rec[mask_proc].copy()
print(f"  Recetas AA (post-filtro procedencia): {len(df_aa):,}")

# ═══════════════════════════════════════════════
# 5. BUILD UNIVERSE AA
# ═══════════════════════════════════════════════
# From ALL historical data
universo_raw = df_aa['Prescripcion_norm'].unique()
universo_set  = {p for p in universo_raw if es_farmaco_aa(p)}
# Agregar forzados (tienen stock AA pero sin recetas en el periodo cargado)
universo_set |= FORZAR_UNIVERSO
print(f"  Universo AA: {len(universo_set):,} medicamentos únicos")
print(f"  (incluye {len(FORZAR_UNIVERSO)} forzados sin recetas historicas)")

# Flag on df_aa
df_aa['en_universo'] = df_aa['Prescripcion_norm'].isin(universo_set)
df_aa_univ = df_aa[df_aa['en_universo']].copy()

# ═══════════════════════════════════════════════
# 6. STOCK PIVOTS
# ═══════════════════════════════════════════════
# Apply homologation to stock descriptions
df_stk['Descripcion_norm'] = df_stk['Descripcion_norm'].map(lambda x: HOMOLOGACION.get(x, x))

# Only bodegas in scope (not excluded)
df_stk_scope = df_stk[~df_stk['Bodega'].isin(BODEGAS_EXCLUIDAS)].copy()

def pivot_stock(bodegas_set, col_name):
    sub = df_stk_scope[df_stk_scope['Bodega'].isin(bodegas_set)]
    return sub.groupby('Descripcion_norm')['Cantidad'].sum().rename(col_name)

stk_farmacia_aa = pivot_stock(BODEGAS_AA_FARMACIA, 'Stock_Farmacia_AA')
stk_bodega_aa   = pivot_stock(BODEGAS_AA_BODEGA,   'Stock_Bodega_AA')
# Full stock per bodega (all in scope)
stk_all = df_stk_scope.pivot_table(
    index='Descripcion_norm', columns='Bodega', values='Cantidad',
    aggfunc='sum', fill_value=0
)

# ═══════════════════════════════════════════════
# 7. CONSUMO & CMP  (Rule 5)
# ═══════════════════════════════════════════════
# Only from 2026-01-01 onwards
df_op = df_aa_univ[df_aa_univ['Fecha_Entrega'] >= FECHA_INICIO_OP].copy()

# ── FILTRO DE ESTADOS VÁLIDOS PARA CONSUMO ───────────────────────────────────
# Definicion CMP: Suma total de Cantidad_Recetada (líneas válidas) / N meses
# Cada línea de prescripción (ID_Receta_Detalle único) cuenta con su cantidad
# completa, independientemente de cuántos períodos se retiren en un mismo día.
# Estados incluidos:
#   ENTREGADO  : entrega confirmada — consumo real histórico
#   PENDIENTE  : prescrito, no entregado aún — demanda real activa
#   SOLICITADO : solicitado a bodega, en proceso
# Estados excluidos:
#   REEMPLAZADO: la receta sustituta ya aparece como ENTREGADO → evita doble conteo
#   ANULADO / RECHAZADO / DEVUELTO: no representan consumo real
ESTADOS_CONSUMO = {'ENTREGADO', 'PENDIENTE', 'SOLICITADO'}
df_op_consumo = df_op[df_op['Estado_Prescripcion'].isin(ESTADOS_CONSUMO)].copy()
n_excluidos_estado = len(df_op) - len(df_op_consumo)
print(f"  Registros excluidos (REEMPLAZADO/ANULADO/RECHAZADO/DEVUELTO): {n_excluidos_estado:,}")
print(f"  Registros para calculo CMP: {len(df_op_consumo):,}")

def consumo_periodo(df, fecha_ini, fecha_fin):
    # Filtra por Fecha_Entrega (fecha real de despacho en farmacia)
    mask = (df['Fecha_Entrega'] >= fecha_ini) & (df['Fecha_Entrega'] <= fecha_fin)
    return df[mask].groupby('Prescripcion_norm')['Cantidad_Recetada'].sum()

# Período completo operacional (Fecha_Entrega: 2026-01-01 -> FECHA_MAX)
c_total = consumo_periodo(df_op_consumo, FECHA_INICIO_OP, FECHA_MAX)

# Consumo de DIÁLISIS por medicamento (recetas de nefrólogos). Se usa para
# SEPARAR el pedido normal: Farmacia AA repone solo lo NO-diálisis, porque el
# pedido de diálisis se hace aparte (mensual, 3ª semana).
df_op_consumo['Prof_Norm'] = (
    df_op_consumo['Nombre Profesional'].fillna('') + ' ' +
    df_op_consumo['Apellido Paterno Profesional'].fillna('') + ' ' +
    df_op_consumo['Apellido Materno Profesional'].fillna('')
).apply(norm_erp)
c_total_dial = consumo_periodo(
    df_op_consumo[df_op_consumo['Prof_Norm'].isin(MEDICOS_DIALISIS)],
    FECHA_INICIO_OP, FECHA_MAX)

# Ventanas para CMP ponderado (base Fecha_Entrega)
c_15 = consumo_periodo(df_op_consumo, FECHA_MAX - timedelta(days=15), FECHA_MAX)
c_30 = consumo_periodo(df_op_consumo, FECHA_MAX - timedelta(days=30), FECHA_MAX)
c_60 = consumo_periodo(df_op_consumo, FECHA_MAX - timedelta(days=60), FECHA_MAX)

# ─── BUILD MASTER TABLE ───
idx = sorted(universo_set)
df_master = pd.DataFrame(index=idx)
df_master.index.name = 'Medicamento'

df_master['Stock_Farmacia_AA']  = stk_farmacia_aa.reindex(idx, fill_value=0)
df_master['Stock_Bodega_AA']    = stk_bodega_aa.reindex(idx,   fill_value=0)
df_master['Stock_AA_Total']     = df_master['Stock_Farmacia_AA'] + df_master['Stock_Bodega_AA']

# Hospital bodegas
for b in ORDEN_TRASPASO:
    col = f'Stock_{b.replace(" ","_")}'
    series = pivot_stock({b}, col)
    df_master[col] = series.reindex(idx, fill_value=0)

df_master['Stock_Hospital_Total'] = df_master[[
    f'Stock_{b.replace(" ","_")}' for b in ORDEN_TRASPASO
]].sum(axis=1)

# Consumo columns (base: Cantidad_Recetada)
df_master['Prescrito_Total_Op']   = c_total.reindex(idx, fill_value=0)
df_master['Prescrito_Dialisis_Op'] = c_total_dial.reindex(idx, fill_value=0)
df_master['Prescrito_15d']        = c_15.reindex(idx, fill_value=0)
df_master['Prescrito_30d']        = c_30.reindex(idx, fill_value=0)
df_master['Prescrito_60d']        = c_60.reindex(idx, fill_value=0)

# ── MÉTRICAS EN DÍAS HÁBILES (todas derivadas de CDL — siempre proporcionales) ──
#
#   CDL          = Prescrito_Total / DIAS_LAB_OP        [unidades / día hábil]
#   CMP_Mensual  = CDL × 22                             [unidades / mes hábil]
#   Consumo_5D   = CDL × 5                              [unidades para 5 días hábiles]
#   Cobertura    = Stock_Farmacia / CDL                 [días hábiles de cobertura]
#   Reposicion   = max(Consumo_5D − Stock_Farmacia, 0)  [unidades a pedir]
#
#   Verificación: CMP_Mensual / 22 = CDL  |  Consumo_5D / 5 = CDL  ✓

# 1. CDL — consumo diario laboral (base: Fecha_Entrega, período completo)
df_master['CDL'] = np.where(
    DIAS_LAB_OP > 0,
    df_master['Prescrito_Total_Op'] / DIAS_LAB_OP,
    0
).round(4)

# 2. CMP_Mensual = CDL × 22  (mes hábil estándar → siempre CMP/22 = CDL)
df_master['CMP_Mensual'] = (df_master['CDL'] * DIAS_LAB_MES).round(2)

# 3. Consumo_5D = CDL × 5  (necesidad para los próximos 5 días hábiles)
df_master['Consumo_5D'] = (df_master['CDL'] * 5).round(1)

# 4. CMP AA Laboral ponderado — tasa diaria reciente (referencia de tendencia)
#    15d×50% + 30d×30% + 60d×20% — NO se usa para Reposicion_5D
cmp_15 = (df_master['Prescrito_15d'] / DIAS_LAB_15) * 0.50
cmp_30 = (df_master['Prescrito_30d'] / DIAS_LAB_30) * 0.30
cmp_60 = (df_master['Prescrito_60d'] / DIAS_LAB_60) * 0.20
df_master['CDL_Pond'] = (cmp_15 + cmp_30 + cmp_60).round(4)  # tasa diaria ponderada

# 5. Cobertura en días hábiles  (stock farmacia / CDL)
df_master['Cobertura_Lab'] = np.where(
    df_master['CDL'] > 0,
    df_master['Stock_Farmacia_AA'] / df_master['CDL'],
    np.where(df_master['Stock_Farmacia_AA'] > 0, 9999, 0)
).round(1)

# 6. Reposición 5D = max(Consumo_5D − Stock_Farmacia_AA, 0)
df_master['Reposicion_Sugerida'] = np.maximum(
    df_master['Consumo_5D'] - df_master['Stock_Farmacia_AA'],
    0
).round(0)

# 7. Quiebre total AA (stock AA = 0 y hay consumo real)
df_master['Quiebre_Total'] = (
    (df_master['Stock_Farmacia_AA'] == 0) &
    (df_master['Stock_Bodega_AA']   == 0) &
    (df_master['CDL']               >  0)
)

df_master = df_master.reset_index()
print(f"  Master table: {len(df_master):,} medicamentos")

# ── VALIDACIÓN CONSISTENCIA (CDL × 22 = CMP_Mensual | CDL × 5 = Consumo_5D) ──
print(f"\n  {'Medicamento':<48} {'CDL':>7} {'CMP(×22)':>10} {'5D(×5)':>8} {'Rep5D':>7}  {'Cob(d)':>6}")
print(f"  {'-'*92}")
top_val = df_master.nlargest(8, 'CDL')
for _, r in top_val.iterrows():
    print(f"  {r['Medicamento'][:46]:<46}  "
          f"{r['CDL']:7.2f}  "
          f"{r['CMP_Mensual']:9.1f}  "
          f"{r['Consumo_5D']:7.1f}  "
          f"{r['Reposicion_Sugerida']:6.0f}  "
          f"{r['Cobertura_Lab']:6.1f}")

# ═══════════════════════════════════════════════
# 8. FALTANTES REALES AA
# ═══════════════════════════════════════════════
mask_falt = (
    df_aa_univ['Estado_Prescripcion'].isin(['PENDIENTE', 'SOLICITADO']) &
    (df_aa_univ['Cantidad_Pendiente'] > 0)
)
df_pend = df_aa_univ[mask_falt].copy()

# Aggregate demand by medication
dem_agg = df_pend.groupby('Prescripcion_norm').agg(
    Cant_Demanda_Activa=('Cantidad_Pendiente', 'sum'),
    Pacientes_Afectados=('RUN',               'nunique'),
    N_Recetas           =('Numero_Receta',     'nunique'),
).reset_index()

# Merge with master stock
df_falt = dem_agg.merge(
    df_master[['Medicamento','Stock_Farmacia_AA','Stock_Bodega_AA','Stock_AA_Total',
               'Stock_Hospital_Total'] +
              [f'Stock_{b.replace(" ","_")}' for b in ORDEN_TRASPASO]],
    left_on='Prescripcion_norm', right_on='Medicamento', how='left'
).fillna(0)

# Only real faltantes: stock AA = 0
df_falt = df_falt[df_falt['Stock_AA_Total'] == 0].copy()
df_falt['Faltante_Neto'] = np.maximum(
    df_falt['Cant_Demanda_Activa'] - df_falt['Stock_AA_Total'], 0
)

# Criticidad — solo Bodega Farmacos es respaldo real para AT Abierta
def criticidad(row):
    bod_farm = row.get('Stock_BODEGA_FARMACOS', 0)
    pax      = row['Pacientes_Afectados']
    if bod_farm == 0 and pax >= 10: return '[CRITICO] CRITICO — SIN RESPALDO'
    if bod_farm == 0 and pax < 10:  return '[ALTO] ALTO — SIN RESPALDO'
    if bod_farm > 0  and pax >= 10: return '[MODERADO] MODERADO — TRASPASO BOD.FARMACOS'
    return '[BAJO] BAJO — TRASPASO BOD.FARMACOS'

df_falt['Criticidad'] = df_falt.apply(criticidad, axis=1)

# Accion sugerida — unico respaldo es Bodega Farmacos
def accion(row):
    stk_bf = row.get('Stock_BODEGA_FARMACOS', 0)
    if stk_bf > 0:
        return f'TRASPASAR DESDE BODEGA FARMACOS ({int(stk_bf)} ud. disponibles)'
    return 'COMPRA URGENTE — SIN RESPALDO EN BODEGA FARMACOS'

df_falt['Accion_Sugerida'] = df_falt.apply(accion, axis=1)
# Orden por severidad real (1=mas urgente) via crit_nivel, no por texto de
# 'Criticidad' — el string '[CRITICO]...' ordena alfabeticamente DESPUES de
# '[ALTO]'/'[BAJO]', lo que sacaba los faltantes mas urgentes del tope.
df_falt['_nivel_orden'] = df_falt['Criticidad'].apply(crit_nivel)
df_falt.sort_values(['_nivel_orden','Pacientes_Afectados'], ascending=[True, False], inplace=True)
df_falt.drop(columns=['_nivel_orden'], inplace=True)

print(f"  Faltantes reales AA: {len(df_falt):,}")

# ═══════════════════════════════════════════════
# 8b. FALTANTES ABSOLUTOS AA (30 días, Atención Abierta)
# ═══════════════════════════════════════════════
# Igual definición que Faltantes Reales AA (sec. 8) — prescripción vigente
# (PENDIENTE/SOLICITADO, no ANULADO/RECHAZADO/DEVUELTO) que aún no se ha podido
# entregar — pero acotado a: (a) los últimos 30 días de Fecha_Atencion, y
# (b) despacho por el mostrador de Atención Abierta (Bodega_Despacha ==
# FARMACIA AT ABIERTA). Stock_AA_Total == 0 confirma que es quiebre real, no
# solo demanda pendiente que ya se puede cubrir con lo que hay en bodega/farmacia.
cutoff_30 = HOY - pd.Timedelta(days=30)
mask_falt30 = (
    df_aa_univ['Estado_Prescripcion'].isin(['PENDIENTE', 'SOLICITADO']) &
    (df_aa_univ['Cantidad_Pendiente'] > 0) &
    (df_aa_univ['Fecha_Atencion'] >= cutoff_30) &
    (df_aa_univ['Bodega_Despacha'] == 'FARMACIA AT ABIERTA')
)
df_pend30 = df_aa_univ[mask_falt30].copy()

dem_agg30 = df_pend30.groupby('Prescripcion_norm').agg(
    Cant_Demanda_Activa=('Cantidad_Pendiente', 'sum'),
    Pacientes_Afectados=('RUN',               'nunique'),
    N_Recetas           =('Numero_Receta',     'nunique'),
).reset_index()

df_falt30 = dem_agg30.merge(
    df_master[['Medicamento','Stock_Farmacia_AA','Stock_Bodega_AA','Stock_AA_Total',
               'Stock_Hospital_Total'] +
              [f'Stock_{b.replace(" ","_")}' for b in ORDEN_TRASPASO]],
    left_on='Prescripcion_norm', right_on='Medicamento', how='left'
).fillna(0)

df_falt30 = df_falt30[df_falt30['Stock_AA_Total'] == 0].copy()
df_falt30['Faltante_Neto'] = np.maximum(
    df_falt30['Cant_Demanda_Activa'] - df_falt30['Stock_AA_Total'], 0
)
df_falt30['Criticidad'] = df_falt30.apply(criticidad, axis=1)
df_falt30['Accion_Sugerida'] = df_falt30.apply(accion, axis=1)
df_falt30['_nivel_orden'] = df_falt30['Criticidad'].apply(crit_nivel)
df_falt30.sort_values(['_nivel_orden','Pacientes_Afectados'], ascending=[True, False], inplace=True)
df_falt30.drop(columns=['_nivel_orden'], inplace=True)

print(f"  Faltantes absolutos AA (30d, Atencion Abierta): {len(df_falt30):,}")

# ═══════════════════════════════════════════════
# 8c. FALTANTES PERSISTENTES (60 días, se mantienen hasta hoy)
# ═══════════════════════════════════════════════
# Faltantes que llevan tiempo sin poder despacharse en el mostrador de Atención
# Abierta: demanda no cubierta (PENDIENTE/SOLICITADO, Cantidad_Pendiente > 0) en
# los últimos 60 días cuyo faltante SIGUE VIGENTE hoy — última prescripción sin
# cubrir dentro de los últimos RECIENTE_FALT_DIAS días. Se basa en la DEMANDA
# pendiente, NO en el stock reportado: casos como Empagliflozina 25 mg muestran
# stock "fantasma" en el reporte (>0) pero acumulan recetas sin entregar por
# semanas — el reporte de stock no refleja lo que hay físicamente en mesón.
VENTANA_FALT_60   = 60
RECIENTE_FALT_DIAS = 15
# Con stock reportado > 0 solo se considera faltante si la demanda pendiente es
# alta respecto de ese stock (fantasma / stock insuficiente que no despacha, como
# Empagliflozina 25 mg). Si el stock reportado cubre de sobra la demanda, los
# pendientes son de tránsito (paciente aún no retira) y NO se listan como quiebre.
FANTASMA_RATIO    = 0.25
cutoff_60       = HOY - pd.Timedelta(days=VENTANA_FALT_60)
cutoff_reciente = HOY - pd.Timedelta(days=RECIENTE_FALT_DIAS)
mask_falt60 = (
    df_aa_univ['Estado_Prescripcion'].isin(['PENDIENTE', 'SOLICITADO']) &
    (df_aa_univ['Cantidad_Pendiente'] > 0) &
    (df_aa_univ['Fecha_Atencion'] >= cutoff_60) &
    (df_aa_univ['Bodega_Despacha'] == 'FARMACIA AT ABIERTA')
)
df_pend60 = df_aa_univ[mask_falt60].copy()

dem_agg60 = df_pend60.groupby('Prescripcion_norm').agg(
    Cant_Demanda_Activa=('Cantidad_Pendiente', 'sum'),
    Pacientes_Afectados=('RUN',               'nunique'),
    N_Recetas           =('Numero_Receta',     'nunique'),
    Primer_Faltante     =('Fecha_Atencion',    'min'),
    Ultimo_Faltante     =('Fecha_Atencion',    'max'),
).reset_index()

# "Se mantiene hasta ahora": el último pendiente cae dentro de la ventana reciente
df_falt60 = dem_agg60[dem_agg60['Ultimo_Faltante'] >= cutoff_reciente].copy()
df_falt60['Dias_En_Falta'] = (HOY - df_falt60['Primer_Faltante']).dt.days

df_falt60 = df_falt60.merge(
    df_master[['Medicamento','Stock_Farmacia_AA','Stock_Bodega_AA','Stock_AA_Total',
               'Stock_Hospital_Total'] +
              [f'Stock_{b.replace(" ","_")}' for b in ORDEN_TRASPASO]],
    left_on='Prescripcion_norm', right_on='Medicamento', how='left'
).fillna(0)

# Filtro fantasma: stock 0 (quiebre real) o demanda pendiente ≥ FANTASMA_RATIO
# del stock reportado (stock que no despacha). Excluye meds bien abastecidos
# cuyos pendientes son solo pacientes que no han retirado.
df_falt60 = df_falt60[
    (df_falt60['Stock_AA_Total'] == 0) |
    (df_falt60['Cant_Demanda_Activa'] >= FANTASMA_RATIO * df_falt60['Stock_AA_Total'])
].copy()

df_falt60['Faltante_Neto'] = np.maximum(
    df_falt60['Cant_Demanda_Activa'] - df_falt60['Stock_AA_Total'], 0
)
df_falt60['Criticidad'] = df_falt60.apply(criticidad, axis=1)

def accion_60(row):
    stk_bf = row.get('Stock_BODEGA_FARMACOS', 0)
    stk_aa = row.get('Stock_AA_Total', 0)
    if stk_aa > 0:
        return ('REVISAR: stock reportado NO se despacha '
                '(posible fantasma) — verificar físico en mesón')
    if stk_bf > 0:
        return f'TRASPASAR DESDE BODEGA FARMACOS ({int(stk_bf)} ud. disponibles)'
    return 'COMPRA URGENTE — SIN RESPALDO EN BODEGA FARMACOS'

df_falt60['Accion_Sugerida'] = df_falt60.apply(accion_60, axis=1)
# Orden: primero los que llevan más días en falta, luego por pacientes afectados
df_falt60.sort_values(['Dias_En_Falta','Pacientes_Afectados'],
                      ascending=[False, False], inplace=True)
df_falt60['Primer_Faltante'] = df_falt60['Primer_Faltante'].dt.strftime('%Y-%m-%d')
df_falt60['Ultimo_Faltante'] = df_falt60['Ultimo_Faltante'].dt.strftime('%Y-%m-%d')

print(f"  Faltantes persistentes AA (60d, vigentes): {len(df_falt60):,}")

# ═══════════════════════════════════════════════
# 9. PENDIENTES 15D
# ═══════════════════════════════════════════════
cutoff_15 = HOY - pd.Timedelta(days=15)
mask_p15 = (
    (df_aa_univ['Estado_Prescripcion'] == 'PENDIENTE') &
    (df_aa_univ['Fecha_Atencion'] >= cutoff_15) &
    (df_aa_univ['Cantidad_Pendiente'] > 0) &
    (df_aa_univ['Bodega_Despacha'] == 'FARMACIA AT ABIERTA')
)
df_p15 = df_aa_univ[mask_p15].copy()
df_p15['Paciente'] = (df_p15['Nombre'].fillna('') + ' ' +
                      df_p15['Apellido_Paterno'].fillna('')).str.strip()
df_p15 = df_p15[['Fecha_Atencion','RUN','Paciente','Prescripcion',
                  'Prescripcion_norm','Cantidad_Pendiente',
                  'Bodega_Despacha','Estado_Prescripcion',
                  'Numero_Receta','Numero_Folio']].copy()
df_p15.sort_values(['Fecha_Atencion'], inplace=True)
print(f"  Pendientes 15d: {len(df_p15):,}")

# ═══════════════════════════════════════════════
# 10. QUIEBRES TOTALES
# ═══════════════════════════════════════════════
df_quiebres = df_master[df_master['Quiebre_Total']].copy()
df_quiebres['Instruccion'] = df_quiebres.apply(
    lambda r: accion(r), axis=1
)
df_quiebres.sort_values('CDL', ascending=False, inplace=True)
print(f"  Quiebres totales AA: {len(df_quiebres):,}")

# ═══════════════════════════════════════════════
# 11. REPOSICION 5D
# ═══════════════════════════════════════════════
df_repos = df_master[df_master['Reposicion_Sugerida'] > 0].copy()
df_repos.sort_values('Reposicion_Sugerida', ascending=False, inplace=True)

# ═══════════════════════════════════════════════
# 12. CONSUMO MENSUAL POR MEDICAMENTO (Fecha_Entrega)
# ═══════════════════════════════════════════════
# Pivot: filas = medicamento, columnas = año-mes de entrega
df_op_consumo['Mes_Entrega'] = df_op_consumo['Fecha_Entrega'].dt.to_period('M')
monthly_raw = (
    df_op_consumo
    .groupby(['Prescripcion_norm', 'Mes_Entrega'])['Cantidad_Recetada']
    .sum()
    .unstack(fill_value=0)
)
# Asegurar que todos los medicamentos del universo estén presentes
monthly_raw = monthly_raw.reindex(idx, fill_value=0)
# Columnas: convertir Period a string legible "Ene-2026" etc.
month_names = {str(p): p.strftime('%b-%Y') for p in monthly_raw.columns}
monthly_raw.columns = [month_names[str(c)] for c in monthly_raw.columns]

# Totales y métricas derivadas
monthly_raw['Total_Periodo']   = monthly_raw.sum(axis=1)
monthly_raw['Meses_Periodo']   = round(MESES_TOT_OP, 4)
monthly_raw['Dias_Lab_Periodo']= DIAS_LAB_OP
monthly_raw['CMP_Mensual']     = (monthly_raw['Total_Periodo'] / MESES_TOT_OP).round(2)
monthly_raw['CDL']             = (monthly_raw['Total_Periodo'] / DIAS_LAB_OP).round(3)

# Añadir stock AA para contexto
monthly_raw = monthly_raw.join(
    df_master.set_index('Medicamento')[['Stock_Farmacia_AA','Stock_Bodega_AA',
                                        'Stock_AA_Total','Cobertura_Lab']],
    how='left'
)

df_consumo_mensual = monthly_raw.reset_index().rename(
    columns={'Prescripcion_norm': 'Medicamento'}
)
# Ordenar por medicamento
df_consumo_mensual.sort_values('Medicamento', inplace=True)
print(f"  Tabla consumo mensual: {len(df_consumo_mensual):,} medicamentos x {len(monthly_raw.columns)} columnas")

# Validacion rapida en consola
val_meds = ['INSULINA GLARGINA 300', 'INSULINA ASPARTA 300']
for vm in val_meds:
    rows_v = df_consumo_mensual[df_consumo_mensual['Medicamento'].str.contains(vm, case=False)]
    for _, rv in rows_v.iterrows():
        mes_cols = [c for c in df_consumo_mensual.columns
                    if c not in ('Medicamento','Total_Periodo','Meses_Periodo',
                                 'Dias_Lab_Periodo','CMP_Mensual','CDL',
                                 'Stock_Farmacia_AA','Stock_Bodega_AA',
                                 'Stock_AA_Total','Cobertura_Lab')]
        detalle = '  |  '.join(f"{c}: {int(rv[c])}" for c in mes_cols)
        print(f"  {rv['Medicamento'][:45]}")
        print(f"    {detalle}")
        print(f"    CMP={rv['CMP_Mensual']:.2f}  CDL={rv['CDL']:.3f}  Total={rv['Total_Periodo']:.0f}")

# ═══════════════════════════════════════════════
# 13. TENDENCIA SEMANAL (semana del mes por Fecha_Entrega)
# ═══════════════════════════════════════════════
# Semana del mes: día 1-7 → S1, 8-14 → S2, 15-21 → S3, 22-31 → S4
df_op_consumo['Semana_Mes']  = ((df_op_consumo['Fecha_Entrega'].dt.day - 1) // 7 + 1).clip(upper=4)
df_op_consumo['Mes_Period']  = df_op_consumo['Fecha_Entrega'].dt.to_period('M')
df_op_consumo['Mes_Label']   = df_op_consumo['Fecha_Entrega'].dt.strftime('%b-%Y')

# ── A) Pivot detallado: medicamento × "Mes-S#" ──────────────────────────────
weekly_detail = (
    df_op_consumo
    .groupby(['Prescripcion_norm', 'Mes_Label', 'Semana_Mes', 'Mes_Period'],
             observed=True)['Cantidad_Recetada']
    .sum()
    .reset_index()
)
weekly_detail['Col_Key']   = list(zip(                             # para ordenar
    weekly_detail['Mes_Period'].astype(str),
    weekly_detail['Semana_Mes']
))
weekly_detail['Col_Label'] = weekly_detail['Mes_Label'] + ' S' + weekly_detail['Semana_Mes'].astype(str)

wp = weekly_detail.pivot_table(
    index='Prescripcion_norm', columns='Col_Label',
    values='Cantidad_Recetada', fill_value=0, aggfunc='sum'
)

# Ordenar columnas cronológicamente (año, mes, semana)
def _col_sort(c):
    try:
        mes_str, s = c.split(' S')
        p = pd.Period(mes_str, freq='M')
        return (p.year, p.month, int(s))
    except Exception:
        return (9999, 99, 9)

sorted_detail_cols = sorted(wp.columns, key=_col_sort)
wp = wp[sorted_detail_cols].reindex(idx, fill_value=0)

# ── B) Patrón global: total por semana (S1-S4 sumados todos los meses) ──────
patron_global = (
    df_op_consumo
    .groupby(['Prescripcion_norm', 'Semana_Mes'], observed=True)['Cantidad_Recetada']
    .sum()
    .unstack(fill_value=0)
    .reindex(idx, fill_value=0)
)
patron_global.columns = [f'Total_S{c}' for c in patron_global.columns]

# ── C) Distribución porcentual por semana ────────────────────────────────────
pat_tot   = patron_global.sum(axis=1).replace(0, np.nan)
patron_pct = patron_global.div(pat_tot, axis=0).fillna(0).mul(100).round(1)
patron_pct.columns = [c.replace('Total_', '%_') for c in patron_global.columns]

# ── D) Semana pico (mayor consumo global) ────────────────────────────────────
# apply condicional para evitar ValueError en filas con todo cero
semana_pico = patron_global.apply(
    lambda r: r.idxmax().replace('Total_', '') if r.sum() > 0 else 'Sin_Consumo',
    axis=1
)

# ── E) Construir DataFrame final ─────────────────────────────────────────────
df_tendencia = pd.concat([
    wp,
    patron_global,
    patron_pct,
    semana_pico.rename('Semana_Pico'),
], axis=1).reset_index().rename(columns={'Prescripcion_norm': 'Medicamento'})

# Unir stock y CMP para contexto
df_tendencia = df_tendencia.merge(
    df_master[['Medicamento', 'CMP_Mensual', 'CDL',
               'Stock_Farmacia_AA', 'Cobertura_Lab']],
    on='Medicamento', how='left'
)
df_tendencia.sort_values('Medicamento', inplace=True)

print(f"  Tabla tendencia semanal: {len(df_tendencia):,} medicamentos x {len(df_tendencia.columns)} columnas")

# Extraer % por semana para usar en pedidos (columnas %_S1 … %_S4)
_pct_cols = [c for c in df_tendencia.columns if c.startswith('%_S')]
df_pct_semana = df_tendencia[['Medicamento'] + _pct_cols].copy()

# Validacion en consola: top 10 medicamentos por CMP + patron semanal
top10 = df_master.nlargest(10, 'CMP_Mensual')['Medicamento'].tolist()
print(f"\n{'Medicamento':<48} {'Sem_Pico':<8} {'%S1':>6} {'%S2':>6} {'%S3':>6} {'%S4':>6}  CMP")
print('-' * 90)
for med in top10:
    row = df_tendencia[df_tendencia['Medicamento'] == med]
    if row.empty:
        continue
    r = row.iloc[0]
    pct_cols = [c for c in df_tendencia.columns if c.startswith('%_')]
    pcts = [f"{r[c]:5.1f}%" for c in pct_cols]
    print(f"  {med[:46]:<46} {str(r.get('Semana_Pico','')):<8} "
          f"{'  '.join(pcts)}  {r['CMP_Mensual']:.1f}")

# ═══════════════════════════════════════════════
# 14. PEDIDOS AA — FARMACIA ↔ BODEGA  (consumo ajustado por tendencia semanal)
# ═══════════════════════════════════════════════
# Objetivos de cobertura (días hábiles)
TARGET_COB_FARMACIA_DIAS  = 5   # 1 semana laboral en farmacia
CICLO_PEDIDO_BODEGA_DIAS  = 10  # Bodega AA pide a Bod. Fármacos cada 2 semanas = 10 días háb.
TARGET_COB_BODEGA_DIAS    = CICLO_PEDIDO_BODEGA_DIAS

# ── Regla de reposición Farmacia AA por ROTACIÓN (decisión usuario 2026-06-18) ──
# Alta rotación (CDL ≥ UMBRAL) → cobertura 2 días hábiles (se repone seguido desde
# Bodega AA, en sitio). Baja rotación → 5 días, con piso de 1 empaque (el redondeo
# por factor de empaque garantiza ≥ 1 empaque cuando hay necesidad > 0).
UMBRAL_ALTA_ROTACION_CDL  = 30   # ud/día hábil
COB_ALTA_ROTACION_DIAS    = 2
COB_BAJA_ROTACION_DIAS    = 5

# Factor de empaque CENABAST (para aproximar todos los pedidos a empaques completos)
FACTOR_EMPAQUE = cargar_factores_empaque(WORK_DIR)
_med_series_fe = df_master['Medicamento'] if 'Medicamento' in df_master.columns else df_master.index.to_series()
df_master['Factor_Empaque'] = [FACTOR_EMPAQUE.get(_clave_empaque(m), 1) for m in _med_series_fe]
_n_con_fe = int((df_master['Factor_Empaque'] > 1).sum())
print(f"  Factor de empaque CENABAST: {len(FACTOR_EMPAQUE)} claves · {_n_con_fe}/{len(df_master)} meds con empaque")

# ── A) CDL ajustado por semana del mes ───────────────────────────────────────
# Fórmula: CDL_Sx = CDL × (% consumo semana x / 25)
# Baseline = 25 % (distribución plana entre 4 semanas)
# Si un medicamento tiene pico S4 (28 %), CDL_S4 = CDL × 1.12 (+12 %)
# Si tiene semana baja S1 (21 %), CDL_S1 = CDL × 0.84 (-16 %)

df_ped = df_master[df_master['CDL'] > 0].copy()
df_ped = df_ped.merge(df_pct_semana, on='Medicamento', how='left')

for s in range(1, 5):
    col_pct = f'%_S{s}'
    if col_pct not in df_ped.columns:
        df_ped[col_pct] = 25.0
    df_ped[col_pct] = df_ped[col_pct].fillna(25.0)
    df_ped[f'CDL_S{s}'] = (df_ped['CDL'] * df_ped[col_pct] / 25.0).round(4)

# ── B) Días hábiles próximos 5 y 10 jornadas → semana del mes ───────────────
def _semana_mes(d):
    return min((d.day - 1) // 7 + 1, 4)      # día 1-7→S1, 8-14→S2, 15-21→S3, 22+→S4

next_5_bdays  = [(d, _semana_mes(d)) for d in pd.bdate_range(HOY, periods=5)]
next_10_bdays = [(d, _semana_mes(d)) for d in pd.bdate_range(HOY, periods=10)]

print(f"\n  Proximos 5 dias hab (tendencia):")
for d, s in next_5_bdays:
    print(f"    {d.strftime('%a %d-%b')}  →  S{s}")

# ── C) Consumo proyectado por tendencia ──────────────────────────────────────
_s2_list  = [s for _, s in next_5_bdays][:2]   # próximos 2 días hábiles (alta rotación)
_s5_list  = [s for _, s in next_5_bdays]
_s10_list = [s for _, s in next_10_bdays]

# Consumo_xD_Trend = suma CDL_Sx para cada uno de los próximos x días hábiles
df_ped['Consumo_2D_Trend']  = df_ped.apply(
    lambda r: round(sum(r[f'CDL_S{s}'] for s in _s2_list), 1), axis=1)
df_ped['Consumo_5D_Trend']  = df_ped.apply(
    lambda r: round(sum(r[f'CDL_S{s}'] for s in _s5_list), 1), axis=1)
df_ped['Consumo_10D_Trend'] = df_ped.apply(
    lambda r: round(sum(r[f'CDL_S{s}'] for s in _s10_list), 1), axis=1)

# Factor de carga actual: cuánto más/menos consume respecto al CDL plano
df_ped['Factor_Carga_5D'] = (df_ped['Consumo_5D_Trend'] /
                               (df_ped['CDL'] * 5).replace(0, np.nan)).round(3).fillna(1)

# ── D) Cobertura actual (en CDL plano — referencia histórica) ────────────────
df_ped['Cob_Farm_Dias']  = (df_ped['Stock_Farmacia_AA'] / df_ped['CDL']).round(1)
df_ped['Cob_Bod_Dias']   = (df_ped['Stock_Bodega_AA']   / df_ped['CDL']).round(1)
df_ped['Cob_Total_Dias'] = (df_ped['Stock_AA_Total']    / df_ped['CDL']).round(1)

# ── E) Necesidad Farmacia — regla por ROTACIÓN + SOLO no-diálisis + empaque ───
# Cobertura objetivo según rotación: 2 días (alta) / 5 días (baja). Usa el consumo
# proyectado con tendencia (semana pico) para esa ventana.
df_ped['Cobertura_Obj_Dias'] = np.where(
    df_ped['CDL'] >= UMBRAL_ALTA_ROTACION_CDL, COB_ALTA_ROTACION_DIAS, COB_BAJA_ROTACION_DIAS)
df_ped['Requerimiento_Farm'] = np.where(
    df_ped['CDL'] >= UMBRAL_ALTA_ROTACION_CDL,
    df_ped['Consumo_2D_Trend'], df_ped['Consumo_5D_Trend'])

# El pedido normal cubre SOLO la demanda no-diálisis (diálisis se pide aparte,
# mensual). Se descuenta la fracción de consumo atribuible a recetas de diálisis.
_frac_dial = (df_ped['Prescrito_Dialisis_Op'] /
              df_ped['Prescrito_Total_Op'].replace(0, np.nan)).fillna(0).clip(0, 1)
df_ped['Requerimiento_Farm_NoDial'] = (df_ped['Requerimiento_Farm'] * (1 - _frac_dial)).round(1)

# Necesidad bruta = requerimiento no-diálisis − stock Farmacia AA; luego redondeo
# al factor de empaque (garantiza ≥ 1 empaque cuando hay necesidad > 0 = el "piso").
_nec_farm_bruta = np.maximum(
    df_ped['Requerimiento_Farm_NoDial'] - df_ped['Stock_Farmacia_AA'], 0).round(1)
df_ped['Necesidad_Farm'] = [
    redondear_empaque(n, m, FACTOR_EMPAQUE)
    for n, m in zip(_nec_farm_bruta, df_ped['Medicamento'])
]

df_ped['Traspaso_Posible'] = np.minimum(
    df_ped['Necesidad_Farm'], df_ped['Stock_Bodega_AA']
).round(0)

df_ped['Stock_Farm_Post']  = (df_ped['Stock_Farmacia_AA'] + df_ped['Traspaso_Posible']).round(0)
df_ped['Cob_Farm_Post']    = (df_ped['Stock_Farm_Post']   / df_ped['CDL']).round(1)
df_ped['Stock_Bod_Post']   = (df_ped['Stock_Bodega_AA']   - df_ped['Traspaso_Posible']).round(0)
df_ped['Cob_Bod_Post']     = (df_ped['Stock_Bod_Post']    / df_ped['CDL']).round(1)
df_ped['Deficit_Farm']     = np.maximum(
    df_ped['Necesidad_Farm'] - df_ped['Stock_Bodega_AA'], 0
).round(0)

# ── F) Criticidad Farmacia ───────────────────────────────────────────────────
def crit_farm(row):
    cob = row['Cob_Farm_Dias']
    nec = row['Necesidad_Farm']
    bod = row['Stock_Bodega_AA']
    if nec == 0:                        return '5-OK'
    if cob == 0 and bod == 0:          return '1-CRITICO'
    if cob == 0 or cob < 2:            return '2-URGENTE'
    if cob < TARGET_COB_FARMACIA_DIAS: return '3-MODERADO'
    return '4-BAJO'

df_ped['Crit_Farm'] = df_ped.apply(crit_farm, axis=1)

def accion_farm_traspaso(row):
    """Accion 1: traspaso desde Bodega AA hacia Farmacia AA."""
    nec  = int(row['Necesidad_Farm'])
    tras = int(row['Traspaso_Posible'])
    if nec == 0:  return ''
    if tras > 0:  return f'TRASPASAR {tras} ud. DESDE BODEGA AA'
    return 'SIN STOCK EN BODEGA AA'

def accion_farm_externo(row):
    """Accion 2: gestion externa (compra/traslado) cuando bodega AA no alcanza."""
    nec  = int(row['Necesidad_Farm'])
    def_ = int(row['Deficit_Farm'])
    if nec == 0 or def_ == 0: return ''
    return f'GESTIONAR EXTERNO {def_} ud.'

df_ped['Accion_Traspaso'] = df_ped.apply(accion_farm_traspaso, axis=1)
df_ped['Accion_Externo']  = df_ped.apply(accion_farm_externo,  axis=1)

# ── G) Necesidad Bodega → Bodega Fármacos (decisión usuario 2026-06-18) ──────
# Pedido = requerimiento de 10 días hábiles (con factor de semana pico)
#          − TODO el stock de Atención Abierta (Farmacia AA + Bodega AA).
# Así no se pide a Bodega Fármacos lo que Atención Abierta ya tiene en mano.
# Luego se redondea al factor de empaque.
df_ped['Req_2_Semanas']    = (df_ped['CDL'] * CICLO_PEDIDO_BODEGA_DIAS).round(0)
df_ped['Target_Stock_Bod'] = df_ped['Consumo_10D_Trend']   # 10 días háb. ajustado por tendencia
_nec_bod_bruta = np.maximum(
    df_ped['Consumo_10D_Trend'] - df_ped['Stock_AA_Total'], 0
).round(1)
df_ped['Necesidad_Bod'] = [
    redondear_empaque(n, m, FACTOR_EMPAQUE)
    for n, m in zip(_nec_bod_bruta, df_ped['Medicamento'])
]

# ── H) Criticidad Bodega — umbrales basados en el ciclo de 2 semanas (10 días háb.) ──
def crit_bod(row):
    cob = row['Cob_Bod_Post']
    nec = row['Necesidad_Bod']
    if nec == 0:                             return '5-OK'
    if cob == 0:                             return '1-CRITICO'  # sin stock
    if cob < 3:                              return '2-URGENTE'  # < 3 días: pedir hoy
    if cob < 5:                              return '3-ALTO'     # < 1 semana laboral
    if cob < CICLO_PEDIDO_BODEGA_DIAS:       return '4-MODERADO' # < 1 ciclo (2 semanas)
    return '5-BAJO'

df_ped['Crit_Bod'] = df_ped.apply(crit_bod, axis=1)

# ── Fuentes para reponer Bodega AA ───────────────────────────────────────────
# Unico respaldo para Bodega AA es BODEGA FARMACOS
# Accion 1: traspaso desde Bodega Farmacos
# Accion 2: compra externa por lo que no cubre Bodega Farmacos
df_ped['Stock_BodFarm_Disp'] = (
    df_ped['Stock_BODEGA_FARMACOS'].fillna(0)
    if 'Stock_BODEGA_FARMACOS' in df_ped.columns
    else pd.Series(0, index=df_ped.index)
)

df_ped['Traspaso_Hospital_Bod'] = np.minimum(
    df_ped['Necesidad_Bod'], df_ped['Stock_BodFarm_Disp']
).round(0)
df_ped['Compra_Externa_Bod'] = np.maximum(
    df_ped['Necesidad_Bod'] - df_ped['Stock_BodFarm_Disp'], 0
).round(0)

def accion_bod_traspaso(row):
    """Accion 1: traspaso desde Bodega Farmacos hacia Bodega AA."""
    nec  = int(row['Necesidad_Bod'])
    tras = int(row['Traspaso_Hospital_Bod'])
    if nec == 0:  return ''
    if tras > 0:
        return f'TRASPASAR {tras} ud. DESDE BODEGA FARMACOS'
    return 'SIN STOCK EN BODEGA FARMACOS'

def accion_bod_compra(row):
    """Accion 2: compra externa cuando Bodega Farmacos no alcanza."""
    comp = int(row['Compra_Externa_Bod'])
    if comp == 0: return ''
    return f'COMPRA EXTERNA {comp} ud.'

df_ped['Accion_Traspaso_Hosp'] = df_ped.apply(accion_bod_traspaso, axis=1)
df_ped['Accion_Compra_Ext']    = df_ped.apply(accion_bod_compra,   axis=1)

# ── I) DataFrames para hojas ─────────────────────────────────────────────────
_farm_cols = {
    'Criticidad'               : 'Crit_Farm',
    'Semana_Pico_Hist'         : 'Semana_Pico',
    'Factor_Carga_5D'          : 'Factor_Carga_5D',
    'Stock_Farm_Actual'        : 'Stock_Farmacia_AA',
    'Cob_Farm_Actual_Dias'     : 'Cob_Farm_Dias',
    'Consumo_5D_Trend'         : 'Consumo_5D_Trend',
    'Consumo_5D_Plano'         : 'Consumo_5D',
    'Necesidad_5D_Farm'        : 'Necesidad_Farm',
    'Stock_Bodega_Disponible'  : 'Stock_Bodega_AA',
    'A_Traspasar'              : 'Traspaso_Posible',
    'Deficit_Post_Traspaso'    : 'Deficit_Farm',
    'Stock_Farm_Post_Traspaso' : 'Stock_Farm_Post',
    'Cob_Farm_Post_Dias'       : 'Cob_Farm_Post',
    'CDL_DiasHab'              : 'CDL',
    'CMP_Mensual_22d'          : 'CMP_Mensual',
    'Accion_1_Traspaso_Bodega' : 'Accion_Traspaso',   # columna 1: traspaso interno
    'Accion_2_Gestion_Externa' : 'Accion_Externo',    # columna 2: gestion externa (solo si hay deficit)
}

# Añadir Semana_Pico desde df_tendencia
_sp_map = df_tendencia.set_index('Medicamento')['Semana_Pico'].to_dict()
df_ped['Semana_Pico'] = df_ped['Medicamento'].map(_sp_map).fillna('-')

df_farm_pedido = (
    df_ped[df_ped['Necesidad_Farm'] > 0]
    .sort_values(['Crit_Farm', 'CDL'], ascending=[True, False])
    [['Medicamento'] + list(_farm_cols.values())]
    .rename(columns={v: k for k, v in _farm_cols.items()})
    .reset_index(drop=True)
)

# Universo completo para la hoja Pedido_Farm_Bodega (consumida por pedido_fusion.py
# --todos): df_farm_pedido de arriba queda filtrado a Necesidad_Farm>0 a propósito
# para el Resumen_Pedidos_AA (solo lo accionable). Acá se arma el equivalente de
# _bod_pedido_sin_consumo pero para Farmacia, para que --todos pueda listar los
# 378 medicamentos igual que ya hace Pedido_Repos_Bodega.
_farm_pedido_con_datos = (
    df_ped[['Medicamento'] + list(_farm_cols.values())]
    .rename(columns={v: k for k, v in _farm_cols.items()})
)
_df_master_sin_datos_farm = df_master[~df_master['Medicamento'].isin(df_ped['Medicamento'])]
_farm_pedido_sin_datos = pd.DataFrame({
    'Medicamento'               : _df_master_sin_datos_farm['Medicamento'].values,
    'Criticidad'                : '5-OK',
    'Semana_Pico_Hist'          : _df_master_sin_datos_farm['Medicamento'].map(_sp_map).fillna('-').values,
    'Factor_Carga_5D'           : 1.0,
    'Stock_Farm_Actual'         : _df_master_sin_datos_farm['Stock_Farmacia_AA'].values,
    'Cob_Farm_Actual_Dias'      : 0.0,
    'Consumo_5D_Trend'          : 0.0,
    'Consumo_5D_Plano'          : 0.0,
    'Necesidad_5D_Farm'         : 0,
    'Stock_Bodega_Disponible'   : _df_master_sin_datos_farm['Stock_Bodega_AA'].values,
    'A_Traspasar'               : 0,
    'Deficit_Post_Traspaso'     : 0,
    'Stock_Farm_Post_Traspaso'  : _df_master_sin_datos_farm['Stock_Farmacia_AA'].values,
    'Cob_Farm_Post_Dias'        : 0.0,
    'CDL_DiasHab'               : 0.0,
    'CMP_Mensual_22d'           : 0.0,
    'Accion_1_Traspaso_Bodega'  : '',
    'Accion_2_Gestion_Externa'  : '',
})
df_farm_pedido_completo = (
    pd.concat([_farm_pedido_con_datos, _farm_pedido_sin_datos], ignore_index=True)
    .sort_values(['Criticidad', 'CDL_DiasHab'], ascending=[True, False])
    .reset_index(drop=True)
)

_bod_cols = {
    'Criticidad'              : 'Crit_Bod',
    'Semana_Pico_Hist'        : 'Semana_Pico',
    'Factor_Carga_5D'         : 'Factor_Carga_5D',
    'Stock_Bod_Actual'        : 'Stock_Bodega_AA',
    'Stock_Farm_Actual'       : 'Stock_Farmacia_AA',    # stock que Atención Abierta ya tiene en Farmacia
    'Cob_Bod_Actual_Dias'     : 'Cob_Bod_Dias',
    'Comprometido_Farm'       : 'Traspaso_Posible',
    'Stock_Bod_Post_Traspaso' : 'Stock_Bod_Post',
    'Cob_Bod_Post_Dias'       : 'Cob_Bod_Post',
    'Req_2_Semanas'           : 'Req_2_Semanas',         # CDL × 10 días = consumo plano 2 semanas
    'Consumo_10D_Trend'       : 'Consumo_10D_Trend',     # consumo 2 semanas ajustado por tendencia
    'Reponer_Bodega'          : 'Necesidad_Bod',
    'Stock_BODEGA_FARMACOS'   : 'Stock_BODEGA_FARMACOS',   # respaldo real (unico origen valido)
    'Stock_Hospital_Total'    : 'Stock_Hospital_Total',    # referencia: suma de todas las bodegas hospital
    'CDL_DiasHab'             : 'CDL',
    'CMP_Mensual_22d'         : 'CMP_Mensual',
    'Accion_1_Traspaso_Hospital' : 'Accion_Traspaso_Hosp',  # desde bodegas hospital
    'Accion_2_Compra_Externa'    : 'Accion_Compra_Ext',     # compra externa por deficit
}
_bod_pedido_con_consumo = (
    df_ped[['Medicamento'] + list(_bod_cols.values())]
    .rename(columns={v: k for k, v in _bod_cols.items()})
)

# Medicamentos del universo AA sin consumo reciente (CDL=0, fuera de df_ped):
# igual deben aparecer en la hoja Bodega AA -> Bodega Fármacos (universo completo)
# para poder iniciar un pedido manual, aunque el sistema no calcule necesidad.
_df_master_sin_consumo = df_master[~df_master['Medicamento'].isin(df_ped['Medicamento'])]
_bod_pedido_sin_consumo = pd.DataFrame({
    'Medicamento'                : _df_master_sin_consumo['Medicamento'].values,
    'Criticidad'                 : '5-OK',
    'Semana_Pico_Hist'           : _df_master_sin_consumo['Medicamento'].map(_sp_map).fillna('-').values,
    'Factor_Carga_5D'            : 1.0,
    'Stock_Bod_Actual'           : _df_master_sin_consumo['Stock_Bodega_AA'].values,
    'Stock_Farm_Actual'          : _df_master_sin_consumo['Stock_Farmacia_AA'].values,
    'Cob_Bod_Actual_Dias'        : 0.0,
    'Comprometido_Farm'          : 0,
    'Stock_Bod_Post_Traspaso'    : _df_master_sin_consumo['Stock_Bodega_AA'].values,
    'Cob_Bod_Post_Dias'          : 0.0,
    'Req_2_Semanas'              : 0,
    'Consumo_10D_Trend'          : 0.0,
    'Reponer_Bodega'             : 0,
    'Stock_BODEGA_FARMACOS'      : _df_master_sin_consumo['Stock_BODEGA_FARMACOS'].values,
    'Stock_Hospital_Total'       : _df_master_sin_consumo['Stock_Hospital_Total'].values,
    'CDL_DiasHab'                : 0.0,
    'CMP_Mensual_22d'            : 0.0,
    'Accion_1_Traspaso_Hospital' : '',
    'Accion_2_Compra_Externa'    : '',
})

df_bod_pedido = (
    pd.concat([_bod_pedido_con_consumo, _bod_pedido_sin_consumo], ignore_index=True)
    .sort_values(['Criticidad', 'CDL_DiasHab'], ascending=[True, False])
    .reset_index(drop=True)
)

print(f"  Pedido Farmacia <- Bodega AA : {len(df_farm_pedido):,} medicamentos")
print(f"  Pedido Reposicion Bodega AA  : {len(df_bod_pedido):,} medicamentos")

# ═══════════════════════════════════════════════
# 15. PEDIDOS DIALISIS — solo recetas de nefrólogos
#     Mismo motor y escala de criticidad que sección 14.
#     La demanda combina: consumo diálisis + consumo farmacia sin diálisis
#     (evita quiebre en stock compartido FARMACIA AT ABIERTA).
# ═══════════════════════════════════════════════
# El universo y el consumo se restringen a las recetas firmadas por los
# medicos de MEDICOS_DIALISIS. El stock es el compartido (no se separa por
# medico). Se reutilizan los helpers crit_farm/crit_bod/accion_* y los mapeos
# de columnas _farm_cols/_bod_cols ya definidos arriba → una sola fuente de
# verdad para criticidad y acciones.

df_dial = df_op_consumo[df_op_consumo['Prof_Norm'].isin(MEDICOS_DIALISIS)].copy()

# Consumo dialisis (Cantidad_Recetada) en el periodo completo, por medicamento
cd_total_dial = consumo_periodo(df_dial, FECHA_INICIO_OP, FECHA_MAX)
idx_dial = sorted([m for m in cd_total_dial.index if cd_total_dial[m] > 0])

# Master dialisis = stock compartido (df_master) + CDL/CMP de SOLO estas recetas
_stock_cols_dial = ['Medicamento','Stock_Farmacia_AA','Stock_Bodega_AA','Stock_AA_Total',
                    'Stock_BODEGA_FARMACOS','Stock_Hospital_Total','Prescrito_Total_Op',
                    'Factor_Empaque']
df_md = df_master[df_master['Medicamento'].isin(idx_dial)][
    [c for c in _stock_cols_dial if c in df_master.columns]
].copy()
df_md = df_md.merge(cd_total_dial.rename('Prescrito_Dialisis'),
                    left_on='Medicamento', right_index=True, how='left')
df_md['Prescrito_Dialisis'] = df_md['Prescrito_Dialisis'].fillna(0)

# ── Demanda COMBINADA (decisión usuario 2026-06-15) ─────────────────────────
# Para disminuir el riesgo de quiebre en diálisis sobre el stock COMPARTIDO,
# el pedido de diálisis cubre: consumo de la farmacia EXCLUYENDO diálisis +
# consumo de diálisis. Como las recetas de diálisis son un subconjunto del
# total, esto equivale al consumo total real (sin doble conteo). Se escribe de
# forma aditiva para que el desglose quede auditable en las hojas de diálisis.
df_md['Prescrito_Farm_NoDial'] = (df_md.get('Prescrito_Total_Op', 0) - df_md['Prescrito_Dialisis']).clip(lower=0)
df_md['Prescrito_Combinado']   = df_md['Prescrito_Farm_NoDial'] + df_md['Prescrito_Dialisis']
df_md['CDL'] = np.where(DIAS_LAB_OP > 0,
                        df_md['Prescrito_Combinado'] / DIAS_LAB_OP, 0).round(4)
df_md['CMP_Mensual'] = (df_md['CDL'] * DIAS_LAB_MES).round(2)

def _pipeline_ped_dial(df_base):
    """Mismo flujo que la seccion 14, aplicado al master de dialisis.
    Reusa los helpers crit_farm/crit_bod/accion_* y _farm_cols/_bod_cols."""
    dp = df_base[df_base['CDL'] > 0].copy()
    dp = dp.merge(df_pct_semana, on='Medicamento', how='left')
    for s in range(1, 5):
        col_pct = f'%_S{s}'
        if col_pct not in dp.columns:
            dp[col_pct] = 25.0
        dp[col_pct] = dp[col_pct].fillna(25.0)
        dp[f'CDL_S{s}'] = (dp['CDL'] * dp[col_pct] / 25.0).round(4)

    dp['Consumo_5D']        = (dp['CDL'] * 5).round(1)   # plano COMBINADO (farm no-dial + dial)
    # Desglose de referencia (solo se muestra en las hojas de diálisis): cuánto
    # del consumo de 5 días corresponde a diálisis y cuánto a la farmacia general.
    _cdl_dial = np.where(DIAS_LAB_OP > 0, dp['Prescrito_Dialisis']     / DIAS_LAB_OP, 0)
    _cdl_fnd  = np.where(DIAS_LAB_OP > 0, dp['Prescrito_Farm_NoDial']  / DIAS_LAB_OP, 0)
    dp['Consumo_5D_Dial']       = (_cdl_dial * 5).round(1)
    dp['Consumo_5D_FarmNoDial'] = (_cdl_fnd  * 5).round(1)
    # Requerimiento mensual de diálisis en DÍAS CORRIDOS (30d), no días hábiles —
    # mismo criterio que pedido_fusion.py Hoja 3 (el pedido de diálisis es mensual
    # calendario, no ligado al ciclo de reposición hábil de Farm_Bod/Bod_Farmacos).
    dp['Requerimiento_Mensual_Dial_30d'] = (dp['Consumo_5D_Dial'] / 5 * 30).round(0)
    dp['Consumo_5D_Trend']  = dp.apply(
        lambda r: round(sum(r[f'CDL_S{s}'] for s in _s5_list), 1), axis=1)
    dp['Consumo_10D_Trend'] = dp.apply(
        lambda r: round(sum(r[f'CDL_S{s}'] for s in _s10_list), 1), axis=1)
    dp['Consumo_2D_Trend']  = dp.apply(
        lambda r: round(sum(r[f'CDL_S{s}'] for s in _s2_list), 1), axis=1)
    dp['Factor_Carga_5D'] = (dp['Consumo_5D_Trend'] /
                              (dp['CDL'] * 5).replace(0, np.nan)).round(3).fillna(1)

    dp['Cob_Farm_Dias'] = (dp['Stock_Farmacia_AA'] / dp['CDL']).round(1)
    dp['Cob_Bod_Dias']  = (dp['Stock_Bodega_AA']   / dp['CDL']).round(1)

    # Alta rotacion (CDL >= UMBRAL_ALTA_ROTACION_CDL) cubre solo 2 dias habiles,
    # no 5 — misma regla que la seccion 14/E (df_ped, Requerimiento_Farm).
    dp['Requerimiento_Farm'] = np.where(
        dp['CDL'] >= UMBRAL_ALTA_ROTACION_CDL,
        dp['Consumo_2D_Trend'], dp['Consumo_5D_Trend'])
    _nec_farm_bruta = np.maximum(dp['Requerimiento_Farm'] - dp['Stock_Farmacia_AA'], 0).round(1)
    dp['Necesidad_Farm'] = [
        redondear_empaque(n, m, FACTOR_EMPAQUE)
        for n, m in zip(_nec_farm_bruta, dp['Medicamento'])
    ]
    dp['Traspaso_Posible'] = np.minimum(dp['Necesidad_Farm'], dp['Stock_Bodega_AA']).round(0)
    dp['Stock_Farm_Post']  = (dp['Stock_Farmacia_AA'] + dp['Traspaso_Posible']).round(0)
    dp['Cob_Farm_Post']    = (dp['Stock_Farm_Post']   / dp['CDL']).round(1)
    dp['Stock_Bod_Post']   = (dp['Stock_Bodega_AA']   - dp['Traspaso_Posible']).round(0)
    dp['Cob_Bod_Post']     = (dp['Stock_Bod_Post']    / dp['CDL']).round(1)
    dp['Deficit_Farm']     = np.maximum(dp['Necesidad_Farm'] - dp['Stock_Bodega_AA'], 0).round(0)
    dp['Crit_Farm']        = dp.apply(crit_farm, axis=1)
    dp['Accion_Traspaso']  = dp.apply(accion_farm_traspaso, axis=1)
    dp['Accion_Externo']   = dp.apply(accion_farm_externo,  axis=1)

    dp['Req_2_Semanas']    = (dp['CDL'] * CICLO_PEDIDO_BODEGA_DIAS).round(0)
    dp['Target_Stock_Bod'] = dp['Consumo_10D_Trend']
    _nec_bod_bruta = np.maximum(dp['Target_Stock_Bod'] - dp['Stock_AA_Total'], 0).round(1)
    dp['Necesidad_Bod'] = [
        redondear_empaque(n, m, FACTOR_EMPAQUE)
        for n, m in zip(_nec_bod_bruta, dp['Medicamento'])
    ]
    dp['Crit_Bod']         = dp.apply(crit_bod, axis=1)

    dp['Stock_BodFarm_Disp'] = (dp['Stock_BODEGA_FARMACOS'].fillna(0)
                                if 'Stock_BODEGA_FARMACOS' in dp.columns
                                else pd.Series(0, index=dp.index))
    dp['Traspaso_Hospital_Bod'] = np.minimum(dp['Necesidad_Bod'], dp['Stock_BodFarm_Disp']).round(0)
    dp['Compra_Externa_Bod']    = np.maximum(dp['Necesidad_Bod'] - dp['Stock_BodFarm_Disp'], 0).round(0)
    dp['Accion_Traspaso_Hosp']  = dp.apply(accion_bod_traspaso, axis=1)
    dp['Accion_Compra_Ext']     = dp.apply(accion_bod_compra,   axis=1)

    dp['Semana_Pico'] = dp['Medicamento'].map(_sp_map).fillna('-')
    return dp

if len(df_md):
    df_ped_dial = _pipeline_ped_dial(df_md)
else:
    df_ped_dial = pd.DataFrame(columns=list(df_md.columns) + ['Necesidad_Farm','Necesidad_Bod'])

# Mapeos específicos de diálisis = mapeo base + desglose de la demanda combinada
# (consumo 5D total, y cuánto es de diálisis vs farmacia no-diálisis). Solo las
# hojas de diálisis llevan estas columnas extra; las hojas generales no cambian.
_farm_cols_dial = {**_farm_cols,
                   'Consumo_5D_Solo_Dialisis' : 'Consumo_5D_Dial',
                   'Consumo_5D_Farm_NoDial'   : 'Consumo_5D_FarmNoDial',
                   'Factor_Empaque'           : 'Factor_Empaque'}
_bod_cols_dial  = {**_bod_cols,
                   'Consumo_5D_Solo_Dialisis' : 'Consumo_5D_Dial',
                   'Consumo_5D_Farm_NoDial'   : 'Consumo_5D_FarmNoDial',
                   'Factor_Empaque'           : 'Factor_Empaque'}

df_dial_farm_pedido = (
    df_ped_dial[df_ped_dial['Necesidad_Farm'] > 0]
    .sort_values(['Crit_Farm', 'CDL'], ascending=[True, False])
    [['Medicamento'] + list(_farm_cols_dial.values())]
    .rename(columns={v: k for k, v in _farm_cols_dial.items()})
    .reset_index(drop=True)
) if len(df_ped_dial) else pd.DataFrame(columns=['Medicamento'] + list(_farm_cols_dial.keys()))

df_dial_bod_pedido = (
    df_ped_dial[df_ped_dial['Necesidad_Bod'] > 0]
    .sort_values(['Crit_Bod', 'CDL'], ascending=[True, False])
    [['Medicamento'] + list(_bod_cols_dial.values())]
    .rename(columns={v: k for k, v in _bod_cols_dial.items()})
    .reset_index(drop=True)
) if len(df_ped_dial) else pd.DataFrame(columns=['Medicamento'] + list(_bod_cols_dial.keys()))

print(f"  DIALISIS — recetas nefrologos: {len(df_dial):,}  |  medicamentos: {len(idx_dial)}")
print(f"  DIALISIS — Pedido Farmacia: {len(df_dial_farm_pedido)}  |  Pedido Bodega: {len(df_dial_bod_pedido)}")

# Universo COMPLETO de medicamentos prescritos por nefrólogos (no solo los que
# necesitan reposición) — Dialisis_Pedido_Farm/Bod filtran a Necesidad>0 a
# propósito (solo lo accionable); esta hoja es el requerimiento íntegro para
# poder responder "cuánto se ocupa de X en diálisis" aunque el stock ya alcance.
_dial_med_cols = {
    'Stock_Farmacia_AA'      : 'Stock_Farmacia_AA',
    'Stock_Bodega_AA'        : 'Stock_Bodega_AA',
    'Stock_BODEGA_FARMACOS'  : 'Stock_BODEGA_FARMACOS',
    'Consumo_5D_Solo_Dialisis'          : 'Consumo_5D_Dial',
    'Consumo_5D_Farm_NoDial'            : 'Consumo_5D_FarmNoDial',
    'CDL_Combinado'                     : 'CDL',
    'Requerimiento_Mensual_Dialisis_30d': 'Requerimiento_Mensual_Dial_30d',
    'Necesidad_Farm'         : 'Necesidad_Farm',
    'Necesidad_Bod'          : 'Necesidad_Bod',
    'Criticidad_Farm'        : 'Crit_Farm',
    'Criticidad_Bod'         : 'Crit_Bod',
}
df_dial_medicamentos = (
    df_ped_dial
    .sort_values('Medicamento')
    [['Medicamento'] + list(_dial_med_cols.values())]
    .rename(columns={v: k for k, v in _dial_med_cols.items()})
    .reset_index(drop=True)
) if len(df_ped_dial) else pd.DataFrame(columns=['Medicamento'] + list(_dial_med_cols.keys()))
print(f"  DIALISIS — universo completo (Dialisis_Medicamentos): {len(df_dial_medicamentos)}")

# ═══════════════════════════════════════════════
# 16. AUDITORÍA DE HOMOLOGACIÓN
#     Registra qué nombres del ERP fueron renombrados por la tabla
#     HOMOLOGACION. Útil para detectar cambios de nombre en SSASUR.
# ═══════════════════════════════════════════════
rows_homo = []
for orig, dest in HOMOLOGACION.items():
    if orig != dest:
        cnt = (df_aa['Prescripcion_norm'] == dest).sum()
        rows_homo.append({'Original_Receta': orig,
                          'Homologado_ERP': dest,
                          'Registros_Afectados': cnt})
df_auditoria = pd.DataFrame(rows_homo)

# Outside-universe items with stock
# Excluye tambien los items descartados explicitamente (EXCLUIR_ESPECIFICOS)
# para que no aparezcan como "inconsistencias" cuando simplemente no son de AA
mask_fuera = (
    ~df_stk_scope['Descripcion_norm'].isin(universo_set) &
    ~df_stk_scope['Descripcion_norm'].isin(EXCLUIR_ESPECIFICOS)
)
df_inconsistencias = df_stk_scope[mask_fuera].groupby(
    ['Descripcion_norm','Bodega']
)['Cantidad'].sum().reset_index()
df_inconsistencias = df_inconsistencias[df_inconsistencias['Cantidad'] != 0]
neg_stk = df_stk_scope[df_stk_scope['Cantidad'] < 0][
    ['Descripcion_norm','Bodega','Cantidad']
].copy()

# ═══════════════════════════════════════════════
# 17. COMPARACIÓN DE STOCKS POR BODEGA
#     Muestra cuánto hay de cada medicamento en cada bodega del hospital
#     (BODEGA FARMACOS, AT CERRADA, URGENCIA...) para facilitar traspasos.
# ═══════════════════════════════════════════════
comp_rows = []
for med in df_master['Medicamento']:
    row = df_master[df_master['Medicamento'] == med].iloc[0]
    for b in ORDEN_TRASPASO:
        col = f'Stock_{b.replace(" ","_")}'
        comp_rows.append({
            'Medicamento': med,
            'Bodega': b,
            'Stock_Bodega': row.get(col, 0),
            'Stock_AA_Total': row['Stock_AA_Total'],
        })
df_comp = pd.DataFrame(comp_rows)
df_comp = df_comp[df_comp['Stock_Bodega'] > 0]

# ═══════════════════════════════════════════════
# 18. RESUMEN EJECUTIVO KPIs
#     Tabla de indicadores globales del proceso (queda en hoja KPIs del Excel).
# ═══════════════════════════════════════════════
kpis = {
    'Fecha Proceso':               _T_INICIO.strftime('%d/%m/%Y %H:%M'),
    'Fecha Inicio Datos':          FECHA_INICIO_OP.strftime('%d/%m/%Y'),
    'Fecha Max Datos':             FECHA_MAX.strftime('%d/%m/%Y'),
    'Días Calendario Período':     DIAS_CAL_OP,
    'Meses Reales Período':        round(MESES_TOT_OP, 4),
    'Días Laborales Período':      DIAS_LAB_OP,
    'Medicamentos Universo AA':    len(universo_set),
    'Total Recetas AA (dedup)':    len(df_aa_univ),
    'Medicamentos con Quiebre':    len(df_quiebres),
    'Faltantes Reales AA':         len(df_falt),
    'Pacientes Afectados (falt.)': int(df_falt['Pacientes_Afectados'].sum()),
    'Faltantes Absolutos AA 30d':  len(df_falt30),
    'Pacientes Afectados (30d)':   int(df_falt30['Pacientes_Afectados'].sum()),
    'Faltantes Persistentes 60d':  len(df_falt60),
    'Pacientes Afectados (60d)':   int(df_falt60['Pacientes_Afectados'].sum()),
    'Pendientes 15d':              len(df_p15),
    'Cobertura Objetivo (días)':   COB_OBJETIVO,
    'Medicamentos a Reponer':      len(df_repos),
    'Dialisis — Medicamentos':     len(idx_dial),
    'Dialisis — Pedido Farmacia':  len(df_dial_farm_pedido),
    'Dialisis — Pedido Bodega':    len(df_dial_bod_pedido),
    'Archivos CSV Cargados':       len(csv_files),
    'Registros Antes Dedup':       antes,
    'Registros Eliminados (dup)':  antes - len(df_rec),
}
df_kpis = pd.DataFrame(list(kpis.items()), columns=['Indicador', 'Valor'])

# ═══════════════════════════════════════════════
# 19. ESCRITURA EXCEL — Consolidado_AA_MAESTRO.xlsx
#     17 hojas: Stock_AA | Pedido_Farm_Bodega | Pedido_Repos_Bodega |
#               Faltas_Farmacia_AA | Faltantes_Detalle_AA |
#               Quiebres_Totales | Tendencia_Semanal |
#               Dialisis_Pedido_Farm | Dialisis_Pedido_Bod |
#               SGLI_Estres | Auditoria_Homologacion | Comparacion_Hospital |
#               KPIs | (más hojas auxiliares)
# ═══════════════════════════════════════════════
print("Generando Excel...")

# ─── Style helpers ───
HEAD_FILL        = PatternFill('solid', fgColor='2F5496')   # misma paleta SGLI
HEAD_FONT        = Font(bold=True, color='FFFFFF', name='Arial', size=11)
ROW_ALT          = PatternFill('solid', fgColor='F7F7F7')   # SGLI alt-row
ROJO_FILL        = PatternFill('solid', fgColor='FFD7D7')   # SGLI FILL_ALERTA
ROJO_CRITICO_FILL= crit_fill('1-CRITICO')  # rojo fuerte estandar — nivel CRITICO (aa_colors)
NARANJA          = PatternFill('solid', fgColor='FFF3E0')
AMARILLO   = PatternFill('solid', fgColor='FFF2CC')   # SGLI FILL_B_ROW
VERDE_FILL = PatternFill('solid', fgColor='E2EFDA')   # SGLI FILL_A_ROW
AZUL_FILL  = PatternFill('solid', fgColor='0D47A1')
AZUL_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=11)

# ─── Leyenda de colores de Criticidad (para que el operador entienda el semaforo) ───
LEYENDA_CRIT = [
    ('1-CRITICO',  'CRITICO  —  sin stock ni respaldo'),
    ('2-URGENTE',  'URGENTE  —  cobertura critica'),
    ('3-ALTO',     'ALTO  —  cobertura < 1 semana (Bodega)'),
    ('3-MODERADO', 'MODERADO  —  reponer pronto (Farmacia)'),
    ('4-MODERADO', 'MODERADO  —  cobertura < 2 sem. (Bodega)'),
    ('4-BAJO',     'BAJO  —  vigilar (Farmacia)'),
    ('5-BAJO',     'BAJO  —  stock suficiente (Bodega)'),
    ('5-OK',       'SUFICIENTE  —  sin necesidad'),
]

def add_leyenda(ws, fila_inicio, col=1):
    """Escribe una leyenda compacta del semaforo de Criticidad debajo de los datos."""
    t = ws.cell(row=fila_inicio, column=col, value='LEYENDA DE COLORES')
    t.font = Font(bold=True, name='Arial', size=10, color='334155')
    for i, (k, label) in enumerate(LEYENDA_CRIT, 1):
        c = ws.cell(row=fila_inicio + i, column=col, value=label)
        c.fill = crit_fill(k)
        c.font = Font(name='Arial', size=10, bold=(k == '1-CRITICO'),
                      color='7F1D1D' if k == '1-CRITICO' else '000000')
        c.alignment = Alignment(vertical='center')

def style_sheet(ws, df, col_widths=None, freeze='A2',
                row_color_fn=None, header_fill=None):
    """Apply standard formatting to a worksheet."""
    hf_hex  = fill_hex(header_fill, '2F5496')
    hf      = PatternFill('solid', fgColor=hf_hex)
    hf_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cols = list(df.columns)
    # Header row (row 1) — fondo color sólido + texto blanco (paleta SGLI)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill    = hf
        cell.font    = hf_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 25

    # Data rows
    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(vertical='center')
            # Numeric format
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = '#,##0'
            # Alt row
            if row_color_fn:
                f = row_color_fn(row_data, ri)
                if f: cell.fill = f
            elif ri % 2 == 0:
                cell.fill = ROW_ALT

    # Column widths
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    else:
        for ci, col in enumerate(cols, 1):
            max_w = max(len(str(col)),
                        max((len(str(v)) for v in df.iloc[:, ci-1].dropna()), default=0))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 3, 40)

    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True

# ═══════════════════════════════════════════════
# 18-bis. SGLI — Reposicion restringida por CAPACIDAD FISICA (motor sgli.py)
#   Baseline persistido con Factor_Carga = 1.15; la pestana SGLI de la app
#   recalcula en vivo con su slider. Incluye TODOS los meds con CDL>0 (la app
#   filtra los accionables) y ecoa Talla/Rotacion/Unidades_Caja + Stock/CDL para
#   que el recalculo no dependa de otras hojas. Talla por forma farmaceutica
#   (overrides por med en cenabast_tallas.csv). Cap_Max en UNIDADES (la farmacia
#   fracciona y ordena por unidad, no por empaque).
# ═══════════════════════════════════════════════
_sgli_overrides = cargar_tallas(WORK_DIR)
df_sgli = calcular_sgli(
    df_ped,
    factor_carga=FACTOR_CARGA_DEFAULT,
    overrides=_sgli_overrides,
    col_crit='Crit_Farm',
    col_cdl_pond='CDL_Pond',
)
print(f"  Tabla SGLI (capacidad): {len(df_sgli):,} meds  ·  baseline FC={FACTOR_CARGA_DEFAULT}  ·  overrides talla={len(_sgli_overrides)}")

# Factor_Empaque confiable para Dialisis_Medicamentos: Unidades_Caja de
# SGLI_Estres (cenabast_tallas.csv, curado a mano) — mismo criterio que
# pedido_fusion.py calc_h3. Reemplaza el Factor_Empaque de df_master, que viene
# de cenabast_intermediacion.csv con un matching de nombre más débil.
if len(df_dial_medicamentos):
    _fe_map_sgli = {
        str(r['Medicamento']).strip(): int(r['Unidades_Caja'])
        for _, r in df_sgli.iterrows()
        if pd.notna(r.get('Unidades_Caja')) and r['Unidades_Caja'] > 0
    }
    df_dial_medicamentos['Factor_Empaque'] = df_dial_medicamentos['Medicamento'].map(_fe_map_sgli).fillna(1).astype(int)

with pd.ExcelWriter(OUTPUT_XLS, engine='openpyxl') as writer:

    # ── 1. Resumen Ejecutivo ──────────────────────
    df_kpis.to_excel(writer, sheet_name='Resumen_Ejecutivo', index=False)
    ws = writer.sheets['Resumen_Ejecutivo']
    style_sheet(ws, df_kpis, col_widths=[38, 25])
    # KPI title banner
    ws.insert_rows(1)
    ws['A1'] = 'CONSOLIDADO OPERACIONAL — FARMACIA ATENCIÓN ABIERTA'
    ws['A1'].fill = PatternFill('solid', fgColor='2F5496')
    ws['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=14)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A3'

    # ── 2. Stock_AA ───────────────────────────────
    # Orden de columnas: stock → consumo (CDL→CMP→5D) → cobertura → reposición
    cols_stk = ['Medicamento',
                'Stock_Farmacia_AA','Stock_Bodega_AA','Stock_AA_Total',
                'CDL','CMP_Mensual','Consumo_5D',
                'Cobertura_Lab','Reposicion_Sugerida',
                'CDL_Pond','Quiebre_Total','Prescrito_Total_Op']
    df_stk_out = df_master[cols_stk].copy()
    df_stk_out.rename(columns={
        'CDL'              : 'CDL_DiasHab',
        'CMP_Mensual'      : 'CMP_Mensual_22d',
        'Consumo_5D'       : 'Consumo_5D_Hab',
        'Reposicion_Sugerida': 'Reposicion_5D',
        'CDL_Pond'         : 'CDL_Ponderado',
        'Prescrito_Total_Op': 'Total_Prescrito',
    }, inplace=True)
    df_stk_out.to_excel(writer, sheet_name='Stock_AA', index=False)
    ws2 = writer.sheets['Stock_AA']

    def color_stock(row_t, ri):
        qt = getattr(row_t, 'Quiebre_Total', False)
        cob = getattr(row_t, 'Cobertura_Lab', 999)
        if qt: return ROJO_FILL
        if cob < COB_OBJETIVO: return NARANJA
        if cob >= COB_OBJETIVO * 2: return VERDE_FILL
        return ROW_ALT if ri % 2 == 0 else None

    style_sheet(ws2, df_stk_out, row_color_fn=color_stock)

    # ── 3. Faltas_Farmacia_AA ─────────────────────
    # Unico respaldo real: Stock_BODEGA_FARMACOS
    cols_f3 = ['Medicamento','Stock_AA_Total','Faltante_Neto',
               'Pacientes_Afectados','N_Recetas',
               'Stock_BODEGA_FARMACOS','Stock_Hospital_Total',
               'Criticidad','Accion_Sugerida']
    if len(df_falt):
        src_cols = ['Prescripcion_norm','Faltante_Neto','Pacientes_Afectados',
                    'N_Recetas','Stock_AA_Total','Stock_BODEGA_FARMACOS',
                    'Stock_Hospital_Total','Criticidad','Accion_Sugerida']
        df_f3 = df_falt[[c for c in src_cols if c in df_falt.columns]].copy()
        df_f3.rename(columns={'Prescripcion_norm':'Medicamento'}, inplace=True)
        # Orden de columnas
        orden_f3 = ['Medicamento','Stock_AA_Total','Faltante_Neto',
                    'Pacientes_Afectados','N_Recetas',
                    'Stock_BODEGA_FARMACOS','Stock_Hospital_Total',
                    'Criticidad','Accion_Sugerida']
        df_f3 = df_f3[[c for c in orden_f3 if c in df_f3.columns]]
    else:
        df_f3 = pd.DataFrame(columns=cols_f3)

    df_f3.to_excel(writer, sheet_name='Faltas_Farmacia_AA', index=False)
    ws3 = writer.sheets['Faltas_Farmacia_AA']
    def color_falt(row_t, ri):
        crit = str(getattr(row_t, 'Criticidad', ''))
        if 'CRITICO'  in crit: return ROJO_CRITICO_FILL  # sin respaldo, muchos pacientes
        if 'ALTO'     in crit: return NARANJA             # sin respaldo, pocos pacientes
        if 'MODERADO' in crit: return AMARILLO            # con respaldo Bod.Farm., muchos pac.
        return VERDE_FILL                                  # con respaldo Bod.Farm., pocos pac.
    style_sheet(ws3, df_f3, row_color_fn=color_falt)

    # ── 4. Faltantes_Detalle_AA ───────────────────
    falt_det_cols = ['Prescripcion_norm','Cant_Demanda_Activa','Faltante_Neto',
                     'Pacientes_Afectados','N_Recetas','Stock_AA_Total',
                     'Stock_Farmacia_AA','Stock_Bodega_AA','Stock_Hospital_Total',
                     'Criticidad','Accion_Sugerida'] + \
                    [f'Stock_{b.replace(" ","_")}' for b in ORDEN_TRASPASO]
    df_falt_det = df_falt[[c for c in falt_det_cols if c in df_falt.columns]].copy()
    df_falt_det.rename(columns={'Prescripcion_norm':'Medicamento'}, inplace=True)
    df_falt_det.to_excel(writer, sheet_name='Faltantes_Detalle_AA', index=False)
    ws4 = writer.sheets['Faltantes_Detalle_AA']
    style_sheet(ws4, df_falt_det, row_color_fn=color_falt)

    # ── 4b. Faltantes_Absolutos_30D ───────────────
    # Faltantes reales (sec.4) acotados a los últimos 30 días y solo mostrador
    # de Atención Abierta — ver sec. 8b más arriba.
    falt30_cols = ['Prescripcion_norm','Cant_Demanda_Activa','Faltante_Neto',
                   'Pacientes_Afectados','N_Recetas','Stock_AA_Total',
                   'Stock_Farmacia_AA','Stock_Bodega_AA','Stock_Hospital_Total',
                   'Criticidad','Accion_Sugerida'] + \
                  [f'Stock_{b.replace(" ","_")}' for b in ORDEN_TRASPASO]
    df_falt30_det = df_falt30[[c for c in falt30_cols if c in df_falt30.columns]].copy()
    df_falt30_det.rename(columns={'Prescripcion_norm':'Medicamento'}, inplace=True)
    df_falt30_det.to_excel(writer, sheet_name='Faltantes_Absolutos_30D', index=False)
    ws4b = writer.sheets['Faltantes_Absolutos_30D']
    style_sheet(ws4b, df_falt30_det, row_color_fn=color_falt)

    # ── 4c. Faltantes_60D_Persistente ─────────────
    # Faltantes del mostrador AA con demanda pendiente en los últimos 60 días
    # que siguen vigentes hoy — ver sec. 8c. Basado en demanda, no en el stock
    # reportado (que puede ser fantasma). Dias_En_Falta = días desde la primera
    # prescripción sin cubrir en la ventana.
    falt60_cols = ['Prescripcion_norm','Dias_En_Falta','Primer_Faltante',
                   'Ultimo_Faltante','Pacientes_Afectados','N_Recetas',
                   'Cant_Demanda_Activa','Faltante_Neto','Stock_AA_Total',
                   'Stock_BODEGA_FARMACOS','Criticidad','Accion_Sugerida']
    df_falt60_det = df_falt60[[c for c in falt60_cols if c in df_falt60.columns]].copy()
    df_falt60_det.rename(columns={'Prescripcion_norm':'Medicamento'}, inplace=True)
    df_falt60_det.to_excel(writer, sheet_name='Faltantes_60D_Persistente', index=False)
    ws4c = writer.sheets['Faltantes_60D_Persistente']
    style_sheet(ws4c, df_falt60_det, row_color_fn=color_falt)

    # ── 5. Reposicion_5D ──────────────────────────
    # Fórmula: Reposicion = max(CDL×5 − Stock_Farmacia, 0)
    # Consistencia: CMP_Mensual = CDL×22  |  Consumo_5D = CDL×5  |  CMP/22 = CDL
    df_repos_out = df_repos[[
        'Medicamento',
        'CDL',             # Consumo Diario Laboral  = Prescrito / DIAS_LAB_OP
        'CMP_Mensual',     # = CDL × 22              → CMP_Mensual / 22 = CDL
        'Consumo_5D',      # = CDL × 5               → Consumo_5D / CMP_Mensual = 5/22
        'Stock_Farmacia_AA',
        'Reposicion_Sugerida',   # = max(Consumo_5D - Stock, 0)
        'Cobertura_Lab',         # = Stock / CDL  [días hábiles]
        'CDL_Pond',        # tasa ponderada 15d×50%+30d×30%+60d×20% (referencia tendencia)
        'Stock_Bodega_AA',
        'Stock_Hospital_Total',
        'Prescrito_Total_Op',
    ]].copy()
    df_repos_out.rename(columns={
        'CDL'               : 'CDL_DiasHab',
        'CMP_Mensual'       : 'CMP_Mensual_22d',
        'Consumo_5D'        : 'Consumo_5D_Hab',
        'Reposicion_Sugerida': 'Reposicion_5D',
        'Cobertura_Lab'     : 'Cobertura_DiasHab',
        'CDL_Pond'          : 'CDL_Ponderado_Ref',
        'Prescrito_Total_Op': 'Total_Prescrito_Periodo',
    }, inplace=True)
    df_repos_out.to_excel(writer, sheet_name='Reposicion_5D', index=False)
    ws5 = writer.sheets['Reposicion_5D']
    style_sheet(ws5, df_repos_out, col_widths=[
        52, 13, 16, 14, 16, 13, 16, 16, 16, 16, 22
    ])

    # ── 6. Pendientes_15D ─────────────────────────
    df_p15_out = df_p15.rename(columns={
        'Fecha_Atencion':'Fecha_Atencion',
        'Prescripcion':'Prescripcion_Original',
        'Prescripcion_norm':'Prescripcion_Norm',
    })
    df_p15_out.to_excel(writer, sheet_name='Pendientes_15D', index=False)
    ws6 = writer.sheets['Pendientes_15D']
    style_sheet(ws6, df_p15_out)

    # ── 7. Comparacion_Hospital ───────────────────
    df_comp.to_excel(writer, sheet_name='Comparacion_Hospital', index=False)
    ws7 = writer.sheets['Comparacion_Hospital']
    style_sheet(ws7, df_comp)

    # ── 8. Quiebres_Totales_AA ────────────────────
    df_q_out = df_quiebres[['Medicamento','Stock_Farmacia_AA','Stock_Bodega_AA',
                             'CDL','CMP_Mensual','Consumo_5D','Cobertura_Lab',
                             'Stock_Hospital_Total','Instruccion']].copy()
    df_q_out.rename(columns={
        'CDL'        : 'CDL_DiasHab',
        'CMP_Mensual': 'CMP_Mensual_22d',
        'Consumo_5D' : 'Consumo_5D_Hab',
    }, inplace=True)
    df_q_out.to_excel(writer, sheet_name='Quiebres_Totales_AA', index=False)
    ws8 = writer.sheets['Quiebres_Totales_AA']
    def color_quiebre(row_t, ri):
        return ROJO_FILL
    style_sheet(ws8, df_q_out, row_color_fn=color_quiebre)

    # ── 9. Auditoria_Homologacion ─────────────────
    # Also include universo trazabilidad
    univ_trace = pd.DataFrame({
        'Medicamento_Norm': list(universo_set),
        'En_Stock_AA': [m in df_stk_scope[df_stk_scope['Bodega'].isin(BODEGAS_AA)]
                        ['Descripcion_norm'].values for m in universo_set],
        'En_Stock_Hospital': [m in df_stk_scope[df_stk_scope['Bodega'].isin(BODEGAS_HOSPITAL)]
                              ['Descripcion_norm'].values for m in universo_set],
    })
    df_excl_esp = pd.DataFrame({
        'Medicamento_Original': EXCLUIR_ESPECIFICOS_RAW,
        'Medicamento_Norm':     [norm_erp(x) for x in EXCLUIR_ESPECIFICOS_RAW],
        'Motivo':               'EXCLUIDO MANUAL — NO ES AT ABIERTA',
    })
    df_auditoria.to_excel(writer, sheet_name='Auditoria_Homologacion', index=False, startrow=0)
    df_excl_esp.to_excel(writer, sheet_name='Auditoria_Homologacion', index=False,
                         startrow=len(df_auditoria)+2)
    univ_trace.to_excel(writer, sheet_name='Auditoria_Homologacion', index=False,
                        startrow=len(df_auditoria)+len(df_excl_esp)+4)
    ws9 = writer.sheets['Auditoria_Homologacion']
    style_sheet(ws9, df_auditoria)

    # ── 10. Inconsistencias_ERP ───────────────────
    neg_stk['Tipo'] = 'STOCK NEGATIVO'
    df_inconsistencias['Tipo'] = 'FUERA UNIVERSO AA'
    df_incons_out = pd.concat([
        df_inconsistencias.rename(columns={'Descripcion_norm':'Descripcion','Cantidad':'Stock_Actual'}),
        neg_stk.rename(columns={'Descripcion_norm':'Descripcion','Cantidad':'Stock_Actual'}),
    ], ignore_index=True)
    df_incons_out.to_excel(writer, sheet_name='Inconsistencias_ERP', index=False)
    ws10 = writer.sheets['Inconsistencias_ERP']
    style_sheet(ws10, df_incons_out)

    # ── 11. Consumo_Mensual ───────────────────────
    # Una fila por medicamento — columnas: mes-a-mes por Fecha_Entrega, CMP, CDL, stock
    df_consumo_mensual.to_excel(writer, sheet_name='Consumo_Mensual', index=False)
    ws11 = writer.sheets['Consumo_Mensual']

    # Columnas de mes (dinamicas) vs columnas fijas al final
    fixed_tail = ['Total_Periodo','Meses_Periodo','Dias_Lab_Periodo',
                  'CMP_Mensual','CDL',
                  'Stock_Farmacia_AA','Stock_Bodega_AA','Stock_AA_Total','Cobertura_Lab']
    mes_cols_out = [c for c in df_consumo_mensual.columns
                    if c not in (['Medicamento'] + fixed_tail)]

    def color_consumo_mensual(row_t, ri):
        # Resalta filas con CMP > 0 y quiebre (stock farmacia = 0)
        cmp_v = getattr(row_t, 'CMP_Mensual', 0)
        stk_v = getattr(row_t, 'Stock_Farmacia_AA', 1)
        if cmp_v > 0 and stk_v == 0:
            return ROJO_FILL
        return ROW_ALT if ri % 2 == 0 else None

    style_sheet(ws11, df_consumo_mensual, row_color_fn=color_consumo_mensual)

    # Ancho columna medicamento amplio; meses angostos
    col_list = list(df_consumo_mensual.columns)
    for ci, col in enumerate(col_list, 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento':
            ws11.column_dimensions[ltr].width = 52
        elif col in mes_cols_out:
            ws11.column_dimensions[ltr].width = 11
        else:
            ws11.column_dimensions[ltr].width = 16
    ws11.freeze_panes = 'B2'

    # ── 12. Tendencia_Semanal ─────────────────────
    # Filas = medicamento | Col_1..N = "Mes S#" (detalle) + Total_S1..4 + %_S1..4 + Semana_Pico
    df_tendencia.to_excel(writer, sheet_name='Tendencia_Semanal', index=False)
    ws12 = writer.sheets['Tendencia_Semanal']

    tend_cols   = list(df_tendencia.columns)
    detail_cols = [c for c in tend_cols
                   if c not in (['Medicamento','Semana_Pico','CMP_Mensual','CDL',
                                  'Stock_Farmacia_AA','Cobertura_Lab'])
                   and not c.startswith('Total_S') and not c.startswith('%_S')]

    def color_tendencia(row_t, ri):
        sp = str(getattr(row_t, 'Semana_Pico', ''))
        if sp == 'S1': return PatternFill('solid', fgColor='E3F2FD')   # azul claro → pico semana 1
        if sp == 'S2': return PatternFill('solid', fgColor='E8F5E9')   # verde claro → pico semana 2
        if sp == 'S3': return PatternFill('solid', fgColor='FFF9C4')   # amarillo    → pico semana 3
        if sp == 'S4': return PatternFill('solid', fgColor='FFF3E0')   # naranja claro→ pico semana 4
        return ROW_ALT if ri % 2 == 0 else None

    style_sheet(ws12, df_tendencia, row_color_fn=color_tendencia)

    for ci, col in enumerate(tend_cols, 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento':
            ws12.column_dimensions[ltr].width = 52
        elif col in detail_cols:          # columnas "Mes S#"
            ws12.column_dimensions[ltr].width = 10
        elif col.startswith('Total_S') or col.startswith('%_S'):
            ws12.column_dimensions[ltr].width = 12
        else:
            ws12.column_dimensions[ltr].width = 14
    ws12.freeze_panes = 'B2'

    # ── 13. Pedido_Farm_Bodega ────────────────────
    # Qué debe pedir Farmacia AT Abierta a Bodega AT Abierta
    # Ordenado por criticidad (1-CRITICO primero) luego por CDL (mayor volumen primero)
    # Cubre ambas escalas (Crit_Farm y Crit_Bod) — ver aa_colors.py
    CRIT_COLOR = {k: crit_fill(k) for k in CRIT_FILL_HEX}
    CRIT_FONT = {
        '1-CRITICO': Font(bold=True, color='7F1D1D', name='Arial', size=11),
    }
    def color_pedido_farm(row_t, ri):
        crit = str(getattr(row_t, 'Criticidad', '5-OK'))
        return CRIT_COLOR.get(crit, ROW_ALT if ri % 2 == 0 else None)

    df_farm_pedido_completo.to_excel(writer, sheet_name='Pedido_Farm_Bodega', index=False)
    ws13 = writer.sheets['Pedido_Farm_Bodega']
    style_sheet(ws13, df_farm_pedido_completo, row_color_fn=color_pedido_farm,
                header_fill=PatternFill('solid', fgColor='880E4F'))
    # Negrita en filas CRITICO
    for ri, row_t in enumerate(df_farm_pedido_completo.itertuples(index=False), 2):
        crit = str(getattr(row_t, 'Criticidad', ''))
        if crit == '1-CRITICO':
            for ci in range(1, len(df_farm_pedido_completo.columns)+1):
                ws13.cell(row=ri, column=ci).font = Font(bold=True, color='7F1D1D',
                                                          name='Arial', size=11)
    # Anchos
    farm_ped_cols = list(df_farm_pedido_completo.columns)
    for ci, col in enumerate(farm_ped_cols, 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento':                ws13.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Bodega': ws13.column_dimensions[ltr].width = 38
        elif col == 'Accion_2_Gestion_Externa': ws13.column_dimensions[ltr].width = 28
        elif col == 'Criticidad':               ws13.column_dimensions[ltr].width = 14
        else:                                   ws13.column_dimensions[ltr].width = 16
    ws13.freeze_panes = 'B2'

    # ── 14. Pedido_Reposicion_Bodega ──────────────
    # Qué debe pedir/comprar Bodega AT Abierta para reponerse
    def color_pedido_bod(row_t, ri):
        crit = str(getattr(row_t, 'Criticidad', '5-OK'))
        return CRIT_COLOR.get(crit, ROW_ALT if ri % 2 == 0 else None)

    df_bod_pedido.to_excel(writer, sheet_name='Pedido_Repos_Bodega', index=False)
    ws14 = writer.sheets['Pedido_Repos_Bodega']
    style_sheet(ws14, df_bod_pedido, row_color_fn=color_pedido_bod,
                header_fill=PatternFill('solid', fgColor='1A237E'))
    for ri, row_t in enumerate(df_bod_pedido.itertuples(index=False), 2):
        crit = str(getattr(row_t, 'Criticidad', ''))
        if crit == '1-CRITICO':
            for ci in range(1, len(df_bod_pedido.columns)+1):
                ws14.cell(row=ri, column=ci).font = Font(bold=True, color='7F1D1D',
                                                          name='Arial', size=11)
    bod_ped_cols = list(df_bod_pedido.columns)
    for ci, col in enumerate(bod_ped_cols, 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento':                   ws14.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Hospital':  ws14.column_dimensions[ltr].width = 40
        elif col == 'Accion_2_Compra_Externa':     ws14.column_dimensions[ltr].width = 26
        elif col == 'Criticidad':                  ws14.column_dimensions[ltr].width = 14
        else:                                      ws14.column_dimensions[ltr].width = 18
    ws14.freeze_panes = 'B2'

    # ── 15. Dialisis_Pedido_Farm ──────────────────
    # Mismo formato que Pedido_Farm_Bodega pero SOLO recetas de nefrologos
    df_dial_farm_pedido.to_excel(writer, sheet_name='Dialisis_Pedido_Farm', index=False)
    ws15 = writer.sheets['Dialisis_Pedido_Farm']
    style_sheet(ws15, df_dial_farm_pedido, row_color_fn=color_pedido_farm,
                header_fill=PatternFill('solid', fgColor='0F766E'))
    for ri, row_t in enumerate(df_dial_farm_pedido.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_dial_farm_pedido.columns)+1):
                ws15.cell(row=ri, column=ci).font = Font(bold=True, color='7F1D1D',
                                                          name='Arial', size=11)
    for ci, col in enumerate(list(df_dial_farm_pedido.columns), 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento':                 ws15.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Bodega':  ws15.column_dimensions[ltr].width = 38
        elif col == 'Accion_2_Gestion_Externa':  ws15.column_dimensions[ltr].width = 28
        elif col == 'Criticidad':                ws15.column_dimensions[ltr].width = 14
        else:                                    ws15.column_dimensions[ltr].width = 16
    ws15.freeze_panes = 'B2'

    # ── 16. Dialisis_Pedido_Bod ───────────────────
    # Mismo formato que Pedido_Repos_Bodega pero SOLO recetas de nefrologos
    df_dial_bod_pedido.to_excel(writer, sheet_name='Dialisis_Pedido_Bod', index=False)
    ws16 = writer.sheets['Dialisis_Pedido_Bod']
    style_sheet(ws16, df_dial_bod_pedido, row_color_fn=color_pedido_bod,
                header_fill=PatternFill('solid', fgColor='0E7490'))
    for ri, row_t in enumerate(df_dial_bod_pedido.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_dial_bod_pedido.columns)+1):
                ws16.cell(row=ri, column=ci).font = Font(bold=True, color='7F1D1D',
                                                          name='Arial', size=11)
    for ci, col in enumerate(list(df_dial_bod_pedido.columns), 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento':                   ws16.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Hospital':  ws16.column_dimensions[ltr].width = 40
        elif col == 'Accion_2_Compra_Externa':     ws16.column_dimensions[ltr].width = 26
        elif col == 'Criticidad':                  ws16.column_dimensions[ltr].width = 14
        else:                                      ws16.column_dimensions[ltr].width = 18
    ws16.freeze_panes = 'B2'

    # ── 16b. Dialisis_Medicamentos ─────────────────
    # Universo COMPLETO prescrito por nefrólogos (no solo Necesidad>0) — permite
    # ver el requerimiento de un medicamento de diálisis aunque el stock alcance.
    df_dial_medicamentos.to_excel(writer, sheet_name='Dialisis_Medicamentos', index=False)
    ws16b = writer.sheets['Dialisis_Medicamentos']
    style_sheet(ws16b, df_dial_medicamentos, row_color_fn=color_pedido_bod,
                header_fill=PatternFill('solid', fgColor='334155'))
    for ci, col in enumerate(list(df_dial_medicamentos.columns), 1):
        ltr = get_column_letter(ci)
        if col == 'Medicamento': ws16b.column_dimensions[ltr].width = 52
        else:                    ws16b.column_dimensions[ltr].width = 18
    ws16b.freeze_panes = 'B2'

    # ── 17. SGLI_Estres ───────────────────────────
    # Reposicion restringida por CAPACIDAD FISICA (motor sgli.py). Baseline
    # FC=1.15; la pestana SGLI de la app recalcula en vivo. Color por NIVEL de
    # criticidad (aa_colors.crit_hex) para que '2-ALTO' (no esta en
    # CRIT_FILL_HEX) tambien tome el color del semaforo.
    def color_sgli(row_t, ri):
        return PatternFill('solid', fgColor=crit_hex(str(getattr(row_t, 'Criticidad', '5-OK'))))

    df_sgli.to_excel(writer, sheet_name='SGLI_Estres', index=False)
    ws_sgli = writer.sheets['SGLI_Estres']
    style_sheet(ws_sgli, df_sgli, row_color_fn=color_sgli,
                header_fill=PatternFill('solid', fgColor='7C2D12'))
    for ri, row_t in enumerate(df_sgli.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_sgli.columns) + 1):
                ws_sgli.cell(row=ri, column=ci).font = Font(bold=True, color='7F1D1D',
                                                            name='Arial', size=11)
    for ci, col in enumerate(list(df_sgli.columns), 1):
        ltr = get_column_letter(ci)
        if   col == 'Medicamento':       ws_sgli.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso': ws_sgli.column_dimensions[ltr].width = 30
        elif col == 'Accion_2_Externa':  ws_sgli.column_dimensions[ltr].width = 26
        elif col == 'Criticidad':        ws_sgli.column_dimensions[ltr].width = 13
        else:                            ws_sgli.column_dimensions[ltr].width = 13
    ws_sgli.freeze_panes = 'B2'

# ═══════════════════════════════════════════════
# 20. RESUMEN SEMANAL — Resumen_Pedidos_AA.xlsx
#     Excel operativo simplificado para imprimir o compartir:
#     5 hojas: Faltantes_Semana | Pedido_Farmacia_AA | Pedido_Bodega_AA |
#              Dialisis_Farmacia | Dialisis_Bodega
# ═══════════════════════════════════════════════
_RES_BASE  = os.path.join(WORK_DIR, "Resumen_Pedidos_AA.xlsx")
_RES_DATED = os.path.join(WORK_DIR,
    f"Resumen_Pedidos_AA_{datetime.now().strftime('%Y%m%d')}.xlsx")

def _resumen_path():
    if not os.path.exists(_RES_BASE):
        return _RES_BASE
    try:
        with open(_RES_BASE, 'a'): pass
        return _RES_BASE
    except PermissionError:
        return _RES_DATED

RESUMEN_XLS = _resumen_path()

semana_actual = _semana_mes(HOY)
periodo_label = f"S{semana_actual}  {HOY.strftime('%d/%m/%Y')}"

# Columnas resumidas — solo lo accionable
_farm_res = ['Medicamento','Criticidad','Stock_Farm_Actual','Cob_Farm_Actual_Dias',
             'Consumo_5D_Trend','Necesidad_5D_Farm','A_Traspasar',
             'Deficit_Post_Traspaso','Accion_1_Traspaso_Bodega','Accion_2_Gestion_Externa']
_bod_res  = ['Medicamento','Criticidad','Stock_Bod_Post_Traspaso','Cob_Bod_Post_Dias',
             'Consumo_10D_Trend','Reponer_Bodega','Stock_BODEGA_FARMACOS',
             'Accion_1_Traspaso_Hospital','Accion_2_Compra_Externa']
_falt_res = ['Medicamento','Stock_AA_Total','Faltante_Neto',
             'Pacientes_Afectados','N_Recetas','Criticidad','Accion_Sugerida']

df_farm_res = df_farm_pedido[[c for c in _farm_res if c in df_farm_pedido.columns]].copy()
df_bod_res  = df_bod_pedido [[c for c in _bod_res  if c in df_bod_pedido.columns]].copy()
df_falt_res = df_f3[[c for c in _falt_res if c in df_f3.columns]].copy() if len(df_f3) else \
              pd.DataFrame(columns=_falt_res)

# Dialisis — columnas resumidas + desglose de la demanda combinada
# (Consumo_5D_Trend ya es la suma farm no-dial + diálisis; se añade el desglose).
_farm_res_dial = _farm_res[:5] + ['Consumo_5D_Solo_Dialisis','Consumo_5D_Farm_NoDial'] + _farm_res[5:]
_bod_res_dial  = _bod_res[:5]  + ['Consumo_5D_Solo_Dialisis','Consumo_5D_Farm_NoDial'] + _bod_res[5:]
df_dial_farm_res = df_dial_farm_pedido[[c for c in _farm_res_dial if c in df_dial_farm_pedido.columns]].copy() \
                   if len(df_dial_farm_pedido) else pd.DataFrame(columns=_farm_res_dial)
df_dial_bod_res  = df_dial_bod_pedido [[c for c in _bod_res_dial  if c in df_dial_bod_pedido.columns]].copy() \
                   if len(df_dial_bod_pedido) else pd.DataFrame(columns=_bod_res_dial)

def _titulo(ws, texto, color_hex):
    ws.insert_rows(1)
    ws['A1'] = texto
    ws['A1'].fill = PatternFill('solid', fgColor=soften(color_hex))
    ws['A1'].font = Font(bold=True, color=darken(color_hex), name='Arial', size=12)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A3'

with pd.ExcelWriter(RESUMEN_XLS, engine='openpyxl') as rw:

    # ── Hoja 1: Faltantes_Semana ──────────────────
    df_falt_res.to_excel(rw, sheet_name='Faltantes_Semana', index=False)
    ws_r1 = rw.sheets['Faltantes_Semana']
    style_sheet(ws_r1, df_falt_res, row_color_fn=color_falt,
                header_fill=PatternFill('solid', fgColor='0D47A1'))
    _titulo(ws_r1, f'FALTANTES REALES — {periodo_label}', 'B71C1C')
    for ci, col in enumerate(list(df_falt_res.columns), 1):
        ltr = get_column_letter(ci)
        if   col == 'Medicamento':    ws_r1.column_dimensions[ltr].width = 52
        elif col == 'Criticidad':     ws_r1.column_dimensions[ltr].width = 36
        elif col == 'Accion_Sugerida':ws_r1.column_dimensions[ltr].width = 36
        else:                         ws_r1.column_dimensions[ltr].width = 16
    add_leyenda(ws_r1, len(df_falt_res) + 4)

    # ── Hoja 2: Pedido_Farmacia_AA ────────────────
    df_farm_res.to_excel(rw, sheet_name='Pedido_Farmacia_AA', index=False)
    ws_r2 = rw.sheets['Pedido_Farmacia_AA']
    style_sheet(ws_r2, df_farm_res, row_color_fn=color_pedido_farm,
                header_fill=PatternFill('solid', fgColor='880E4F'))
    _titulo(ws_r2, f'PEDIDO FARMACIA AA → BODEGA AA — {periodo_label}', '880E4F')
    for ci, col in enumerate(list(df_farm_res.columns), 1):
        ltr = get_column_letter(ci)
        if   col == 'Medicamento':                ws_r2.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Bodega':   ws_r2.column_dimensions[ltr].width = 38
        elif col == 'Accion_2_Gestion_Externa':   ws_r2.column_dimensions[ltr].width = 28
        elif col == 'Criticidad':                 ws_r2.column_dimensions[ltr].width = 14
        else:                                     ws_r2.column_dimensions[ltr].width = 16
    # Negrita filas CRITICO — fuente oscura sobre fondo rosa claro
    for ri, row_t in enumerate(df_farm_res.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_farm_res.columns)+1):
                ws_r2.cell(row=ri+1, column=ci).font = Font(bold=True, color='7F1D1D',
                                                             name='Arial', size=11)
    add_leyenda(ws_r2, len(df_farm_res) + 4)

    # ── Hoja 3: Pedido_Bodega_AA ──────────────────
    df_bod_res.to_excel(rw, sheet_name='Pedido_Bodega_AA', index=False)
    ws_r3 = rw.sheets['Pedido_Bodega_AA']
    style_sheet(ws_r3, df_bod_res, row_color_fn=color_pedido_bod,
                header_fill=PatternFill('solid', fgColor='1A237E'))
    _titulo(ws_r3, f'PEDIDO REPOSICION BODEGA AA — {periodo_label}', '1A237E')
    for ci, col in enumerate(list(df_bod_res.columns), 1):
        ltr = get_column_letter(ci)
        if   col == 'Medicamento':                   ws_r3.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Hospital':    ws_r3.column_dimensions[ltr].width = 40
        elif col == 'Accion_2_Compra_Externa':       ws_r3.column_dimensions[ltr].width = 26
        elif col == 'Criticidad':                    ws_r3.column_dimensions[ltr].width = 14
        else:                                        ws_r3.column_dimensions[ltr].width = 18
    for ri, row_t in enumerate(df_bod_res.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_bod_res.columns)+1):
                ws_r3.cell(row=ri+1, column=ci).font = Font(bold=True, color='7F1D1D',
                                                             name='Arial', size=11)
    add_leyenda(ws_r3, len(df_bod_res) + 4)

    # ── Hoja 4: Dialisis_Farmacia (Farmacia AA ← Bodega AA, solo nefrologos) ──
    df_dial_farm_res.to_excel(rw, sheet_name='Dialisis_Farmacia', index=False)
    ws_r4 = rw.sheets['Dialisis_Farmacia']
    style_sheet(ws_r4, df_dial_farm_res, row_color_fn=color_pedido_farm,
                header_fill=PatternFill('solid', fgColor='0F766E'))
    _titulo(ws_r4, f'DIALISIS · PEDIDO FARMACIA AA → BODEGA AA — {periodo_label}', '0F766E')
    for ci, col in enumerate(list(df_dial_farm_res.columns), 1):
        ltr = get_column_letter(ci)
        if   col == 'Medicamento':                ws_r4.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Bodega':   ws_r4.column_dimensions[ltr].width = 38
        elif col == 'Accion_2_Gestion_Externa':   ws_r4.column_dimensions[ltr].width = 28
        elif col == 'Criticidad':                 ws_r4.column_dimensions[ltr].width = 14
        else:                                     ws_r4.column_dimensions[ltr].width = 16
    for ri, row_t in enumerate(df_dial_farm_res.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_dial_farm_res.columns)+1):
                ws_r4.cell(row=ri+1, column=ci).font = Font(bold=True, color='7F1D1D',
                                                             name='Arial', size=11)
    add_leyenda(ws_r4, len(df_dial_farm_res) + 4)

    # ── Hoja 5: Dialisis_Bodega (Bodega AA ← Bodega Farmacos, solo nefrologos) ──
    df_dial_bod_res.to_excel(rw, sheet_name='Dialisis_Bodega', index=False)
    ws_r5 = rw.sheets['Dialisis_Bodega']
    style_sheet(ws_r5, df_dial_bod_res, row_color_fn=color_pedido_bod,
                header_fill=PatternFill('solid', fgColor='0E7490'))
    _titulo(ws_r5, f'DIALISIS · PEDIDO BODEGA AA → BODEGA FARMACOS — {periodo_label}', '0E7490')
    for ci, col in enumerate(list(df_dial_bod_res.columns), 1):
        ltr = get_column_letter(ci)
        if   col == 'Medicamento':                   ws_r5.column_dimensions[ltr].width = 52
        elif col == 'Accion_1_Traspaso_Hospital':    ws_r5.column_dimensions[ltr].width = 40
        elif col == 'Accion_2_Compra_Externa':       ws_r5.column_dimensions[ltr].width = 26
        elif col == 'Criticidad':                    ws_r5.column_dimensions[ltr].width = 14
        else:                                        ws_r5.column_dimensions[ltr].width = 18
    for ri, row_t in enumerate(df_dial_bod_res.itertuples(index=False), 2):
        if str(getattr(row_t, 'Criticidad', '')) == '1-CRITICO':
            for ci in range(1, len(df_dial_bod_res.columns)+1):
                ws_r5.cell(row=ri+1, column=ci).font = Font(bold=True, color='7F1D1D',
                                                             name='Arial', size=11)
    add_leyenda(ws_r5, len(df_dial_bod_res) + 4)

print(f"  [OK] Resumen semanal: {os.path.basename(RESUMEN_XLS)}")

_duracion = datetime.now() - _T_INICIO
print(f"\n{'='*60}")
print(f"  [OK] Excel generado : {os.path.basename(OUTPUT_XLS)}")

# ─── Final validation ───
wb = load_workbook(OUTPUT_XLS)
print(f"  [OK] Hojas ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")
print(f"  Duracion total      : {_duracion.seconds}s")
print(f"  Carpeta             : {WORK_DIR}")
print(f"{'='*60}")

# ─── Log en archivo ───────────────────────────────────────────────────────────
LOGS_DIR = os.path.join(WORK_DIR, 'logs')
os.makedirs(LOGS_DIR, exist_ok=True)
log_file = os.path.join(LOGS_DIR, f"maestro_{_T_INICIO.strftime('%Y%m%d_%H%M%S')}.log")
with open(log_file, 'w', encoding='utf-8') as _lf:
    _lf.write(f"Ejecucion: {_T_INICIO}\n")
    _lf.write(f"Stock   : {os.path.basename(stk_file)}\n")
    _lf.write(f"Recetas : {len(csv_files)} archivos ({len(df_rec):,} recetas dedup)\n")
    _lf.write(f"Universo: {len(universo_set)} medicamentos\n")
    _lf.write(f"Salida  : {os.path.basename(OUTPUT_XLS)}\n")
    _lf.write(f"Duracion: {_duracion.seconds}s\n")
print(f"  Log guardado en: logs/{os.path.basename(log_file)}")

# ─── Generar planilla SGLI histórica ─────────────────────────────────────────
try:
    import sgli_historico
    print("\n[SGLI] Generando planilla histórica SGLI...")
    sgli_historico.main()
except Exception as _e_sgli:
    print(f"  [WARN] SGLI histórico no generado: {_e_sgli}")
