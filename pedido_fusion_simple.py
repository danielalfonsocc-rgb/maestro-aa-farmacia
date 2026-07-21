# -*- coding: utf-8 -*-
"""
pedido_fusion_simple.py — Planilla simplificada Farm_Bodega + Faltantes_AA
===========================================================================
Version reducida de Pedido_Fusion_AA pensada para Google Sheets: solo 4 hojas.

  1 "Farm_Bodega"   Mismo universo/cálculo que la hoja Farm_Bod de
                     pedido_fusion.py, pero solo muestra Medicamento,
                     Criticidad, Stock Bodega A, Stock Farmacia y
                     Cantidad a Reponer (el resto de columnas/estilo igual).
  2 "Faltantes_AA"  Idéntica a la hoja 4 de pedido_fusion.py.
  3 "Por_Agotarse"  Idéntica a la hoja 5 de pedido_fusion.py (Bodega AA en 0,
                     farmacia con cobertura ≤ UMBRAL_PREQUIEBRE días).
  4 "Faltantes_60D" Idéntica a la hoja 6 de pedido_fusion.py (faltantes
                     persistentes de los últimos 60 días, vigentes hoy).

Sin llamadas a IA — solo pandas + openpyxl. Reutiliza el cálculo y estilos de
pedido_fusion.py (no duplica lógica de negocio).

Uso:
    py pedido_fusion_simple.py
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter

import pedido_fusion as pf

WORK_DIR = pf.WORK_DIR

HDRS_SIMPLE = [
    ('Medicamento',              44),
    ('Criticidad',               12),
    ('Stock Bodega A (ud)',      16),
    ('Stock Farmacia (ud)',      16),
    ('Cantidad a Reponer (ud)',  18),
]


def calc_simple(df_farm, fe_map, dias_ef, rep_h2_map):
    rows = pf._calc_h1_rows(df_farm, fe_map, dias_ef, todos=False, rep_h2_map=rep_h2_map)
    out = [{'v': (r['v'][0], r['v'][1], r['_sbod'], r['v'][2], r['v'][5]), '_nv': r['_nv']}
           for r in rows]
    return [x['v'] for x in out]


def write_simple(ws, rows, dias_ef, hoy, semana):
    ndia = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][hoy.weekday()]
    pf._titulo(ws,
        f'FARMACIA AA → BODEGA AA  ·  {ndia} {hoy.strftime("%d/%m/%Y")}  ·  '
        f'S{semana}  ·  {dias_ef} día(s) hábil(es) restante(s)',
        len(HDRS_SIMPLE))
    pf._subtit(ws,
        'Stock Bodega A = disponible en Bodega AA para traspaso | '
        'Stock Farmacia = stock actual en Farmacia AA | '
        f'Cantidad a Reponer = CDL×{dias_ef}d+SS−Stock Farmacia, redondeado al empaque CENABAST',
        len(HDRS_SIMPLE))
    pf._hdr(ws, 3, HDRS_SIMPLE)
    for i, vals in enumerate(rows, 4):
        pf._fila_crit(ws, i, vals, vals[1], {2, 3, 4, 5})
    ws.freeze_panes = 'A4'
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS_SIMPLE))}{last}'
        pf._totals(ws, 4, last, len(HDRS_SIMPLE), {5})
        ws.print_area = f'A1:{get_column_letter(len(HDRS_SIMPLE))}{last + 1}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


def main():
    hoy = pf.dt.date.today()
    fer = pf._feriados()
    mae = pf._maestro()
    sem = pf._semana(hoy)
    def_ = pf._dias_ef(hoy, fer)

    data = pf._leer(mae)

    fe_map = {}
    if len(data['sgli']):
        for _, r in data['sgli'].iterrows():
            m = str(r.get('Medicamento', '')).strip()
            fe_map[m] = int(pf._n(r.get('Unidades_Caja', 1))) or 1

    _, r2 = pf.calc_h2(data['bod'], fe_map, hoy, fer)
    rep_h2_map = {v[0]: v[7] for v in r2}
    r1 = calc_simple(data['farm'], fe_map, def_, rep_h2_map)
    r4 = pf.calc_h4(data['falt30'])
    r4b = pf.calc_h4b(data['stock'], data['bod'])
    r4c = pf.calc_h4c(data['falt60'])

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Farm_Bodega'
    write_simple(ws1, r1, def_, hoy, sem)
    pf.write_h4(wb.create_sheet('Faltantes_AA'), r4, hoy)
    pf.write_h4b(wb.create_sheet('Por_Agotarse'), r4b, hoy)
    pf.write_h4c(wb.create_sheet('Faltantes_60D'), r4c, hoy)

    sal = os.path.join(WORK_DIR, f'Pedido_Fusion_Simple_AA_{hoy.strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(sal)

    n1 = len(r1)
    print(f'Farm_Bodega     : {n1} meds con cantidad a reponer')
    print(f'Faltantes AA 30d: {len(r4)} meds sin poder despachar en Atencion Abierta')
    print(f'Por agotarse    : {len(r4b)} meds con Bodega AA en 0 y cobertura '
          f'farmacia <= {pf.UMBRAL_PREQUIEBRE} dias')
    print(f'Faltantes 60d   : {len(r4c)} meds con faltante persistente vigente')
    print(f'\nExcel: {os.path.basename(sal)}')


if __name__ == '__main__':
    main()
