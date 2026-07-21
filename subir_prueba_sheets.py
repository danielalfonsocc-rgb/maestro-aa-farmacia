#!/usr/bin/env python3
"""Prueba puntual: sube el estado actual (post gt_maestro.py) de una COPIA de
prueba de la planilla GT a un Google Sheet nuevo y privado, reusando el token
OAuth ya emitido para Drive (token_drive.json, scope "drive" — válido también
para Sheets API). No toca el archivo real ni ningún Sheet existente.

Formato: mismo lenguaje visual que las planillas maestro (skill_gt/scripts/
generar.py + gt_maestro.aplicar_formato_maestro), con vuelta extra de diseño:
banda de título navy, filas coloreadas por establecimiento de destino
(incluida la columna Estado), encabezado azul, bordes, alineación por tipo
de columna, fila Pendiente resaltada, filtro básico y todo congelado arriba."""
import datetime
import os
import shutil
import sys

MAESTRO_DIR = r"C:\Users\danie\Downloads\maestro"
sys.path.insert(0, MAESTRO_DIR)
os.chdir(MAESTRO_DIR)

import openpyxl  # noqa: E402
import gt_maestro as GM  # noqa: E402
import generar as G  # noqa: E402 (paleta NAVY/BLUE/GREY/... — skill_gt/scripts ya está en sys.path via gt_maestro)

from google.oauth2.credentials import Credentials  # noqa: E402
from google.auth.transport.requests import Request  # noqa: E402
import gspread  # noqa: E402

TOKEN_FILE = os.path.join(MAESTRO_DIR, "token_drive.json")
SCOPES = ["https://www.googleapis.com/auth/drive"]

creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
if not creds.valid:
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(TOKEN_FILE, "w", encoding="utf-8") as fh:
            fh.write(creds.to_json())
    else:
        raise SystemExit("Token inválido y sin refresh_token — hay que re-autorizar con publicar_drive.py --setup")

gc = gspread.authorize(creds)


def _hex_a_float(hexval):
    if not hexval or len(str(hexval)) < 6:
        return None
    h = str(hexval)[-6:]
    return {"red": int(h[0:2], 16) / 255, "green": int(h[2:4], 16) / 255, "blue": int(h[4:6], 16) / 255}


BLANCO = {"red": 1, "green": 1, "blue": 1}
NAVY = _hex_a_float(G.NAVY)
BLUE = _hex_a_float(G.BLUE)
GREY = _hex_a_float(G.GREY)
TEXTO_OSCURO = {"red": 0x1F / 255, "green": 0x38 / 255, "blue": 0x64 / 255}
BORDE_GRIS = {"style": "SOLID", "color": {"red": 0xBF / 255, "green": 0xBF / 255, "blue": 0xBF / 255}}

# Columnas que se ven mejor centradas (códigos, fechas cortas, números) vs. las
# que se ven mejor a la izquierda (nombres, texto libre).
CENTRADAS = {"nreceta", "ndereceta", "rut", "periodoreceta", "numeroprescripciones", "estado",
             "refrigerado", "controlado", "fechadesolicitud", "fechaderetiroenfarmacia", "telefono"}

# El color por establecimiento lo decide gt_maestro.color_establecimiento()
# (hash estable, sin estado compartido) — se reusa tal cual para que el Excel
# real y esta prueba de Sheets pinten siempre el mismo color por
# establecimiento. No se duplica la paleta acá.

# 1) Prepara los datos de prueba en una copia descartable (no toca el archivo real)
test_path = r"C:\Users\danie\Downloads\_TEST_gt_maestro_sheets.xlsx"
shutil.copy2(r"C:\Users\danie\Downloads\GT PITRUFQUEN 2026 (2).xlsx", test_path)
wb, path = GM.cargar_maestro(test_path)
resumen, hojas = GM.sincronizar_gt_report(
    wb, r"C:\Users\danie\Downloads\04_Farmacia_Gestion_Territorial\reporteGestionTerritorial_16-07-2026_30-07-2026.xlsx"
)
GM.guardar(wb, path)
print("Sincronizado en copia de prueba:", resumen, hojas)

wb2 = openpyxl.load_workbook(path, data_only=True)
ws2 = wb2["JULIO 2026"]

# 2) Extrae encabezado (fila 2 del Excel — la 1 es la banda de título) + filas de
#    datos + el color de fila (por establecimiento) para replicarlo en el Sheet.
headers = [c.value for c in ws2[2]]
ncols = len(headers)
idx_estado = next((i for i, h in enumerate(headers) if GM._norm(h) == "estado"), None)
idx_pendiente = next((i for i, h in enumerate(headers) if GM._norm(h) == "pendiente"), None)

filas_datos = []
fills_fila = []
fuentes_fila = []
for r in range(3, ws2.max_row + 1):
    vals = [ws2.cell(row=r, column=c).value for c in range(1, ncols + 1)]
    if all(v in (None, "") for v in vals):
        continue
    fila = []
    for v in vals:
        if hasattr(v, "strftime"):
            fila.append(v.strftime("%d/%m/%Y"))
        else:
            fila.append("" if v is None else v)
    filas_datos.append(fila)
    # Fondo y texto de fila = lo que aplicar_formato_maestro ya calculó por
    # establecimiento (uniforme en toda la fila salvo negrita en Estado) — se
    # toman de la columna 1 (nunca es Estado).
    celda_ref = ws2.cell(row=r, column=1)
    fills_fila.append(_hex_a_float(celda_ref.fill.fgColor.rgb))
    fuentes_fila.append(_hex_a_float(celda_ref.font.color.rgb) if celda_ref.font.color else None)

print(f"Filas de datos: {len(filas_datos)}")

# 3) Crea el Sheet de prueba (privado, solo visible para la cuenta autenticada)
titulo_doc = f"GT PITRUFQUEN - PRUEBA (no oficial) {datetime.date.today().isoformat()}"
sh = gc.create(titulo_doc)
wsheet = sh.sheet1
wsheet.update_title("JULIO 2026")

FILA_TITULO, FILA_LEYENDA, FILA_HEADER, FILA_DATOS_INICIO = 1, 2, 3, 4
total_filas = len(filas_datos) + 3  # título + leyenda + encabezado
wsheet.resize(rows=max(total_filas + 5, 20), cols=max(ncols, 1))

titulo_banda = f"🏥 GESTIÓN TERRITORIAL — FARMACIA HOSPITAL DE PITRUFQUÉN — {wsheet.title}"
wsheet.update("A1", [[titulo_banda]], value_input_option="USER_ENTERED")
wsheet.update(f"A{FILA_HEADER}", [headers] + filas_datos, value_input_option="USER_ENTERED")

sheet_id = wsheet.id
requests = [
    # ── Banda de título ──────────────────────────────────────────────
    {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                               "startColumnIndex": 0, "endColumnIndex": ncols}, "mergeType": "MERGE_ALL"}},
    {"repeatCell": {
        "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": ncols},
        "cell": {"userEnteredFormat": {
            "backgroundColor": NAVY,
            "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE",
            "textFormat": {"bold": True, "fontSize": 14, "fontFamily": "Calibri", "foregroundColor": BLANCO},
        }},
        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)",
    }},
    {"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 0, "endIndex": 1},
        "properties": {"pixelSize": 40}, "fields": "pixelSize",
    }},
    # ── Fila leyenda (fondo navy clarito de base) ───────────────────
    {"repeatCell": {
        "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": ncols},
        "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.93, "green": 0.94, "blue": 0.97},
                                        "verticalAlignment": "MIDDLE"}},
        "fields": "userEnteredFormat(backgroundColor,verticalAlignment)",
    }},
    {"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 1, "endIndex": 2},
        "properties": {"pixelSize": 26}, "fields": "pixelSize",
    }},
    # ── Encabezado ───────────────────────────────────────────────────
    {"repeatCell": {
        "range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": 3, "startColumnIndex": 0, "endColumnIndex": ncols},
        "cell": {"userEnteredFormat": {
            "backgroundColor": BLUE,
            "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE", "wrapStrategy": "WRAP",
            "textFormat": {"bold": True, "fontSize": 10, "fontFamily": "Calibri", "foregroundColor": BLANCO},
        }},
        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,wrapStrategy,textFormat)",
    }},
    {"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": 2, "endIndex": 3},
        "properties": {"pixelSize": 34}, "fields": "pixelSize",
    }},
    # ── Fuente base de los datos (la cebra la pone addBanding aparte) ─
    {"repeatCell": {
        "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": ncols},
        "cell": {"userEnteredFormat": {"textFormat": {"fontFamily": "Calibri", "fontSize": 10}, "verticalAlignment": "MIDDLE"}},
        "fields": "userEnteredFormat(textFormat,verticalAlignment)",
    }},
    {"addBanding": {
        "bandedRange": {
            "range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": ncols},
            "rowProperties": {"headerColor": BLUE, "firstBandColor": BLANCO, "secondBandColor": GREY},
        }
    }},
    {"updateBorders": {
        "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": total_filas, "startColumnIndex": 0, "endColumnIndex": ncols},
        "top": BORDE_GRIS, "bottom": BORDE_GRIS, "left": BORDE_GRIS, "right": BORDE_GRIS,
        "innerHorizontal": BORDE_GRIS, "innerVertical": BORDE_GRIS,
    }},
    {"updateSheetProperties": {
        "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 3}},
        "fields": "gridProperties.frozenRowCount",
    }},
    # ── Filtro básico sobre encabezado + datos, para ordenar/filtrar fácil ─
    {"setBasicFilter": {
        "filter": {"range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": total_filas,
                              "startColumnIndex": 0, "endColumnIndex": ncols}}
    }},
    # ── Dropdown en Estado: mismas 9 opciones ya consolidadas en gt_maestro.py
    #    (las 6 propias + las 3 que ya existían en la planilla real: Validado QF,
    #    sin receta vigente, pendiente por falta stock) ─
    {"setDataValidation": {
        "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas,
                   "startColumnIndex": idx_estado, "endColumnIndex": idx_estado + 1},
        "rule": {
            "condition": {"type": "ONE_OF_LIST", "values": [{"userEnteredValue": v} for v in [
                "EN REVISIÓN", "EN PREPARACIÓN", "LISTA PARA RETIRO", "RETIRO EN VENTANILLA",
                "ENVIADA", "ENTREGADA", "VALIDADO QF", "SIN RECETA VIGENTE", "PENDIENTE POR FALTA STOCK",
            ]]},
            "showCustomUi": True, "strict": False,
        },
    }} if idx_estado is not None else None,
    # ── Formato condicional: Pendiente en negrita/naranjo cuando no está vacío ─
    {"addConditionalFormatRule": {
        "rule": {
            "ranges": [{"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas,
                        "startColumnIndex": idx_pendiente, "endColumnIndex": idx_pendiente + 1}],
            "booleanRule": {
                "condition": {"type": "NOT_BLANK"},
                "format": {"textFormat": {"bold": True, "foregroundColor": {"red": 0xC5 / 255, "green": 0x5A / 255, "blue": 0x11 / 255}}},
            },
        },
        "index": 0,
    }} if idx_pendiente is not None else None,
]
requests = [r for r in requests if r is not None]

# Alineación + anchos: el ancho lo decide directamente el Excel (columna ya
# auto-ajustada por aplicar_formato_maestro al contenido real) — se convierte
# de unidades de carácter de Excel a píxeles de Sheets (~7px/carácter + margen)
# en vez de mantener una tabla estática aparte que se puede desincronizar.
for i, h in enumerate(headers):
    clave = GM._norm(h)
    ancho_excel = ws2.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width or 14
    ancho = max(round(ancho_excel * 7 + 12), 50)
    requests.append({"updateDimensionProperties": {
        "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": i, "endIndex": i + 1},
        "properties": {"pixelSize": ancho}, "fields": "pixelSize",
    }})
    if clave in CENTRADAS:
        requests.append({"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": total_filas, "startColumnIndex": i, "endColumnIndex": i + 1},
            "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
            "fields": "userEnteredFormat.horizontalAlignment",
        }})

sh.batch_update({"requests": requests})

# 4) Colorea cada fila completa (incluida Estado) del color de su
#    establecimiento de destino — fondo Y texto, ya calculados por
#    aplicar_formato_maestro en el Excel (estilo pedido_fusion.py).
color_requests = []
for i in range(len(filas_datos)):
    if fills_fila[i] is None:
        continue
    fila_sheet = FILA_DATOS_INICIO + i
    cell_format = {"backgroundColor": fills_fila[i]}
    fields = "userEnteredFormat.backgroundColor"
    if fuentes_fila[i] is not None:
        cell_format["textFormat"] = {"foregroundColor": fuentes_fila[i]}
        fields = "userEnteredFormat(backgroundColor,textFormat)"
    color_requests.append({
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": fila_sheet - 1, "endRowIndex": fila_sheet,
                       "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": cell_format},
            "fields": fields,
        }
    })
if color_requests:
    TANDA = 400
    for ini in range(0, len(color_requests), TANDA):
        sh.batch_update({"requests": color_requests[ini:ini + TANDA]})

# 5) Hoja HISTORIAL (si existe en el maestro) — mismo contenido, formato
#    liviano (encabezado azul + cebra), sin color por establecimiento (no
#    aplica: es un log cronológico, no una hoja de recetas por mes).
if GM.HOJA_HISTORIAL in wb2.sheetnames:
    ws_hist = wb2[GM.HOJA_HISTORIAL]
    hist_headers = [c.value for c in ws_hist[1]]
    hist_ncols = len(hist_headers)
    filas_hist = []
    for r in range(2, ws_hist.max_row + 1):
        vals = [ws_hist.cell(row=r, column=c).value for c in range(1, hist_ncols + 1)]
        if all(v in (None, "") for v in vals):
            continue
        fila = [v.strftime("%d/%m/%Y %H:%M") if hasattr(v, "strftime") else ("" if v is None else v) for v in vals]
        filas_hist.append(fila)

    ws_hist_sheet = sh.add_worksheet(title=GM.HOJA_HISTORIAL, rows=max(len(filas_hist) + 5, 20), cols=hist_ncols)
    ws_hist_sheet.update("A1", [hist_headers] + filas_hist, value_input_option="USER_ENTERED")
    hist_sheet_id = ws_hist_sheet.id
    hist_total_filas = len(filas_hist) + 1
    hist_requests = [
        {"repeatCell": {
            "range": {"sheetId": hist_sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": hist_ncols},
            "cell": {"userEnteredFormat": {
                "backgroundColor": BLUE, "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE",
                "textFormat": {"bold": True, "fontSize": 10, "fontFamily": "Calibri", "foregroundColor": BLANCO},
            }},
            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)",
        }},
        {"addBanding": {
            "bandedRange": {
                "range": {"sheetId": hist_sheet_id, "startRowIndex": 0, "endRowIndex": hist_total_filas, "startColumnIndex": 0, "endColumnIndex": hist_ncols},
                "rowProperties": {"headerColor": BLUE, "firstBandColor": BLANCO, "secondBandColor": GREY},
            }
        }},
        {"updateBorders": {
            "range": {"sheetId": hist_sheet_id, "startRowIndex": 0, "endRowIndex": hist_total_filas, "startColumnIndex": 0, "endColumnIndex": hist_ncols},
            "top": BORDE_GRIS, "bottom": BORDE_GRIS, "left": BORDE_GRIS, "right": BORDE_GRIS,
            "innerHorizontal": BORDE_GRIS, "innerVertical": BORDE_GRIS,
        }},
        {"updateSheetProperties": {
            "properties": {"sheetId": hist_sheet_id, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount",
        }},
        {"setBasicFilter": {
            "filter": {"range": {"sheetId": hist_sheet_id, "startRowIndex": 0, "endRowIndex": hist_total_filas,
                                  "startColumnIndex": 0, "endColumnIndex": hist_ncols}}
        }},
    ]
    sh.batch_update({"requests": hist_requests})
    print(f"HISTORIAL: {len(filas_hist)} fila(s)")

print("URL:", sh.url)

os.remove(test_path)
if os.path.exists(test_path + ".bak"):
    os.remove(test_path + ".bak")
