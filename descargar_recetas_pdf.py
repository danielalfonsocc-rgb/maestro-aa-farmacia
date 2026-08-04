#!/usr/bin/env python3
"""
descargar_recetas_pdf.py — Descarga de SSASUR el PDF OFICIAL de cada receta
(por Nº) — "RECETA MÉDICA N°..." con firma del médico y checklist de
retiro, el mismo formato que produce el sitio al hacer clic en "Imprimir" —
para completar una Revisión de Solicitudes con los documentos reales, no
solo la lista de números que arma revision_solicitudes.py --xlsx.

Confirmado en vivo 23-07-2026, contrastado byte a byte contra el PDF de
referencia de Toltén (18-07-2026): el botón "Imprimir" del sitio abre el
diálogo de impresión NATIVO del sistema operativo (window.print()), que
Playwright no puede controlar — pero ese botón en realidad pide, por debajo,
el endpoint https://www.ssasur.cl/receta/impresion/pdf/<receta>/undefined,
que devuelve el PDF oficial directo. Se pide ese endpoint con la sesión ya
autenticada, sin necesidad de navegar ni clicar nada del formulario.

Uso:
  py descargar_recetas_pdf.py --estab "CESFAM FREIRE" --recetas 45858462,46045176 --debug
  py descargar_recetas_pdf.py --estab "CESFAM FREIRE" --feedback "C:/ruta/Feedback_Solicitud_CESFAM_FREIRE_2026-07-23.xlsx"
  py descargar_recetas_pdf.py --estab "CESFAM FREIRE" --rut "6.385.207-4"
      (busca en vivo las recetas vigentes de ese RUT — para cuando la receta
      es tan nueva que ni siquiera aparece en el CSV local; confirmado en
      vivo 23-07-2026 vía Consultar Receta -> pestaña Run, filtrando Estado
      distinto de ENTREGADA/CERRADA-INCOMPLETA/ANULADA)
"""
import argparse
import asyncio
import datetime
import os
import re
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from pypdf import PdfWriter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import AUTO_SSASUR as AS
from agregar_gt_manual import _CARPETA_LOCAL, GT_SOLICITUDES_DIR
from revision_solicitudes import _carpeta_salida

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))


def _leer_recetas_de_feedback(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(c).strip() == "N° Receta" for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"No encontré la columna 'N° Receta' en {path}")
    header = [str(c).strip() if c else "" for c in list(ws.iter_rows(values_only=True))[header_idx - 1]]
    col_receta = header.index("N° Receta")
    col_pac = header.index("Paciente") if "Paciente" in header else None
    out = []
    for row in list(ws.iter_rows(values_only=True))[header_idx:]:
        receta = row[col_receta] if col_receta < len(row) else None
        if not receta or str(receta).startswith("SIN_RECETA_"):
            continue
        paciente = row[col_pac] if col_pac is not None and col_pac < len(row) else ""
        out.append((str(receta).strip(), str(paciente or "").strip()))
    return out


ESTADOS_NO_VIGENTES = {"ENTREGADA", "CERRADA / INCOMPLETA", "ANULADA"}


def _es_vigente(estado):
    """Corregido 31-07-2026 (caso CESFAM Teodoro Schmidt): el sitio también
    devuelve 'CERRADA POR VENCIMIENTO' — una variante de cerrada que la
    igualdad exacta contra ESTADOS_NO_VIGENTES no capturaba, así que se
    colaban recetas vencidas al combinado para despachar. Cualquier estado
    que empiece con CERRADA, o sea ENTREGADA/ANULADA, se considera no
    vigente."""
    e = (estado or "").strip().upper()
    return e not in ESTADOS_NO_VIGENTES and not e.startswith("CERRADA")


def _leer_ruts_de_feedback(path):
    """Lee Paciente+RUT de un Feedback_Solicitud_*.xlsx (columnas ya
    conocidas por revision_solicitudes._escribir_feedback) para hacer
    búsqueda EN VIVO en SSASUR por RUT — usado cuando el Nº de receta que
    trae el CSV local ya está ENTREGADA/no existe (receta de una cuota tan
    nueva que el histórico local todavía no la tiene, caso frecuente en
    solicitudes de CESFAM Teodoro Schmidt 31-07-2026)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(c).strip() == "RUT" for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"No encontré la columna 'RUT' en {path}")
    header = [str(c).strip() if c else "" for c in list(ws.iter_rows(values_only=True))[header_idx - 1]]
    col_rut = header.index("RUT")
    col_pac = header.index("Paciente") if "Paciente" in header else None
    out = []
    vistos = set()
    for row in list(ws.iter_rows(values_only=True))[header_idx:]:
        rut = row[col_rut] if col_rut < len(row) else None
        if not rut:
            continue
        rut = str(rut).strip()
        if rut in vistos:
            continue
        vistos.add(rut)
        paciente = row[col_pac] if col_pac is not None and col_pac < len(row) else ""
        out.append((rut, str(paciente or "").strip()))
    return out


def _split_rut(rut):
    """'6.385.207-4' / '6385207-4' -> ('6385207', '4')."""
    limpio = re.sub(r"[.\s]", "", str(rut)).upper()
    numero, dv = limpio.rsplit("-", 1) if "-" in limpio else (limpio, "")
    return numero, dv


DEBUG_DIR = Path(MAESTRO_DIR) / "debug_ssasur_receta"


async def _volcar_debug(page, tag):
    """Guarda screenshot + HTML del estado actual de la página — para
    diagnosticar con evidencia real un fallo de búsqueda en vivo (SPA de
    SSASUR) en vez de seguir adivinando selectores a ciegas."""
    try:
        DEBUG_DIR.mkdir(exist_ok=True)
        await page.screenshot(path=str(DEBUG_DIR / f"{tag}.png"), full_page=True)
        html = await page.content()
        (DEBUG_DIR / f"{tag}.html").write_text(html, encoding="utf-8")
        print(f"  [DEBUG] Guardado {tag}.png / {tag}.html en {DEBUG_DIR}")
    except Exception as e:
        print(f"  [DEBUG] No pude volcar debug para {tag}: {e}")


async def _esperar_tabla_resultados(page, timeout=20_000):
    """Espera el ciclo de carga AJAX de #tablaResultados (aparece el
    indicador 'Cargando', luego desaparece).

    04-08-2026: esperar solo a que 'Cargando' DESAPAREZCA es una carrera —
    si el AJAX todavía no insertó el indicador cuando se evalúa la función,
    la condición ya es verdadera y la espera termina de inmediato, antes de
    que la tabla realmente cargue (confirmado con captura: tabla en '0 de 0
    registros' con el spinner recién apareciendo). Se espera primero a que
    'Cargando' APAREZCA (timeout corto, se ignora si la respuesta fue tan
    rápida que nunca se alcanzó a ver) y solo después a que desaparezca."""
    try:
        await page.wait_for_function(
            "() => document.body.innerText.includes('Cargando')",
            timeout=2_000,
        )
    except Exception:
        pass
    await page.wait_for_function(
        "() => !document.body.innerText.includes('Cargando')",
        timeout=timeout,
    )
    await page.wait_for_timeout(500)


async def _buscar_vigentes_por_rut(page, rut, origen_filtro="PITRUFQUEN HOSP.", debug_tag=None):
    """Busca por RUN o ID Único de Paciente (pestaña activa por defecto en
    Consultar Receta — confirmado en vivo 23-07-2026) y devuelve las
    recetas NO entregadas del establecimiento de origen indicado — el caso
    de uso real: encontrar la receta más nueva de un RUT que SSASUR aún no
    despachó, cuando ni siquiera aparece todavía en el CSV local. La tabla
    de resultados viene ordenada por Nº de receta descendente (las más
    nuevas primero), por eso basta con mirar la primera página.

    03-08-2026: confirmado en vivo con captura de pantalla del usuario — la
    tabla de resultados SÍ es la esperada (columnas Establecimiento
    Origen/.../Estado/.../Fecha Ingreso/Fecha Entrega, con "Estado" en el
    índice 9 tal como asumía este código: n_receta=r[0], origen=r[1],
    estado=r[9]) y la ruta "Resultados > por Run o ID Único de Paciente" es
    el destino legítimo tras buscar — no es una "salida" a una pantalla
    ajena. El fallo real es que, sin clickear explícitamente la pestaña
    "Run" antes de llenar el formulario, la búsqueda no toma el campo
    correcto y devuelve 0 filas para prácticamente todos los RUT (probado
    revirtiendo el clic: mismo patrón de "0 encontrado" que sin el fix).
    Clickear la pestaña SÍ ayuda (probado: pasa de 0 a encontrar casos
    reales) — lo que rompía todo en un intento anterior fue exigir que
    #rut quedara "visible" según el criterio estricto de Playwright antes
    de llenar; sin esa espera, fill() funciona bien."""
    for sel in ('a:has-text("Run")', 'button:has-text("Run")', '[role=tab]:has-text("Run")'):
        try:
            await page.click(sel, timeout=3_000)
            await page.wait_for_timeout(300)
            break
        except Exception:
            continue
    if debug_tag:
        await _volcar_debug(page, f"{debug_tag}_1_tras_click_run")
    numero, dv = _split_rut(rut)
    await page.fill("#rut", numero)
    await page.fill("#dv", dv)
    if debug_tag:
        await _volcar_debug(page, f"{debug_tag}_2_tras_llenar")
    await AS._click_primero(page, ('button:has-text("Buscar"):visible', 'a:has-text("Buscar"):visible'), "Buscar")
    await page.wait_for_load_state("networkidle")
    if debug_tag:
        await _volcar_debug(page, f"{debug_tag}_3_tras_buscar")
    try:
        # Corregido 31-07-2026: exigir >=1 fila para dar por "cargada" la
        # tabla es incorrecto — un RUT con CERO recetas vigentes (resultado
        # legítimo, no una falla) nunca cumple esa condición y siempre
        # agotaba el timeout, disparando el [AVISO] aunque la búsqueda haya
        # funcionado bien. Ahora solo se espera a que desaparezca el
        # indicador "Cargando"; la tabla vacía se lee después como 0 filas.
        await _esperar_tabla_resultados(page)
    except Exception:
        print(f"  [AVISO] RUT {rut}: no cargó la tabla de resultados a tiempo (url={page.url}).")
        if debug_tag:
            await _volcar_debug(page, debug_tag)
        return None

    existe_tabla = await page.evaluate("() => !!document.querySelector('#tablaResultados')")
    if not existe_tabla:
        print(f"  [AVISO] RUT {rut}: no encontré #tablaResultados en la página actual (url={page.url}).")
        if debug_tag:
            await _volcar_debug(page, debug_tag)
        return None

    # Pedido por el usuario 04-08-2026: en vez de traer TODAS las filas del
    # RUT (hasta 300+ recetas, mezclando otros establecimientos de origen
    # entre medio) y filtrar en Python, usar el propio filtro de columna
    # "Establecimiento Origen" del sitio (select2 sobre un <select> oculto
    # con opciones fijas como "PITRUFQUEN HOSP.", confirmado en vivo con
    # captura de pantalla). Más simple y evita perder una receta vigente de
    # Pitrufquén que quedara fuera de las primeras filas de una tabla sin
    # filtrar. Se ubica el <select> por su opción (no por id, que el sitio
    # genera dinámicamente y cambia entre cargas).
    if origen_filtro:
        filtro_aplicado = await page.evaluate(
            """(valor) => {
                const t = document.querySelector('#tablaResultados');
                if (!t) return false;
                const wrapper = document.querySelector('#tablaResultados_wrapper') || t;
                const selects = wrapper.querySelectorAll('select');
                for (const s of selects) {
                    if ([...s.options].some(o => o.value === valor)) {
                        s.value = valor;
                        if (window.jQuery) { window.jQuery(s).trigger('change'); }
                        s.dispatchEvent(new Event('change', {bubbles: true}));
                        return true;
                    }
                }
                return false;
            }""",
            origen_filtro,
        )
        if filtro_aplicado:
            try:
                await _esperar_tabla_resultados(page)
            except Exception:
                print(f"  [AVISO] RUT {rut}: el filtro de Establecimiento Origen no recargó la tabla a tiempo (url={page.url}).")
                if debug_tag:
                    await _volcar_debug(page, debug_tag)
                return None
        elif debug_tag:
            print(f"  [AVISO] RUT {rut}: no encontré el filtro de Establecimiento Origen — se sigue filtrando en Python.")

    if debug_tag:
        await _volcar_debug(page, f"{debug_tag}_4_tras_filtro_origen")

    filas = await page.evaluate(r"""() => {
      const t = document.querySelector('#tablaResultados');
      return [...t.querySelectorAll('tbody tr')].map(tr =>
        [...tr.querySelectorAll('td')].map(td => td.innerText.trim()));
    }""")
    vigentes = []
    for r in filas:
        if len(r) <= 9:
            continue
        n_receta, origen, estado = r[0], r[1], r[9]
        if not n_receta or not n_receta.isdigit():
            continue
        if not _es_vigente(estado):
            continue
        if origen_filtro and origen_filtro.upper() not in origen.upper():
            continue
        # Pedido por el usuario 03-08-2026: exigir que "Fecha Entrega"
        # (columna índice 13, confirmada por captura de pantalla —
        # inmediatamente después de "Fecha Ingreso") caiga en 2026 — una
        # receta vigente pero con fecha de entrega de otro año no cuenta.
        fecha_entrega = r[13].strip() if len(r) > 13 else ""
        if not fecha_entrega.endswith("2026"):
            continue
        vigentes.append((n_receta, estado, fecha_entrega))
    return vigentes


PDF_ENDPOINT = "https://www.ssasur.cl/receta/impresion/pdf/{n_receta}/undefined"


async def _descargar_una(page, n_receta, dest_path, debug):
    """Confirmado en vivo 23-07-2026 (contrastado byte a byte contra el
    formato de referencia de Toltén 18-07-2026): los botones "Imprimir*" del
    sitio abren el diálogo de impresión NATIVO del sistema operativo
    (window.print()), que Playwright no puede controlar. Pero ese botón
    dispara internamente una petición a un endpoint que SÍ devuelve el PDF
    oficial directo — "RECETA MÉDICA N°..." con firma del médico y checklist
    de retiro, no la vista web de consulta. Se pide ese endpoint directo con
    la sesión ya autenticada, sin necesidad de navegar el formulario.

    Corregido 31-07-2026 (caso CESFAM Teodoro Schmidt, corrida de 36 RUT):
    una falla de red transitoria (DNS/VPN) en request.get() no está atrapada
    por ningún try/except más arriba — sin este bloque, una sola receta con
    mala suerte de red mata TODO el proceso en asyncio.run() y se pierde el
    trabajo de las recetas ya descargadas (el combinado solo se arma al
    final del bucle). Se reintenta una vez tras una pausa corta antes de
    darse por vencido con esa receta puntual."""
    for intento in (1, 2):
        try:
            resp = await page.context.request.get(PDF_ENDPOINT.format(n_receta=n_receta))
        except Exception as e:
            if intento == 1:
                print(f"  [AVISO] {n_receta}: error de red ({e}) — reintentando...")
                await page.wait_for_timeout(3_000)
                continue
            print(f"  [ERROR] {n_receta}: error de red persistente — {e}")
            return False
        break
    if resp.status != 200 or "pdf" not in (resp.headers.get("content-type") or "").lower():
        print(f"  [ERROR] {n_receta}: el endpoint no devolvió un PDF (status {resp.status}).")
        if debug:
            body = await resp.body()
            (AS.MAESTRO_DIR / f"debug_receta_{n_receta}_respuesta.bin").write_bytes(body)
        return False
    body = await resp.body()
    if not body.startswith(b"%PDF"):
        print(f"  [ERROR] {n_receta}: la respuesta no es un PDF válido.")
        return False
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    dest_path.write_bytes(body)
    print(f"  ✓ {dest_path.name}  ({dest_path.stat().st_size // 1024:,} KB)")
    return True


async def _main_async(estab, recetas, debug, rut=None, ruts=None):
    carpeta_local = _CARPETA_LOCAL.get(estab.upper())
    if not carpeta_local:
        print(f"[ERROR] '{estab}' no está en el mapeo de establecimientos.")
        return
    base_dir = os.path.join(GT_SOLICITUDES_DIR, carpeta_local, "Revisión de Solicitudes")
    salida_dir = _carpeta_salida(base_dir, datetime.date.today())

    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--start-maximized"])
        if AS.SESSION_FILE.exists():
            context = await browser.new_context(accept_downloads=True, storage_state=str(AS.SESSION_FILE))
            print("(Sesión guardada encontrada — puede que no necesites logarte)")
        else:
            context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        print("Logéate en SSASUR (tienes 5 minutos)...")
        await page.goto(AS.DASHBOARD_URL)
        await page.wait_for_selector('button:has-text("ABASTECIMIENTO"), div:has-text("ABASTECIMIENTO")',
                                      timeout=AS.TIMEOUT_LOGIN)
        await context.storage_state(path=str(AS.SESSION_FILE))
        print("✓ Sesión detectada")

        await AS.entrar_receta(page)

        async def _ir_a_consultar_receta():
            """Corregido 31-07-2026 (caso CESFAM Teodoro Schmidt, 36 RUT en
            lote): reusar la misma pantalla de resultados para buscar un
            segundo RUT no funciona — tras la primera búsqueda el sitio deja
            #rut como type=hidden con el valor viejo pegado y las siguientes
            fill() cuelgan 30s cada una. Hay que volver a entrar a Consultar
            Receta (vía AS.entrar_receta, que ya sabe resetear al inicio del
            módulo) ANTES de cada búsqueda, no solo una vez al principio."""
            await AS.entrar_receta(page)
            for sel in ('a:has-text("Reportes")', 'button:has-text("Reportes")'):
                try:
                    await page.click(sel, timeout=3_000)
                    await page.wait_for_timeout(700)
                    break
                except Exception:
                    continue
            for sel in ('a:has-text("Consultar Receta")', 'button:has-text("Consultar Receta")'):
                try:
                    await page.click(sel, timeout=3_000)
                    await page.wait_for_load_state("networkidle")
                    break
                except Exception:
                    continue

        rut_targets = ([(rut, "")] if rut else []) + list(ruts or [])
        encontrados_por_rut = {}
        fallidos_tecnicos = set()
        debug_restantes = 3  # capar volcados de debug para no llenar disco
        for rut_i, nombre_i in rut_targets:
            etiqueta = nombre_i or rut_i
            print(f"\nBuscando recetas vigentes de {etiqueta} (RUT {rut_i}, origen PITRUFQUEN HOSP.)...")
            vigentes = None
            # 3 intentos: el sitio a veces deja la pantalla en un estado
            # "stale" (campo oculto, tabla anterior, redirección a
            # resultados) que un solo reintento no siempre alcanza a
            # limpiar — confirmado en vivo 03-08-2026, mismo caso.
            for intento in (1, 2, 3):
                try:
                    await _ir_a_consultar_receta()
                    tag = None
                    if debug_restantes > 0:
                        tag = re.sub(r"[^A-Za-z0-9]+", "_", f"{rut_i}_{nombre_i}_i{intento}")
                        debug_restantes -= 1
                    vigentes = await _buscar_vigentes_por_rut(page, rut_i, debug_tag=tag)
                except Exception as e:
                    print(f"  [ERROR] {rut_i}: {e}")
                    vigentes = None
                if vigentes is not None:
                    break
            if vigentes is None:
                print(f"  [ERROR] {rut_i}: la búsqueda no se completó tras {intento} intentos — requiere revisión manual.")
                vigentes = []
                fallidos_tecnicos.add(rut_i)
            elif not vigentes:
                print("  No se encontraron recetas vigentes (no ENTREGADA/CERRADA/ANULADA) para ese RUT.")
            encontrados_por_rut[rut_i] = vigentes
            for n_receta, estado, fecha_entrega in vigentes:
                print(f"  Encontrada: receta {n_receta} — Estado: {estado} — Fecha Entrega: {fecha_entrega}")
            recetas = list(recetas) + [(n, nombre_i) for n, _, _ in vigentes]

        if not recetas:
            print("No hay recetas para descargar.")
            await browser.close()
            return

        # _descargar_una pide el endpoint PDF directo — no navega el
        # formulario, así que no hace falta "volver a la pantalla de
        # consulta" entre una receta y otra como antes. Ya trae su propio
        # reintento de red; igual se envuelve en try/except para que una
        # falla imprevista en una receta puntual no tumbe el resto del lote.
        ok, fallidos, descargados = 0, [], []
        for n_receta, paciente in recetas:
            nombre_archivo = re.sub(r"\s+", "", paciente) or "Paciente"
            dest = Path(os.path.join(salida_dir, f"Receta_{n_receta}_{nombre_archivo}.pdf"))
            print(f"\nReceta {n_receta} ({paciente})...")
            try:
                exito = await _descargar_una(page, n_receta, dest, debug)
            except Exception as e:
                print(f"  [ERROR] {n_receta}: {e}")
                exito = False
            if exito:
                ok += 1
                descargados.append(dest)
            else:
                fallidos.append(n_receta)

        print(f"\n{ok}/{len(recetas)} PDF descargados. Fallidos: {fallidos or 'ninguno'}")
        await browser.close()

    if descargados:
        fecha_str = datetime.date.today().strftime("%Y-%m-%d")
        combinado = os.path.join(salida_dir, f"Recetas_Combinadas_{carpeta_local}_{fecha_str}.pdf")
        writer = PdfWriter()
        for p in descargados:
            writer.append(str(p))
        with open(combinado, "wb") as fh:
            writer.write(fh)
        print(f"Guardado: {combinado}  ({len(descargados)} receta(s))")

        pdfs_dir = os.path.join(salida_dir, "PDFs individuales")
        os.makedirs(pdfs_dir, exist_ok=True)
        for p in descargados:
            os.replace(str(p), os.path.join(pdfs_dir, p.name))
        print(f"{len(descargados)} PDF individual(es) movido(s) a 'PDFs individuales'.")

    if ruts:
        fecha_str = datetime.date.today().strftime("%Y-%m-%d")
        resumen_path = os.path.join(salida_dir, f"Resumen_Busqueda_SSASUR_{carpeta_local}_{fecha_str}.xlsx")
        _escribir_resumen_busqueda(resumen_path, rut_targets, encontrados_por_rut, fallidos_tecnicos)
        print(f"Guardado: {resumen_path}")
        if fallidos_tecnicos:
            print(f"[AVISO] {len(fallidos_tecnicos)} RUT no se pudieron buscar por falla técnica del sitio "
                  f"(no es un '0 recetas' real) — quedan marcados en el Resumen para reintentar.")

    # Pedido por el usuario 04-08-2026: subir a Drive en la MISMA corrida —
    # antes había que acordarse de correr publicar_drive.py --solo-gt a
    # mano y el caso real fue que no se subió nada. Mismo patrón que usa
    # Clozapina/Centinela en AUTO_SSASUR.py.
    if descargados:
        pub_py = os.path.join(MAESTRO_DIR, "publicar_drive.py")
        if os.path.exists(pub_py) and os.path.exists(os.path.join(MAESTRO_DIR, "token_drive.json")):
            print("\n[Drive] Sincronizando Revisión de Solicitudes a Drive...")
            pret = subprocess.run(
                [sys.executable, pub_py, "--solo-gt"],
                cwd=MAESTRO_DIR,
                env={**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
            )
            if pret.returncode != 0:
                print(f"  [aviso] publicar_drive.py --solo-gt terminó con código {pret.returncode}")


def _escribir_resumen_busqueda(path, rut_targets, encontrados_por_rut, fallidos_tecnicos=None):
    """Feedback de la búsqueda EN VIVO por RUT en SSASUR — una fila por
    paciente con las recetas vigentes (no ENTREGADA/CERRADA/ANULADA)
    encontradas en origen Pitrufquén, o el aviso de que no se encontró
    ninguna (para revisión manual del QF).

    Corregido 03-08-2026: distingue "se buscó y no había ninguna vigente"
    (resultado real) de "la búsqueda no se completó por falla técnica del
    sitio" (fallidos_tecnicos) — antes ambos casos se escribían con el mismo
    texto de "sin recetas vigentes", lo que hacía indistinguible un
    resultado confiable de uno que había que reintentar."""
    fallidos_tecnicos = fallidos_tecnicos or set()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Paciente", "RUT", "Recetas vigentes encontradas (Nº — Estado — Fecha Entrega)", "Observación"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for rut_i, nombre_i in rut_targets:
        vigentes = encontrados_por_rut.get(rut_i, [])
        if vigentes:
            texto = "; ".join(f"{n} — {e} — {fe}" for n, e, fe in vigentes)
            obs = "PDF descargado y combinado." if len(vigentes) == 1 else "Varias vigentes — se descargaron todas, revisar cuál corresponde."
        elif rut_i in fallidos_tecnicos:
            texto = ""
            obs = "FALLA TÉCNICA: la búsqueda en SSASUR no se completó — reintentar, NO asumir que no tiene receta vigente."
        else:
            texto = ""
            obs = "Sin recetas vigentes en SSASUR (origen Pitrufquén) — revisar manualmente."
        ws.append([nombre_i, rut_i, texto, obs])
    anchos = [30, 14, 40, 45]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--estab", required=True)
    ap.add_argument("--feedback", help="Ruta a un Feedback_Solicitud_*.xlsx generado por revision_solicitudes.py")
    ap.add_argument("--recetas", help="Lista de N° de receta separados por coma (alternativa a --feedback)")
    ap.add_argument("--rut", help="Busca las recetas VIGENTES (no entregadas) de este RUT en vivo en SSASUR — "
                                   "para cuando la receta es tan nueva que todavía no está en el CSV local")
    ap.add_argument("--rut-live-desde-feedback",
                     help="Ruta a un Feedback_Solicitud_*.xlsx: busca EN VIVO en SSASUR, por cada RUT de la "
                          "planilla, las recetas vigentes (no ENTREGADA/CERRADA/ANULADA) de origen Pitrufquén, "
                          "en una sola sesión — para solicitudes donde el CSV local no tiene la cuota nueva")
    ap.add_argument("--debug", action="store_true", help="Vuelca formularios y screenshots para ajustar selectores")
    a = ap.parse_args()

    ruts = None
    if a.rut_live_desde_feedback:
        ruts = _leer_ruts_de_feedback(a.rut_live_desde_feedback)
        recetas = []
    elif a.feedback:
        recetas = _leer_recetas_de_feedback(a.feedback)
    elif a.recetas:
        recetas = [(n.strip(), "") for n in a.recetas.split(",") if n.strip()]
    elif a.rut:
        recetas = []
    else:
        print("[ERROR] Pasa --feedback, --recetas, --rut o --rut-live-desde-feedback.")
        return
    if not recetas and not a.rut and not ruts:
        print("No hay recetas para procesar.")
        return
    if recetas:
        print(f"{len(recetas)} receta(s) a descargar.")
    if ruts:
        print(f"{len(ruts)} RUT(s) a buscar en vivo en SSASUR.")
    asyncio.run(_main_async(a.estab, recetas, a.debug, rut=a.rut, ruts=ruts))


if __name__ == "__main__":
    main()
