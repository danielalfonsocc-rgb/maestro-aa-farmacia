#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
publicar_escritorio.py — Centraliza en el ESCRITORIO las salidas de todos los
procesos de la Farmacia AT Abierta, para revisar los resultados de un vistazo
sin entrar a la carpeta del repositorio.

  Escritorio\Farmacia AA\
    ├── Abrir App de Pedidos.lnk          (acceso directo → ABRIR_APP.bat)
    ├── AUTO_SSASUR.lnk                   (acceso directo → AUTO_SSASUR.bat)
    ├── Gestion Territorial.lnk           (acceso directo → GT.bat)
    ├── Recetas Cheque ISP.lnk            (acceso directo → RECETAS_CHEQUE.bat)
    ├── LEEME.txt
    ├── _ultima_sync.txt                  (qué se copió y cuándo)
    ├── 1 - App Pedidos\        Consolidado_AA_MAESTRO.xlsx + Resumen_Pedidos_AA.xlsx
    ├── 2 - Gestion Territorial\  ÚLTIMO rango al frente + Historial\<rango>\
    ├── 3 - Recetas Cheque\     SOLO un acceso directo a la carpeta local (datos de
    │                           pacientes NO se copian a la nube de OneDrive)
    ├── 4 - Auditoria Prescripcion\  Auditoria_Prescripcion_Resumen.xlsx (legible)
    ├── 5 - Reposicion\         Reposicion_DiasHabiles_AA.xlsx (plan días hábiles)
    ├── 6 - Pedido Fusionado\   Pedido_Fusion_AA.xlsx (Farm_Bod + Bod_Farmacos + Dialisis)
    ├── 7 - Centinela\          Centinela_Reportes\<Sxx>\ (json + pdf) por semana
    └── 8 - Auditoria Duplicados\  Accesos a Agente Duplicados IA + Auditoria Profunda
                                    (corren a demanda; el Excel con nombres de pacientes
                                    NO se copia aquí — queda solo en la carpeta local)

IMPORTANTE: este script COPIA, no mueve. El repositorio sigue siendo la fuente de
verdad — la app Streamlit lee el Consolidado del repo y PUBLICAR_DATOS.bat publica
desde el repo. Aquí solo dejamos copias legibles, ordenadas por proceso.

Uso:
    py publicar_escritorio.py            # sincroniza TODO
    py publicar_escritorio.py --app          # solo Consolidado/Resumen (maestro)
    py publicar_escritorio.py --gt           # solo Gestion Territorial (out_gt)
    py publicar_escritorio.py --rch          # solo el acceso directo de recetas cheque
    py publicar_escritorio.py --auditoria
    py publicar_escritorio.py --reposicion   # solo Reposicion_DiasHabiles_AA.xlsx
    py publicar_escritorio.py --centinela    # solo Centinela_Reportes\
    py publicar_escritorio.py --duplicados   # solo accesos de Agente/Auditoria Duplicados
    py publicar_escritorio.py --enlaces      # solo (re)crea carpetas, LEEME y accesos
"""
import os
import re
import sys
import glob
import json
import shutil
import subprocess
from datetime import datetime

# Nombre de carpeta de un rango GT: DD-MM-AAAA_DD-MM-AAAA
_RANGO_RE = re.compile(r"^\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{4}$")

# Detecta el establecimiento de destino en nombres de archivo GT
_TIPO_GT_RE = re.compile(
    r"^(.+?)_(Planilla|Letrero|Controlados_Planilla|Verificacion)\.(xlsx|pdf)$",
    re.IGNORECASE,
)

def _destino_de_archivo(nombre):
    m = _TIPO_GT_RE.match(nombre)
    return m.group(1).replace("_", " ") if m else None

# ── Rutas base ───────────────────────────────────────────────────────────────
WORK_DIR = os.path.dirname(os.path.abspath(__file__))

# Carpeta del formulario ISP: fuente única = utils_aa.py (configurable por
# variable de entorno MAESTRO_RCH_DIR — ver utils_aa.py).
sys.path.insert(0, WORK_DIR)
from utils_aa import RCH_DIR, setup_stdout
setup_stdout()  # evita UnicodeEncodeError en consolas cp1252 (mensajes usan →, tildes)
PREFIJO_FORM = "Formulario-Notificacion-Recetas-Cheque"

NOMBRE_CARPETA = "Farmacia AA"
SUB_APP   = "1 - App Pedidos"
SUB_GT    = "2 - Gestion Territorial"
SUB_RCH    = "3 - Recetas Cheque"
SUB_AUDIT  = "4 - Auditoria Prescripcion"
SUB_REP    = "5 - Reposicion"
SUB_PEDIDO = "6 - Pedido Fusionado"
SUB_CENTINELA = "7 - Centinela"
SUB_DUP    = "8 - Auditoria Duplicados"

# Iconos para distinguir los accesos directos (shell32.dll, índices clásicos).
_SHELL32 = os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "System32", "shell32.dll")
def _icon(idx):
    return f"{_SHELL32},{idx}"
_ICON_APP     = _icon(13)   # globo de red  → app web
_ICON_REFRESH = _icon(238)  # flechas sync  → actualizar/publicar
_ICON_DOC     = _icon(1)    # documento     → planillas GT
_ICON_RUN     = _icon(2)    # aplicación    → proceso recetas cheque
_ICON_FOLDER  = _icon(4)    # carpeta       → abrir carpeta local
_ICON_LUPA    = _icon(23)   # lupa/buscar   → agentes de duplicados


def detectar_escritorio():
    """Ruta del Escritorio probando las ubicaciones habituales (OneDrive lo
    redirige en este equipo)."""
    perfil = os.environ.get("USERPROFILE", os.path.expanduser("~"))
    candidatos = []
    od = os.environ.get("OneDrive")
    if od:
        candidatos += [os.path.join(od, "Desktop"), os.path.join(od, "Escritorio")]
    candidatos += [
        os.path.join(perfil, "OneDrive", "Desktop"),
        os.path.join(perfil, "OneDrive", "Escritorio"),
        os.path.join(perfil, "Desktop"),
        os.path.join(perfil, "Escritorio"),
    ]
    for c in candidatos:
        if os.path.isdir(c):
            return c
    destino = os.path.join(perfil, "Desktop")
    os.makedirs(destino, exist_ok=True)
    return destino


DESKTOP = detectar_escritorio()
BASE    = os.path.join(DESKTOP, NOMBRE_CARPETA)


# ── Reporte / log ────────────────────────────────────────────────────────────
class _Reporte:
    def __init__(self):
        self.ok = self.skip = self.fail = 0
        self.log = []

    def say(self, msg):
        print(msg)
        self.log.append(msg)


REP = _Reporte()


# ── Utilidades de copia (incremental, tolerante a archivos abiertos) ──────────
def _es_temporal(nombre):
    return nombre.startswith("~$") or nombre.endswith(".tmp")


def _igual(src, dst):
    """True si dst ya existe y coincide en tamaño y fecha (±2s) → no recopiar."""
    try:
        a, b = os.stat(src), os.stat(dst)
    except OSError:
        return False
    return a.st_size == b.st_size and abs(a.st_mtime - b.st_mtime) <= 2


def _copiar(src, dst_dir, *, nuevo_nombre=None):
    """Copia src a dst_dir conservando metadatos. Salta si ya está igual.
    No aborta si el archivo está abierto en Excel o bloqueado por OneDrive."""
    if not src or not os.path.isfile(src):
        return "none"
    os.makedirs(dst_dir, exist_ok=True)
    destino = os.path.join(dst_dir, nuevo_nombre or os.path.basename(src))
    if _igual(src, destino):
        REP.skip += 1
        return "skip"
    try:
        shutil.copy2(src, destino)
        REP.ok += 1
        return "ok"
    except PermissionError:
        REP.fail += 1
        REP.say(f"  [aviso] abierto/bloqueado, no copiado: {os.path.basename(src)}")
        return "fail"
    except OSError as e:
        REP.fail += 1
        REP.say(f"  [aviso] error copiando {os.path.basename(src)}: {e}")
        return "fail"


def _mas_reciente(patron):
    cand = [f for f in glob.glob(patron) if not _es_temporal(os.path.basename(f))]
    return max(cand, key=os.path.getmtime) if cand else None


def _espejo(src_dir, dst_dir):
    """Copia recursiva incremental de src_dir → dst_dir (ignora temporales)."""
    n = 0
    for raiz, _dirs, files in os.walk(src_dir):
        rel = os.path.relpath(raiz, src_dir)
        destino = dst_dir if rel == "." else os.path.join(dst_dir, rel)
        for f in files:
            if _es_temporal(f):
                continue
            if _copiar(os.path.join(raiz, f), destino) in ("ok", "skip"):
                n += 1
    return n


def _limpiar_archivos_sueltos(dirpath):
    """Borra solo los archivos (no subcarpetas) que cuelgan de dirpath."""
    if not os.path.isdir(dirpath):
        return
    for nombre in os.listdir(dirpath):
        p = os.path.join(dirpath, nombre)
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass


# ── Sincronizadores por proceso ──────────────────────────────────────────────
def sync_app():
    dst = os.path.join(BASE, SUB_APP)
    # El Consolidado puede tener respaldo fechado si el base estaba abierto;
    # publicamos siempre la versión MÁS reciente con el nombre canónico.
    _copiar(_mas_reciente(os.path.join(WORK_DIR, "Consolidado_AA_MAESTRO*.xlsx")),
            dst, nuevo_nombre="Consolidado_AA_MAESTRO.xlsx")
    _copiar(_mas_reciente(os.path.join(WORK_DIR, "Resumen_Pedidos_AA*.xlsx")),
            dst, nuevo_nombre="Resumen_Pedidos_AA.xlsx")
    REP.say(f"[App Pedidos] Consolidado y Resumen → «{SUB_APP}»")


def _copiar_rango_por_estab(rango_dir, dst_base):
    """Copia archivos de un rango GT en dst_base/<ESTABLECIMIENTO>/.
    Archivos sin establecimiento conocido van directamente a dst_base."""
    archivos = sorted(
        glob.glob(os.path.join(rango_dir, "*.xlsx")) +
        glob.glob(os.path.join(rango_dir, "*.pdf"))
    )
    estabs = set()
    for f in archivos:
        nb = os.path.basename(f)
        if _es_temporal(nb):
            continue
        destino = _destino_de_archivo(nb)
        if destino:
            estabs.add(destino)
            _copiar(f, os.path.join(dst_base, destino))
        else:
            _copiar(f, dst_base)
    return sorted(estabs)


def sync_gt():
    src = os.path.join(WORK_DIR, "out_gt")
    base_gt = os.path.join(BASE, SUB_GT)
    if not os.path.isdir(src):
        REP.say("[Gestion Territorial] (sin out_gt todavía)")
        return
    rangos = sorted(
        [d for d in glob.glob(os.path.join(src, "*")) if os.path.isdir(d)],
        key=os.path.basename,
    )
    if not rangos:
        REP.say("[Gestion Territorial] (sin rangos en out_gt)")
        return

    # 1) Historial: cada rango en su carpeta, organizado por establecimiento
    hist = os.path.join(base_gt, "Historial")
    for d in rangos:
        _copiar_rango_por_estab(d, os.path.join(hist, os.path.basename(d)))

    # 2) Último rango → carpetas por establecimiento en la raíz GT (al frente)
    ultimo = max(rangos, key=os.path.getmtime)
    rango_actual = os.path.basename(ultimo)
    marca = os.path.join(base_gt, "_rango_actual.txt")
    previo = None
    if os.path.isfile(marca):
        try:
            previo = open(marca, encoding="utf-8").read().strip()
        except OSError:
            pass
    if previo != rango_actual:
        # Nuevo rango: limpia archivos sueltos y carpetas de establecimientos viejos
        _limpiar_archivos_sueltos(base_gt)
        for nombre in os.listdir(base_gt):
            p = os.path.join(base_gt, nombre)
            if os.path.isdir(p) and nombre not in ("Historial",) and not _RANGO_RE.match(nombre):
                shutil.rmtree(p, ignore_errors=True)
    estabs = _copiar_rango_por_estab(ultimo, base_gt)
    try:
        with open(marca, "w", encoding="utf-8") as fh:
            fh.write(rango_actual)
    except OSError:
        pass
    estabs_str = ", ".join(estabs) if estabs else "(ninguno)"
    REP.say(f"[Gestion Territorial] último «{rango_actual}» · "
            f"establecimientos: {estabs_str} · {len(rangos)} rango(s) en Historial")


def sync_rch():
    """NO copia el formulario (datos de pacientes) a la nube de OneDrive: deja
    solo un acceso directo a la carpeta local original, y purga cualquier copia
    previa que se hubiera subido."""
    dst = os.path.join(BASE, SUB_RCH)
    os.makedirs(dst, exist_ok=True)
    # Purga: quita formularios que pudieran haberse copiado antes (privacidad).
    purgados = 0
    for viejo in glob.glob(os.path.join(dst, PREFIJO_FORM + "*.xlsx")):
        try:
            os.remove(viejo)
            purgados += 1
        except OSError:
            pass
    if purgados:
        REP.say(f"  [privacidad] quitada(s) {purgados} copia(s) de la nube")
    # Acceso directo a la carpeta local del formulario.
    if os.path.isdir(RCH_DIR):
        _crear_lnk(os.path.join(dst, "Abrir carpeta Recetas Cheque.lnk"), RCH_DIR,
                   "Abre la carpeta LOCAL del formulario ISP (no se sube a la nube)",
                   icono=_ICON_FOLDER)
        REP.say("[Recetas Cheque] acceso directo a la carpeta local (sin subir a la nube)")
    else:
        REP.say(f"[Recetas Cheque] (no existe la carpeta local {RCH_DIR})")
    nota = (
        "RECETAS CHEQUE — datos de pacientes\n"
        "===================================\n\n"
        "El formulario ISP NO se copia a esta carpeta porque contiene datos de\n"
        "pacientes y el Escritorio se sincroniza a la nube de OneDrive.\n\n"
        "Usa el acceso directo «Abrir carpeta Recetas Cheque» para abrir el\n"
        "formulario en su carpeta LOCAL (no sincronizada):\n"
        f"  {RCH_DIR}\n"
    )
    try:
        with open(os.path.join(dst, "LEEME.txt"), "w", encoding="utf-8") as fh:
            fh.write(nota)
    except OSError:
        pass


def sync_reposicion():
    dst = os.path.join(BASE, SUB_REP)
    src = _mas_reciente(os.path.join(WORK_DIR, "Reposicion_DiasHabiles_AA*.xlsx"))
    if not src:
        REP.say("[Reposición] (aún no se ha generado el plan — corre reposicion_dias_habiles.py)")
        return
    _copiar(src, dst, nuevo_nombre="Reposicion_DiasHabiles_AA.xlsx")
    REP.say(f"[Reposición] {os.path.basename(src)} → «{SUB_REP}»")


def sync_pedido():
    dst = os.path.join(BASE, SUB_PEDIDO)
    src = _mas_reciente(os.path.join(WORK_DIR, "Pedido_Fusion_AA*.xlsx"))
    if not src:
        REP.say("[Pedido Fusionado] (aún no generado — corre pedido_fusion.py)")
        return
    _copiar(src, dst, nuevo_nombre="Pedido_Fusion_AA.xlsx")
    REP.say(f"[Pedido Fusionado] {os.path.basename(src)} → «{SUB_PEDIDO}»")


_LEEME_DUPLICADOS = """\
AUDITORIA DE DUPLICADOS — Farmacia AT Abierta
==============================================

Estos DOS accesos NO se ejecutan solos con AUTO_SSASUR (llaman a la API de
Claude y cuestan tokens) — corren a demanda, cuando tú los abres.

  Agente Duplicados IA.lnk
  -------------------------
  Para que sirve: revisa las recetas de HOY (ventana de 90 dias hacia atras)
  y detecta pacientes con el mismo medicamento prescrito por duplicado. Una
  IA (Claude) razona caso a caso y prioriza cuales investigar primero.
  Que ayuda al ejecutarlo: pillar a tiempo un doble retiro reciente (mismo
  paciente retirando el mismo medicamento dos veces) ANTES de que se
  acumule, especialmente util con medicamentos controlados o de alto costo.
  Genera un Excel con el razonamiento de cada caso. Los RUT de pacientes
  NUNCA se envian a la IA (se anonimizan con SHA-256 antes de la llamada).

  Auditoria Duplicados Profunda.lnk
  ----------------------------------
  Para que sirve: audita TODO el historico de recetas (no solo hoy), marca
  si el doble retiro sigue ACTIVO en este momento, desde que fecha empezo
  y cuantos dias lleva acumulado. La IA sugiere una accion por caso:
  URGENTE / REVISAR / INFORMAR / MONITOREAR.
  Que ayuda al ejecutarlo: da la foto completa y priorizada de TODOS los
  casos de duplicidad (activos e historicos), util para una revision
  periodica de fondo o cuando se sospecha de un paciente en particular.
  Al abrirla pide elegir 1-4 (completa con IA / rapida sin IA / rapida con
  IA / salir) — la opcion 1 (completa) es la recomendada.

  Ambas abren una ventana de consola: se demoran unos minutos y quedan
  esperando que presiones una tecla al terminar (revisa el resultado ahi
  antes de cerrar).

  IMPORTANTE — privacidad: el Excel generado (trae nombre completo del
  paciente junto al medicamento) NO se copia a esta carpeta ni a ningun
  otro lugar del Escritorio, porque el Escritorio esta sincronizado a
  OneDrive (nube). Queda SOLO en la carpeta local del programa:
  {work}
  Abrelo desde ahi cuando lo necesites.
"""


def sync_duplicados():
    dst = os.path.join(BASE, SUB_DUP)
    os.makedirs(dst, exist_ok=True)
    n = 0
    if _crear_lnk(os.path.join(dst, "Agente Duplicados IA.lnk"),
                  os.path.join(WORK_DIR, "AGENTE_DUPLICADOS.bat"),
                  "Revisa las recetas de hoy y detecta duplicados con IA (Claude)",
                  icono=_ICON_LUPA):
        n += 1
    if _crear_lnk(os.path.join(dst, "Auditoria Duplicados Profunda.lnk"),
                  os.path.join(WORK_DIR, "AUDITAR_DUPLICADOS_PROFUNDO.bat"),
                  "Audita todo el historico de duplicados, vigencia y accion sugerida",
                  icono=_ICON_LUPA):
        n += 1
    try:
        with open(os.path.join(dst, "LEEME.txt"), "w", encoding="utf-8") as fh:
            fh.write(_LEEME_DUPLICADOS.format(work=WORK_DIR))
    except OSError:
        pass
    REP.say(f"[Auditoria Duplicados] {n} acceso(s) directo(s) → «{SUB_DUP}»")


def sync_centinela():
    dst = os.path.join(BASE, SUB_CENTINELA)
    src = os.path.join(WORK_DIR, "Centinela_Reportes")
    if not os.path.isdir(src):
        REP.say("[Centinela] (aún no se ha generado ningún reporte)")
        return
    n = _espejo(src, dst)
    semanas = sorted(d for d in os.listdir(src) if os.path.isdir(os.path.join(src, d)))
    REP.say(f"[Centinela] {n} archivo(s) · semanas: {', '.join(semanas) if semanas else '(ninguna)'} → «{SUB_CENTINELA}»")


def sync_auditoria():
    dst = os.path.join(BASE, SUB_AUDIT)
    src = os.path.join(WORK_DIR, "auditoria_prescripcion.json")
    destino_xlsx = os.path.join(dst, "Auditoria_Prescripcion_Resumen.xlsx")
    if not os.path.isfile(src):
        REP.say("[Auditoría] (aún no se ha generado el JSON)")
        return
    # Preferimos el Excel legible: quita el JSON crudo que dejaba una versión
    # anterior (si el Excel falla, el fallback de abajo lo vuelve a copiar).
    viejo_json = os.path.join(dst, "auditoria_prescripcion.json")
    if os.path.isfile(viejo_json):
        try:
            os.remove(viejo_json)
        except OSError:
            pass
    # Incremental: si el Excel ya está al día respecto al JSON, no regenerar.
    if os.path.isfile(destino_xlsx) and os.path.getmtime(destino_xlsx) >= os.path.getmtime(src):
        REP.skip += 1
        REP.say("[Auditoría] Excel ya está al día (sin cambios)")
        return
    try:
        with open(src, encoding="utf-8") as f:
            payload = json.load(f)
        if _excel_auditoria(payload, destino_xlsx):
            REP.say(f"[Auditoría] Excel legible generado "
                    f"({payload.get('n_medicamentos', '?')} medicamentos)")
            return
    except Exception as e:
        REP.say(f"  [aviso] no se pudo generar el Excel de auditoría: {e}")
    _copiar(src, dst)  # respaldo: el JSON crudo
    REP.say("[Auditoría] copiado el JSON (no se pudo generar el Excel)")


def _excel_auditoria(payload, ruta):
    """Convierte el JSON de auditoría en un Excel de una hoja, legible y filtrable."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    data = payload.get("data", {})
    if not data:
        return False
    filas = []
    for med, d in data.items():
        pres = d.get("prescriptores") or []
        diag = d.get("diagnosticos") or []
        p1 = pres[0] if pres else {}
        d1 = diag[0] if diag else {}
        filas.append([
            med,
            d.get("cmp_dispensado", 0),
            d.get("total_prescrito", 0),
            d.get("total_dispensado", 0),
            d.get("pct_dispensado", 0),
            d.get("pacientes", 0),
            d.get("dup_2mas_medicos", 0),
            d.get("dup_2mas_recetas_mes", 0),
            d.get("sin_diagnostico", 0),
            (f"{p1.get('medico','')} · {p1.get('esp','')} · {p1.get('pct','')}%" if p1 else ""),
            (d1.get("dx", "") if d1 else ""),
        ])
    # Más dispensado primero.
    filas.sort(key=lambda r: r[3], reverse=True)

    cols = ["Medicamento", "CMP dispensado (mes)", "Total prescrito", "Total dispensado",
            "% dispensado", "Pacientes", "Pac. >=2 medicos", "Pac. >=2 recetas/mes",
            "Sin diagnostico", "Top prescriptor", "Top diagnostico"]
    anchos = [40, 16, 14, 15, 12, 11, 15, 18, 14, 38, 38]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(cols)
    for fila in filas:
        ws.append(fila)

    hdr = PatternFill("solid", fgColor="0F766E")
    for c, ancho in enumerate(anchos, 1):
        celda = ws.cell(1, c)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = hdr
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[ws.cell(1, c).column_letter].width = ancho
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(ruta)
    return True


# ── Estructura, LEEME y accesos directos ─────────────────────────────────────
LEEME = """\
========================================================================
  FARMACIA AT ABIERTA — Resultados (Hospital de Pitrufquén)
========================================================================

Esta carpeta reúne, en el Escritorio, las SALIDAS de todos los procesos
para revisarlas sin abrir la carpeta del programa. Se actualizan solas
cada vez que corres cada proceso.

  Abrir App de Pedidos        ← abre el tablero en el navegador
  AUTO_SSASUR                 ← descarga de SSASUR + recalcula + publica (todo)
  Gestion Territorial         ← solo descarga y genera planillas GT
  Recetas Cheque ISP          ← solo actualiza el registro ISP del mes

  1 - App Pedidos          Consolidado_AA_MAESTRO.xlsx y Resumen_Pedidos_AA.xlsx
  2 - Gestion Territorial  Lo del ÚLTIMO rango queda al frente; lo anterior, en Historial\\
  3 - Recetas Cheque       Acceso directo a la carpeta LOCAL (no sube datos de pacientes a la nube)
  4 - Auditoria Prescripcion  Auditoria_Prescripcion_Resumen.xlsx (ordena/filtra en Excel)
  5 - Reposicion           Reposicion_DiasHabiles_AA.xlsx (plan días hábiles con feriados)
  6 - Pedido Fusionado     Pedido_Fusion_AA.xlsx (Farm_Bod + Bod_Farmacos + Dialisis)
  7 - Centinela             Reportes semanales (json + pdf) por semana epidemiológica
  8 - Auditoria Duplicados Accesos a Agente Duplicados IA y Auditoria Profunda (a demanda,
                            con su propio LEEME.txt explicando cada uno). El Excel con
                            nombres de pacientes NO se copia aquí (Escritorio = OneDrive).

------------------------------------------------------------------------
Nota: estas son COPIAS para consulta. El programa original sigue en
  {work}
No edites los archivos aquí esperando que cambien los cálculos; los
datos oficiales se generan y publican desde esa carpeta.
========================================================================
"""


def _ps_quote(s):
    """Comilla simple de PowerShell (duplica las comillas simples internas)."""
    return "'" + str(s).replace("'", "''") + "'"


def _crear_lnk(ruta_lnk, target, descripcion, icono=None):
    """Crea un acceso directo .lnk (a archivo o carpeta) con WScript.Shell."""
    if not os.path.exists(target):
        REP.say(f"  [aviso] no existe el destino del acceso: {target}")
        return False
    workdir = target if os.path.isdir(target) else os.path.dirname(target)
    ps = (
        "$ws = New-Object -ComObject WScript.Shell; "
        f"$s = $ws.CreateShortcut({_ps_quote(ruta_lnk)}); "
        f"$s.TargetPath = {_ps_quote(target)}; "
        f"$s.WorkingDirectory = {_ps_quote(workdir)}; "
        f"$s.Description = {_ps_quote(descripcion)}; "
    )
    if icono:
        ps += f"$s.IconLocation = {_ps_quote(icono)}; "
    ps += "$s.WindowStyle = 1; $s.Save()"
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
            capture_output=True, text=True,
        )
        if r.returncode != 0:
            REP.say(f"  [aviso] no se pudo crear {os.path.basename(ruta_lnk)}: {r.stderr.strip()}")
            return False
        return True
    except OSError as e:
        REP.say(f"  [aviso] PowerShell no disponible para accesos: {e}")
        return False


_ACCESOS = [
    ("Abrir App de Pedidos.lnk",        "ABRIR_APP.bat",      "Abre el tablero de Pedidos AA en el navegador", _ICON_APP),
    ("AUTO_SSASUR.lnk",                 "AUTO_SSASUR.bat",    "Descarga de SSASUR, recalcula todo y publica",  _ICON_REFRESH),
    ("Gestion Territorial.lnk",         "GT.bat",             "Descarga y genera las planillas de GT",         _ICON_DOC),
    ("Recetas Cheque ISP.lnk",          "RECETAS_CHEQUE.bat", "Actualiza el registro ISP del mes",             _ICON_RUN),
]


def crear_estructura(forzar_lnk=False):
    """Crea carpetas, LEEME y (si faltan o forzar_lnk) los accesos directos."""
    for sub in (SUB_APP, SUB_GT, SUB_RCH, SUB_AUDIT, SUB_REP, SUB_PEDIDO, SUB_CENTINELA, SUB_DUP):
        os.makedirs(os.path.join(BASE, sub), exist_ok=True)
    try:
        with open(os.path.join(BASE, "LEEME.txt"), "w", encoding="utf-8") as fh:
            fh.write(LEEME.format(work=WORK_DIR))
    except OSError as e:
        REP.say(f"  [aviso] no se pudo escribir LEEME.txt: {e}")

    creados = 0
    for nombre_lnk, bat, desc, icono in _ACCESOS:
        ruta = os.path.join(BASE, nombre_lnk)
        if forzar_lnk or not os.path.exists(ruta):
            if _crear_lnk(ruta, os.path.join(WORK_DIR, bat), desc, icono=icono):
                creados += 1
    if creados:
        REP.say(f"[Accesos] {creados} acceso(s) directo(s) actualizados.")


def escribir_log():
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    except Exception:
        ts = "(sin fecha)"
    cuerpo = (
        f"Última sincronización: {ts}\n"
        f"Copiados: {REP.ok} · Sin cambios: {REP.skip} · Con problemas: {REP.fail}\n"
        + "-" * 56 + "\n" + "\n".join(REP.log) + "\n"
    )
    try:
        with open(os.path.join(BASE, "_ultima_sync.txt"), "w", encoding="utf-8") as fh:
            fh.write(cuerpo)
    except OSError:
        pass


# ── CLI ──────────────────────────────────────────────────────────────────────
def main():
    args = set(a.lower() for a in sys.argv[1:])
    print("=" * 60)
    print(f"  Publicando resultados en: {BASE}")
    print("=" * 60)

    crear_estructura(forzar_lnk="--enlaces" in args)
    if args == {"--enlaces"}:
        escribir_log()
        print("\nListo: carpetas y accesos directos actualizados.")
        return

    selectivo = args & {"--app", "--gt", "--rch", "--auditoria", "--reposicion",
                         "--pedido", "--centinela", "--duplicados"}
    todo = not selectivo

    if todo or "--app" in args:
        sync_app()
    if todo or "--gt" in args:
        sync_gt()
    if todo or "--rch" in args:
        sync_rch()
    if todo or "--auditoria" in args:
        sync_auditoria()
    if todo or "--reposicion" in args:
        sync_reposicion()
    if todo or "--pedido" in args:
        sync_pedido()
    if todo or "--centinela" in args:
        sync_centinela()
    if todo or "--duplicados" in args:
        sync_duplicados()

    escribir_log()
    print(f"\nListo ({REP.ok} copiados, {REP.skip} sin cambios). "
          f"Abre «{NOMBRE_CARPETA}» en tu Escritorio.")


if __name__ == "__main__":
    main()
