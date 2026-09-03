#!/usr/bin/env python3
"""
crear_acta_vencimiento.py — Genera automáticamente una hoja "ACTA SALIDA" nueva
en la planilla oficial ACTAS DE VENCIMIENTO 2026.xlsx, a partir del PDF
"Datos de Ajuste por AJUSTE DE SALIDA" (motivo VENCIMIENTO) que exporta SSASUR.

Qué hace:
  1. Lee el PDF del ajuste: bodega, fecha, folio SSASUR, quién solicita, y el
     detalle de medicamentos/lotes/cantidades dados de baja.
  2. Para cada medicamento busca el Precio Unitario REAL (costo promedio
     ponderado por Código+Bodega) en el reporte_de_stock de SSASUR más cercano
     a la fecha del ajuste — el PDF solo trae un precio redondeado a pesos
     enteros que NO es el que se usa en el acta.
  3. Clona (vía Excel/COM, preservando logos y formato) la hoja más reciente
     del mismo tipo de bodega en ACTAS DE VENCIMIENTO 2026.xlsx, la renombra
     con el folio siguiente y la llena con los datos del ajuste.
  4. Antes de tocar el archivo real, guarda un respaldo con fecha/hora.

Uso:
  py crear_acta_vencimiento.py "ruta\\al\\ajuste.pdf"
  py crear_acta_vencimiento.py "ajuste.pdf" --stock "reporte_de_stock_XXXX.xlsx"
  py crear_acta_vencimiento.py "ajuste.pdf" --periodo "correspondiente al mes de julio 2026"
  py crear_acta_vencimiento.py "ajuste.pdf" --dry-run
"""
import argparse
import glob
import os
import re
import shutil
import sys
from datetime import datetime

import openpyxl
import pdfplumber

from utils_aa import ACTAS_VENCIMIENTO_PATH, norm_erp, setup_stdout

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))

MESES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre",
    12: "diciembre",
}

# (keywords que deben estar TODAS en la Bodega del ajuste, tipo_code, texto
#  narrativo, cadencia). Se evalúa en orden — el primer match gana.
BODEGA_TIPOS = [
    (["URGENCIA"],              "URGENCIA",       "farmacia urgencia",              "mensual", "correspondiente al mes"),
    (["FARMACIA", "ABIERTA"],   "F.ABIERTA",      "farmacia atención abierta",      "mensual", "del mes"),
    (["FARMACIA", "CERRADA"],   "F.CERRADA",      "farmacia de atención cerrada",   "trimestral", None),
    (["CUARENTENA"],            "B.A.CUARENTENA", "bodega de cuarentena",           "adhoc", None),
    (["BODEGA", "ABIERTA"],     "B.A.ABIERTA",    "bodega de atención abierta",     "mensual", "correspondiente al mes"),
    (["BODEGA", "CERRADA"],     "B.A.CERRADA",    "bodega de atención cerrada",     "trimestral", None),
]

# El QF firmante de la Farmacia AT Abierta es siempre Daniel Castro, sin
# importar quién figure como "Solicita" en el ajuste de SSASUR (a veces lo
# tramita otra persona pero el acta la firma él igual). Pedido explícito del
# usuario 14-08-2026 — no usar ajuste["solicita"] para el tipo F.ABIERTA.
QF_FIJO_POR_TIPO = {
    "F.ABIERTA": "Daniel Castro Cortes",
}

UNIDAD_ABBR = {
    "COMPRIMIDO": "CM", "TABLETA": "CM", "CAPSULA": "CP", "AMPOLLA": "AMP",
    "FRASCO AMPOLLA": "FA", "FRASCO": "FC", "OVULO": "OV", "SOBRE": "SOB",
    "UNGUENTO OFTALMICO": "UNG OFT", "UNGUENTO": "UNG", "GEL": "GEL",
    "JERINGA PRELLENADA": "JRP", "MATRAZ": "MTZ", "POTE": "PT", "BOLSA": "BOL",
    "UNIDAD": "UD", "PARCHE": "PCH", "SUPOSITORIO": "SUP",
}


def unidad_abrev(texto_unidad):
    if not texto_unidad:
        return "UD"
    # SSASUR a veces trae unidades compuestas ("CAPSULA/COMPRIMIDO") — la
    # unidad de dispensación real es siempre la primera del listado.
    primera = str(texto_unidad).split("/")[0]
    t = norm_erp(primera)
    if t in UNIDAD_ABBR:
        return UNIDAD_ABBR[t]
    for palabra, abr in UNIDAD_ABBR.items():
        if palabra in t:
            return abr
    return t[:4]


# ── 1. Parseo del PDF de ajuste ───────────────────────────────────────────────
def parse_ajuste_pdf(path):
    with pdfplumber.open(path) as pdf:
        page = pdf.pages[0]
        texto = page.extract_text() or ""
        tablas = page.extract_tables()

    def grab(patron, flags=re.IGNORECASE):
        m = re.search(patron, texto, flags)
        return m.group(1).strip() if m else None

    folio_ajuste = grab(r"FOLIO:\s*([\w\-]+)")
    fecha_doc_str = grab(r"Fecha Documento:\s*(\d{2}/\d{2}/\d{4})")
    motivo = grab(r"Motivo Ajuste:\s*([^\n]+)")
    bodega_raw = grab(r"Bodega:\s*([^\n]+)")
    solicita = grab(r"Solicita:\s*([^\n]+?)\s*Autoriza")
    autoriza = grab(r"Autoriza:?\s*([^\n]+)")
    total_str = grab(r"Total Ajuste:\s*\$?\s*(-?[\d.,]+)")
    observacion = grab(r"Observaci[oó]n:\s*\n?(.*?)\n(?:Detalle de productos|$)", flags=re.IGNORECASE | re.DOTALL)

    if not fecha_doc_str:
        raise ValueError("No encontré 'Fecha Documento' en el PDF — ¿es un ajuste de SSASUR válido?")
    fecha_doc = datetime.strptime(fecha_doc_str, "%d/%m/%Y").date()

    # El mes/año "real" del vencimiento va en la Observación ("VENCIMIENTOS JUNIO 2026...")
    # y casi siempre es distinto al mes en que se tramita el ajuste (Fecha Documento).
    periodo_mes, periodo_anio = None, None
    if observacion:
        obs_norm = norm_erp(observacion)
        anio_m = re.search(r"(20\d{2})", obs_norm)
        periodo_anio = int(anio_m.group(1)) if anio_m else None
        for num, nombre in MESES.items():
            if norm_erp(nombre) in obs_norm:
                periodo_mes = num
                break

    if motivo and "VENCIMIENTO" not in motivo.upper():
        print(f"  [AVISO] Motivo Ajuste = '{motivo}' (no dice VENCIMIENTO) — revisa que sea el documento correcto.")

    total_declarado = None
    if total_str:
        limpio = total_str.replace(".", "").replace(",", ".")
        try:
            total_declarado = abs(float(limpio))
        except ValueError:
            total_declarado = None

    # localizar la tabla de detalle (encabezado con 'Código' y 'Producto')
    detail = None
    for t in tablas:
        if not t or not t[0]:
            continue
        hdr = [norm_erp(c) for c in t[0] if c]
        if any("CODIGO" in h for h in hdr) and any("PRODUCTO" in h for h in hdr):
            detail = t
            break
    if detail is None:
        raise ValueError("No encontré la tabla 'Detalle de productos' en el PDF.")

    hdr_norm = [norm_erp(c).replace("\n", " ") if c else "" for c in detail[0]]

    def col_idx(*claves):
        for i, h in enumerate(hdr_norm):
            if all(k in h for k in claves):
                return i
        return None

    i_cod = col_idx("CODIGO")
    i_prod = col_idx("PRODUCTO")
    i_lote = col_idx("LOTE")
    i_venc = col_idx("FECHA")
    i_cant = col_idx("CANTIDAD")

    items = []
    for row in detail[1:]:
        if not row or not row[i_cod]:
            continue
        if str(row[i_cod]).strip().lower().startswith("total"):
            continue
        codigo = re.sub(r"\s+", "", str(row[i_cod]))
        producto = re.sub(r"\s+", " ", str(row[i_prod])).strip()
        lote_txt = str(row[i_lote]).strip() if row[i_lote] else ""
        fecha_venc_str = str(row[i_venc]).strip() if row[i_venc] else None
        cant_txt = str(row[i_cant]).strip() if row[i_cant] else "0"
        cant_txt = cant_txt.split()[0] if cant_txt.split() else "0"
        cant_txt = re.sub(r"[^\d\-]", "", cant_txt)
        if not fecha_venc_str or not cant_txt:
            continue
        items.append(dict(
            codigo=codigo,
            producto=producto,
            lote_txt=lote_txt,
            fecha_venc=datetime.strptime(fecha_venc_str, "%d/%m/%Y").date(),
            cantidad=int(cant_txt),
        ))

    if not items:
        raise ValueError("La tabla de detalle no arrojó ningún ítem — revisa el PDF manualmente.")

    return dict(
        folio_ajuste=folio_ajuste, fecha_doc=fecha_doc, motivo=motivo,
        bodega_raw=bodega_raw, solicita=solicita, autoriza=autoriza,
        total_declarado=total_declarado, items=items,
        periodo_mes=periodo_mes, periodo_anio=periodo_anio,
    )


# ── 2. Clasificación de bodega ────────────────────────────────────────────────
def clasificar_bodega(bodega_raw):
    b = norm_erp(bodega_raw or "")
    for keywords, tipo_code, desc, cadencia, conector in BODEGA_TIPOS:
        if all(kw in b for kw in keywords):
            return dict(tipo_code=tipo_code, desc=desc, cadencia=cadencia, conector=conector)
    raise ValueError(f"No pude clasificar la Bodega '{bodega_raw}' en ninguno de los 6 tipos conocidos "
                      f"(F.ABIERTA/F.CERRADA/URGENCIA/B.A.ABIERTA/B.A.CERRADA/B.A.CUARENTENA).")


# ── 3. Precio real: buscar reporte_de_stock más cercano a la fecha del ajuste ─
def buscar_reporte_stock(fecha_doc, override=None):
    if override:
        return override, None
    candidatos = []
    patrones = [
        os.path.join(MAESTRO_DIR, "reporte_de_stock_*.xlsx"),
        os.path.join(MAESTRO_DIR, "Datos_Extraidos", "reporte_de_stock_historico", "reporte_de_stock_*.xlsx"),
    ]
    for pat in patrones:
        for f in glob.glob(pat):
            m = re.search(r"reporte_de_stock_(\d{8})(\d{6})\.xlsx$", os.path.basename(f))
            if not m:
                continue
            fdate = datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
            candidatos.append((fdate, f))
    if not candidatos:
        return None, None

    def score(item):
        fdate, _ = item
        diff = (fdate.date() - fecha_doc).days
        return (0, diff) if diff >= 0 else (1, -diff)

    candidatos.sort(key=score)
    mejor_fecha, mejor_archivo = candidatos[0]
    return mejor_archivo, mejor_fecha


def cargar_precios(stock_path):
    wb = openpyxl.load_workbook(stock_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    precios = {}
    hdr_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] and norm_erp(row[0]) == "CODIGO":
            hdr_row = i
            break
    if hdr_row is None:
        wb.close()
        raise ValueError(f"'{stock_path}' no tiene el formato esperado de reporte_de_stock (sin fila 'Código').")
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or row[0] is None:
            continue
        row = list(row) + [None] * (9 - len(row))
        codigo, bodega, _desc, _zgen, _cant, precio, _total, unidad, _lote = row[:9]
        if codigo is None or bodega is None:
            continue
        precios[(str(codigo).strip(), str(bodega).strip().upper())] = (precio, unidad)
    wb.close()
    return precios


# ── 4. Narrativa ───────────────────────────────────────────────────────────
def construir_narrativa(bodega_info, fecha_doc, folio_ajuste, periodo_mes=None, periodo_anio=None, override=None):
    # El período (mes/trimestre) es el de la Observación del ajuste ("VENCIMIENTOS JUNIO 2026"),
    # NO el de Fecha Documento (que es cuando se tramitó el ajuste, casi siempre el mes siguiente).
    mes_ref = periodo_mes or fecha_doc.month
    anio_ref = periodo_anio or fecha_doc.year

    if override:
        cuerpo = override
    elif bodega_info["cadencia"] == "trimestral":
        trimestre = (mes_ref - 1) // 3 + 1
        cuerpo = f"vencidos de {bodega_info['desc']} correspondiente al {trimestre}° trimestre {anio_ref}"
    elif bodega_info["cadencia"] == "adhoc":
        cuerpo = f"vencidos de {bodega_info['desc']} correspondiente"
    else:
        mes = MESES[mes_ref]
        cuerpo = f"vencidos de {bodega_info['desc']} {bodega_info['conector']} de {mes} {anio_ref}"

    texto = (
        f"Con fecha {fecha_doc.strftime('%d/%m/%Y')} el Hospital Pitrufquén a través de su área de "
        f"Farmacia procede a dar de baja los siguientes medicamentos. Correspondientes a {cuerpo}."
    )
    if folio_ajuste:
        texto += f" Folio {folio_ajuste}."
    return texto


# ── 5. Escritura en Excel vía COM (preserva logos/formato) ───────────────────
def col_letter(idx):
    return openpyxl.utils.get_column_letter(idx)


def buscar_texto(grid, row0, col0, predicate):
    """grid: tupla de tuplas 1-indexada relativa a (row0,col0). Devuelve (fila,col) absolutos o None."""
    for r, fila in enumerate(grid):
        for c, val in enumerate(fila):
            if val is not None and predicate(val):
                return row0 + r, col0 + c
    return None


def clonar_y_llenar_acta(actas_path, ajuste, bodega_info, items_final, folio_nuevo, narrativa, dry_run=False):
    import win32com.client

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        wb = xl.Workbooks.Open(actas_path)
        nombres_originales = {sh.Name for sh in wb.Sheets}

        # Buscar la hoja plantilla: la más reciente del mismo tipo de bodega
        candidatas = []
        for sh in wb.Sheets:
            m = re.match(r"\s*(\d+)\.?\s*(.+)$", sh.Name)
            if not m:
                continue
            folio_txt, resto = m.group(1), m.group(2)
            resto_norm = norm_erp(resto).replace(" ", "").replace(".", "")
            tipo_norm = bodega_info["tipo_code"].replace(".", "").replace(" ", "")
            if resto_norm == tipo_norm:
                candidatas.append((int(folio_txt), sh))
        if not candidatas:
            raise ValueError(f"No hay ninguna hoja existente del tipo '{bodega_info['tipo_code']}' para usar como plantilla.")
        candidatas.sort(key=lambda x: x[0])
        _, plantilla = candidatas[-1]
        nombre_plantilla = plantilla.Name

        # separador del nombre ("20. F.ABIERTA" vs "21.B.A.ABIERTA")
        m = re.match(r"\s*\d+(\.?\s*)", nombre_plantilla)
        separador = m.group(1) if m else ". "
        nombre_nuevo = f"{folio_nuevo}{separador}{bodega_info['tipo_code']}"

        n_hojas_antes = wb.Sheets.Count
        # OJO: win32com (dynamic dispatch) NO soporta bien args nombrados tipo
        # After=... en Copy() — los ignora en silencio y copia la hoja a un
        # libro nuevo aparte, dejando wb intacto. Hay que pasar posicional
        # (Before, After) y verificar que el conteo de hojas subió en 1.
        plantilla.Copy(None, wb.Sheets(wb.Sheets.Count))
        if wb.Sheets.Count != n_hojas_antes + 1:
            raise RuntimeError(
                f"Copy() no agregó una hoja nueva (antes {n_hojas_antes}, después {wb.Sheets.Count}) — "
                f"abortando para no arriesgar sobrescribir una hoja existente."
            )
        nueva = wb.Sheets(wb.Sheets.Count)
        nueva.Name = nombre_nuevo

        # Leer la grilla usada para ubicar celdas dinámicamente (evita hardcodear filas/columnas)
        used = nueva.UsedRange
        row0, col0 = used.Row, used.Column
        grid = used.Value  # tupla de tuplas

        def es(txt_objetivo):
            t = norm_erp(txt_objetivo)
            return lambda v: isinstance(v, str) and norm_erp(v) == t

        def contiene(sub):
            s = norm_erp(sub)
            return lambda v: isinstance(v, str) and s in norm_erp(v)

        pos_folio_lbl = buscar_texto(grid, row0, col0, contiene("FOLIO N"))
        pos_medic = buscar_texto(grid, row0, col0, es("Medicamento"))
        pos_narrativa = buscar_texto(grid, row0, col0, contiene("Con fecha"))
        pos_qf_lbl = buscar_texto(grid, row0, col0, contiene("Quimico Farmaceutico"))

        if not (pos_folio_lbl and pos_medic and pos_narrativa and pos_qf_lbl):
            raise ValueError(
                f"No pude ubicar todas las celdas clave en la hoja plantilla '{nombre_plantilla}' "
                f"(folio={bool(pos_folio_lbl)}, medicamento={bool(pos_medic)}, "
                f"narrativa={bool(pos_narrativa)}, qf={bool(pos_qf_lbl)}). Reviso manualmente."
            )

        r_folio, c_folio = pos_folio_lbl
        r_hdr, c_med = pos_medic
        r_narr, c_narr = pos_narrativa
        r_qf_lbl, c_qf = pos_qf_lbl

        # columnas de la tabla, leyendo el encabezado de esa fila
        hdr_vals = {c: nueva.Cells(r_hdr, c).Value for c in range(c_med, c_med + 8)}

        def hallar_col(clave):
            for c, v in hdr_vals.items():
                if isinstance(v, str) and clave in norm_erp(v):
                    return c
            raise ValueError(f"No encontré la columna '{clave}' en el encabezado de la tabla.")

        c_unidad = hallar_col("UNIDAD")
        c_cant = hallar_col("CANTIDAD")
        c_venc = hallar_col("VENCIMIENTO")
        c_lote = hallar_col("LOTE")
        c_motivo = hallar_col("MOTIVO")
        c_valor = hallar_col("VALOR")
        c_total = hallar_col("TOTAL")

        # fila TOTAL: primera fila debajo del encabezado, en la columna Medicamento, con texto "TOTAL"
        r = r_hdr + 1
        r_total_row = None
        while r < r_hdr + 200:
            v = nueva.Cells(r, c_med).Value
            if isinstance(v, str) and v.strip().upper() == "TOTAL":
                r_total_row = r
                break
            r += 1
        if r_total_row is None:
            raise ValueError("No encontré la fila TOTAL de la tabla en la hoja plantilla.")

        n_items_actual = r_total_row - (r_hdr + 1)
        n_items_nuevo = len(items_final)

        if n_items_nuevo > n_items_actual:
            fila_formato_origen = r_total_row - 1
            for _ in range(n_items_nuevo - n_items_actual):
                nueva.Rows(fila_formato_origen).Copy()
                nueva.Rows(r_total_row).Insert(Shift=-4121)  # xlShiftDown
            r_total_row = r_hdr + 1 + n_items_nuevo
        elif n_items_nuevo < n_items_actual:
            primera_a_borrar = r_hdr + 1 + n_items_nuevo
            ultima_a_borrar = r_total_row - 1
            nueva.Range(f"{primera_a_borrar}:{ultima_a_borrar}").Delete(Shift=-4162)  # xlShiftUp
            r_total_row = r_hdr + 1 + n_items_nuevo

        # Todo lo que estaba DEBAJO de la tabla original (narrativa, firmas) se
        # movió junto con las filas al insertar/borrar — hay que corregir las
        # posiciones que ubicamos ANTES de tocar filas, o se escribe en la fila vieja.
        delta = n_items_nuevo - n_items_actual
        r_narr += delta
        r_qf_lbl += delta

        # FOLIO N°
        nueva.Cells(r_folio, c_folio + 1).Value = folio_nuevo

        # Ítems
        for i, it in enumerate(items_final):
            fila = r_hdr + 1 + i
            nueva.Cells(fila, c_med).Value = it["producto"]
            nueva.Cells(fila, c_unidad).Value = it["unidad"]
            nueva.Cells(fila, c_cant).Value = it["cantidad"]
            nueva.Cells(fila, c_venc).Value = datetime.combine(it["fecha_venc"], datetime.min.time())
            nueva.Cells(fila, c_lote).Value = it["lote"]
            nueva.Cells(fila, c_motivo).Value = "VENCIMIENTO"
            nueva.Cells(fila, c_valor).Value = it["valor_unitario"]
            nueva.Cells(fila, c_total).Formula = (
                f"={col_letter(c_valor)}{fila}*{col_letter(c_cant)}{fila}"
            )

        nueva.Cells(r_total_row, c_total).Formula = (
            f"=SUM({col_letter(c_total)}{r_hdr + 1}:{col_letter(c_total)}{r_total_row - 1})"
        )

        # Narrativa
        nueva.Cells(r_narr, c_narr).Value = narrativa

        # Nombre del QF firmante (fila justo arriba del label "Químico Farmacéutico").
        # Algunos tipos de bodega tienen un QF fijo (ver QF_FIJO_POR_TIPO) que no
        # depende de quién figure como "Solicita" en el ajuste de SSASUR.
        nombre_qf = QF_FIJO_POR_TIPO.get(bodega_info["tipo_code"]) or (
            ajuste["solicita"].title() if ajuste.get("solicita") else None
        )
        if nombre_qf:
            nueva.Cells(r_qf_lbl - 1, c_qf).Value = nombre_qf

        # Salvaguarda final: ninguna hoja pre-existente debe haber desaparecido/cambiado de nombre.
        nombres_finales = {sh.Name for sh in wb.Sheets}
        faltantes = nombres_originales - nombres_finales
        if faltantes:
            raise RuntimeError(
                f"Abortando sin guardar: desaparecieron hojas existentes ({faltantes}) — "
                f"esto no debería pasar nunca, revisa manualmente antes de reintentar."
            )

        if dry_run:
            wb.Close(False)
        else:
            wb.Save()
            wb.Close(True)
        return nombre_nuevo, nombre_plantilla
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        xl.Quit()


# ── main ───────────────────────────────────────────────────────────────────
def main():
    setup_stdout()
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("ajuste_pdf", help="PDF 'Datos de Ajuste por AJUSTE DE SALIDA' descargado de SSASUR")
    ap.add_argument("--actas", default=ACTAS_VENCIMIENTO_PATH, help="Ruta a ACTAS DE VENCIMIENTO 2026.xlsx")
    ap.add_argument("--stock", default=None, help="Forzar un reporte_de_stock_*.xlsx específico para el precio")
    ap.add_argument("--periodo", default=None, help="Texto de período a usar en la narrativa (sobrescribe el heurístico)")
    ap.add_argument("--dry-run", action="store_true", help="No guarda cambios; solo muestra qué haría")
    args = ap.parse_args()

    if not os.path.exists(args.ajuste_pdf):
        sys.exit(f"[ERROR] No existe el PDF: {args.ajuste_pdf}")
    if not os.path.exists(args.actas):
        sys.exit(f"[ERROR] No existe la planilla de actas: {args.actas}\n"
                  f"        (configúrala con --actas o la variable MAESTRO_ACTAS_VENCIMIENTO)")
    lock = os.path.join(os.path.dirname(args.actas), "~$" + os.path.basename(args.actas))
    if os.path.exists(lock):
        sys.exit(f"[ERROR] '{os.path.basename(args.actas)}' está abierta en Excel ahora mismo. "
                  f"Ciérrala y vuelve a correr el script.")

    print(f"Leyendo ajuste: {args.ajuste_pdf}")
    ajuste = parse_ajuste_pdf(args.ajuste_pdf)
    bodega_info = clasificar_bodega(ajuste["bodega_raw"])
    print(f"  Bodega: {ajuste['bodega_raw']}  ->  {bodega_info['tipo_code']}")
    print(f"  Fecha documento: {ajuste['fecha_doc'].strftime('%d/%m/%Y')}   Folio SSASUR: {ajuste['folio_ajuste']}")
    print(f"  Solicita: {ajuste['solicita']}    {len(ajuste['items'])} ítem(s)")

    stock_path, stock_fecha = buscar_reporte_stock(ajuste["fecha_doc"], override=args.stock)
    if not stock_path:
        sys.exit("[ERROR] No encontré ningún reporte_de_stock_*.xlsx (ni en el repo ni en Datos_Extraidos\\reporte_de_stock_historico)."
                  " Descarga uno de SSASUR cercano a la fecha del ajuste, o pásalo con --stock.")
    gap_dias = abs((stock_fecha.date() - ajuste["fecha_doc"]).days) if stock_fecha else None
    print(f"  Reporte de stock usado para precios: {os.path.basename(stock_path)}"
          + (f" ({gap_dias} día(s) de diferencia con el ajuste)" if gap_dias is not None else ""))
    if gap_dias is not None and gap_dias > 3:
        print(f"  [AVISO] El reporte de stock está a {gap_dias} días del ajuste — el precio puede no calzar "
              f"exactamente con el total de la guía. Si tienes uno más cercano, pásalo con --stock.")

    precios = cargar_precios(stock_path)

    items_final = []
    faltantes = []
    for it in ajuste["items"]:
        key = (it["codigo"], ajuste["bodega_raw"].strip().upper())
        precio, unidad = precios.get(key, (None, None))
        if precio is None:
            # fallback: mismo código en cualquier bodega
            alt = [(k, v) for k, v in precios.items() if k[0] == it["codigo"]]
            if alt:
                (_, bodega_alt), (precio, unidad) = alt[0]
                print(f"  [AVISO] '{it['producto']}': sin precio en '{ajuste['bodega_raw']}', "
                      f"usando el de '{bodega_alt}' (${precio:.2f}) — revisa antes de firmar.")
        if precio is None:
            faltantes.append(it["producto"])
            valor_unitario = 0
        else:
            valor_unitario = round(float(precio), 2)

        lote_txt = it["lote_txt"]
        lote = int(lote_txt) if lote_txt.isdigit() else lote_txt

        items_final.append(dict(
            producto=it["producto"],
            unidad=unidad_abrev(unidad),
            cantidad=it["cantidad"],
            fecha_venc=it["fecha_venc"],
            lote=lote,
            valor_unitario=valor_unitario,
        ))

    if faltantes:
        sys.exit(f"[ERROR] No encontré precio para: {', '.join(faltantes)} en '{stock_path}'. "
                  f"No genero el acta con precios en $0 — busca un reporte de stock donde sí aparezcan "
                  f"(o pásalo con --stock) y vuelve a intentar.")

    total_calculado = sum(round(it["valor_unitario"] * it["cantidad"], 2) for it in items_final)
    print(f"  Total calculado: ${total_calculado:,.0f}".replace(",", "."))
    if ajuste["total_declarado"] is not None:
        diff = abs(round(total_calculado) - round(ajuste["total_declarado"]))
        if diff > 1:
            print(f"  [AVISO] El total calculado (${total_calculado:,.0f}) no calza con el 'Total Ajuste' del PDF "
                  f"(${ajuste['total_declarado']:,.0f}), diferencia ${diff:,.0f}. "
                  f"Revisa el reporte de stock usado antes de firmar el acta.".replace(",", "."))
        else:
            print(f"  Calza con el Total Ajuste del PDF (${ajuste['total_declarado']:,.0f}).".replace(",", "."))

    wb_actual = openpyxl.load_workbook(args.actas, read_only=True)
    folio_nuevo = len(wb_actual.sheetnames) + 1
    wb_actual.close()

    narrativa = construir_narrativa(
        bodega_info, ajuste["fecha_doc"], ajuste["folio_ajuste"],
        periodo_mes=ajuste.get("periodo_mes"), periodo_anio=ajuste.get("periodo_anio"),
        override=args.periodo,
    )
    print(f"  Narrativa: {narrativa}")
    print(f"  Folio de acta nuevo: {folio_nuevo}")

    if not args.dry_run:
        bak = args.actas + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(args.actas, bak)
        print(f"  Respaldo guardado en: {bak}")

    nombre_nuevo, nombre_plantilla = clonar_y_llenar_acta(
        args.actas, ajuste, bodega_info, items_final, folio_nuevo, narrativa, dry_run=args.dry_run
    )

    if args.dry_run:
        print(f"\n[DRY-RUN] Se habría creado la hoja '{nombre_nuevo}' (clonando '{nombre_plantilla}'). No se guardó nada.")
    else:
        print(f"\nHoja '{nombre_nuevo}' creada en {args.actas} (clonada de '{nombre_plantilla}').")
        print("Revisa el nombre del QF firmante y los montos antes de imprimir/firmar.")


if __name__ == "__main__":
    main()
