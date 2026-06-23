#!/usr/bin/env python3
"""
agente_gt_pendientes.py — Clasifica PENDIENTES de gt_enriquecido.json
con Claude Haiku, asignando prioridad: URGENTE / RUTINARIO / DIFERIBLE.

Los RUTs NUNCA se envían a la API (Ley 19.628). La API solo recibe el
número de receta (interno), establecimiento destino y nombres de medicamentos.

Uso:
    py agente_gt_pendientes.py
    py agente_gt_pendientes.py out_gt/gt_enriquecido.json
    py agente_gt_pendientes.py out_gt/gt_enriquecido.json --salida out_gt/
"""
import argparse, json, os, sys
from collections import Counter

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils_aa import setup_stdout

setup_stdout()
MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
MODELO = "claude-haiku-4-5-20251001"
BATCH  = 25   # recetas por llamada API

# ── System prompt (cacheado con ephemeral) ─────────────────────────────────────
SYSTEM_PROMPT = """\
Eres un asistente de farmacia del Hospital de Pitrufquén (SSASur, Chile).
Clasificas medicamentos PENDIENTES de despacho en Gestión Territorial,
asignando prioridad de seguimiento según el tipo de medicamento.

URGENTE — No puede esperar al siguiente ciclo GT:
  Insulinas (cualquier tipo), anticoagulantes (acenocumarol, warfarina, heparina,
  enoxaparina), anticonvulsivantes (fenitoína, valproato, carbamazepina,
  levetiracetam, lamotrigina, clonazepam en epilepsia), antibióticos en curso activo,
  corticoides sistémicos, inmunosupresores, antirretrovirales, oncológicos,
  opioides (morfina, tramadol, oxicodona, tapentadol), digoxina, levotiroxina,
  furosemida alta dosis, litio, metotrexato, ciclosporina, tacrolimus.

RUTINARIO — Puede esperar al próximo ciclo (≈ 1 semana):
  Antihipertensivos (enalapril, losartan, amlodipino, hidroclorotiazida,
  bisoprolol, valsartan), hipoglicemiantes orales (metformina, glibenclamida,
  sitagliptina, empagliflozina, dapagliflozina), estatinas (atorvastatina,
  rosuvastatina), omeprazol, pantoprazol, AAS ≤100 mg, ácido fólico, calcio
  oral, hierro oral, antidepresivos estables (sertralina, fluoxetina,
  venlafaxina), ansiolíticos crónicos, antiepilépticos estables con stock previo.

DIFERIBLE — Baja urgencia o no esencial:
  Vitaminas (A, C, D, E, B12), suplementos (magnesio, zinc, omega-3),
  antihistamínicos no urgentes (loratadina tópica), lubricantes oculares,
  laxantes suaves, emolientes, antiflatulentos, protectores solares.

Responde ÚNICAMENTE usando la herramienta clasificar_pendientes.
No incluyas RUTs, nombres de pacientes ni datos identificatorios.
"""

TOOL = {
    "name": "clasificar_pendientes",
    "description": "Clasifica recetas con medicamentos pendientes por urgencia de seguimiento.",
    "input_schema": {
        "type": "object",
        "properties": {
            "clasificaciones": {
                "type": "array",
                "description": "Una entrada por cada receta enviada en el batch",
                "items": {
                    "type": "object",
                    "properties": {
                        "id_receta": {
                            "type": "string",
                            "description": "Número de receta tal como fue enviado",
                        },
                        "prioridad": {
                            "type": "string",
                            "enum": ["URGENTE", "RUTINARIO", "DIFERIBLE"],
                        },
                        "med_critico": {
                            "type": "string",
                            "description": "Nombre del medicamento más crítico (vacío si DIFERIBLE)",
                        },
                        "razon": {
                            "type": "string",
                            "description": "Razón breve, máximo 12 palabras",
                        },
                    },
                    "required": ["id_receta", "prioridad", "razon"],
                },
            }
        },
        "required": ["clasificaciones"],
    },
}


def clasificar_batch(client, batch):
    """Envía un batch de recetas a Haiku y devuelve (clasificaciones, usage)."""
    texto = "\n".join(
        f"receta={r['receta']} | destino={r['destino']} | pendiente={r['pendiente']}"
        for r in batch
    )
    resp = client.messages.create(
        model=MODELO,
        max_tokens=2048,
        system=[{"type": "text", "text": SYSTEM_PROMPT,
                 "cache_control": {"type": "ephemeral"}}],
        tools=[TOOL],
        tool_choice={"type": "tool", "name": "clasificar_pendientes"},
        messages=[{"role": "user", "content": (
            f"Clasifica estos {len(batch)} registros con medicamentos pendientes:\n\n{texto}"
        )}],
    )
    for blk in resp.content:
        if blk.type == "tool_use" and blk.name == "clasificar_pendientes":
            return blk.input.get("clasificaciones", []), resp.usage
    return [], resp.usage


def escribir_excel(resultados, ruta):
    ORDER  = {"URGENTE": 0, "RUTINARIO": 1, "DIFERIBLE": 2}
    FILLS  = {"URGENTE": "FFD7D7", "RUTINARIO": "FFF2CC", "DIFERIBLE": "E2EFDA"}
    FTEXTS = {"URGENTE": "C00000", "RUTINARIO": "C55A11", "DIFERIBLE": "375623"}

    thin = Side(style="thin", color="BFBFBF")
    bd   = Border(thin, thin, thin, thin)
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Pendientes GT"

    cols   = ["Prioridad", "Destino", "Nº Receta", "Paciente",
              "Medicamentos pendientes", "Med. crítico", "Razón", "Especialidad"]
    widths = [12, 26, 11, 28, 40, 24, 36, 20]

    for c, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c, h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.border    = bd
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    sorted_res = sorted(resultados, key=lambda x: (ORDER.get(x.get("prioridad", ""), 3),
                                                    x.get("destino", ""),
                                                    x.get("paciente", "")))
    for i, r in enumerate(sorted_res, 2):
        pri  = r.get("prioridad", "")
        vals = [
            pri,
            r.get("destino", ""),
            r.get("receta", ""),
            r.get("paciente", ""),
            r.get("pendiente", ""),
            r.get("med_critico", ""),
            r.get("razon", ""),
            r.get("especialidad", ""),
        ]
        fg = FILLS.get(pri, "FFFFFF")
        tc = FTEXTS.get(pri, "000000")
        for c, v in enumerate(vals, 1):
            cell           = ws.cell(i, c, v)
            cell.border    = bd
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.fill      = PatternFill("solid", fgColor=fg)
            cell.font      = Font(size=9, bold=(c == 1), color=(tc if c == 1 else "000000"))
        ws.row_dimensions[i].height = 42

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(sorted_res) + 1}"
    wb.save(ruta)
    return len(sorted_res)


def main():
    ap = argparse.ArgumentParser(
        description="Clasifica PENDIENTES de gt_enriquecido.json con Claude Haiku."
    )
    ap.add_argument(
        "json_path",
        nargs="?",
        default=os.path.join(MAESTRO_DIR, "out_gt", "gt_enriquecido.json"),
        help="Ruta al gt_enriquecido.json (por defecto: out_gt/gt_enriquecido.json)",
    )
    ap.add_argument(
        "--salida",
        default=None,
        help="Directorio de salida (por defecto: mismo dir que el JSON)",
    )
    a = ap.parse_args()

    if not os.path.exists(a.json_path):
        print(f"[ERROR] No encontré: {a.json_path}")
        print("  Ejecuta primero: py cruce_gt.py <reporteGT.xlsx>  o  GT.bat")
        sys.exit(1)

    salida_dir = a.salida or os.path.dirname(os.path.abspath(a.json_path))
    os.makedirs(salida_dir, exist_ok=True)

    with open(a.json_path, encoding="utf-8") as f:
        data = json.load(f)

    registros  = data.get("registros", [])
    pendientes = [r for r in registros if r.get("pendiente", "").strip()]

    print(f"gt_enriquecido.json : {len(registros)} registros totales")
    print(f"Con pendientes      : {len(pendientes)}")

    if not pendientes:
        print("  Sin pendientes — nada que clasificar.")
        sys.exit(0)

    # Batches — NO se envía run/nombre del paciente (Ley 19.628)
    batches = [pendientes[i:i + BATCH] for i in range(0, len(pendientes), BATCH)]
    client  = anthropic.Anthropic()
    result_map: dict[str, dict] = {}

    total_in = total_out = total_cache_r = total_cache_w = 0

    for bi, batch in enumerate(batches, 1):
        payload = [
            {"receta": r["receta"], "destino": r["estab_destino"], "pendiente": r["pendiente"]}
            for r in batch
        ]
        print(f"  Batch {bi}/{len(batches)}: {len(payload)} recetas...", end=" ", flush=True)
        clases, usage = clasificar_batch(client, payload)
        print(f"ok ({len(clases)} clasificadas)")

        total_in      += usage.input_tokens
        total_out     += usage.output_tokens
        total_cache_r += getattr(usage, "cache_read_input_tokens",  0) or 0
        total_cache_w += getattr(usage, "cache_creation_input_tokens", 0) or 0

        for c in clases:
            result_map[str(c["id_receta"])] = c

    # Fusiona clasificaciones con datos locales (paciente/especialidad no viajan a la API)
    resultados = []
    sin_match  = 0
    for r in pendientes:
        cl = result_map.get(str(r["receta"]), {})
        if not cl:
            sin_match += 1
        resultados.append({
            **cl,
            "receta":      r["receta"],
            "destino":     r["estab_destino"],
            "paciente":    r["paciente"],       # solo local
            "pendiente":   r["pendiente"],
            "especialidad": r.get("especialidad", ""),
        })

    if sin_match:
        print(f"  [aviso] {sin_match} recetas sin clasificación — quedan sin prioridad")

    # Excel
    xlsx_path = os.path.join(salida_dir, "Pendientes_GT_Prioridad.xlsx")
    n = escribir_excel(resultados, xlsx_path)
    print(f"\n  -> {xlsx_path}  ({n} registros)")

    # Resumen
    cnt = Counter(r.get("prioridad", "?") for r in resultados)
    print(f"\n  URGENTE   : {cnt.get('URGENTE', 0):>3}")
    print(f"  RUTINARIO : {cnt.get('RUTINARIO', 0):>3}")
    print(f"  DIFERIBLE : {cnt.get('DIFERIBLE', 0):>3}")

    # Costo estimado Haiku: $0.80/MTok in, $4/MTok out, $0.08/MTok cache_read
    costo = (total_in * 0.80 + total_out * 4.0 + total_cache_r * 0.08) / 1_000_000
    print(f"\n  Tokens: in={total_in}  out={total_out}  "
          f"cache_r={total_cache_r}  cache_w={total_cache_w}")
    print(f"  Costo estimado: US${costo:.4f}")


if __name__ == "__main__":
    main()
