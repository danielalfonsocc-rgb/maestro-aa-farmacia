#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
centinela_inyectables_sm.py — Centinela de Inyectables de Salud Mental Ambulatoria
═══════════════════════════════════════════════════════════════════════════════
Lee el Consolidado_AA_MAESTRO.xlsx y el reporte_de_stock_*.xlsx más reciente
(ambos ya descargados/generados por AUTO_SSASUR.py y maestro_aa.py) y vigila
el stock de los antipsicóticos de depósito (LAI) usados en salud mental
ambulatoria:

    Haloperidol decanoato 50 mg, Zuclopentixol decanoato 200 mg/ml,
    Risperidona 25/37,5 mg (microesferas), Paliperidona palmitato 75/100/150 mg

Deliberadamente NO incluye benzodiacepinas (diazepam, lorazepam, midazolam)
ni metadona inyectables: son psicotrópicos/estupefacientes sujetos a control
legal (DS 405 y DS 404) — de uso transversal a otras especialidades y con
circuito de receta retenida propio, no específicos de salud mental ambulatoria.

Alcance de stock: Farmacia AT Abierta + Bodega AT Abierta + Bodega de Fármacos
(el único respaldo real para AT Abierta, ver maestro_aa.py líneas ~799-815).
NUNCA incluye Atención Cerrada.

Dos fuentes de datos por medicamento:
  · "consolidado"   — Haloperidol y Zuclopentixol decanoato SÍ forman parte del
    universo de 378 medicamentos AA (recetados vía AT Abierta con volumen
    suficiente): se leen del Consolidado, con CMP/cobertura/consumo mensual.
  · "stock_directo" — Risperidona y Paliperidona NO alcanzan a formar parte de
    ese universo (bajo volumen histórico vía AT Abierta): se leen directo del
    reporte_de_stock_*.xlsx más reciente, solo con la foto de stock actual
    (sin CMP/cobertura calculados).

maestro_aa.py llama a este módulo automáticamente al final de su propio main()
(mismo patrón que sgli_historico). El reporte HTML solo se regenera cuando:
  · han pasado >= 7 días desde el último reporte generado (cadencia semanal), o
  · el stock (Farmacia AA / Bodega AA / Bodega Fármacos) de alguno de los
    medicamentos vigilados SUBIÓ respecto del último reporte (ingreso detectado).
En cualquier otra corrida, el módulo revisa y sale sin generar nada.

Uso:
  py centinela_inyectables_sm.py             # lógica normal (semanal / ingreso)
  py centinela_inyectables_sm.py --forzar     # genera el reporte igual
  py centinela_inyectables_sm.py --no-pause
"""
import sys
import json
import glob
import argparse
import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

MAESTRO_DIR = Path(__file__).parent
sys.path.insert(0, str(MAESTRO_DIR))
from utils_aa import setup_stdout, norm_erp, HOMOLOGACION

CONSOLIDADO_PATH = MAESTRO_DIR / "Consolidado_AA_MAESTRO.xlsx"
OUT_DIR = MAESTRO_DIR / "Centinela_Inyectables_SM"
ESTADO_PATH = OUT_DIR / "_estado.json"

CADENCIA_DIAS = 7

# Umbrales de cobertura (días hábiles, campo Cobertura_Lab del Consolidado)
UMBRAL_VIGILAR_DIAS = 21
UMBRAL_SOBRESTOCK_DIAS = 250
# Los antipsicóticos de depósito se aplican en dosis discretas por paciente:
# un stock absoluto bajo es frágil aunque la cobertura calculada luzca alta
# (consumo histórico muy bajo).
UMBRAL_STOCK_BAJO = 20

# Mismos nombres canónicos de bodega que usa maestro_aa.py (líneas ~380-390).
BODEGA_FARMACIA_AA = "FARMACIA AT ABIERTA"
BODEGA_BODEGA_AA = "BODEGA AT ABIERTA"
BODEGA_FARMACOS = "BODEGA FARMACOS"

# ── Universo vigilado (juicio clínico, no un filtro automático de datos —
#    ver nota en el reporte). Solo antipsicóticos de depósito: se excluyen a
#    propósito los psicotrópicos/estupefacientes sujetos a control legal.
#    Nombres EXACTOS tal como aparecen en el ERP. ────────────────────────────
TARGET_MEDS = [
    {"nombre": "HALOPERIDOL DECANOATO 50 MG. A", "fuente": "consolidado",
     "uso": "Antipsicótico de depósito — mantención ambulatoria"},
    {"nombre": "ZUCLOPENTIXOL DECANOATO 200 MG/ML INYECTABLE", "fuente": "consolidado",
     "uso": "Antipsicótico de depósito — mantención ambulatoria"},
    {"nombre": "RISPERIDONA 25 MG FA", "fuente": "stock_directo",
     "uso": "Antipsicótico de depósito (microesferas) — mantención ambulatoria"},
    {"nombre": "RISPERIDONA  37,5 MG FA", "fuente": "stock_directo",
     "uso": "Antipsicótico de depósito (microesferas) — mantención ambulatoria"},
    {"nombre": "PALIPERIDONA 75 MG JER.PC", "fuente": "stock_directo",
     "uso": "Antipsicótico de depósito (palmitato) — mantención ambulatoria"},
    {"nombre": "PALIPERIDONA 100 MG JER.PC UNIDAD", "fuente": "stock_directo",
     "uso": "Antipsicótico de depósito (palmitato) — mantención ambulatoria"},
    {"nombre": "PALIPERIDONA 150 MG/ML JER.PC UNIDAD", "fuente": "stock_directo",
     "uso": "Antipsicótico de depósito (palmitato, dosis de carga) — mantención ambulatoria"},
]


def _sheet_rows(wb, nombre_hoja):
    ws = wb[nombre_hoja]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, r)) for r in rows[1:]]


def _localizar_stock_xlsx():
    archivos = glob.glob(str(MAESTRO_DIR / "reporte_de_stock_*.xlsx"))
    if not archivos:
        return None
    return Path(max(archivos, key=lambda p: Path(p).stat().st_mtime))


def _leer_stock_directo(datos, stock_xlsx_path):
    """Rellena farmacia/bodega_aa/bodega_farmacos para los medicamentos
    fuente='stock_directo' directamente desde el reporte de stock crudo
    (mismo criterio de columnas/normalización que maestro_aa.py)."""
    pendientes = {norm_erp(m["nombre"]): m["nombre"]
                  for m in TARGET_MEDS if m["fuente"] == "stock_directo"}
    if not pendientes or stock_xlsx_path is None:
        return

    df = pd.read_excel(stock_xlsx_path, header=2, engine="openpyxl")
    cols = list(df.columns)
    col_desc = next(c for c in cols if "DESCRIPCI" in str(c).upper())
    col_bodega = next(c for c in cols if "BODEGA" in str(c).upper())
    col_cant = next(c for c in cols if str(c).strip().upper() == "CANTIDAD")

    df = df[[col_desc, col_bodega, col_cant]].dropna(subset=[col_desc, col_bodega])
    df[col_cant] = pd.to_numeric(df[col_cant], errors="coerce").fillna(0)
    df[col_bodega] = df[col_bodega].astype(str).str.strip().str.upper()
    df["_norm"] = df[col_desc].astype(str).apply(norm_erp).map(lambda x: HOMOLOGACION.get(x, x))

    for norm_nombre, nombre_original in pendientes.items():
        sub = df[df["_norm"] == norm_nombre]
        datos[nombre_original]["farmacia"] = int(sub[sub[col_bodega] == BODEGA_FARMACIA_AA][col_cant].sum())
        datos[nombre_original]["bodega_aa"] = int(sub[sub[col_bodega] == BODEGA_BODEGA_AA][col_cant].sum())
        datos[nombre_original]["bodega_farmacos"] = int(sub[sub[col_bodega] == BODEGA_FARMACOS][col_cant].sum())


def leer_datos(consolidado_path=CONSOLIDADO_PATH):
    """Cruza Consolidado + reporte de stock crudo y arma un dict por medicamento vigilado."""
    datos = {m["nombre"]: {
        **m,
        "farmacia": 0, "bodega_aa": 0, "bodega_farmacos": 0,
        "cdl": 0.0, "cmp": 0.0, "cobertura": 0.0,
        "consumo_mensual": {}, "demanda_activa": None,
        "pacientes_afectados": None, "criticidad_falt": None,
        "accion_falt": None,
    } for m in TARGET_MEDS}

    if consolidado_path.exists():
        wb = load_workbook(consolidado_path, read_only=True, data_only=True)

        for row in _sheet_rows(wb, "Stock_AA"):
            med = row.get("Medicamento")
            if med in datos and datos[med]["fuente"] == "consolidado":
                datos[med]["farmacia"] = int(row.get("Stock_Farmacia_AA") or 0)
                datos[med]["bodega_aa"] = int(row.get("Stock_Bodega_AA") or 0)
                datos[med]["cdl"] = float(row.get("CDL_DiasHab") or 0)
                datos[med]["cmp"] = float(row.get("CMP_Mensual_22d") or 0)
                datos[med]["cobertura"] = float(row.get("Cobertura_Lab") or 0)

        if "Pedido_Repos_Bodega" in wb.sheetnames:
            for row in _sheet_rows(wb, "Pedido_Repos_Bodega"):
                med = row.get("Medicamento")
                if med in datos and datos[med]["fuente"] == "consolidado":
                    datos[med]["bodega_farmacos"] = int(row.get("Stock_BODEGA_FARMACOS") or 0)

        mes_cols = []
        if "Consumo_Mensual" in wb.sheetnames:
            rows = _sheet_rows(wb, "Consumo_Mensual")
            if rows:
                excluir = {"Medicamento", "Total_Periodo", "Meses_Periodo", "Dias_Lab_Periodo",
                           "CMP_Mensual", "CDL", "Stock_Farmacia_AA", "Stock_Bodega_AA",
                           "Stock_AA_Total", "Cobertura_Lab"}
                mes_cols = [k for k in rows[0].keys() if k not in excluir]
            for row in rows:
                med = row.get("Medicamento")
                if med in datos and datos[med]["fuente"] == "consolidado":
                    datos[med]["consumo_mensual"] = {c: (row.get(c) or 0) for c in mes_cols}

        if "Faltantes_Absolutos_30D" in wb.sheetnames:
            for row in _sheet_rows(wb, "Faltantes_Absolutos_30D"):
                med = row.get("Medicamento")
                if med in datos and datos[med]["fuente"] == "consolidado":
                    datos[med]["demanda_activa"] = row.get("Cant_Demanda_Activa")
                    datos[med]["pacientes_afectados"] = row.get("Pacientes_Afectados")
                    datos[med]["criticidad_falt"] = row.get("Criticidad")
                    datos[med]["accion_falt"] = row.get("Accion_Sugerida")

        wb.close()

    _leer_stock_directo(datos, _localizar_stock_xlsx())
    return datos


def clasificar(d):
    """Devuelve (nivel, etiqueta) según el mismo criterio que arma el reporte."""
    total_aa = d["farmacia"] + d["bodega_aa"]

    if total_aa == 0 and d["bodega_farmacos"] == 0:
        return "critical", "Quiebre total"
    if total_aa == 0 and d["bodega_farmacos"] > 0:
        return "warning", "Traspasar desde Bodega Fármacos"
    if total_aa < UMBRAL_STOCK_BAJO:
        return "warning", "Stock bajo — vigilar"

    if d["fuente"] == "stock_directo":
        # Sin CMP/cobertura calculados (bajo volumen histórico vía AT Abierta):
        # solo se evalúa el stock absoluto, ya cubierto arriba.
        return "neutral", "Normal (sin cobertura calculada)"

    # Sin consumo -> Cobertura_Lab llega como 9999 (sentinel, no un valor real de
    # días) en el Consolidado. Hay que descartar "sin rotación" ANTES de leer el
    # umbral de sobre-stock, o el sentinel se confunde con cobertura real enorme.
    if d["consumo_mensual"] and sum(d["consumo_mensual"].values()) == 0:
        return "neutral", "Sin rotación"
    if d["cobertura"] and d["cobertura"] < UMBRAL_VIGILAR_DIAS:
        return "warning", "Vigilar"
    if d["cobertura"] and d["cobertura"] > UMBRAL_SOBRESTOCK_DIAS:
        return "info", "Sobre-stock"
    return "neutral", "Normal"


def cargar_estado():
    if not ESTADO_PATH.exists():
        return None
    try:
        return json.loads(ESTADO_PATH.read_text(encoding="utf-8"))
    except Exception:
        return None


def guardar_estado(datos, fecha):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    estado = {
        "ultima_fecha_reporte": fecha.isoformat(),
        "ultimo_stock": {
            nombre: {
                "farmacia": d["farmacia"],
                "bodega_aa": d["bodega_aa"],
                "bodega_farmacos": d["bodega_farmacos"],
            } for nombre, d in datos.items()
        },
    }
    ESTADO_PATH.write_text(json.dumps(estado, ensure_ascii=False, indent=2), encoding="utf-8")


def debe_generar(datos, estado_prev):
    if estado_prev is None:
        return True, "primera corrida — sin historial previo"

    try:
        ultima = datetime.date.fromisoformat(estado_prev["ultima_fecha_reporte"])
    except Exception:
        return True, "estado previo ilegible"

    dias = (datetime.date.today() - ultima).days
    if dias >= CADENCIA_DIAS:
        return True, f"cadencia semanal cumplida ({dias} días desde el último reporte)"

    prev_stock = estado_prev.get("ultimo_stock", {})
    for nombre, d in datos.items():
        prev = prev_stock.get(nombre, {"farmacia": 0, "bodega_aa": 0, "bodega_farmacos": 0})
        for campo, etiqueta in (("farmacia", "Farmacia AA"), ("bodega_aa", "Bodega AA"),
                                 ("bodega_farmacos", "Bodega Fármacos")):
            actual = d[campo]
            anterior = prev.get(campo, 0)
            if actual > anterior:
                return True, (f"ingreso detectado en {nombre} — {etiqueta}: "
                               f"{anterior} → {actual} unidades")

    return False, None


# ── Render HTML ───────────────────────────────────────────────────────────────
def _miles(n):
    """Formatea un entero con punto como separador de miles (formato local)."""
    return f"{int(n):,}".replace(",", ".")


def _spark_html(consumo_mensual):
    meses = list(consumo_mensual.items())
    if not meses:
        return ""
    pico = max((v for _, v in meses), default=0) or 1
    barras = []
    for m, v in meses:
        alto = max(3, min(26, int(v / pico * 26)))
        clase = " zero" if v == 0 else ""
        barras.append(f'<i class="bar{clase}" style="height:{alto}px" title="{m}: {v}"></i>')
    return "".join(barras)


def _fila_html(d):
    nivel, etiqueta = clasificar(d)
    spark = _spark_html(d["consumo_mensual"])
    cobertura_txt = f'{d["cobertura"]:.1f}'.replace(".", ",") if d["cobertura"] else "—"
    detalle = d["uso"]
    if d["fuente"] == "consolidado":
        total_prescrito = sum(d["consumo_mensual"].values()) if d["consumo_mensual"] else 0
        detalle += f" &middot; {int(total_prescrito)} u. prescritas en el período &middot; CMP {d['cmp']:.1f}/mes"
    else:
        detalle += " &middot; fuera del universo AA calculado (bajo volumen histórico) — sin CMP/cobertura"
    demanda_txt = ""
    if d["demanda_activa"]:
        demanda_txt = (f'<div class="note-inline">Demanda activa 30d: {int(d["demanda_activa"])} u. '
                        f'&middot; {int(d["pacientes_afectados"] or 0)} paciente(s) afectado(s)</div>')
    total = d["farmacia"] + d["bodega_aa"] + d["bodega_farmacos"]
    return f"""
          <tr>
            <td class="name">{d['nombre'].title()}<span>{detalle}{demanda_txt}</span></td>
            <td><div class="spark">{spark}</div></td>
            <td class="num">{_miles(d['farmacia'])}</td>
            <td class="num">{_miles(d['bodega_aa'])}</td>
            <td class="num">{_miles(d['bodega_farmacos'])}</td>
            <td class="num total">{_miles(total)}</td>
            <td class="num">{cobertura_txt}</td>
            <td><span class="pill {nivel}">{etiqueta}</span></td>
          </tr>"""


def generar_html(datos, fecha):
    total_quiebre = sum(1 for d in datos.values() if clasificar(d)[0] == "critical")
    total_atencion = sum(1 for d in datos.values() if clasificar(d)[0] == "warning")
    pacientes_afectados = sum(int(d["pacientes_afectados"] or 0) for d in datos.values())
    filas = "".join(_fila_html(d) for d in datos.values())

    return f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Centinela Inyectables Salud Mental — {fecha.strftime('%d-%m-%Y')}</title>
<style>
  :root{{
    --bg:#f5f8f8; --surface:#ffffff; --surface-2:#eef3f3; --ink:#152426;
    --ink-muted:#54666a; --border:#dbe4e4; --accent:#2b6777;
    --critical:#b23a2e; --critical-bg:#fbe9e6; --critical-border:#e3b3ab;
    --warning:#93641b; --warning-bg:#faf0dd; --warning-border:#e6cd97;
    --info:#2f6b82; --info-bg:#e7f1f4; --info-border:#b9d6e0;
    --neutral-bg:#eef2f2; --neutral-border:#d3dede;
    --bar:#2b6777; --bar-track:#dde8e8;
  }}
  *{{box-sizing:border-box;}}
  body{{margin:0;background:var(--bg);color:var(--ink);
       font-family:Georgia,"Times New Roman",serif;line-height:1.5;}}
  .page{{max-width:960px;margin:0 auto;padding:2.2rem 1.5rem 3rem;
        font-family:"Segoe UI",Arial,sans-serif;}}
  header{{border-bottom:1px solid var(--border);padding-bottom:1.2rem;margin-bottom:1.6rem;}}
  .eyebrow{{font-family:Consolas,"Courier New",monospace;font-size:.72rem;
           letter-spacing:.08em;text-transform:uppercase;color:var(--accent);font-weight:700;}}
  h1{{font-family:Georgia,serif;font-size:1.9rem;margin:.3rem 0;}}
  .meta{{font-size:.85rem;color:var(--ink-muted);display:flex;gap:1.2rem;flex-wrap:wrap;}}
  .summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:.8rem;margin-bottom:1.8rem;}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:10px;
        padding:.85rem 1rem;}}
  .stat .n{{font-family:Consolas,monospace;font-size:1.5rem;font-weight:700;}}
  .stat .l{{font-size:.76rem;color:var(--ink-muted);}}
  h2{{font-family:Georgia,serif;font-size:1.2rem;margin:1.6rem 0 .5rem;}}
  .sub{{font-size:.83rem;color:var(--ink-muted);margin-bottom:.6rem;}}
  .table-wrap{{background:var(--surface);border:1px solid var(--border);border-radius:10px;
              overflow-x:auto;}}
  table{{width:100%;border-collapse:collapse;min-width:820px;}}
  thead th{{text-align:left;font-size:.7rem;text-transform:uppercase;letter-spacing:.05em;
           color:var(--ink-muted);padding:.75rem .9rem;border-bottom:1px solid var(--border);
           background:var(--surface-2);white-space:nowrap;}}
  tbody td{{padding:.75rem .9rem;border-bottom:1px solid var(--border);font-size:.85rem;
           vertical-align:middle;}}
  tbody tr:last-child td{{border-bottom:none;}}
  td.num{{font-family:Consolas,monospace;text-align:right;white-space:nowrap;}}
  td.num.total{{font-weight:700;color:var(--ink);}}
  td.name{{font-weight:600;}}
  td.name span{{display:block;font-weight:400;font-size:.75rem;color:#849699;margin-top:.15rem;}}
  .note-inline{{color:var(--critical);font-weight:600;margin-top:.2rem;}}
  .pill{{display:inline-flex;align-items:center;gap:.35rem;padding:.25rem .6rem;
        border-radius:999px;font-size:.72rem;font-weight:700;white-space:nowrap;
        border:1px solid transparent;}}
  .pill::before{{content:"";width:.4rem;height:.4rem;border-radius:50%;
               background:currentColor;flex:none;}}
  .pill.critical{{background:var(--critical-bg);color:var(--critical);border-color:var(--critical-border);}}
  .pill.warning{{background:var(--warning-bg);color:var(--warning);border-color:var(--warning-border);}}
  .pill.info{{background:var(--info-bg);color:var(--info);border-color:var(--info-border);}}
  .pill.neutral{{background:var(--neutral-bg);color:var(--ink-muted);border-color:var(--neutral-border);}}
  .spark{{display:flex;align-items:flex-end;gap:3px;height:26px;}}
  .spark i{{display:block;width:8px;border-radius:2px 2px 0 0;background:var(--bar);opacity:.85;}}
  .spark i.zero{{height:2px!important;background:var(--bar-track);opacity:1;}}
  footer{{border-top:1px solid var(--border);padding-top:1rem;margin-top:2rem;
         font-size:.76rem;color:var(--ink-muted);}}
  footer p{{margin:.2rem 0;}}
</style>
</head>
<body>
<div class="page">
  <header>
    <div class="eyebrow">Vigilancia de stock &middot; Farmacia AT Abierta</div>
    <h1>Antipsicóticos de Depósito — Salud Mental Ambulatoria</h1>
    <div class="meta">
      <span>Hospital de Pitrufquén (SSASur)</span>
      <span>Corte: {fecha.strftime('%d-%m-%Y')}</span>
      <span>Alcance: Farmacia AA + Bodega AA + Bodega Fármacos (excluye Atención Cerrada)</span>
    </div>
  </header>

  <div class="summary">
    <div class="stat"><div class="n">{total_quiebre}</div><div class="l">en quiebre total</div></div>
    <div class="stat"><div class="n">{total_atencion}</div><div class="l">en nivel de atención</div></div>
    <div class="stat"><div class="n">{len(datos)}</div><div class="l">inyectables vigilados</div></div>
    <div class="stat"><div class="n">{pacientes_afectados}</div><div class="l">pacientes con demanda activa sin surtir</div></div>
  </div>

  <h2>Antipsicóticos de depósito (LAI)</h2>
  <div class="sub">Uso exclusivo en salud mental: mantención ambulatoria vía inyección intramuscular de depósito. No incluye benzodiacepinas ni metadona inyectables — quedan fuera por estar sujetas a control legal (DS 405/404) y por ser de uso transversal a otras especialidades.</div>
  <div class="table-wrap">
    <table>
      <thead><tr>
        <th>Medicamento</th><th>Tendencia</th><th>Farmacia AA</th><th>Bodega AA</th>
        <th>Bodega Fármacos</th><th>Total</th><th>Cobertura (días háb.)</th><th>Estado</th>
      </tr></thead>
      <tbody>{filas}
      </tbody>
    </table>
  </div>

  <footer>
    <p><strong>Fuente:</strong> Consolidado_AA_MAESTRO.xlsx (Haloperidol/Zuclopentixol decanoato) + reporte_de_stock_*.xlsx más reciente (Risperidona/Paliperidona, fuera del universo de 378 medicamentos AA por bajo volumen histórico vía AT Abierta).</p>
    <p><strong>Cobertura (días hábiles):</strong> stock actual ÷ consumo diario promedio hábil. Solo calculada para los medicamentos que sí forman parte del universo AA.</p>
    <p><strong>Bodega Fármacos:</strong> único respaldo real reconocido para Farmacia/Bodega AT Abierta (ver criterio en maestro_aa.py). Atención Cerrada queda siempre excluida.</p>
    <p><strong>Excluidos a propósito:</strong> diazepam, lorazepam, midazolam y metadona inyectables — psicotrópicos/estupefacientes sujetos a control legal (receta retenida), de uso no exclusivo de salud mental.</p>
    <p><strong>Generado automáticamente</strong> por centinela_inyectables_sm.py al ejecutar maestro_aa.py — cadencia semanal o al detectar ingreso de stock.</p>
  </footer>
</div>
</body>
</html>"""


def main(forzar=False, no_pause=False):
    setup_stdout()
    print()
    print("=" * 62)
    print("  CENTINELA · Antipsicóticos de Depósito — Salud Mental")
    print("=" * 62)

    if not CONSOLIDADO_PATH.exists() and _localizar_stock_xlsx() is None:
        print("  [AVISO] No se encontró Consolidado_AA_MAESTRO.xlsx ni reporte_de_stock_*.xlsx — omitiendo.")
        return

    datos = leer_datos()
    estado_prev = cargar_estado()
    generar, motivo = debe_generar(datos, estado_prev)
    if forzar and not generar:
        generar, motivo = True, "forzado con --forzar"

    if not generar:
        print("  Sin cambios que ameriten un reporte nuevo (ni cadencia semanal cumplida "
              "ni ingreso de stock detectado).")
        print("=" * 62)
        return

    print(f"  Generando reporte — motivo: {motivo}")
    fecha = datetime.date.today()
    out_sub = OUT_DIR / fecha.isoformat()
    out_sub.mkdir(parents=True, exist_ok=True)
    out_html = out_sub / f"reporte_inyectables_sm_{fecha.isoformat()}.html"
    out_html.write_text(generar_html(datos, fecha), encoding="utf-8")
    guardar_estado(datos, fecha)

    print(f"  -> {out_html}")
    for d in datos.values():
        nivel, etiqueta = clasificar(d)
        ico = {"critical": "[CRITICO]", "warning": "[ATENCION]",
               "info": "[INFO]     ", "neutral": "[OK]       "}.get(nivel, "")
        total = d["farmacia"] + d["bodega_aa"] + d["bodega_farmacos"]
        print(f"  {ico} {d['nombre'][:45]:<45} stock_total={total:>5}  {etiqueta}")
    print("=" * 62)

    if not no_pause and __name__ == "__main__":
        try:
            input("Presiona Enter para cerrar...")
        except EOFError:
            pass


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Centinela de Antipsicóticos de Depósito — Salud Mental Ambulatoria")
    parser.add_argument("--forzar", action="store_true", help="Genera el reporte aunque no corresponda por cadencia/ingreso")
    parser.add_argument("--no-pause", action="store_true")
    args = parser.parse_args()
    main(forzar=args.forzar, no_pause=args.no_pause)
