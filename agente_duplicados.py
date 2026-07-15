#!/usr/bin/env python3
"""
agente_duplicados.py
────────────────────────────────────────────────────────────────────────────
Agente IA (Claude · tool use) que lee las recetas del día, detecta
prescripciones duplicadas (mismo paciente + mismo medicamento dentro de una
ventana de N días) y genera un reporte Excel con razonamiento clínico.

Diferencia con auditoria_duplicados_profunda.py (auditoría histórica completa):
  - Este agente es operacional: foco en LAS RECETAS DE HOY y sus antecedentes.
  - Claude decide qué casos investigar más y cómo priorizarlos.
  - Los RUTs nunca salen del proceso local; a la API se envían IDs anónimos.

Uso:
    py agente_duplicados.py
    py agente_duplicados.py --fecha 2026-06-22
    py agente_duplicados.py --fecha 2026-06-22 --ventana 60
    py agente_duplicados.py --fecha 2026-06-22 --salida reporte.xlsx
    py agente_duplicados.py --fecha 2026-06-22 --modelo claude-haiku-4-5-20251001
"""
import argparse
import glob
import hashlib
import os
import re
import sys
from datetime import datetime, timedelta

import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aa_colors import TEAL, ROJO, NARANJA, AMBAR, GRIS_CLR
from utils_aa import norm_erp, HOMOLOGACION, setup_stdout

setup_stdout()
WORK = os.path.dirname(os.path.abspath(__file__))

BODEGA_OBJETIVO = "FARMACIA AT ABIERTA"
ESTADOS_EXCLUIDOS = {"ANULADO", "RECHAZADO", "REEMPLAZADO"}
DIAS_POR_CUOTA = 30
MODELO_DEFAULT = "claude-haiku-4-5-20251001"   # Haiku: tarea mecánica-analítica

THIN   = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Mapa RUN ↔ ID anónimo (los RUTs nunca llegan a la API) ─────────────────
_run_map: dict[str, str] = {}   # anon_id → run_real
_run_rev: dict[str, str] = {}   # run_real → anon_id


def _anon(run: str) -> str:
    if run not in _run_rev:
        h = hashlib.sha256(run.encode()).hexdigest()[:10]
        _run_rev[run] = h
        _run_map[h] = run
    return _run_rev[run]




def parse_cuotas(serie) -> int:
    n = 1
    for v in serie.dropna():
        m = re.search(r"de\s+(\d+)", str(v))
        if m:
            n = max(n, int(m.group(1)))
    return n


# ── Carga y preparación ─────────────────────────────────────────────────────
_cache_df: pd.DataFrame | None = None


def _df() -> pd.DataFrame:
    global _cache_df
    if _cache_df is not None:
        return _cache_df

    files = sorted(glob.glob(os.path.join(WORK, "informe_completo_recetas*.csv")))
    if not files:
        raise FileNotFoundError(
            "No hay archivos informe_completo_recetas*.csv en la carpeta maestro.\n"
            "Ejecuta AUTO_SSASUR.bat primero."
        )

    cols = [
        "ID Receta Detalle", "RUN", "Nombre", "Apellido Paterno", "Apellido Materno",
        "Prescripción", "Número Receta", "Periodo",
        "Fecha Atención", "Fecha Entrega Receta", "Estado Prescripción",
        "RUN Profesional", "Nombre Profesional",
        "Apellido Paterno Profesional", "Apellido Materno Profesional",
        "Especialidad", "Bodega Despacha",
        "Cantidad Recetada", "Cantidad Entregada",
    ]
    chunks = []
    for f in files:
        df = pd.read_csv(f, encoding="latin1", sep=";", on_bad_lines="skip",
                         dtype=str, usecols=lambda c: c in cols)
        chunks.append(df)

    rec = pd.concat(chunks, ignore_index=True)
    rec = rec.drop_duplicates(subset=["ID Receta Detalle"], keep="first")

    bod = rec["Bodega Despacha"].fillna("").apply(norm_erp)
    est = rec["Estado Prescripción"].fillna("").str.upper().str.strip()
    rec = rec[(bod == norm_erp(BODEGA_OBJETIVO)) & (~est.isin(ESTADOS_EXCLUIDOS))].copy()

    rec["_med"] = rec["Prescripción"].fillna("").apply(norm_erp).map(
        lambda x: HOMOLOGACION.get(x, x)
    )
    fa  = pd.to_datetime(rec["Fecha Atención"],       dayfirst=True, errors="coerce")
    fe  = pd.to_datetime(rec["Fecha Entrega Receta"], dayfirst=True, errors="coerce")
    rec["_fecha"]    = fa.fillna(fe)
    rec["_paciente"] = (
        rec["Nombre"].fillna("") + " " + rec["Apellido Paterno"].fillna("") +
        " " + rec["Apellido Materno"].fillna("")
    ).str.strip().str.title()
    rec["_medico"] = (
        rec["Nombre Profesional"].fillna("") + " " +
        rec["Apellido Paterno Profesional"].fillna("") + " " +
        rec["Apellido Materno Profesional"].fillna("")
    ).str.strip().str.title()
    rec["_run_prof"] = rec["RUN Profesional"].fillna("").str.strip()
    rec["_esp"]      = rec["Especialidad"].fillna("").str.strip()
    rec["_est"]      = est
    rec["_cant_e"]   = pd.to_numeric(rec["Cantidad Entregada"], errors="coerce").fillna(0)
    rec["_cant_r"]   = pd.to_numeric(rec["Cantidad Recetada"],  errors="coerce").fillna(0)

    rec = rec.dropna(subset=["_fecha"])
    rec = rec[rec["_med"] != ""]
    rec = rec[rec["RUN"].fillna("").str.strip() != ""]

    _cache_df = rec
    print(f"  Recetas cargadas: {len(rec):,}  "
          f"| rango {rec['_fecha'].min():%d-%m-%Y} – {rec['_fecha'].max():%d-%m-%Y}",
          file=sys.stderr)
    return rec


def _nivel_evento(sub: pd.DataFrame) -> pd.DataFrame:
    """Colapsa cuotas mensuales de una misma indicación al nivel de evento
    (mismo RUN + medicamento + médico + fecha)."""
    def _primero(s):
        s2 = s[s.astype(str).str.strip() != ""]
        return s2.mode().iloc[0] if not s2.empty else ""

    g = sub.groupby(["RUN", "_med", "_run_prof", "_fecha"]).agg(
        paciente  = ("_paciente", _primero),
        med_nombre= ("Prescripción", _primero),
        cuotas    = ("Periodo", parse_cuotas),
        medico    = ("_medico", _primero),
        esp       = ("_esp", _primero),
        cant_r    = ("_cant_r", "sum"),
        cant_e    = ("_cant_e", "sum"),
    ).reset_index()
    g["cob_dias"] = g["cuotas"].clip(lower=1) * DIAS_POR_CUOTA
    return g


# ═══════════════════════════════════════════════════════════════════════════
# HERRAMIENTAS DEL AGENTE
# ═══════════════════════════════════════════════════════════════════════════

def tool_cargar_recetas_dia(fecha: str) -> dict:
    """Estadísticas de las prescripciones del día (sin RUTs)."""
    rec  = _df()
    tgt  = pd.Timestamp(fecha)
    dia  = rec[rec["_fecha"].dt.date == tgt.date()]

    if dia.empty:
        return {"fecha": fecha, "n_lineas": 0, "n_pacientes": 0,
                "n_medicamentos_distintos": 0, "top_medicamentos": [],
                "aviso": "No hay recetas para esta fecha en los CSV disponibles."}

    evts = _nivel_evento(dia)
    top  = (
        evts.groupby("_med")
            .agg(n_eventos=("RUN", "size"), n_pacientes=("RUN", "nunique"))
            .sort_values("n_eventos", ascending=False)
            .head(15)
            .reset_index()
    )
    return {
        "fecha"                  : fecha,
        "n_lineas_csv"           : int(len(dia)),
        "n_eventos_prescripcion" : int(len(evts)),
        "n_pacientes"            : int(evts["RUN"].nunique()),
        "n_medicamentos_distintos": int(evts["_med"].nunique()),
        "top_medicamentos"       : [
            {"medicamento": r["_med"], "eventos": int(r["n_eventos"]),
             "pacientes": int(r["n_pacientes"])}
            for _, r in top.iterrows()
        ],
    }


def tool_detectar_duplicados(fecha: str, ventana_dias: int = 90) -> dict:
    """Encuentra prescripciones del día cuyo paciente ya tenía el mismo
    medicamento en los últimos ventana_dias días. Devuelve IDs anónimos."""
    rec  = _df()
    tgt  = pd.Timestamp(fecha)
    lim  = tgt - timedelta(days=ventana_dias)

    dia  = rec[rec["_fecha"].dt.date == tgt.date()]
    hist = rec[(rec["_fecha"] < tgt) & (rec["_fecha"] >= lim)]

    if dia.empty:
        return {"fecha": fecha, "ventana_dias": ventana_dias,
                "n_casos": 0, "casos": [],
                "aviso": "No hay recetas para esta fecha."}

    evts_dia  = _nivel_evento(dia)
    evts_hist = _nivel_evento(hist)

    casos = []
    for _, row in evts_dia.iterrows():
        run = row["RUN"]
        med = row["_med"]
        previas = evts_hist[(evts_hist["RUN"] == run) & (evts_hist["_med"] == med)]
        if previas.empty:
            continue

        run_anon    = _anon(run)
        profs_prev  = set(p for p in previas["_run_prof"] if p)
        mismo_med   = bool(row["_run_prof"]) and row["_run_prof"] in profs_prev
        dias_ult    = int((tgt - previas["_fecha"].max()).days)
        n_prescr    = len(previas)

        # cobertura aún activa: ¿la última prescripción previa sigue cubriendo hoy?
        ult_evento  = previas.loc[previas["_fecha"].idxmax()]
        fin_cob     = ult_evento["_fecha"] + timedelta(days=int(ult_evento["cob_dias"]))
        cob_activa  = bool(fin_cob >= tgt)

        casos.append({
            "paciente_id"              : run_anon,
            "medicamento"              : med,
            "tipo"                     : "mismo_medico" if mismo_med else "distintos_medicos",
            "dias_desde_ultima_previa" : dias_ult,
            "cobertura_aun_activa"     : cob_activa,
            "n_prescripciones_previas" : n_prescr,
            "hoy": {
                "medico"       : row["medico"],
                "especialidad" : row["esp"] or "—",
                "cant_entregada": int(row["cant_e"]),
            },
        })

    # Ordenar: mayor riesgo primero
    # (distintos médicos + cobertura activa = más urgente)
    def _score(c):
        return (
            0 if (c["tipo"] == "distintos_medicos" and c["cobertura_aun_activa"]) else
            1 if c["tipo"] == "distintos_medicos" else
            2 if c["cobertura_aun_activa"] else 3
        )
    casos.sort(key=_score)

    return {
        "fecha"       : fecha,
        "ventana_dias": ventana_dias,
        "n_casos"     : len(casos),
        "casos"       : casos,
    }


def tool_ver_historial_paciente(paciente_id: str, medicamento: str) -> dict:
    """Historial completo de un paciente para un medicamento.
    El paciente_id es el ID anónimo que devolvió detectar_duplicados."""
    run = _run_map.get(paciente_id)
    if not run:
        return {"error": f"paciente_id desconocido: {paciente_id!r}"}

    rec  = _df()
    hist = rec[(rec["RUN"] == run) & (rec["_med"] == medicamento)].copy()
    if hist.empty:
        return {"paciente_id": paciente_id, "medicamento": medicamento,
                "n_eventos": 0, "prescripciones": []}

    evts = _nivel_evento(hist).sort_values("_fecha", ascending=False)
    prescripciones = []
    for _, r in evts.iterrows():
        fin = r["_fecha"] + timedelta(days=int(r["cob_dias"]))
        prescripciones.append({
            "fecha"           : r["_fecha"].strftime("%d-%m-%Y"),
            "medico"          : r["medico"],
            "especialidad"    : r["esp"] or "—",
            "cuotas"          : int(r["cuotas"]),
            "cobertura_dias"  : int(r["cob_dias"]),
            "fin_cobertura_est": fin.strftime("%d-%m-%Y"),
            "cant_recetada"   : int(r["cant_r"]),
            "cant_entregada"  : int(r["cant_e"]),
        })

    return {
        "paciente_id"  : paciente_id,
        "medicamento"  : medicamento,
        "n_eventos"    : len(prescripciones),
        "prescripciones": prescripciones,
    }


def tool_generar_reporte(casos: list, resumen_ejecutivo: str, fecha: str, ventana_dias: int,
                          dest: str) -> dict:
    """Crea el Excel de reporte. Solo se llama cuando el análisis está listo."""
    rec = _df()

    PRIORIDAD_COLOR = {"ALTA": ROJO, "MEDIA": NARANJA, "BAJA": AMBAR}

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Reporte del día"

    # ── Encabezado ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"Auditoría de Prescripciones Duplicadas — {fecha}  ·  ventana {ventana_dias} días"
    c.font  = Font(bold=True, color="FFFFFF", size=13)
    c.fill  = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Resumen ejecutivo ────────────────────────────────────────────────────
    ws.merge_cells("A2:K2")
    c2 = ws["A2"]
    c2.value = resumen_ejecutivo
    c2.font  = Font(italic=True, size=9, color="374151")
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = max(40, min(len(resumen_ejecutivo) // 5, 90))

    # ── Cabecera de tabla ────────────────────────────────────────────────────
    cabecera = [
        "Prioridad", "Medicamento", "Tipo duplicación",
        "Días desde anterior", "Cobertura activa",
        "N° prescr. previas", "Médico hoy", "Especialidad hoy",
        "Cant. entregada hoy", "Razonamiento IA",
        "Paciente",
    ]
    ws.append([""] * len(cabecera))   # fila 3 vacía (separador)
    ws.append(cabecera)               # fila 4

    fill_hdr = PatternFill("solid", fgColor=TEAL)
    font_hdr = Font(bold=True, color="FFFFFF", size=10)
    for col_i, _ in enumerate(cabecera, 1):
        cell = ws.cell(row=4, column=col_i)
        cell.fill      = fill_hdr
        cell.font      = font_hdr
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = ws["A5"]

    # ── Filas de datos ────────────────────────────────────────────────────────
    for row_i, caso in enumerate(casos, 5):
        pid   = caso["paciente_id"]
        run   = _run_map.get(pid, pid)
        prev  = rec[(rec["RUN"] == run) & (rec["_med"] == caso["medicamento"])].copy()
        pac   = prev["_paciente"].mode().iloc[0] if not prev.empty else "—"

        tipo_label = (
            "Distintos médicos" if caso.get("tipo") == "distintos_medicos" else "Mismo médico"
        )
        cob_label = "Sí" if caso.get("cobertura_aun_activa") else "No"

        vals = [
            caso.get("prioridad", "—"),
            caso.get("medicamento", "—"),
            tipo_label,
            caso.get("dias_desde_ultima_previa", "—"),
            cob_label,
            caso.get("n_prescripciones_previas", "—"),
            caso.get("hoy", {}).get("medico", "—"),
            caso.get("hoy", {}).get("especialidad", "—"),
            caso.get("hoy", {}).get("cant_entregada", "—"),
            caso.get("razonamiento", ""),
            pac,   # columna K = nombre real (local, no enviado a API)
        ]
        ws.append(vals)

        prio    = caso.get("prioridad", "BAJA")
        fgcolor = PRIORIDAD_COLOR.get(prio, AMBAR)
        for col_i, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_i == 10))
            if col_i == 1:
                cell.font  = Font(bold=True, color=fgcolor)
            if row_i % 2 == 0:
                cell.fill  = PatternFill("solid", fgColor=GRIS_CLR)
        ws.row_dimensions[row_i].height = 50 if len(str(vals[9])) > 80 else 30

    # ── Anchos de columna ─────────────────────────────────────────────────────
    anchos = [10, 38, 18, 16, 14, 16, 28, 20, 16, 54, 28]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Auto-filtro ───────────────────────────────────────────────────────────
    if casos:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(cabecera))}{4 + len(casos)}"

    # ── Hoja metodología ──────────────────────────────────────────────────────
    wm = wb.create_sheet("Metodología")
    lineas = [
        ("AUDITORÍA OPERACIONAL DE PRESCRIPCIONES DUPLICADAS", True),
        (f"Generado: {datetime.now():%d-%m-%Y %H:%M}  ·  Fecha auditada: {fecha}  ·  Ventana: {ventana_dias} días", False),
        ("", False),
        ("Fuente y filtros", True),
        ("• Archivos: informe_completo_recetas*.csv (descargados por AUTO_SSASUR.bat)", False),
        (f"• Bodega Despacha = {BODEGA_OBJETIVO}", False),
        ("• Excluidos: ANULADO / RECHAZADO / REEMPLAZADO", False),
        ("", False),
        ("Definición de 'duplicado operacional' (este reporte)", True),
        ("• Una prescripción del día D se marca si el mismo paciente (RUN) recibió", False),
        ("  el mismo medicamento (normalizado) en los últimos N días antes de D.", False),
        ("• Se trabaja a nivel EVENTO de prescripción: las cuotas mensuales de una", False),
        ("  indicación anual (ej. '1 de 12') se colapsan en 1 evento para ese día.", False),
        ("• Cobertura activa: la última prescripción previa aún cubre la fecha de hoy", False),
        (f"  (se estima a razón de {DIAS_POR_CUOTA} días/cuota).", False),
        ("", False),
        ("Priorización por el agente IA (Claude)", True),
        ("• ALTA: distintos médicos + cobertura aún activa (riesgo máximo de sobrestock).", False),
        ("• MEDIA: distintos médicos O cobertura activa (revisar).", False),
        ("• BAJA: mismo médico, sin cobertura activa (probable renovación normal).", False),
        ("", False),
        ("Privacidad", True),
        ("• Los RUTs nunca se envían al modelo de IA: se trabaja con IDs anónimos.", False),
        ("• La columna 'Paciente' de este Excel se reconstruye localmente al generar el reporte.", False),
        ("• Este archivo no debe publicarse fuera de la institución (Ley 19.628).", False),
        ("", False),
        ("Nota clínica", True),
        ("• Este listado es un TAMIZAJE. Cada caso priorizado como ALTA o MEDIA", False),
        ("  debe revisarse en la ficha clínica antes de tomar acción.", False),
    ]
    for i, (txt, bold) in enumerate(lineas, 1):
        cell = wm.cell(row=i, column=1, value=txt)
        cell.font = Font(bold=bold, size=12 if (bold and i == 1) else 10,
                         color=TEAL if bold else "1F2937")
    wm.column_dimensions["A"].width = 92

    wb.save(dest)
    return {"ok": True, "ruta": dest, "n_casos": len(casos),
            "kb": os.path.getsize(dest) // 1024}


# ═══════════════════════════════════════════════════════════════════════════
# DEFINICIÓN DE HERRAMIENTAS PARA LA API
# ═══════════════════════════════════════════════════════════════════════════

TOOLS_SCHEMA = [
    {
        "name": "cargar_recetas_dia",
        "description": (
            "Carga las prescripciones emitidas en una fecha específica en Farmacia AT Abierta. "
            "Devuelve estadísticas resumidas (sin RUTs ni datos de pacientes)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "fecha": {"type": "string",
                          "description": "Fecha a analizar (YYYY-MM-DD)."}
            },
            "required": ["fecha"],
        },
    },
    {
        "name": "detectar_duplicados",
        "description": (
            "Detecta prescripciones del día cuyo paciente ya tenía el MISMO medicamento "
            "en los últimos `ventana_dias` días. Devuelve una lista de casos con IDs "
            "anónimos (nunca RUTs reales). Incluye: tipo (mismo_medico / distintos_medicos), "
            "días desde la última prescripción previa, y si la cobertura anterior aún está activa."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "fecha": {"type": "string",
                          "description": "Fecha del día a auditar (YYYY-MM-DD)."},
                "ventana_dias": {"type": "integer", "default": 90,
                                 "description": "Días hacia atrás a revisar (default 90)."},
            },
            "required": ["fecha"],
        },
    },
    {
        "name": "ver_historial_paciente",
        "description": (
            "Muestra el historial completo de un paciente para un medicamento concreto "
            "(fecha, médico, cuotas, cobertura estimada). Útil para evaluar si un duplicado "
            "es una renovación válida o una prescripción problemática. "
            "Usa el paciente_id anónimo que devuelve detectar_duplicados."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "paciente_id": {"type": "string",
                                "description": "ID anónimo del paciente."},
                "medicamento": {"type": "string",
                                "description": "Nombre normalizado del medicamento (exactamente como aparece en el resultado de detectar_duplicados)."},
            },
            "required": ["paciente_id", "medicamento"],
        },
    },
    {
        "name": "generar_reporte",
        "description": (
            "Genera el reporte Excel con los casos priorizados y el razonamiento clínico. "
            "Llamar SOLO cuando el análisis esté completo. "
            "Cada caso debe incluir prioridad (ALTA / MEDIA / BAJA) y un razonamiento breve."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "casos": {
                    "type": "array",
                    "description": "Lista de casos a incluir, en orden de prioridad (mayor primero).",
                    "items": {
                        "type": "object",
                        "properties": {
                            "paciente_id"              : {"type": "string"},
                            "medicamento"              : {"type": "string"},
                            "prioridad"                : {"type": "string", "enum": ["ALTA", "MEDIA", "BAJA"]},
                            "tipo"                     : {"type": "string"},
                            "dias_desde_ultima_previa" : {"type": "integer"},
                            "cobertura_aun_activa"     : {"type": "boolean"},
                            "n_prescripciones_previas" : {"type": "integer"},
                            "hoy"                      : {
                                "type": "object",
                                "properties": {
                                    "medico"         : {"type": "string"},
                                    "especialidad"   : {"type": "string"},
                                    "cant_entregada" : {"type": "integer"},
                                },
                            },
                            "razonamiento": {
                                "type": "string",
                                "description": "Justificación clínica breve (1-3 oraciones) del nivel de prioridad.",
                            },
                        },
                        "required": ["paciente_id", "medicamento", "prioridad", "razonamiento"],
                    },
                },
                "resumen_ejecutivo": {
                    "type": "string",
                    "description": (
                        "Párrafo breve (3-5 oraciones) con el resumen del día: "
                        "cuántos casos se encontraron, distribución por prioridad, "
                        "principales patrones y recomendación para el farmacéutico."
                    ),
                },
            },
            "required": ["casos", "resumen_ejecutivo"],
        },
    },
]

SYSTEM_PROMPT = """\
Eres un agente farmacéutico clínico del Hospital de Pitrufquén (SSASur, Chile).
Tu tarea es auditar las prescripciones de HOY en busca de posibles DUPLICADOS:
mismo paciente, mismo medicamento, dentro de una ventana de días.

Proceso que debes seguir (sin saltarte pasos):
1. Llama a cargar_recetas_dia para ver qué llegó hoy.
2. Llama a detectar_duplicados para obtener los posibles casos.
3. Para cada caso con tipo=distintos_medicos O cobertura_aun_activa=true,
   llama a ver_historial_paciente para entender el patrón antes de opinar.
   IMPORTANTE: puedes llamar ver_historial_paciente para MÚLTIPLES pacientes
   en la MISMA respuesta (una llamada por paciente en paralelo). No las hagas
   de una en una si puedes agruparlas; ahorra rondas cuando hay varios casos.
4. Prioriza los casos:
   - ALTA: distintos médicos + cobertura activa (sobrestock real o descoordinación).
   - MEDIA: distintos médicos SIN cobertura activa, O mismo médico CON cobertura activa.
   - BAJA: mismo médico, sin cobertura activa (renovación normal con algún solape menor).
5. Llama a generar_reporte con TODOS los casos encontrados (aunque sean BAJA),
   el resumen_ejecutivo y el razonamiento para cada uno.

Criterios clínicos a aplicar:
- Una receta crónica anual genera ~12 números de receta el mismo día: ya están
  colapsados en 1 evento, así que no los cuentes como 12 duplicados.
- Medicamentos de alta alerta (anticoagulantes, hipoglicemiantes, insulinas,
  psicotrópicos, opioides) merecen prioridad más alta aún si el médico es el mismo.
- Si días_desde_ultima_previa > 25 y mismo médico y no hay cobertura activa,
  es casi seguro una renovación normal → prioridad BAJA.
- El razonamiento debe ser en español, conciso, orientado a la acción del farmacéutico.

Restricciones:
- NUNCA menciones RUTs reales en tus mensajes. Solo usa los paciente_id anónimos.
- Si no hay duplicados, igual llama a generar_reporte con lista vacía y resumen ejecutivo.
"""


# ═══════════════════════════════════════════════════════════════════════════
# LOOP DEL AGENTE
# ═══════════════════════════════════════════════════════════════════════════

def _ejecutar_tool(name: str, inputs: dict, fecha: str, ventana_dias: int, dest: str) -> str:
    import json
    if name == "cargar_recetas_dia":
        result = tool_cargar_recetas_dia(**inputs)
    elif name == "detectar_duplicados":
        result = tool_detectar_duplicados(
            fecha=inputs.get("fecha", fecha),
            ventana_dias=inputs.get("ventana_dias", ventana_dias),
        )
    elif name == "ver_historial_paciente":
        result = tool_ver_historial_paciente(**inputs)
    elif name == "generar_reporte":
        result = tool_generar_reporte(
            casos=inputs["casos"],
            resumen_ejecutivo=inputs["resumen_ejecutivo"],
            fecha=fecha,
            ventana_dias=ventana_dias,
            dest=dest,
        )
    else:
        result = {"error": f"Herramienta desconocida: {name}"}
    return json.dumps(result, ensure_ascii=False, default=str)


def run_agent(fecha: str, ventana_dias: int, dest: str, modelo: str) -> None:
    client = anthropic.Anthropic()

    mensaje_inicial = (
        f"Audita las prescripciones del día {fecha} con una ventana de {ventana_dias} días. "
        f"Sigue el proceso completo y genera el reporte."
    )

    messages = [{"role": "user", "content": mensaje_inicial}]
    n_iter  = 0
    reporte_generado = False

    print(f"\n  Agente: modelo={modelo}  fecha={fecha}  ventana={ventana_dias}d")
    print(f"  Cargando datos...", file=sys.stderr)
    _df()   # precarga para que el primer tool call sea instantáneo

    while not reporte_generado:
        n_iter += 1
        if n_iter > 20:
            print("  [AVISO] Se alcanzó el límite de iteraciones (20).")
            break

        resp = client.messages.create(
            model     = modelo,
            max_tokens= 4096,
            system    = [{"type": "text", "text": SYSTEM_PROMPT,
                          "cache_control": {"type": "ephemeral"}}],
            tools     = TOOLS_SCHEMA,
            messages  = messages,
        )

        # Agregar respuesta del asistente al historial
        messages.append({"role": "assistant", "content": resp.content})

        # ── Mostrar texto del asistente ──────────────────────────────────
        for blk in resp.content:
            if blk.type == "text" and blk.text.strip():
                print(f"\n  [Agente] {blk.text.strip()}")

        # ── Fin si el modelo no pidió más tools ──────────────────────────
        if resp.stop_reason == "end_turn":
            break

        # ── Procesar tool_use ────────────────────────────────────────────
        tool_results = []
        for blk in resp.content:
            if blk.type != "tool_use":
                continue
            tool_name = blk.name
            tool_id   = blk.id
            tool_in   = blk.input

            print(f"  → Tool: {tool_name}({', '.join(f'{k}={v!r}' for k, v in tool_in.items() if k != 'casos')})")

            result_str = _ejecutar_tool(tool_name, tool_in, fecha, ventana_dias, dest)

            if tool_name == "generar_reporte":
                import json
                r = json.loads(result_str)
                if r.get("ok"):
                    reporte_generado = True
                    print(f"\n  Excel generado: {r['ruta']}  ({r['kb']} KB, {r['n_casos']} casos)")

            tool_results.append({
                "type"       : "tool_result",
                "tool_use_id": tool_id,
                "content"    : result_str,
            })

        messages.append({"role": "user", "content": tool_results})

    if not reporte_generado:
        print("  [AVISO] El agente no generó el reporte. Prueba con --modelo claude-sonnet-4-6.")


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Agente IA — Duplicados del día (Farmacia AT Abierta)")
    ap.add_argument("--fecha",   default=datetime.now().strftime("%Y-%m-%d"),
                    help="Fecha a auditar (YYYY-MM-DD). Default: hoy.")
    ap.add_argument("--ventana", type=int, default=90,
                    help="Días hacia atrás a revisar. Default: 90.")
    ap.add_argument("--salida",  default=None,
                    help="Ruta del Excel de salida. Default: Duplicados_Dia_AA_<fecha>.xlsx")
    ap.add_argument("--modelo",  default=MODELO_DEFAULT,
                    help=f"Modelo Claude a usar. Default: {MODELO_DEFAULT}")
    args = ap.parse_args()

    dest = args.salida or os.path.join(
        WORK, f"Duplicados_Dia_AA_{args.fecha.replace('-', '')}.xlsx"
    )

    print("=" * 66)
    print("  Agente duplicados — Farmacia AT Abierta")
    print("=" * 66)

    try:
        run_agent(args.fecha, args.ventana, dest, args.modelo)
    except anthropic.APIStatusError as e:
        print(f"\n  [ERROR API] {e.status_code}: {e.message}")
        print("  Verifica que ANTHROPIC_API_KEY esté configurada.")
        sys.exit(1)
    except FileNotFoundError as e:
        print(f"\n  [ERROR] {e}")
        sys.exit(1)

    print("\n  Listo.")


if __name__ == "__main__":
    main()
