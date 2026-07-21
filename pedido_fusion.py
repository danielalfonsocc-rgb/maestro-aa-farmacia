# -*- coding: utf-8 -*-
"""
pedido_fusion.py v2 — Planilla Simplificada de Pedidos AA
==========================================================
Genera  Pedido_Fusion_AA_<fecha>.xlsx  con 4 hojas:

  1 "Farm_Bod"       Farmacia AA → Bodega AA
      Factor_Empaque = Unidades_Caja de SGLI_Estres (ICP CENABAST)
      Pedir Hoy ajustado al día de la semana (no hay columna Req.5d)
  2 "Bod_Farmacos"   Bodega AA → Bodega Fármacos (ciclo 2 semanas)
  3 "Dialisis"       Universo COMPLETO de diálisis (98 meds), todos los días —
      Consumo = Consumo_5D_Solo_Dialisis / 5 × 30 (días naturales)
      El pedido MENSUAL real solo se ejecuta en S3 (--forzar-dialisis marca
      la nota de "semana de pedido" fuera de S3, ya no oculta la hoja)
  4 "Faltantes_AA"   Faltantes absolutos AT Abierta (últimos 30 días) — quiebre
      real de stock (Farm.AA + Bod.AA = 0) con prescripción vigente pendiente
  5 "Por_Agotarse"   Preventiva: Bodega AA en 0 con farmacia aún despachando
      pero cobertura ≤ UMBRAL_PREQUIEBRE días (aún no es quiebre, pero lo será)
  6 "Faltantes_60D"  Retrospectiva: faltantes con demanda pendiente en los
      últimos 60 días que siguen vigentes hoy (incluye stock fantasma que no
      despacha, ej. Empagliflozina 25 mg)

Sin llamadas a IA — solo pandas + openpyxl.

Uso:
    py pedido_fusion.py
    py pedido_fusion.py --forzar-dialisis
    py pedido_fusion.py --todos
"""
import os, math, datetime as dt, argparse, glob, json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils_aa import norm_erp

WORK_DIR     = os.path.dirname(os.path.abspath(__file__))
FERIADOS_CSV = os.path.join(WORK_DIR, 'feriados_chile.csv')
LISTA_MANUAL_JSON = os.path.join(WORK_DIR, 'lista_manual_faltantes.json')
BUFFER_SS    = 1    # días de safety stock (blindaje reapertura lunes)
EXTRA_CRIT   = 1    # días extra SS para criticidad ≤ 2
UMBRAL_PREQUIEBRE = 10   # días de cobertura farmacia bajo los cuales una bodega
                         # AA en 0 se reporta como POR AGOTARSE en Faltantes_AA
CICLO_INICIO = dt.date(2026, 7, 13)   # inicio ciclo Bod→BodFarm; repite cada 10d hábiles
# Recalibrado 2026-07-13: el ancla anterior (2026-06-29) cayó en feriado
# (San Pedro y San Pablo), lo que restó 1 día hábil al 1er ciclo y corrió
# el límite del 2° ciclo de lunes 13-jul a martes 14-jul. Confirmado con
# el usuario que el nuevo período de pedido Bod→BodFarm arranca esta semana.

# ─────────────── helpers ────────────────────────────────────────────────────

def _maestro():
    c = [f for f in glob.glob(os.path.join(WORK_DIR, 'Consolidado_AA_MAESTRO*.xlsx'))
         if not os.path.basename(f).startswith('~$')]
    return max(c, key=os.path.getmtime) if c else \
           os.path.join(WORK_DIR, 'Consolidado_AA_MAESTRO.xlsx')

def _feriados():
    f = {}
    try:
        with open(FERIADOS_CSV, encoding='utf-8') as fh:
            next(fh)
            for ln in fh:
                p = ln.rstrip().split(';')
                if p[0].strip():
                    f[dt.date.fromisoformat(p[0].strip())] = p[1].strip() if len(p) > 1 else ''
    except FileNotFoundError:
        pass
    return f

def _habil(d, fer):
    return d.weekday() < 5 and d not in fer

def _dias_ef(hoy, fer):
    """Días hábiles desde hoy (inclusive) hasta el viernes; 5 si es finde/feriado."""
    if not _habil(hoy, fer):
        return 5
    d, n = hoy, 0
    while d.weekday() < 5:
        if _habil(d, fer):
            n += 1
        d += dt.timedelta(days=1)
    return max(1, n)

def _dias_ciclo(hoy, fer):
    """Días hábiles restantes en el ciclo Bod→BodFarm (10d hábiles, inicio CICLO_INICIO, repite cada 2 semanas)."""
    if hoy < CICLO_INICIO:
        return 10
    d = CICLO_INICIO
    habiles = 0
    while d < hoy:
        if _habil(d, fer):
            habiles += 1
        d += dt.timedelta(days=1)
    pos = habiles % 10        # posición 0-9 dentro del ciclo
    return max(1, 10 - pos)  # días restantes en el ciclo actual

def _semana(d):
    return min((d.day - 1) // 7 + 1, 4)

def _nivel(c):
    s = str(c).strip().upper()
    if len(s) >= 2 and s[0] in '12345' and s[1] == '-':
        return int(s[0])
    if 'CRITICO' in s: return 1
    if 'URGENTE' in s or 'ALTO' in s: return 2
    if 'MODERADO' in s: return 3
    if 'BAJO' in s: return 4
    return 5

def _n(v, d=0.0):
    v = pd.to_numeric(v, errors='coerce')
    return d if pd.isna(v) else float(v)

def _ceil_fe(raw, fe):
    fe = max(1, int(fe))
    if raw <= 0:
        return 0
    return int(math.ceil(raw / fe)) * fe

# ─────────────── paleta de colores tenues ───────────────────────────────────

# (fondo muy claro, texto discreto) — sin colores saturados en urgentes
CPAL = {
    1: ('FFF0EF', 'B91C1C'),   # rojo pálido / vino
    2: ('FFF5EC', 'B45309'),   # durazno pálido / ocre
    3: ('FFFDE7', '856404'),   # crema / dorado
    4: ('F0FDF4', '166534'),   # verde muy claro / esmeralda
    5: ('F9FAFB', '374151'),   # gris casi blanco / gris oscuro
}
DIAL_BG = ('EFF6FF', '1E40AF')   # azul muy claro para diálisis
THIN    = Side(style='thin', color='DCDCDC')
BRD     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFILL   = PatternFill('solid', fgColor='D1FAF5')
HFONT   = Font(bold=True, color='065F46', name='Arial', size=10)

def _pfill(hx):
    h = str(hx).lstrip('#')
    if len(h) == 6:
        h = 'FF' + h   # openpyxl necesita ARGB de 8 dígitos; sin FF queda transparente
    return PatternFill('solid', fgColor=h)

def _totals(ws, data_start, data_end, ncols, sum_cols):
    """Fila de totales en negrita inmediatamente después de los datos."""
    tr = data_end + 1
    for j in range(1, ncols + 1):
        c = ws.cell(tr, j)
        c.fill = _pfill('D1FAF5'); c.border = BRD
        if j == 1:
            c.value = 'TOTAL'
            c.font = Font(bold=True, name='Arial', size=10, color='065F46')
        elif j in sum_cols:
            c.value = f'=SUM({get_column_letter(j)}{data_start}:{get_column_letter(j)}{data_end})'
            c.font = Font(bold=True, name='Arial', size=10, color='065F46')
            c.alignment = Alignment(horizontal='center')


def _hdr(ws, row, hdrs):
    for j, (label, w) in enumerate(hdrs, 1):
        c = ws.cell(row, j, label)
        c.fill = HFILL; c.font = HFONT; c.border = BRD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 26

def _titulo(ws, txt, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, txt)
    t.font = Font(bold=True, size=12, color='065F46', name='Arial')
    ws.row_dimensions[1].height = 22

def _subtit(ws, txt, ncols, height=36):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, txt)
    c.font = Font(italic=True, size=9, color='555555', name='Arial')
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = height

FMT1D = '0.0'   # 1 decimal para coberturas y CDL

def _fila_crit(ws, row_i, vals, crit, cols_center, cols_fmt1d=None):
    nv = _nivel(crit)
    bg, fg = CPAL.get(nv, CPAL[5])
    for j, v in enumerate(vals, 1):
        c = ws.cell(row_i, j, v)
        c.border = BRD
        c.fill = _pfill(bg)
        c.font = Font(name='Arial', size=10, color=fg)
        if j in cols_center:
            c.alignment = Alignment(horizontal='center')
        if cols_fmt1d and j in cols_fmt1d:
            c.number_format = FMT1D

def _fila_dial(ws, row_i, vals, cols_center, cols_fmt1d=None):
    bg, fg = DIAL_BG
    for j, v in enumerate(vals, 1):
        c = ws.cell(row_i, j, v)
        c.border = BRD
        c.fill = _pfill(bg)
        c.font = Font(name='Arial', size=10, color=fg)
        if j in cols_center:
            c.alignment = Alignment(horizontal='center')
        if cols_fmt1d and j in cols_fmt1d:
            c.number_format = FMT1D

# ─────────────── carga de datos ─────────────────────────────────────────────

def _leer(maestro):
    def sh(name):
        try:
            return pd.read_excel(maestro, sheet_name=name, engine='openpyxl')
        except Exception:
            return pd.DataFrame()
    return {k: sh(v) for k, v in {
        'sgli'    : 'SGLI_Estres',
        'farm'    : 'Pedido_Farm_Bodega',
        'bod'     : 'Pedido_Repos_Bodega',
        'dialmed' : 'Dialisis_Medicamentos',
        'falt30'  : 'Faltantes_Absolutos_30D',
        'falt60'  : 'Faltantes_60D_Persistente',
        'stock'   : 'Stock_AA',
    }.items()}

# ─────────────── Hoja 1: Farmacia AA → Bodega AA ────────────────────────────

HDRS1 = [
    ('Medicamento',              44),
    ('Criticidad',               12),
    ('Stock Farm. AA (ud)',      14),   # Stock_Farm_Actual
    ('Cob. actual (días)',       13),   # Stock / CDL_Trend
    ('CDL Tend. (ud/d)',         12),
    ('A Pedir (ud)',             13),   # redondeado al Fe ICP CENABAST (invisible)
    ('Farm → Bod AA',            34),   # traspaso Bodega AA → Farmacia
    ('Bod AA ← Bod.Fármacos',   34),   # reposición que Bodega AA necesita de BodFarm
]

def calc_h1(df_farm, fe_map, dias_ef, todos=False, rep_h2_map=None):
    return [x['v'] for x in _calc_h1_rows(df_farm, fe_map, dias_ef, todos, rep_h2_map)]


def _calc_h1_rows(df_farm, fe_map, dias_ef, todos=False, rep_h2_map=None):
    """Filas completas de la hoja Farm_Bod, incluyendo campos internos (p.ej.
    '_sbod') que calc_h1() no expone en su tupla pública 'v' pero que otras
    salidas (pedido_fusion_simple.py) sí necesitan."""
    rep_h2_map = rep_h2_map or {}
    rows = []
    for _, r in df_farm.iterrows():
        med   = str(r.get('Medicamento', '')).strip()
        crit  = str(r.get('Criticidad', '5-OK'))
        nv    = _nivel(crit)
        # CDL de tendencia (Consumo_5D_Trend ya tiene factor de carga semanal aplicado)
        trend5d = _n(r.get('Consumo_5D_Trend', 0))
        cdl     = (trend5d / 5) if trend5d > 0 else _n(r.get('CDL_DiasHab', 0))
        sfarm   = int(_n(r.get('Stock_Farm_Actual', 0)))
        sbod    = int(_n(r.get('Stock_Bodega_Disponible', 0)))

        # Cobertura actual (antes del pedido)
        cob_actual = round(sfarm / cdl, 1) if cdl > 0 else 0.0

        ss_dias = BUFFER_SS + (EXTRA_CRIT if nv <= 2 else 0)
        ss      = math.ceil(cdl * ss_dias) if cdl > 0 else 0
        raw     = max(0, math.ceil(cdl * dias_ef) + ss - sfarm) if cdl > 0 else 0

        # Redondear al múltiplo del factor de empaque ICP CENABAST — Bodega AA
        # es bodega activa, toda cantidad que se pida (no la que ya está físicamente
        # en el estante) se ajusta siempre a caja completa.
        fe  = int(fe_map.get(med, 1)) or 1
        ud  = _ceil_fe(raw, fe)

        if not todos and ud <= 0:
            continue

        # Lo que Bodega AA le falta a Farmacia se netea contra lo que el ciclo
        # Bod_Farmacos ya trae en camino para este medicamento (rep_h2_map), igual
        # que se hace en la hoja Diálisis — evita pedir dos veces a Bod.Fármacos
        # contra el mismo déficit. El remanente se re-redondea al Fe.
        rep_bod       = int(rep_h2_map.get(med, 0))
        deficit_bruto = max(0, ud - sbod)
        deficit_neto  = _ceil_fe(max(0, deficit_bruto - rep_bod), fe) if deficit_bruto > 0 else 0

        if ud <= 0:
            accion1, accion2 = '', ''
        elif sbod >= ud:
            accion1 = f'Traspasar {ud} ud → Farmacia'
            accion2 = ''
        else:
            accion1 = f'Traspasar {sbod} ud → Farmacia' if sbod > 0 else 'SIN STOCK en Bodega AA'
            if deficit_neto > 0:
                etiqueta = 'COMPRA URGENTE' if sbod <= 0 else 'Reponer Bod.AA:'
                accion2  = f'{etiqueta} {deficit_neto} ud de Bod.Fármacos'
                if rep_bod > 0:
                    accion2 += f' (neteado con {rep_bod} ud que ya trae el ciclo Bod_Farmacos)'
            else:
                accion2 = f'Cubierto por ciclo Bod_Farmacos en camino ({rep_bod} ud)'

        rows.append({
            'v': (med, crit, sfarm, cob_actual, round(cdl, 1), ud, accion1, accion2),
            '_nv': nv, '_ud': ud, '_sbod': sbod,
        })

    rows.sort(key=lambda x: x['v'][0])
    return rows


def write_h1(ws, rows, dias_ef, hoy, semana):
    ndia = ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom'][hoy.weekday()]
    _titulo(ws,
        f'FARMACIA AA → BODEGA AA  ·  {ndia} {hoy.strftime("%d/%m/%Y")}  ·  '
        f'S{semana}  ·  {dias_ef} día(s) hábil(es) restante(s)',
        len(HDRS1))
    _subtit(ws,
        f'Stock Farm.AA = stock actual en Farmacia | Cob.actual = Stock/CDL (días antes del pedido) | '
        f'A Pedir = CDL×{dias_ef}d+SS−Stock, redondeado al empaque CENABAST | '
        f'Col 7: traspaso Farm←Bod.AA | Col 8: reposición Bod.AA←Bod.Fármacos, neteada con lo que '
        f'ya trae el ciclo Bod_Farmacos y re-redondeada al empaque '
        f'(rosa = sin stock, ámbar = compra urgente, verde = cubierto por el ciclo en camino)',
        len(HDRS1))
    hdrs = list(HDRS1)
    hdrs[5] = (f'A Pedir ({dias_ef}d hab., ud)', 13)
    _hdr(ws, 3, hdrs)
    for i, vals in enumerate(rows, 4):
        # vals: (med, crit, sfarm, cob_actual, cdl, ud, accion1, accion2)
        _fila_crit(ws, i, vals, vals[1], {2, 3, 4, 5, 6}, cols_fmt1d={4, 5})
        a1, a2 = str(vals[6]), str(vals[7])
        if 'SIN STOCK' in a1:
            c = ws.cell(i, 7)
            c.fill = _pfill('FFDAD6')
            c.font = Font(name='Arial', size=10, color='9B1C1C', bold=True)
        if 'COMPRA URGENTE' in a2:
            c = ws.cell(i, 8)
            c.fill = _pfill('FEF08A')
            c.font = Font(name='Arial', size=10, color='854D0E', bold=True)
        elif 'Reponer' in a2:
            c = ws.cell(i, 8)
            c.fill = _pfill('FFF9C4')
            c.font = Font(name='Arial', size=10, color='78350F')
        elif 'Cubierto' in a2:
            c = ws.cell(i, 8)
            c.fill = _pfill('F0FDF4')
            c.font = Font(name='Arial', size=10, color='166534')
    ws.freeze_panes = 'A4'
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS1))}{last}'
        _totals(ws, 4, last, len(HDRS1), {6})
        ws.print_area = f'A1:{get_column_letter(len(HDRS1))}{last + 1}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


# ─────────────── Hoja 2: Bodega AA → Bodega Fármacos ────────────────────────

HDRS2 = [
    ('Medicamento',              44),
    ('Criticidad',               12),
    ('Stock Bod. AA',            12),
    ('Cob. Bod.AA (días)',       15),   # Stock_Bod_Actual / CDL (antes del pedido)
    ('Stock Farm. AA',           12),
    ('Stock Bod. Fármacos',      14),
    ('Req. ciclo (ud)',          13),   # CDL × dias_ciclo; header dinámico en write_h2
    ('A Reponer (ud)',           13),   # max(0, req_ciclo - (sbod+sfarm)), redondeado al Fe
    ('Accion',                   44),
]

def calc_h2(df_bod, fe_map, hoy, fer):
    dc = _dias_ciclo(hoy, fer)
    rows = []
    for _, r in df_bod.iterrows():
        med    = str(r.get('Medicamento', '')).strip()
        crit   = str(r.get('Criticidad', '5-OK'))
        sbod   = int(_n(r.get('Stock_Bod_Actual', 0)))
        sfarm  = int(_n(r.get('Stock_Farm_Actual', 0)))
        sbfarm = int(_n(r.get('Stock_BODEGA_FARMACOS', 0)))
        cons10 = _n(r.get('Consumo_10D_Trend', 0))
        req2   = _n(r.get('Req_2_Semanas', 0))

        # CDL de tendencia (Consumo_10D_Trend ya trae el ajuste por semana del mes,
        # igual que Consumo_5D_Trend en Farm_Bod); cae al req plano de 2 semanas
        # (10d hábiles) solo si no hay dato de tendencia.
        cdl = (cons10 / 10) if cons10 > 0 else (req2 / 10 if req2 > 0 else 0.0)

        # Cobertura actual de Bodega AA (antes del pedido)
        cob_bod = round(sbod / cdl, 1) if cdl > 0 else 0.0

        # Stock requerido para el ciclo actual (días restantes). Se descuenta TODO
        # el stock de Atención Abierta (Bodega AA + Farmacia AA) — el de Bodega
        # Fármacos vive en otra ubicación física y solo decide la acción (traspaso
        # vs. compra externa), no reduce la necesidad.
        req_ciclo   = math.ceil(cdl * dc) if cdl > 0 else 0
        fe          = int(fe_map.get(med, 1)) or 1
        rep         = _ceil_fe(max(0, req_ciclo - (sbod + sfarm)), fe)

        if rep <= 0:
            accion = ''
        elif sbfarm >= rep:
            accion = f'Pedir {rep} ud a Bod.Fármacos'
        else:
            falt = rep - sbfarm
            accion = f'Bod.Fármacos: {sbfarm} ud disponibles | COMPRA EXTERNA: {falt} ud'

        rows.append({
            'v': (med, crit, sbod, cob_bod, sfarm, sbfarm, req_ciclo, rep, accion),
            '_nv': _nivel(crit), '_rep': rep,
        })

    rows.sort(key=lambda x: x['v'][0])
    return dc, [x['v'] for x in rows]


def write_h2(ws, rows, hoy, semana, dias_ciclo):
    _titulo(ws,
        f'BODEGA AA → BODEGA FÁRMACOS  ·  Ciclo {dias_ciclo}d hábiles restantes  ·  '
        f'{hoy.strftime("%d/%m/%Y")}  ·  S{semana}',
        len(HDRS2))
    _subtit(ws,
        f'Ciclo 10d hábiles (inicio {CICLO_INICIO.strftime("%d-%m")}, repite c/2 semanas) | Quedan {dias_ciclo}d | '
        f'Cob. Bod.AA = Stock Bod.AA ÷ CDL (días de cobertura actual, antes del pedido) | '
        f'A Reponer = max(0, CDL×{dias_ciclo}d − (Stock Bod.AA + Stock Farm.AA)), '
        f'redondeado al empaque CENABAST | '
        f'Stock Bod.Fármacos solo decide la acción: pedir traspaso o compra externa | '
        f'Ámbar = compra externa a Bod.Fármacos',
        len(HDRS2))
    hdrs = list(HDRS2)
    hdrs[6] = (f'Req. ciclo ({dias_ciclo}d, ud)', 13)
    _hdr(ws, 3, hdrs)
    for i, vals in enumerate(rows, 4):
        _fila_crit(ws, i, vals, vals[1], {2, 3, 4, 5, 6, 7, 8}, cols_fmt1d={4})
        ac = str(vals[8])
        if 'COMPRA EXTERNA' in ac:
            c = ws.cell(i, 9)
            c.fill = _pfill('FEF08A')
            c.font = Font(name='Arial', size=10, color='854D0E', bold=True)
    ws.freeze_panes = 'A4'
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS2))}{last}'
        _totals(ws, 4, last, len(HDRS2), {8})
        ws.print_area = f'A1:{get_column_letter(len(HDRS2))}{last + 1}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


# ─────────────── Hoja 3: Diálisis (Farm + Bod) ──────────────────────────────

HDRS3 = [
    ('Medicamento',              44),
    ('Fe (ud/envase)',           13),   # Factor de empaque CENABAST usado para redondear
    ('Cons. Mensual Diál. (ud)', 18),
    ('Cob.Farm (mes)',           12),   # Stock Farm / Cons.Mensual
    ('A Pedir Farm. (ud)',       14),
    ('Cob.Bod (mes)',            12),   # Stock Bod / Cons.Mensual
    ('A Pedir Bod. (ud)',        14),
    ('Obs',                      46),
]

def calc_h3(df_dialmed, fe_map, rep_h2_map=None):
    """Universo COMPLETO de medicamentos de diálisis (hoja Dialisis_Medicamentos
    de maestro_aa.py, sin filtrar por Necesidad>0) — a diferencia de la versión
    anterior (que solo veía Dialisis_Pedido_Farm/Bod, ya filtradas a lo
    accionable), esto muestra TODO medicamento con consumo de diálisis aunque
    el stock alcance (ej. Furosemida), consistente todos los días, no solo S3."""
    rep_h2_map = rep_h2_map or {}
    rows = []
    for _, r in df_dialmed.iterrows():
        med = str(r.get('Medicamento', '')).strip()
        c5d = _n(r.get('Consumo_5D_Solo_Dialisis', 0))
        if c5d <= 0:
            continue
        # Unidades_Caja de SGLI_Estres (cenabast_tallas.csv, curado a mano) es la
        # fuente confiable de empaque — cae a 1 (sin redondeo) si el medicamento
        # no tiene talla conocida ahí.
        fe = int(fe_map.get(med, 0)) or 1
        mensual = _ceil_fe(c5d / 5 * 30, fe)
        sfarm   = int(_n(r.get('Stock_Farmacia_AA', 0)))
        sbod    = int(_n(r.get('Stock_Bodega_AA', 0)))
        apfarm  = _ceil_fe(max(0, mensual - sfarm), fe)

        # El CDL de Bod_Farmacos ya es COMBINADO (incluye diálisis, ver calc_h2),
        # así que el "A Reponer" de esa hoja ya trae en camino parte (o todo) de lo
        # que diálisis necesita de Bodega AA. Si ambas hojas se piden la misma
        # semana (coincide con S3), sin netear se pediría dos veces a Bod.Fármacos
        # contra el mismo stock base. Se descuenta el rep del ciclo antes de pedir.
        rep_bod   = int(rep_h2_map.get(med, 0))
        sbod_proy = sbod + rep_bod
        apbod     = _ceil_fe(max(0, mensual - sbod_proy), fe)
        cob_farm  = round(sfarm / mensual, 1) if mensual > 0 else None
        cob_bod   = round(sbod  / mensual, 1) if mensual > 0 else None

        obs = []
        if apfarm > 0: obs.append(f'Farm: solicitar {apfarm} ud a Bodega AA')
        if apbod  > 0: obs.append(f'Bod:  solicitar {apbod} ud a Bod.Fármacos')
        if rep_bod > 0 and (apbod > 0 or max(0, mensual - sbod) > 0):
            obs.append(f'(neteado con {rep_bod} ud que ya trae el ciclo Bod_Farmacos)')
        if not obs:    obs = ['Stock suficiente en ambos niveles']

        # (med, fe, mensual, cob_farm, apfarm, cob_bod, apbod, obs)
        rows.append((med, fe, mensual, cob_farm, apfarm, cob_bod, apbod, ' / '.join(obs)))

    rows.sort(key=lambda x: x[0])
    return rows


def write_h3(ws, rows, hoy, semana, es_semana_pedido):
    _titulo(ws,
        f'DIÁLISIS MENSUAL — Farm. AA + Bodega AA  ·  {hoy.strftime("%d/%m/%Y")}  ·  S{semana}',
        len(HDRS3))
    nota_semana = ('✅ S3 — semana de pedido: ejecutar los traspasos/compras de esta hoja.'
                   if es_semana_pedido else
                   f'ℹ️ Solo consulta — el pedido mensual de diálisis se ejecuta en la 3ª semana '
                   f'del mes (S3). Semana actual: S{semana}. Esta hoja se genera todos los días '
                   f'para ver el consumo, pero no se debe pedir fuera de S3.')
    _subtit(ws,
        f'{nota_semana}  |  '
        'Cons.Mensual = C5D_Diál÷5×30, redondeado al empaque CENABAST (días naturales) | '
        'Cob.Farm/Bod = Stock÷Cons.Mensual (en meses, sobre stock actual) | '
        'A Pedir Bod. = max(0, Mensual − (Stock Bod.AA + lo que ya trae el ciclo Bod_Farmacos)) '
        '— evita pedir dos veces a Bod.Fármacos contra el mismo stock',
        len(HDRS3), height=48)
    _hdr(ws, 3, HDRS3)
    for i, vals in enumerate(rows, 4):
        # fe@col2, mensual@col3, cob_farm@col4, apfarm@col5, cob_bod@col6, apbod@col7
        _fila_dial(ws, i, vals, {2, 3, 4, 5, 6, 7}, cols_fmt1d={4, 6})
    ws.freeze_panes = 'A4'
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS3))}{last}'
        _totals(ws, 4, last, len(HDRS3), {3, 5, 7})
        ws.print_area = f'A1:{get_column_letter(len(HDRS3))}{last + 1}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


# ─────────────── Hoja 4: Faltantes Absolutos AA (30 días) ──────────────────

HDRS4 = [
    ('Medicamento',              44),
    ('Criticidad',               34),
    ('Pacientes Afectados',      15),
    ('Recetas',                  10),
    ('Faltante (ud)',            13),
    ('Stock Bod. Fármacos',      16),
    ('Acción',                   50),
]

def calc_h4(df_falt30):
    """Filas ya vienen filtradas y agregadas desde maestro_aa.py (hoja
    Faltantes_Absolutos_30D): prescripción vigente de los últimos 30 días en
    el mostrador de Atención Abierta con Stock Farm.AA + Bod.AA = 0 (quiebre
    real, no solo demanda pendiente cubierta por stock existente)."""
    if df_falt30 is None or not len(df_falt30):
        return []
    rows = []
    for _, r in df_falt30.iterrows():
        med    = str(r.get('Medicamento', '')).strip()
        crit   = str(r.get('Criticidad', ''))
        pax    = int(_n(r.get('Pacientes_Afectados', 0)))
        nrec   = int(_n(r.get('N_Recetas', 0)))
        falt   = int(_n(r.get('Faltante_Neto', 0)))
        sbf    = int(_n(r.get('Stock_BODEGA_FARMACOS', 0)))
        accion = str(r.get('Accion_Sugerida', ''))
        rows.append({
            'v': (med, crit, pax, nrec, falt, sbf, accion),
            '_nv': _nivel(crit),
        })
    rows.sort(key=lambda x: (x['_nv'], -x['v'][2]))
    return [x['v'] for x in rows]


HDRS4B = [
    ('Medicamento',              44),
    ('Cobertura Farm. (días)',   34),
    ('Stock Farm. AA (ud)',      15),
    ('Consumo mensual (ud)',     10),
    ('CDL (ud/día)',             13),
    ('Stock Bod. Fármacos',      16),
    ('Acción',                   50),
]

def calc_h4b(df_stock, df_bod):
    """Pre-quiebres: Bodega AA en 0 con farmacia aún despachando pero con
    cobertura ≤ UMBRAL_PREQUIEBRE días. Incorporado tras el cruce con la
    detección manual del 21-07-2026: estos casos (ej. Empagliflozina 25 mg
    con 1,6 días de cobertura) no aparecían en ninguna hoja hasta consumar
    el quiebre total. Devuelve [(vals, criticidad_sintetica), ...]."""
    if df_stock is None or not len(df_stock):
        return []
    sbf_map = {}
    if df_bod is not None and len(df_bod):
        for _, r in df_bod.iterrows():
            sbf_map[str(r.get('Medicamento', '')).strip()] = \
                int(_n(r.get('Stock_BODEGA_FARMACOS', 0)))
    rows = []
    for _, r in df_stock.iterrows():
        med   = str(r.get('Medicamento', '')).strip()
        sbod  = _n(r.get('Stock_Bodega_AA', 0))
        sfarm = _n(r.get('Stock_Farmacia_AA', 0))
        cob   = _n(r.get('Cobertura_Lab', 0))
        cmp_  = _n(r.get('CMP_Mensual_22d', 0))
        cdl   = _n(r.get('CDL_DiasHab', 0))
        if sbod > 0 or sfarm <= 0 or cmp_ <= 0 or cob > UMBRAL_PREQUIEBRE:
            continue
        sbf = sbf_map.get(med, 0)
        if sbf > 0:
            accion = f'REPONER BODEGA AA DESDE BODEGA FARMACOS ({sbf} ud. disponibles)'
        else:
            accion = 'GESTIONAR COMPRA ANTES DEL QUIEBRE — SIN RESPALDO EN BODEGA FARMACOS'
        crit = '1-CRITICO' if cob <= 3 else ('2-ALTO' if cob <= 5 else '3-MODERADO')
        rows.append(((med, round(cob, 1), int(sfarm), int(round(cmp_)),
                      round(cdl, 1), sbf, accion), crit))
    rows.sort(key=lambda x: x[0][1])
    return rows


def write_h4(ws, rows, hoy):
    _titulo(ws,
        f'FALTANTES ABSOLUTOS AT ABIERTA — últimos 30 días  ·  {hoy.strftime("%d/%m/%Y")}',
        len(HDRS4))
    _subtit(ws,
        'Medicamentos con prescripción vigente (PENDIENTE/SOLICITADO) de los últimos 30 días en '
        'el mostrador de Atención Abierta que NO se ha podido despachar por quiebre real de stock '
        '(Stock Farmacia AA + Bodega AA = 0 en este momento) | Faltante = unidades pendientes de '
        'entrega | Stock Bod. Fármacos indica si hay respaldo para traspaso o si se requiere compra '
        'urgente',
        len(HDRS4), height=48)
    _hdr(ws, 3, HDRS4)
    for i, vals in enumerate(rows, 4):
        _fila_crit(ws, i, vals, vals[1], {2, 3, 4, 5})
        ac = str(vals[6])
        if 'COMPRA URGENTE' in ac:
            c = ws.cell(i, 7)
            c.fill = _pfill('FFDAD6')
            c.font = Font(name='Arial', size=10, color='9B1C1C', bold=True)
        elif 'TRASPASAR' in ac:
            c = ws.cell(i, 7)
            c.fill = _pfill('F0FDF4')
            c.font = Font(name='Arial', size=10, color='166534')
    ws.freeze_panes = 'A4'
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS4))}{last}'
        _totals(ws, 4, last, len(HDRS4), {3, 4, 5})
        ws.print_area = f'A1:{get_column_letter(len(HDRS4))}{last + 1}'
    else:
        ws.cell(4, 1, 'Sin faltantes absolutos en los últimos 30 días.')
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


def write_h4b(ws, rows_pre, hoy):
    """Hoja Por_Agotarse: Bodega AA en 0 con farmacia aún despachando."""
    _titulo(ws,
        f'POR AGOTARSE — BODEGA AA EN 0  ·  {hoy.strftime("%d/%m/%Y")}',
        len(HDRS4B))
    _subtit(ws,
        f'Medicamentos SIN stock en Bodega AA cuya Farmacia AA aún despacha pero con cobertura '
        f'≤ {UMBRAL_PREQUIEBRE} días (aún NO es quiebre: anticipar reposición o compra antes de que '
        f'caigan a Faltantes_AA) | Orden: cobertura ascendente | Acción verde = hay respaldo en '
        f'Bodega Fármacos para traspaso, roja = requiere gestión de compra',
        len(HDRS4B), height=48)
    _hdr(ws, 3, HDRS4B)
    for i, (vals, crit) in enumerate(rows_pre, 4):
        _fila_crit(ws, i, vals, crit, {2, 3, 4, 5, 6}, cols_fmt1d={2, 5})
        ac = str(vals[6])
        c = ws.cell(i, 7)
        if 'GESTIONAR COMPRA' in ac:
            c.fill = _pfill('FFDAD6')
            c.font = Font(name='Arial', size=10, color='9B1C1C', bold=True)
        elif 'REPONER' in ac:
            c.fill = _pfill('F0FDF4')
            c.font = Font(name='Arial', size=10, color='166534')
    ws.freeze_panes = 'A4'
    if rows_pre:
        last = 3 + len(rows_pre)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS4B))}{last}'
        ws.print_area = f'A1:{get_column_letter(len(HDRS4B))}{last}'
    else:
        ws.cell(4, 1, 'Sin medicamentos por agotarse (Bodega AA en 0) hoy.')
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


# ─────────────── Hoja 6: Faltantes Persistentes (60 días) ──────────────────

HDRS4C = [
    ('Medicamento',              44),
    ('Lista Manual',            11),
    ('Días en Falta',           11),
    ('Desde',                   12),
    ('Pacientes Afectados',     13),
    ('Recetas',                  9),
    ('Pendiente (ud)',          12),
    ('Stock AA Reportado',      14),
    ('Stock Bod. Fármacos',     16),
    ('Acción',                  52),
]

def _lista_manual():
    """Set de nombres normalizados de la lista de faltantes detectados a mano
    por la QF (lista_manual_faltantes.json). Vacío si el archivo no existe."""
    try:
        with open(LISTA_MANUAL_JSON, encoding='utf-8') as fh:
            data = json.load(fh)
        return {norm_erp(m) for m in data.get('medicamentos', [])}
    except (FileNotFoundError, json.JSONDecodeError):
        return set()


def calc_h4c(df_falt60, manual=None):
    """Filas de la hoja Faltantes_60D_Persistente del consolidado (calculada en
    maestro_aa.py sec.8c): faltantes del mostrador AA con demanda pendiente en
    los últimos 60 días que siguen vigentes hoy. Ya viene ordenada por días en
    falta descendente. `manual` = set de nombres normalizados de la lista manual
    para marcar la columna 'Lista Manual'."""
    if df_falt60 is None or not len(df_falt60):
        return []
    manual = manual or set()
    rows = []
    for _, r in df_falt60.iterrows():
        med    = str(r.get('Medicamento', '')).strip()
        crit   = str(r.get('Criticidad', ''))
        dias   = int(_n(r.get('Dias_En_Falta', 0)))
        desde  = str(r.get('Primer_Faltante', ''))
        pax    = int(_n(r.get('Pacientes_Afectados', 0)))
        nrec   = int(_n(r.get('N_Recetas', 0)))
        pend   = int(_n(r.get('Cant_Demanda_Activa', 0)))
        saa    = int(_n(r.get('Stock_AA_Total', 0)))
        sbf    = int(_n(r.get('Stock_BODEGA_FARMACOS', 0)))
        accion = str(r.get('Accion_Sugerida', ''))
        en_lista = 'SI' if norm_erp(med) in manual else '—'
        rows.append(((med, en_lista, dias, desde, pax, nrec, pend, saa, sbf, accion), crit))
    return rows


def write_h4c(ws, rows, hoy):
    """Hoja Faltantes_60D: faltantes persistentes de los últimos 60 días."""
    _titulo(ws,
        f'FALTANTES PERSISTENTES AT ABIERTA — últimos 60 días  ·  {hoy.strftime("%d/%m/%Y")}',
        len(HDRS4C))
    _subtit(ws,
        'Medicamentos con demanda pendiente en el mostrador de Atención Abierta durante los últimos '
        '60 días que SIGUEN sin resolverse hoy (última receta sin cubrir dentro de los últimos 15 '
        'días) | Lista Manual = SI si estaba en la detección manual de la QF | Días en Falta = días '
        'desde la primera receta sin cubrir | Se basa en la DEMANDA pendiente, no en el stock '
        'reportado: "Stock AA Reportado > 0" con acción REVISAR indica stock que no se despacha '
        '(posible fantasma) — verificar físico en mesón',
        len(HDRS4C), height=60)
    _hdr(ws, 3, HDRS4C)
    for i, (vals, crit) in enumerate(rows, 4):
        _fila_crit(ws, i, vals, crit, {2, 3, 5, 6, 7, 8, 9})
        # columna Lista Manual: azul/negrita si SI
        if str(vals[1]) == 'SI':
            cm = ws.cell(i, 2)
            cm.fill = _pfill('DBEAFE'); cm.font = Font(name='Arial', size=10, color='1E40AF', bold=True)
        ac  = str(vals[9])
        saa = vals[7]
        # columna Stock AA Reportado: ámbar si hay stock fantasma
        if saa > 0:
            cs = ws.cell(i, 8)
            cs.fill = _pfill('FEF3C7'); cs.font = Font(name='Arial', size=10, color='92400E', bold=True)
        c = ws.cell(i, 10)
        if 'COMPRA URGENTE' in ac:
            c.fill = _pfill('FFDAD6'); c.font = Font(name='Arial', size=10, color='9B1C1C', bold=True)
        elif 'REVISAR' in ac:
            c.fill = _pfill('FEF3C7'); c.font = Font(name='Arial', size=10, color='92400E', bold=True)
        elif 'TRASPASAR' in ac:
            c.fill = _pfill('F0FDF4'); c.font = Font(name='Arial', size=10, color='166534')
    ws.freeze_panes = 'A4'
    if rows:
        last = 3 + len(rows)
        ws.auto_filter.ref = f'A3:{get_column_letter(len(HDRS4C))}{last}'
        _totals(ws, 4, last, len(HDRS4C), {5, 6, 7})
        ws.print_area = f'A1:{get_column_letter(len(HDRS4C))}{last + 1}'
    else:
        ws.cell(4, 1, 'Sin faltantes persistentes en los últimos 60 días.')
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'


# ─────────────── main ────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description='Planilla simplificada Pedidos AA — 4 hojas, sin IA.')
    ap.add_argument('--forzar-dialisis', action='store_true',
                    help='Genera hoja Diálisis aunque no sea S3')
    ap.add_argument('--todos', action='store_true',
                    help='Incluye meds sin necesidad de pedido en Farm_Bod '
                         '(Bod_Farmacos siempre lista el universo completo)')
    args = ap.parse_args()

    hoy  = dt.date.today()
    fer  = _feriados()
    mae  = _maestro()
    sem  = _semana(hoy)
    def_ = _dias_ef(hoy, fer)
    ndia = ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom'][hoy.weekday()]

    print(f'HOY = {hoy} ({ndia}) | S{sem} | dias hab. restantes: {def_}')
    print(f'Maestro: {os.path.basename(mae)}\n')

    data = _leer(mae)

    # fe_map: Medicamento → Unidades_Caja (ICP CENABAST desde SGLI_Estres)
    fe_map = {}
    if len(data['sgli']):
        for _, r in data['sgli'].iterrows():
            m = str(r.get('Medicamento', '')).strip()
            fe_map[m] = int(_n(r.get('Unidades_Caja', 1))) or 1

    dc, r2 = calc_h2(data['bod'], fe_map, hoy, fer)
    # rep_h2_map: lo que la hoja Bod_Farmacos ya trae para cada med (CDL combinado,
    # incluye diálisis) — se usa para netear el pedido urgente de Farm_Bod y el de
    # Diálisis a Bod.Fármacos, y no pedir dos veces contra el mismo déficit.
    rep_h2_map = {v[0]: v[7] for v in r2}
    r1 = calc_h1(data['farm'], fe_map, def_, args.todos, rep_h2_map)
    # La hoja Dialisis se genera TODOS los días con el universo completo (para
    # poder consultar consumo, ej. Furosemida, aunque el stock alcance) — el
    # pedido real de diálisis sigue siendo mensual y solo se ejecuta en S3;
    # --forzar-dialisis ahora solo fuerza la marca de "semana de pedido" fuera
    # de S3, ya no controla si la hoja tiene datos (write_h3 se lo indica al
    # usuario en la nota de la hoja en vez de vaciarla).
    es_semana_pedido = args.forzar_dialisis or sem == 3
    r3 = calc_h3(data['dialmed'], fe_map, rep_h2_map)
    r4 = calc_h4(data['falt30'])
    r4b = calc_h4b(data['stock'], data['bod'])
    r4c = calc_h4c(data['falt60'], _lista_manual())

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = 'Farm_Bod'
    write_h1(ws1,                             r1, def_, hoy, sem)
    write_h2(wb.create_sheet('Bod_Farmacos'), r2, hoy, sem, dc)
    write_h3(wb.create_sheet('Dialisis'),     r3, hoy, sem, es_semana_pedido)
    write_h4(wb.create_sheet('Faltantes_AA'), r4, hoy)
    write_h4b(wb.create_sheet('Por_Agotarse'), r4b, hoy)
    write_h4c(wb.create_sheet('Faltantes_60D'), r4c, hoy)

    sal = os.path.join(WORK_DIR,
        f'Pedido_Fusion_AA_{hoy.strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(sal)

    # índices: h1=(med,crit,sfarm,cob_actual,cdl,ud,accion1,accion2)                  → ud@5
    #          h2=(med,crit,sbod,cob_bod,sfarm,sbfarm,req_ciclo,rep,accion)          → rep@7
    #          h3=(med,fe,mensual,cob_farm,apfarm,cob_bod,apbod,obs)                 → apfarm@4, apbod@6
    n1 = sum(1 for v in r1 if v[5] > 0)
    n2 = sum(1 for v in r2 if v[7] > 0)
    n3 = sum(1 for v in r3 if (v[4] + v[6]) > 0)
    print(f'Farm->Bod       : {len(r1)} meds ({n1} con pedido)')
    print(f'Bod->Farmacos   : {len(r2)} meds ({n2} con reposicion)')
    print(f'Dialisis        : {len(r3)} meds ({n3} con faltante)'
          f'{"  [S3 — semana de pedido]" if es_semana_pedido else "  [solo consulta, pedido real en S3]"}')
    print(f'Faltantes AA 30d: {len(r4)} meds sin poder despachar en Atencion Abierta')
    print(f'Por agotarse    : {len(r4b)} meds con Bodega AA en 0 y cobertura '
          f'farmacia <= {UMBRAL_PREQUIEBRE} dias')
    print(f'Faltantes 60d   : {len(r4c)} meds con faltante persistente vigente')
    print(f'\nExcel: {os.path.basename(sal)}')


if __name__ == '__main__':
    main()
