#!/usr/bin/env python3
"""
centinela_reporte.py  —  Reporte Medicamentos Centinela (Campaña Invierno 2026)
═══════════════════════════════════════════════════════════════════════════════
Auto-detecta los archivos de recetas (todos los bloques informe_completo*.csv)
y stock (reporte_de_stock_*.xlsx) en la carpeta del proyecto y genera el PDF
del reporte semanal para el MINSAL.

Uso:
  py centinela_reporte.py                  # auto-detección
  py centinela_reporte.py --csv ruta.csv --xlsx ruta.xlsx
  py centinela_reporte.py --semana 25      # fuerza semana epidemiológica
"""
import sys, math, json, datetime, warnings, re, argparse
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

MAESTRO_DIR = Path(__file__).parent

# verificar_frescura(): blindaje compartido contra datos auto-detectados
# desactualizados (incidente S.52, ver utils_aa.py para el detalle).
sys.path.insert(0, str(MAESTRO_DIR))
from utils_aa import verificar_frescura

# ── HOMOLOGACIÓN (ZGEN como clave primaria) ──────────────────────────────────
FM = [
    {"minsal": "SALBUTAMOL 5MG/ML SOL.NEBU FRA 20ML",
     "zgen": "100002314", "xlsx": "SALBUTAMOL SOL. P/NEBULIZAR 5 MG/ML FC X 20 ML",
     "ci": ["SALBUTAMOL"], "ce": ["100 MCG", "INHALADOR"], "na": False},
    {"minsal": "FLUTICAS/SALMET 125/25 UG/DO X 120DO FRA",
     "zgen": "100004610", "xlsx": "SALMETEROL /FLUTICASONA 125 MG/25 MG INHALADOR X 120 DO",
     "ci": ["SALMETEROL", "125"], "ce": ["250"], "na": False},
    {"minsal": "FLUTICAS/SALMET 250/25MCG 120A150DO FRA",
     "zgen": "100001962", "xlsx": "SALMETEROL /FLUTICASONA 250 MG/25 MG INH UD",
     "ci": ["SALMETEROL", "250"], "ce": ["125"], "na": False},
    {"minsal": "PREDNISONA 20MG/5ML SUSP. FRA 60 ML",
     "zgen": "100002019", "xlsx": "PREDNISONA 20 MG/5  ML SUSPENSION ORAL",
     "ci": ["PREDNISONA", "SUSPENSION"], "ce": ["COMPRIMIDO"], "na": False},
    {"minsal": "OSELTAMIVIR 75 MG CP",
     "zgen": "100001108", "xlsx": "OSELTAMIVIR 75 MG CAPSULA",
     "ci": ["OSELTAMIVIR", "75"], "ce": ["MG/ML"], "na": False},
    {"minsal": "OSELTAMIVIR 12 MG/ML FRASCO",
     "zgen": "100001107", "xlsx": "OSELTAMIVIR 12 MG/ML JAR 75 ML",
     "ci": ["OSELTAMIVIR", "12"], "ce": [], "na": False},
    {"minsal": "OSELTAMIVIR 6 MG/ML FRASCO",
     "zgen": None, "xlsx": None, "ci": [], "ce": [], "na": True},
    {"minsal": "IPRATROPIO 0,25MG/ML SOL P/NEB FRA 20ML",
     "zgen": "100000829", "xlsx": "BROMURO DE IPRATROPIO 0,25 MG/ML SOL. P/NEBULIZAR FC",
     "ci": ["IPRATROPIO", "0,25"], "ce": ["20 MCG", "AEROSOL", "FENOTEROL"], "na": False},
]


def detect_csv(nombre):
    if not nombre:
        return None
    n = str(nombre).upper()
    for f in FM:
        if f["na"] or not f["ci"]:
            continue
        if all(k.upper() in n for k in f["ci"]) and not any(e.upper() in n for e in f["ce"]):
            return f
    return None


def _domingo_semana1(year):
    """Domingo de inicio de la semana epidemiológica 1 (MINSAL/SaludResponde):
    la semana domingo-sábado que contiene el 4 de enero."""
    d = datetime.date(year, 1, 4)
    return d - datetime.timedelta(days=(d.weekday() + 1) % 7)


def iso_week(fecha):
    """Semana epidemiológica oficial (NO es semana ISO 8601 pese al nombre,
    se conserva por compatibilidad). MINSAL define la semana epidemiológica
    como domingo 00:00 a sábado 24:00 — un día antes que la semana ISO
    (lunes-domingo), lo que desplaza en -1 el número de semana de Python
    durante casi todo el año. Ver Calendario Epidemiológico SaludResponde."""
    try:
        ts = pd.to_datetime(fecha, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return None
        d = ts.date()
        dom1 = _domingo_semana1(d.year)
        if d < dom1:
            anio = d.year - 1
            dom1 = _domingo_semana1(anio)
        else:
            dom1_sig = _domingo_semana1(d.year + 1)
            if d >= dom1_sig:
                anio, dom1 = d.year + 1, dom1_sig
            else:
                anio = d.year
        semana = (d - dom1).days // 7 + 1
        return (anio, semana)
    except Exception:
        return None


def _encontrar_archivos():
    """Localiza automáticamente los archivos de datos en MAESTRO_DIR."""
    csvs = sorted(MAESTRO_DIR.glob("informe_completo_recetas*.csv"),
                  key=lambda p: p.stat().st_mtime)
    xlsxs = sorted(MAESTRO_DIR.glob("reporte_de_stock_*.xlsx"),
                   key=lambda p: p.stat().st_mtime)
    if not csvs:
        raise FileNotFoundError(
            "No se encontró informe_completo_recetas*.csv en " + str(MAESTRO_DIR)
        )
    if not xlsxs:
        raise FileNotFoundError(
            "No se encontró reporte_de_stock_*.xlsx en " + str(MAESTRO_DIR)
        )
    return csvs, xlsxs[-1]


def leer_recetas(csv_paths):
    """Lee y concatena todos los bloques CSV de recetas."""
    frames = []
    for p in csv_paths:
        try:
            df = pd.read_csv(
                p, encoding="latin-1", sep=None, engine="python",
                usecols=["Prescripción", "Estado Prescripción", "Cantidad Entregada",
                         "Fecha Entrega Receta", "ID Receta Detalle", "Número Receta"],
            )
            frames.append(df)
        except Exception as e:
            print(f"  [AVISO] No se pudo leer {p.name}: {e}")
    if not frames:
        raise RuntimeError("No se pudo leer ningún CSV de recetas.")
    df = pd.concat(frames, ignore_index=True)
    total_bruto = len(df)
    df = df[df["Estado Prescripción"].str.upper().str.strip() == "ENTREGADO"].copy()
    # SSASUR repite el mismo "ID Receta Detalle" en varias filas-snapshot del
    # mismo trámite (p.ej. una con Cantidad Entregada=0 cuando aún estaba
    # pendiente de dispensar, y otra posterior ya con la cantidad real). Si
    # drop_duplicates se queda con la primera por orden de archivo, descarta
    # en silencio el egreso real cuando el snapshot vacío aparece antes.
    # Ordenamos por cantidad entregada descendente para quedarnos con la fila
    # que sí refleja la entrega (y su fecha real), no la del trámite vacío.
    df["Cantidad Entregada"] = pd.to_numeric(df["Cantidad Entregada"], errors="coerce").fillna(0)
    df = df.sort_values("Cantidad Entregada", ascending=False)
    df = df.drop_duplicates(subset=["ID Receta Detalle"], keep="first")
    df["semana"] = df["Fecha Entrega Receta"].apply(iso_week)
    df = df.dropna(subset=["semana"])
    if df.empty:
        raise RuntimeError("No hay entregas con datos de fecha/semana válidos tras los filtros aplicados.")
    # Filtrar al año epidemiológico más reciente para evitar colisión S.52/2025 vs S.25/2026
    max_epi_year = max(t[0] for t in df["semana"])
    df = df[df["semana"].apply(lambda t: t[0] == max_epi_year)].copy()
    df["semana"] = df["semana"].apply(lambda t: t[1]).astype(int)
    num_recetas = df["Número Receta"].nunique() if "Número Receta" in df.columns else 0
    return df, total_bruto, num_recetas


def leer_stock(xlsx_path):
    """Lee el reporte de stock y extrae cantidades por ZGEN."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    wb.close()

    h_row = c_desc = c_zgen = c_cant = -1
    for i, row in enumerate(raw[:8]):
        row_s = [str(c or "").strip() for c in row]
        di = next((j for j, v in enumerate(row_s) if v.lower().startswith("descripci")), -1)
        ci = next((j for j, v in enumerate(row_s) if v.lower() == "cantidad"), -1)
        if di >= 0 and ci >= 0:
            h_row, c_desc, c_cant = i, di, ci
            c_zgen = next((j for j, v in enumerate(row_s) if "zgen" in v.lower()), 3)
            break

    xlsx_fecha = ""
    for row in raw[:3]:
        txt = " ".join(str(c or "") for c in row)
        m = re.search(r"(\d{2}/\d{2}/\d{4})", txt)
        if m:
            xlsx_fecha = m.group(1)
            break

    zgen_idx = {f["zgen"]: f for f in FM if not f["na"] and f["zgen"]}
    stock = {f["zgen"]: 0 for f in FM if not f["na"] and f["zgen"]}

    for row in raw[h_row + 1:]:
        if not row or len(row) <= max(c_desc, c_cant):
            continue
        desc = str(row[c_desc] or "").strip()
        zgen = str(row[c_zgen] or "").strip() if c_zgen >= 0 else ""
        try:
            v = float(str(row[c_cant] or "0").replace(",", "."))
            cant = v if (v == int(v) and 0 < v < 100_000) else 0
        except Exception:
            cant = 0
        if cant == 0:
            continue
        f = zgen_idx.get(zgen)
        if not f:
            f = next((x for x in FM if not x["na"] and x["xlsx"] and
                      (desc.upper() == x["xlsx"].upper() or
                       (x["ci"] and all(k.upper() in desc.upper() for k in x["ci"]) and
                        not any(e.upper() in desc.upper() for e in x["ce"])))), None)
        if f and f["zgen"]:
            stock[f["zgen"]] = stock.get(f["zgen"], 0) + cant
    return stock, xlsx_fecha


def proyeccion_inteligente(eg_dict, sem_min, sem_max):
    todas_sems = list(range(sem_min, sem_max + 1))
    valores = [eg_dict.get(s, 0) for s in todas_sems]
    activos = [(s, eg_dict[s]) for s in todas_sems if eg_dict.get(s, 0) > 0]
    if not activos:
        return 0, "Sin movimiento", "0 egresos en el período", []
    vals_activos = [v for _, v in activos]
    n = len(vals_activos)
    media = np.mean(vals_activos)
    std = np.std(vals_activos) if n > 1 else 0
    cv = std / media if media > 0 else 0
    x = np.array(todas_sems)
    y = np.array(valores, dtype=float)
    slope = np.polyfit(x, y, 1)[0] if len(x) >= 2 else 0

    if cv < 0.25:
        return math.ceil(media), "Promedio simple", f"CV={cv:.2f} → patrón estable ({n} sem. activas)", vals_activos
    if slope > 1.0:
        recientes = vals_activos[-3:] if n >= 3 else vals_activos
        pesos = list(range(1, len(recientes) + 1))
        wma = sum(v * p for v, p in zip(recientes, pesos)) / sum(pesos)
        return math.ceil(wma), "Prom. móvil ponderado", f"Tendencia ↑ (slope={slope:.1f}) → últimas {len(recientes)} sem.", vals_activos
    if slope < -1.0:
        recientes = vals_activos[-2:] if n >= 2 else vals_activos
        return math.ceil(np.mean(recientes)), "Semanas recientes", f"Tendencia ↓ (slope={slope:.1f}) → últimas {len(recientes)} sem.", vals_activos
    if n <= 3:
        return math.ceil(np.median(vals_activos)), "Mediana", f"Datos escasos ({n} sem. activas) → mediana", vals_activos
    alpha = 0.4
    suav = vals_activos[0]
    for v in vals_activos[1:]:
        suav = alpha * v + (1 - alpha) * suav
    return math.ceil(suav), "Suavizado exp.", f"CV={cv:.2f} (volátil) → α=0.4, {n} sem. activas", vals_activos


def calcular(df, stock, semana_override=None):
    egresos = {f["minsal"]: {} for f in FM}
    for _, row in df.iterrows():
        f = detect_csv(row["Prescripción"])
        if not f:
            continue
        s = row["semana"]
        egresos[f["minsal"]][s] = egresos[f["minsal"]].get(s, 0) + row["Cantidad Entregada"]

    sem_min = int(df["semana"].min()) if len(df) else 0
    sem_max_datos = int(df["semana"].max()) if len(df) else 0
    srep = semana_override if semana_override else sem_max_datos
    # No usar semanas posteriores a la reportada (p.ej. la semana en curso, con
    # datos parciales) para las proyecciones: distorsionaría el promedio/tendencia.
    sem_max = min(sem_max_datos, srep) if sem_max_datos else srep

    resultados = []
    for f in FM:
        if f["na"]:
            resultados.append({**f, "stk": 0, "consumo": 0, "proy": 0, "alg": "N/A",
                                "alg_desc": "No existe en este establecimiento",
                                "var_sem": None, "var_hist": None, "alerta": "na",
                                "rot": None, "vals_activos": [], "eg_dict": {},
                                "eg_sem_ant": 0, "hist_avg": 0})
            continue

        stk = round(stock.get(f["zgen"], 0))
        eg = egresos[f["minsal"]]
        consumo = eg.get(srep, 0)
        proy, alg, alg_desc, vals_activos = proyeccion_inteligente(eg, sem_min, sem_max)
        eg_sem_ant = eg.get(srep - 1, 0)
        hist_avg = np.mean(vals_activos) if vals_activos else 0
        var_sem = round((consumo - eg_sem_ant) / eg_sem_ant * 100, 1) if eg_sem_ant > 0 else None
        var_hist = round((consumo - hist_avg) / hist_avg * 100, 1) if hist_avg > 0 else None
        rot = round(stk / proy, 1) if proy > 0 else None
        sinmov = not vals_activos
        if stk == 0 and not sinmov:    alerta = "q"
        elif sinmov:                   alerta = "w"
        elif rot is not None and rot < 4:  alerta = "a"
        elif rot is not None and rot < 8:  alerta = "m"
        else:                          alerta = "s"

        resultados.append({
            **f, "stk": stk, "consumo": int(consumo), "proy": proy,
            "alg": alg, "alg_desc": alg_desc, "var_sem": var_sem, "var_hist": var_hist,
            "alerta": alerta, "rot": rot,
            "vals_activos": [int(v) for v in vals_activos],
            "eg_dict": {int(k): int(v) for k, v in eg.items()},
            "eg_sem_ant": int(eg_sem_ant), "hist_avg": round(hist_avg, 1),
        })
    return resultados, srep, sem_min, sem_max


def generar_pdf(data, pdf_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER

    srep = data["srep"]
    xlsx_fecha = data["xlsx_fecha"]
    resultados = data["resultados"]

    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                            leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    esti = {
        "titulo":    ParagraphStyle("t",   fontSize=15, fontName="Helvetica-Bold",  spaceAfter=4, alignment=TA_CENTER),
        "subtitulo": ParagraphStyle("s",   fontSize=9,  fontName="Helvetica",       spaceAfter=2, alignment=TA_CENTER, textColor=colors.HexColor("#666666")),
        "seccion":   ParagraphStyle("sec", fontSize=9,  fontName="Helvetica-Bold",  spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#6b3fa0")),
        "normal":    ParagraphStyle("n",   fontSize=8,  fontName="Helvetica",       spaceAfter=2),
        "alerta_q":  ParagraphStyle("aq",  fontSize=8,  fontName="Helvetica-Bold",  textColor=colors.HexColor("#c42e2e")),
        "alerta_a":  ParagraphStyle("aa",  fontSize=8,  fontName="Helvetica-Bold",  textColor=colors.HexColor("#c47e2e")),
    }
    ALERTA_LABELS = {"q": "QUIEBRE", "a": "ATENCION", "m": "MODERADO", "s": "SIN RIESGO", "w": "SIN MOV.", "na": "N/A"}
    ALERTA_COLORS = {
        "q": colors.HexColor("#fdf0f0"), "a": colors.HexColor("#fdf6f0"),
        "m": colors.HexColor("#fdfaf0"), "s": colors.HexColor("#f0fdf4"),
        "w": colors.HexColor("#f5f3f0"), "na": colors.HexColor("#f5f3f0"),
    }

    story = []
    story.append(Paragraph("REPORTE MEDICAMENTOS CENTINELA", esti["titulo"]))
    story.append(Paragraph("Campaña de Invierno 2026 — Farmacia Asistencial Hospital Pitrufquén", esti["subtitulo"]))
    story.append(Paragraph(
        f"Semana epidemiológica S.{srep} &nbsp;|&nbsp; Stock al {xlsx_fecha} &nbsp;|&nbsp; "
        f"Generado {datetime.date.today().strftime('%d/%m/%Y')}",
        esti["subtitulo"]))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#6b3fa0"), spaceAfter=10))

    story.append(Paragraph(f"TABLA MINSAL — STOCK | CONSUMO | PROYECCIÓN S.{srep + 1}", esti["seccion"]))
    enc = [Paragraph(f"<b>{h}</b>", styles["Normal"]) for h in
           ["Fármaco / Presentación", "Stock", f"Consumo S.{srep}", f"Proyección S.{srep+1}",
            "Algoritmo", "Rotación", "Alerta"]]
    filas = [enc]
    row_colors = [colors.HexColor("#f5f0fc")]
    for r in resultados:
        rot_txt = f"{r['rot']} sem." if r["rot"] is not None else "—"
        filas.append([
            Paragraph(r["minsal"], styles["Normal"]),
            Paragraph(str(r["stk"]), styles["Normal"]),
            Paragraph(str(r["consumo"]) if not r["na"] else "—", styles["Normal"]),
            Paragraph(str(r["proy"]) + " u." if not r["na"] else "—", styles["Normal"]),
            Paragraph(r["alg"], styles["Normal"]),
            Paragraph(rot_txt, styles["Normal"]),
            Paragraph(ALERTA_LABELS.get(r["alerta"], ""), styles["Normal"]),
        ])
        row_colors.append(ALERTA_COLORS.get(r["alerta"], colors.white))

    col_w = [5.8 * cm, 1.3 * cm, 1.6 * cm, 1.8 * cm, 2.8 * cm, 1.5 * cm, 2 * cm]
    tbl = Table(filas, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTSIZE",     (0, 0), (-1, -1), 7.5),
        ("FONTNAME",     (0, 0), (-1,  0), "Helvetica-Bold"),
        ("BACKGROUND",   (0, 0), (-1,  0), colors.HexColor("#6b3fa0")),
        ("TEXTCOLOR",    (0, 0), (-1,  0), colors.white),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#d4cfc7")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ] + [("BACKGROUND", (0, i + 1), (-1, i + 1), row_colors[i + 1]) for i in range(len(resultados))]))
    story.append(tbl)
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("ANÁLISIS DE VARIACIONES", esti["seccion"]))
    var_data = [[Paragraph(f"<b>{h}</b>", styles["Normal"]) for h in
                 ["Fármaco", f"Consumo S.{srep}", f"Sem. anterior S.{srep-1}",
                  "Var. sem.", "Prom. histórico", "Var. vs hist."]]]
    for r in resultados:
        if r["na"]:
            continue
        def fmt_var(v):
            if v is None:
                return Paragraph("—", styles["Normal"])
            signo = "+" if v >= 0 else ""
            col = "#c42e2e" if v < -30 else "#c47e2e" if v < 0 else "#2e8c4a" if v > 30 else "#1a1714"
            return Paragraph(f'<font color="{col}"><b>{signo}{v}%</b></font>', styles["Normal"])
        var_data.append([
            Paragraph(r["minsal"][:30], styles["Normal"]),
            Paragraph(str(r["consumo"]), styles["Normal"]),
            Paragraph(str(r["eg_sem_ant"]), styles["Normal"]),
            fmt_var(r["var_sem"]),
            Paragraph(str(r["hist_avg"]), styles["Normal"]),
            fmt_var(r["var_hist"]),
        ])
    tbl2 = Table(var_data, colWidths=[4.8 * cm, 1.6 * cm, 2.2 * cm, 1.8 * cm, 2.2 * cm, 1.8 * cm], repeatRows=1)
    tbl2.setStyle(TableStyle([
        ("FONTSIZE",    (0, 0), (-1, -1), 7.5),
        ("FONTNAME",    (0, 0), (-1,  0), "Helvetica-Bold"),
        ("BACKGROUND",  (0, 0), (-1,  0), colors.HexColor("#6b3fa0")),
        ("TEXTCOLOR",   (0, 0), (-1,  0), colors.white),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#d4cfc7")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0, 0), (-1, -1), 4),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
    ]))
    story.append(tbl2)
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("HALLAZGOS Y ACCIONES RECOMENDADAS", esti["seccion"]))
    quiebres = [r for r in resultados if r["alerta"] == "q"]
    atencion = [r for r in resultados if r["alerta"] == "a"]
    if quiebres:
        story.append(Paragraph("QUIEBRES ACTIVOS — Acción inmediata requerida:", esti["alerta_q"]))
        for r in quiebres:
            story.append(Paragraph(
                f"  * {r['minsal']} — Stock: 0 u. | Consumo S.{srep}: {r['consumo']} u. | Proyección: {r['proy']} u./sem.",
                esti["alerta_q"]))
    if atencion:
        story.append(Paragraph("RIESGO DE QUIEBRE (&lt;4 semanas):", esti["alerta_a"]))
        for r in atencion:
            story.append(Paragraph(
                f"  * {r['minsal']} — Stock: {r['stk']} u. | Rotación: {r['rot']} sem.",
                esti["alerta_a"]))
    if not quiebres and not atencion:
        story.append(Paragraph("Sin alertas críticas en este período.", esti["normal"]))

    story.append(Spacer(1, 0.6 * cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#d4cfc7")))
    story.append(Paragraph(
        f"Farmacia Asistencial Hospital Pitrufquén — Servicio Salud Araucanía Sur &nbsp;|&nbsp; "
        f"Datos: {data['total_bruto']:,} registros procesados, {data['num_recetas']:,} recetas únicas",
        esti["subtitulo"]))

    doc.build(story)


def main():
    parser = argparse.ArgumentParser(description="Reporte Centinela Campaña Invierno 2026")
    parser.add_argument("--csv",   nargs="+", help="Ruta(s) al CSV de recetas (opcional, auto-detecta si no se indica)")
    parser.add_argument("--xlsx",  help="Ruta al XLSX de stock (opcional, auto-detecta si no se indica)")
    parser.add_argument("--semana", type=int,
                         help="Forzar semana epidemiológica (por defecto: la última semana "
                              "COMPLETA, es decir la anterior a la actual)")
    parser.add_argument("--no-pause", action="store_true")
    args = parser.parse_args()

    print()
    print("═" * 62)
    print("  CENTINELA  ·  Reporte Medicamentos Campaña Invierno 2026")
    print("═" * 62)

    # ── Localizar archivos ────────────────────────────────────────────
    if args.csv:
        csv_paths = [Path(p) for p in args.csv]
    else:
        csv_paths, _ = _encontrar_archivos()
        print(f"  CSV: {len(csv_paths)} bloque(s) → {', '.join(p.name for p in csv_paths)}")

    if args.xlsx:
        xlsx_path = Path(args.xlsx)
    else:
        _, xlsx_path = _encontrar_archivos()
        print(f"  XLSX stock: {xlsx_path.name}")

    # ── Procesar ──────────────────────────────────────────────────────
    print("\n  Leyendo recetas...")
    df, total_bruto, num_recetas = leer_recetas(csv_paths)
    print(f"    {total_bruto:,} registros brutos → {len(df):,} entregas únicas · {num_recetas:,} recetas")

    if not args.csv:
        fecha_max_recetas = pd.to_datetime(df["Fecha Entrega Receta"], dayfirst=True, errors="coerce").max()
        fecha_max_recetas = fecha_max_recetas.date() if pd.notna(fecha_max_recetas) else None
        verificar_frescura(fecha_max_recetas, "sábana de recetas (auto-detectada)")

    print("  Leyendo stock...")
    stock, xlsx_fecha = leer_stock(xlsx_path)
    print(f"    Stock al {xlsx_fecha}")

    if not args.xlsx:
        fecha_stock = datetime.datetime.strptime(xlsx_fecha, "%d/%m/%Y").date() if xlsx_fecha else None
        verificar_frescura(fecha_stock, "reporte de stock (auto-detectado)")

    # El lunes que se genera el reporte, la semana en curso recién empieza
    # (datos parciales). Por defecto reportamos la última semana COMPLETA:
    # la que contiene la fecha de hace 7 días, sea cual sea el día de hoy.
    if args.semana:
        semana_objetivo = args.semana
    else:
        _anio_ant, semana_objetivo = iso_week(datetime.date.today() - datetime.timedelta(days=7))
        print(f"  Semana no indicada (--semana): se usa la última semana completa → S.{semana_objetivo}")

    print("  Calculando proyecciones...")
    resultados, srep, sem_min, sem_max = calcular(df, stock, semana_objetivo)
    print(f"    Semana epidemiológica reportada: S.{srep} (rango S.{sem_min}–S.{sem_max})")

    # ── Guardar JSON ──────────────────────────────────────────────────
    out_dir = MAESTRO_DIR / "Centinela_Reportes" / f"S{srep}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_json = out_dir / f"centinela_S{srep}.json"
    data = {
        "srep": srep, "sem_min": sem_min, "sem_max": sem_max,
        "xlsx_fecha": xlsx_fecha, "total_bruto": total_bruto,
        "num_recetas": num_recetas, "resultados": resultados,
    }
    with open(out_json, "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)

    # ── Generar PDF ───────────────────────────────────────────────────
    try:
        pdf_path = out_dir / f"centinela_S{srep}.pdf"
        print(f"  Generando PDF: {pdf_path.name}...")
        generar_pdf(data, pdf_path)
        print(f"    ✓ {pdf_path.name}  ({pdf_path.stat().st_size // 1024:,} KB)")
    except ImportError:
        print("  [AVISO] reportlab no instalado — omitiendo PDF.")
        print("          Ejecuta: py -m pip install reportlab")
        pdf_path = None

    # ── Resumen ───────────────────────────────────────────────────────
    q = [r["minsal"] for r in resultados if r["alerta"] == "q"]
    a = [r["minsal"] for r in resultados if r["alerta"] == "a"]
    print()
    print("═" * 62)
    print(f"  RESUMEN — Semana S.{srep}")
    for r in resultados:
        ico = {"q": "🔴 QUIEBRE ", "a": "🟠 ATENCION", "m": "🟡 MODERADO",
               "s": "🟢 SIN RIESGO", "w": "⚫ SIN MOV.", "na": "  N/A      "}.get(r["alerta"], "   ")
        rot = f"  rotación {r['rot']} sem." if r["rot"] is not None else ""
        print(f"  {ico}  {r['minsal'][:42]:<42}  stk={r['stk']:>5}  proy={r['proy']:>4}{rot}")
    if q:
        print(f"\n  ⚠ QUIEBRES: {', '.join(q)}")
    if a:
        print(f"  ⚠ ATENCIÓN: {', '.join(a)}")
    print("═" * 62)
    if pdf_path:
        print(f"  PDF → {pdf_path}")
    print()

    if not args.no_pause:
        try:
            input("  Presiona Enter para cerrar...")
        except EOFError:
            pass


if __name__ == "__main__":
    main()
