#!/usr/bin/env python3
"""
cruce_gt.py — Cruza el reporte de Gestión Territorial (modalidad de despacho) con
el histórico de recetas (informe_completo_recetas*.csv) para clasificar, por
receta, sus medicamentos en:

  · REFRIGERADOS  (insulinas / cadena de frío)        + cantidad recetada
  · CONTROLADOS   (Tipo Receta == CONTROLADA en SSASUR) + cantidad recetada
  · PENDIENTES    (Cantidad Pendiente > 0)             + cantidad pendiente

Join por "Nº Receta" (GT) == "Número Receta" (histórico). Deduplica las líneas de
prescripción por "ID Receta Detalle". Produce:
  · resumen por consola
  · <salida>/gt_enriquecido.json   (formato registros del skill gestion-territorial)
  · <salida>/Cruce_GT_Clasificacion.xlsx  (3 categorías con cantidad, por destino)

Uso:
  py cruce_gt.py <reporteGT.xlsx> --salida ./out_gt [--hist-glob "informe_completo_recetas*.csv"]
"""
import argparse, csv, glob, json, os, re, sys, tempfile, unicodedata
from collections import OrderedDict, defaultdict

for _s in (sys.stdout, sys.stderr):
    try: _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception: pass

MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Listas de clasificación ────────────────────────────────────────────────────
TIPOS_INSULINA = [("GLARGINA","Glargina"),("ASPART","Asparta"),("GLULISINA","Glulisina"),
                  ("CRISTALINA","Cristalina"),("LISPRO","Lispro"),("DEGLUDEC","Degludec"),
                  ("DETEMIR","Detemir"),("NPH","NPH")]
# Otros refrigerados frecuentes (cadena de frío) además de insulinas:
REFRIG_OTROS = ["ANALOGO GLP", "DULAGLUTIDA", "SEMAGLUTIDA", "LIRAGLUTIDA",
                "ENOXAPARINA", "VACUNA", "TOXINA BOTULINICA", "OCTREOTIDA",
                "FILGRASTIM", "ERITROPOYETINA", "EPOETINA", "SOMATROPINA",
                "TERIPARATIDA", "INTERFERON"]
# Controlados — INN extraídos de reporte_de_stock_20260618093729.xlsx (columna Descripción).
# Fuente: stock filtrado a psicotrópicos/opioides de Farmacia Hospital de Pitrufquén (18-06-2026).
# SENSIDISCO ERTAPENEM excluido (disco antibiótico, no psicotrópico).
CONTROL_INN = {
    "ALPRAZOLAM","BUPRENORFINA","CLOBAZAM","CLONAZEPAM","CODEINA","DIAZEPAM",
    "FENOBARBITAL","FENTANILO","KETAMINA","LISDEXANFETAMINA","LORAZEPAM",
    "METADONA","METILFENIDATO","MIDAZOLAM","MORFINA","OXICODONA",
    "PETIDINA","REMIFENTANILO","TAPENTADOL","ZOLPIDEM",
}


def _key(h):
    h = unicodedata.normalize("NFKD", str(h or "")).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", h)


def insulina_label(prod):
    u = str(prod or "").upper()
    for kw, lab in TIPOS_INSULINA:
        if kw in u: return f"Insulina {lab}"
    if "INSULINA" in u and "JERINGA" not in u and "AGUJA" not in u:
        return "Insulina"
    return None


def es_refrigerado(prod):
    lab = insulina_label(prod)
    if lab: return lab
    u = str(prod or "").upper()
    for kw in REFRIG_OTROS:
        if kw in u:
            return prod.strip().title()
    return None


def es_controlado_oficial(prod):
    u = str(prod or "").upper()
    for inn in CONTROL_INN:
        if inn in u: return True
    return False


def _num(v):
    try: return int(float(str(v).replace(",", ".").strip() or 0))
    except Exception: return 0


# ── Lectura del reporte GT (.xlsx) ─────────────────────────────────────────────
def leer_reporte_gt(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()
    # fila de encabezado = primera con >=5 celdas no vacías
    hi = next((i for i, r in enumerate(rows) if sum(1 for c in r if c not in (None, "")) >= 5), None)
    if hi is None:
        raise ValueError(f"No se encontró fila de encabezado válida en {path} (se esperaban >=5 columnas con datos)")
    hdr = [str(c).strip() if c is not None else "" for c in rows[hi]]
    K = {_key(h): i for i, h in enumerate(hdr)}

    def col(*keys, default=None):
        for k in keys:
            if k in K: return K[k]
        return default

    c_rec  = col("nreceta", "noreceta", "numeroreceta", default=0)
    c_pac  = col("paciente")
    c_run  = col("runpaciente", "run", "rut")
    c_edad = col("edad")
    c_dir  = col("direccion")
    c_com  = col("comuna")
    c_tel  = col("telefono", "fono")
    c_org  = col("estaborigen")
    c_dst  = col("estabdestino")
    c_fen  = col("fechaentrega")
    c_per  = col("periodoreceta", "periodo")
    c_esp  = col("especialidad")
    c_np   = col("numeroprescripciones", "nprescripciones")
    c_tret = col("tiporetiro")

    def g(r, i):
        return str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else ""

    regs = OrderedDict()
    for r in rows[hi+1:]:
        if not r:
            continue
        rec = g(r, c_rec)
        if not rec or rec.upper() in ("TOTAL", "NONE"):
            continue
        if rec not in regs:
            edad = re.sub(r"\D", "", g(r, c_edad))
            regs[rec] = {
                "receta": rec, "paciente": g(r, c_pac), "run": g(r, c_run),
                "edad": int(edad) if edad else None, "direccion": g(r, c_dir),
                "comuna": g(r, c_com), "telefono": g(r, c_tel),
                "estab_origen": g(r, c_org) or "Pitrufquén Hosp.",
                "estab_destino": g(r, c_dst), "fecha_entrega_rep": g(r, c_fen),
                "periodo": g(r, c_per), "especialidad": g(r, c_esp),
                "n_presc": _num(g(r, c_np)),
                "ventanilla": g(r, c_tret).upper() == "PACIENTE",
            }
    return regs, hdr


# ── Cruce con el histórico ─────────────────────────────────────────────────────
def cruzar_historico(recetas_set, archivos):
    """Devuelve {n_receta: {"tipo_receta","lineas":[{id,prod,recetada,pendiente}]}}.
    Deduplica líneas por ID Receta Detalle."""
    det = defaultdict(lambda: {"tipo_receta": "", "lineas": OrderedDict()})
    vistos = set()
    for fp in archivos:
        try:
            f = open(fp, encoding="latin-1", newline="")
        except Exception as e:
            print(f"  [aviso] no pude abrir {os.path.basename(fp)}: {e}")
            continue
        with f:
            rd = csv.reader(f, delimiter=";")
            try:
                hdr = next(rd)
            except StopIteration:
                continue
            K = {_key(h): i for i, h in enumerate(hdr)}
            ix_rec = K.get("numeroreceta"); ix_tr = K.get("tiporeceta")
            ix_id = K.get("idrecetadetalle")
            ix_pre = K.get("prescripcion")
            ix_cr = K.get("cantidadrecetada"); ix_cp = K.get("cantidadpendiente")
            if ix_rec is None or ix_pre is None:
                print(f"  [aviso] {os.path.basename(fp)} sin columnas esperadas — omito")
                continue
            n = len(hdr)
            for row in rd:
                if len(row) < n:
                    continue
                rec = (row[ix_rec] or "").strip()
                if rec not in recetas_set:
                    continue
                idd = (row[ix_id] or "").strip() if ix_id is not None else ""
                vk = idd or f"{rec}|{row[ix_pre]}"
                if vk in vistos:
                    continue
                vistos.add(vk)
                d = det[rec]
                if not d["tipo_receta"] and ix_tr is not None: d["tipo_receta"] = (row[ix_tr] or "").strip()
                d["lineas"][vk] = {
                    "prod": (row[ix_pre] or "").strip(),
                    "recetada": _num(row[ix_cr]) if ix_cr is not None else 0,
                    "pendiente": _num(row[ix_cp]) if ix_cp is not None else 0,
                }
    return det


def clasificar(reg, d):
    """Rellena refrigerado/controlado/pendiente (texto con cantidad) en reg."""
    refri, control, pend = [], [], []
    tipo_controlada = (d["tipo_receta"] or "").upper() == "CONTROLADA"
    for ln in d["lineas"].values():
        prod = ln["prod"]
        if not prod:
            continue
        rlab = es_refrigerado(prod)
        if rlab:
            refri.append(f"{rlab} x{ln['recetada']}")
        # controlado: receta CONTROLADA y producto que matchea lista, o (si ninguno
        # matchea) todos los fármacos de una receta CONTROLADA
        if es_controlado_oficial(prod):
            control.append(f"{prod.title()} x{ln['recetada']}")
        # pendiente: cantidad pendiente > 0
        if ln["pendiente"] > 0:
            pend.append(f"{prod.title()} x{ln['pendiente']}")
    if tipo_controlada and not control:
        # receta marcada CONTROLADA pero sin match de palabra clave → listar fármacos no insulina
        for ln in d["lineas"].values():
            if ln["prod"] and not es_refrigerado(ln["prod"]):
                control.append(f"{ln['prod'].title()} x{ln['recetada']}")
    reg["refrigerado"] = "; ".join(dict.fromkeys(refri))
    reg["controlado"]  = "; ".join(dict.fromkeys(control))
    reg["pendiente"]   = "; ".join(dict.fromkeys(pend))
    reg["tipo_receta"] = d["tipo_receta"]
    reg["_en_historico"] = bool(d["lineas"])
    return reg


# ── Salidas ────────────────────────────────────────────────────────────────────
def escribir_excel(regs, ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color="BFBFBF"); bd = Border(thin, thin, thin, thin)
    wb = Workbook(); ws = wb.active; ws.title = "Cruce GT"
    cols = ["Estab. Destino","Nº Receta","Paciente","RUN","Especialidad","Período",
            "Tipo Receta","Refrigerados (cant.)","Controlados (cant.)","Pendientes (cant.)","En histórico"]
    widths = [22,11,30,14,24,8,14,34,34,34,11]
    for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = w
    for c, h in enumerate(cols, 1):
        cell = ws.cell(1, c, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E5496"); cell.border = bd
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    r = 2
    for g in sorted(regs.values(), key=lambda x: (x["estab_destino"], x["paciente"])):
        vals = [g["estab_destino"], g["receta"], g["paciente"], g["run"], g["especialidad"],
                g["periodo"], g.get("tipo_receta",""), g.get("refrigerado",""),
                g.get("controlado",""), g.get("pendiente",""),
                "Sí" if g.get("_en_historico") else "NO"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.border = bd
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(size=9)
            if c == 8 and v: cell.font = Font(size=9, bold=True, color="1F6F3D")   # refrig verde
            if c == 9 and v: cell.font = Font(size=9, bold=True, color="C00000")   # control rojo
            if c == 10 and v: cell.font = Font(size=9, bold=True, color="C55A11")  # pend naranjo
        if not g.get("_en_historico"):
            for c in range(1, len(cols)+1): ws.cell(r, c).fill = PatternFill("solid", fgColor="FFF2CC")
        r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r-1}"
    wb.save(ruta)


def _buscar_generar_py():
    """Busca generar.py del skill gestion-territorial-pitrufquen.
    Prioridad: copia local skill_gt/ > AppData (solo accesible fuera del sandbox)."""
    # 1) Copia local incluida en el repositorio maestro (siempre accesible)
    local = os.path.join(MAESTRO_DIR, "skill_gt", "scripts", "generar.py")
    if os.path.exists(local):
        return local
    return None


GT_DIR_DEFAULT = os.path.join(os.path.dirname(MAESTRO_DIR), "04_Farmacia_Gestion_Territorial")


def _recetas_en_gt_previos(reporte_actual, gt_dir):
    """Carga los Nº Receta de todos los GT anteriores (excluye el actual).
    Retorna un set de números ya procesados."""
    from openpyxl import load_workbook
    ya = set()
    patron = os.path.join(gt_dir, "reporteGestionTerritorial_*.xlsx")
    for fp in glob.glob(patron):
        if os.path.abspath(fp) == os.path.abspath(reporte_actual):
            continue
        try:
            wb = load_workbook(fp, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            continue
        hi = next((i for i, r in enumerate(rows)
                   if sum(1 for c in r if c not in (None, "")) >= 5), None)
        if hi is None:
            continue
        hdr_r = [str(c).strip() if c is not None else "" for c in rows[hi]]
        K = {_key(h): i for i, h in enumerate(hdr_r)}
        c_rec = next((K[k] for k in ("nreceta", "noreceta", "numeroreceta") if k in K), None)
        if c_rec is None:
            continue
        for r in rows[hi + 1:]:
            if not r:
                continue
            v = str(r[c_rec]).strip() if c_rec < len(r) and r[c_rec] is not None else ""
            if v and v.upper() not in ("TOTAL", "NONE", ""):
                ya.add(v)
    return ya


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("reporte", help="reporteGestionTerritorial*.xlsx")
    ap.add_argument("--salida", default="./out_gt")
    ap.add_argument("--hist-glob", default=os.path.join(MAESTRO_DIR, "informe_completo_recetas*.csv"))
    ap.add_argument("--generar", action="store_true", help="Invocar generar.py del skill al terminar el cruce")
    ap.add_argument("--no-pdf", action="store_true", help="No generar PDFs (pasa --no-pdf a generar.py)")
    ap.add_argument("--no-dedup", action="store_true", help="No filtrar recetas ya procesadas en GT anteriores")
    ap.add_argument("--gt-dir", default=GT_DIR_DEFAULT,
                    help="Carpeta con reporteGestionTerritorial_*.xlsx para detectar previos")
    a = ap.parse_args()

    os.makedirs(a.salida, exist_ok=True)
    regs, hdr = leer_reporte_gt(a.reporte)
    print(f"Reporte GT: {len(regs)} recetas únicas | columnas detectadas OK")

    # Dedup: excluir recetas ya listadas en reportes GT anteriores
    if not a.no_dedup and os.path.isdir(a.gt_dir):
        ya_procesadas = _recetas_en_gt_previos(a.reporte, a.gt_dir)
        if ya_procesadas:
            antes = len(regs)
            regs = {k: v for k, v in regs.items() if k not in ya_procesadas}
            omitidas = antes - len(regs)
            if omitidas:
                print(f"  [dedup GT] {omitidas} receta(s) omitidas por ya estar en reportes anteriores")
    recetas_set = set(regs.keys())

    archivos = sorted(glob.glob(a.hist_glob))
    print(f"Histórico: {len(archivos)} archivo(s) — cruzando por Nº receta (puede tardar)...")
    det = cruzar_historico(recetas_set, archivos)
    encontradas = sum(1 for r in recetas_set if r in det and det[r]["lineas"])
    print(f"  Cruce: {encontradas}/{len(recetas_set)} recetas encontradas en el histórico")

    for rec, reg in regs.items():
        clasificar(reg, det.get(rec, {"tipo_receta":"","estado":"","gt":"","lineas":OrderedDict()}))

    # Resumen
    n_ref = sum(1 for g in regs.values() if g["refrigerado"])
    n_con = sum(1 for g in regs.values() if g["controlado"])
    n_pen = sum(1 for g in regs.values() if g["pendiente"])
    n_nohist = sum(1 for g in regs.values() if not g["_en_historico"])
    print(f"\n  Clasificación (recetas con al menos uno):")
    print(f"    ❄  Refrigerados : {n_ref}")
    print(f"    ⚠  Controlados  : {n_con}")
    print(f"    ⏳ Pendientes   : {n_pen}")
    if n_nohist:
        print(f"    • No halladas en histórico: {n_nohist} (sin clasificar)")
    print(f"\n  Por establecimiento de destino:")
    por_dest = defaultdict(lambda: [0,0,0,0])
    for g in regs.values():
        d = por_dest[g["estab_destino"]]
        d[0]+=1; d[1]+=bool(g["refrigerado"]); d[2]+=bool(g["controlado"]); d[3]+=bool(g["pendiente"])
    for dest, (t,rf,co,pe) in sorted(por_dest.items()):
        print(f"    {dest:<28} {t:>3} recetas | ❄{rf}  ⚠{co}  ⏳{pe}")

    # JSON enriquecido (formato registros del skill)
    CAMPOS_PUBLICOS = ("receta","paciente","edad","direccion","comuna","telefono",
                        "estab_origen","estab_destino","periodo","especialidad","n_presc",
                        "ventanilla","refrigerado","pendiente","controlado")
    out_json = os.path.join(a.salida, "gt_enriquecido.json")
    data = {
        "fecha_entrega": next((g["fecha_entrega_rep"] for g in regs.values() if g.get("fecha_entrega_rep")), ""),
        "origen": "Farmacia Hospital de Pitrufquén",
        "registros": [{k: g[k] for k in CAMPOS_PUBLICOS} for g in regs.values()],
    }
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n  → {out_json}")

    out_xlsx = os.path.join(a.salida, "Cruce_GT_Clasificacion.xlsx")
    escribir_excel(regs, out_xlsx)
    print(f"  → {out_xlsx}")

    if a.generar:
        generar = _buscar_generar_py()
        if not generar:
            print("\n  [aviso] No se encontró generar.py del skill — genera manualmente con:")
            print(f"    py <ruta>/generar.py {out_json} --salida {a.salida}")
        else:
            print(f"\n[GT] Generando planillas → {a.salida} ...")
            import subprocess as _sp
            # generar.py necesita "run" para la columna RUN de la planilla impresa,
            # pero el JSON persistido (gt_enriquecido.json) no debe llevar RUT (Ley 19.628).
            # Se pasa un JSON temporal con "run" que se borra apenas termina el subproceso.
            data_con_run = {**data, "registros": [
                {**{k: g[k] for k in CAMPOS_PUBLICOS}, "run": g.get("run", "")}
                for g in regs.values()
            ]}
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".json", prefix="gt_tmp_")
            try:
                with os.fdopen(tmp_fd, "w", encoding="utf-8") as f:
                    json.dump(data_con_run, f, ensure_ascii=False, indent=2)
                cmd = [sys.executable, generar, os.path.abspath(tmp_path),
                       "--salida", os.path.abspath(a.salida)]
                if a.no_pdf:
                    cmd.append("--no-pdf")
                env = os.environ.copy()
                env["PYTHONUTF8"] = "1"; env["PYTHONIOENCODING"] = "utf-8"
                _sp.run(cmd, env=env)
            finally:
                try: os.remove(tmp_path)
                except OSError: pass


if __name__ == "__main__":
    main()
