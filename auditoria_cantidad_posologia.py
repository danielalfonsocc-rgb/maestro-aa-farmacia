#!/usr/bin/env python3
"""
auditoria_cantidad_posologia.py
════════════════════════════════════════════════════════════════════════════════
Auditoría de CANTIDAD RECETADA vs. POSOLOGÍA — todo el universo AA (excepto
insulinas, que tienen su propia auditoría: auditoria_insulinas.py).

Para cada prescripción del mes (despachada o pendiente) intenta calcular cuántas
unidades (comprimidos/cápsulas/etc.) debería cubrir 30 días de tratamiento según
la posología escrita por el médico, y lo compara contra la Cantidad Recetada.

El cálculo SOLO se intenta cuando:
  1. La forma farmacéutica es un sólido oral simple (comprimido, cápsula,
     tableta, gragea, sobre, óvulo, supositorio) — 1 unidad = 1 dosis, sin
     conversión de volumen/concentración de por medio.
  2. La receta es de un tratamiento CRÓNICO multi-cuota (campo Periodo = "X de
     N" con N ≥ 2), igual que auditoria_duplicados_profunda.py. Las recetas de
     una sola entrega ("1 de 1") suelen ser cursos cortos o SOS con cantidades
     legítimamente variables — calcularlas igual produciría falsas alarmas.
  3. La posología permite calcular una dosis diaria inequívoca.

Todo lo demás (otras formas farmacéuticas, cursos cortos, posología ambigua o
ausente) queda en la categoría "SIN DATOS": se muestra igual (no se oculta
ninguna línea del universo) pero sin cálculo ni urgencia.

Hojas Excel:
  1. Medicamentos Este Mes — 1 fila por prescripción, ordenada por prioridad
  2. Histórico por Paciente — evolución de la posología de cada paciente/medicamento
  3. Sin Datos               — casos no evaluables, con el motivo
  4. Metodología

Uso:
    py auditoria_cantidad_posologia.py
    py auditoria_cantidad_posologia.py --mes 2026-07
    py auditoria_cantidad_posologia.py --salida mi_reporte.xlsx
"""
import argparse
import math
import os
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aa_colors import TEAL, ROJO, GRIS
from utils_aa import cargar_recetas_csv, setup_stdout

setup_stdout()
WORK = os.path.dirname(os.path.abspath(__file__))

BODEGA_OBJETIVO    = "FARMACIA AT ABIERTA"
ESTADOS_EXCLUIDOS  = {"ANULADO", "RECHAZADO", "REEMPLAZADO", "DEVUELTO"}
ESTADOS_DESPACHADO = {"ENTREGADO"}
ESTADOS_PENDIENTE  = {"PENDIENTE", "SOLICITADO"}
DIAS_CICLO         = 30   # 1 cuota mensual = 30 días, igual que el resto del proyecto.
MIN_CUOTAS_CRONICO = 2    # Periodo "X de N" con N < 2 = entrega única (curso corto/SOS).

FORMA_SOPORTADA_RE = re.compile(
    r"\bCOMPRIMIDOS?\b|\bTABLETAS?\b|\bGRAGEAS?\b|\bCAPSULAS?\b|\bCAP\b|\bCM\b|\bCP\b|"
    r"\bSOBRES?\b|\bOVULOS?\b|\bSUPOSITORIOS?\b"
)
UNIT_WORDS_RE = (r"COMPRIMIDOS?|COMP\b|CMP\b|TABLETAS?|GRAGEAS?|CAPSULAS?|CAP\b|CP\b|"
                 r"SOBRES?|OVULOS?|SUPOSITORIOS?|UNIDAD(?:ES)?")
MEAL_KW       = ["DESAYUNO", "ALMUERZO", "CENA", "ONCE"]
MULTI_TOMA_RE = re.compile(
    r"CADA\s+COMIDA|TODAS\s+LAS\s+COMIDAS|ANTES\s+(?:DE\s+)?(?:LAS\s+)?COMIDAS?|"
    r"PRANDIAL|EN\s+CADA\s+COMIDA"
)
NUM_PALABRAS = {
    "MEDIA": "0.5", "MEDIO": "0.5", "CUARTO": "0.25",
    "UNA": "1", "UNO": "1", "UN": "1", "OTRO": "1", "OTRA": "1",
    "DOS": "2", "TRES": "3", "CUATRO": "4",
}
NUM_PALABRAS_RE = re.compile(r"\b(" + "|".join(NUM_PALABRAS) + r")\b")
NO_ES_DOSIS = {"MIN", "MINUTOS", "MINUTO", "HRS", "HORAS", "HORA", "AM", "PM",
              "MG", "MCG", "GR", "G", "ML", "UI", "%"}
DIAS_SEMANA = r"LUNES|MARTES|MI[EÉ]RCOLES|JUEVES|VIERNES|S[AÁ]BADOS?|DOMINGOS?"
SEMANAL_RE = re.compile(
    rf"CADA\s+SEMANA|POR\s+SEMANA|A\s+LA\s+SEMANA|X\s+SEMANA|/\s*SEMANA|"
    rf"SEMANAL(?:MENTE)?|LOS\s+D[IÍ]AS?\s+(?:{DIAS_SEMANA})|"
    rf"TODOS\s+LOS\s+(?:{DIAS_SEMANA})|C/\s*(?:{DIAS_SEMANA})|\(\s*(?:{DIAS_SEMANA})\s*\)|"
    rf"^(?:{DIAS_SEMANA})\b"
)
TIEMPO_LISTA_RE = re.compile(
    r"(?:(?<![.,\d])\d+\s*(?:AM|PM)\s*[,Y]\s*){1,}(?<![.,\d])\d+\s*(?:AM|PM)"
)
VECES_SEMANA_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*VECES\s*(?:POR|A\s+LA)\s*SEMANA")

PRIO_COLOR = {
    "URGENTE":   ("DC2626", "FEE2E2"),
    "REVISAR":   ("C2410C", "FFF3E0"),
    "OK":        ("15803D", "F0FDF4"),
    "SIN DATOS": ("6B7280", "F3F4F6"),
}
_ORDEN_PRIO = {"URGENTE": 0, "REVISAR": 1, "OK": 2, "SIN DATOS": 3}

THIN   = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    "ID Receta Detalle", "RUN", "Nombre", "Apellido Paterno", "Apellido Materno",
    "Prescripción", "Número Receta", "Periodo", "Bodega Despacha",
    "Fecha Atención", "Fecha Entrega Receta", "Estado Prescripción",
    "Nombre Profesional", "Apellido Paterno Profesional",
    "Apellido Materno Profesional", "Especialidad",
    "Cod. Diagnóstico 1", "Diagnóstico 1",
    "Cantidad Recetada", "Cantidad Entregada",
    "Observación Médica Prescripción",
]


def _forma_soportada(presc: str) -> bool:
    return bool(FORMA_SOPORTADA_RE.search(str(presc).upper()))


def _cuotas_periodo(periodo: str):
    m = re.search(r"de\s*(\d+)", str(periodo), re.IGNORECASE)
    return int(m.group(1)) if m else 1


def cargar_y_preparar() -> pd.DataFrame:
    try:
        rec = cargar_recetas_csv(WORK, cols=COLS, solo_ultimo=False)
    except FileNotFoundError:
        print("[AVISO] No hay CSV de recetas. Ejecuta AUTO_SSASUR.bat primero.")
        sys.exit(0)

    rec = rec.copy()
    presc = rec["Prescripción"].fillna("").str.upper()
    bod = rec["Bodega Despacha"].fillna("").str.upper().str.strip()
    est = rec["Estado Prescripción"].fillna("").str.upper().str.strip()

    rec = rec[(bod == BODEGA_OBJETIVO) & (~est.isin(ESTADOS_EXCLUIDOS)) &
              (~presc.str.contains("INSULINA", regex=False))].copy()
    if rec.empty:
        print("[AVISO] No se encontraron prescripciones en Farmacia AT Abierta.")
        sys.exit(0)

    fa = pd.to_datetime(rec["Fecha Atención"],       dayfirst=True, errors="coerce")
    fe = pd.to_datetime(rec["Fecha Entrega Receta"], dayfirst=True, errors="coerce")
    rec["_fecha"] = fa.fillna(fe)
    rec = rec.dropna(subset=["_fecha"])
    rec["_mes"]   = rec["_fecha"].dt.to_period("M")

    rec["_paciente"] = (
        rec["Nombre"].fillna("") + " " + rec["Apellido Paterno"].fillna("") +
        " " + rec["Apellido Materno"].fillna("")
    ).str.strip().str.title()
    rec["_medico"] = (
        rec["Nombre Profesional"].fillna("") + " " +
        rec["Apellido Paterno Profesional"].fillna("") + " " +
        rec["Apellido Materno Profesional"].fillna("")
    ).str.strip().str.title()
    rec["_pos"]         = rec["Observación Médica Prescripción"].fillna("").str.strip()
    rec["_cant_r"]      = pd.to_numeric(rec["Cantidad Recetada"],  errors="coerce").fillna(0)
    rec["_cant_e"]      = pd.to_numeric(rec["Cantidad Entregada"], errors="coerce").fillna(0)
    rec["_estado_norm"] = est.reindex(rec.index)
    rec["_forma_ok"]    = presc.reindex(rec.index).apply(_forma_soportada)
    rec["_cuotas"]      = rec["Periodo"].apply(_cuotas_periodo)

    rec = rec.sort_values("_fecha")
    return rec


def posologia_historica(rec: pd.DataFrame, run: str, presc: str, antes_de: pd.Timestamp):
    """Última posología no vacía del mismo paciente+medicamento exacto, antes de `antes_de`."""
    prev = rec[(rec["RUN"] == run) & (rec["Prescripción"] == presc) &
               (rec["_fecha"] < antes_de) & (rec["_pos"] != "")]
    if prev.empty:
        return None, None
    ult = prev.iloc[-1]
    return ult["_pos"], ult["_fecha"]


MARCADORES_DIARIOS = ("POR DIA", "AL DIA", "DIARIO", "DIARIA", "NOCHE", "MAÑANA", "MANANA",
                     "TARDE", "ACOSTARSE", "DESAYUNO", "ALMUERZO", "CENA", "ONCE",
                     "VEZ AL DIA", "X DIA")


ENTRE_SEMANA_RE = r"LU-?VI\b|L-V\b|LUNES\s*A\s*VIERNES"
FIN_DE_SEMANA_RE = r"SA-?DO\b|S-D\b|S[AÁ]BADO\s*(?:Y|A)\s*DOMINGO"


def _dosis_semana_partida(t_base: str):
    """Detecta 'N entre semana + M fin de semana' (ej. levotiroxina) y retorna
    (promedio_diario, detalle) usando 5 días hábiles + 2 de fin de semana, o None."""
    m1 = re.search(rf"(\d+(?:[.,]\d+)?)\s*[^0-9]{{0,30}}?(?:{ENTRE_SEMANA_RE})|"
                  rf"(?:{ENTRE_SEMANA_RE})\s*[=:]?\s*(\d+(?:[.,]\d+)?)", t_base)
    m2 = re.search(rf"(\d+(?:[.,]\d+)?)\s*[^0-9]{{0,30}}?(?:{FIN_DE_SEMANA_RE})|"
                  rf"(?:{FIN_DE_SEMANA_RE})\s*[=:]?\s*(\d+(?:[.,]\d+)?)", t_base)
    if not (m1 and m2):
        return None
    d1 = float((m1.group(1) or m1.group(2)).replace(",", "."))
    d2 = float((m2.group(1) or m2.group(2)).replace(",", "."))
    promedio = (d1 * 5 + d2 * 2) / 7
    return promedio, f"{d1:g} L-V + {d2:g} S-D → promedio {promedio:.2f}/día"


def _dosis_con_marcador(dosis: float, t_base: str, es_semanal: bool, dias_ciclo: int):
    """Confirma una dosis única (ya extraída) SOLO si hay un marcador de frecuencia
    reconocible junto a ella — nunca asume 'una vez al día' por defecto."""
    if es_semanal:
        return dosis, True, f"{dosis:g}, 1 vez/semana", dias_ciclo, 7
    n_comidas = sum(1 for kw in MEAL_KW if kw in t_base)
    if n_comidas >= 2 or MULTI_TOMA_RE.search(t_base):
        return (None, False, (f"Posología menciona {dosis:g} en varias comidas sin "
                              "detallar cada toma — verificar dosis diaria real"), dias_ciclo, 1)
    if any(mk in t_base for mk in MARCADORES_DIARIOS):
        return dosis, True, f"{dosis:g}, 1 vez/día", dias_ciclo, 1
    return (None, False, "No se reconoce un marcador de frecuencia diaria/semanal junto al número "
            "(posible indicación 'a necesidad' sin horario fijo)", dias_ciclo, 1)


def _dosis_diaria_generica(texto: str):
    """
    Estima la dosis diaria (o semanal) total desde el texto libre de posología,
    para formas farmacéuticas simples (1 unidad = 1 dosis).
    Retorna (dosis_valor | None, confiable: bool, detalle: str, dias_ciclo: int,
    periodo_dias: int) — periodo_dias es 1 si dosis_valor es diaria, 7 si es semanal.
    Solo calcula cuando el patrón es inequívoco; en cualquier caso ambiguo
    devuelve confiable=False para que el QF lo revise manualmente.
    """
    t = str(texto).upper().strip()
    if not t:
        return None, False, "Sin texto de posología", DIAS_CICLO, 1

    # Esquema de dosis cambiante en el tiempo ("...LUEGO AUMENTAR A...", titulación):
    # no hay una tasa fija que calcular.
    if re.search(r"\bLUEGO\b", t):
        return None, False, "Posología cambia en el tiempo (titulación) — verificar manualmente", DIAS_CICLO, 1

    # "DÍA POR MEDIO" / "X DIA X MEDIO" = cada 48 horas (día sí, día no) — debe resolverse
    # ANTES de que "MEDIO" se sustituya por 0.5 más abajo, o se pierde el significado.
    t = re.sub(r"(?:X|POR)\s*D[IÍ]AS?\s*(?:X|POR)\s*MEDIO\b", " CADA 48 HORAS ", t)

    # "N ... Y MEDIO/MEDIA" -> N+0.5 (antes de sustituir palabras sueltas)
    def _mas_medio(m):
        return f"{float(m.group(1).replace(',', '.')) + 0.5:g}"
    t = re.sub(rf"(\d+(?:[.,]\d+)?)\s*(?:{UNIT_WORDS_RE})?\s*Y\s*MEDI[OA]\b", _mas_medio, t)

    # "UNA/UNO/DOS... Y MEDIO/MEDIA" (número en palabra) -> N+0.5
    def _mas_medio_palabra(m):
        return f"{float(NUM_PALABRAS[m.group(1)]) + 0.5:g}"
    t = re.sub(r"\b(UNA|UNO|DOS|TRES|CUATRO)\s*Y\s*MEDI[OA]\b", _mas_medio_palabra, t)

    # Fracción "1/2" (con o sin entero antes): "11/2", "4 1/2", "1/2" -> N+0.5
    t = re.sub(r"(\d+)?\s*1/2\b", lambda m: f"{int(m.group(1) or 0) + 0.5:g}", t)

    # Cualquier otra fracción no reconocida ("1/5", "1/3", etc.) es demasiado ambigua
    if re.search(r"\d+\s*/\s*\d+", t):
        return None, False, "Posología usa una fracción no reconocida — verificar manualmente", DIAS_CICLO, 1

    # Números escritos en palabra: "UNA", "MEDIA", "DOS", etc. (ya resueltos "Y MEDIO"/"1/2")
    t = NUM_PALABRAS_RE.sub(lambda m: NUM_PALABRAS[m.group(1)], t)

    # Quita restituciones de concentración/volumen ("40 MG", "5 ML"): no son cantidad de unidades
    t = re.sub(r"\b\d+(?:[.,]\d+)?\s*(?:MG|MCG|GR|ML|UI)\b", " ", t)

    # Cualquier indicación SOS/"a necesidad" en CUALQUIER parte del texto vuelve toda la
    # posología condicional (no una dosis recurrente comparable) — a diferencia de una nota
    # de corrección puntual, aquí no se puede rescatar una frecuencia dicha antes del "SOS":
    # "1 X DIA SOS" significa "hasta 1 al día, a necesidad", no "1 al día" fijo.
    if re.search(r"\bSOS\b|EN\s+CASO\s+DE|SI\s+HAY|SI\s+NECESITA|A\s+NECESIDAD|"
                r"SEGUN\s+NECESIDAD|S\.O\.S\.?", t):
        return None, False, "Es SOS/a necesidad — no tiene una dosis diaria fija que calcular", DIAS_CICLO, 1
    t_base = t

    # Dosis de carga / total de un tratamiento puntual: no es una dosis diaria recurrente
    if re.search(r"\bCARGA\b|\bTOTAL\b", t_base):
        return (None, False, "Menciona dosis de carga/total (no recurrente) — no aplica cálculo mensual",
                DIAS_CICLO, 1)

    # Duración explícita distinta al ciclo estándar: "POR 10 DIAS", "POR 8 SEMANAS"
    dias_ciclo = DIAS_CICLO
    m_dur_dias = re.search(r"(?:POR|X)\s+(\d+)\s*D[IÍ]AS?", t_base)
    m_dur_sem  = re.search(r"(?:POR|X)\s+(\d+)\s*SEMANAS?", t_base)
    if m_dur_sem:
        dias_ciclo = int(m_dur_sem.group(1)) * 7
    elif m_dur_dias:
        dias_ciclo = int(m_dur_dias.group(1))

    # Dosis distinta entre semana / fin de semana (común en levotiroxina): calcula el
    # promedio diario ponderado en vez de sumarlas como si fueran dos tomas del mismo día.
    partida = _dosis_semana_partida(t_base)
    if partida:
        return partida[0], True, partida[1], dias_ciclo, 1

    # Lista de horarios de reloj ("A LAS 8 AM, 14 PM Y 20 PM"): son horas, no dosis —
    # demasiado ambiguo cuántas veces se repite la dosis inicial mencionada una sola vez.
    if TIEMPO_LISTA_RE.search(t_base):
        return (None, False, "Posología enumera horarios sin repetir la dosis en cada uno "
                "— verificar manualmente", dias_ciclo, 1)

    es_semanal = bool(SEMANAL_RE.search(t_base))

    # 1) Lista de tomas al inicio del texto: "1-0-1", "2-0-0", "0-0-1" (solo diarias)
    if not es_semanal:
        m = re.match(r"^(\d+(?:[.,]\d+)?(?:\s*[-,]\s*\d+(?:[.,]\d+)?){1,4})(?=\s|$|\.)", t_base)
        if m:
            nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", m.group(1))]
            return (sum(nums), True, f"Suma de {len(nums)} tomas ({'+'.join(f'{n:g}' for n in nums)})",
                    dias_ciclo, 1)

        # 2) Tomas separadas por "+" o " Y " (lenient: primer número de cada segmento)
        #    ej. "1 EN LA MAÑANA Y UNO EN LA NOCHE" (ya con "UNO"->"1" sustituido).
        #    Para " Y " exige que cada segmento tenga un marcador horario, para no
        #    confundir una nota aparte ("Y CONTROL EN 1 MES") con una segunda toma.
        conector = "+" if "+" in t_base else (r"\bY\b" if re.search(r"\bY\b", t_base) else None)
        if conector:
            partes = t_base.split("+") if conector == "+" else re.split(conector, t_base)
            marcador_hora = ("MAÑANA", "MANANA", "NOCHE", "TARDE", "AM", "PM", "ALMUERZO",
                             "DESAYUNO", "CENA", "ONCE", "ACOSTARSE")
            nums = []
            for seg in partes:
                m_seg = re.search(r"(\d+(?:[.,]\d+)?)", seg)
                if m_seg and (conector == "+" or any(mk in seg for mk in marcador_hora)):
                    nums.append(float(m_seg.group(1).replace(",", ".")))
            if len(nums) >= 2:
                sep = "+" if conector == "+" else "Y"
                return (sum(nums), True, f"Suma de {len(nums)} tomas ({sep}) ({'+'.join(f'{n:g}' for n in nums)})",
                        dias_ciclo, 1)

    dose_all = re.findall(rf"(\d+(?:[.,]\d+)?)\s*(?:{UNIT_WORDS_RE})", t_base)

    # 3) ≥2 menciones explícitas con unidad, sin conector (solo diarias)
    if not es_semanal and len(dose_all) >= 2:
        nums = [float(x.replace(",", ".")) for x in dose_all]
        return sum(nums), True, f"Suma de {len(nums)} tomas ({'+'.join(f'{n:g}' for n in nums)})", dias_ciclo, 1

    lead_m = re.match(r"^(\d+(?:[.,]\d+)?)\b", t_base)

    if not es_semanal:
        # 4) "CADA N HORAS" / "C/N H(RS)" con dosis inicial
        freq_m = re.search(r"CADA\s+(\d+)\s*(?:HORAS?|HRS?)\b|C/\s*(\d+)\s*H(?:RS?)?\b", t_base)
        if freq_m and lead_m:
            n_horas = int(freq_m.group(1) or freq_m.group(2))
            veces_dia = 24 / n_horas
            dosis_toma = float(lead_m.group(1).replace(",", "."))
            return (dosis_toma * veces_dia, True,
                    f"{dosis_toma:g} c/{n_horas}h → {veces_dia:g} veces/día", dias_ciclo, 1)

        # 5) Multiplicador explícito "X VECES AL DIA"
        mult_m = re.search(r"(\d+)\s*VECES\s*(?:AL|POR)\s*D[IÍ]A", t_base)
        if mult_m and lead_m:
            dosis_toma = float(lead_m.group(1).replace(",", "."))
            veces = int(mult_m.group(1))
            return dosis_toma * veces, True, f"{dosis_toma:g} × {veces} veces/día", dias_ciclo, 1

    # 6) Multiplicador semanal explícito: "N VECES POR SEMANA"
    veces_sem_m = VECES_SEMANA_RE.search(t_base)
    if es_semanal and veces_sem_m and lead_m:
        dosis_toma = float(lead_m.group(1).replace(",", "."))
        veces = float(veces_sem_m.group(1).replace(",", "."))
        return dosis_toma * veces, True, f"{dosis_toma:g} × {veces:g} veces/semana", dias_ciclo, 7

    # 7) Una sola mención con unidad explícita — solo confirma con marcador de frecuencia
    if len(dose_all) == 1:
        return _dosis_con_marcador(float(dose_all[0].replace(",", ".")), t_base, es_semanal, dias_ciclo)

    # 8) Número inicial sin unidad explícita — solo si va seguido de un marcador de
    #    frecuencia reconocible (evita confundir "30 MIN ANTES..." o "40 MG" con una dosis)
    if lead_m:
        resto = t_base[lead_m.end():].strip()
        primera_palabra = resto.split(" ", 1)[0] if resto else ""
        if primera_palabra in NO_ES_DOSIS:
            return (None, False, "El número detectado es parte de una instrucción de tiempo/concentración, "
                    "no una dosis", dias_ciclo, 1)
        return _dosis_con_marcador(float(lead_m.group(1).replace(",", ".")), t_base, es_semanal, dias_ciclo)

    return None, False, "No se detectó una dosis numérica en el texto", dias_ciclo, 1


def _evaluar_cantidad(dosis_valor, confiable: bool, detalle: str, dias_ciclo: int, periodo_dias: int,
                     recetados: int, forma_ok: bool, es_cronico: bool):
    """Retorna (prioridad, observación, esperados|None)."""
    if not forma_ok:
        return ("SIN DATOS", "Forma farmacéutica no soportada para este cálculo (no es sólido oral simple).",
                None)
    if not es_cronico:
        return ("SIN DATOS", ("Receta de una sola entrega (curso corto o SOS posible) — "
                              "no se calcula para evitar falsas alarmas."), None)
    if not confiable or dosis_valor is None:
        return "SIN DATOS", f"{detalle}.", None

    esperados = math.ceil(round(dosis_valor / periodo_dias * dias_ciclo, 6))
    diff = int(recetados) - esperados
    unidad = "semana" if periodo_dias == 7 else "día"
    base = (f"{detalle} → {dosis_valor:g} unid./{unidad} × {dias_ciclo} días = {esperados} "
            f"unidad(es) esperada(s); recetadas: {int(recetados)}.")

    if diff == 0:
        return "OK", base, esperados
    if abs(diff) == 1:
        return "REVISAR", base + " Diferencia de 1 unidad (margen habitual) — revisar cantidad.", esperados
    falta_o_exceso = "faltan unidades para cubrir el ciclo" if diff < 0 else "exceso de unidades recetadas"
    return "URGENTE", base + f" No concuerda ({falta_o_exceso}) — revisar cantidad recetada.", esperados


def construir_reportes(rec: pd.DataFrame, mes: pd.Period):
    del_mes = rec[rec["_mes"] == mes].copy()
    if del_mes.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    filas = []
    for _, r in del_mes.iterrows():
        estado_norm = r["_estado_norm"]
        if estado_norm in ESTADOS_DESPACHADO or r["_cant_e"] > 0:
            estado_disp = "Despachado"
        elif estado_norm in ESTADOS_PENDIENTE:
            estado_disp = "Pendiente"
        else:
            estado_disp = estado_norm.title() or "—"

        pos_mes = r["_pos"]
        if pos_mes:
            pos_final, fuente = pos_mes, "Receta de este mes"
        else:
            pos_hist, fecha_hist = posologia_historica(rec, r["RUN"], r["Prescripción"], r["_fecha"])
            if pos_hist:
                pos_final, fuente = pos_hist, f"Histórico ({fecha_hist:%d-%m-%Y})"
            else:
                pos_final, fuente = "", "Sin registro"

        es_cronico = r["_cuotas"] >= MIN_CUOTAS_CRONICO
        dosis_valor, confiable, detalle_dosis, dias_ciclo, periodo_dias = _dosis_diaria_generica(pos_final)
        prioridad, observacion, esperados = _evaluar_cantidad(
            dosis_valor, confiable, detalle_dosis, dias_ciclo, periodo_dias,
            r["_cant_r"], r["_forma_ok"], es_cronico)

        filas.append({
            "Prioridad revisión"    : prioridad,
            "RUN"                  : r["RUN"],
            "Paciente"              : r["_paciente"],
            "Medicamento"           : r["Prescripción"],
            "Estado"                : estado_disp,
            "Periodo"               : r["Periodo"],
            "N° Receta"             : r["Número Receta"],
            "Fecha"                 : r["_fecha"],
            "Cant. recetada"        : int(r["_cant_r"]),
            "Cant. entregada"       : int(r["_cant_e"]),
            "Dosis detectada"       : dosis_valor if dosis_valor is not None else "—",
            "Frecuencia"            : ("por semana" if periodo_dias == 7 else "por día")
                                       if dosis_valor is not None else "—",
            "Días de ciclo usados"  : dias_ciclo if dosis_valor is not None else "—",
            "Cantidad esperada"     : esperados if esperados is not None else "—",
            "Diferencia"            : (int(r["_cant_r"]) - esperados) if esperados is not None else "—",
            "Observación"           : observacion,
            "Médico"                : r["_medico"],
            "Especialidad"          : r["Especialidad"] or "—",
            "Diagnóstico"           : (str(r["Cod. Diagnóstico 1"]).strip() + " · " +
                                        str(r["Diagnóstico 1"]).strip()).strip(" ·") or "—",
            "Posología (este mes)"  : pos_mes,
            "Posología detectada"   : pos_final,
            "Fuente posología"      : fuente,
        })

    resumen = pd.DataFrame(filas)
    resumen["_o"] = resumen["Prioridad revisión"].map(_ORDEN_PRIO)
    resumen = resumen.sort_values(
        ["_o", "Estado", "Paciente"], ascending=[True, False, True]).drop(columns=["_o"])

    # ── Histórico por paciente (para pacientes del mes con URGENTE/REVISAR) ─
    pares = resumen.loc[resumen["Prioridad revisión"].isin(["URGENTE", "REVISAR"]),
                        ["RUN", "Medicamento"]].drop_duplicates()
    hist_rows = []
    for _, p in pares.iterrows():
        sub = rec[(rec["RUN"] == p["RUN"]) & (rec["Prescripción"] == p["Medicamento"]) &
                  (rec["_pos"] != "")].sort_values("_fecha", ascending=False)
        for _, r in sub.iterrows():
            hist_rows.append({
                "RUN"        : r["RUN"],
                "Paciente"   : r["_paciente"],
                "Medicamento": r["Prescripción"],
                "Fecha"      : r["_fecha"],
                "Periodo"    : r["Periodo"],
                "Estado"     : r["_estado_norm"].title(),
                "Posología"  : r["_pos"],
                "Cant. recetada": int(r["_cant_r"]),
                "Médico"     : r["_medico"],
            })
    historico = pd.DataFrame(hist_rows)

    sin_datos = resumen[resumen["Prioridad revisión"] == "SIN DATOS"].copy()

    return resumen, historico, sin_datos


# ── Escritura Excel ──────────────────────────────────────────────────────────

def _encabezado_hoja(ws, texto: str, color: str, n_cols: int):
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = texto
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _escribir_tabla(ws, df: pd.DataFrame, fila_inicio: int,
                    anchos: dict | None = None, fechas: list | None = None,
                    wrap_cols: tuple = (), estado_col: str | None = None,
                    prio_col: str | None = None):
    fechas = fechas or []
    cols   = list(df.columns)
    col_ix = {c: i + 1 for i, c in enumerate(cols)}

    for c_i, col in enumerate(cols, 1):
        cell = ws.cell(row=fila_inicio, column=c_i, value=col)
        cell.fill      = PatternFill("solid", fgColor=TEAL)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[fila_inicio].height = 28
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)

    for r_off, (_, row) in enumerate(df.iterrows()):
        r = fila_inicio + 1 + r_off
        ws.append(list(row.values))
        for c_i, col in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c_i)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col in wrap_cols))
            if prio_col and col == prio_col:
                txt_c, bg_c = PRIO_COLOR.get(str(cell.value), ("1F2937", "FFFFFF"))
                cell.font = Font(bold=True, color=txt_c)
                cell.fill = PatternFill("solid", fgColor=bg_c)
            elif estado_col and col == estado_col:
                if str(cell.value) == "Pendiente":
                    cell.font = Font(bold=True, color=ROJO)
                elif str(cell.value) == "Despachado":
                    cell.font = Font(bold=True, color="15803D")
            elif r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRIS)
        ws.row_dimensions[r].height = 32 if any(col in wrap_cols for col in cols) else 20

    for fcol in fechas:
        if fcol in col_ix:
            for r in range(fila_inicio + 1, ws.max_row + 1):
                ws.cell(row=r, column=col_ix[fcol]).number_format = "DD-MM-YYYY"

    for i, col in enumerate(cols, 1):
        w = (anchos or {}).get(col)
        if w is None:
            vals = df[col].astype(str)
            w = min(max(len(str(col)) + 2, vals.str.len().max() + 2 if len(df) else 12), 55)
        ws.column_dimensions[get_column_letter(i)].width = w

    ultima = ws.max_row
    ws.auto_filter.ref = f"A{fila_inicio}:{get_column_letter(len(cols))}{ultima}"


ANCHOS_RES = {
    "Prioridad revisión": 13, "Paciente": 26, "Medicamento": 44, "Estado": 12,
    "Periodo": 9, "N° Receta": 12, "Fecha": 13,
    "Dosis detectada": 13, "Frecuencia": 11, "Días de ciclo usados": 14,
    "Cantidad esperada": 14, "Diferencia": 11, "Observación": 55,
    "Médico": 30, "Especialidad": 20,
    "Diagnóstico": 34, "Posología (este mes)": 40, "Posología detectada": 40,
    "Fuente posología": 20,
}
ANCHOS_HIST = {
    "Paciente": 26, "Medicamento": 44, "Fecha": 13, "Periodo": 9,
    "Estado": 12, "Posología": 40, "Cant. recetada": 14, "Médico": 30,
}


def exportar_excel(resumen, historico, sin_datos, mes: pd.Period, dest: str):
    wb = Workbook()

    n_total = len(resumen)
    n_desp  = int((resumen["Estado"] == "Despachado").sum()) if n_total else 0
    n_pend  = int((resumen["Estado"] == "Pendiente").sum()) if n_total else 0
    n_urg   = int((resumen["Prioridad revisión"] == "URGENTE").sum()) if n_total else 0
    n_rev   = int((resumen["Prioridad revisión"] == "REVISAR").sum()) if n_total else 0
    n_ok    = int((resumen["Prioridad revisión"] == "OK").sum()) if n_total else 0
    n_sin   = len(sin_datos)

    ws1 = wb.active
    ws1.title = "Medicamentos Este Mes"
    res_disp = resumen.drop(columns=["RUN"], errors="ignore")
    _encabezado_hoja(ws1,
                     f"CANTIDAD RECETADA VS. POSOLOGÍA · {mes}  ·  {n_total} prescripciones "
                     f"({n_desp} despachadas, {n_pend} pendientes)  ·  "
                     f"URGENTE: {n_urg}  ·  REVISAR: {n_rev}  ·  OK: {n_ok}  ·  Sin datos: {n_sin}",
                     TEAL, max(len(res_disp.columns), 10))
    _escribir_tabla(ws1, res_disp, fila_inicio=2, anchos=ANCHOS_RES, fechas=["Fecha"],
                    wrap_cols=("Posología (este mes)", "Posología detectada", "Médico", "Observación"),
                    estado_col="Estado", prio_col="Prioridad revisión")

    ws2 = wb.create_sheet("Histórico por Paciente")
    hist_disp = historico.drop(columns=["RUN"], errors="ignore")
    _encabezado_hoja(ws2, "EVOLUCIÓN DE POSOLOGÍA — pacientes con caso URGENTE o REVISAR",
                     "0F766E", max(len(hist_disp.columns), 8) if len(hist_disp) else 8)
    if len(hist_disp):
        _escribir_tabla(ws2, hist_disp, fila_inicio=2, anchos=ANCHOS_HIST, fechas=["Fecha"],
                        wrap_cols=("Posología", "Médico"))
    else:
        ws2.cell(row=2, column=1, value="Sin histórico disponible.").font = Font(italic=True, color="6B7280")

    ws3 = wb.create_sheet("Sin Datos")
    sd_disp = sin_datos.drop(columns=["RUN", "Dosis detectada", "Frecuencia", "Días de ciclo usados",
                                       "Cantidad esperada", "Diferencia"], errors="ignore")
    _encabezado_hoja(ws3, f"CASOS NO EVALUABLES (forma no soportada, curso corto o posología "
                          f"ambigua) · {n_sin} caso(s)", GRIS, max(len(sd_disp.columns), 8) if len(sd_disp) else 8)
    if len(sd_disp):
        _escribir_tabla(ws3, sd_disp, fila_inicio=2, anchos=ANCHOS_RES, fechas=["Fecha"],
                        wrap_cols=("Posología (este mes)", "Posología detectada", "Observación", "Médico"),
                        estado_col="Estado")
    else:
        ws3.cell(row=2, column=1, value="Todos los casos fueron evaluables.").font = Font(
            italic=True, color="15803D")

    ws4 = wb.create_sheet("Metodología")
    txt = [
        ("AUDITORÍA DE CANTIDAD RECETADA VS. POSOLOGÍA", True),
        (f"Generado: {datetime.now():%d-%m-%Y %H:%M}  ·  Hospital de Pitrufquén (SSASur, Chile)", False),
        ("", False),
        ("Alcance", True),
        (f"• Mes analizado: {mes}  ·  Bodega: {BODEGA_OBJETIVO}", False),
        ("• Incluye TODO el universo AA excepto insulinas (ver auditoria_insulinas.py, que usa", False),
        ("  su propia lógica de UI/lápices).", False),
        (f"• Excluidos estados: {', '.join(sorted(ESTADOS_EXCLUIDOS))}", False),
        ("• Despachado = Estado 'Entregado' o Cantidad Entregada > 0.", False),
        ("• Pendiente  = Estado 'Pendiente' o 'Solicitado' (aún no retirado).", False),
        ("", False),
        ("Cuándo SÍ se calcula la cantidad esperada", True),
        ("• Forma farmacéutica: sólido oral simple (comprimido, cápsula, tableta, gragea, sobre,", False),
        ("  óvulo, supositorio) — 1 unidad = 1 dosis, sin conversión de volumen/concentración.", False),
        (f"• Tratamiento CRÓNICO multi-cuota: campo Periodo = 'X de N' con N ≥ {MIN_CUOTAS_CRONICO}", False),
        ("  (mismo criterio que auditoria_duplicados_profunda.py). Las recetas de una sola entrega", False),
        ("  ('1 de 1') no se calculan: suelen ser cursos cortos o SOS con cantidades legítimamente", False),
        ("  variables — calcularlas igual generaría falsas alarmas.", False),
        ("• Posología inequívoca: dosis calculada desde texto explícito (lista de tomas '1-0-1',", False),
        ("  suma de tomas separadas por '+' o 'Y', 'cada N horas'/'C/N hrs', 'X veces al día', o una", False),
        ("  sola toma con marcador de frecuencia reconocible). Se reconocen números en palabra ('UNA',", False),
        ("  'MEDIA', 'DOS'), fracciones '1/2', 'Y MEDIO', y una duración distinta ('POR 10 DIAS') si", False),
        ("  aparece. Medicamentos semanales (metotrexato, ácido fólico, vitamina D — 'LOS DIAS", False),
        ("  MIERCOLES', 'A LA SEMANA', 'C/JUEVES') se calculan sobre base semanal, no diaria.", False),
        ("• Dosis distinta entre semana y fin de semana (típico de levotiroxina: 'LU-VI' / 'SA-DO')", False),
        ("  se promedia con 5 días hábiles + 2 de fin de semana, no se suman como si fueran del", False),
        ("  mismo día.", False),
        ("• Cualquier mención de SOS, 'a necesidad', 'en caso de', 'si hay/necesita' en CUALQUIER", False),
        ("  parte del texto anula el cálculo por completo (no se rescata una frecuencia dicha antes:", False),
        ("  '1 x día SOS' no es 'una vez al día', es 'a necesidad'). Igual con dosis de carga/total.", False),
        ("", False),
        ("Cuándo queda en 'SIN DATOS' (no implica urgencia, solo que no se pudo verificar)", True),
        ("• Forma farmacéutica no soportada (jarabes, gotas, colirios, cremas, parches, inhaladores,", False),
        ("  inyectables no-insulina, alimentos, etc.) — no calculados en esta versión.", False),
        ("• Receta de entrega única (curso corto/SOS posible).", False),
        ("• Medicamento SOS/'a necesidad', dosis de carga, o esquema que cambia en el tiempo", False),
        ("  ('luego aumentar a...').", False),
        ("• Posología vacía ('SEGÚN INDICACIÓN'), ambigua (varias comidas sin detallar cada toma,", False),
        ("  fracción no reconocida tipo '1/5', horarios enumerados sin repetir la dosis) o formato", False),
        ("  no reconocido.", False),
        ("• Limitación conocida: errores de tipeo en la posología (ej. 'CASI' por 'CASO') y algunos", False),
        ("  formatos poco frecuentes pueden no detectarse — revisar igual si el caso llama la atención.", False),
        ("", False),
        ("Prioridad de revisión (columna 'Prioridad revisión', planilla ordenada por esta columna)", True),
        ("• OK       — la cantidad recetada coincide exactamente con lo esperado.", False),
        ("• REVISAR  — difiere en 1 unidad (margen habitual, igual conviene mirarlo).", False),
        ("• URGENTE  — difiere en ≥2 unidades (faltan para cubrir el ciclo, o exceso importante).", False),
        ("• SIN DATOS — no se pudo calcular (ver motivo en la columna Observación).", False),
        ("• Esto es un TAMIZAJE automático, no un cálculo clínico definitivo — siempre confirmar con", False),
        ("  la ficha clínica y el criterio del QF antes de ajustar una receta.", False),
        ("", False),
        ("Resultados", True),
        (f"• Prescripciones del mes    : {n_total:,}", False),
        (f"• Despachadas               : {n_desp:,}", False),
        (f"• Pendientes                : {n_pend:,}", False),
        (f"• URGENTE                   : {n_urg:,}", False),
        (f"• REVISAR                   : {n_rev:,}", False),
        (f"• OK                        : {n_ok:,}", False),
        (f"• SIN DATOS                 : {n_sin:,}", False),
        ("", False),
        ("Privacidad", True),
        ("• Este archivo contiene datos de salud de pacientes (Ley 19.628).", False),
        ("• No distribuir fuera de la institución ni publicar en redes compartidas.", False),
    ]
    for i, (line, bold) in enumerate(txt, 1):
        c = ws4.cell(row=i, column=1, value=line)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 10, color=TEAL if bold else "1F2937")
    ws4.column_dimensions["A"].width = 100

    wb.save(dest)


def main():
    ap = argparse.ArgumentParser(
        description="Auditoría de cantidad recetada vs. posología — todo el universo AA (excepto insulinas).")
    ap.add_argument("--mes", default=None, help="Mes a analizar, formato YYYY-MM (default: mes actual).")
    ap.add_argument("--salida", default=None, help="Ruta del Excel de salida.")
    args = ap.parse_args()

    mes = pd.Period(args.mes) if args.mes else pd.Period(datetime.now(), freq="M")
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    dest = args.salida or os.path.join(WORK, f"Auditoria_Cantidad_Posologia_{mes}_{ts}.xlsx")

    print("=" * 70)
    print("  AUDITORÍA DE CANTIDAD RECETADA VS. POSOLOGÍA")
    print(f"  Farmacia AT Abierta · Hospital de Pitrufquén · Mes: {mes}")
    print("=" * 70)

    rec = cargar_y_preparar()
    print(f"Líneas (histórico, Farmacia AT Abierta, sin insulinas): {len(rec):,}")

    resumen, historico, sin_datos = construir_reportes(rec, mes)
    if resumen.empty:
        print(f"\n[AVISO] No hay prescripciones en {mes}.")
        sys.exit(0)

    n_desp = int((resumen["Estado"] == "Despachado").sum())
    n_pend = int((resumen["Estado"] == "Pendiente").sum())
    n_urg  = int((resumen["Prioridad revisión"] == "URGENTE").sum())
    n_rev  = int((resumen["Prioridad revisión"] == "REVISAR").sum())
    n_ok   = int((resumen["Prioridad revisión"] == "OK").sum())

    print(f"\nPrescripciones del mes : {len(resumen):,}")
    print(f"  Despachadas          : {n_desp:,}")
    print(f"  Pendientes           : {n_pend:,}")
    print(f"  URGENTE              : {n_urg:,}")
    print(f"  REVISAR              : {n_rev:,}")
    print(f"  OK                   : {n_ok:,}")
    print(f"  Sin datos            : {len(sin_datos):,}")

    print("\n  Exportando Excel...")
    exportar_excel(resumen, historico, sin_datos, mes, dest)
    print(f"\n[OK] Excel generado: {dest}  ({os.path.getsize(dest) // 1024} KB)")


if __name__ == "__main__":
    main()
