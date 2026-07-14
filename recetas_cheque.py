#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
recetas_cheque.py  —  Registro ISP de Recetas Cheque (Farmacia AT Abierta)
═══════════════════════════════════════════════════════════════════════════
Actualiza el formulario oficial ISP de recetas cheque (estupefacientes y
psicotrópicos) agregando SOLO los folios nuevos. Consume la MISMA sábana
(`informe_completo_recetas*.csv`) que AUTO_SSASUR ya descarga del módulo RECETA.

Qué hace:
  1. Toma la sábana más reciente (o la indicada con --csv).
  2. Filtra:  Tipo Receta = CONTROLADA  ·  Estado = ENTREGADA
              ·  Bodega Despacha = "FARMACIA AT ABIERTA"
  3. Mapea el código HHHA → F-código ISP + nombre + presentación.
  4. Agrega a la planilla SOLO los folios que no estén ya (dedup por N° Folio).
     No toca filas existentes. La QF completa a mano DV QF y Nombre QF.

Versus el script original de la QF, esta versión:
  · Normaliza los nombres de columna con unicodedata (mapea tildes a la letra
    base). El original borraba la vocal acentuada y reventaba con la sábana de
    SSASUR (Número Folio → "Nmero Folio" → KeyError).
  · Autodescubre la sábana (carpeta maestro). Sin --form, procesa TODOS los
    meses de dispensación presentes en la sábana (no solo el más reciente):
    si un mes no tiene formulario en la carpeta RCh, lo crea desde la
    plantilla en blanco (PLANTILLA_BLANCO) con B7/B8 fijados a ese mes.
  · Filtra por el mes/año de dispensación declarado en cada formulario
    (B7/B8), para no mezclar meses cuando la sábana cruza el cambio de mes.
  · Hace copia de respaldo (.bak) antes de escribir.
  · Maneja el caso "Excel abierto" (PermissionError) sin reventar.

Uso:
  py recetas_cheque.py                      # autodescubre todo, pausa al final
  py recetas_cheque.py --csv <sabana.csv> --form <formulario.xlsx> --no-pause
  py recetas_cheque.py --no-backup
═══════════════════════════════════════════════════════════════════════════
NUNCA toca credenciales. Solo lee la sábana ya descargada y escribe la planilla.
"""
import sys
import os
import argparse
import shutil
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime, date
from glob import glob

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ── CONFIGURACIÓN ────────────────────────────────────────────────────────────
MAESTRO_DIR = os.path.dirname(os.path.abspath(__file__))
PREFIJO_CSV = "informe_completo_recetas"

# verificar_frescura(): blindaje compartido contra datos auto-detectados
# desactualizados (incidente S.52, ver utils_aa.py para el detalle).
sys.path.insert(0, MAESTRO_DIR)
from utils_aa import verificar_frescura

# Carpeta donde la QF mantiene el formulario ISP mensual. Se elige el .xlsx más
# reciente que empiece con este prefijo (maneja el rollover v11_junio → v12_julio).
RCH_DIR     = r"C:\Users\danie\Downloads\Farmacia_AT_Abierta_RCh\Farmacia_AT_Abierta_RCh"
PREFIJO_FORM = "Formulario-Notificacion-Recetas-Cheque"
HOJA_RCH    = "Registro de RCh"
FILA_DATOS  = 12   # los registros empiezan en la fila 12

# Plantilla ISP en blanco (B7/B8 vacíos, sin registros) usada para crear
# automáticamente el formulario del mes siguiente cuando la sábana trae datos
# de un mes que todavía no tiene formulario en RCH_DIR.
PLANTILLA_BLANCO = r"C:\Users\danie\Downloads\02_Farmacia_Recetas_e_Informes_CSV\Formulario-Notificacion-Recetas-Cheque_v11.xlsx"

# ── MAPEO HHHA → F-código ISP + Nombre + Presentación ────────────────────────
MAPEO_HHHA = {
    "212-0009": ("F-1373",  "FENOBARBITAL COMPRIMIDOS 100 mg",                                                                    "COMPRIMIDOS"),
    "212-0052": ("F-24748", "FENTADUR PARCHE TRANSDERMICO 25 mcg/hora (FENTANILO)",                                              "PARCHES"),
    "212-0073": ("F-19539", "PALEXIS RETARD COMPRIMIDOS RECUBIERTOS DE LIBERACION PROLONGADA 50 mg (TAPENTADOL CLORHIDRATO)",    "COMPRIMIDOS"),
    "212-0031": ("F-22274", "VENDAL RETARD COMPRIMIDOS RECUBIERTOS DE LIBERACION PROLONGADA 30 mg (MORFINA CLORHIDRATO TRIHIDRATO)", "CAPSULAS"),
    "214-0597": ("F-7553",  "ARADIX RETARD COMPRIMIDOS DE LIBERACION PROLONGADA 10 mg",                                          "COMPRIMIDOS"),
    "212-0059": ("F-17744", "METILFENIDATO CLORHIDRATO COMPRIMIDOS 10 mg",                                                       "COMPRIMIDOS"),
    "212-0034": ("F-19391", "METADONA CLORHIDRATO COMPRIMIDOS 10 mg",                                                            "COMPRIMIDOS"),
    "212-0083": ("F-27066", "NEUROK CAPSULAS 50 mg (LISDEXANFETAMINA DIMESILATO)",                                               "COMPRIMIDOS"),
    "212-0085": ("F-27069", "NEUROK CAPSULAS 70 mg (LISDEXANFETAMINA DIMESILATO)",                                               "CAPSULAS"),
    "212-0068": ("F-19518", "MORFINA SULFATO SOLUCION ORAL 2%",                                                                  "FRASCO x 20 mL"),
}

# Respaldo por nombre de prescripción cuando el código HHHA no esté en el mapeo.
MAPEO_NOMBRE = {
    "FENTANILO  25 MCG/HORA PARCHE": ("F-24748", "FENTADUR PARCHE TRANSDERMICO 25 mcg/hora (FENTANILO)", "PARCHES"),
}

MESES_ES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# ── NORMALIZACIÓN / UTILIDADES ───────────────────────────────────────────────
def norm_col(c):
    """Tilde → letra base (Número→Numero), sin borrar la vocal."""
    return unicodedata.normalize("NFKD", str(c).strip()).encode("ascii", "ignore").decode()


def autodescubrir_csv(carpeta):
    archivos = glob(os.path.join(carpeta, PREFIJO_CSV + "*.csv"))
    if not archivos:
        raise FileNotFoundError(
            f"No se encontró sábana '{PREFIJO_CSV}*.csv' en:\n  {carpeta}")
    return max(archivos, key=os.path.getmtime)


def get_producto(row):
    hhha = str(row["Codigo Prescripcion HHHA"]).strip() if pd.notna(row["Codigo Prescripcion HHHA"]) else ""
    if hhha in MAPEO_HHHA:
        return MAPEO_HHHA[hhha]
    nombre = str(row["Prescripcion"]).strip()
    if nombre in MAPEO_NOMBRE:
        return MAPEO_NOMBRE[nombre]
    return (None, nombre, None)


def split_rut(s):
    if pd.isna(s):
        return None, None
    s = str(s).strip().replace(" ", "")
    if "-" in s:
        partes = s.split("-")
        rut = partes[0].replace(".", "")
        dv  = partes[1].upper()
        try:
            rut = int(rut)
        except ValueError:
            pass
        return rut, dv
    return s, None


def get_posologia(row):
    dosis  = str(row["Dosis"]).strip()    if pd.notna(row["Dosis"])    else ""
    interv = str(row["Intervalo"]).strip() if pd.notna(row["Intervalo"]) else ""
    per    = str(row["Periodo.1"]).strip() if pd.notna(row["Periodo.1"]) else ""
    obs    = str(row["Observacion Medica Prescripcion"]).strip() if pd.notna(row["Observacion Medica Prescripcion"]) else ""
    partes = [p for p in [dosis, interv, per] if p and p not in ("nan", "NaN")]
    return " ".join(partes) if partes else (obs if obs and obs not in ("nan", "NaN") else "SEGUN INDICACION")


def sanitizar_folio(folio):
    """Corrige un glitch conocido de SSASUR: en el rollover de mes, algunos
    folios quedan exportados con el año pegado al final sin separador
    (ej. 42894 + 2026 -> 428942026). Se detectó el 2026-07-13: de 44 folios
    en la sábana de julio, 2 tenían 9 dígitos y ambos correspondían a
    registros digitados el día 1 del mes.
    Devuelve (folio_saneado, alerta:str|None)."""
    s = str(folio)
    if len(s) <= 6:
        return folio, None
    for anio_str in ("2024", "2025", "2026", "2027", "2028", "2029"):
        if s.endswith(anio_str) and 1 <= len(s) - 4 <= 6:
            corregido = int(s[:-4])
            return corregido, f"Folio {folio} -> {corregido} (año '{anio_str}' pegado, glitch SSASUR)"
    return folio, f"Folio {folio} tiene {len(s)} dígitos (fuera de rango normal, revisar a mano)"


EXCEL_EPOCH = date(1899, 12, 30)

def parse_fecha(s):
    if pd.isna(s) or str(s).strip() in ("", "nan", "NaT"):
        return None
    try:
        return datetime.strptime(str(s).strip()[:10], "%d/%m/%Y").date()
    except Exception:
        return None


def to_excel_date(s):
    d = parse_fecha(s)
    return (d - EXCEL_EPOCH).days if d else None


def leer_folios_existentes(ws):
    folios = set()
    for row in ws.iter_rows(min_row=FILA_DATOS, values_only=True):
        if row[0] is not None:
            try:
                folios.add(int(row[0]))
            except (ValueError, TypeError):
                pass
    return folios


def ultima_fila_con_datos(ws):
    """Última fila con Folio real (columna A). La plantilla en blanco trae
    fórmulas precargadas hasta la fila ~5026, así que ws.max_row NO sirve
    para ubicar dónde termina la data — hay que buscar la última con valor."""
    ultima = FILA_DATOS - 1
    for row in ws.iter_rows(min_row=FILA_DATOS, values_only=False):
        if row[0].value is not None:
            ultima = row[0].row
    return ultima


def leer_periodo_form(ws):
    """Año/mes de dispensación declarados en el formulario (B7/B8). (año, mes)|(None,None)."""
    try:
        anio = ws["B7"].value
        mes  = ws["B8"].value
        anio = int(str(anio).strip()) if anio is not None else None
        mes  = MESES_ES.get(str(mes).strip().upper()) if mes is not None else None
        return anio, mes
    except Exception:
        return None, None


def listar_formularios(carpeta):
    """(path, año, mes) de cada formulario ISP con periodo declarado en la carpeta."""
    out = []
    for f in glob(os.path.join(carpeta, PREFIJO_FORM + "*.xlsx")):
        if os.path.basename(f).startswith("~$"):
            continue
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            anio, mes = leer_periodo_form(wb[HOJA_RCH])
            wb.close()
            if anio and mes:
                out.append((f, anio, mes))
        except Exception:
            pass
    return out


def crear_formulario_mes(anio, mes_num, carpeta):
    """Copia la plantilla en blanco, fija B7/B8 y la deja lista en `carpeta`."""
    if not os.path.exists(PLANTILLA_BLANCO):
        raise FileNotFoundError("No se encontró la plantilla en blanco:\n  " + PLANTILLA_BLANCO)
    mes_nombre = next(k for k, v in MESES_ES.items() if v == mes_num)
    destino = os.path.join(carpeta, f"{PREFIJO_FORM}_v11_{mes_nombre.lower()}.xlsx")
    if os.path.exists(destino):
        return destino
    shutil.copy2(PLANTILLA_BLANCO, destino)
    wb = load_workbook(destino)
    ws = wb[HOJA_RCH]
    ws["B7"] = anio
    ws["B8"] = mes_nombre
    wb.save(destino)
    print(f"  [nuevo] Formulario creado para {mes_nombre} {anio}: {os.path.basename(destino)}")
    return destino


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Registro ISP de recetas cheque — Farmacia AT Abierta")
    ap.add_argument("--csv",  help="Sábana CSV (default: la más reciente en la carpeta maestro)")
    ap.add_argument("--form", help="Formulario ISP .xlsx (default: el más reciente en la carpeta RCh)")
    ap.add_argument("--carpeta-csv",  default=MAESTRO_DIR, help="Carpeta donde buscar la sábana")
    ap.add_argument("--carpeta-form", default=RCH_DIR,     help="Carpeta donde buscar el formulario")
    ap.add_argument("--no-backup", action="store_true", help="No crear copia .bak antes de escribir")
    ap.add_argument("--no-pause",  action="store_true", help="No esperar Enter al final (modo automático)")
    ap.add_argument("--sin-filtro-mes", action="store_true", help="No filtrar por el mes del formulario")
    a = ap.parse_args()

    print("=" * 62)
    print("  Registro ISP Recetas Cheque — Farmacia AT Abierta")
    print("=" * 62)

    csv_path  = a.csv  or autodescubrir_csv(a.carpeta_csv)
    form_path = a.form
    print("  Sábana     : " + os.path.basename(csv_path))
    if form_path:
        print("  Formulario : " + os.path.basename(form_path))
    else:
        print("  Formulario : (autodetectado por mes de dispensación)")

    # 1) Leer y normalizar columnas de la sábana.
    df = pd.read_csv(csv_path, encoding="latin-1", sep=None, engine="python")
    df = df.rename(columns={c: norm_col(c) for c in df.columns})

    requeridas = ["Tipo Receta", "Estado", "Numero Folio", "Bodega Despacha",
                  "Codigo Prescripcion HHHA", "Prescripcion"]
    faltan = [c for c in requeridas if c not in df.columns]
    if faltan:
        raise KeyError("Faltan columnas en la sábana tras normalizar: " + ", ".join(faltan))

    # Sábana auto-detectada (sin --csv explícito) → blindaje contra descargas
    # viejas/fallidas de AUTO_SSASUR (mismo guard que centinela_reporte.py,
    # incidente S.52): si el registro más reciente tiene más de 10 días,
    # abortar en vez de reportar en silencio "sin folios nuevos" cuando en
    # realidad la sábana nunca se actualizó.
    if not a.csv and "Fecha Entrega Receta" in df.columns:
        fecha_max_sabana = pd.to_datetime(df["Fecha Entrega Receta"], dayfirst=True, errors="coerce").max()
        fecha_max_sabana = fecha_max_sabana.date() if pd.notna(fecha_max_sabana) else None
        verificar_frescura(fecha_max_sabana, "sábana de recetas (auto-detectada)")

    # 2) Filtrar recetas cheque AT Abierta con dispensación real.
    # El Estado de la receta (PENDIENTE/CERRADA INCOMPLETA/ENTREGADA) refleja
    # el ciclo de vida administrativo, no si esta línea ya se despachó. Una
    # receta crónica puede quedar abierta o cerrarse incompleta sin que eso
    # anule la entrega que ya ocurrió: si Cantidad Entregada > 0 hubo
    # dispensación real de un controlado y debe quedar en el libro ISP.
    rch = df[
        (df["Tipo Receta"]     == "CONTROLADA")       &
        (df["Cantidad Entregada"].fillna(0) > 0)       &
        (df["Numero Folio"].notna())                   &
        (df["Bodega Despacha"] == "FARMACIA AT ABIERTA")
    ].copy()
    print("\n  Recetas cheque AT Abierta en la sábana: " + str(len(rch)))
    if rch.empty:
        print("\n  No hay recetas cheque AT Abierta en esta sábana.")
        if not a.no_pause:
            input("\nPresiona Enter para cerrar...")
        return

    # 3) Derivar campos.
    rch[["F_COD", "NOMBRE_PROD", "PRESENTACION"]] = rch.apply(lambda r: pd.Series(get_producto(r)), axis=1)
    rch[["RUT_MED", "DV_MED"]] = rch["RUN Profesional"].apply(lambda x: pd.Series(split_rut(x)))
    rch[["RUT_PAC", "DV_PAC"]] = rch["RUN"].apply(lambda x: pd.Series(split_rut(x)))
    rch[["RUT_ADQ", "DV_ADQ"]] = rch["Run Persona Retira"].apply(lambda x: pd.Series(split_rut(x)))
    rch["RUT_QF"]        = rch["Usuario Creacion Registro"].apply(lambda x: int(str(x).strip()) if pd.notna(x) else None)
    rch["POSOLOGIA"]     = rch.apply(get_posologia, axis=1)
    rch["FECHA_PRESC_N"] = rch["Fecha Atencion"].apply(to_excel_date)
    rch["FECHA_DISP_N"]  = rch["Fecha Entrega Receta"].apply(to_excel_date)
    rch["FECHA_DISP_D"]  = rch["Fecha Entrega Receta"].apply(parse_fecha)

    folios_raw = rch["Numero Folio"].apply(lambda x: int(x) if pd.notna(x) else None)
    saneados = folios_raw.apply(lambda f: sanitizar_folio(f) if f is not None else (None, None))
    rch["FOLIO_INT"] = saneados.apply(lambda t: t[0])
    for _, alerta in saneados:
        if alerta:
            print("  [ALERTA folio] " + alerta)

    # 4) Modo single-form (--form explícito): comportamiento clásico, un solo
    #    formulario, filtrado por su propio periodo B7/B8.
    if a.form:
        actualizar_formulario(form_path, rch, sin_filtro_mes=a.sin_filtro_mes, no_backup=a.no_backup)
        print("\n  Recuerda completar a mano: DV QF y Nombre QF.")
        if not a.no_pause:
            input("\nPresiona Enter para cerrar...")
        return

    # 4') Modo autodescubrimiento: procesar TODOS los meses presentes en la
    #     sábana, no solo el del formulario más reciente. Si un mes no tiene
    #     formulario todavía, se crea desde la plantilla en blanco.
    meses_presentes = sorted({(d.year, d.month) for d in rch["FECHA_DISP_D"] if d})
    if not meses_presentes:
        print("\n  No pude determinar el mes de dispensación de ningún registro.")
        if not a.no_pause:
            input("\nPresiona Enter para cerrar...")
        return

    formularios = listar_formularios(a.carpeta_form)
    for anio, mes in meses_presentes:
        nombre_mes = next(k for k, v in MESES_ES.items() if v == mes)
        path_existente = next((f for f, ay, m in formularios if ay == anio and m == mes), None)
        ruta = path_existente or crear_formulario_mes(anio, mes, a.carpeta_form)
        print(f"\n  ── {nombre_mes} {anio} " + "─" * (40 - len(nombre_mes)))
        print("  Formulario : " + os.path.basename(ruta))
        actualizar_formulario(ruta, rch, sin_filtro_mes=False, no_backup=a.no_backup)

    print("\n  Recuerda completar a mano: DV QF y Nombre QF.")
    if not a.no_pause:
        input("\nPresiona Enter para cerrar...")


def actualizar_formulario(form_path, rch, sin_filtro_mes, no_backup):
    """Llena `form_path` con los folios nuevos de `rch` que caen en su periodo B7/B8."""
    if not os.path.exists(form_path):
        raise FileNotFoundError("No se encontró el formulario:\n  " + form_path)
    wb = load_workbook(form_path)
    if HOJA_RCH not in wb.sheetnames:
        raise KeyError(f"El formulario no tiene la hoja '{HOJA_RCH}'. Hojas: {wb.sheetnames}")
    ws = wb[HOJA_RCH]

    folios_existentes = leer_folios_existentes(ws)
    print("  Registros ya en la planilla: " + str(len(folios_existentes)))

    # Filtrar por mes/año del formulario (evita mezclar meses).
    anio_f, mes_f = leer_periodo_form(ws)
    if not sin_filtro_mes and anio_f and mes_f:
        antes = len(rch)
        rch = rch[rch["FECHA_DISP_D"].apply(
            lambda d: bool(d) and d.year == anio_f and d.month == mes_f)].copy()
        omit = antes - len(rch)
        nombre_mes = next((k for k, v in MESES_ES.items() if v == mes_f), str(mes_f))
        print(f"  Filtro periodo formulario: {nombre_mes} {anio_f}  (omitidas {omit} de otro mes)")
    elif not sin_filtro_mes:
        print("  [aviso] No pude leer el periodo del formulario (B7/B8) — agrego sin filtrar por mes.")

    # Quedarnos con los folios nuevos.
    nuevos = rch[~rch["FOLIO_INT"].isin(folios_existentes)].copy()
    nuevos = nuevos.drop_duplicates(subset=["FOLIO_INT"])
    print("  Registros NUEVOS a agregar: " + str(len(nuevos)))
    if nuevos.empty:
        print("  La planilla ya está al día.")
        return

    # Respaldo antes de escribir.
    if not no_backup:
        bak = form_path + ".bak"
        try:
            shutil.copy2(form_path, bak)
            print("  Respaldo: " + os.path.basename(bak))
        except OSError as e:
            print(f"  [aviso] No pude crear respaldo: {e}")

    # Escribir filas nuevas.
    sample_font = Font(name="Arial", size=10)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    primera = ultima_fila_con_datos(ws) + 1
    sin_fcod = 0

    for idx, (_, row) in enumerate(nuevos.iterrows()):
        er = primera + idx
        cant = int(row["Cantidad Entregada"]) if pd.notna(row["Cantidad Entregada"]) else None
        if not row["F_COD"]:
            sin_fcod += 1
        valores = [
            row["FOLIO_INT"],
            row["RUT_MED"], row["DV_MED"],
            row["RUT_PAC"], row["DV_PAC"],
            str(row["Direccion"]) if pd.notna(row["Direccion"]) else None,
            str(row["Comuna"])    if pd.notna(row["Comuna"])    else None,
            row["F_COD"],
            "=VLOOKUP(H" + str(er) + ",OCULTA!A:B,2,0)",
            row["PRESENTACION"],
            cant,
            row["POSOLOGIA"],
            row["FECHA_PRESC_N"],
            row["RUT_ADQ"], row["DV_ADQ"],
            row["FECHA_DISP_N"],
            row["RUT_QF"], None, None,
            cant,
            row["PRESENTACION"],
            1,
        ]
        for ci, val in enumerate(valores):
            cell = ws.cell(row=er, column=ci + 1)
            cell.value  = val
            cell.font   = sample_font
            cell.border = border
            if ci in (12, 15) and val is not None:
                cell.number_format = "DD/MM/YYYY"

    # Guardar (manejar Excel abierto).
    try:
        wb.save(form_path)
    except PermissionError:
        print("\n  [ERROR] No pude guardar: el formulario está ABIERTO en Excel.")
        print("          Ciérralo y vuelve a ejecutar.")
        return

    print("  ✓ Se agregaron " + str(len(nuevos)) + " registros nuevos.")
    print("  Total en planilla ahora: " + str(len(folios_existentes) + len(nuevos)))
    if sin_fcod:
        print(f"  [revisar] {sin_fcod} sin F-código (HHHA no mapeado) — completar a mano en la planilla.")
    print("  Registros agregados por producto:")
    for prod, cnt in nuevos["Prescripcion"].value_counts().items():
        print("    " + str(prod) + ": " + str(cnt))
    print("  Planilla: " + os.path.abspath(form_path))


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, KeyError) as e:
        print("\nERROR: " + str(e))
        if "--no-pause" not in sys.argv:
            input("\nPresiona Enter para cerrar...")
    except Exception as e:
        print("\nERROR inesperado: " + str(e))
        import traceback
        traceback.print_exc()
        if "--no-pause" not in sys.argv:
            input("\nPresiona Enter para cerrar...")
