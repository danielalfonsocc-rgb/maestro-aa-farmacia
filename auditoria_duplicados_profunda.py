#!/usr/bin/env python3
"""
auditoria_duplicados_profunda.py
════════════════════════════════════════════════════════════════════════════════
Auditoría PROFUNDA de prescripciones duplicadas — Farmacia AT Abierta
Hospital de Pitrufquén (SSASur)

Qué agrega sobre recetas_duplicadas.py:
  1. VIGENCIA ACTUAL — detecta si el doble retiro está ACTIVO HOY
     (≥2 prescripciones del cluster con cobertura que llega a hoy).
  2. INICIO DEL DOBLE RETIRO — fecha exacta en que comenzó el solapamiento.
  3. DÍAS EN DOBLE RETIRO — acumulado desde el inicio hasta hoy (o hasta cierre).
  4. PROPUESTAS IA (Claude Haiku) — acción concreta por cada caso activo:
       URGENTE / REVISAR / INFORMAR / MONITOREAR
     Con plazo, propuesta y alertas clínicas específicas.
     Los RUTs nunca salen del proceso local; se envían IDs anónimos (SHA-256).

Hojas Excel:
  1. Activos Hoy       — casos vigentes hoy (prioridad máxima)
  2. Histórico Completo — todos los casos ordenados por prioridad
  3. Detalle Prescripciones — cada prescripción dentro de cada caso
  4. Por Medicamento   — agregados por fármaco
  5. Metodología       — documentación

Uso:
    py auditoria_duplicados_profunda.py
    py auditoria_duplicados_profunda.py --salida mi_reporte.xlsx
    py auditoria_duplicados_profunda.py --rapido      # solo CSV más reciente
    py auditoria_duplicados_profunda.py --sin-ia      # omite llamadas a Claude
════════════════════════════════════════════════════════════════════════════════
"""
import argparse
import hashlib
import json
import os
import re
import sys
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aa_colors import TEAL, ROJO, NARANJA, AMBAR, VERDE, GRIS
from utils_aa import norm_erp, HOMOLOGACION, cargar_recetas_csv, setup_stdout

setup_stdout()
WORK = os.path.dirname(os.path.abspath(__file__))

BODEGA_OBJETIVO    = "FARMACIA AT ABIERTA"
ESTADOS_EXCLUIDOS  = {"ANULADO", "RECHAZADO", "REEMPLAZADO"}
DIAS_POR_CUOTA     = 30
MIN_EVENTOS        = 2
MODELO_HAIKU       = "claude-haiku-4-5-20251001"
HOY                = date.today()

THIN   = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Colores por prioridad: (texto, fondo)
PRIO_COLOR = {
    "URGENTE":   ("DC2626", "FEE2E2"),
    "REVISAR":   ("C2410C", "FFF3E0"),
    "INFORMAR":  ("B45309", "FFFBEB"),
    "HISTORIAL": ("6B7280", "F9FAFB"),
}


# ── Carga y preparación ──────────────────────────────────────────────────────

def cargar_y_preparar(solo_ultimo: bool) -> pd.DataFrame:
    cols = [
        "ID Receta Detalle", "RUN", "Nombre", "Apellido Paterno", "Apellido Materno",
        "Prescripción", "Número Receta", "Periodo",
        "Fecha Atención", "Fecha Entrega Receta", "Estado Prescripción",
        "RUN Profesional", "Nombre Profesional",
        "Apellido Paterno Profesional", "Apellido Materno Profesional",
        "Especialidad", "Bodega Despacha",
        "Cantidad Recetada", "Cantidad Entregada",
    ]
    try:
        rec = cargar_recetas_csv(WORK, cols=cols, solo_ultimo=solo_ultimo)
    except FileNotFoundError:
        print("[AVISO] No hay CSV de recetas. Ejecuta AUTO_SSASUR.bat primero.")
        sys.exit(0)

    rec = rec.copy()
    _bod = rec["Bodega Despacha"].fillna("").apply(norm_erp)
    _est = rec["Estado Prescripción"].fillna("").str.upper().str.strip()
    rec  = rec[(_bod == norm_erp(BODEGA_OBJETIVO)) & (~_est.isin(ESTADOS_EXCLUIDOS))].copy()

    rec["_med"]      = rec["Prescripción"].fillna("").apply(norm_erp).map(
                           lambda x: HOMOLOGACION.get(x, x))
    rec["_est"]      = _est.loc[rec.index]
    fa               = pd.to_datetime(rec["Fecha Atención"],       dayfirst=True, errors="coerce")
    fe               = pd.to_datetime(rec["Fecha Entrega Receta"], dayfirst=True, errors="coerce")
    rec["_fecha"]    = fa.fillna(fe)
    rec["_paciente"] = (
        rec["Nombre"].fillna("") + " " + rec["Apellido Paterno"].fillna("") +
        " " + rec["Apellido Materno"].fillna("")
    ).str.strip().str.title()
    rec["_medico"]   = (
        rec["Nombre Profesional"].fillna("") + " " +
        rec["Apellido Paterno Profesional"].fillna("") + " " +
        rec["Apellido Materno Profesional"].fillna("")
    ).str.strip().str.title()
    rec["_run_prof"] = rec["RUN Profesional"].fillna("").str.strip()
    rec["_esp"]      = rec["Especialidad"].fillna("").str.strip()
    rec["_cant_r"]   = pd.to_numeric(rec["Cantidad Recetada"],  errors="coerce").fillna(0)
    rec["_cant_e"]   = pd.to_numeric(rec["Cantidad Entregada"], errors="coerce").fillna(0)

    rec = rec.dropna(subset=["_fecha"])
    rec = rec[rec["Número Receta"].fillna("").str.strip() != ""]
    rec = rec[rec["_med"] != ""]

    print(f"Filas Farmacia AT Abierta válidas: {len(rec):,}  "
          f"| rango: {rec['_fecha'].min():%d-%m-%Y} a {rec['_fecha'].max():%d-%m-%Y}")
    return rec


def _parse_cuotas(serie) -> int:
    n = 1
    for v in serie.dropna():
        m = re.search(r"de\s+(\d+)", str(v))
        if m:
            n = max(n, int(m.group(1)))
    return n


def nivel_evento(rec: pd.DataFrame) -> pd.DataFrame:
    """Colapsa cuotas mensuales de una misma indicación crónica a 1 evento."""
    def _primero(s):
        s2 = s[s.astype(str).str.strip() != ""]
        return s2.mode().iloc[0] if not s2.empty else ""

    g = rec.groupby(["RUN", "_med", "_run_prof", "_fecha"]).agg(
        paciente   = ("_paciente", _primero),
        med_nombre = ("Prescripción", _primero),
        cuotas     = ("Periodo",    _parse_cuotas),
        medico     = ("_medico",    _primero),
        esp        = ("_esp",       _primero),
        cant_r     = ("_cant_r",    "sum"),
        cant_e     = ("_cant_e",    "sum"),
        n_cuotas   = ("Número Receta", "nunique"),
        estado     = ("_est",       _primero),
    ).reset_index().rename(columns={"_run_prof": "run_prof", "_fecha": "start"})
    g["cob_dias"] = g["cuotas"].clip(lower=1) * DIAS_POR_CUOTA
    return g


# ── Detección de solapamientos ───────────────────────────────────────────────

def _clusters_solapados(sub: pd.DataFrame):
    """Devuelve listas de eventos con coberturas solapadas (≥2 eventos)."""
    sub  = sub.sort_values("start")
    recs = sub.to_dict("records")
    out, cur, end = [], [], None
    for r in recs:
        r_end = r["start"] + timedelta(days=int(r["cob_dias"]))
        if not cur:
            cur, end = [r], r_end
        elif r["start"] <= end:
            cur.append(r)
            end = max(end, r_end)
        else:
            if len(cur) >= MIN_EVENTOS:
                out.append(cur)
            cur, end = [r], r_end
    if len(cur) >= MIN_EVENTOS:
        out.append(cur)
    return out


def _fin_cob(evento):
    return pd.Timestamp(evento["start"]) + timedelta(days=int(evento["cob_dias"]))


def _vigente_hoy(cl) -> bool:
    """True si ≥2 prescripciones del cluster cubren hasta hoy (doble retiro activo)."""
    hoy = pd.Timestamp(HOY)
    return sum(1 for r in cl if _fin_cob(r) >= hoy) >= 2


def _inicio_doble_retiro(cl) -> pd.Timestamp:
    """Fecha en que comenzó el solapamiento = inicio de la 2ª prescripción."""
    return sorted(cl, key=lambda r: r["start"])[1]["start"]


def _dias_en_doble_retiro(cl) -> int:
    """Días acumulados en doble retiro (desde inicio hasta hoy o hasta cierre)."""
    inicio = _inicio_doble_retiro(cl)
    hoy    = pd.Timestamp(HOY)
    sorted_cl = sorted(cl, key=lambda r: r["start"])
    # Fin aproximado = cuando el primero de los dos primeros deja de cubrir
    fin0 = _fin_cob(sorted_cl[0])
    fin1 = _fin_cob(sorted_cl[1])
    fin_solape = min(fin0, fin1)
    fin = min(fin_solape, hoy)
    return max(0, (fin - inicio).days)


def _prioridad_base(vigente: bool, n_medicos: int) -> str:
    if vigente and n_medicos >= 2:
        return "URGENTE"
    if vigente:
        return "REVISAR"
    if n_medicos >= 2:
        return "INFORMAR"
    return "HISTORIAL"


# ── Construcción de tablas ───────────────────────────────────────────────────

def construir_casos(rl: pd.DataFrame):
    resumen  = []
    detalle  = []
    cid      = 0

    for (run, med), sub in rl.groupby(["RUN", "_med"]):
        if len(sub) < MIN_EVENTOS:
            continue
        for cl in _clusters_solapados(sub):
            cid       += 1
            run_profs  = sorted({r["run_prof"] for r in cl if r["run_prof"]})
            medicos    = sorted({r["medico"]   for r in cl if r["medico"]})
            esps       = sorted({r["esp"]      for r in cl if r["esp"]})
            n_med      = max(len(run_profs), 1)

            fechas    = [r["start"]   for r in cl]
            entregas  = [r["cant_e"]  for r in cl]
            con_dup   = int(sum(entregas))
            sin_dup   = int(max(entregas)) if entregas else 0

            vigente    = _vigente_hoy(cl)
            inicio_dup = _inicio_doble_retiro(cl)
            dias_dup   = _dias_en_doble_retiro(cl)

            sorted_cl  = sorted(cl, key=lambda r: r["start"])
            ultima_fin = _fin_cob(sorted_cl[-1])
            prio       = _prioridad_base(vigente, n_med)

            resumen.append({
                "CID"                  : cid,
                "RUN"                  : run,
                "Paciente"             : cl[0]["paciente"],
                "Medicamento"          : cl[0]["med_nombre"],
                "Prioridad"            : prio,
                "Vigente hoy"          : "Sí" if vigente else "No",
                "Tipo duplicación"     : "Distintos médicos" if n_med >= 2 else "Mismo médico",
                "N° médicos"           : n_med,
                "Médico(s)"            : " ; ".join(medicos),
                "Especialidad(es)"     : " ; ".join(e for e in esps if e) or "—",
                "1ª prescripción"      : min(fechas),
                "Inicio doble retiro"  : inicio_dup,
                "Días en doble retiro" : dias_dup,
                "Fin est. última receta": ultima_fin,
                "N° prescripciones"    : len(cl),
                "Total recetado"       : int(sum(r["cant_r"] for r in cl)),
                "Consumo c/ dup"       : con_dup,
                "Consumo s/ dup"       : sin_dup,
                "Exceso (uds)"         : con_dup - sin_dup,
                "Acción IA"            : "",
                "Propuesta IA"         : "",
            })

            for r in sorted(cl, key=lambda x: x["start"]):
                fin_r  = _fin_cob(r)
                activa = "Sí" if fin_r >= pd.Timestamp(HOY) else "No"
                detalle.append({
                    "CID"               : cid,
                    "RUN"               : run,
                    "Paciente"          : r["paciente"],
                    "Medicamento"       : r["med_nombre"],
                    "Fecha prescripción": r["start"],
                    "Médico"            : r["medico"],
                    "RUN médico"        : r["run_prof"],
                    "Especialidad"      : r["esp"] or "—",
                    "Cuotas"            : int(r["cuotas"]),
                    "Cobertura (días)"  : int(r["cob_dias"]),
                    "Fin est. cobertura": fin_r,
                    "Vigente hoy"       : activa,
                    "Cant. recetada"    : int(r["cant_r"]),
                    "Cant. entregada"   : int(r["cant_e"]),
                })

    return pd.DataFrame(resumen), pd.DataFrame(detalle)


# ── Propuestas IA (Claude Haiku) ─────────────────────────────────────────────

_SYSTEM_HAIKU = """\
Eres un Químico Farmacéutico (QF) clínico del Hospital de Pitrufquén (SSASur, Chile).
Se te presentan casos de PRESCRIPCIÓN DUPLICADA detectados en Farmacia AT Abierta.
Genera UNA propuesta de acción concreta por cada caso.

INSTRUCCIONES:
- Devuelve SOLO un array JSON, sin texto adicional ni bloques de código.
- Un objeto por caso con exactamente estas claves:
  {
    "cid": <entero>,
    "accion": "URGENTE" | "REVISAR" | "INFORMAR" | "MONITOREAR",
    "plazo": "Inmediato (hoy)" | "Esta semana" | "Próxima visita" | "Registro",
    "propuesta": "<acción concreta, máx 220 caracteres>",
    "alertas": ["<alerta1>", ...]  // 0 a 3 alertas clínicas específicas
  }

CRITERIOS:
- URGENTE + Inmediato: distintos médicos + vigente hoy → riesgo real de sobrestock o interacción.
- REVISAR + Esta semana: vigente hoy + mismo médico, o distintos médicos recientes.
- INFORMAR + Próxima visita: duplicado histórico resuelto, distintos médicos.
- MONITOREAR + Registro: duplicado histórico cerrado, mismo médico, bajo exceso.

Medicamentos de alta alerta (psicotrópicos, opioides, insulinas, anticoagulantes,
hipoglicemiantes, estatinas en dosis altas) → subir un nivel de urgencia.

Si días_doble_retiro > 60 y vigente_hoy=true → mencionar "patrón sostenido" en propuesta.

No menciones pac_id, RUTs ni ningún identificador personal. La propuesta debe ser
concisa, orientada a la acción del QF, y en español.
"""


def _anon_run(run: str, run_map: dict, anon_map: dict) -> str:
    if run not in run_map:
        h = hashlib.sha256(str(run).encode()).hexdigest()[:8]
        run_map[run] = h
        anon_map[h]  = run
    return run_map[run]


def generar_propuestas_haiku(df_res: pd.DataFrame) -> dict:
    """
    Llama a Haiku con los casos URGENTE/REVISAR/INFORMAR (activos o distintos médicos).
    Retorna dict: cid -> {accion, plazo, propuesta, alertas}
    """
    import anthropic

    filtro = df_res["Prioridad"].isin({"URGENTE", "REVISAR", "INFORMAR"})
    subset = df_res[filtro]
    if subset.empty:
        return {}

    run_map  = {}
    anon_map = {}

    casos_anon = []
    for _, row in subset.iterrows():
        casos_anon.append({
            "cid"             : int(row["CID"]),
            "pac_id"          : _anon_run(str(row["RUN"]), run_map, anon_map),
            "medicamento"     : str(row["Medicamento"])[:70],
            "tipo"            : str(row["Tipo duplicación"]),
            "vigente_hoy"     : row["Vigente hoy"] == "Sí",
            "n_medicos"       : int(row["N° médicos"]),
            "n_prescripciones": int(row["N° prescripciones"]),
            "dias_doble_retiro": int(row["Días en doble retiro"]),
            "exceso_unidades" : int(row["Exceso (uds)"]),
            "especialidades"  : str(row["Especialidad(es)"])[:60],
        })

    client    = anthropic.Anthropic()
    resultado = {}
    batch_sz  = 25

    for i in range(0, len(casos_anon), batch_sz):
        lote = casos_anon[i:i + batch_sz]
        n_lote = (len(casos_anon) - 1) // batch_sz + 1
        print(f"  [IA] Lote {i // batch_sz + 1}/{n_lote} ({len(lote)} casos)...")

        resp = client.messages.create(
            model      = MODELO_HAIKU,
            max_tokens = 4096,
            system     = _SYSTEM_HAIKU,
            messages   = [{"role": "user", "content":
                           f"Casos ({len(lote)}):\n{json.dumps(lote, ensure_ascii=False, indent=2)}"}],
        )
        texto = resp.content[0].text.strip()
        try:
            m = re.search(r"\[.*\]", texto, re.DOTALL)
            if m:
                for obj in json.loads(m.group()):
                    resultado[obj["cid"]] = obj
        except (json.JSONDecodeError, KeyError) as e:
            print(f"  [AVISO IA] Error parseando lote {i // batch_sz + 1}: {e}")

    return resultado


def aplicar_propuestas(df_res: pd.DataFrame, propuestas: dict) -> pd.DataFrame:
    df = df_res.copy()
    for idx, row in df.iterrows():
        p = propuestas.get(int(row["CID"]))
        if not p:
            continue
        df.at[idx, "Acción IA"]   = p.get("accion", "")
        alertas = p.get("alertas", [])
        texto   = p.get("propuesta", "")
        plazo   = p.get("plazo",    "")
        if plazo:
            texto = f"[{plazo}] {texto}"
        if alertas:
            texto += " | " + " · ".join(alertas)
        df.at[idx, "Propuesta IA"] = texto[:500]
    return df


# ── Escritura Excel ──────────────────────────────────────────────────────────

def _encabezado_hoja(ws, texto: str, color: str, n_cols: int):
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = texto
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _fila_stats(ws, stats: list[tuple], n_cols: int, fila: int = 2):
    """Escribe una fila de estadísticas rápidas tipo 'URGENTE: 5 | REVISAR: 12 ...'"""
    txt = "   |   ".join(f"{k}: {v}" for k, v in stats)
    ws.merge_cells(f"A{fila}:{get_column_letter(n_cols)}{fila}")
    c = ws.cell(row=fila, column=1, value=txt)
    c.font      = Font(bold=True, color="374151", size=10)
    c.fill      = PatternFill("solid", fgColor="F1F5F9")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 20


def _escribir_tabla(ws, df: pd.DataFrame, fila_inicio: int,
                    anchos: dict | None = None,
                    fechas: list | None = None,
                    prio_col: str | None = None,
                    vigente_col: str | None = None):
    """Escribe un DataFrame como tabla openpyxl desde fila_inicio."""
    fechas = fechas or []
    cols   = list(df.columns)
    col_ix = {c: i + 1 for i, c in enumerate(cols)}

    # Cabecera
    for c_i, col in enumerate(cols, 1):
        cell = ws.cell(row=fila_inicio, column=c_i, value=col)
        cell.fill      = PatternFill("solid", fgColor=TEAL)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[fila_inicio].height = 28
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)

    # Filas de datos
    for r_off, (_, row) in enumerate(df.iterrows()):
        r = fila_inicio + 1 + r_off
        ws.append(list(row.values))

        prio         = str(row.get(prio_col, "")) if prio_col else ""
        txt_c, bg_c  = PRIO_COLOR.get(prio, ("1F2937", "FFFFFF"))

        for c_i, col in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c_i)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(col in ("Propuesta IA", "Médico(s)", "Médico")))
            if prio_col and col == prio_col:
                cell.font  = Font(bold=True, color=txt_c)
                cell.fill  = PatternFill("solid", fgColor=bg_c)
            elif vigente_col and col == vigente_col:
                if str(cell.value) == "Sí":
                    cell.font = Font(bold=True, color=ROJO)
                else:
                    cell.fill = PatternFill("solid", fgColor="F3F4F6")
            elif r % 2 == 0:
                cell.fill  = PatternFill("solid", fgColor=GRIS)

        altura = 45 if (prio_col and row.get("Propuesta IA", "")) else 22
        ws.row_dimensions[r].height = altura

    # Fechas
    for fcol in fechas:
        if fcol in col_ix:
            for r in range(fila_inicio + 1, ws.max_row + 1):
                ws.cell(row=r, column=col_ix[fcol]).number_format = "DD-MM-YYYY"

    # Anchos de columna
    for i, col in enumerate(cols, 1):
        w = (anchos or {}).get(col)
        if w is None:
            vals = df[col].astype(str)
            w = min(max(len(str(col)) + 2, vals.str.len().max() + 2 if len(df) else 12), 52)
        ws.column_dimensions[get_column_letter(i)].width = w

    # Autofiltro
    ultima = ws.max_row
    ws.auto_filter.ref = f"A{fila_inicio}:{get_column_letter(len(cols))}{ultima}"


ANCHOS_RES = {
    "CID": 6, "Paciente": 26, "Medicamento": 40,
    "Prioridad": 12, "Vigente hoy": 10, "Tipo duplicación": 18,
    "N° médicos": 9, "Médico(s)": 36, "Especialidad(es)": 22,
    "Inicio doble retiro": 16, "Días en doble retiro": 16,
    "Fin est. última receta": 18, "N° prescripciones": 13,
    "Consumo c/ dup": 14, "Consumo s/ dup": 14, "Exceso (uds)": 12,
    "Acción IA": 12, "Propuesta IA": 62,
}

FECHAS_RES  = ["1ª prescripción", "Inicio doble retiro", "Fin est. última receta"]
FECHAS_DET  = ["Fecha prescripción", "Fin est. cobertura"]


def exportar_excel(resumen: pd.DataFrame, detalle: pd.DataFrame,
                   n_lineas: int, rango: tuple, dest: str):
    wb = Workbook()

    n_total    = len(resumen)
    n_activos  = int((resumen["Vigente hoy"] == "Sí").sum()) if n_total else 0
    n_urgente  = int((resumen["Prioridad"]   == "URGENTE").sum()) if n_total else 0
    n_revisar  = int((resumen["Prioridad"]   == "REVISAR").sum()) if n_total else 0
    n_informar = int((resumen["Prioridad"]   == "INFORMAR").sum()) if n_total else 0
    n_hist     = int((resumen["Prioridad"]   == "HISTORIAL").sum()) if n_total else 0
    n_dist     = int((resumen["Tipo duplicación"] == "Distintos médicos").sum()) if n_total else 0
    exceso_tot = int(resumen["Exceso (uds)"].sum()) if n_total else 0

    # ── Hoja 1: Activos Hoy ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Activos Hoy"

    activos = resumen[resumen["Vigente hoy"] == "Sí"].copy()
    _orden  = {"URGENTE": 0, "REVISAR": 1, "INFORMAR": 2, "HISTORIAL": 3}
    activos["_o"] = activos["Prioridad"].map(_orden)
    activos = activos.sort_values(["_o", "Días en doble retiro"], ascending=[True, False])
    activos = activos.drop(columns=["_o", "RUN"], errors="ignore")

    n_cols1 = max(len(activos.columns), 10)
    _encabezado_hoja(ws1,
                     f"PRESCRIPCIONES DUPLICADAS ACTIVAS HOY  ·  {HOY:%d-%m-%Y}  ·  "
                     f"{n_activos} caso(s) vigentes",
                     ROJO, n_cols1)
    _fila_stats(ws1,
                [("URGENTE", n_urgente), ("REVISAR", n_revisar),
                 ("Distintos médicos", n_dist), ("Exceso total ud", exceso_tot)],
                n_cols1, fila=2)

    if len(activos):
        _escribir_tabla(ws1, activos, fila_inicio=3,
                        anchos=ANCHOS_RES, fechas=FECHAS_RES,
                        prio_col="Prioridad", vigente_col="Vigente hoy")
    else:
        ws1.cell(row=3, column=1, value="No hay casos vigentes hoy.").font = Font(
            italic=True, color="6B7280", size=11)

    # ── Hoja 2: Histórico Completo ─────────────────────────────────────────
    ws2 = wb.create_sheet("Histórico Completo")
    _encabezado_hoja(ws2,
                     f"HISTORIAL COMPLETO DE DUPLICADOS  ·  {n_total} casos  ·  "
                     f"Rango: {rango[0]:%d-%m-%Y} → {rango[1]:%d-%m-%Y}",
                     TEAL, max(len(resumen.columns), 10))
    _fila_stats(ws2,
                [("Total casos", n_total), ("Activos hoy", n_activos),
                 ("URGENTE", n_urgente), ("REVISAR", n_revisar),
                 ("INFORMAR", n_informar), ("HISTORIAL", n_hist)],
                max(len(resumen.columns), 10), fila=2)

    res_sorted = resumen.copy()
    res_sorted["_o"] = res_sorted["Prioridad"].map(_orden)
    res_sorted = res_sorted.sort_values(["_o", "Exceso (uds)"], ascending=[True, False])
    res_sorted = res_sorted.drop(columns=["_o", "RUN"], errors="ignore")

    _escribir_tabla(ws2, res_sorted, fila_inicio=3,
                    anchos=ANCHOS_RES, fechas=FECHAS_RES,
                    prio_col="Prioridad", vigente_col="Vigente hoy")

    # ── Hoja 3: Detalle Prescripciones ────────────────────────────────────
    ws3 = wb.create_sheet("Detalle Prescripciones")
    _encabezado_hoja(ws3, "DETALLE DE PRESCRIPCIONES POR CASO DE DUPLICACIÓN",
                     "7C3AED", max(len(detalle.columns), 10))

    det_disp = detalle.drop(columns=["RUN"], errors="ignore")
    _escribir_tabla(ws3, det_disp, fila_inicio=2,
                    anchos={"Paciente": 26, "Medicamento": 40, "Médico": 30,
                            "Especialidad": 22, "Fin est. cobertura": 18},
                    fechas=FECHAS_DET, vigente_col="Vigente hoy")

    # ── Hoja 4: Por Medicamento ────────────────────────────────────────────
    ws4 = wb.create_sheet("Por Medicamento")
    if n_total:
        pm_rows = []
        for med, grp in resumen.groupby("Medicamento"):
            pm_rows.append({
                "Medicamento"       : med,
                "Casos totales"     : len(grp),
                "Pacientes"         : grp["Paciente"].nunique(),
                "Activos hoy"       : int((grp["Vigente hoy"] == "Sí").sum()),
                "URGENTE"           : int((grp["Prioridad"] == "URGENTE").sum()),
                "REVISAR"           : int((grp["Prioridad"] == "REVISAR").sum()),
                "INFORMAR"          : int((grp["Prioridad"] == "INFORMAR").sum()),
                "Distintos médicos" : int((grp["Tipo duplicación"] == "Distintos médicos").sum()),
                "Mismo médico"      : int((grp["Tipo duplicación"] == "Mismo médico").sum()),
                "Exceso total (uds)": int(grp["Exceso (uds)"].sum()),
                "Días dup promedio" : int(grp["Días en doble retiro"].mean()),
            })
        pm = pd.DataFrame(pm_rows).sort_values(
            ["Activos hoy", "URGENTE", "Exceso total (uds)"], ascending=False)
        _encabezado_hoja(ws4, "RESUMEN POR MEDICAMENTO", "0F766E",
                         max(len(pm.columns), 8))
        _escribir_tabla(ws4, pm, fila_inicio=2,
                        anchos={"Medicamento": 44, "Exceso total (uds)": 16,
                                "Días dup promedio": 16})

    # ── Hoja 5: Metodología ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Metodología")
    txt = [
        ("AUDITORÍA PROFUNDA DE PRESCRIPCIONES DUPLICADAS — Farmacia AT Abierta", True),
        (f"Generado: {datetime.now():%d-%m-%Y %H:%M}  ·  Hospital de Pitrufquén (SSASur, Chile)", False),
        ("", False),
        ("Fuente de datos", True),
        (f"• CSV: informe_completo_recetas*.csv (descargados por AUTO_SSASUR.bat)", False),
        (f"• Filtro: Bodega Despacha = {BODEGA_OBJETIVO}", False),
        (f"• Excluidos estados: {', '.join(sorted(ESTADOS_EXCLUIDOS))}", False),
        (f"• Líneas analizadas: {n_lineas:,}  ·  Rango: {rango[0]:%d-%m-%Y} a {rango[1]:%d-%m-%Y}", False),
        ("", False),
        ("Metodología de detección de duplicados", True),
        ("• Unidad de análisis: EVENTO de prescripción (RUN + medicamento normalizado + médico + fecha).", False),
        ("  Las cuotas mensuales de una receta crónica anual se colapsan en 1 evento (ej. '1 de 12' → 1 evento).", False),
        ("• DUPLICADO = ≥2 EVENTOS de un mismo paciente+medicamento cuyas coberturas se SOLAPAN en el tiempo.", False),
        (f"  Cobertura estimada: denominador del campo 'Periodo' × {DIAS_POR_CUOTA} días/cuota.", False),
        ("  Ej. 'X de 12' → 12 cuotas × 30 días = 360 días de cobertura estimada.", False),
        ("• Se EXCLUYEN renovaciones secuenciales (la nueva receta empieza DESPUÉS de que la anterior vence).", False),
        ("", False),
        ("Indicadores añadidos (exclusivos de esta auditoría)", True),
        ("• VIGENTE HOY: el cluster tiene ≥2 prescripciones con cobertura que alcanza la fecha de hoy.", False),
        ("  Es el indicador más importante: representa doble retiro actualmente en curso.", False),
        ("• INICIO DOBLE RETIRO: fecha en que comenzó el solapamiento (= fecha de la 2ª prescripción del cluster).", False),
        ("• DÍAS EN DOBLE RETIRO: desde el inicio hasta hoy (si vigente) o hasta el fin del solapamiento.", False),
        ("", False),
        ("Niveles de prioridad (pre-IA)", True),
        ("• URGENTE   — Vigente hoy + distintos médicos → riesgo máximo (sobrestock / interacción / descoordinación).", False),
        ("• REVISAR   — Vigente hoy + mismo médico → re-prescripción antes de agotar la previa.", False),
        ("• INFORMAR  — Cerrado + distintos médicos → registrar y prevenir recurrencia.", False),
        ("• HISTORIAL — Cerrado + mismo médico → baja urgencia, solo registro.", False),
        ("", False),
        ("Propuestas IA — Claude Haiku", True),
        (f"• Modelo: {MODELO_HAIKU}. Solo casos URGENTE / REVISAR / INFORMAR.", False),
        ("• Los RUTs NUNCA se envían a la API: se usan IDs anónimos (SHA-256, primeros 8 caracteres).", False),
        ("• La propuesta incluye: nivel de acción, plazo sugerido, texto concreto para el QF y alertas clínicas.", False),
        ("• Las propuestas son un TAMIZAJE. Siempre confirmar con la ficha clínica antes de actuar.", False),
        ("", False),
        ("Resultados", True),
        (f"• Casos totales detectados  : {n_total:,}", False),
        (f"• Activos hoy               : {n_activos:,}", False),
        (f"• URGENTES                  : {n_urgente:,}", False),
        (f"• REVISAR                   : {n_revisar:,}", False),
        (f"• INFORMAR                  : {n_informar:,}", False),
        (f"• HISTORIAL                 : {n_hist:,}", False),
        (f"• Con distintos médicos     : {n_dist:,}", False),
        (f"• Exceso total estimado     : {exceso_tot:,} unidades dispensadas", False),
        ("", False),
        ("Privacidad", True),
        ("• Este archivo contiene datos de salud de pacientes (Ley 19.628).", False),
        ("• No distribuir fuera de la institución ni publicar en redes compartidas.", False),
    ]
    for i, (line, bold) in enumerate(txt, 1):
        c = ws5.cell(row=i, column=1, value=line)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 10,
                      color=TEAL if bold else "1F2937")
    ws5.column_dimensions["A"].width = 105

    wb.save(dest)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Auditoría profunda de prescripciones duplicadas — Farmacia AT Abierta")
    ap.add_argument("--salida",  default=None,
                    help="Ruta del Excel de salida.")
    ap.add_argument("--rapido",  action="store_true",
                    help="Usar solo el CSV más reciente (más rápido, menos histórico).")
    ap.add_argument("--sin-ia",  action="store_true",
                    help="Omitir llamadas a Claude Haiku (propuestas IA vacías).")
    args = ap.parse_args()

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    dest = args.salida or os.path.join(WORK, f"Auditoria_Duplicados_Profunda_{ts}.xlsx")

    print("=" * 70)
    print("  AUDITORÍA PROFUNDA DE PRESCRIPCIONES DUPLICADAS")
    print(f"  Farmacia AT Abierta · Hospital de Pitrufquén · {HOY:%d-%m-%Y}")
    print("=" * 70)

    rec    = cargar_y_preparar(args.rapido)
    rango  = (rec["_fecha"].min().date(), rec["_fecha"].max().date())
    n_lin  = len(rec)

    rl = nivel_evento(rec)
    print(f"Eventos de prescripción: {len(rl):,}")

    resumen, detalle = construir_casos(rl)

    if not len(resumen):
        print("\n[OK] No se detectaron casos de prescripción duplicada.")
        sys.exit(0)

    n_act  = int((resumen["Vigente hoy"] == "Sí").sum())
    n_urg  = int((resumen["Prioridad"]   == "URGENTE").sum())
    n_rev  = int((resumen["Prioridad"]   == "REVISAR").sum())
    n_dist = int((resumen["Tipo duplicación"] == "Distintos médicos").sum())

    print(f"\n{'='*70}")
    print(f"  Casos detectados         : {len(resumen):,}")
    print(f"  Activos hoy              : {n_act:,}")
    print(f"  URGENTES                 : {n_urg:,}")
    print(f"  REVISAR                  : {n_rev:,}")
    print(f"  Con distintos médicos    : {n_dist:,}")
    print(f"  Exceso estimado total    : {int(resumen['Exceso (uds)'].sum()):,} uds")
    print(f"{'='*70}")

    if not args.sin_ia:
        print("\n  Generando propuestas IA (Claude Haiku)...")
        try:
            import anthropic
            propuestas = generar_propuestas_haiku(resumen)
            print(f"  Propuestas generadas: {len(propuestas)}")
            resumen = aplicar_propuestas(resumen, propuestas)
        except ImportError:
            print("  [AVISO] anthropic no instalado. Continúa sin propuestas IA.")
        except Exception as e:
            print(f"  [AVISO IA] {e}. Continúa sin propuestas IA.")
    else:
        print("\n  [--sin-ia] Propuestas IA desactivadas.")

    print("\n  Exportando Excel...")
    exportar_excel(resumen, detalle, n_lin, rango, dest)
    print(f"\n[OK] Excel generado: {dest}  ({os.path.getsize(dest) // 1024} KB)")


if __name__ == "__main__":
    main()
