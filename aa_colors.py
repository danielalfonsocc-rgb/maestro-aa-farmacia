"""Paleta de colores compartida para los niveles de Criticidad de AA.

Fuente unica de verdad usada por maestro_aa.py y app_pedidos.py, para que
todos los Excel generados (Consolidado, Resumen, descargas del dashboard)
muestren siempre el mismo color para el mismo nivel de criticidad.

Escalas de Criticidad usadas en el proyecto:
  Crit_Farm (crit_farm): 5-OK, 1-CRITICO, 2-URGENTE, 3-MODERADO, 4-BAJO
  Crit_Bod  (crit_bod) : 5-OK, 1-CRITICO, 2-URGENTE, 3-ALTO, 4-MODERADO, 5-BAJO

Los pares 3-MODERADO/4-MODERADO y 4-BAJO/5-BAJO comparten color: ocupan la
misma posicion relativa (penultimo / ultimo escalon de urgencia) en cada
escala, aunque tengan distinto numero por tener Bodega un escalon mas.
"""

from openpyxl.styles import PatternFill, Font

# Paleta TENUE pensada para impresion tamano carta (economia de tinta): todos
# los fondos son claros/pastel — ninguno es un bloque oscuro saturado. La senal
# de "CRITICO" se da con el TEXTO en vino oscuro y negrita (ver FONT_CRITICO),
# no pintando la celda de rojo fuerte. Asi el papel queda casi en blanco.
CRIT_FILL_HEX = {
    '1-CRITICO' : 'F4B3B3',  # rosa pastel    - sin stock / sin respaldo (texto vino+negrita)
    '2-URGENTE' : 'FFCDD2',  # rosa claro
    '3-ALTO'    : 'FFE0B2',  # naranja medio  (Bodega: cobertura < 1 semana laboral)
    '3-MODERADO': 'FFF3E0',  # naranja claro  (Farmacia)
    '4-MODERADO': 'FFF3E0',  # naranja claro  (Bodega: cobertura < 1 ciclo de 2 semanas)
    '4-BAJO'    : 'FFF9C4',  # amarillo       (Farmacia)
    '5-BAJO'    : 'FFF9C4',  # amarillo       (Bodega: cobertura >= 1 ciclo de 2 semanas)
    '5-OK'      : 'E8F5E9',  # verde          - sin necesidad
}

# Texto del nivel CRITICO: vino oscuro + negrita sobre el rosa pastel. La "alarma"
# vive en el texto (gasta poquisima tinta), no en un fondo rojo solido.
FONT_CRITICO = Font(bold=True, color='7F1D1D', name='Arial', size=11)


def crit_fill(crit, default='FFFFFF'):
    """PatternFill solido para un nivel de Criticidad (1-CRITICO, 2-URGENTE, ...)."""
    return PatternFill('solid', fgColor=CRIT_FILL_HEX.get(str(crit), default))


# ─────────────────────────────────────────────
# NIVEL DE CRITICIDAD — normaliza las DOS escalas del proyecto a 1..5
# ─────────────────────────────────────────────
# El proyecto usa dos formas de escribir la Criticidad:
#   Pedidos   : '1-CRITICO', '2-URGENTE', '3-ALTO'/'3-MODERADO', '4-...', '5-...'
#   Faltantes : '[CRITICO] CRITICO - SIN RESPALDO', '[ALTO] ...',
#               '[MODERADO] ...', '[BAJO] ...'
# crit_nivel() las reduce a un solo numero (1 = mas urgente, 5 = sin necesidad)
# para que la app y los Excel coloreen y ordenen igual sin importar la escala.
_NIVEL_HEX = {1: 'F4B3B3', 2: 'FFCDD2', 3: 'FFE0B2', 4: 'FFF9C4', 5: 'E8F5E9'}


def crit_nivel(crit) -> int:
    """Devuelve 1..5 para cualquier etiqueta de Criticidad (ambas escalas)."""
    c = str(crit).strip().upper()
    # Escala numerica de pedidos: el prefijo "N-" manda
    if len(c) >= 2 and c[0] in '12345' and c[1] == '-':
        return int(c[0])
    # Escala de faltantes (texto): CRITICO > ALTO > MODERADO > BAJO
    if 'CRITICO'  in c: return 1
    if 'URGENTE'  in c: return 2
    if 'ALTO'     in c: return 2
    if 'MODERADO' in c: return 3
    if 'BAJO'     in c: return 4
    return 5


def crit_hex(crit, default='F8F8F8'):
    """Color hex representativo para cualquier Criticidad (via crit_nivel)."""
    return _NIVEL_HEX.get(crit_nivel(crit), default)


def is_critico(crit):
    return crit_nivel(crit) == 1


# ─────────────────────────────────────────────
# ATENUADO DE BANDAS — headers y titulos tinta-economicos
# ─────────────────────────────────────────────
# Los encabezados/titulos del proyecto usaban bandas oscuras saturadas (azul
# 1F4E78, magenta 880E4F, etc.) con texto blanco: al imprimir en carta eso es
# un bloque solido de tinta por cada hoja. soften() aclara el color hasta un
# pastel (conserva el matiz que identifica cada hoja, pero casi sin tinta) y
# darken()/text_on() eligen un texto legible y oscuro encima.

def _rgb(hexstr):
    h = str(hexstr).lstrip('#')[-6:].rjust(6, '0')
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def soften(hexstr, keep=0.20):
    """Version pastel (clara) de un color de banda — mezcla keep del color con blanco.
    keep=0.20 deja ~20% del color: el matiz se reconoce, el fondo casi no gasta tinta."""
    r, g, b = _rgb(hexstr)
    f = lambda v: round(255 - (255 - v) * keep)
    return f'{f(r):02X}{f(g):02X}{f(b):02X}'


def darken(hexstr, keep=0.55):
    """Version oscura del mismo matiz — para texto/negrita sobre el pastel de soften()."""
    r, g, b = _rgb(hexstr)
    f = lambda v: round(v * keep)
    return f'{f(r):02X}{f(g):02X}{f(b):02X}'


def text_on(hexstr, dark='1F2937', light='FFFFFF'):
    """Color de texto legible segun la luminancia del fondo (oscuro sobre claro)."""
    r, g, b = _rgb(hexstr)
    return dark if (0.2126 * r + 0.7152 * g + 0.0722 * b) > 150 else light


def fill_hex(fill, default='FFFFFF'):
    """Extrae el hex 'RRGGBB' de un PatternFill (o devuelve el string tal cual)."""
    if fill is None:
        return default
    if isinstance(fill, str):
        return fill.lstrip('#')[-6:].rjust(6, '0')
    try:
        return str(fill.fgColor.rgb)[-6:]
    except Exception:
        return default
