#!/usr/bin/env python3
"""
AUDITORÍA DE PRESCRIPCIÓN DUPLICADA — Farmacia AT Abierta (Hospital Pitrufquén)
==============================================================================
Genera un Excel con los PACIENTES que tienen prescripciones DUPLICADAS de un
mismo medicamento, distinguiendo si la duplicación fue del MISMO médico o de
médicos DISTINTOS.

Definición de "prescripción duplicada" (documentada en la hoja Metodología):
  - Mismo paciente (RUN) + mismo medicamento (Prescripción normalizada).
  - ≥2 Números de Receta DISTINTOS cuyos períodos de cobertura se SOLAPAN en el
    tiempo. La cobertura de cada receta se estima con las cuotas del campo
    "Periodo" ("X de N"  →  N cuotas mensuales  →  N·30 días). Una receta
    "1 de 1" cubre ~30 días; una "X de 12" cubre ~360 días.
  - Esto evita marcar como duplicado las RENOVACIONES SECUENCIALES (p.ej. la
    misma receta crónica re-emitida 8 meses después, sin solape real).

Alcance: SOLO Bodega Despacha = FARMACIA AT ABIERTA (ambulatorio/crónicos).
Se excluyen prescripciones ANULADO / RECHAZADO / REEMPLAZADO.

Fuente: las mismas sábanas informe_completo_recetas*.csv que ya baja AUTO_SSASUR
(no se publican a la nube; este Excel queda local con datos de pacientes).

Uso:   py recetas_duplicadas.py [--salida <archivo.xlsx>]
"""
import os
import re
import sys
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aa_colors import TEAL, ROJO, NARANJA, GRIS
from utils_aa import norm_erp, HOMOLOGACION, cargar_recetas_csv, setup_stdout

setup_stdout()
WORK = os.path.dirname(os.path.abspath(__file__))

BODEGA_OBJETIVO = "FARMACIA AT ABIERTA"
ESTADOS_EXCLUIDOS = {"ANULADO", "RECHAZADO", "REEMPLAZADO"}
DIAS_POR_CUOTA = 30
MIN_EVENTOS = 2


def parse_cuotas(serie) -> int:
    """Máximo denominador 'de N' del campo Periodo (= nº de cuotas del tratamiento)."""
    n = 1
    for v in serie.dropna():
        m = re.search(r"de\s+(\d+)", str(v))
        if m:
            n = max(n, int(m.group(1)))
    return n


def cargar_recetas() -> pd.DataFrame:
    solo_ultimo = "--rapido" in sys.argv
    cols = [
        "ID Receta Detalle", "RUN", "Nombre", "Apellido Paterno", "Apellido Materno",
        "Prescripción", "Número Receta", "Periodo",
        "Fecha Atención", "Fecha Entrega Receta", "Estado Prescripción",
        "RUN Profesional", "Nombre Profesional",
        "Apellido Paterno Profesional", "Apellido Materno Profesional",
        "Especialidad", "Bodega Despacha",
        "Cantidad Recetada", "Cantidad Entregada",
    ]
    try:
        rec = cargar_recetas_csv(WORK, cols=cols, solo_ultimo=solo_ultimo)
    except FileNotFoundError:
        print("[AVISO] No hay CSV de recetas — nada que auditar.")
        sys.exit(0)
    if solo_ultimo:
        print("[--rapido] usando solo el CSV más reciente")
    print(f"Recetas (líneas) dedup por ID Receta Detalle: {len(rec):,}")
    return rec


def preparar(rec: pd.DataFrame) -> pd.DataFrame:
    rec = rec.copy()
    rec["_bod"] = rec["Bodega Despacha"].fillna("").apply(norm_erp)
    rec["_est"] = rec["Estado Prescripción"].fillna("").str.upper().str.strip()
    rec = rec[(rec["_bod"] == BODEGA_OBJETIVO) & (~rec["_est"].isin(ESTADOS_EXCLUIDOS))].copy()

    rec["_med"] = rec["Prescripción"].fillna("").apply(norm_erp).map(lambda x: HOMOLOGACION.get(x, x))
    fa = pd.to_datetime(rec["Fecha Atención"], dayfirst=True, errors="coerce")
    fe = pd.to_datetime(rec["Fecha Entrega Receta"], dayfirst=True, errors="coerce")
    rec["_fecha"] = fa.fillna(fe)
    rec["_paciente"] = (rec["Nombre"].fillna("") + " " + rec["Apellido Paterno"].fillna("")
                        + " " + rec["Apellido Materno"].fillna("")).str.strip().str.title()
    rec["_medico"] = (rec["Nombre Profesional"].fillna("") + " "
                      + rec["Apellido Paterno Profesional"].fillna("") + " "
                      + rec["Apellido Materno Profesional"].fillna("")).str.strip().str.title()
    rec["_run_prof"] = rec["RUN Profesional"].fillna("").str.strip()
    rec["_esp"] = rec["Especialidad"].fillna("").str.strip()
    rec["_cant_r"] = pd.to_numeric(rec["Cantidad Recetada"], errors="coerce").fillna(0)
    rec["_cant_e"] = pd.to_numeric(rec["Cantidad Entregada"], errors="coerce").fillna(0)
    rec = rec.dropna(subset=["_fecha"])
    rec = rec[rec["Número Receta"].fillna("").str.strip() != ""]
    rec = rec[rec["_med"] != ""]
    print(f"Filas Farmacia AT Abierta válidas: {len(rec):,}  "
          f"| rango: {rec['_fecha'].min():%d-%m-%Y} a {rec['_fecha'].max():%d-%m-%Y}")
    return rec


def nivel_evento(rec: pd.DataFrame) -> pd.DataFrame:
    """Colapsa a 1 fila por EVENTO de prescripción = (RUN, med, médico, fecha).

    En el ERP SSASUR una receta crónica anual se emite como N° de receta SEPARADO
    por cada cuota mensual (p.ej. una indicación 'de 12' genera ~12 N° de receta el
    mismo día). Por eso NO se puede usar 'N° de receta distinto' como 'prescripción
    distinta': hay que agrupar las cuotas que comparten paciente+medicamento+médico+
    fecha de atención en un solo evento de prescripción y sumar sus cantidades."""
    def primero(s):
        s = s[s.astype(str).str.strip() != ""]
        return s.mode().iloc[0] if not s.empty else ""
    g = rec.groupby(["RUN", "_med", "_run_prof", "_fecha"]).agg(
        paciente=("_paciente", primero),
        med_nombre=("Prescripción", primero),
        cuotas=("Periodo", parse_cuotas),
        medico=("_medico", primero),
        esp=("_esp", primero),
        cant_r=("_cant_r", "sum"),
        cant_e=("_cant_e", "sum"),
        n_cuotas=("Número Receta", "nunique"),
        estado=("_est", primero),
    ).reset_index().rename(columns={"_run_prof": "run_prof", "_fecha": "start"})
    g["cob_dias"] = g["cuotas"].clip(lower=1) * DIAS_POR_CUOTA
    return g


def clusters_solapados(sub: pd.DataFrame):
    """Agrupa recetas (de un mismo RUN+med) cuyas coberturas se solapan. Devuelve
    listas con ≥2 recetas (= eventos de duplicación)."""
    sub = sub.sort_values("start")
    recs = sub.to_dict("records")
    out, cur, end = [], [], None
    for r in recs:
        r_end = r["start"] + timedelta(days=int(r["cob_dias"]))
        if not cur:
            cur, end = [r], r_end
        elif r["start"] <= end:
            cur.append(r)
            end = max(end, r_end)
        else:
            if len(cur) >= MIN_EVENTOS:
                out.append(cur)
            cur, end = [r], r_end
    if len(cur) >= MIN_EVENTOS:
        out.append(cur)
    return out


def solape_max_dias(cl):
    """Máximo solape temporal (días) entre recetas consecutivas del cluster."""
    cl = sorted(cl, key=lambda r: r["start"])
    mx = 0
    for i in range(1, len(cl)):
        prev = cl[i - 1]
        ov = (prev["start"] + timedelta(days=int(prev["cob_dias"])) - cl[i]["start"]).days
        mx = max(mx, ov)
    return max(mx, 0)


def construir(rl: pd.DataFrame):
    resumen, detalle = [], []
    cid = 0
    for (run, med), sub in rl.groupby(["RUN", "_med"]):
        if len(sub) < MIN_EVENTOS:          # < 2 eventos de prescripción → no hay duplicado
            continue
        for cl in clusters_solapados(sub):
            cid += 1
            run_profs = sorted({r["run_prof"] for r in cl if r["run_prof"]})
            medicos = sorted({r["medico"] for r in cl if r["medico"]})
            esps = sorted({r["esp"] for r in cl if r["esp"]})
            n_med = max(len(run_profs), 1)
            tipo = "🔴 Distintos médicos" if n_med >= 2 else "🟠 Mismo médico"
            fechas = [r["start"] for r in cl]
            entregas = [r["cant_e"] for r in cl]
            # Consumo (unidades entregadas): "con duplicados" suma todas las recetas
            # del episodio; "sin duplicados" deja sólo la de mayor cantidad (legítima)
            # y descarta el resto como duplicado. El exceso = lo evitable.
            con_dup = int(sum(entregas))
            sin_dup = int(max(entregas)) if entregas else 0
            resumen.append({
                "RUN": run,
                "Paciente": cl[0]["paciente"],
                "Medicamento": cl[0]["med_nombre"],
                "N° prescripciones": len(cl),
                "Tipo duplicación": tipo,
                "N° médicos": n_med,
                "Médico(s)": " ; ".join(medicos),
                "Especialidad(es)": " ; ".join([e for e in esps if e]) or "—",
                "1ª prescripción": min(fechas),
                "Última prescripción": max(fechas),
                "Días 1ª→última": (max(fechas) - min(fechas)).days,
                "Solape estimado (días)": solape_max_dias(cl),
                "Total recetado": int(sum(r["cant_r"] for r in cl)),
                "Consumo c/ duplicados": con_dup,
                "Consumo s/ duplicados": sin_dup,
                "Exceso por duplicación": con_dup - sin_dup,
                "N° receta-cuotas": int(sum(r["n_cuotas"] for r in cl)),
                "_cid": cid,
                "_medkey": med,
                "_dist": n_med >= 2,
            })
            for r in sorted(cl, key=lambda x: x["start"]):
                detalle.append({
                    "Caso": cid,
                    "RUN": run,
                    "Paciente": r["paciente"],
                    "Medicamento": r["med_nombre"],
                    "Fecha prescripción": r["start"],
                    "Médico": r["medico"],
                    "RUN médico": r["run_prof"],
                    "Especialidad": r["esp"] or "—",
                    "Cuotas (Periodo)": int(r["cuotas"]),
                    "Cobertura est. (días)": int(r["cob_dias"]),
                    "N° receta-cuotas": int(r["n_cuotas"]),
                    "Cant. recetada": int(r["cant_r"]),
                    "Cant. entregada": int(r["cant_e"]),
                    "Estado": r["estado"],
                })
    return pd.DataFrame(resumen), pd.DataFrame(detalle)


# ── Escritura Excel (openpyxl, estilo clínico-profesional) ────────────────────
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _estilizar_encabezado(ws, ncols, fila=1):
    fill = PatternFill("solid", fgColor=TEAL)
    font = Font(bold=True, color="FFFFFF", size=10)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=fila, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[fila].height = 30
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def _escribir_df(ws, df, anchos=None, fechas=None, tipo_col=None):
    fechas = fechas or []
    df = df.copy()
    for col in df.columns:
        if str(col).startswith("_"):
            df = df.drop(columns=col)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row.values))
    _estilizar_encabezado(ws, len(df.columns))
    # formato de fechas
    col_idx = {c: i + 1 for i, c in enumerate(df.columns)}
    for fcol in fechas:
        if fcol in col_idx:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx[fcol]).number_format = "DD-MM-YYYY"
    # coloreo del tipo (rojo distinto / naranja mismo)
    if tipo_col and tipo_col in col_idx:
        ci = col_idx[tipo_col]
        for r in range(2, ws.max_row + 1):
            val = str(ws.cell(row=r, column=ci).value or "")
            color = ROJO if "Distintos" in val else NARANJA
            ws.cell(row=r, column=ci).font = Font(bold=True, color=color)
    # bordes + zebra
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GRIS)
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)
    # anchos
    for i, col in enumerate(df.columns, 1):
        w = (anchos or {}).get(col)
        if w is None:
            w = min(max(len(str(col)) + 2, df[col].astype(str).str.len().max() + 2 if len(df) else 12), 46)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"


def exportar(resumen: pd.DataFrame, detalle: pd.DataFrame, med_tot: dict, rango, n_lineas, dest):
    wb = Workbook()

    # Hoja 1 — Resumen (un caso de duplicación por fila)
    ws = wb.active
    ws.title = "Resumen"
    if len(resumen):
        res = resumen.sort_values(["_dist", "Exceso por duplicación", "N° prescripciones"],
                                  ascending=[False, False, False]).drop(columns=["_cid", "_dist"])
    else:
        res = resumen.drop(columns=["_cid", "_dist"], errors="ignore")
    _escribir_df(
        ws, res,
        anchos={"Paciente": 26, "Medicamento": 40, "Médico(s)": 34, "Especialidad(es)": 22,
                "N° receta-cuotas": 14, "Tipo duplicación": 20,
                "Consumo c/ duplicados": 16, "Consumo s/ duplicados": 16,
                "Exceso por duplicación": 16},
        fechas=["1ª prescripción", "Última prescripción"],
        tipo_col="Tipo duplicación",
    )

    # Hoja 2 — Detalle de recetas
    ws2 = wb.create_sheet("Detalle recetas")
    _escribir_df(
        ws2, detalle,
        anchos={"Paciente": 26, "Medicamento": 40, "Médico": 30, "Especialidad": 22},
        fechas=["Fecha prescripción"],
    )

    # Hoja 3 — Por medicamento (consumo con vs sin duplicados)
    if len(resumen):
        pm = resumen.groupby("_medkey").agg(
            Medicamento=("Medicamento", lambda s: s.mode().iloc[0] if not s.mode().empty else s.iloc[0]),
            Casos=("RUN", "size"),
            Pacientes=("RUN", "nunique"),
            Distintos=("_dist", "sum"),
            Exceso=("Exceso por duplicación", "sum"),
        ).reset_index()
        pm["Mismo médico"] = (pm["Casos"] - pm["Distintos"]).astype(int)
        pm["Consumo total c/ dup"] = pm["_medkey"].map(med_tot).fillna(0).astype(int)
        pm["Exceso por duplicación"] = pm["Exceso"].astype(int)
        pm["Consumo total s/ dup"] = (pm["Consumo total c/ dup"] - pm["Exceso por duplicación"]).clip(lower=0)
        _tot = pm["Consumo total c/ dup"].astype(float).replace(0, float("nan"))
        pm["% exceso"] = (100 * pm["Exceso por duplicación"] / _tot).fillna(0).round(1)
        pm = pm.rename(columns={"Distintos": "Distintos médicos"})
        pm = pm[["Medicamento", "Casos", "Pacientes", "Distintos médicos", "Mismo médico",
                 "Consumo total c/ dup", "Consumo total s/ dup", "Exceso por duplicación", "% exceso"]]
        pm = pm.sort_values(["Exceso por duplicación", "Casos"], ascending=False)
        ws3 = wb.create_sheet("Por medicamento")
        _escribir_df(ws3, pm, anchos={"Medicamento": 44, "Consumo total c/ dup": 16,
                                      "Consumo total s/ dup": 16, "Exceso por duplicación": 16})

    # Hoja 4 — Metodología
    ws4 = wb.create_sheet("Metodología")
    n_dist = int(resumen["_dist"].sum()) if len(resumen) else 0
    n_mismo = (len(resumen) - n_dist) if len(resumen) else 0
    n_pac = resumen["RUN"].nunique() if len(resumen) else 0
    g_con = int(resumen["Consumo c/ duplicados"].sum()) if len(resumen) else 0
    g_sin = int(resumen["Consumo s/ duplicados"].sum()) if len(resumen) else 0
    txt = [
        ("AUDITORÍA DE PRESCRIPCIÓN DUPLICADA — Farmacia AT Abierta", True),
        (f"Generado: {datetime.now():%d-%m-%Y %H:%M}", False),
        ("", False),
        ("Unidad de análisis = EVENTO de prescripción (no N° de receta):", True),
        ("• En el ERP una indicación crónica anual se emite como un N° de receta SEPARADO", False),
        ("    por cada cuota mensual (una indicación 'de 12' → ~12 N° de receta el mismo día).", False),
        ("• Por eso las cuotas que comparten paciente+medicamento+médico+fecha se agrupan en", False),
        ("    UN evento de prescripción y se suman sus cantidades. Una receta anual = 1 evento", False),
        ("    (NO se cuenta como 12 duplicados). 'N° receta-cuotas' indica cuántos N° agrupó.", False),
        ("", False),
        ("Definición de prescripción duplicada:", True),
        ("• Mismo paciente (RUN) + mismo medicamento (Prescripción normalizada).", False),
        ("• ≥2 EVENTOS de prescripción distintos cuyas coberturas se SOLAPAN en el tiempo.", False),
        ("• Cobertura de cada evento estimada por cuotas del campo Periodo ('X de N'):", False),
        (f"    N cuotas × {DIAS_POR_CUOTA} días.  '1 de 1' ≈ 30 días; 'X de 12' ≈ 360 días.", False),
        ("• Así se EXCLUYEN renovaciones secuenciales (sin solape real de suministro).", False),
        ("", False),
        ("Consumo CON vs SIN duplicados (unidades entregadas):", True),
        ("• Consumo c/ duplicados = suma de las unidades entregadas de TODOS los eventos", False),
        ("    del episodio (lo que realmente salió de farmacia).", False),
        ("• Consumo s/ duplicados = unidades del evento de MAYOR cantidad del episodio", False),
        ("    (se considera legítimo ese y se descarta el resto como duplicado).", False),
        ("• Exceso por duplicación = c/ duplicados − s/ duplicados = consumo evitable.", False),
        ("• 'Por medicamento' compara el consumo TOTAL del fármaco contra el mismo total", False),
        ("    descontando ese exceso, y el % que representa.", False),
        ("", False),
        ("Tipo de duplicación:", True),
        ("• 🔴 Distintos médicos: los eventos solapados fueron emitidos por ≥2 prescriptores", False),
        ("    distintos (posible descoordinación entre nivel primario y especialista).", False),
        ("• 🟠 Mismo médico: re-prescripción por el mismo profesional antes de agotar la previa.", False),
        ("", False),
        ("Filtros aplicados:", True),
        (f"• Bodega Despacha = {BODEGA_OBJETIVO}.", False),
        ("• Excluidos estados de prescripción: " + ", ".join(sorted(ESTADOS_EXCLUIDOS)) + ".", False),
        ("• Fecha de prescripción = Fecha Atención (fallback Fecha Entrega Receta).", False),
        ("", False),
        ("Resultados:", True),
        (f"• Líneas de receta analizadas (Farmacia AT Abierta): {n_lineas:,}", False),
        (f"• Rango de fechas: {rango[0]:%d-%m-%Y} a {rango[1]:%d-%m-%Y}", False),
        (f"• Casos de duplicación detectados: {len(resumen):,}", False),
        (f"• Pacientes involucrados: {n_pac:,}", False),
        (f"• Por DISTINTOS médicos: {n_dist:,}", False),
        (f"• Por MISMO médico: {n_mismo:,}", False),
        (f"• Consumo en casos duplicados — CON duplicados: {g_con:,} ud entregadas", False),
        (f"• Consumo en casos duplicados — SIN duplicados: {g_sin:,} ud entregadas", False),
        (f"• Exceso de consumo atribuible a duplicación: {g_con - g_sin:,} ud", False),
        ("", False),
        ("Nota clínica: este listado es un TAMIZAJE. Cada caso debe revisarse con la ficha", False),
        ("para confirmar si la duplicación es real (terapia duplicada / sobre-stock) o", False),
        ("justificada (cambio de dosis, ajuste, traslape administrativo).", False),
    ]
    for i, (line, bold) in enumerate(txt, 1):
        c = ws4.cell(row=i, column=1, value=line)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 10,
                      color=TEAL if bold else "1F2937")
    ws4.column_dimensions["A"].width = 95

    wb.save(dest)


def main():
    dest = None
    if "--salida" in sys.argv:
        dest = sys.argv[sys.argv.index("--salida") + 1]
    if not dest:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        dest = os.path.join(WORK, f"Duplicidad_Recetas_AA_{ts}.xlsx")

    rec = cargar_recetas()
    rec = preparar(rec)
    rango = (rec["_fecha"].min(), rec["_fecha"].max())
    n_lineas = len(rec)
    rl = nivel_evento(rec)
    med_tot = rl.groupby("_med")["cant_e"].sum().astype(int).to_dict()  # consumo total por med
    resumen, detalle = construir(rl)

    n_dist = int(resumen["_dist"].sum()) if len(resumen) else 0
    print("\n" + "=" * 64)
    print(f"  Casos de prescripción duplicada : {len(resumen):,}")
    print(f"  Pacientes involucrados          : {resumen['RUN'].nunique() if len(resumen) else 0:,}")
    print(f"  🔴 Distintos médicos            : {n_dist:,}")
    print(f"  🟠 Mismo médico                 : {len(resumen) - n_dist:,}")
    print("=" * 64)

    exportar(resumen, detalle, med_tot, rango, n_lineas, dest)
    print(f"\n[OK] Excel generado: {dest}  ({os.path.getsize(dest) // 1024} KB)")


if __name__ == "__main__":
    main()
